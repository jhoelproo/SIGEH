from __future__ import annotations

import hashlib
import json
import os
import re
import tempfile
import threading
import uuid
from copy import deepcopy
from datetime import date, datetime, timezone
from decimal import Decimal
from pathlib import Path
from typing import Any

import psycopg2.extras

from historical_documents import parse_shift_closure_source_key
from report_engine import ReportHTMLRenderer

STORAGE_LEGACY = "LEGACY_PDF"
STORAGE_HYBRID = "HYBRID"
STORAGE_SNAPSHOT = "SNAPSHOT"
STORAGE_MODES = (STORAGE_LEGACY, STORAGE_HYBRID, STORAGE_SNAPSHOT)
REPORT_TEMPLATE_VERSION = "report_html_v1"
REPORT_SNAPSHOT_SCHEMA_VERSION = 2
SUPPORTED_REPORT_SNAPSHOT_SCHEMA_VERSIONS = (1, REPORT_SNAPSHOT_SCHEMA_VERSION)


class ReportDocumentError(RuntimeError):
    pass


class ReportSnapshotMissingError(ReportDocumentError):
    pass


class ReportSnapshotHashError(ReportDocumentError):
    pass


class ReportTemplateError(ReportDocumentError):
    pass


REPORT_DOCUMENT_MIGRATION_SQL = """
ALTER TABLE report_history
  ADD COLUMN IF NOT EXISTS document_storage_mode TEXT;
ALTER TABLE report_history
  ADD COLUMN IF NOT EXISTS revision_version INTEGER NOT NULL DEFAULT 0;
ALTER TABLE report_history
  ADD COLUMN IF NOT EXISTS legacy_pdf_backup_reference TEXT;
ALTER TABLE report_history
  ADD COLUMN IF NOT EXISTS legacy_pdf_checksum TEXT;
ALTER TABLE report_history
  ADD COLUMN IF NOT EXISTS legacy_pdf_size BIGINT;
ALTER TABLE report_history
  ADD COLUMN IF NOT EXISTS is_deleted BOOLEAN NOT NULL DEFAULT FALSE;

UPDATE report_history
SET document_storage_mode='LEGACY_PDF'
WHERE document_storage_mode IS NULL OR BTRIM(document_storage_mode)='';

ALTER TABLE report_history
  ALTER COLUMN document_storage_mode SET DEFAULT 'SNAPSHOT';
ALTER TABLE report_history
  ALTER COLUMN document_storage_mode SET NOT NULL;

ALTER TABLE daily_reports
  ADD COLUMN IF NOT EXISTS document_storage_mode TEXT;
ALTER TABLE daily_reports
  ADD COLUMN IF NOT EXISTS revision_version INTEGER NOT NULL DEFAULT 0;
ALTER TABLE daily_reports
  ADD COLUMN IF NOT EXISTS legacy_pdf_backup_reference TEXT;
ALTER TABLE daily_reports
  ADD COLUMN IF NOT EXISTS legacy_pdf_checksum TEXT;
ALTER TABLE daily_reports
  ADD COLUMN IF NOT EXISTS legacy_pdf_size BIGINT;
ALTER TABLE daily_reports
  ADD COLUMN IF NOT EXISTS is_deleted BOOLEAN NOT NULL DEFAULT FALSE;

UPDATE daily_reports
SET document_storage_mode='LEGACY_PDF'
WHERE document_storage_mode IS NULL OR BTRIM(document_storage_mode)='';

ALTER TABLE daily_reports
  ALTER COLUMN document_storage_mode SET DEFAULT 'SNAPSHOT';
ALTER TABLE daily_reports
  ALTER COLUMN document_storage_mode SET NOT NULL;

ALTER TABLE billing_shift_closures
  ADD COLUMN IF NOT EXISTS document_storage_mode TEXT;
ALTER TABLE billing_shift_closures
  ADD COLUMN IF NOT EXISTS revision_version INTEGER NOT NULL DEFAULT 0;
ALTER TABLE billing_shift_closures
  ADD COLUMN IF NOT EXISTS legacy_pdf_backup_reference TEXT;
ALTER TABLE billing_shift_closures
  ADD COLUMN IF NOT EXISTS legacy_pdf_checksum TEXT;
ALTER TABLE billing_shift_closures
  ADD COLUMN IF NOT EXISTS legacy_pdf_size BIGINT;

UPDATE billing_shift_closures
SET document_storage_mode='LEGACY_PDF'
WHERE document_storage_mode IS NULL OR BTRIM(document_storage_mode)='';

ALTER TABLE billing_shift_closures
  ALTER COLUMN document_storage_mode SET DEFAULT 'SNAPSHOT';
ALTER TABLE billing_shift_closures
  ALTER COLUMN document_storage_mode SET NOT NULL;

DO $$
DECLARE
  target_table TEXT;
  constraint_name TEXT;
BEGIN
  FOREACH target_table IN ARRAY ARRAY[
    'report_history','daily_reports','billing_shift_closures'
  ]
  LOOP
    constraint_name := target_table || '_document_storage_mode_check';
    IF NOT EXISTS(
      SELECT 1 FROM pg_constraint
      WHERE conname=constraint_name
        AND conrelid=target_table::regclass
    ) THEN
      EXECUTE FORMAT(
        'ALTER TABLE %I ADD CONSTRAINT %I CHECK(document_storage_mode IN (''LEGACY_PDF'',''HYBRID'',''SNAPSHOT''))',
        target_table,constraint_name
      );
    END IF;
  END LOOP;
END $$;

CREATE TABLE IF NOT EXISTS report_document_versions(
  id BIGSERIAL PRIMARY KEY,
  source_table TEXT NOT NULL,
  source_key TEXT NOT NULL,
  report_id INTEGER,
  version INTEGER NOT NULL CHECK(version > 0),
  report_type TEXT NOT NULL,
  report_title TEXT NOT NULL,
  period_start DATE,
  period_end DATE,
  filters_jsonb JSONB NOT NULL DEFAULT '{}'::jsonb,
  financial_basis_jsonb JSONB NOT NULL DEFAULT '{}'::jsonb,
  dataset_snapshot_jsonb JSONB NOT NULL DEFAULT '{}'::jsonb,
  summary_snapshot_jsonb JSONB NOT NULL DEFAULT '{}'::jsonb,
  charts_snapshot_jsonb JSONB NOT NULL DEFAULT '{}'::jsonb,
  guided_reading_snapshot_jsonb JSONB NOT NULL DEFAULT '{}'::jsonb,
  render_context_jsonb JSONB NOT NULL,
  snapshot_jsonb JSONB NOT NULL,
  snapshot_hash TEXT NOT NULL CHECK(snapshot_hash ~ '^[0-9a-f]{64}$'),
  template_version TEXT NOT NULL,
  schema_version INTEGER NOT NULL CHECK(schema_version > 0),
  created_at TIMESTAMPTZ NOT NULL DEFAULT NOW(),
  created_by TEXT NOT NULL,
  generated_by_user_id TEXT,
  created_from_module TEXT,
  report_uuid TEXT,
  report_generation_uuid TEXT,
  is_current BOOLEAN NOT NULL DEFAULT TRUE,
  CONSTRAINT report_document_versions_source_check CHECK(
    source_table IN('report_history','daily_reports','billing_shift_closures')
  ),
  CONSTRAINT report_document_versions_report_fk
    FOREIGN KEY(report_id) REFERENCES report_history(id) ON DELETE RESTRICT,
  CONSTRAINT report_document_versions_source_version_uq
    UNIQUE(source_table,source_key,version)
);
ALTER TABLE report_document_versions
  ADD COLUMN IF NOT EXISTS generated_by_user_id TEXT;
ALTER TABLE report_document_versions
  ADD COLUMN IF NOT EXISTS created_from_module TEXT;
ALTER TABLE report_document_versions
  ADD COLUMN IF NOT EXISTS report_uuid TEXT;
ALTER TABLE report_document_versions
  ADD COLUMN IF NOT EXISTS report_generation_uuid TEXT;

CREATE INDEX IF NOT EXISTS idx_report_document_versions_source
  ON report_document_versions(source_table,source_key);
CREATE INDEX IF NOT EXISTS idx_report_document_versions_report
  ON report_document_versions(report_id)
  WHERE report_id IS NOT NULL;
CREATE UNIQUE INDEX IF NOT EXISTS uq_report_document_versions_current
  ON report_document_versions(source_table,source_key)
  WHERE is_current=TRUE;

CREATE TABLE IF NOT EXISTS report_document_migration(
  source_table TEXT NOT NULL,
  source_key TEXT NOT NULL,
  migration_status TEXT NOT NULL DEFAULT 'PENDING',
  classification TEXT,
  source_pdf_filename TEXT,
  source_pdf_size BIGINT,
  source_pdf_hash TEXT,
  backup_location TEXT,
  backup_hash TEXT,
  backup_verified BOOLEAN NOT NULL DEFAULT FALSE,
  snapshot_version INTEGER,
  snapshot_hash TEXT,
  render_verified BOOLEAN NOT NULL DEFAULT FALSE,
  validation_jsonb JSONB,
  deletion_batch_id TEXT,
  pdf_deleted_at TIMESTAMPTZ,
  attempts INTEGER NOT NULL DEFAULT 0,
  last_error TEXT,
  updated_at TIMESTAMPTZ NOT NULL DEFAULT NOW(),
  PRIMARY KEY(source_table,source_key),
  CONSTRAINT report_document_migration_source_check CHECK(
    source_table IN('report_history','daily_reports','billing_shift_closures')
  )
);
CREATE INDEX IF NOT EXISTS idx_report_document_migration_status
  ON report_document_migration(migration_status,source_table,source_key);

CREATE TABLE IF NOT EXISTS document_external_files(
  id BIGSERIAL PRIMARY KEY,
  filename TEXT NOT NULL,
  document_type TEXT NOT NULL,
  source_table TEXT,
  source_key TEXT,
  sha256 TEXT NOT NULL CHECK(sha256 ~ '^[0-9a-f]{64}$'),
  size_bytes BIGINT NOT NULL CHECK(size_bytes >= 0),
  archive_relative_path TEXT NOT NULL,
  archived_at TIMESTAMPTZ NOT NULL DEFAULT NOW(),
  verified_at TIMESTAMPTZ NOT NULL,
  status TEXT NOT NULL DEFAULT 'AVAILABLE',
  UNIQUE(filename,sha256)
);
CREATE INDEX IF NOT EXISTS idx_document_external_files_source
  ON document_external_files(source_table,source_key);
CREATE INDEX IF NOT EXISTS idx_document_external_files_filename
  ON document_external_files(filename);

CREATE TABLE IF NOT EXISTS database_capacity_history(
  id BIGSERIAL PRIMARY KEY,
  captured_at TIMESTAMPTZ NOT NULL DEFAULT NOW(),
  database_size_bytes BIGINT NOT NULL,
  database_limit_bytes BIGINT NOT NULL,
  usage_percent NUMERIC(8,3) NOT NULL,
  largest_relations_jsonb JSONB NOT NULL DEFAULT '[]'::jsonb,
  dead_rows_jsonb JSONB NOT NULL DEFAULT '[]'::jsonb,
  captured_by TEXT NOT NULL,
  notes TEXT,
  pdf_storage_bytes BIGINT NOT NULL DEFAULT 0,
  sync_events_bytes BIGINT NOT NULL DEFAULT 0,
  import_staging_bytes BIGINT NOT NULL DEFAULT 0,
  patient_events_bytes BIGINT NOT NULL DEFAULT 0
);
ALTER TABLE database_capacity_history
  ADD COLUMN IF NOT EXISTS pdf_storage_bytes BIGINT NOT NULL DEFAULT 0;
ALTER TABLE database_capacity_history
  ADD COLUMN IF NOT EXISTS sync_events_bytes BIGINT NOT NULL DEFAULT 0;
ALTER TABLE database_capacity_history
  ADD COLUMN IF NOT EXISTS import_staging_bytes BIGINT NOT NULL DEFAULT 0;
ALTER TABLE database_capacity_history
  ADD COLUMN IF NOT EXISTS patient_events_bytes BIGINT NOT NULL DEFAULT 0;
CREATE INDEX IF NOT EXISTS idx_database_capacity_history_captured
  ON database_capacity_history(captured_at DESC);

CREATE TABLE IF NOT EXISTS document_maintenance_config(
  config_key TEXT PRIMARY KEY,
  config_value_jsonb JSONB NOT NULL,
  updated_at TIMESTAMPTZ NOT NULL DEFAULT NOW(),
  updated_by TEXT NOT NULL
);
INSERT INTO document_maintenance_config(
  config_key,config_value_jsonb,updated_by
) VALUES
  ('closed_session_retention_days','90'::jsonb,'SYSTEM'),
  ('technical_log_retention_days','30'::jsonb,'SYSTEM'),
  ('capacity_warning_thresholds','{"warning":80,"review":85,"critical":90,"nonessential_block":95,"emergency":98}'::jsonb,'SYSTEM')
ON CONFLICT(config_key) DO NOTHING;

CREATE OR REPLACE FUNCTION enforce_report_document_version_immutable()
RETURNS TRIGGER
LANGUAGE plpgsql
AS $$
BEGIN
  IF TG_OP='DELETE' THEN
    RAISE EXCEPTION 'Las versiones documentales de reportes son inmutables';
  END IF;
  IF NEW.source_table IS DISTINCT FROM OLD.source_table
     OR NEW.source_key IS DISTINCT FROM OLD.source_key
     OR NEW.report_id IS DISTINCT FROM OLD.report_id
     OR NEW.version IS DISTINCT FROM OLD.version
     OR NEW.report_type IS DISTINCT FROM OLD.report_type
     OR NEW.report_title IS DISTINCT FROM OLD.report_title
     OR NEW.period_start IS DISTINCT FROM OLD.period_start
     OR NEW.period_end IS DISTINCT FROM OLD.period_end
     OR NEW.filters_jsonb IS DISTINCT FROM OLD.filters_jsonb
     OR NEW.financial_basis_jsonb IS DISTINCT FROM OLD.financial_basis_jsonb
     OR NEW.dataset_snapshot_jsonb IS DISTINCT FROM OLD.dataset_snapshot_jsonb
     OR NEW.summary_snapshot_jsonb IS DISTINCT FROM OLD.summary_snapshot_jsonb
     OR NEW.charts_snapshot_jsonb IS DISTINCT FROM OLD.charts_snapshot_jsonb
     OR NEW.guided_reading_snapshot_jsonb IS DISTINCT FROM OLD.guided_reading_snapshot_jsonb
     OR NEW.render_context_jsonb IS DISTINCT FROM OLD.render_context_jsonb
     OR NEW.snapshot_jsonb IS DISTINCT FROM OLD.snapshot_jsonb
     OR NEW.snapshot_hash IS DISTINCT FROM OLD.snapshot_hash
     OR NEW.template_version IS DISTINCT FROM OLD.template_version
     OR NEW.schema_version IS DISTINCT FROM OLD.schema_version
     OR NEW.created_at IS DISTINCT FROM OLD.created_at
     OR NEW.created_by IS DISTINCT FROM OLD.created_by
     OR NEW.generated_by_user_id IS DISTINCT FROM OLD.generated_by_user_id
     OR NEW.created_from_module IS DISTINCT FROM OLD.created_from_module
     OR NEW.report_uuid IS DISTINCT FROM OLD.report_uuid
     OR NEW.report_generation_uuid IS DISTINCT FROM OLD.report_generation_uuid THEN
    RAISE EXCEPTION 'El contenido de una versión documental de reporte no puede modificarse';
  END IF;
  RETURN NEW;
END;
$$;

DO $$
BEGIN
  IF NOT EXISTS(
    SELECT 1 FROM pg_trigger
    WHERE tgname='trg_report_document_version_immutable'
      AND tgrelid='report_document_versions'::regclass
      AND NOT tgisinternal
  ) THEN
    CREATE TRIGGER trg_report_document_version_immutable
    BEFORE UPDATE OR DELETE ON report_document_versions
    FOR EACH ROW EXECUTE FUNCTION enforce_report_document_version_immutable();
  END IF;
END $$;

ALTER TABLE report_document_versions ENABLE ROW LEVEL SECURITY;
ALTER TABLE report_document_migration ENABLE ROW LEVEL SECURITY;
ALTER TABLE document_external_files ENABLE ROW LEVEL SECURITY;
ALTER TABLE database_capacity_history ENABLE ROW LEVEL SECURITY;
ALTER TABLE document_maintenance_config ENABLE ROW LEVEL SECURITY;
"""


_render_lock = threading.Lock()


def apply_report_document_migration(connection) -> None:
    connection.executescript(REPORT_DOCUMENT_MIGRATION_SQL)


def _json_safe(value):
    if isinstance(value, dict):
        return {str(key): _json_safe(item) for key, item in value.items()}
    if isinstance(value, (list, tuple)):
        return [_json_safe(item) for item in value]
    if isinstance(value, Decimal):
        return float(value)
    if isinstance(value, (datetime, date)):
        return value.isoformat()
    if isinstance(value, bytes):
        return value.decode("utf-8", "replace")
    if value is None or isinstance(value, (str, int, float, bool)):
        return value
    return str(value)


def canonical_report_snapshot_json(snapshot: dict[str, Any]) -> str:
    return json.dumps(
        _json_safe(snapshot),
        ensure_ascii=False,
        sort_keys=True,
        separators=(",", ":"),
        allow_nan=False,
    )


def calculate_report_snapshot_hash(snapshot: dict[str, Any]) -> str:
    canonical = canonical_report_snapshot_json(snapshot)
    return hashlib.sha256(canonical.encode("utf-8")).hexdigest()


def source_key(source_table: str, record: dict[str, Any]) -> str:
    if source_table == "billing_shift_closures":
        return (
            f"{record.get('source_instance_id') or ''!s}|"
            f"{int(record.get('turn_id') or record.get('record_id') or 0)}"
        )
    return str(int(record.get("id") or record.get("record_id") or 0))


def build_report_snapshot(
    *,
    source_table: str,
    source_key_value: str,
    report_id: int | None,
    report_type: str,
    report_title: str,
    period_start: str | None,
    period_end: str | None,
    generated_at: str,
    generated_by: str,
    filters: dict | None,
    financial_basis: dict | None,
    dataset: dict | list | None,
    summary: dict | None,
    charts: dict | list | None,
    guided_reading: dict | list | str | None,
    render_context: dict,
    generated_by_user_id: str | int | None = None,
    created_from_module: str = "",
    report_generation_uuid: str = "",
) -> dict[str, Any]:
    context = _snapshot_render_context(render_context)
    render_contract = _snapshot_render_contract(render_context)
    snapshot = {
        "identity": {
            "source_table": source_table,
            "source_key": str(source_key_value),
            "report_id": int(report_id) if report_id else None,
            "report_type": str(report_type or "Reporte"),
            "report_title": str(report_title or report_type or "Reporte"),
            "report_uuid": str(
                uuid.uuid5(
                    uuid.NAMESPACE_URL,
                    f"hospital-report/{source_table}/{source_key_value}",
                )
            ),
            "generated_at": str(generated_at or ""),
            "generated_by": str(generated_by or ""),
            "generated_by_user_id": str(generated_by_user_id or ""),
            "created_from_module": str(created_from_module or ""),
            "report_generation_uuid": str(report_generation_uuid or ""),
        },
        "period": {
            "start": str(period_start or ""),
            "end": str(period_end or ""),
        },
        "filters": _json_safe(filters or {}),
        "financial_basis": _json_safe(financial_basis or {}),
        "dataset": _json_safe(dataset or {}),
        "summary": _json_safe(summary or {}),
        "charts": _json_safe(charts or {}),
        "guided_reading": _json_safe(guided_reading or {}),
        "render_context": context,
        "render_contract": render_contract,
        "template_version": REPORT_TEMPLATE_VERSION,
        "schema_version": REPORT_SNAPSHOT_SCHEMA_VERSION,
    }
    return snapshot


def _snapshot_render_context(render_context: dict | None) -> dict[str, Any]:
    """Persist presentation metadata without duplicating the report dataset."""
    context = _json_safe(dict(render_context or {}))
    for key in (
        "logo_path",
        "data",
        "totals",
        "category_rows",
        "ars_rows",
        "user_rows",
    ):
        context.pop(key, None)
    return context


def _snapshot_render_contract(render_context: dict | None) -> dict[str, Any]:
    """Keep only small, renderer-specific projections not derivable from data."""
    context = dict(render_context or {})
    return _json_safe(
        {
            key: context[key]
            for key in ("category_rows", "ars_rows", "user_rows")
            if key in context
        }
    )


def _snapshot_from_record(document_record: dict[str, Any]) -> dict[str, Any]:
    snapshot = document_record.get("snapshot") or document_record.get("snapshot_jsonb")
    if isinstance(snapshot, str):
        snapshot = json.loads(snapshot)
    if not isinstance(snapshot, dict):
        raise ReportSnapshotMissingError(
            "Los datos históricos de este reporte no están disponibles."
        )
    return snapshot


def _validate_snapshot(snapshot: dict[str, Any]) -> None:
    identity = snapshot.get("identity")
    if not isinstance(identity, dict) or not str(identity.get("source_table") or ""):
        raise ReportDocumentError("La instantánea histórica tiene un contrato inválido.")
    schema_version = int(snapshot.get("schema_version") or 1)
    if schema_version not in SUPPORTED_REPORT_SNAPSHOT_SCHEMA_VERSIONS:
        raise ReportTemplateError(
            "La versión de datos de este reporte todavía no es compatible."
        )


def _load_document_record(row, *, latest: bool) -> dict[str, Any]:
    if not row:
        message = (
            "El reporte no contiene una versión estructurada histórica."
            if latest
            else "El reporte no tiene una instantánea documental disponible."
        )
        raise ReportSnapshotMissingError(message)
    record = dict(row)
    snapshot = _snapshot_from_record(record)
    _validate_snapshot(snapshot)
    actual_hash = calculate_report_snapshot_hash(snapshot)
    if actual_hash != str(record["snapshot_hash"]):
        raise ReportSnapshotHashError(
            "La instantánea documental del reporte no superó la validación de integridad."
        )
    template_version = str(record.get("template_version") or snapshot.get("template_version") or "")
    if template_version != REPORT_TEMPLATE_VERSION:
        raise ReportTemplateError(
            "La versión histórica de la plantilla del reporte no está disponible."
        )
    record["snapshot"] = snapshot
    record["render_context"] = _render_context_from_snapshot(snapshot)
    return record


def _render_context_from_snapshot(snapshot: dict[str, Any]) -> dict[str, Any]:
    """Rebuild the renderer input solely from immutable snapshot fields."""
    context = deepcopy(dict(snapshot.get("render_context") or {}))
    dataset = deepcopy(snapshot.get("dataset") or {})
    contract = deepcopy(dict(snapshot.get("render_contract") or {}))
    identity = dict(snapshot.get("identity") or {})
    context.setdefault("title", identity.get("report_title") or "Reporte")
    context.setdefault("generated_by", identity.get("generated_by") or "Sistema")
    context.setdefault("generated_at", identity.get("generated_at") or "")
    mode = str(context.get("mode") or "standard")
    if mode in ("panel", "comparison", "shift_closure"):
        # Version 1 embedded a full render payload in context. Preserve it for
        # backward compatibility; v2+ reconstruct strictly from dataset.
        if int(snapshot.get("schema_version") or 1) >= 2 or not context.get("data"):
            context["data"] = dataset
    else:
        context["totals"] = dataset
        context.update({key: value for key, value in contract.items() if key not in context})
        context.setdefault("category_rows", [])
        context.setdefault("ars_rows", [])
        context.setdefault("user_rows", [])
    return context


def _source_update_sql(source_table: str) -> str:
    if source_table == "report_history":
        return (
            "UPDATE report_history SET document_storage_mode=%s,"
            "revision_version=%s WHERE id=%s"
        )
    if source_table == "daily_reports":
        return (
            "UPDATE daily_reports SET document_storage_mode=%s,"
            "revision_version=%s WHERE id=%s"
        )
    if source_table == "billing_shift_closures":
        return (
            "UPDATE billing_shift_closures SET document_storage_mode=%s,"
            "revision_version=%s WHERE source_instance_id=%s AND turn_id=%s"
        )
    raise ValueError("Fuente de reporte no permitida.")


def _source_update_params(
    source_table: str, source_key_value: str, mode: str, version: int
):
    if source_table in ("report_history", "daily_reports"):
        return (mode, int(version), int(source_key_value))
    instance_id, turn_id = parse_shift_closure_source_key(source_key_value)
    return (mode, int(version), instance_id, int(turn_id))


def save_report_document_snapshot(
    connection,
    *,
    source_table: str,
    source_key_value: str,
    report_id: int | None,
    report_type: str,
    report_title: str,
    period_start: str | None,
    period_end: str | None,
    generated_at: str,
    generated_by: str,
    filters: dict | None,
    financial_basis: dict | None,
    dataset: dict | list | None,
    summary: dict | None,
    charts: dict | list | None,
    guided_reading: dict | list | str | None,
    render_context: dict,
    storage_mode: str = STORAGE_SNAPSHOT,
    generated_by_user_id: str | int | None = None,
    created_from_module: str = "",
    report_generation_uuid: str = "",
) -> dict[str, Any]:
    snapshot = build_report_snapshot(
        source_table=source_table,
        source_key_value=source_key_value,
        report_id=report_id,
        report_type=report_type,
        report_title=report_title,
        period_start=period_start,
        period_end=period_end,
        generated_at=generated_at,
        generated_by=generated_by,
        filters=filters,
        financial_basis=financial_basis,
        dataset=dataset,
        summary=summary,
        charts=charts,
        guided_reading=guided_reading,
        render_context=render_context,
        generated_by_user_id=generated_by_user_id,
        created_from_module=created_from_module,
        report_generation_uuid=report_generation_uuid,
    )
    snapshot_hash = calculate_report_snapshot_hash(snapshot)
    current = connection.execute(
        """SELECT id,version,snapshot_hash
           FROM report_document_versions
           WHERE source_table=%s AND source_key=%s AND is_current=TRUE
           FOR UPDATE""",
        (source_table, str(source_key_value)),
    ).fetchone()
    if current and str(current["snapshot_hash"]) == snapshot_hash:
        mode = storage_mode if storage_mode in STORAGE_MODES else STORAGE_SNAPSHOT
        connection.execute(
            _source_update_sql(source_table),
            _source_update_params(
                source_table, source_key_value, mode, int(current["version"])
            ),
        )
        return {
            "id": int(current["id"]),
            "version": int(current["version"]),
            "snapshot_hash": snapshot_hash,
            "created": False,
            "snapshot": snapshot,
        }
    next_version = int(current["version"]) + 1 if current else 1
    if current:
        connection.execute(
            "UPDATE report_document_versions SET is_current=FALSE WHERE id=%s",
            (int(current["id"]),),
        )
    row = connection.execute(
        """INSERT INTO report_document_versions(
             source_table,source_key,report_id,version,report_type,report_title,
             period_start,period_end,filters_jsonb,financial_basis_jsonb,
             dataset_snapshot_jsonb,summary_snapshot_jsonb,charts_snapshot_jsonb,
             guided_reading_snapshot_jsonb,render_context_jsonb,snapshot_jsonb,snapshot_hash,
             template_version,schema_version,created_at,created_by,generated_by_user_id,
             created_from_module,report_uuid,report_generation_uuid,is_current
           ) VALUES(
             %s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,TRUE
           ) RETURNING id""",
        (
            source_table,
            str(source_key_value),
            int(report_id) if report_id else None,
            next_version,
            str(report_type or "Reporte"),
            str(report_title or report_type or "Reporte"),
            period_start or None,
            period_end or None,
            psycopg2.extras.Json(_json_safe(filters or {})),
            psycopg2.extras.Json(_json_safe(financial_basis or {})),
            psycopg2.extras.Json(_json_safe(dataset or {})),
            psycopg2.extras.Json(_json_safe(summary or {})),
            psycopg2.extras.Json(_json_safe(charts or {})),
            psycopg2.extras.Json(_json_safe(guided_reading or {})),
            psycopg2.extras.Json(_json_safe(render_context or {})),
            psycopg2.extras.Json(snapshot),
            snapshot_hash,
            REPORT_TEMPLATE_VERSION,
            REPORT_SNAPSHOT_SCHEMA_VERSION,
            str(generated_at or datetime.now(timezone.utc).isoformat()),
            str(generated_by or "SYSTEM"),
            str(generated_by_user_id or "") or None,
            str(created_from_module or "") or None,
            str((snapshot.get("identity") or {}).get("report_uuid") or "") or None,
            str(report_generation_uuid or "") or None,
        ),
    ).fetchone()
    mode = storage_mode if storage_mode in STORAGE_MODES else STORAGE_SNAPSHOT
    connection.execute(
        _source_update_sql(source_table),
        _source_update_params(source_table, source_key_value, mode, next_version),
    )
    return {
        "id": int(row["id"]),
        "version": next_version,
        "snapshot_hash": snapshot_hash,
        "created": True,
        "snapshot": snapshot,
    }


def load_current_report_snapshot(
    connection, source_table: str, source_key_value: str
) -> dict[str, Any]:
    row = connection.execute(
        """SELECT *
           FROM report_document_versions
           WHERE source_table=%s AND source_key=%s AND is_current=TRUE""",
        (source_table, str(source_key_value)),
    ).fetchone()
    return _load_document_record(row, latest=False)


def load_latest_report_snapshot(
    connection, source_table: str, source_key_value: str
) -> dict[str, Any]:
    """Read the newest valid snapshot even when a legacy current flag is lost."""
    row = connection.execute(
        """SELECT *
           FROM report_document_versions
           WHERE source_table=%s AND source_key=%s
           ORDER BY is_current DESC,version DESC
           LIMIT 1""",
        (source_table, str(source_key_value)),
    ).fetchone()
    return _load_document_record(row, latest=True)


def default_archive_root() -> Path:
    configured = str(os.getenv("DOCUMENT_ARCHIVE_ROOT") or "").strip()
    if configured:
        return Path(configured).expanduser().resolve()
    onedrive = str(os.getenv("OneDrive") or "").strip()
    if onedrive:
        return (Path(onedrive) / "HospitalDocumentArchive").resolve()
    return (Path.home() / "HospitalDocumentArchive").resolve()


def report_cache_root() -> Path:
    root = Path(tempfile.gettempdir()) / "hospital_document_cache" / "reports"
    root.mkdir(parents=True, exist_ok=True)
    return root


def _safe_name(value: str) -> str:
    return re.sub(r"[^A-Za-z0-9_.-]+", "_", str(value or "report"))[:120]


def render_report_snapshot_pdf(
    document_record: dict[str, Any], logo_path: str = ""
) -> str:
    snapshot = _snapshot_from_record(document_record)
    _validate_snapshot(snapshot)
    context = _render_context_from_snapshot(snapshot)
    context["logo_path"] = logo_path or ""
    source = _safe_name(
        f"{document_record.get('source_table')}_{document_record.get('source_key')}"
    )
    version = int(document_record.get("version") or 1)
    short_hash = str(document_record.get("snapshot_hash") or "")[:12]
    output_path = report_cache_root() / f"{source}_v{version}_{short_hash}.pdf"
    with _render_lock:
        if output_path.is_file() and output_path.stat().st_size > 0:
            return str(output_path)
        temporary = output_path.with_name(
            f".{output_path.stem}_{os.getpid()}_{threading.get_ident()}.tmp.pdf"
        )
        try:
            ReportHTMLRenderer().render_pdf(
                context,
                str(temporary),
                landscape=bool(context.get("landscape", False)),
            )
            if not temporary.is_file() or temporary.stat().st_size <= 0:
                raise ReportDocumentError(
                    "El motor documental no generó un reporte válido."
                )
            os.replace(str(temporary), str(output_path))
        finally:
            if temporary.exists():
                temporary.unlink(missing_ok=True)
    return str(output_path)


def _spreadsheet_value(value: Any) -> Any:
    if isinstance(value, str) and value[:1] in ("=", "+", "-", "@"):
        return "'" + value
    return value


def _flatten_snapshot_values(value: Any, prefix: str = "") -> list[tuple[str, Any]]:
    rows: list[tuple[str, Any]] = []
    if isinstance(value, dict):
        for key in sorted(value, key=lambda item: str(item)):
            label = f"{prefix} / {key}" if prefix else str(key)
            rows.extend(_flatten_snapshot_values(value[key], label))
    elif isinstance(value, list):
        for index, item in enumerate(value, start=1):
            label = f"{prefix} #{index}" if prefix else f"Registro #{index}"
            rows.extend(_flatten_snapshot_values(item, label))
    else:
        rows.append((prefix or "Valor", _spreadsheet_value(value)))
    return rows


def export_report_snapshot_xlsx(
    document_record: dict[str, Any], output_path: str, logo_path: str = ""
) -> str:
    """Materialize a historical Excel file from the immutable report snapshot."""
    snapshot = _snapshot_from_record(document_record)
    _validate_snapshot(snapshot)
    context = _render_context_from_snapshot(snapshot)
    dataset = deepcopy(snapshot.get("dataset") or {})
    output_path = os.path.abspath(str(output_path))
    os.makedirs(os.path.dirname(output_path), exist_ok=True)
    if str(context.get("mode") or "") == "panel":
        from report_engine.excel_exporter import export_panel_xlsx

        return export_panel_xlsx(
            dataset,
            output_path,
            str((snapshot.get("identity") or {}).get("generated_by") or "Sistema"),
            logo_path,
        )
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
    except ImportError as exc:
        raise ReportDocumentError(
            "La exportación Excel requiere el componente openpyxl."
        ) from exc

    identity = dict(snapshot.get("identity") or {})
    period = dict(snapshot.get("period") or {})
    workbook = Workbook()
    summary_sheet = workbook.active
    summary_sheet.title = "Resumen"
    data_sheet = workbook.create_sheet("Datos históricos")
    filters_sheet = workbook.create_sheet("Filtros")
    for sheet in (summary_sheet, data_sheet, filters_sheet):
        sheet.sheet_view.showGridLines = False
        sheet.freeze_panes = "A5"
        sheet.column_dimensions["A"].width = 34
        sheet.column_dimensions["B"].width = 72
    header_fill = PatternFill("solid", fgColor="123F83")
    title_font = Font(name="Aptos Display", size=16, bold=True, color="FFFFFF")
    section_font = Font(bold=True, color="123F83")
    summary_sheet.merge_cells("A1:B1")
    title = summary_sheet["A1"]
    title.value = _spreadsheet_value(identity.get("report_title") or "Reporte histórico")
    title.fill = header_fill
    title.font = title_font
    title.alignment = Alignment(horizontal="center")
    metadata = (
        ("Tipo", identity.get("report_type") or ""),
        ("Período", f"{period.get('start') or '—'} al {period.get('end') or '—'}"),
        ("Generado", identity.get("generated_at") or ""),
        ("Usuario", identity.get("generated_by") or ""),
        ("Versión de datos", snapshot.get("schema_version") or 1),
        ("Hash", document_record.get("snapshot_hash") or calculate_report_snapshot_hash(snapshot)),
    )
    for row_number, (label, value) in enumerate(metadata, start=3):
        summary_sheet.cell(row_number, 1, label).font = section_font
        summary_sheet.cell(row_number, 2, _spreadsheet_value(value))
    summary_sheet.cell(10, 1, "Resumen histórico").font = section_font
    for row_number, (label, value) in enumerate(
        _flatten_snapshot_values(snapshot.get("summary") or {}), start=11
    ):
        summary_sheet.cell(row_number, 1, label)
        summary_sheet.cell(row_number, 2, value)

    data_sheet.append(["Campo", "Valor"])
    for cell in data_sheet[1]:
        cell.fill = header_fill
        cell.font = Font(bold=True, color="FFFFFF")
    for label, value in _flatten_snapshot_values(dataset):
        data_sheet.append([label, value])

    filters_sheet.append(["Filtro", "Valor"])
    for cell in filters_sheet[1]:
        cell.fill = header_fill
        cell.font = Font(bold=True, color="FFFFFF")
    filters_payload = {
        "filters": snapshot.get("filters") or {},
        "financial_basis": snapshot.get("financial_basis") or {},
        "period": period,
    }
    for label, value in _flatten_snapshot_values(filters_payload):
        filters_sheet.append([label, value])
    for sheet in (summary_sheet, data_sheet, filters_sheet):
        for row in sheet.iter_rows():
            for cell in row:
                cell.alignment = Alignment(vertical="top", wrap_text=True)
    workbook.save(output_path)
    return output_path


def find_external_document(
    connection,
    *,
    filename: str,
    source_table: str = "",
    source_key_value: str = "",
    archive_root: Path | None = None,
) -> Path:
    root = archive_root or default_archive_root()
    row = None
    if source_table and source_key_value:
        row = connection.execute(
            """SELECT archive_relative_path,sha256,size_bytes
               FROM document_external_files
               WHERE source_table=%s AND source_key=%s AND filename=%s
                 AND status='AVAILABLE'
               ORDER BY verified_at DESC LIMIT 1""",
            (source_table, str(source_key_value), os.path.basename(filename)),
        ).fetchone()
    if not row:
        row = connection.execute(
            """SELECT archive_relative_path,sha256,size_bytes
               FROM document_external_files
               WHERE filename=%s AND status='AVAILABLE'
               ORDER BY verified_at DESC LIMIT 1""",
            (os.path.basename(filename),),
        ).fetchone()
    if not row:
        raise ReportDocumentError(
            "El documento histórico no tiene un respaldo externo registrado."
        )
    candidate = (root / str(row["archive_relative_path"])).resolve()
    if not candidate.is_file():
        raise ReportDocumentError(
            "No se encontró el respaldo externo. Configure DOCUMENT_ARCHIVE_ROOT "
            f"con la carpeta que contiene: {row['archive_relative_path']}"
        )
    if candidate.stat().st_size != int(row["size_bytes"]):
        raise ReportDocumentError(
            "El respaldo externo no coincide con el tamaño registrado."
        )
    digest = hashlib.sha256(candidate.read_bytes()).hexdigest()
    if digest != str(row["sha256"]):
        raise ReportDocumentError(
            "El respaldo externo no superó la validación de integridad."
        )
    return candidate
