"""Generación del expediente mensual de Emergencias para ARS privadas.

El archivo contiene dos documentos imprimibles y vinculados:

* Relación Emergencias: detalle paciente por paciente.
* Factura Global: resumen fiscal del lote completo.

Los datos clínicos/originales no se modifican aquí. El exportador recibe las
fotografías auditables guardadas en el lote de Facturación.
"""

from __future__ import annotations

import re
from datetime import date, datetime
from pathlib import Path
from typing import Any, Iterable

from openpyxl import Workbook, load_workbook
from openpyxl.drawing.image import Image as ExcelImage
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


BLUE = "075985"
CYAN = "00A6CE"
LIGHT_BLUE = "DDEBF7"
PALE_BLUE = "EAF4FA"
LIGHT_GRAY = "D9E2F3"
INPUT_YELLOW = "FFF2CC"
WHITE = "FFFFFF"
BLACK = "000000"
RED = "D71920"

THIN_BLACK = Side(style="thin", color=BLACK)
MEDIUM_BLUE = Side(style="medium", color=BLUE)


def _text(value: Any) -> str:
    return str(value or "").strip()


def _money(value: Any) -> float:
    try:
        return float(value or 0)
    except (TypeError, ValueError):
        return 0.0


def _date_value(value: Any) -> date | None:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    text = _text(value)[:10]
    for pattern in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y"):
        try:
            return datetime.strptime(text, pattern).date()
        except ValueError:
            continue
    return None


def long_date_es(value: Any) -> str:
    parsed = _date_value(value)
    if parsed is None:
        return _text(value)
    days = (
        "lunes", "martes", "miércoles", "jueves",
        "viernes", "sábado", "domingo",
    )
    months = (
        "enero", "febrero", "marzo", "abril", "mayo", "junio",
        "julio", "agosto", "septiembre", "octubre", "noviembre", "diciembre",
    )
    return (
        f"{days[parsed.weekday()]}, {parsed.day} de "
        f"{months[parsed.month - 1]} de {parsed.year}"
    )


def safe_export_filename(ars: str, year: int, month: int) -> str:
    normalized = re.sub(r"[^A-Z0-9]+", "_", _text(ars).upper()).strip("_")
    return f"LISTADO_EMERGENCIAS_{normalized or 'ARS'}_{year:04d}_{month:02d}.xlsx"


def validate_export_payload(
    batch: dict[str, Any],
    receipts: Iterable[dict[str, Any]],
) -> list[str]:
    """Devuelve problemas que deben resolverse antes de emitir el expediente."""
    rows = list(receipts)
    problems: list[str] = []
    required_batch = (
        ("invoice_date", "Fecha de factura"),
        ("invoice_number", "Número de factura"),
        ("ncf", "NCF"),
        ("provider_code", "Código de prestador"),
        ("provider_name", "Nombre del prestador"),
        ("provider_rnc", "RNC del centro"),
        ("ars_rnc", "RNC de la ARS"),
        ("ars_address", "Dirección de la ARS"),
        ("director_name", "Nombre de la directora"),
    )
    for key, label in required_batch:
        if not _text(batch.get(key)):
            problems.append(f"Falta {label}.")
    if not rows:
        problems.append("El listado no contiene pacientes.")
        return problems

    for index, row in enumerate(rows, start=1):
        prefix = f"Fila {index}"
        if row.get("estado_facturacion") and str(row.get("estado_facturacion")).upper() != "FACTURADO":
            problems.append(f"{prefix}: la atención todavía no está facturada.")
        if not _text(row.get("patient_snapshot") or row.get("nombre")):
            problems.append(f"{prefix}: falta el nombre del paciente.")
        if not _text(row.get("document_number_snapshot")):
            problems.append(f"{prefix}: falta NSS o cédula.")
        if not _date_value(row.get("service_date_snapshot")):
            problems.append(f"{prefix}: falta una fecha de servicio válida.")
        if not _text(row.get("authorization_snapshot")):
            problems.append(f"{prefix}: falta el número de autorización.")
        if _money(row.get("total_snapshot")) <= 0:
            problems.append(f"{prefix}: el valor reclamado debe ser mayor que cero.")
    return problems


def _set_border(range_cells, border: Border) -> None:
    for row in range_cells:
        for cell in row:
            cell.border = border


def _add_logo(
    sheet,
    logo_path: str | Path | None,
    anchor: str,
    width: int,
    *,
    max_height: int | None = None,
) -> None:
    if not logo_path:
        return
    path = Path(logo_path)
    if not path.is_file():
        return
    excel_png = path.with_name("logo_excel.png")
    if excel_png.is_file():
        path = excel_png
    image = ExcelImage(str(path))
    ratio = float(image.height or 1) / float(image.width or 1)
    image.width = width
    image.height = max(45, int(width * ratio))
    if max_height and image.height > max_height:
        image.height = max_height
        image.width = max(1, int(max_height / ratio))
    sheet.add_image(image, anchor)


def _short_date(value: Any) -> str:
    parsed = _date_value(value)
    return parsed.strftime("%d-%m-%Y") if parsed else _text(value)


def _configure_page(sheet, *, landscape: bool) -> None:
    sheet.sheet_view.showGridLines = False
    sheet.freeze_panes = "A11" if landscape else None
    sheet.page_setup.orientation = "landscape" if landscape else "portrait"
    sheet.page_setup.paperSize = sheet.PAPERSIZE_LETTER
    sheet.page_setup.fitToWidth = 1
    sheet.page_setup.fitToHeight = 0
    sheet.sheet_properties.pageSetUpPr.fitToPage = True
    sheet.page_margins.left = 0.25
    sheet.page_margins.right = 0.25
    sheet.page_margins.top = 0.35
    sheet.page_margins.bottom = 0.35
    sheet.oddFooter.center.text = (
        "Documento generado por el Sistema de Facturación Médica"
    )
    sheet.oddFooter.center.size = 8


def _write_relation_sheet(
    sheet,
    batch: dict[str, Any],
    receipts: list[dict[str, Any]],
    logo_path: str | Path | None,
) -> tuple[int, int]:
    _configure_page(sheet, landscape=True)
    sheet.title = "Relación Emergencias"
    sheet.freeze_panes = "A15"
    sheet.print_title_rows = "1:14"
    for index, width in enumerate(
        (8.57, 28, 94.43, 23.29, 40.14, 25.43, 49.43, 35.29), start=1
    ):
        sheet.column_dimensions[get_column_letter(index)].width = width
    for row in range(1, 7):
        sheet.row_dimensions[row].height = 18
    _add_logo(sheet, logo_path, "A1", 260, max_height=105)

    sheet.merge_cells("F7:H12")
    sheet["F7"] = (
        "RELACIÓN DE EMERGENCIAS\n"
        f"PACIENTES ARS {_text(batch.get('ars_display_name') or batch.get('ars')).upper()}"
    )
    sheet["F7"].font = Font(name="Arial", size=12, bold=True)
    sheet["F7"].alignment = Alignment(
        horizontal="center", vertical="center", wrap_text=True
    )

    metadata = (
        ("Fecha Facturación", _short_date(batch.get("invoice_date"))),
        ("Código Prestador", _text(batch.get("provider_code"))),
        ("Nombre Prestador", _text(batch.get("provider_name")).upper()),
        ("RNC Centro", _text(batch.get("provider_rnc"))),
        ("No. Factura", _text(batch.get("invoice_number"))),
        ("NCF", _text(batch.get("ncf"))),
    )
    for row, (label, value) in enumerate(metadata, start=7):
        sheet.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
        sheet.cell(row, 1, label)
        sheet.cell(row, 3, value)
        sheet.cell(row, 1).font = Font(name="Arial", bold=True, size=10)
        sheet.cell(row, 3).font = Font(name="Arial", size=10)
        sheet.cell(row, 1).alignment = Alignment(vertical="center")
        sheet.cell(row, 3).alignment = Alignment(vertical="center")
        sheet.row_dimensions[row].height = 19.5 if row in (7, 11) else 18
    sheet.merge_cells("D12:E12")
    sheet["D12"] = "Valores En RD$"
    sheet["D12"].font = Font(name="Arial", bold=True, size=10)
    sheet["D12"].alignment = Alignment(horizontal="center", vertical="center")

    headers = (
        "NO.", "TIPO DOC.\nNSS / CÉDULA", "NOMBRE DEL PACIENTE", "FECHA",
        "NO. DE AUTORIZACIÓN", "VALOR RECLAMADO", "ESPECIALIDAD MÉDICA",
    )
    header_row = 14
    sheet.row_dimensions[13].height = 21
    sheet.row_dimensions[header_row].height = 39
    for column, header in enumerate(headers, start=1):
        cell = sheet.cell(header_row, column, header)
        cell.font = Font(name="Arial", bold=True, size=10)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = Border(
            left=THIN_BLACK, right=THIN_BLACK, top=THIN_BLACK, bottom=THIN_BLACK
        )

    first_data_row = header_row + 1
    for index, data in enumerate(receipts, start=1):
        row = first_data_row + index - 1
        values = (
            index,
            f"{_text(data.get('document_type_snapshot') or 'NSS').upper()}\n"
            f"{_text(data.get('document_number_snapshot'))}",
            _text(data.get("patient_snapshot") or data.get("nombre")).upper(),
            _date_value(data.get("service_date_snapshot")),
            _text(data.get("authorization_snapshot")),
            _money(data.get("total_snapshot")),
            _text(
                data.get("specialty_snapshot")
                or batch.get("specialty_default")
                or "EMERGENCIOLOGÍA"
            ).upper(),
        )
        for column, value in enumerate(values, start=1):
            cell = sheet.cell(row, column, value)
            cell.font = Font(name="Arial", size=10)
            cell.border = Border(
                left=THIN_BLACK, right=THIN_BLACK, top=THIN_BLACK, bottom=THIN_BLACK
            )
            cell.alignment = Alignment(
                horizontal="left" if column in (2, 3, 7) else "center",
                vertical="center", wrap_text=True,
            )
        sheet.cell(row, 4).number_format = "dd-mm-yyyy"
        sheet.cell(row, 5).font = Font(name="Arial", size=10, color=RED)
        sheet.cell(row, 6).number_format = '"RD$" #,##0.00'
        sheet.row_dimensions[row].height = 28.5

    last_data_row = first_data_row + len(receipts) - 1
    total_row = last_data_row + 2
    sheet.row_dimensions[total_row - 1].height = 20
    sheet.row_dimensions[total_row].height = 28.5
    sheet.cell(total_row, 5, "TOTAL")
    sheet.cell(total_row, 6, f"=SUM(F{first_data_row}:F{last_data_row})")
    for column in range(1, 8):
        cell = sheet.cell(total_row, column)
        cell.font = Font(
            name="Arial", bold=True, color=RED if column == 5 else BLACK
        )
        cell.border = Border(
            left=THIN_BLACK, right=THIN_BLACK, top=THIN_BLACK, bottom=THIN_BLACK
        )
        cell.alignment = Alignment(horizontal="center", vertical="center")
    sheet.cell(total_row, 6).number_format = '"RD$" #,##0.00'
    sheet.auto_filter.ref = f"A{header_row}:G{last_data_row}"
    sheet.print_area = f"A1:H{total_row}"
    return first_data_row, total_row


def _write_global_invoice_sheet(
    sheet,
    batch: dict[str, Any],
    receipt_count: int,
    relation_total_row: int,
    logo_path: str | Path | None,
) -> None:
    _configure_page(sheet, landscape=False)
    sheet.title = "Factura Global"
    for index, width in enumerate((38.71, 41.29, 53, 22, 22, 22), start=1):
        sheet.column_dimensions[get_column_letter(index)].width = width
    for row in range(1, 34):
        sheet.row_dimensions[row].height = 23.25
    _add_logo(sheet, logo_path, "A1", 350, max_height=145)

    sheet.merge_cells("A6:C6")
    sheet["A6"] = _text(batch.get("provider_name")).upper()
    sheet["A7"] = f"RNC CENTRO: {_text(batch.get('provider_rnc'))}"
    sheet["A9"] = f"CÓDIGO: {_text(batch.get('provider_code'))}"
    sheet["A10"] = f"FECHA: {_short_date(batch.get('invoice_date'))}"
    sheet.merge_cells("D7:F7")
    sheet["D7"] = "NÚMERO DE COMPROBANTE FISCAL"
    sheet.merge_cells("D8:F8")
    sheet["D8"] = f"NCF: {_text(batch.get('ncf'))}"
    sheet.merge_cells("D9:F9")
    sheet["D9"] = f"NO. FACTURA: {_text(batch.get('invoice_number'))}"
    sheet.merge_cells("D10:F10")
    expiration = _short_date(batch.get("ncf_expiration_date"))
    sheet["D10"] = f"FECHA VENCIMIENTO NCF: {expiration or '—'}"
    for cell_ref in ("A6", "A7", "A9", "A10", "D7", "D8", "D9", "D10"):
        sheet[cell_ref].font = Font(name="Arial", bold=True, size=11)
        sheet[cell_ref].alignment = Alignment(vertical="center", wrap_text=True)
    sheet.row_dimensions[6].height = 34
    sheet["A10"].fill = PatternFill("solid", fgColor="FFFF00")

    sheet.merge_cells("A12:B12")
    sheet["A12"] = f"ARS {_text(batch.get('ars_display_name') or batch.get('ars')).upper()}"
    sheet["A12"].font = Font(name="Arial", size=14, bold=True)
    sheet["A13"] = f"RNC: {_text(batch.get('ars_rnc'))}"
    sheet["A14"] = _text(batch.get("ars_address")).upper()
    sheet["A14"].font = Font(name="Arial", color=BLUE, size=11)
    sheet["A14"].alignment = Alignment(vertical="top", wrap_text=True)
    sheet.row_dimensions[14].height = 72.75

    for column, value in enumerate(("CANTIDAD", "DESCRIPCIÓN", "VALOR"), start=1):
        cell = sheet.cell(16, column, value)
        cell.font = Font(name="Arial", bold=True)
        cell.fill = PatternFill("solid", fgColor=LIGHT_GRAY)
        cell.alignment = Alignment(horizontal="center", vertical="center")
    sheet["A17"] = receipt_count
    sheet["B17"] = _text(batch.get("service_description") or "EMERGENCIA").upper()
    sheet["C17"] = f"='Relación Emergencias'!F{relation_total_row}"
    sheet["C17"].number_format = '"RD$" #,##0.00'
    for cell_ref in ("A17", "B17", "C17"):
        sheet[cell_ref].alignment = Alignment(horizontal="center", vertical="center")

    summary = (
        (19, "SUB-TOTAL:", f"='Relación Emergencias'!F{relation_total_row}"),
        (20, "DESC:", _money(batch.get("discount"))),
        (21, "ITBIS:", _money(batch.get("itbis"))),
        (22, "TOTAL GENERAL", "=C19-C20+C21"),
    )
    for row, label, value in summary:
        sheet.cell(row, 2, label)
        sheet.cell(row, 3, value)
        sheet.cell(row, 2).font = Font(name="Arial", bold=True)
        sheet.cell(row, 3).font = Font(name="Arial", bold=row in (19, 22))
        sheet.cell(row, 3).number_format = '"RD$" #,##0.00'
        fill = LIGHT_GRAY if row in (20, 22) else WHITE
        for column in range(2, 4):
            sheet.cell(row, column).fill = PatternFill("solid", fgColor=fill)
            sheet.cell(row, column).alignment = Alignment(horizontal="center")

    sheet["B30"] = "____________________________________________"
    sheet["B30"].alignment = Alignment(horizontal="center")
    sheet.merge_cells("A31:C31")
    sheet["A31"] = _text(batch.get("director_name")).upper()
    sheet["B32"] = _text(batch.get("director_title") or "DIRECTORA").upper()
    sheet.merge_cells("A33:C33")
    sheet["A33"] = _text(batch.get("provider_name")).upper()
    for cell_ref in ("A31", "B32", "A33"):
        sheet[cell_ref].font = Font(name="Arial", bold=True, size=11)
        sheet[cell_ref].alignment = Alignment(horizontal="center")

    _set_border(
        sheet["A16:C17"],
        Border(left=THIN_BLACK, right=THIN_BLACK, top=THIN_BLACK, bottom=THIN_BLACK),
    )
    _set_border(
        sheet["B19:C22"],
        Border(left=THIN_BLACK, right=THIN_BLACK, top=THIN_BLACK, bottom=THIN_BLACK),
    )
    sheet.print_area = "A1:F33"


def create_private_ars_workbook(
    output_path: str | Path,
    batch: dict[str, Any],
    receipts: Iterable[dict[str, Any]],
    *,
    logo_path: str | Path | None = None,
    allow_incomplete: bool = False,
) -> Path:
    rows = list(receipts)
    problems = validate_export_payload(batch, rows)
    if problems and not allow_incomplete:
        raise ValueError("\n".join(problems))

    destination = Path(output_path)
    destination.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    relation = workbook.active
    first_data_row, relation_total_row = _write_relation_sheet(
        relation, batch, rows, logo_path
    )
    invoice = workbook.create_sheet("Factura Global")
    _write_global_invoice_sheet(
        invoice,
        batch,
        len(rows),
        relation_total_row,
        logo_path,
    )
    workbook.calculation.fullCalcOnLoad = True
    workbook.calculation.forceFullCalc = True
    workbook.calculation.calcMode = "auto"
    workbook.save(destination)

    # Apertura inmediata para comprobar que el contenedor XLSX no quedó corrupto.
    check = load_workbook(destination, read_only=False, data_only=False)
    try:
        expected = {"Relación Emergencias", "Factura Global"}
        if set(check.sheetnames) != expected:
            raise ValueError("El expediente exportado no contiene las dos hojas esperadas.")
        if check["Relación Emergencias"].max_row < first_data_row:
            raise ValueError("La relación exportada no contiene pacientes.")
    finally:
        check.close()
    return destination
