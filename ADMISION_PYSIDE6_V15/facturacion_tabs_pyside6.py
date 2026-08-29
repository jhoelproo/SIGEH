# facturacion_tabs.py
# Sistema de Formularios de Emergencia - Hospital General
# Version: 4.1.9-Qt-V10 - PySide6 con paridad visual validada del inicio
# Python 3.14 compatible
import os
import re
import sys
import json
import io
import sqlite3
from difflib import SequenceMatcher, get_close_matches
from contextlib import closing
from datetime import datetime, timedelta, date, time
import platform
import subprocess
import tempfile
import shutil
import zipfile
import zlib
import hashlib
import ctypes
import time as _time
import threading
import logging
from dataclasses import dataclass
from types import MappingProxyType
V15_SOURCE_BUILD_ID = "20260821_ADMIN_REP_FORCE_V1"
from logging.handlers import RotatingFileHandler

# Preparación del entorno del proyecto antes de importar emergency_core.
# No altera la lógica de negocio; únicamente localiza la raíz existente del sistema.
try:
    from .project_bootstrap import bootstrap_project_root
except ImportError:  # Ejecución directa del entrypoint standalone.
    from project_bootstrap import bootstrap_project_root
PROJECT_ROOT = bootstrap_project_root()

from sqlite_write_coordinator import connect_local_sqlite, prepare_sqlite_database
from admission_refresh_coordinator import CoalescedRefreshGate, history_rows_fingerprint
from display_layout import (
    DENSITY_AUTO,
    PROFILE_AUTO,
    resolve_admission_layout_profile,
)

SELF_TEST_MODE = "--self-test" in sys.argv
OFFLINE_MODE = str(os.environ.get("HOSPITAL_OFFLINE", "")).strip().lower() in {
    "1", "true", "yes", "si", "sí",
}
SELF_TEST_DATA_DIR = ""
if SELF_TEST_MODE:
    SELF_TEST_DATA_DIR = tempfile.mkdtemp(prefix="generador_hojas_selftest_")
    os.environ["EMERGENCIAS_DATA_DIR"] = SELF_TEST_DATA_DIR

from emergency_core.backup import BackupManager
from emergency_core.db_migrations import LATEST_SCHEMA_VERSION, migrate_database
from emergency_core.integration_events import (
    build_attention_event_ref,
    build_shift_event_ref,
    enqueue_billing_event,
    enqueue_shift_closure_event,
)
from emergency_core.main_app_gateway import MainAppGateway, MainAppGatewayError
from emergency_core.io_utils import ConfigError, atomic_write_json, load_json_file
from emergency_core.paths import data_root, harden_windows_acl, migrate_legacy_files
from emergency_core.security import AdminSecurity, SecurityError
from emergency_core.session_context import (
    CAP_EDIT_RECORDS,
    CAP_INTERNAL_CONFIG,
    CAP_OPEN_EXCEL,
    CAP_VIEW_REPORTS,
    CAP_VOID_RECORDS,
    ROLE_ADMIN,
    load_session_context,
    normalize_role,
)
from emergency_core.single_instance import SingleInstanceGuard

# PDF / Excel
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import letter
from PyPDF2 import PdfReader, PdfWriter
from PyPDF2.generic import ArrayObject, DecodedStreamObject, DictionaryObject, NameObject
import openpyxl
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font as XLFont, Alignment as XLAlignment, Border as XLBorder, Side as XLSide

from admission_statistical_reports import (  # noqa: E402 - PROJECT_ROOT bootstrap requerido
    ARS_ALL,
    ARS_EXCLUDE,
    ARS_INCLUDE,
    COVERAGE_ALL,
    COVERAGE_INSURED,
    COVERAGE_UNINSURED,
    SPECIALTY_ALL,
    AdmissionReportFilters,
    ReportSnapshotStore,
    SnapshotStaleError,
    build_admission_report_dataset,
    build_operational_period,
    build_turn_operational_period,
    search_ars_catalog,
)


# UI - MIGRADA A PySide6
# La lógica funcional permanece en este archivo; qt_compat sustituye únicamente
# el subconjunto de Tk/ttkbootstrap utilizado por la aplicación sobre widgets Qt.
try:
    from . import qt_compat
    from .qt_compat import (
        tk, ttk, tb, messagebox, Toplevel, filedialog, simpledialog, TBDateEntry,
        PRIMARY, SECONDARY, SUCCESS, INFO, WARNING, DANGER, LIGHT, DARK,
        create_standalone_application,
    )
except ImportError:  # Ejecución directa del entrypoint standalone.
    import qt_compat
    from qt_compat import (
        tk, ttk, tb, messagebox, Toplevel, filedialog, simpledialog, TBDateEntry,
        PRIMARY, SECONDARY, SUCCESS, INFO, WARNING, DANGER, LIGHT, DARK,
        create_standalone_application,
    )
from PySide6.QtCore import Qt, QSize, QSignalBlocker, QTimer  # type: ignore
from PySide6.QtWidgets import QSizePolicy  # type: ignore
from PySide6.QtGui import QColor, QFont, QIcon, QImage, QPainter, QPixmap  # type: ignore


@dataclass(frozen=True)
class AttentionOutputResult:
    """One rendered immediate sheet shared by its open and print dispatches."""

    attention_id: int
    pdf_path: str
    render_ms: float
    flow_started_at: float
    should_open: bool
    should_print: bool

try:
    from .admission_context import AdmissionContext, create_standalone_context
except ImportError:  # Ejecución directa del entrypoint standalone.
    from admission_context import AdmissionContext, create_standalone_context


# ============================================================
# FASE 13 - Colores institucionales sobrios
# ============================================================
COLOR_PRIMARY = "#2563EB"
COLOR_SUCCESS = "#1F7A4D"
COLOR_WARNING = "#A16207"
COLOR_DANGER  = "#B42318"
COLOR_INFO    = "#0E7490"


# -------------------------------
# HELPERS DE RUTAS PARA .PY Y .EXE
# -------------------------------
def resource_path(relative_path):
    relative_path = str(relative_path or "").lstrip("/\\")
    module_root = os.path.dirname(os.path.abspath(__file__))
    bundle_root = getattr(sys, "_MEIPASS", "")
    candidates = []
    if bundle_root:
        candidates.extend([
            os.path.join(bundle_root, "ADMISION_PYSIDE6_V15", relative_path),
            os.path.join(bundle_root, relative_path),
        ])
    candidates.append(os.path.join(module_root, relative_path))
    return next((path for path in candidates if os.path.exists(path)), candidates[-1])


_THEMED_ICON_CACHE = {}


def _theme_icon_device_pixel_ratio() -> float:
    """Use the active screen DPR without depending on an already shown window."""
    from PySide6.QtWidgets import QApplication

    application = QApplication.instance()
    screen = application.primaryScreen() if application is not None else None
    return max(1.0, float(screen.devicePixelRatio() if screen is not None else 1.0))


def _render_themed_svg_icon(path, foreground, size, device_pixel_ratio):
    """Rasterize an SVG only after its semantic foreground is known."""
    from PySide6.QtSvg import QSvgRenderer

    renderer = QSvgRenderer(str(path))
    if not renderer.isValid():
        return QPixmap()
    pixel_size = max(1, int(round(int(size) * float(device_pixel_ratio))))
    image = QImage(
        pixel_size,
        pixel_size,
        QImage.Format.Format_ARGB32_Premultiplied,
    )
    image.fill(Qt.transparent)
    painter = QPainter(image)
    renderer.render(painter)
    painter.setCompositionMode(QPainter.CompositionMode_SourceIn)
    painter.fillRect(image.rect(), QColor(str(foreground)))
    painter.end()
    pixmap = QPixmap.fromImage(image)
    pixmap.setDevicePixelRatio(float(device_pixel_ratio))
    return pixmap


def theme_icon(
    icon_name,
    foreground,
    *,
    size=18,
    theme_mode="",
    role="",
    device_pixel_ratio=None,
):
    """Return a readable, DPR-aware semantic icon for the current theme.

    SVGs in V15 are intentionally monochrome source masks.  The cache key
    keeps every output dimension that can alter a rendered asset, preventing a
    cold-start light icon from reusing a pixmap made for another role, mode or
    scale factor.
    """
    path = resource_path(os.path.join("assets", str(icon_name or "")))
    dpr = (
        _theme_icon_device_pixel_ratio()
        if device_pixel_ratio is None
        else max(1.0, float(device_pixel_ratio))
    )
    key = (
        os.path.normcase(os.path.abspath(path)),
        str(theme_mode or ""),
        str(role or ""),
        str(foreground),
        int(size),
        round(dpr, 3),
    )
    cached = _THEMED_ICON_CACHE.get(key)
    if cached is not None:
        return cached
    pixmap = _render_themed_svg_icon(path, foreground, size, dpr)
    if pixmap.isNull():
        # Do not reintroduce the historical fixed-white SVG fallback: a
        # damaged asset must not silently become white-on-light.  Bundled
        # resources are validated by the production smoke test.
        return QIcon()
    icon = QIcon(pixmap)
    _THEMED_ICON_CACHE[key] = icon
    return icon


def apply_admission_button_icon(button, icon_name, palette, role, *, size=18):
    """Apply one semantic V15 icon without changing the button's geometry."""
    role_name = str(role or "primary").lower()
    foreground = str(
        palette.get(f"button_{role_name}_text")
        or palette.get("button_fg")
        or "#FFFFFF"
    )
    theme_mode = str(palette.get("mode") or "")
    icon = theme_icon(
        icon_name,
        foreground,
        size=int(size),
        theme_mode=theme_mode,
        role=role_name,
    )
    button.setProperty("preserveOriginalIcons", True)
    button.setProperty("admissionVisualRole", role_name)
    button.setProperty("admissionIconSource", str(icon_name or ""))
    button.setProperty("admissionThemeMode", theme_mode)
    button.setProperty(
        "admissionIconFingerprint",
        f"{theme_mode}:{role_name}:{icon_name}:{foreground}:{int(size)}",
    )
    button.setIcon(icon)
    button.setIconSize(QSize(int(size), int(size)))
    return icon


def app_data_path(*paths):
    final_path = os.path.join(str(data_root()), *paths)

    if os.path.splitext(final_path)[1]:
        os.makedirs(os.path.dirname(final_path), exist_ok=True)
    else:
        os.makedirs(final_path, exist_ok=True)

    return final_path


def output_report_path(filename):
    return app_data_path("REPORTES", filename)


# -------------------------------
# ARCHIVO DIARIO DE REPORTES Y LISTADOS
# -------------------------------
MESES_ES = {
    1: "enero", 2: "febrero", 3: "marzo", 4: "abril",
    5: "mayo", 6: "junio", 7: "julio", 8: "agosto",
    9: "septiembre", 10: "octubre", 11: "noviembre", 12: "diciembre"
}

DIAS_ES = {
    0: "lunes", 1: "martes", 2: "miércoles", 3: "jueves",
    4: "viernes", 5: "sábado", 6: "domingo"
}


def limpiar_nombre_archivo(nombre: str) -> str:
    """
    Limpia caracteres no permitidos en nombres de carpetas/archivos de Windows.
    Mantiene acentos y espacios para que el nombre se vea natural en español.
    """
    nombre = (nombre or "").strip()
    nombre = re.sub(r'[<>:"/\\|?*]+', " ", nombre)
    nombre = re.sub(r"\s+", " ", nombre).strip()
    return nombre


def fecha_larga_es(fecha_base: date) -> str:
    return f"{DIAS_ES[fecha_base.weekday()]} {fecha_base.day:02d} de {MESES_ES[fecha_base.month]} de {fecha_base.year}"


def etiqueta_turno_archivo(turno_cfg: dict) -> str:
    """
    Para días normales devuelve solo la fecha en español.
    Para turnos divididos, especialmente domingo, agrega diurno/nocturno.
    """
    if not turno_cfg:
        return datetime.now().strftime("%d-%m-%Y")

    fecha_base = turno_cfg.get("fecha_base")
    if not isinstance(fecha_base, date):
        fecha_base = fecha_base_operativa_actual()

    base = fecha_larga_es(fecha_base)
    codigo = normalizar_turno_codigo(turno_cfg.get("turno_codigo", "8AM_8AM"))

    if codigo == "8AM_8PM":
        return f"{base} - diurno"
    if codigo == "8PM_8AM":
        return f"{base} - nocturno"

    return base


def carpeta_archivo_turno(turno_cfg: dict) -> str:
    etiqueta = limpiar_nombre_archivo(etiqueta_turno_archivo(turno_cfg))
    carpeta = os.path.join(ARCHIVO_DIARIO_DIR, etiqueta)
    os.makedirs(carpeta, exist_ok=True)
    return carpeta


def ruta_unica_si_existe(ruta: str) -> str:
    """
    Evita sobrescribir si por error se genera más de una vez el mismo archivo.
    """
    if not os.path.exists(ruta):
        return ruta

    base, ext = os.path.splitext(ruta)
    ts = datetime.now().strftime("%H%M%S")
    candidato = f"{base}_{ts}{ext}"
    contador = 2
    while os.path.exists(candidato):
        candidato = f"{base}_{ts}_{contador}{ext}"
        contador += 1
    return candidato


def guardar_copia_reporte_turno(ruta_pdf: str, turno_cfg: dict) -> str:
    """
    Guarda una copia organizada del reporte PDF del turno saliente.
    """
    if not ruta_pdf or not os.path.exists(ruta_pdf) or not turno_cfg:
        return ""

    carpeta = carpeta_archivo_turno(turno_cfg)
    etiqueta = limpiar_nombre_archivo(etiqueta_turno_archivo(turno_cfg))
    destino = ruta_unica_si_existe(os.path.join(carpeta, f"Reporte - {etiqueta}.pdf"))
    shutil.copy2(ruta_pdf, destino)
    return destino


def guardar_copia_excel_turno(turno_cfg: dict, ruta_excel=None) -> str:
    """
    Guarda una copia del listado Excel actual ANTES de limpiarlo/reconstruirlo.
    Se usa ruta_excel=None para evitar que EXCEL_PATH se evalúe antes de definirse.
    """
    if ruta_excel is None:
        ruta_excel = EXCEL_PATH

    if not turno_cfg or not os.path.exists(ruta_excel):
        return ""

    if not excel_tiene_registros(ruta_excel):
        return ""

    carpeta = carpeta_archivo_turno(turno_cfg)
    etiqueta = limpiar_nombre_archivo(etiqueta_turno_archivo(turno_cfg))
    destino = ruta_unica_si_existe(os.path.join(carpeta, f"Listado de pacientes - {etiqueta}.xlsx"))
    shutil.copy2(ruta_excel, destino)
    return destino


# -------------------------------
# RUTAS
# -------------------------------
BASE_DIR = os.path.dirname(os.path.abspath(__file__))

HOJAS_DIR = resource_path("HOJAS")
LOGO_PATH = resource_path("logo.jpg")

REPORTES_DIR = app_data_path("REPORTES")
ARCHIVO_DIARIO_DIR = app_data_path("ARCHIVO_DIARIO")
TURNOS_CFG = app_data_path("turnos_config.json")
EXCEL_PATH = app_data_path("LISTADO DE PACIENTES EN EMERGENCIA.xlsx")
EXCEL_LATEST_PATH = app_data_path("LISTADO.latest.xlsx")
EXCEL_EXPORT_STATE_PATH = app_data_path("excel_export_state.json")
EXCEL_VERSIONED_DIR = app_data_path("EXCEL_POR_TURNO")
EXCEL_EXPORT_QUEUE_PATH = app_data_path("excel_export_jobs.sqlite3")
APP_SETTINGS_PATH = app_data_path("app_settings.json")
ARS_CATALOGO_PATH = app_data_path("ars_catalogo.json")
NSS_FORMATOS_PATH = app_data_path("nss_formatos_ars.json")
REPRESENTANTES_PATH = app_data_path("representantes.json")
SECURITY_CONFIG_PATH = app_data_path("security.json")
LOGS_DIR = app_data_path("LOGS")
BACKUPS_DIR = app_data_path("BACKUPS")
DOCUMENTOS_DIR = app_data_path("DOCUMENTOS")
SUMATRA_PATH_CACHE = None
RESUMEN_TURNO_PATH = app_data_path("resumen_turno.json")

APP_LOG = logging.getLogger("emergencias")
APP_LOG.setLevel(logging.INFO)
if not APP_LOG.handlers:
    _log_path = os.path.join(LOGS_DIR, "app.log")
    try:
        _log_handler = RotatingFileHandler(
            _log_path,
            maxBytes=5 * 1024 * 1024,
            backupCount=5,
            encoding="utf-8",
        )
    except PermissionError:
        harden_windows_acl(data_root())
        try:
            _log_handler = RotatingFileHandler(
                _log_path,
                maxBytes=5 * 1024 * 1024,
                backupCount=5,
                encoding="utf-8",
            )
        except (OSError, PermissionError):
            _fallback_log = os.path.join(tempfile.gettempdir(), "GeneradorHojasEmergencia.log")
            _log_handler = RotatingFileHandler(
                _fallback_log,
                maxBytes=2 * 1024 * 1024,
                backupCount=2,
                encoding="utf-8",
            )
    _log_handler.setFormatter(logging.Formatter("%(asctime)s | %(levelname)s | %(message)s"))
    APP_LOG.addHandler(_log_handler)
APP_LOG.propagate = False

RUTA_HOJAS = {
    "GENERAL": os.path.join(HOJAS_DIR, "EMERGENCIA GENERAL.pdf"),
    "GINECOLOGIA": os.path.join(HOJAS_DIR, "EMERGENCIA GINECOLOGIA.pdf"),
    "PEDIATRIA": os.path.join(HOJAS_DIR, "EMERGENCIA PEDIATRICA.pdf"),
}

for key, ruta in RUTA_HOJAS.items():
    if not os.path.exists(ruta):
        print(f"[AVISO] Plantilla no encontrada para {key}: {ruta}")
    else:
        print(f"[OK] Plantilla {key} -> {ruta}")


# -------------------------------
# AJUSTE GLOBAL PARA REDUCCIÓN DE TEXTO
# -------------------------------
THRESHOLD_LEN = 38


# -------------------------------
# SEGUROS
# -------------------------------
SEGUROS_DISPLAY = {
    "SENASA SUBSIDIADO": "SUB",
    "SENASA CONTRIBUTIVO": "S.CONTRIBUTIVO",
    "SENASA PENSIONADOS": "S.PENSIONADOS",
    "APS": "APS",
    "ASEMAP": "ASEMAP",
    "CMD": "CMD",
    "GMA": "GMA",
    "RENACER": "RENACER",
    "RESERVAS": "RESERVAS",
    "SEMMA": "SEMMA",
    "FUTURO": "FUTURO",
    "HUMANO": "HUMANO",
    "PRIMERA": "PRIMERA",
    "ABEL GONZALEZ/SIMAG": "ABEL GONZALEZ/SIMAG",
    "METASALUD": "METASALUD",
    "MONUMENTAL": "MONUMENTAL",
    "MAPFRE/PALIC": "MAPFRE/PALIC",
    "UNIVERSAL": "UNIVERSAL",
    "BANCO CENTRAL": "BANCO CENTRAL",
    "YUNEN": "YUNEN",
    "SIN SEGURO": "SIN SEGURO",
}


# Catálogo editable de ARS y alias. Permite corregir equivalencias sin tocar el código.
DEFAULT_ARS_CATALOGO = {
    "SENASA SUBSIDIADO": [
        "SUB", "SUBS", "SUBSI", "SUBSID", "SUBSIDIADO", "SUNB",
        "SENASA SUBSIDIADO", "SENASA SUB", "SENASA REGIMEN SUBSIDIADO",
        "ARS SENASA SUBSIDIADO", "SESANA SUBSIADO", "SENASA SUBSIADO", "SUBSIADO"
    ],
    "SENASA CONTRIBUTIVO": [
        "CONTRIBUTIVO", "CONTRIB", "CONTRI", "CONT", "COTIZANTE",
        "SENASA CONTRIBUTIVO", "SENASA CONTRIB", "ARS SENASA CONTRIBUTIVO",
        "SENASA AVANZADA", "AVANZADA",
        "SENASA MAXIMO", "MAXIMO", "MÁXIMO",
        "SENASA ESPECIAL", "ESPECIAL"
    ],
    "SENASA PENSIONADOS": [
        "PENSIONADO", "PENSIONADOS", "PENS",
        "SENASA PENSIONADO", "SENASA PENSIONADOS", "ARS SENASA PENSIONADOS"
    ],
    "HUMANO": [
        "HUMANO", "ARS HUMANO", "HUM", "HUMANA", "HUMAO"
    ],
    "MAPFRE/PALIC": [
        "MAPFRE", "PALIC", "MAPFRE PALIC", "MAPFRE/PALIC",
        "ARS MAPFRE", "ARS PALIC", "MAPHRE"
    ],
    "UNIVERSAL": [
        "UNIVERSAL", "ARS UNIVERSAL", "UNI"
    ],
    "RESERVAS": [
        "RESERVAS", "ARS RESERVAS", "BANRESERVAS", "BANRESERVA", "RESERVA"
    ],
    "MONUMENTAL": [
        "MONUMENTAL", "ARS MONUMENTAL", "MONU"
    ],
    "PRIMERA": [
        "PRIMERA", "ARS PRIMERA", "PRIMERA ARS"
    ],
    "RENACER": [
        "RENACER", "ARS RENACER"
    ],
    "GMA": [
        "GMA"
    ],
    "YUNEN": [
        "YUNEN", "ARS YUNEN"
    ],
    "ABEL GONZALEZ/SIMAG": [
        "SIMAG", "ABEL GONZALEZ", "ABEL GONZÁLEZ", "ABEL",
        "ARS ABEL GONZALEZ", "CENTRO MEDICO ABEL GONZALEZ",
        "CENTRO MÉDICO ABEL GONZÁLEZ"
    ],
    "CMD": [
        "CMD", "COLEGIO MEDICO", "COLEGIO MÉDICO", "COLEGIO MEDICO DOMINICANO"
    ],
    "SEMMA": [
        "SEMMA", "ARS SEMMA", "SEGURO MAESTROS", "MAESTROS"
    ],
    "FUTURO": [
        "FUTURO", "ARS FUTURO"
    ],
    "APS": [
        "APS", "ARS APS"
    ],
    "ASEMAP": [
        "ASEMAP"
    ],
    "METASALUD": [
        "METASALUD", "META SALUD"
    ],
    "BANCO CENTRAL": [
        "BANCO CENTRAL", "BC", "BANCENTRAL"
    ],
    "MEDICA": [
        "MEDICA", "MÉDICA", "ARS MEDICA", "ARS MÉDICA"
    ],
    "SIN SEGURO": [
        "SIN SEGURO", "NO TIENE", "NO", "N/S", "NS", "N\\S",
        "INACTIVO", "INACTIVA", "PENDIENTE", "PENDIENTES",
        "NO ACTIVO", "NO VIGENTE", "VENCIDO", "CANCELADO",
        "DESAFILIADO", "PARTICULAR", "PRIVADO", "NINGUNO", "NINGUNA",
        "N/A", "NA", "S/N", "SN", "NO APLICA", "NO USA", "NO POSEE",
        "NINGUN SEGURO"
    ],
}

# FASE 1: Límites de paginación del historial.
DEFAULT_APP_SETTINGS = {
    # Visual básico
    "font_size": 11,
    "theme": "oscuro",
    "high_contrast": False,
    "window_size": "1280x740",
    "auto_print": True,

    # Preferencias de impresión
    "print_auto_hoja": True,
    "print_auto_reporte_turno": True,
    "print_auto_excel_turno": True,
    "print_copies_hoja": 1,
    "print_copies_reporte": 2,
    "print_copies_excel": 2,
    "print_pdf_orientation": "Horizontal",
    "print_excel_orientation": "Horizontal",
    "print_behavior_hoja": "Imprimir y abrir PDF",

    # Preferencias de validación
    "validation_confirm_before_generate": True,
    "validation_warn_nss_incomplete": True,
    "validation_warn_ars_sin_seguro": True,
    "validation_block_short_ars": True,
    "validation_allow_missing_cedula": True,
    "validation_allow_missing_phone": False,
    "validation_warn_duplicate_turno": True,

    # Preferencias de recién nacido
    "rn_strip_db": True,
    "rn_show_pdf": True,
    "rn_warn": True,
    "rn_format_display": "RN- NOMBRE DE LA MADRE",

    # Preferencias visuales avanzadas
    "button_size": "Normal",
    "table_row_height": 29,
    "compact_mode": False,
    "small_screen_mode": False,
    "show_side_panel": True,
    "show_turno_summary": True,
    "accent_color": "Azul hospitalario",

    # Preferencias de historial (FASE 1: 100 / 150)
    "hist_initial_limit": 100,
    "hist_next_limit": 150,
    "hist_default_filter": "Todos",
    "hist_order": "Más reciente primero",

    # Preferencias de turnos
    "turno_default": "8AM_8AM",
    "turnos_ask_representante_start": False,
    "turnos_generate_report": True,
    "turnos_save_excel_copy": True,
    "turnos_print_empty_report": False,
    "turnos_open_archive_folder": False,

    # Preferencias de PDF
    "pdf_nss_guiones": True,
    "pdf_ars_display_mode": "Abreviada",
    "pdf_nombre_font_size": 12,
    "pdf_direccion_font_size": 12,
    "pdf_open_after_generate": True,
    "pdf_keep_temp": False,
}

ACCENT_COLOR_PRESETS = {
    "Profesional sobrio": "#4f6472",
    "Azul hospitalario": "#68A9D8",
    "Azul profundo": "#4A8CC3",
    "Celeste suave": "#76BDE8",
    "Turquesa clínico": "#55B7B0",
    "Verde salud": "#5BAA70",
    "Gris profesional": "#7D8DA1",
}

def mezclar_color_hex(color_a, color_b, porcentaje=0.15):
    """
    Mezcla dos colores HEX.
    porcentaje=0.15 significa 15% del color_a sobre 85% del color_b.
    Se usa para que el color principal afecte fondos de forma suave.
    """
    try:
        color_a = str(color_a or "").strip().lstrip("#")
        color_b = str(color_b or "").strip().lstrip("#")
        if len(color_a) != 6 or len(color_b) != 6:
            return "#" + color_b

        porcentaje = max(0.0, min(1.0, float(porcentaje)))
        ra, ga, ba = int(color_a[0:2], 16), int(color_a[2:4], 16), int(color_a[4:6], 16)
        rb, gb, bb = int(color_b[0:2], 16), int(color_b[2:4], 16), int(color_b[4:6], 16)

        r = int((ra * porcentaje) + (rb * (1 - porcentaje)))
        g = int((ga * porcentaje) + (gb * (1 - porcentaje)))
        b = int((ba * porcentaje) + (bb * (1 - porcentaje)))
        return f"#{r:02X}{g:02X}{b:02X}"
    except Exception:
        return str(color_b if str(color_b).startswith("#") else "#" + str(color_b))


def resolver_color_principal(valor):
    val = str(valor or "Azul hospitalario").strip()
    if val in ACCENT_COLOR_PRESETS:
        return ACCENT_COLOR_PRESETS[val]
    if re.match(r"^#[0-9A-Fa-f]{6}$", val):
        return val
    return ACCENT_COLOR_PRESETS["Azul hospitalario"]

def nombre_color_principal(valor):
    val = str(valor or "Azul hospitalario").strip()
    if val in ACCENT_COLOR_PRESETS:
        return val
    for nombre, hexv in ACCENT_COLOR_PRESETS.items():
        if val.lower() == hexv.lower():
            return nombre
    return "Azul hospitalario"

def cargar_app_settings():
    try:
        data = load_json_file(APP_SETTINGS_PATH, default={})
        out = dict(DEFAULT_APP_SETTINGS)
        if isinstance(data, dict):
            out.update(data)
        return out
    except ConfigError as exc:
        logging.getLogger("emergencias").error("Configuracion invalida: %s", exc)
    return dict(DEFAULT_APP_SETTINGS)

def guardar_app_settings(settings: dict):
    try:
        data = dict(DEFAULT_APP_SETTINGS)
        data.update(settings or {})
        atomic_write_json(APP_SETTINGS_PATH, data)
        return True
    except (OSError, TypeError, ValueError) as exc:
        logging.getLogger("emergencias").exception("No se pudieron guardar las preferencias: %s", exc)
        return False

def app_setting(key, default=None):
    try:
        return cargar_app_settings().get(key, DEFAULT_APP_SETTINGS.get(key, default))
    except Exception:
        return DEFAULT_APP_SETTINGS.get(key, default)

def cargar_catalogo_ars():
    try:
        merged = {k: list(v) for k, v in DEFAULT_ARS_CATALOGO.items()}

        if os.path.exists(ARS_CATALOGO_PATH):
            with open(ARS_CATALOGO_PATH, "r", encoding="utf-8") as f:
                data = json.load(f)

            if isinstance(data, dict):
                for k, v in data.items():
                    if not k:
                        continue

                    key = _limpiar_texto_seguro(k)
                    if isinstance(v, str):
                        aliases = [x.strip() for x in v.split(",") if x.strip()]
                    elif isinstance(v, list):
                        aliases = [str(x).strip() for x in v if str(x).strip()]
                    else:
                        aliases = []

                    base_aliases = merged.get(key, [])
                    vistos = set(_limpiar_texto_seguro(a) for a in base_aliases)
                    final_aliases = list(base_aliases)
                    for alias in aliases:
                        alias_clean = _limpiar_texto_seguro(alias)
                        if alias_clean and alias_clean not in vistos:
                            final_aliases.append(alias)
                            vistos.add(alias_clean)
                    merged[key] = final_aliases

        return merged
    except Exception:
        pass
    return {k: list(v) for k, v in DEFAULT_ARS_CATALOGO.items()}

def guardar_catalogo_ars(catalogo: dict):
    try:
        data = {}
        for k, v in (catalogo or {}).items():
            key = _limpiar_texto_seguro(k)
            if not key: continue
            aliases = v if isinstance(v, list) else str(v).split(",")
            data[key] = [str(a).strip() for a in aliases if str(a).strip()]
        with open(ARS_CATALOGO_PATH, "w", encoding="utf-8") as f:
            json.dump(data, f, ensure_ascii=False, indent=2)
        return True
    except Exception:
        return False


def cargar_formatos_nss_ars():
    default = {
        "RENACER": "5-5-2"
    }
    try:
        if os.path.exists(NSS_FORMATOS_PATH):
            with open(NSS_FORMATOS_PATH, "r", encoding="utf-8") as f:
                data = json.load(f)
            if isinstance(data, dict):
                for k, v in data.items():
                    key = _limpiar_texto_seguro(k)
                    patron = normalizar_patron_nss(v)
                    if key and patron:
                        default[key] = patron
    except Exception:
        pass
    return default


def guardar_formatos_nss_ars(formatos: dict):
    try:
        data = {}
        for k, v in (formatos or {}).items():
            key = _limpiar_texto_seguro(k)
            patron = normalizar_patron_nss(v)
            if key and patron:
                data[key] = patron
        with open(NSS_FORMATOS_PATH, "w", encoding="utf-8") as f:
            json.dump(data, f, ensure_ascii=False, indent=2)
        return True
    except Exception:
        return False


def normalizar_patron_nss(patron) -> str:
    raw = str(patron or "").strip()
    if not raw:
        return ""
    nums = re.findall(r"\d+", raw)
    if not nums:
        return ""
    valores = []
    for n in nums:
        try:
            val = int(n)
            if val <= 0:
                return ""
            valores.append(str(val))
        except Exception:
            return ""
    return "-".join(valores)


def patron_desde_ejemplo_nss(ejemplo: str) -> str:
    raw = str(ejemplo or "").strip()
    if "-" not in raw:
        return ""

    partes = [p for p in raw.split("-") if p != ""]
    if not partes:
        return ""

    if not all(re.sub(r"\D", "", p) for p in partes):
        return ""

    return "-".join(str(len(re.sub(r"\D", "", p))) for p in partes)


def aplicar_patron_nss(nss: str, patron: str) -> str:
    original = str(nss or "").strip().upper()
    digitos = re.sub(r"\D", "", original)
    patron_norm = normalizar_patron_nss(patron)
    if not digitos or not patron_norm:
        return original

    partes = [int(x) for x in patron_norm.split("-")]
    if sum(partes) != len(digitos):
        return original

    out = []
    pos = 0
    for size in partes:
        out.append(digitos[pos:pos + size])
        pos += size
    return "-".join(out)


def formatear_nss_para_pdf(nss: str, ars: str) -> str:
    original = str(nss or "").strip().upper()
    if not original:
        return ""

    ars_canon = normalizar_seguro(ars or "", original)
    formatos = cargar_formatos_nss_ars()
    patron = formatos.get(ars_canon)

    if not patron:
        ars_limpia = _limpiar_texto_seguro(ars or "")
        for key, value in formatos.items():
            if key == ars_limpia or key in ars_limpia or ars_limpia in key:
                patron = value
                break

    return aplicar_patron_nss(original, patron) if patron else original


def limpiar_nombre_rn_para_db(nombre: str) -> str:
    txt = str(nombre or "").strip()
    txt = re.sub(r"^\s*RN\s*[-–—:]\s*", "", txt, flags=re.IGNORECASE)
    return re.sub(r"\s+", " ", txt).strip()


def nombre_tiene_prefijo_rn(nombre: str) -> bool:
    return bool(re.match(r"^\s*RN\s*[-–—:]\s*", str(nombre or ""), flags=re.IGNORECASE))


# -------------------------------
# HELPERS DE VALIDACIÓN
# -------------------------------
def is_all_zeros(s: str) -> bool:
    s = (s or "").strip()
    return len(s) > 0 and set(s) == {"0"}


def is_valid_nss_key(nss: str) -> bool:
    if not nss:
        return False
    nss = nss.strip().upper()
    if nss in ["N/S", r"N\\S", "NS", "NO", "SIN SEGURO", ""]:
        return False
    if not nss.isdigit():
        return False
    if is_all_zeros(nss):
        return False
    if len(nss) < 3:
        return False
    return True


def is_valid_cedula_key(cedula: str) -> bool:
    if not cedula:
        return False
    ced = cedula.strip()
    return ced.isdigit() and len(ced) == 11 and not is_all_zeros(ced)


def normalizar_nombre_clave(nombre: str) -> str:
    return re.sub(r"\s+", " ", str(nombre or "").strip()).upper()


def get_patient_key(nss: str, cedula: str):
    if is_valid_nss_key(nss):
        return nss.strip().upper()
    if is_valid_cedula_key(cedula):
        return cedula.strip()
    return None


def patient_identity_key_from_row(row: dict):
    nss = (row.get("nss") or "").strip().upper()
    cedula = (row.get("cedula") or "").strip()
    if is_valid_nss_key(nss):
        return ("NSS", nss)
    if is_valid_cedula_key(cedula):
        return ("CEDULA", cedula)
    return ("ROW", row.get("id"))


def _detectar_campos_invertidos(nss_raw: str, ars_raw: str):
    nss = (nss_raw or "").strip()
    ars = (ars_raw or "").strip()

    nss_valid_text = ["N/S", r"N\\S", "NS", "NO", "SIN SEGURO", ""]
    nss_is_text = any(c.isalpha() for c in nss) and nss.upper() not in nss_valid_text
    ars_sin_espacios = re.sub(r"\s+", "", ars)
    ars_is_numeric = ars_sin_espacios.isdigit() and len(ars_sin_espacios) >= 3

    if nss_is_text and ars_is_numeric:
        return (
            True,
            "Campos invertidos:\n\nColocaste el nombre del seguro en el campo 'NSS' "
            "y el número en el campo 'Aseguradora'.\n\nPor favor, intercámbialos."
        )
    if nss_is_text:
        return (
            True,
            f"Error en NSS:\n\nEl campo NSS debe ser estrictamente numérico. "
            f"Si no tiene NSS, déjalo vacío o escribe 'SIN SEGURO'. No coloques el nombre de la aseguradora aquí."
        )
    if ars_is_numeric:
        return (
            True,
            f"Error en Aseguradora (ARS):\n\nEl campo Aseguradora debe contener el nombre "
            f"del seguro (texto), no números largos."
        )

    return (False, "")


# -------------------------------
# FECHAS / DÍA OPERATIVO
# -------------------------------
def parse_fecha_ddmmyyyy(fecha_str: str):
    try:
        return datetime.strptime((fecha_str or "").strip(), "%d/%m/%Y").date()
    except Exception:
        return None


def parse_hora_12h(hora_str: str):
    try:
        return datetime.strptime((hora_str or "").strip(), "%I:%M %p").time()
    except Exception:
        try:
            txt = (hora_str or "").strip().upper().replace(".", "")
            match = re.search(r"(\d{1,2}):(\d{2})\s*(A\s*M|P\s*M|A|P)", txt)
            if match:
                h = int(match.group(1))
                m = int(match.group(2))
                p = match.group(3).replace(" ", "")
                if p.startswith("P") and h < 12:
                    h += 12
                elif p.startswith("A") and h == 12:
                    h = 0
                return time(h, m)
        except Exception:
            pass
        return None


def parse_datetime_local(dt_str: str):
    try:
        return datetime.strptime((dt_str or "").strip(), "%d/%m/%Y %I:%M %p")
    except Exception:
        return None


def format_datetime_local(dt: datetime) -> str:
    return dt.strftime("%d/%m/%Y %I:%M %p")


def construir_datetime_operativo(fecha_str: str, hora_str: str):
    f = parse_fecha_ddmmyyyy(fecha_str)
    h = parse_hora_12h(hora_str)
    if not f or not h:
        return None

    dt = datetime.combine(f, h)
    if h < time(8, 0):
        dt = dt - timedelta(days=1)
    return dt


def construir_datetime_real(fecha_str: str, hora_str: str):
    f = parse_fecha_ddmmyyyy(fecha_str)
    h = parse_hora_12h(hora_str)
    if not f or not h:
        return None
    return datetime.combine(f, h)


def obtener_rango_operativo_desde_fecha(base_date: date):
    inicio = datetime.combine(base_date, time(8, 0))
    fin = inicio + timedelta(days=1)
    return inicio, fin


def fecha_base_operativa_actual(momento: datetime = None) -> date:
    actual = momento or datetime.now()
    return actual.date() - timedelta(days=1) if actual.time() < time(8, 0) else actual.date()


def normalizar_turno_codigo(turno_codigo: str) -> str:
    raw = str(turno_codigo or "").strip()
    if raw in ("8AM_8AM", "8AM_8PM", "8PM_8AM"):
        return raw

    up = raw.upper()
    up = up.replace("→", " A ").replace("-", " A ").replace("–", " A ").replace("—", " A ")
    up = up.replace(".", "")
    up = up.replace(":00", "")
    up = re.sub(r"\s+", "", up)

    if "8PM" in up and "8AM" in up:
        return "8PM_8AM"
    if "8AM" in up and "8PM" in up:
        return "8AM_8PM"
    if "8AM" in up:
        return "8AM_8AM"

    return "8AM_8AM"


def limpiar_nombre_representante(valor: str) -> str:
    txt = str(valor or "").strip()
    if not txt:
        return ""

    txt = re.sub(r"\b\d{1,2}/\d{1,2}/\d{4}\b", " ", txt)
    txt = re.sub(r"\bAL\b", " ", txt, flags=re.IGNORECASE)
    txt = re.sub(r"\bDEL\b", " ", txt, flags=re.IGNORECASE)
    txt = re.sub(r"\bDESDE\b", " ", txt, flags=re.IGNORECASE)
    txt = re.sub(r"\bHASTA\b", " ", txt, flags=re.IGNORECASE)

    txt = re.sub(r"[-–—:|]+", " ", txt)
    txt = re.sub(r"\s+", " ", txt).strip()

    if txt.upper() in {"AL", "DEL", "DESDE", "HASTA"}:
        return ""
    return txt


REPRESENTANTES_NO_VALIDOS = {
    "NO DISPONIBLE",
    "NO CONFIGURADO",
    "NOMBRE DEL REPRESENTANTE",
    "NOMBRE REPRESENTANTE",
    "SIN REPRESENTANTE",
}


def es_representante_valido(valor: str) -> bool:
    limpio = limpiar_nombre_representante(valor)
    return bool(limpio and limpio.upper() not in REPRESENTANTES_NO_VALIDOS)


def cargar_representantes(db=None, incluir_actual=True):
    nombres = set()
    catalogo_existente = os.path.exists(REPRESENTANTES_PATH)
    try:
        if catalogo_existente:
            with open(REPRESENTANTES_PATH, "r", encoding="utf-8") as archivo:
                data = json.load(archivo)
            for nombre in data if isinstance(data, list) else []:
                limpio = limpiar_nombre_representante(nombre)
                if es_representante_valido(limpio):
                    nombres.add(limpio)
    except Exception:
        APP_LOG.exception("No se pudo leer el catálogo de representantes")

    if incluir_actual:
        try:
            cfg = cargar_turno_config(permitir_vencido=True)
            actual = limpiar_nombre_representante((cfg or {}).get("representante", ""))
            if es_representante_valido(actual):
                nombres.add(actual)
        except Exception:
            APP_LOG.exception("No se pudo recuperar el representante del turno actual")

    # La BD histórica solo se usa para inicializar el catálogo una vez. Después,
    # un usuario eliminado no debe reaparecer por existir en turnos antiguos.
    if db is not None and not catalogo_existente:
        try:
            nombres.update(
                nombre for nombre in db.listar_representantes()
                if es_representante_valido(nombre)
            )
        except Exception:
            APP_LOG.exception("No se pudieron consultar representantes históricos")

    return sorted(nombres, key=lambda valor: valor.casefold())


def guardar_catalogo_representantes(nombres) -> bool:
    limpios = {}
    for nombre in nombres or []:
        limpio = limpiar_nombre_representante(nombre)
        if es_representante_valido(limpio):
            limpios.setdefault(limpio.casefold(), limpio)
    try:
        atomic_write_json(
            REPRESENTANTES_PATH,
            sorted(limpios.values(), key=lambda valor: valor.casefold()),
        )
        return True
    except (OSError, TypeError, ValueError):
        APP_LOG.exception("No se pudo guardar el catálogo de representantes")
        return False


def guardar_representante_catalogo(nombre: str, db=None):
    limpio = limpiar_nombre_representante(nombre)
    if not es_representante_valido(limpio):
        return ""

    nombres = cargar_representantes(db)
    if limpio.casefold() not in {valor.casefold() for valor in nombres}:
        nombres.append(limpio)
    return limpio if guardar_catalogo_representantes(nombres) else ""

def descripcion_turno_config(turno_cfg: dict) -> str:
    if not turno_cfg:
        return "No configurado"
    try:
        fecha_base = turno_cfg.get("fecha_base")
        if not isinstance(fecha_base, date):
            return "No configurado"
        codigo = normalizar_turno_codigo(turno_cfg.get("turno_codigo", "8AM_8AM"))
        return obtener_datos_turno_visual(fecha_base, codigo)["turno_resumen"]
    except Exception:
        return "No configurado"


def obtener_rango_turno_real(fecha_base: date, turno_codigo: str):
    turno_codigo = normalizar_turno_codigo(turno_codigo)
    if turno_codigo == "8AM_8PM":
        inicio = datetime.combine(fecha_base, time(8, 0))
        fin = datetime.combine(fecha_base, time(20, 0))
        return inicio, fin

    if turno_codigo == "8PM_8AM":
        inicio = datetime.combine(fecha_base, time(20, 0))
        fin = datetime.combine(fecha_base + timedelta(days=1), time(8, 0))
        return inicio, fin

    inicio = datetime.combine(fecha_base, time(8, 0))
    fin = datetime.combine(fecha_base + timedelta(days=1), time(8, 0))
    return inicio, fin


def obtener_datos_turno_visual(fecha_base: date, turno: str):
    turno = normalizar_turno_codigo(turno)
    siguiente = fecha_base + timedelta(days=1)

    if turno == "8AM_8PM":
        return {
            "turno_label": "DESDE 8:00 AM A 8:00 PM",
            "fecha_label": fecha_base.strftime("%d/%m/%Y"),
            "turno_resumen": "8:00 AM → 8:00 PM",
        }

    if turno == "8PM_8AM":
        return {
            "turno_label": "DESDE 8:00 PM A 8:00 AM",
            "fecha_label": f"{fecha_base.strftime('%d/%m/%Y')} AL {siguiente.strftime('%d/%m/%Y')}",
            "turno_resumen": "8:00 PM → 8:00 AM",
        }

    return {
        "turno_label": "DESDE 8:00 AM A 8:00 AM",
        "fecha_label": f"{fecha_base.strftime('%d/%m/%Y')} AL {siguiente.strftime('%d/%m/%Y')}",
        "turno_resumen": "8:00 AM → 8:00 AM",
    }


def obtener_inicio_real_turno(turno_cfg: dict):
    if not turno_cfg:
        return None

    inicio_nominal, _ = obtener_rango_turno_real(turno_cfg["fecha_base"], turno_cfg["turno_codigo"])
    inicio_real_guardado = turno_cfg.get("inicio_real_dt")

    if inicio_real_guardado and inicio_real_guardado > inicio_nominal:
        return inicio_real_guardado
    return inicio_nominal


def obtener_rango_turno_efectivo(turno_cfg: dict, fin_override: datetime = None):
    if not turno_cfg:
        return None, None
    inicio_nominal, fin_nominal = obtener_rango_turno_real(turno_cfg["fecha_base"], turno_cfg["turno_codigo"])
    inicio_real = obtener_inicio_real_turno(turno_cfg)
    fin_real = fin_override if fin_override else fin_nominal
    if fin_real < inicio_real:
        fin_real = inicio_real
    return inicio_real, fin_real


def turno_config_es_vigente(turno_cfg: dict, momento: datetime = None) -> bool:
    if not turno_cfg:
        return False
    try:
        inicio, fin = obtener_rango_turno_efectivo(turno_cfg)
        actual = momento or datetime.now()
        return bool(inicio and fin and inicio <= actual < fin)
    except (KeyError, TypeError, ValueError):
        return False


# -------------------------------
# NORMALIZACIÓN DE SEGUROS
# -------------------------------
def _limpiar_texto_seguro(txt: str) -> str:
    txt = (txt or "").strip().upper()
    txt = txt.replace("Á", "A").replace("É", "E").replace("Í", "I").replace("Ó", "O").replace("Ú", "U")
    txt = re.sub(r"\bARS\b", "", txt)
    txt = re.sub(r"[^A-Z0-9/ ]+", " ", txt)
    txt = re.sub(r"\s+", " ", txt).strip()
    return txt


def _compact(txt: str) -> str:
    return re.sub(r"[^A-Z0-9]+", "", _limpiar_texto_seguro(txt))


def _score(a: str, b: str) -> float:
    return SequenceMatcher(None, _compact(a), _compact(b)).ratio()


def _es_sin_seguro_por_texto(seguro_raw: str) -> bool:
    txt = _limpiar_texto_seguro(seguro_raw)
    valores = {
        "", "N/S", "NS", "NO", "SIN SEGURO", "NINGUNO", "NINGUNA",
        "N/A", "NA", "S/N", "SN", "NO APLICA", "NO TIENE",
        "NO USA", "NO POSEE", "NINGUN SEGURO",
        "INACTIVO", "INACTIVA", "PENDIENTE", "PENDIENTES",
        "NO ACTIVO", "NO VIGENTE", "VENCIDO", "CANCELADO",
        "DESAFILIADO", "PARTICULAR", "PRIVADO"
    }
    if txt in valores:
        return True
    if txt.startswith("SIN SEG"):
        return True
    return False


def _mejor_seguro_por_similitud(txt: str):
    txt_clean = _limpiar_texto_seguro(txt)
    txt_comp = _compact(txt_clean)

    if not txt_comp:
        return None

    candidatos = {
        "SENASA SUBSIDIADO": [
            "SUB", "SUNB", "SUBSIDIADO", "SENASA SUB", "SENASA SUBSIDIADO",
            "ARS SENASA SUBSIDIADO", "SESANA SUBSIADO", "SENASA SUBSIADO",
            "SUBSIADO", "SUBSI", "SENASA SUBSIDIAD", "SUBS"
        ],
        "SENASA CONTRIBUTIVO": [
            "CONTRIBUTIVO", "SENASA CONTRIBUTIVO", "ARS SENASA CONTRIBUTIVO",
            "CONTRI", "CONTRIB", "CONTRIBUT", "CONT", "COTIZANTE",
            "SENASA AVANZADA", "AVANZADA", "SENASA MAXIMO", "MAXIMO", "MÁXIMO",
            "SENASA ESPECIAL", "ESPECIAL"
        ],
        "SENASA PENSIONADOS": [
            "PENSIONADOS", "SENASA PENSIONADOS", "ARS SENASA PENSIONADOS", "PENSIONADO"
        ],
        "APS": ["APS"],
        "ASEMAP": ["ASEMAP"],
        "CMD": ["CMD"],
        "GMA": ["GMA"],
        "RENACER": ["RENACER"],
        "RESERVAS": ["RESERVAS", "ARS RESERVAS"],
        "SEMMA": ["SEMMA"],
        "FUTURO": ["FUTURO"],
        "HUMANO": ["HUMANO", "ARS HUMANO", "HUMAO"],
        "PRIMERA": ["PRIMERA", "ARS PRIMERA"],
        "ABEL GONZALEZ/SIMAG": ["ABEL GONZALEZ", "SIMAG", "ABEL", "ARS ABEL GONZALEZ"],
        "METASALUD": ["METASALUD"],
        "MONUMENTAL": ["MONUMENTAL"],
        "MAPFRE/PALIC": ["MAPFRE", "PALIC", "ARS MAPFRE", "ARS PALIC", "MAPHRE"],
        "UNIVERSAL": ["UNIVERSAL"],
        "BANCO CENTRAL": ["BANCO CENTRAL"],
        "YUNEN": ["YUNEN"],
    }

    try:
        for canonico, aliases in cargar_catalogo_ars().items():
            canon = _limpiar_texto_seguro(canonico)
            if not canon: continue
            candidatos.setdefault(canon, [])
            for alias in aliases:
                if alias and alias not in candidatos[canon]:
                    candidatos[canon].append(alias)
    except Exception:
        pass

    if txt_comp.startswith("SU") or txt_comp in {"SUNB", "SUB", "SUBS", "SUBSI", "SUBSIDIADO", "SUBSIADO"}:
        return "SENASA SUBSIDIADO"

    if txt_comp in {"AVANZADA", "SENASAAVANZADA", "MAXIMO", "MAXIMA", "SENASAMAXIMO", "SENASAMAXIMA", "ESPECIAL", "SENASAESPECIAL"}:
        return "SENASA CONTRIBUTIVO"

    if "SENASA" in txt_clean and ("SUB" in txt_clean or "SUBSI" in txt_clean or "SUBSID" in txt_clean):
        return "SENASA SUBSIDIADO"
    if "SENASA" in txt_clean and ("CONTRI" in txt_clean or "CONTRIB" in txt_clean):
        return "SENASA CONTRIBUTIVO"
    if "SENASA" in txt_clean and "PENSION" in txt_clean:
        return "SENASA PENSIONADOS"

    if "MAPFRE" in txt_clean or "PALIC" in txt_clean:
        return "MAPFRE/PALIC"
    if "ABEL" in txt_clean or "SIMAG" in txt_clean:
        return "ABEL GONZALEZ/SIMAG"
    if "BANCO CENTRAL" in txt_clean:
        return "BANCO CENTRAL"

    mejor_nombre = None
    mejor_score = 0.0

    for canonico, aliases in candidatos.items():
        for alias in aliases:
            s = _score(txt_clean, alias)
            if s > mejor_score:
                mejor_score = s
                mejor_nombre = canonico

    if mejor_score >= 0.72:
        return mejor_nombre

    if len(txt_comp) <= 8 and mejor_score >= 0.58:
        return mejor_nombre

    flat_aliases = []
    alias_to_canon = {}
    for canonico, aliases in candidatos.items():
        for alias in aliases:
            flat_aliases.append(alias)
            alias_to_canon[alias] = canonico

    match = get_close_matches(txt_clean, flat_aliases, n=1, cutoff=0.55)
    if match:
        return alias_to_canon[match[0]]

    return None


def ars_es_corta_invalida(seguro_raw: str) -> bool:
    txt = _limpiar_texto_seguro(seguro_raw)
    comp = _compact(txt)
    if not comp or _es_sin_seguro_por_texto(txt):
        return False
    if comp.isdigit() or len(comp) == 1:
        return True
    if _mejor_seguro_por_similitud(txt):
        return False
    return len(comp) < 4

from functools import lru_cache
@lru_cache(maxsize=5000)
def normalizar_seguro(seguro_raw: str, nss_raw: str = "") -> str:
    txt = _limpiar_texto_seguro(seguro_raw)
    nss = (nss_raw or "").strip().upper()

    if _es_sin_seguro_por_texto(txt):
        return "SIN SEGURO"

    comp_txt = _compact(txt)
    if comp_txt.isdigit() or len(comp_txt) == 1:
        return "SIN SEGURO"

    parecido = _mejor_seguro_por_similitud(txt)
    if parecido:
        return parecido

    if ars_es_corta_invalida(txt):
        return "SIN SEGURO"

    if not is_valid_nss_key(nss):
        return "SIN SEGURO"

    return txt if txt else "SIN SEGURO"


def seguro_para_mostrar(seguro_canonico: str) -> str:
    return SEGUROS_DISPLAY.get(seguro_canonico, (seguro_canonico or "").strip().upper() or "SIN SEGURO")


# -------------------------------
# CONFIG TURNO
# -------------------------------
def cargar_turno_config(permitir_vencido=False):
    try:
        data = load_json_file(TURNOS_CFG, default=None)
        if not isinstance(data, dict):
            return None
        fecha_base = parse_fecha_ddmmyyyy(data.get("fecha_base", ""))
        if not fecha_base:
            return None

        inicio_real_dt = parse_datetime_local(data.get("inicio_real", ""))

        representante = limpiar_nombre_representante(data.get("representante", ""))
        if not es_representante_valido(representante):
            representante = ""
        config = {
            "representante": representante,
            "turno_codigo": normalizar_turno_codigo(data.get("turno_codigo", "8AM_8AM")),
            "fecha_base": fecha_base,
            "inicio_real": data.get("inicio_real", ""),
            "inicio_real_dt": inicio_real_dt,
            "administrative_override": bool(data.get("administrative_override")),
            "override_reason": str(data.get("override_reason") or "")[:240],
        }
        if (
            not permitir_vencido
            and not config["administrative_override"]
            and not turno_config_es_vigente(config)
        ):
            APP_LOG.warning("Se ignoro una configuracion de turno vencida: %s", data)
            return None
        return config
    except ConfigError as exc:
        APP_LOG.error("No se pudo leer la configuracion del turno: %s", exc)
        return None


def guardar_turno_config(
    representante: str,
    turno_codigo: str,
    fecha_base: date,
    inicio_real: datetime = None,
    *,
    administrative_override: bool = False,
    override_reason: str = "",
):
    try:
        representante = limpiar_nombre_representante(representante)
        if not es_representante_valido(representante):
            APP_LOG.warning(
                "Se rechazó un representante inválido para el turno: %r",
                representante,
            )
            return False
        payload = {
            "representante": representante,
            "turno_codigo": normalizar_turno_codigo(turno_codigo),
            "fecha_base": fecha_base.strftime("%d/%m/%Y"),
            "inicio_real": format_datetime_local(inicio_real or datetime.now()),
            "administrative_override": bool(administrative_override),
            "override_reason": str(override_reason or "")[:240],
        }
        atomic_write_json(TURNOS_CFG, payload)
        return True
    except (OSError, TypeError, ValueError) as exc:
        APP_LOG.exception("No se pudo guardar la configuracion del turno: %s", exc)
        return False


def excel_requiere_turno_manual() -> bool:
    """Detecta un listado conservado sin crear ni cambiar turnos automáticamente."""
    try:
        if cargar_turno_config(permitir_vencido=True):
            return False
        return bool(
            os.path.exists(EXCEL_PATH)
            and excel_tiene_registros(EXCEL_PATH)
        )
    except Exception:
        return False


# -------------------------------
# DB MANAGER
# -------------------------------
class TurnoNoVigenteError(RuntimeError):
    pass


class DatabaseManager:
    SCHEMA_VERSION = LATEST_SCHEMA_VERSION
    _BOOTSTRAP_READY_PATHS = set()

    def __init__(self, db_name='pacientes.db', *, session_context, event_bus=None):
        self.session_context = session_context
        self.event_bus = event_bus
        self.db_name = app_data_path(db_name)
        self.backup_manager = BackupManager(
            self.db_name,
            BACKUPS_DIR,
            related_paths=(TURNOS_CFG, APP_SETTINGS_PATH, EXCEL_PATH),
            retention_days=4,
        )
        normalized_path = os.path.normcase(os.path.abspath(self.db_name))
        if normalized_path not in type(self)._BOOTSTRAP_READY_PATHS:
            self._init_db()
            type(self)._BOOTSTRAP_READY_PATHS.add(normalized_path)

    def _emit_contract_event(self, signal_name: str, payload: dict) -> None:
        """Emit only after commit; a listener failure cannot roll back clinical data."""
        signal = getattr(self.event_bus, str(signal_name), None)
        if signal is None or not hasattr(signal, "emit"):
            return
        try:
            signal.emit(dict(payload or {}))
        except Exception:
            APP_LOG.exception("No se pudo emitir el evento %s de Admisión", signal_name)

    def notify_detail_sheet_generated(self, attention_id: int) -> None:
        try:
            hybrid_store = getattr(self, "hybrid_store", None)
            if hybrid_store is not None:
                hybrid_store.queue_detail_sheet_generated(int(attention_id))
            with closing(self._connect()) as conn:
                reference = build_attention_event_ref(
                    conn,
                    attention_id=int(attention_id),
                    event_type="HOJA_DETALLE_GENERADA",
                )
            self._emit_contract_event("detail_sheet_generated", reference)
        except Exception:
            APP_LOG.exception(
                "No se pudo notificar la hoja de la atención #%s", attention_id
            )

    def notify_shift_changed(self, turn_id: int) -> None:
        try:
            with closing(self._connect()) as conn:
                reference = build_shift_event_ref(conn, turn_id=int(turn_id))
            self._emit_contract_event("shift_changed", reference)
        except Exception:
            APP_LOG.exception("No se pudo notificar el cambio al turno #%s", turn_id)

    def _connect(self):
        # Una conexión independiente por consumidor; el coordinador global
        # serializa únicamente las transacciones de escritura de V15/sync/PDF.
        # WAL se prepara una vez en bootstrap y no se renegocia en cada connect.
        return connect_local_sqlite(self.db_name, operation="v15-local-write")

    def _init_db(self):
        result = migrate_database(self.db_name, self.backup_manager, APP_LOG)
        journal_mode = prepare_sqlite_database(self.db_name)
        APP_LOG.info("Base de datos lista en esquema v%s: %s", self.SCHEMA_VERSION, result)
        APP_LOG.info(
            "SQLite local privado preparado journal_mode=%s busy_timeout_ms=5000",
            journal_mode,
        )

    def _crear_respaldo_migracion(self, conn, version_actual):
        if version_actual >= self.SCHEMA_VERSION or not os.path.exists(self.db_name):
            return ""
        tiene_datos = conn.execute(
            "SELECT 1 FROM sqlite_master WHERE type='table' AND name='atenciones'"
        ).fetchone()
        if not tiene_datos:
            return ""
        marca = datetime.now().strftime("%Y%m%d_%H%M%S")
        destino = os.path.join(
            BACKUPS_DIR,
            f"pacientes_pre_migracion_v{version_actual}_a_v{self.SCHEMA_VERSION}_{marca}.db",
        )
        with closing(sqlite3.connect(destino)) as copia:
            conn.backup(copia)
        APP_LOG.info("Respaldo previo a migración creado: %s", destino)
        return destino

    @staticmethod
    def _asegurar_columna(cursor, tabla, columna, definicion):
        existentes = {fila[1] for fila in cursor.execute(f"PRAGMA table_info({tabla})")}
        if columna not in existentes:
            cursor.execute(f"ALTER TABLE {tabla} ADD COLUMN {columna} {definicion}")

    def _init_db_legacy_unused(self):
        with closing(self._connect()) as conn:
            cursor = conn.cursor()
            cursor.execute('''
                CREATE TABLE IF NOT EXISTS schema_version (
                    id INTEGER PRIMARY KEY CHECK (id = 1),
                    version INTEGER NOT NULL
                );
            ''')
            cursor.execute("INSERT OR IGNORE INTO schema_version (id, version) VALUES (1, 0);")
            conn.commit()

            cursor.execute('SELECT version FROM schema_version WHERE id = 1;')
            row = cursor.fetchone()
            current_version = int(row[0]) if row else 0

            self._crear_respaldo_migracion(conn, current_version)

            cursor.execute('''
                CREATE TABLE IF NOT EXISTS pacientes (
                    cedula TEXT CHECK(cedula IS NULL OR LENGTH(cedula) = 11),
                    nombre TEXT NOT NULL,
                    telefono TEXT CHECK(telefono IS NULL OR LENGTH(telefono) = 10),
                    direccion TEXT,
                    nacionalidad TEXT,
                    ars TEXT,
                    nss TEXT UNIQUE,
                    PRIMARY KEY (nss)
                )
            ''')
            cursor.execute('''
                CREATE TABLE IF NOT EXISTS atenciones (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    nss TEXT,
                    nombre TEXT NOT NULL,
                    sexo TEXT,
                    edad_num INTEGER,
                    unidad TEXT,
                    cedula TEXT,
                    telefono TEXT,
                    direccion TEXT,
                    nacionalidad TEXT,
                    ars TEXT,
                    hoja TEXT,
                    fecha TEXT,
                    hora TEXT,
                    tipo_atencion TEXT DEFAULT 'EMERGENCIA',
                    created_at TEXT DEFAULT (datetime('now','localtime')),
                    updated_at TEXT,
                    turno_id INTEGER,
                    nss_clean TEXT,
                    cedula_clean TEXT,
                    telefono_clean TEXT
                );
            ''')
            cursor.execute('''
                CREATE TABLE IF NOT EXISTS turnos (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    fecha_inicio TEXT NOT NULL,
                    fecha_fin TEXT NOT NULL,
                    representante TEXT NOT NULL,
                    tipo_turno TEXT NOT NULL,
                    estado TEXT NOT NULL DEFAULT 'ABIERTO',
                    fecha_cierre TEXT,
                    created_at TEXT DEFAULT (datetime('now','localtime')),
                    UNIQUE(fecha_inicio, tipo_turno)
                );
            ''')
            cursor.execute('''
                CREATE TABLE IF NOT EXISTS atenciones_auditoria (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    atencion_id INTEGER NOT NULL,
                    accion TEXT NOT NULL,
                    motivo TEXT,
                    usuario TEXT,
                    snapshot_json TEXT NOT NULL,
                    created_at TEXT DEFAULT (datetime('now','localtime'))
                );
            ''')

            for columna, definicion in (
                ("tipo_atencion", "TEXT DEFAULT 'EMERGENCIA'"),
                ("edad_num", "INTEGER"),
                ("unidad", "TEXT"),
                ("updated_at", "TEXT"),
                ("turno_id", "INTEGER"),
                ("nss_clean", "TEXT"),
                ("cedula_clean", "TEXT"),
                ("telefono_clean", "TEXT"),
            ):
                self._asegurar_columna(cursor, "atenciones", columna, definicion)

            cursor.execute("""
                UPDATE atenciones
                SET tipo_atencion = COALESCE(NULLIF(TRIM(tipo_atencion), ''), 'EMERGENCIA'),
                    nss_clean = REPLACE(REPLACE(REPLACE(IFNULL(nss,''),'-',''),' ',''),'.',''),
                    cedula_clean = REPLACE(REPLACE(REPLACE(IFNULL(cedula,''),'-',''),' ',''),'.',''),
                    telefono_clean = REPLACE(REPLACE(REPLACE(IFNULL(telefono,''),'-',''),' ',''),'.','')
                WHERE tipo_atencion IS NULL OR TRIM(tipo_atencion) = ''
                   OR nss_clean IS NULL OR cedula_clean IS NULL OR telefono_clean IS NULL
            """)

            indices = (
                "CREATE INDEX IF NOT EXISTS idx_atenciones_fecha ON atenciones(fecha)",
                "CREATE INDEX IF NOT EXISTS idx_atenciones_nombre ON atenciones(nombre)",
                "CREATE INDEX IF NOT EXISTS idx_atenciones_nss ON atenciones(nss)",
                "CREATE INDEX IF NOT EXISTS idx_atenciones_cedula ON atenciones(cedula)",
                "CREATE INDEX IF NOT EXISTS idx_atenciones_ars ON atenciones(ars)",
                "CREATE INDEX IF NOT EXISTS idx_atenciones_hoja ON atenciones(hoja)",
                "CREATE INDEX IF NOT EXISTS idx_atenciones_tipo ON atenciones(tipo_atencion)",
                "CREATE INDEX IF NOT EXISTS idx_atenciones_fecha_id ON atenciones(fecha, id)",
                "CREATE INDEX IF NOT EXISTS idx_atenciones_fecha_hora ON atenciones(fecha, hora)",
                "CREATE INDEX IF NOT EXISTS idx_atenciones_telefono ON atenciones(telefono)",
                "CREATE INDEX IF NOT EXISTS idx_atenciones_created_at ON atenciones(created_at)",
                "CREATE INDEX IF NOT EXISTS idx_atenciones_turno_id ON atenciones(turno_id)",
                "CREATE INDEX IF NOT EXISTS idx_atenciones_nss_clean ON atenciones(nss_clean)",
                "CREATE INDEX IF NOT EXISTS idx_atenciones_cedula_clean ON atenciones(cedula_clean)",
                "CREATE INDEX IF NOT EXISTS idx_atenciones_telefono_clean ON atenciones(telefono_clean)",
                "CREATE INDEX IF NOT EXISTS idx_turnos_estado ON turnos(estado)",
                "CREATE INDEX IF NOT EXISTS idx_auditoria_atencion ON atenciones_auditoria(atencion_id)",
                "CREATE UNIQUE INDEX IF NOT EXISTS uq_atencion_turno_cedula ON atenciones(turno_id, cedula_clean) "
                "WHERE turno_id IS NOT NULL AND LENGTH(cedula_clean) = 11",
                "CREATE UNIQUE INDEX IF NOT EXISTS uq_atencion_turno_nss ON atenciones(turno_id, nss_clean) "
                "WHERE turno_id IS NOT NULL AND LENGTH(nss_clean) >= 3",
            )
            for sql in indices:
                cursor.execute(sql)

            cursor.execute(
                "UPDATE schema_version SET version = ? WHERE id = 1",
                (self.SCHEMA_VERSION,),
            )
            conn.commit()

            integrity = cursor.execute("PRAGMA integrity_check").fetchone()
            if not integrity or integrity[0] != "ok":
                raise sqlite3.DatabaseError(f"Falló integrity_check: {integrity}")
            APP_LOG.info("Base de datos lista en esquema v%s", self.SCHEMA_VERSION)

    def listar_representantes(self):
        with closing(self._connect()) as conn:
            filas = conn.execute(
                "SELECT DISTINCT representante FROM turnos "
                "WHERE TRIM(IFNULL(representante,'')) <> '' ORDER BY representante"
            ).fetchall()
        return [
            limpiar_nombre_representante(fila[0])
            for fila in filas
            if es_representante_valido(fila[0])
        ]

    def actualizar_representante_turno(self, turno_id: int, representante: str) -> bool:
        representante = limpiar_nombre_representante(representante)
        if not es_representante_valido(representante):
            raise ValueError(
                "Escriba un nombre de representante válido; 'No disponible' no se admite."
            )
        with closing(self._connect()) as conn:
            cur = conn.execute(
                """
                UPDATE turnos SET representante=?,updated_at=datetime('now','localtime')
                WHERE id=?
                """,
                (representante, int(turno_id)),
            )
            conn.commit()
            return cur.rowcount == 1

    def obtener_o_crear_turno(
        self, turno_cfg, conn=None, *, administrative_override=False
    ):
        is_administrative_override = bool(
            administrative_override or (turno_cfg or {}).get("administrative_override")
        )
        if not turno_cfg or (
            not is_administrative_override and not turno_config_es_vigente(turno_cfg)
        ):
            raise TurnoNoVigenteError("Debe abrir un turno vigente antes de registrar atenciones.")
        inicio, fin = obtener_rango_turno_efectivo(turno_cfg)
        representante = limpiar_nombre_representante(turno_cfg.get("representante", ""))
        if not es_representante_valido(representante):
            raise ValueError("El turno requiere un representante válido.")
        tipo = normalizar_turno_codigo(turno_cfg.get("turno_codigo", "8AM_8AM"))
        fecha_base = turno_cfg["fecha_base"]
        dia_inicio = datetime.combine(fecha_base, time(8, 0))
        dia_fin = dia_inicio + timedelta(days=1)
        propia = conn is None
        conexion = conn or self._connect()
        try:
            conexion.execute(
                """
                INSERT INTO dias_operativos(fecha_base,fecha_inicio,fecha_fin,estado,origen,requiere_revision)
                VALUES (?,?,?,'ABIERTO','OPERATIVO',0)
                ON CONFLICT(fecha_base) DO UPDATE SET
                    fecha_inicio=excluded.fecha_inicio,
                    fecha_fin=excluded.fecha_fin,
                    estado='ABIERTO'
                """,
                (
                    fecha_base.isoformat(),
                    dia_inicio.strftime("%Y-%m-%d %H:%M:%S"),
                    dia_fin.strftime("%Y-%m-%d %H:%M:%S"),
                ),
            )
            dia_row = conexion.execute(
                "SELECT id FROM dias_operativos WHERE fecha_base=?",
                (fecha_base.isoformat(),),
            ).fetchone()
            dia_id = int(dia_row[0])
            inicio_txt = inicio.strftime("%Y-%m-%d %H:%M:%S")
            fin_txt = fin.strftime("%Y-%m-%d %H:%M:%S")
            cierre_txt = (turno_cfg.get("inicio_real_dt") or datetime.now()).strftime(
                "%Y-%m-%d %H:%M:%S"
            )
            conexion.execute(
                """
                UPDATE turnos SET estado='CERRADO', fecha_cierre=?, updated_at=datetime('now','localtime')
                WHERE estado='ABIERTO'
                  AND NOT (dia_operativo_id=? AND fecha_inicio=? AND tipo_turno=?)
                """,
                (cierre_txt, dia_id, inicio_txt, tipo),
            )
            conexion.execute(
                "UPDATE dias_operativos SET estado='CERRADO' WHERE estado='ABIERTO' AND id<>?",
                (dia_id,),
            )
            conexion.execute('''
                INSERT INTO turnos (
                    dia_operativo_id,fecha_inicio,fecha_fin,fecha_inicio_real,representante,
                    tipo_turno,estado,origen,requiere_revision
                ) VALUES (?, ?, ?, ?, ?, ?, 'ABIERTO', 'OPERATIVO', 0)
                ON CONFLICT(dia_operativo_id,fecha_inicio,tipo_turno) DO UPDATE SET
                    fecha_fin=excluded.fecha_fin,
                    fecha_inicio_real=excluded.fecha_inicio_real,
                    representante=excluded.representante,
                    estado='ABIERTO',
                    fecha_cierre=NULL,
                    updated_at=datetime('now','localtime')
            ''', (
                dia_id,
                inicio_txt,
                fin_txt,
                (turno_cfg.get("inicio_real_dt") or inicio).strftime("%Y-%m-%d %H:%M:%S"),
                representante,
                tipo,
            ))
            fila = conexion.execute(
                "SELECT id FROM turnos WHERE dia_operativo_id=? AND fecha_inicio=? AND tipo_turno=?",
                (dia_id, inicio_txt, tipo),
            ).fetchone()
            if propia:
                conexion.commit()
            return int(fila[0]) if fila else None
        finally:
            if propia:
                conexion.close()

    def obtener_contexto_turno(self, turno_cfg, conn=None):
        turno_id = self.obtener_o_crear_turno(turno_cfg, conn=conn)
        propia = conn is None
        conexion = conn or self._connect()
        try:
            fila = conexion.execute(
                "SELECT id, dia_operativo_id, fecha_inicio, fecha_fin, representante, tipo_turno FROM turnos WHERE id=?",
                (turno_id,),
            ).fetchone()
            if not fila:
                raise TurnoNoVigenteError("No se pudo resolver el turno vigente.")
            return {
                "turno_id": int(fila[0]),
                "dia_operativo_id": int(fila[1]),
                "fecha_inicio": fila[2],
                "fecha_fin": fila[3],
                "representante": fila[4],
                "tipo_turno": fila[5],
            }
        finally:
            if propia:
                conexion.close()

    def buscar_contexto_turno_existente(self, turno_cfg):
        if not turno_cfg:
            return None
        inicio, _fin = obtener_rango_turno_efectivo(turno_cfg)
        with closing(self._connect()) as conn:
            conn.row_factory = sqlite3.Row
            row = conn.execute(
                """
                SELECT t.id AS turno_id,t.dia_operativo_id,t.fecha_inicio,t.fecha_fin,
                       t.representante,t.tipo_turno,t.estado,d.fecha_base
                FROM turnos t
                JOIN dias_operativos d ON d.id=t.dia_operativo_id
                WHERE d.fecha_base=? AND t.fecha_inicio=? AND t.tipo_turno=?
                ORDER BY t.id DESC LIMIT 1
                """,
                (
                    turno_cfg["fecha_base"].isoformat(),
                    inicio.strftime("%Y-%m-%d %H:%M:%S"),
                    normalizar_turno_codigo(turno_cfg.get("turno_codigo", "8AM_8AM")),
                ),
            ).fetchone()
            if not row:
                # La hora real puede variar si la configuración fue recuperada o
                # guardada nuevamente. El día y el tipo identifican el mismo turno.
                row = conn.execute(
                    """
                    SELECT t.id AS turno_id,t.dia_operativo_id,t.fecha_inicio,t.fecha_fin,
                           t.representante,t.tipo_turno,t.estado,d.fecha_base
                    FROM turnos t
                    JOIN dias_operativos d ON d.id=t.dia_operativo_id
                    WHERE d.fecha_base=? AND t.tipo_turno=?
                    ORDER BY CASE WHEN t.estado='ABIERTO' THEN 0 ELSE 1 END,t.id DESC
                    LIMIT 1
                    """,
                    (
                        turno_cfg["fecha_base"].isoformat(),
                        normalizar_turno_codigo(turno_cfg.get("turno_codigo", "8AM_8AM")),
                    ),
                ).fetchone()
        return dict(row) if row else None

    def cerrar_turno_existente(
        self,
        turno_cfg,
        momento_cierre=None,
        *,
        actor="",
        actor_role="",
        session_id="",
    ):
        contexto = self.buscar_contexto_turno_existente(turno_cfg)
        if not contexto:
            return False
        cierre = (momento_cierre or datetime.now()).strftime("%Y-%m-%d %H:%M:%S")
        with closing(self._connect()) as conn:
            conn.execute("BEGIN IMMEDIATE")
            conn.execute(
                """
                UPDATE turnos SET estado='CERRADO',fecha_cierre=?,updated_at=datetime('now','localtime')
                WHERE id=?
                """,
                (cierre, int(contexto["turno_id"])),
            )
            abiertos = int(
                conn.execute(
                    "SELECT COUNT(*) FROM turnos WHERE dia_operativo_id=? AND estado='ABIERTO'",
                    (int(contexto["dia_operativo_id"]),),
                ).fetchone()[0]
            )
            if not abiertos:
                conn.execute(
                    "UPDATE dias_operativos SET estado='CERRADO' WHERE id=?",
                    (int(contexto["dia_operativo_id"]),),
                )
            event_ref = enqueue_shift_closure_event(
                conn,
                shift=contexto,
                closed_at=cierre,
                actor=actor or contexto.get("representante") or "Sistema",
                actor_role=actor_role,
                session_id=session_id,
            )
            conn.commit()
        self._emit_contract_event("shift_closed", event_ref)
        return True

    def _dedupe_rows_keep_latest(self, rows):
        latest = {}
        for row in sorted(rows, key=lambda r: int(r.get("id", 0))):
            latest[patient_identity_key_from_row(row)] = row
        result = list(latest.values())
        result.sort(key=lambda r: int(r.get("id", 0)), reverse=True)
        return result

    def buscar_atencion_en_turno(
        self,
        nss,
        cedula,
        inicio_turno,
        fin_turno,
        turno_id=None,
        nombre="",
        telefono="",
        dia_operativo_id=None,
    ):
        nss_limpio = re.sub(r"\D", "", (nss or "").strip().upper())
        cedula_limpia = re.sub(r"\D", "", (cedula or "").strip())
        telefono_limpio = re.sub(r"\D", "", (telefono or "").strip())
        nombre_limpio = normalizar_nombre_clave(nombre)

        with closing(self._connect()) as conn:
            conn.row_factory = sqlite3.Row
            paciente_id = None
            if is_valid_cedula_key(cedula_limpia):
                row = conn.execute(
                    """
                    SELECT i.paciente_id
                    FROM paciente_identificadores i
                    JOIN pacientes p ON p.id=i.paciente_id
                    WHERE i.tipo='CEDULA' AND i.valor_normalizado=? AND i.activo=1
                      AND p.estado='ACTIVO'
                    ORDER BY i.conflicto,COALESCE(p.updated_at,p.created_at) DESC,p.id DESC
                    LIMIT 1
                    """,
                    (cedula_limpia,),
                ).fetchone()
                paciente_id = int(row["paciente_id"]) if row else None
            elif is_valid_nss_key(nss_limpio):
                rows = conn.execute(
                    """
                    SELECT p.id,p.nombre,p.telefono_clean,p.telefono
                    FROM paciente_identificadores i
                    JOIN pacientes p ON p.id=i.paciente_id
                    WHERE i.tipo='NSS' AND i.valor_normalizado=? AND i.activo=1
                      AND p.estado='ACTIVO'
                    ORDER BY i.conflicto,COALESCE(p.updated_at,p.created_at) DESC,p.id DESC
                    """,
                    (nss_limpio,),
                ).fetchall()
                for row in rows:
                    mismo_nombre = bool(
                        nombre_limpio
                        and nombre_limpio == normalizar_nombre_clave(row["nombre"])
                    )
                    telefono_row = re.sub(
                        r"\D", "", row["telefono_clean"] or row["telefono"] or ""
                    )
                    mismo_telefono = bool(
                        len(telefono_limpio) == 10 and telefono_limpio == telefono_row
                    )
                    if mismo_nombre or mismo_telefono:
                        paciente_id = int(row["id"])
                        break
            if paciente_id is None and not (nombre_limpio and len(telefono_limpio) == 10):
                return None

            if dia_operativo_id is None and turno_id is not None:
                turno = conn.execute("SELECT dia_operativo_id FROM turnos WHERE id=?", (int(turno_id),)).fetchone()
                dia_operativo_id = int(turno[0]) if turno else None

            parametros = []
            if dia_operativo_id is not None:
                rango_sql = "dia_operativo_id = ?"
                parametros.append(int(dia_operativo_id))
            else:
                rango_sql = "datetime(created_at) >= datetime(?) AND datetime(created_at) < datetime(?)"
                parametros.extend([
                    inicio_turno.strftime("%Y-%m-%d %H:%M:%S"),
                    fin_turno.strftime("%Y-%m-%d %H:%M:%S"),
                ])

            if paciente_id is not None:
                identidad_sql = "paciente_id = ?"
                parametros.insert(0, int(paciente_id))
            else:
                identidad_sql = "UPPER(TRIM(nombre)) = ? AND telefono_clean = ?"
                parametros[0:0] = [nombre_limpio, telefono_limpio]

            sql = f'''
                SELECT * FROM atenciones
                WHERE estado='ACTIVA' AND {identidad_sql} AND {rango_sql}
                ORDER BY id DESC LIMIT 1
            '''
            fila = conn.execute(sql, parametros).fetchone()
        return dict(fila) if fila else None

    def buscar_paciente(self, cedula):
        cedula_limpia = re.sub(r"\D", "", (cedula or ""))
        if not is_valid_cedula_key(cedula_limpia):
            return None
        with closing(self._connect()) as conn:
            conn.row_factory = sqlite3.Row
            return conn.execute(
                """
                SELECT p.* FROM pacientes p
                JOIN paciente_identificadores i ON i.paciente_id=p.id
                WHERE i.tipo='CEDULA' AND i.valor_normalizado=? AND i.activo=1
                ORDER BY p.updated_at DESC, p.id DESC LIMIT 1
                """,
                (cedula_limpia,),
            ).fetchone()

    def buscar_por_nss(self, nss):
        nss_limpio = re.sub(r"\D", "", (nss or ""))
        if not is_valid_nss_key(nss_limpio):
            return None
        with closing(self._connect()) as conn:
            conn.row_factory = sqlite3.Row
            return conn.execute(
                """
                SELECT p.* FROM pacientes p
                JOIN paciente_identificadores i ON i.paciente_id=p.id
                WHERE i.tipo='NSS' AND i.valor_normalizado=? AND i.activo=1
                ORDER BY p.updated_at DESC, p.id DESC LIMIT 1
                """,
                (nss_limpio,),
            ).fetchone()

    def _resolver_o_crear_paciente_conn(self, conn, datos, revisiones_nss=None):
        conn.row_factory = sqlite3.Row
        revisiones_nss = revisiones_nss if revisiones_nss is not None else []
        nss = (datos.get("NSS", "") or "").strip().upper()
        cedula = (datos.get("Cédula", "") or "").strip()
        nss_clean = re.sub(r"\D", "", nss)
        cedula_clean = re.sub(r"\D", "", cedula)
        telefono = (datos.get("Teléfono", "") or "").strip()
        telefono_clean = re.sub(r"\D", "", telefono)
        nombre = (datos.get("Nombre", "") or "").strip() or "SIN NOMBRE"
        nombre_clean = normalizar_nombre_clave(nombre)
        cedula_valida = is_valid_cedula_key(cedula_clean)
        nss_valido = is_valid_nss_key(nss_clean)

        def propietarios(tipo, valor):
            if not valor:
                return []
            return conn.execute(
                """
                SELECT p.*,i.conflicto
                FROM paciente_identificadores i
                JOIN pacientes p ON p.id=i.paciente_id
                WHERE i.tipo=? AND i.valor_normalizado=? AND i.activo=1
                  AND p.estado='ACTIVO'
                ORDER BY i.conflicto,COALESCE(p.updated_at,p.created_at) DESC,p.id DESC
                """,
                (tipo, valor),
            ).fetchall()

        def coincide_demografia(row):
            mismo_nombre = bool(
                nombre_clean
                and nombre_clean != "SIN NOMBRE"
                and nombre_clean == normalizar_nombre_clave(row["nombre"])
            )
            mismo_telefono = bool(
                len(telefono_clean) == 10
                and telefono_clean == re.sub(r"\D", "", row["telefono"] or "")
            )
            return mismo_nombre or mismo_telefono

        cedula_rows = propietarios("CEDULA", cedula_clean) if cedula_valida else []
        nss_rows = propietarios("NSS", nss_clean) if nss_valido else []
        patient_id = int(cedula_rows[0]["id"]) if cedula_rows else None

        # Una ficha creada inicialmente sin cédula puede completarse después.
        if patient_id is None and nss_rows:
            coincidencias = [row for row in nss_rows if coincide_demografia(row)]
            if coincidencias:
                patient_id = int(coincidencias[0]["id"])

        revision_nss = bool(not cedula_valida and nss_valido and nss_rows and patient_id is None)
        referencia_id = int(nss_rows[0]["id"]) if revision_nss else None

        direccion = (datos.get("Dirección", "") or "").strip()
        nacionalidad = (datos.get("Nacionalidad", "") or "").strip()
        ars_canonico = normalizar_seguro(datos.get("Aseguradora (ARS)", ""), nss)
        cedula_db = cedula_clean if cedula_valida else None
        nss_db = nss_clean if nss_valido else None
        telefono_db = telefono_clean if len(telefono_clean) == 10 else None

        if patient_id is not None:
            if cedula_valida:
                # La cédula manda: el NSS nuevo reemplaza todos los NSS anteriores
                # de esta ficha, incluso cuando el campo se deja vacío.
                conn.execute(
                    "DELETE FROM paciente_identificadores WHERE paciente_id=? AND tipo='NSS'",
                    (patient_id,),
                )
            conn.execute(
                """
                UPDATE pacientes SET
                    nombre=?,cedula=COALESCE(?,cedula),telefono=COALESCE(?,telefono),
                    direccion=COALESCE(NULLIF(?,''),direccion),
                    nacionalidad=COALESCE(NULLIF(?,''),nacionalidad),ars=?,
                    nss=?,nss_clean=?,cedula_clean=COALESCE(?,cedula_clean),
                    telefono_clean=COALESCE(?,telefono_clean),provisional=0,
                    requiere_revision=0,updated_at=datetime('now','localtime')
                WHERE id=?
                """,
                (
                    nombre,cedula_db,telefono_db,direccion,nacionalidad,ars_canonico,
                    nss_db,nss_db,cedula_db,telefono_db,patient_id,
                ),
            )
        else:
            provisional = int(not cedula_valida and not nss_valido)
            cursor = conn.execute(
                """
                INSERT INTO pacientes(
                    nombre,cedula,telefono,direccion,nacionalidad,ars,nss,
                    nss_clean,cedula_clean,telefono_clean,provisional,requiere_revision
                ) VALUES (?,?,?,?,?,?,?,?,?,?,?,?)
                """,
                (
                    nombre,cedula_db,telefono_db,direccion,nacionalidad,ars_canonico,
                    nss_db,nss_db,cedula_db,telefono_db,provisional,int(revision_nss),
                ),
            )
            patient_id = int(cursor.lastrowid)

        identificadores = []
        if cedula_valida:
            identificadores.append(("CEDULA", cedula_clean, 0))
        if nss_valido:
            identificadores.append(("NSS", nss_clean, int(revision_nss)))
        if revision_nss:
            conn.execute(
                "UPDATE paciente_identificadores SET conflicto=1 "
                "WHERE tipo='NSS' AND valor_normalizado=? AND activo=1",
                (nss_clean,),
            )
        for tipo, valor, conflicto in identificadores:
            conn.execute(
                """
                INSERT INTO paciente_identificadores(
                    paciente_id,tipo,valor_normalizado,activo,conflicto
                ) VALUES (?,?,?,1,?)
                ON CONFLICT(paciente_id,tipo,valor_normalizado)
                DO UPDATE SET activo=1,conflicto=excluded.conflicto
                """,
                (patient_id,tipo,valor,conflicto),
            )

        if revision_nss:
            revisiones_nss.append(
                {
                    "nss": nss_clean,
                    "paciente_nuevo_id": patient_id,
                    "paciente_referencia_id": referencia_id,
                    "detalle": (
                        "El NSS fue registrado sin cédula para datos demográficos diferentes. "
                        "La atención continuó y requiere revisión administrativa."
                    ),
                }
            )
        return patient_id

    def guardar_paciente(self, datos):
        with closing(self._connect()) as conn:
            conn.execute("BEGIN IMMEDIATE")
            revisiones_nss = []
            patient_id = self._resolver_o_crear_paciente_conn(conn, datos, revisiones_nss)
            for revision in revisiones_nss:
                conn.execute(
                    """
                    INSERT INTO nss_conflictos(
                        nss_normalizado,paciente_nuevo_id,paciente_referencia_id,detalle
                    ) VALUES (?,?,?,?)
                    """,
                    (
                        revision["nss"],revision["paciente_nuevo_id"],
                        revision["paciente_referencia_id"],revision["detalle"],
                    ),
                )
            conn.commit()
            return patient_id

    def _registrar_auditoria_conn(
        self,
        conn,
        atencion_id,
        accion,
        motivo,
        usuario,
        snapshot_before,
        snapshot_after=None,
        actor_rol="ADMISION",
    ):
        previous_row = conn.execute(
            "SELECT event_hash FROM atenciones_auditoria WHERE event_hash IS NOT NULL ORDER BY id DESC LIMIT 1"
        ).fetchone()
        previous_hash = str(previous_row[0] or "") if previous_row else ""
        before_json = json.dumps(snapshot_before or {}, ensure_ascii=False, sort_keys=True, default=str)
        after_json = json.dumps(snapshot_after, ensure_ascii=False, sort_keys=True, default=str) if snapshot_after is not None else None
        payload = {
            "atencion_id": int(atencion_id) if atencion_id is not None else None,
            "accion": accion,
            "motivo": (motivo or "").strip(),
            "usuario": limpiar_nombre_representante(usuario),
            "actor_rol": actor_rol,
            "snapshot_before": before_json,
            "snapshot_after": after_json,
            "previous_hash": previous_hash,
            "workstation": platform.node(),
        }
        event_hash = hashlib.sha256(
            json.dumps(payload, ensure_ascii=False, sort_keys=True, separators=(",", ":")).encode("utf-8")
        ).hexdigest()
        conn.execute(
            """
            INSERT INTO atenciones_auditoria(
                atencion_id,accion,motivo,usuario,actor_rol,snapshot_json,snapshot_after_json,
                previous_hash,event_hash,workstation
            ) VALUES (?,?,?,?,?,?,?,?,?,?)
            """,
            (
                payload["atencion_id"], payload["accion"], payload["motivo"], payload["usuario"],
                actor_rol, before_json, after_json, previous_hash, event_hash, platform.node(),
            ),
        )
        return event_hash

    def guardar_atencion(self, datos, hoja, turno_cfg=None):
        administrative_override = bool((turno_cfg or {}).get("administrative_override"))
        if not turno_cfg or (
            not administrative_override and not turno_config_es_vigente(turno_cfg)
        ):
            raise TurnoNoVigenteError("El turno no existe o está vencido. Abra el turno actual.")
        ars_canonico = normalizar_seguro(
            datos.get('Aseguradora (ARS)', ''),
            datos.get('NSS', '')
        )

        nss = (datos.get('NSS', '') or '').strip().upper()
        cedula = (datos.get('Cédula', '') or '').strip()
        tipo_atencion = (datos.get("TipoAtencion") or datos.get("tipo_atencion") or "EMERGENCIA").strip().upper()
        if tipo_atencion not in ("EMERGENCIA", "URGENCIA", "CONSULTA"):
            tipo_atencion = "EMERGENCIA"

        with closing(self._connect()) as conn:
            conn.row_factory = sqlite3.Row
            conn.execute("BEGIN IMMEDIATE")
            cur = conn.cursor()
            contexto = self.obtener_contexto_turno(turno_cfg, conn=conn)
            turno_id = contexto["turno_id"]
            dia_operativo_id = contexto["dia_operativo_id"]
            nss_clean = re.sub(r"\D", "", nss)
            cedula_clean = re.sub(r"\D", "", cedula)
            telefono = (datos.get('Teléfono', '') or '').strip()
            telefono_clean = re.sub(r"\D", "", telefono)
            es_reingreso = int(bool(datos.get("EsReingreso") or datos.get("es_reingreso")))
            atencion_origen_id = datos.get("AtencionOrigenId") or datos.get("atencion_origen_id")
            motivo_reingreso = (datos.get("MotivoReingreso") or datos.get("motivo_reingreso") or "").strip()
            autorizado_por = limpiar_nombre_representante(
                datos.get("AutorizadoPor") or datos.get("autorizado_por") or ""
            )
            revisiones_nss = []
            if es_reingreso and (not atencion_origen_id or not motivo_reingreso or not autorizado_por):
                raise ValueError("El reingreso requiere atención original, motivo y autorización.")
            if es_reingreso:
                origen = cur.execute(
                    """
                    SELECT * FROM atenciones
                    WHERE id=? AND dia_operativo_id=? AND estado='ACTIVA'
                    """,
                    (int(atencion_origen_id), dia_operativo_id),
                ).fetchone()
                if not origen:
                    raise ValueError("La atención original no está activa en este día operativo.")
                paciente_id = int(origen["paciente_id"])
            else:
                paciente_id = self._resolver_o_crear_paciente_conn(
                    conn, datos, revisiones_nss
                )

            try:
                cur.execute('''
                    INSERT INTO atenciones (
                        paciente_id,dia_operativo_id,turno_id,nss,nombre,sexo,edad_num,unidad,
                        cedula,telefono,direccion,nacionalidad,ars,hoja,fecha,hora,tipo_atencion,
                        estado,es_reingreso,atencion_origen_id,motivo_reingreso,autorizado_por,
                        identidad_estado,requiere_revision,nss_clean,cedula_clean,telefono_clean
                    ) VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,'ACTIVA',?,?,?,?,?,?,?,?,?)
                ''', (
                    paciente_id,
                    dia_operativo_id,
                    turno_id,
                    nss,
                    datos.get('Nombre', ''),
                    datos.get('Sexo', ''),
                    int(datos.get('Edad_num', 0) or 0),
                    datos.get('Unidad', ''),
                    cedula,
                    telefono,
                    datos.get('Dirección', ''),
                    datos.get('Nacionalidad', ''),
                    ars_canonico,
                    hoja,
                    datos.get('Fecha', ''),
                    datos.get('Hora', ''),
                    tipo_atencion,
                    es_reingreso,
                    int(atencion_origen_id) if atencion_origen_id else None,
                    motivo_reingreso or None,
                    autorizado_por or None,
                    "NSS_EN_REVISION" if revisiones_nss else "VALIDADA",
                    int(bool(revisiones_nss)),
                    nss_clean or None,
                    cedula_clean or None,
                    telefono_clean or None,
                ))
            except sqlite3.IntegrityError as exc:
                conn.rollback()
                if "uq_atencion_dia_paciente" in str(exc) or "UNIQUE constraint failed" in str(exc):
                    raise sqlite3.IntegrityError(
                        "Este paciente ya tiene una atención activa en el día operativo."
                    ) from exc
                raise
            atencion_id = int(cur.lastrowid)
            for revision in revisiones_nss:
                conn.execute(
                    """
                    INSERT INTO nss_conflictos(
                        nss_normalizado,paciente_nuevo_id,paciente_referencia_id,
                        atencion_id,detalle
                    ) VALUES (?,?,?,?,?)
                    """,
                    (
                        revision["nss"],revision["paciente_nuevo_id"],
                        revision["paciente_referencia_id"],atencion_id,
                        revision["detalle"],
                    ),
                )
            snapshot = dict(conn.execute("SELECT * FROM atenciones WHERE id=?", (atencion_id,)).fetchone())
            self._registrar_auditoria_conn(
                conn,
                atencion_id,
                "CREACION",
                "Registro desde formulario principal",
                self.session_context.audit_actor,
                {},
                snapshot,
                self.session_context.role,
            )
            cur.execute(
                """INSERT INTO trabajos_salida(
                       atencion_id,pdf_estado,impresion_estado
                   ) VALUES (?,'NO_APLICA','NO_APLICA')""",
                (atencion_id,),
            )
            event_ref = enqueue_billing_event(
                conn,
                attention_id=atencion_id,
                actor=self.session_context.audit_actor,
                actor_role=self.session_context.role,
                session_id=self.session_context.session_id,
            )
            conn.commit()
            self._emit_contract_event("attention_created", event_ref)
            return atencion_id

    def obtener_revision_nss_atencion(self, atencion_id):
        with closing(self._connect()) as conn:
            row = conn.execute(
                "SELECT id FROM nss_conflictos WHERE atencion_id=? AND estado='PENDIENTE' LIMIT 1",
                (int(atencion_id),),
            ).fetchone()
        return int(row[0]) if row else None

    def listar_revisiones_nss(self, solo_pendientes=True, limite=500):
        where = "WHERE c.estado='PENDIENTE'" if solo_pendientes else ""
        with closing(self._connect()) as conn:
            conn.row_factory = sqlite3.Row
            rows = conn.execute(
                f"""
                SELECT c.*,pn.nombre AS nombre_nuevo,pr.nombre AS nombre_referencia,
                       a.fecha,a.hora,a.estado AS atencion_estado
                FROM nss_conflictos c
                LEFT JOIN pacientes pn ON pn.id=c.paciente_nuevo_id
                LEFT JOIN pacientes pr ON pr.id=c.paciente_referencia_id
                LEFT JOIN atenciones a ON a.id=c.atencion_id
                {where}
                ORDER BY CASE WHEN c.estado='PENDIENTE' THEN 0 ELSE 1 END,c.id DESC
                LIMIT ?
                """,
                (max(1, int(limite)),),
            ).fetchall()
        return [dict(row) for row in rows]

    def resolver_revision_nss(self, revision_id, resolucion, actor, motivo):
        resolucion = str(resolucion or "").strip().upper()
        if resolucion not in {
            "MANTENER_AMBOS", "DESVINCULAR_NSS", "FUSIONAR_CON_EXISTENTE"
        }:
            raise ValueError("Resolución NSS desconocida.")
        actor = limpiar_nombre_representante(actor)
        motivo = str(motivo or "").strip()
        if not actor or len(motivo) < 8:
            raise ValueError("Indique el responsable y un motivo de al menos 8 caracteres.")

        with closing(self._connect()) as conn:
            conn.row_factory = sqlite3.Row
            conn.execute("BEGIN IMMEDIATE")
            revision = conn.execute(
                "SELECT * FROM nss_conflictos WHERE id=?",
                (int(revision_id),),
            ).fetchone()
            if not revision or revision["estado"] != "PENDIENTE":
                conn.rollback()
                raise ValueError("La revisión ya no está pendiente.")
            nuevo_id = revision["paciente_nuevo_id"]
            referencia_id = revision["paciente_referencia_id"]
            nss = revision["nss_normalizado"]

            if resolucion == "MANTENER_AMBOS":
                pass
            elif resolucion == "DESVINCULAR_NSS":
                if nuevo_id:
                    conn.execute(
                        "DELETE FROM paciente_identificadores "
                        "WHERE paciente_id=? AND tipo='NSS' AND valor_normalizado=?",
                        (int(nuevo_id), nss),
                    )
                    conn.execute(
                        "UPDATE pacientes SET nss=NULL,nss_clean=NULL,requiere_revision=0,"
                        "updated_at=datetime('now','localtime') "
                        "WHERE id=? AND nss_clean=?",
                        (int(nuevo_id), nss),
                    )
                if revision["atencion_id"]:
                    aid = int(revision["atencion_id"])
                    before = conn.execute("SELECT * FROM atenciones WHERE id=?", (aid,)).fetchone()
                    conn.execute(
                        "UPDATE atenciones SET nss=NULL,nss_clean=NULL,identidad_estado='VALIDADA',"
                        "requiere_revision=0,updated_at=datetime('now','localtime') WHERE id=?",
                        (aid,),
                    )
                    after = conn.execute("SELECT * FROM atenciones WHERE id=?", (aid,)).fetchone()
                    if before and after:
                        self._registrar_auditoria_conn(
                            conn,aid,"CORRECCION_NSS",motivo,actor,dict(before),dict(after),"ADMINISTRADOR"
                        )
            else:
                if not nuevo_id or not referencia_id:
                    conn.rollback()
                    raise ValueError("No existen ambas fichas para realizar la fusión.")
                atenciones = conn.execute(
                    "SELECT * FROM atenciones WHERE paciente_id=? ORDER BY id",
                    (int(nuevo_id),),
                ).fetchall()
                for atencion in atenciones:
                    aid = int(atencion["id"])
                    existente = conn.execute(
                        """
                        SELECT id FROM atenciones
                        WHERE paciente_id=? AND dia_operativo_id=? AND estado='ACTIVA'
                          AND es_reingreso=0 AND id<>? ORDER BY id LIMIT 1
                        """,
                        (int(referencia_id),atencion["dia_operativo_id"],aid),
                    ).fetchone()
                    conn.execute(
                        """
                        UPDATE atenciones SET paciente_id=?,es_reingreso=?,atencion_origen_id=?,
                            motivo_reingreso=?,autorizado_por=?,identidad_estado='VALIDADA',
                            requiere_revision=0,updated_at=datetime('now','localtime')
                        WHERE id=?
                        """,
                        (
                            int(referencia_id),int(bool(existente)),
                            int(existente[0]) if existente else atencion["atencion_origen_id"],
                            motivo if existente else atencion["motivo_reingreso"],
                            actor if existente else atencion["autorizado_por"],aid,
                        ),
                    )
                    after = conn.execute("SELECT * FROM atenciones WHERE id=?", (aid,)).fetchone()
                    self._registrar_auditoria_conn(
                        conn,aid,"FUSION_NSS",motivo,actor,dict(atencion),dict(after),"ADMINISTRADOR"
                    )
                conn.execute(
                    "DELETE FROM paciente_identificadores WHERE paciente_id=?",
                    (int(nuevo_id),),
                )
                conn.execute("DELETE FROM pacientes WHERE id=?", (int(nuevo_id),))

            otros_pendientes = int(
                conn.execute(
                    "SELECT COUNT(*) FROM nss_conflictos "
                    "WHERE nss_normalizado=? AND estado='PENDIENTE' AND id<>?",
                    (nss, int(revision_id)),
                ).fetchone()[0]
            )
            if not otros_pendientes:
                conn.execute(
                    "UPDATE paciente_identificadores SET conflicto=0 "
                    "WHERE tipo='NSS' AND valor_normalizado=? AND activo=1",
                    (nss,),
                )

            if nuevo_id:
                conn.execute(
                    "UPDATE pacientes SET requiere_revision=0 WHERE id=?",
                    (int(nuevo_id),),
                )
            if revision["atencion_id"]:
                conn.execute(
                    "UPDATE atenciones SET identidad_estado='VALIDADA',requiere_revision=0 WHERE id=?",
                    (int(revision["atencion_id"]),),
                )
            conn.execute(
                """
                UPDATE nss_conflictos SET estado='RESUELTO',resolucion=?,motivo_resolucion=?,
                    resuelto_por=?,resuelto_at=datetime('now','localtime') WHERE id=?
                """,
                (resolucion,motivo,actor,int(revision_id)),
            )
            conn.commit()
        return True

    def actualizar_trabajo_salida(
        self,
        atencion_id,
        etapa,
        estado,
        *,
        error="",
        pdf_path=None,
        pdf_sha256=None,
        incrementar_intento=False,
    ):
        columnas = {
            "excel": "excel_estado",
            "pdf": "pdf_estado",
            "impresion": "impresion_estado",
        }
        columna = columnas.get(str(etapa).lower())
        if not columna:
            raise ValueError(f"Etapa de salida desconocida: {etapa}")
        estado = str(estado or "").upper()
        if estado not in {"PENDIENTE", "PROCESANDO", "COMPLETADO", "FALLIDO", "ENVIADO", "ENVIADO_A_IMPRESORA", "DESCONOCIDO", "NO_APLICA"}:
            raise ValueError(f"Estado de salida desconocido: {estado}")
        assignments = [f"{columna}=?", "updated_at=datetime('now','localtime')"]
        params = [estado]
        if error:
            assignments.append("ultimo_error=?")
            params.append(str(error)[:2000])
        elif estado in {"COMPLETADO", "ENVIADO", "ENVIADO_A_IMPRESORA", "NO_APLICA"}:
            assignments.append("ultimo_error=NULL")
        if pdf_path is not None:
            assignments.append("pdf_path=?")
            params.append(str(pdf_path))
        if pdf_sha256 is not None:
            assignments.append("pdf_sha256=?")
            params.append(str(pdf_sha256))
        if incrementar_intento:
            assignments.append("intentos=intentos+1")
        params.append(int(atencion_id))
        with closing(self._connect()) as conn:
            conn.execute(
                f"UPDATE trabajos_salida SET {', '.join(assignments)} WHERE atencion_id=?",
                params,
            )
            conn.commit()
        APP_LOG.info("Salida atención #%s · %s=%s", int(atencion_id), etapa, estado)

    def limpiar_error_trabajo_salida(self, atencion_id):
        with closing(self._connect()) as conn:
            conn.execute(
                """
                UPDATE trabajos_salida SET ultimo_error=NULL,updated_at=datetime('now','localtime')
                WHERE atencion_id=?
                """,
                (int(atencion_id),),
            )
            conn.commit()

    def obtener_trabajo_salida(self, atencion_id):
        with closing(self._connect()) as conn:
            conn.row_factory = sqlite3.Row
            row = conn.execute("SELECT * FROM trabajos_salida WHERE atencion_id=?", (int(atencion_id),)).fetchone()
            return dict(row) if row else None

    def listar_trabajos_salida_pendientes(self, limite=100):
        with closing(self._connect()) as conn:
            conn.row_factory = sqlite3.Row
            conn.execute(
                """
                UPDATE trabajos_salida SET
                    excel_estado=CASE WHEN excel_estado='PROCESANDO' THEN 'PENDIENTE' ELSE excel_estado END,
                    pdf_estado=CASE WHEN pdf_estado='PROCESANDO' THEN 'PENDIENTE' ELSE pdf_estado END,
                    impresion_estado=CASE WHEN impresion_estado='PROCESANDO' THEN 'PENDIENTE' ELSE impresion_estado END,
                    updated_at=datetime('now','localtime')
                WHERE excel_estado='PROCESANDO' OR pdf_estado='PROCESANDO' OR impresion_estado='PROCESANDO'
                """
            )
            conn.commit()
            rows = conn.execute(
                """
                SELECT t.*, a.nombre, a.hoja, a.fecha, a.hora
                FROM trabajos_salida t
                JOIN atenciones a ON a.id=t.atencion_id
                WHERE a.estado='ACTIVA' AND (
                    t.excel_estado IN ('PENDIENTE','FALLIDO') OR
                    t.pdf_estado IN ('PENDIENTE','FALLIDO') OR
                    (
                        t.impresion_estado IN ('PENDIENTE','FALLIDO')
                        AND t.pdf_estado<>'NO_APLICA'
                    )
                )
                ORDER BY t.updated_at, t.atencion_id
                LIMIT ?
                """,
                (max(1, int(limite)),),
            ).fetchall()
            return [dict(row) for row in rows]

    def registrar_documento(self, atencion_id, tipo, ruta, plantilla=""):
        """Compatibilidad: las hojas individuales ya no se archivan ni se registran."""
        APP_LOG.info("Se omitió el registro permanente de la hoja de atención #%s", atencion_id)
        return None

    def obtener_documento_atencion(self, atencion_id, tipo="HOJA_EMERGENCIA"):
        with closing(self._connect()) as conn:
            conn.row_factory = sqlite3.Row
            row = conn.execute(
                """
                SELECT * FROM documentos
                WHERE atencion_id=? AND tipo=?
                ORDER BY id DESC LIMIT 1
                """,
                (int(atencion_id), str(tipo).upper()),
            ).fetchone()
        if not row:
            return None
        documento = dict(row)
        ruta = documento.get("ruta") or ""
        if not os.path.isfile(ruta):
            return None
        digest = hashlib.sha256()
        with open(ruta, "rb") as stream:
            for block in iter(lambda: stream.read(1024 * 1024), b""):
                digest.update(block)
        if digest.hexdigest() != (documento.get("sha256") or ""):
            APP_LOG.error("El PDF archivado de la atención #%s no supera SHA-256", atencion_id)
            return None
        return documento

    def borrar_atencion(self, atencion_id: int, motivo="Eliminada desde el historial", usuario="") -> bool:
        motivo = str(motivo or "").strip()
        usuario = limpiar_nombre_representante(usuario)
        if len(motivo) < 5:
            raise ValueError("La anulación requiere un motivo de al menos 5 caracteres.")
        if not usuario:
            raise ValueError("La anulación requiere identificar al operador.")
        with closing(self._connect()) as conn:
            conn.row_factory = sqlite3.Row
            conn.execute("BEGIN IMMEDIATE")
            cur = conn.cursor()
            fila = cur.execute('SELECT * FROM atenciones WHERE id = ?', (atencion_id,)).fetchone()
            if not fila or str(fila["estado"] or "ACTIVA").upper() == "ANULADA":
                conn.rollback()
                return False
            reingresos_activos = int(
                cur.execute(
                    """
                    SELECT COUNT(*) FROM atenciones
                    WHERE atencion_origen_id=? AND estado='ACTIVA'
                    """,
                    (int(atencion_id),),
                ).fetchone()[0]
            )
            if reingresos_activos:
                conn.rollback()
                raise ValueError(
                    f"No se puede anular: existen {reingresos_activos} reingreso(s) activo(s) vinculados."
                )
            before = dict(fila)
            cur.execute(
                """
                UPDATE atenciones SET
                    estado='ANULADA', anulada_at=datetime('now','localtime'),
                    anulada_por=?, anulada_motivo=?, updated_at=datetime('now','localtime')
                WHERE id=? AND estado='ACTIVA'
                """,
                (usuario, motivo, int(atencion_id)),
            )
            changed = cur.rowcount > 0
            after = dict(cur.execute("SELECT * FROM atenciones WHERE id=?", (atencion_id,)).fetchone())
            self._registrar_auditoria_conn(
                conn,
                atencion_id,
                "ANULACION",
                motivo,
                usuario,
                before,
                after,
                "OPERADOR",
            )
            event_ref = enqueue_billing_event(
                conn,
                attention_id=int(atencion_id),
                actor=usuario,
                actor_role=self.session_context.role,
                session_id=self.session_context.session_id,
            )
            conn.commit()
        if changed:
            self._emit_contract_event("attention_cancelled", event_ref)
        return changed

    def reordenar_ids_atenciones_despues_de_eliminar(self, atencion_id_eliminada: int) -> bool:
        """Los ID son referencias permanentes y nunca se renumeran."""
        APP_LOG.warning("Se ignoró una solicitud de renumeración después de eliminar #%s", atencion_id_eliminada)
        return False

    def listar_atenciones(self, filtro_texto=None, limite=200, offset=0):
        limite = max(1, int(limite or 200))
        offset = max(0, int(offset or 0))
        q = (filtro_texto or "").strip()
        q_digits = re.sub(r"\D", "", q)

        with closing(self._connect()) as conn:
            conn.row_factory = sqlite3.Row
            cur = conn.cursor()

            if q:
                like = f"%{q.upper()}%"
                like_digits = f"%{q_digits or q}%"

                cur.execute('''
                    SELECT id, fecha, hora, nombre, hoja, ars, nss, cedula, edad_num, unidad, tipo_atencion
                    FROM atenciones
                    WHERE estado='ACTIVA' AND (
                        UPPER(IFNULL(nombre,'')) LIKE ?
                        OR UPPER(IFNULL(ars,'')) LIKE ?
                        OR UPPER(IFNULL(nss,'')) LIKE ?
                        OR UPPER(IFNULL(nss,'')) LIKE ?
                        OR IFNULL(cedula,'') LIKE ?
                        OR IFNULL(telefono,'') LIKE ?
                    )
                    ORDER BY id DESC
                    LIMIT ? OFFSET ?
                ''', (like, like, like, like_digits, like_digits, like_digits, limite, offset))
            else:
                cur.execute('''
                    SELECT id, fecha, hora, nombre, hoja, ars, nss, cedula, edad_num, unidad, tipo_atencion
                    FROM atenciones
                    WHERE estado='ACTIVA'
                    ORDER BY id DESC
                    LIMIT ? OFFSET ?
                ''', (limite, offset))

            return [dict(r) for r in cur.fetchall()]

    def obtener_atencion_por_id(self, atencion_id):
        with closing(self._connect()) as conn:
            conn.row_factory = sqlite3.Row
            cur = conn.cursor()
            cur.execute('SELECT * FROM atenciones WHERE id = ?', (atencion_id,))
            r = cur.fetchone()
            return dict(r) if r else None

    def obtener_turno_config_atencion(self, atencion_id):
        with closing(self._connect()) as conn:
            conn.row_factory = sqlite3.Row
            row = conn.execute(
                """
                SELECT d.fecha_base,t.tipo_turno,t.representante,t.fecha_inicio
                FROM atenciones a
                JOIN dias_operativos d ON d.id=a.dia_operativo_id
                JOIN turnos t ON t.id=a.turno_id
                WHERE a.id=?
                """,
                (int(atencion_id),),
            ).fetchone()
        if not row:
            return None
        fecha_base = date.fromisoformat(row["fecha_base"])
        inicio_real = parse_datetime_local(row["fecha_inicio"])
        return {
            "representante": limpiar_nombre_representante(row["representante"] or ""),
            "turno_codigo": normalizar_turno_codigo(row["tipo_turno"] or "8AM_8AM"),
            "fecha_base": fecha_base,
            "inicio_real": format_datetime_local(inicio_real) if inicio_real else "",
            "inicio_real_dt": inicio_real,
        }

    def listar_atenciones_sin_seguro(self, filtro_texto=None, limite=200, offset=0):
        """
        FASE 2: Historial sin seguro dedicado, sin pasar por listar_atenciones_filtradas.
        SQLite filtra directamente. Sin dedupe (id DESC es único).
        """
        limite = max(1, int(limite or 200))
        offset = max(0, int(offset or 0))

        q = (filtro_texto or "").strip()
        q_digits = re.sub(r"\D", "", q)

        sin_seguro_aliases = [
            "SIN SEGURO", "NO TIENE", "NO", "N/S", "NS", "N\\S",
            "INACTIVO", "INACTIVA", "PENDIENTE", "PENDIENTES",
            "NO ACTIVO", "NO VIGENTE", "VENCIDO", "CANCELADO",
            "DESAFILIADO", "PARTICULAR", "PRIVADO", "NINGUNO", "NINGUNA",
            "N/A", "NA", "S/N", "SN", "NO APLICA", "NO USA", "NO POSEE",
            "NINGUN SEGURO"
        ]

        where = [
            "estado='ACTIVA'",
            """(
                TRIM(IFNULL(ars,'')) = ''
                OR UPPER(TRIM(IFNULL(ars,''))) IN ({})
                OR TRIM(IFNULL(nss,'')) = ''
                OR UPPER(TRIM(IFNULL(nss,''))) IN ('SIN SEGURO','NO','N/S','NS','N\\S')
                OR LENGTH(TRIM(IFNULL(ars,''))) = 1
                OR (
                    TRIM(IFNULL(ars,'')) <> ''
                    AND TRIM(IFNULL(ars,'')) NOT GLOB '*[^0-9]*'
                )
            )""".format(",".join(["?"] * len(sin_seguro_aliases)))
        ]
        params = [a.upper() for a in sin_seguro_aliases]

        if q:
            like = f"%{q.upper()}%"
            like_digits = f"%{q_digits or q}%"
            where.append("""(
                UPPER(IFNULL(nombre,'')) LIKE ?
        OR UPPER(IFNULL(ars,'')) LIKE ?
        OR UPPER(IFNULL(nss,'')) LIKE ?
        OR IFNULL(nss_clean,'') LIKE ?
        OR IFNULL(cedula_clean,'') LIKE ?
        OR IFNULL(telefono_clean,'') LIKE ?            
            )""")
            params.extend([like, like, like, like_digits, like_digits, like_digits])

        sql = f"""
            SELECT id, fecha, hora, nombre, hoja, ars, nss, cedula, edad_num, unidad, tipo_atencion
            FROM atenciones
            WHERE {' AND '.join(where)}
            ORDER BY id DESC
            LIMIT ? OFFSET ?
        """
        params.extend([limite, offset])

        with closing(self._connect()) as conn:
            conn.row_factory = sqlite3.Row
            cur = conn.cursor()
            cur.execute(sql, params)
            filas = [dict(r) for r in cur.fetchall()]

        for f in filas:
            f["ars"] = "SIN SEGURO"
        return filas

    def listar_ars_distintas(self):
        with closing(self._connect()) as conn:
            conn.row_factory = sqlite3.Row
            cur = conn.cursor()
            cur.execute("""
                SELECT DISTINCT ars
                FROM atenciones
                WHERE estado='ACTIVA' AND TRIM(IFNULL(ars,'')) <> ''
                ORDER BY ars ASC
            """)
            valores_unicos = [r["ars"] for r in cur.fetchall()]
            valores_norm = set()
            for ars_raw in valores_unicos:
                canon = normalizar_seguro(ars_raw, "9999")
                if canon and canon != "SIN SEGURO":
                    valores_norm.add(canon)
            return sorted(valores_norm)

    def buscar_paciente_para_edicion(self, identidad: str):
        ident = (identidad or "").strip().replace("-", "").upper()
        if not ident:
            return None

        with closing(self._connect()) as conn:
            conn.row_factory = sqlite3.Row
            patient_id = None
            attention = None
            if ident.startswith("A:") and ident[2:].isdigit():
                attention = conn.execute(
                    "SELECT * FROM atenciones WHERE id=? LIMIT 1", (int(ident[2:]),)
                ).fetchone()
                if attention:
                    patient_id = int(attention["paciente_id"])
            elif ident.startswith("P:") and ident[2:].isdigit():
                patient_id = int(ident[2:])
            else:
                ident_normalizado = re.sub(r"\D", "", ident)
                rows = conn.execute(
                    """
                    SELECT paciente_id FROM paciente_identificadores
                    WHERE activo=1 AND valor_normalizado=?
                    ORDER BY conflicto,id DESC LIMIT 1
                    """,
                    (ident_normalizado,),
                ).fetchall()
                if rows:
                    patient_id = int(rows[0][0])
            if patient_id is None:
                return None

            patient = conn.execute("SELECT * FROM pacientes WHERE id=?", (patient_id,)).fetchone()
            if not patient:
                return None
            if attention is None:
                attention = conn.execute(
                    "SELECT * FROM atenciones WHERE paciente_id=? ORDER BY id DESC LIMIT 1",
                    (patient_id,),
                ).fetchone()
            data = dict(patient)
            data["paciente_id"] = patient_id
            if attention:
                data.update(dict(attention))
                data["paciente_id"] = patient_id
            return data

    def _resolver_paciente_para_eliminacion(self, conn, paciente_id: int):
        try:
            paciente_id = int(paciente_id)
        except (TypeError, ValueError):
            return None
        conn.row_factory = sqlite3.Row
        paciente = conn.execute("SELECT * FROM pacientes WHERE id=?", (paciente_id,)).fetchone()
        if not paciente:
            return None
        atenciones = [
            dict(row) for row in conn.execute(
                "SELECT * FROM atenciones WHERE paciente_id=? ORDER BY id", (paciente_id,)
            ).fetchall()
        ]
        ids = [int(row["id"]) for row in atenciones]
        auditorias = 0
        documentos = []
        if ids:
            marks = ",".join("?" for _ in ids)
            auditorias = int(conn.execute(
                f"SELECT COUNT(*) FROM atenciones_auditoria WHERE atencion_id IN ({marks})", ids
            ).fetchone()[0])
            documentos = [row[0] for row in conn.execute(
                f"SELECT ruta FROM documentos WHERE atencion_id IN ({marks})", ids
            ).fetchall()]
        return {
            "seguro": True,
            "paciente_id": paciente_id,
            "paciente": dict(paciente),
            "atenciones": atenciones,
            "fichas": 1,
            "auditorias": auditorias,
            "documentos": documentos,
        }

    def previsualizar_eliminacion_paciente(self, paciente_id: int):
        with closing(self._connect()) as conn:
            return self._resolver_paciente_para_eliminacion(conn, paciente_id)

    def eliminar_paciente_completo(
        self,
        paciente_id: int,
        motivo: str,
        actor: str,
        *,
        confirmado_prueba: bool = False,
    ):
        motivo = (motivo or "").strip()
        actor = limpiar_nombre_representante(actor)
        if not confirmado_prueba:
            raise PermissionError("La purga física solo está permitida para datos confirmados como prueba.")
        if len(motivo) < 8:
            raise ValueError("Debe indicar un motivo de eliminación de al menos 8 caracteres.")
        if not actor:
            raise PermissionError("La purga requiere un actor administrativo identificado.")

        preview = self.previsualizar_eliminacion_paciente(paciente_id)
        if not preview:
            return None
        backup_folder = self.backup_manager.create(
            "antes_purga_paciente",
            label=f"paciente_id={int(paciente_id)}; actor={actor}",
        )
        cuarentena = os.path.join(
            BACKUPS_DIR,
            "CUARENTENA_PURGA",
            f"paciente_{int(paciente_id)}_{datetime.now().strftime('%Y%m%d_%H%M%S_%f')}",
        )
        os.makedirs(cuarentena, exist_ok=False)
        archivos_movidos = []
        try:
            for indice, ruta in enumerate(preview.get("documentos", []), start=1):
                if not ruta or not os.path.isfile(ruta):
                    continue
                destino = os.path.join(cuarentena, f"{indice:04d}_{os.path.basename(ruta)}")
                os.replace(os.path.abspath(ruta), destino)
                archivos_movidos.append((os.path.abspath(ruta), destino))
        except Exception:
            for original, temporal in reversed(archivos_movidos):
                try:
                    if os.path.isfile(temporal):
                        os.makedirs(os.path.dirname(original), exist_ok=True)
                        os.replace(temporal, original)
                except OSError:
                    APP_LOG.critical("No se pudo revertir la cuarentena de %s", original)
            shutil.rmtree(cuarentena, ignore_errors=True)
            raise RuntimeError(
                "No se pudieron aislar todos los documentos; la base no fue modificada."
            )

        try:
            with closing(self._connect()) as conn:
                conn.row_factory = sqlite3.Row
                conn.execute("BEGIN IMMEDIATE")
                try:
                    resumen = self._resolver_paciente_para_eliminacion(conn, int(paciente_id))
                    if not resumen or not resumen.get("seguro"):
                        raise RuntimeError("La ficha dejó de estar disponible antes de la purga.")

                    ids = [int(row["id"]) for row in resumen["atenciones"]]
                    auditoria_ids = []
                    if ids:
                        marks = ",".join("?" for _ in ids)
                        auditoria_ids = [
                            int(row[0])
                            for row in conn.execute(
                                f"SELECT id FROM atenciones_auditoria WHERE atencion_id IN ({marks})",
                                ids,
                            ).fetchall()
                        ]
                    conn.execute("DELETE FROM atenciones WHERE paciente_id=?", (int(paciente_id),))
                    atenciones_eliminadas = int(conn.execute("SELECT changes()").fetchone()[0])
                    conn.execute("DELETE FROM pacientes WHERE id=?", (int(paciente_id),))
                    fichas_eliminadas = int(conn.execute("SELECT changes()").fetchone()[0])

                    previous_row = conn.execute(
                        "SELECT event_hash FROM purga_eventos ORDER BY id DESC LIMIT 1"
                    ).fetchone()
                    previous_hash = str(previous_row[0] or "") if previous_row else ""
                    paciente_hash = hashlib.sha256(
                        os.urandom(32) + str(int(paciente_id)).encode("ascii")
                    ).hexdigest()
                    event_payload = {
                        "paciente_hash": paciente_hash,
                        "motivo": motivo,
                        "actor": actor,
                        "actor_rol": "ADMINISTRADOR",
                        "backup_path": str(backup_folder),
                        "atenciones_eliminadas": atenciones_eliminadas,
                        "fichas_eliminadas": fichas_eliminadas,
                        "previous_hash": previous_hash,
                        "workstation": platform.node(),
                    }
                    event_hash = hashlib.sha256(
                        json.dumps(event_payload, sort_keys=True, separators=(",", ":")).encode("utf-8")
                    ).hexdigest()
                    conn.execute(
                        """
                        INSERT INTO purga_eventos(
                            paciente_hash,motivo,actor,actor_rol,backup_path,
                            atenciones_eliminadas,fichas_eliminadas,previous_hash,event_hash,workstation
                        ) VALUES (?,?,?,?,?,?,?,?,?,?)
                        """,
                        (
                            paciente_hash, motivo, actor, "ADMINISTRADOR", str(backup_folder),
                            atenciones_eliminadas, fichas_eliminadas,
                            previous_hash, event_hash, platform.node(),
                        ),
                    )
                    auditorias_redactadas = 0
                    if auditoria_ids:
                        audit_marks = ",".join("?" for _ in auditoria_ids)
                        conn.execute(
                            f"""
                            UPDATE atenciones_auditoria SET
                                accion='REDACTADO_POR_PURGA',motivo=?,usuario=?,
                                snapshot_json='{{}}',snapshot_after_json='{{}}'
                            WHERE id IN ({audit_marks})
                            """,
                            [f"Purga administrativa: {event_hash}", actor, *auditoria_ids],
                        )
                        auditorias_redactadas = int(
                            conn.execute("SELECT changes()").fetchone()[0]
                        )
                    conn.commit()
                except Exception:
                    conn.rollback()
                    raise
        except Exception:
            for original, temporal in reversed(archivos_movidos):
                try:
                    if os.path.isfile(temporal):
                        os.makedirs(os.path.dirname(original), exist_ok=True)
                        os.replace(temporal, original)
                except OSError:
                    APP_LOG.critical("No se pudo restaurar el documento %s", original)
            shutil.rmtree(cuarentena, ignore_errors=True)
            APP_LOG.exception("Falló la eliminación total de un paciente")
            raise

        archivos_pendientes = []
        for _original, temporal in archivos_movidos:
            try:
                eliminar_archivo_sensible(temporal)
            except OSError:
                archivos_pendientes.append(temporal)
                APP_LOG.critical("Documento de purga pendiente de eliminar: %s", temporal)
        try:
            if not os.listdir(cuarentena):
                os.rmdir(cuarentena)
        except OSError:
            pass

        resumen.pop("paciente", None)
        resumen.pop("atenciones", None)
        resumen.pop("documentos", None)
        resumen.update(
            {
                "atenciones_eliminadas": atenciones_eliminadas,
                "fichas_eliminadas": fichas_eliminadas,
                "auditorias_redactadas": auditorias_redactadas,
                "documentos_eliminados": len(archivos_movidos) - len(archivos_pendientes),
                "documentos_pendientes": archivos_pendientes,
                "backup_path": str(backup_folder),
                "purga_event_hash": event_hash,
            }
        )
        APP_LOG.warning(
            "Purga administrativa completada: event_hash=%s, atenciones=%s, fichas=%s, pendientes=%s",
            event_hash,
            atenciones_eliminadas,
            fichas_eliminadas,
            len(archivos_pendientes),
        )
        return resumen

    def actualizar_datos_paciente_por_identidad(self, identidad_original: str, nuevos: dict, actualizar_ficha=True):
        ident = (identidad_original or "").strip().upper()
        if not ident:
            return 0, 0

        nombre = (nuevos.get("Nombre") or "").strip()
        cedula = (nuevos.get("Cédula") or "").strip().replace("-", "")
        telefono = (nuevos.get("Teléfono") or "").strip().replace("-", "")
        telefono_db = telefono if telefono.isdigit() and len(telefono) == 10 else None
        direccion = (nuevos.get("Dirección") or "").strip()
        nacionalidad = (nuevos.get("Nacionalidad") or "").strip()
        nss = (nuevos.get("NSS") or "").strip().upper()
        ars = normalizar_seguro(nuevos.get("Aseguradora (ARS)", ""), nss)

        with closing(self._connect()) as conn:
            conn.row_factory = sqlite3.Row
            conn.execute("BEGIN IMMEDIATE")
            cur = conn.cursor()
            objetivo = None
            patient_id = None
            if ident.startswith("P:") and ident[2:].isdigit():
                patient_id = int(ident[2:])
            elif ident.startswith("A:") and ident[2:].isdigit():
                objetivo = cur.execute(
                    "SELECT * FROM atenciones WHERE id=? LIMIT 1", (int(ident[2:]),)
                ).fetchone()
            else:
                clean_ident = re.sub(r"\D", "", ident)
                patient_rows = cur.execute(
                    "SELECT paciente_id FROM paciente_identificadores "
                    "WHERE valor_normalizado=? AND activo=1 "
                    "ORDER BY conflicto,id DESC LIMIT 1",
                    (clean_ident,),
                ).fetchall()
                patient_id = int(patient_rows[0][0]) if patient_rows else None
            if objetivo:
                patient_id = int(objetivo["paciente_id"])
            elif patient_id is not None:
                objetivo = cur.execute('''
                    SELECT * FROM atenciones
                    WHERE paciente_id=?
                    ORDER BY id DESC LIMIT 1
                ''', (patient_id,)).fetchone()
            if patient_id is None:
                conn.rollback()
                return 0, 0

            nss_clean = re.sub(r"\D", "", nss)
            cedula_clean = re.sub(r"\D", "", cedula)
            nuevos_ids = []
            if is_valid_nss_key(nss_clean):
                nuevos_ids.append(("NSS", nss_clean))
            if is_valid_cedula_key(cedula_clean):
                nuevos_ids.append(("CEDULA", cedula_clean))
            atenciones_actualizadas = 0
            if objetivo:
                objetivo_id = int(objetivo["id"])
                cur.execute('''
                    UPDATE atenciones
                    SET nombre=?, cedula=?, telefono=?, direccion=?, nacionalidad=?, nss=?, ars=?,
                        nss_clean=?, cedula_clean=?, telefono_clean=?, identidad_estado=?,
                        requiere_revision=?, updated_at=datetime('now','localtime')
                    WHERE id=?
                ''', (
                    nombre, cedula, telefono, direccion, nacionalidad, nss, ars,
                    nss_clean or None, cedula_clean or None,
                    re.sub(r"\D", "", telefono) or None, "VALIDADA", 0, objetivo_id,
                ))
                atenciones_actualizadas = cur.rowcount
                after = dict(cur.execute("SELECT * FROM atenciones WHERE id=?", (objetivo_id,)).fetchone())
                self._registrar_auditoria_conn(
                    conn,
                    objetivo_id,
                    "MODIFICACION",
                    "Edición de una atención y/o ficha del paciente",
                    limpiar_nombre_representante((cargar_turno_config(permitir_vencido=True) or {}).get("representante", "")),
                    dict(objetivo),
                    after,
                    "OPERADOR",
                )

            pacientes_actualizados = 0
            if actualizar_ficha:
                for kind, value in nuevos_ids:
                    if kind != "CEDULA":
                        continue
                    conflict = cur.execute(
                        "SELECT paciente_id FROM paciente_identificadores "
                        "WHERE tipo=? AND valor_normalizado=? AND activo=1 AND paciente_id<>? LIMIT 1",
                        (kind, value, patient_id),
                    ).fetchone()
                    if conflict:
                        conn.rollback()
                        raise ValueError(
                            f"El {kind} indicado ya está asignado a otra ficha."
                        )
                cur.execute('''
                    UPDATE pacientes SET
                        cedula=?, nombre=?, telefono=?, direccion=?, nacionalidad=?, ars=?, nss=?,
                        nss_clean=?, cedula_clean=?, telefono_clean=?, provisional=?,
                        updated_at=datetime('now','localtime')
                    WHERE id=?
                ''', (
                    cedula_clean if is_valid_cedula_key(cedula_clean) else None,
                    nombre, telefono_db, direccion, nacionalidad, ars,
                    nss_clean if is_valid_nss_key(nss_clean) else None,
                    nss_clean if is_valid_nss_key(nss_clean) else None,
                    cedula_clean if is_valid_cedula_key(cedula_clean) else None,
                    re.sub(r"\D", "", telefono) or None,
                    int(not nuevos_ids),
                    patient_id,
                ))
                pacientes_actualizados = cur.rowcount
                for kind in ("NSS", "CEDULA"):
                    new_value = next((value for item_kind, value in nuevos_ids if item_kind == kind), None)
                    cur.execute(
                        "DELETE FROM paciente_identificadores WHERE paciente_id=? AND tipo=?",
                        (patient_id, kind),
                    )
                    if new_value:
                        cur.execute(
                            "INSERT INTO paciente_identificadores(" 
                            "paciente_id,tipo,valor_normalizado,activo,conflicto) VALUES (?,?,?,1,0) "
                            "ON CONFLICT(paciente_id,tipo,valor_normalizado) "
                            "DO UPDATE SET activo=1,conflicto=0",
                            (patient_id, kind, new_value),
                        )

            conn.commit()
            return atenciones_actualizadas, pacientes_actualizados

    def listar_ars_conteo(self):
        conteo = {}
        with closing(self._connect()) as conn:
            conn.row_factory = sqlite3.Row
            cur = conn.cursor()
            cur.execute(
                "SELECT ars, nss, COUNT(*) AS cantidad "
                "FROM atenciones "
                "WHERE estado='ACTIVA' AND TRIM(IFNULL(ars,'')) <> '' "
                "GROUP BY ars, nss"
            )
            for r in cur.fetchall():
                canon = normalizar_seguro(r["ars"], r["nss"])
                if canon:
                    conteo[canon] = conteo.get(canon, 0) + int(r["cantidad"] or 0)
        return sorted(conteo.items(), key=lambda x: (-x[1], x[0]))

    def reemplazar_ars_global(self, ars_actual: str, ars_nueva: str):
        actual = normalizar_seguro(ars_actual, "9999")
        nueva = normalizar_seguro(ars_nueva, "9999")
        if not actual or not nueva:
            return 0

        total = 0
        with closing(self._connect()) as conn:
            conn.row_factory = sqlite3.Row
            cur = conn.cursor()
            cur.execute("SELECT DISTINCT ars, nss FROM pacientes WHERE TRIM(IFNULL(ars,'')) <> ''")
            raw_to_update = [r["ars"] for r in cur.fetchall() if normalizar_seguro(r["ars"], r["nss"]) == actual]
            for raw in set(raw_to_update):
                cur.execute("UPDATE pacientes SET ars = ? WHERE ars = ?", (nueva, raw))
                total += cur.rowcount

            conn.commit()
        return total

    def normalizar_todas_ars(self):
        total = 0
        with closing(self._connect()) as conn:
            conn.row_factory = sqlite3.Row
            cur = conn.cursor()
            cur.execute("SELECT DISTINCT ars, nss FROM pacientes WHERE TRIM(IFNULL(ars,'')) <> ''")
            for r in cur.fetchall():
                canon = normalizar_seguro(r["ars"], r["nss"])
                if canon and canon != r["ars"]:
                    cur.execute("UPDATE pacientes SET ars = ?,updated_at=datetime('now','localtime') WHERE ars = ?", (canon, r["ars"]))
                    total += cur.rowcount
            conn.commit()
        return total

    def limpiar_ars_cortas_invalidas(self):
        total = 0
        with closing(self._connect()) as conn:
            conn.row_factory = sqlite3.Row
            cur = conn.cursor()
            cur.execute("SELECT DISTINCT ars FROM pacientes WHERE TRIM(IFNULL(ars,'')) <> ''")
            for (raw,) in cur.fetchall():
                if ars_es_corta_invalida(raw):
                    cur.execute("UPDATE pacientes SET ars='SIN SEGURO',updated_at=datetime('now','localtime') WHERE ars=?", (raw,))
                    total += cur.rowcount
            conn.commit()
        return total

    def obtener_turnos_historial(self):
        """Devuelve el turno abierto y su predecesor desde los registros de turnos."""
        with closing(self._connect()) as conn:
            conn.row_factory = sqlite3.Row
            actual = conn.execute(
                """
                SELECT id,fecha_inicio,fecha_fin,tipo_turno,representante,estado
                FROM turnos WHERE estado='ABIERTO'
                ORDER BY datetime(fecha_inicio) DESC,id DESC LIMIT 1
                """
            ).fetchone()
            if not actual:
                return {"actual": None, "anterior": None}
            actual = dict(actual)
            anterior = conn.execute(
                """
                SELECT id,fecha_inicio,fecha_fin,tipo_turno,representante,estado
                FROM turnos
                WHERE id<>? AND datetime(fecha_inicio)<=datetime(?)
                ORDER BY datetime(fecha_inicio) DESC,id DESC LIMIT 1
                """,
                (int(actual["id"]), actual["fecha_inicio"]),
            ).fetchone()
        return {"actual": actual, "anterior": dict(anterior) if anterior else None}

    def listar_atenciones_filtradas(self, filtro_texto=None, modo="Todos", ars=None, especialidad=None, fecha_txt=None, limite=200, offset=0, turno_id=None):
        """
        FASE 5: Sin dedupe (id DESC ya es único).
        """
        limite = max(1, int(limite or 200))
        offset = max(0, int(offset or 0))
        modo = (modo or "Todos").strip()
        ars = (ars or "").strip()
        especialidad = (especialidad or "").strip().upper()
        fecha_obj = parse_fecha_ddmmyyyy(fecha_txt) if fecha_txt else None

        where = ["estado='ACTIVA'"]
        params = []

        q = (filtro_texto or "").strip()
        if q:
            q_digits = re.sub(r"\D", "", q)
            like = f"%{q.upper()}%"
            like_digits = f"%{q_digits or q}%"
            where.append('''(
                UPPER(IFNULL(nombre,'')) LIKE ?
                OR UPPER(IFNULL(ars,'')) LIKE ?
                OR UPPER(IFNULL(nss,'')) LIKE ?
                OR UPPER(IFNULL(nss,'')) LIKE ?
                OR IFNULL(cedula,'') LIKE ?
                OR IFNULL(telefono,'') LIKE ?
            )''')
            params.extend([like, like, like, like_digits, like_digits, like_digits])

        if modo == "Hoy":
            where.append(
                "dia_operativo_id=(SELECT id FROM dias_operativos WHERE fecha_base=? LIMIT 1)"
            )
            params.append(fecha_base_operativa_actual().isoformat())

        if modo == "Por fecha" and fecha_obj:
            where.append(
                "dia_operativo_id=(SELECT id FROM dias_operativos WHERE fecha_base=? LIMIT 1)"
            )
            params.append(fecha_obj.isoformat())

        if modo == "Por especialidad" and especialidad and especialidad != "(TODAS)":
            where.append("UPPER(IFNULL(hoja,'')) = ?")
            params.append(especialidad)

        if modo == "Por ARS" and ars and ars != "(Todas)":
            canon_ars = normalizar_seguro(ars, "9999")
            alias_values = [canon_ars]
            try:
                alias_values.extend(cargar_catalogo_ars().get(canon_ars, []))
                alias_values.append(seguro_para_mostrar(canon_ars))
            except Exception:
                pass
            alias_values = sorted({str(a).strip().upper() for a in alias_values if str(a).strip()})
            if alias_values:
                where.append("(" + " OR ".join(["UPPER(IFNULL(ars,'')) LIKE ?"] * len(alias_values)) + ")")
                params.extend([f"%{a}%" for a in alias_values])

        if turno_id is None and modo == "Turno actual":
            turno_id = (self.obtener_turnos_historial().get("actual") or {}).get("id")
            if turno_id is None:
                return []
        if turno_id is not None:
            where.append("turno_id=?")
            params.append(int(turno_id))

        sql = '''
            SELECT id, fecha, hora, nombre, hoja, ars, nss, cedula, edad_num, unidad, tipo_atencion
            FROM atenciones
        '''
        if where:
            sql += " WHERE " + " AND ".join(where)
        sql += " ORDER BY id DESC LIMIT ? OFFSET ?"

        necesita_python = modo == "Sin seguro"
        fetch_limit = limite * 3 if necesita_python else limite
        params_sql = params + [fetch_limit, offset]

        with closing(self._connect()) as conn:
            conn.row_factory = sqlite3.Row
            cur = conn.cursor()
            cur.execute(sql, params_sql)
            filas = [dict(r) for r in cur.fetchall()]

        if not necesita_python:
            return filas
        
        result = []
        for f in filas:
            canon = normalizar_seguro(f.get("ars", ""), f.get("nss", ""))
            if modo == "Sin seguro" and canon != "SIN SEGURO":
                continue
            result.append(f)
            if len(result) >= limite:
                break

        return result

    def buscar_pacientes_avanzado(self, texto: str, limite=80):
        q = (texto or "").strip().replace("-", "")
        if not q: return []
        like = f"%{q}%"
        with closing(self._connect()) as conn:
            conn.row_factory = sqlite3.Row
            cur = conn.cursor()
            cur.execute("""
                SELECT id, paciente_id, fecha, hora, nombre, hoja, ars, nss, cedula, telefono, direccion, nacionalidad
                FROM atenciones
                WHERE estado='ACTIVA' AND (nombre LIKE ? OR nss LIKE ? OR cedula LIKE ? OR telefono LIKE ?)
                ORDER BY id DESC
                LIMIT ?
            """, (like, like, like, like, limite))
            rows = [dict(r) for r in cur.fetchall()]
        latest_by_patient = {}
        for row in rows:
            patient_id = int(row.get("paciente_id") or 0)
            if patient_id not in latest_by_patient:
                latest_by_patient[patient_id] = row
        return list(latest_by_patient.values())

    def actualizar_atencion_especifica(
        self,
        atencion_id: int,
        nuevos: dict,
        actualizar_ficha=False,
        usuario="",
        motivo="Corrección de atención específica",
    ):
        correction_reason = str(motivo or "").strip()
        if len(correction_reason) < 5:
            raise ValueError("Debe indicar un motivo de rectificación de al menos 5 caracteres.")
        nombre = (nuevos.get("Nombre") or "").strip()
        fecha = (nuevos.get("Fecha") or "").strip()
        hora = (nuevos.get("Hora") or "").strip()
        hoja = (nuevos.get("Hoja") or "").strip().upper()
        ars = normalizar_seguro(nuevos.get("Aseguradora (ARS)", ""), nuevos.get("NSS", ""))
        nss = (nuevos.get("NSS") or "").strip().upper()
        cedula = (nuevos.get("Cédula") or "").strip().replace("-", "")
        telefono = (nuevos.get("Teléfono") or "").strip().replace("-", "")
        telefono_db = telefono if telefono.isdigit() and len(telefono) == 10 else None
        direccion = (nuevos.get("Dirección") or "").strip()
        nacionalidad = (nuevos.get("Nacionalidad") or "").strip()
        sexo = (nuevos.get("Sexo") or "").strip()
        if sexo not in ("Masculino", "Femenino"):
            raise ValueError("El sexo debe ser Masculino o Femenino.")
        edad_num = int(nuevos.get("Edad_num", 0) or 0)
        unidad = (nuevos.get("Unidad") or "Años").strip()
        tipo_atencion = (nuevos.get("TipoAtencion") or "EMERGENCIA").strip().upper()
        if tipo_atencion not in ("EMERGENCIA", "URGENCIA", "CONSULTA"):
            tipo_atencion = "EMERGENCIA"
        actor = limpiar_nombre_representante(
            usuario
            or (cargar_turno_config(permitir_vencido=True) or {}).get(
                "representante", ""
            )
        )

        with closing(self._connect()) as conn:
            conn.row_factory = sqlite3.Row
            conn.execute("BEGIN IMMEDIATE")
            cur = conn.cursor()
            snapshot = cur.execute("SELECT * FROM atenciones WHERE id=?", (int(atencion_id),)).fetchone()
            if not snapshot:
                conn.rollback()
                return 0
            if str(snapshot["estado"] or "ACTIVA").upper() != "ACTIVA":
                conn.rollback()
                raise ValueError("No se puede editar una atención anulada.")

            correction_after = {
                "nombre": nombre,
                "sexo": sexo,
                "fecha": fecha,
                "hora": hora,
                "hoja": hoja,
                "ars": ars,
                "nss": nss,
                "cedula": cedula,
                "telefono": telefono,
                "direccion": direccion,
                "nacionalidad": nacionalidad,
                "edad_num": edad_num,
                "unidad": unidad,
                "tipo_atencion": tipo_atencion,
            }
            changed_fields = sorted(
                key
                for key, value in correction_after.items()
                if str(snapshot[key] or "") != str(value or "")
            )

            nss_clean = re.sub(r"\D", "", nss)
            cedula_clean = re.sub(r"\D", "", cedula)
            telefono_clean = re.sub(r"\D", "", telefono)
            paciente_id = int(snapshot["paciente_id"])
            identificadores = []
            if is_valid_nss_key(nss_clean):
                identificadores.append(("NSS", nss_clean))
            if is_valid_cedula_key(cedula_clean):
                identificadores.append(("CEDULA", cedula_clean))
            revision_nss = None
            if not is_valid_cedula_key(cedula_clean) and is_valid_nss_key(nss_clean):
                revision_nss = cur.execute(
                    """
                    SELECT p.id,p.nombre
                    FROM paciente_identificadores i
                    JOIN pacientes p ON p.id=i.paciente_id
                    WHERE i.tipo='NSS' AND i.valor_normalizado=?
                      AND i.activo=1 AND p.id<>?
                    ORDER BY i.conflicto,p.id DESC LIMIT 1
                    """,
                    (nss_clean, paciente_id),
                ).fetchone()
            nss_clean_atencion = None if revision_nss else (nss_clean or None)
            identidad_estado = "NSS_EN_REVISION" if revision_nss else "VALIDADA"
            cur.execute("""
                UPDATE atenciones
                SET nombre=?, sexo=?, fecha=?, hora=?, hoja=?, ars=?, nss=?, cedula=?, telefono=?, direccion=?, nacionalidad=?,
                    edad_num=?, unidad=?, tipo_atencion=?, nss_clean=?, cedula_clean=?, telefono_clean=?,
                    identidad_estado=?, requiere_revision=?,
                    correction_reason=?,correction_actor=?,correction_at=datetime('now'),
                    correction_before_json=?,correction_after_json=?,
                    correction_changed_fields_json=?,
                    updated_at=datetime('now','localtime')
                WHERE id=?
            """, (nombre, sexo, fecha, hora, hoja, ars, nss, cedula, telefono, direccion, nacionalidad,
                  edad_num, unidad, tipo_atencion, nss_clean_atencion,
                  cedula_clean or None, telefono_clean or None, identidad_estado,
                  int(bool(revision_nss)), correction_reason, actor,
                  json.dumps(dict(snapshot), ensure_ascii=False, sort_keys=True),
                  json.dumps(correction_after, ensure_ascii=False, sort_keys=True),
                  json.dumps(changed_fields, ensure_ascii=False), int(atencion_id)))
            rowcount = cur.rowcount

            # Una edición tampoco debe detener el flujo. Si queda un NSS sin
            # cédula compartido con otra ficha, se registra para revisión superior.
            cur.execute(
                """
                UPDATE nss_conflictos SET
                    estado='RESUELTO',resolucion='CORREGIDO_EN_EDICION',
                    motivo_resolucion='La atención fue editada posteriormente',
                    resuelto_por='SISTEMA',resuelto_at=datetime('now','localtime')
                WHERE atencion_id=? AND estado='PENDIENTE'
                """,
                (int(atencion_id),),
            )
            if revision_nss:
                detalle = (
                    f"NSS compartido tras edición con la ficha "
                    f"#{int(revision_nss['id'])} ({revision_nss['nombre'] or 'SIN NOMBRE'})."
                )
                cur.execute(
                    """
                    INSERT INTO nss_conflictos(
                        nss_normalizado,paciente_nuevo_id,paciente_referencia_id,
                        atencion_id,detalle
                    ) VALUES (?,?,?,?,?)
                    """,
                    (
                        nss_clean,
                        paciente_id,
                        int(revision_nss["id"]),
                        int(atencion_id),
                        detalle,
                    ),
                )
            if actualizar_ficha:
                for tipo, valor in identificadores:
                    if tipo != "CEDULA":
                        continue
                    ajenos = cur.execute(
                        """
                        SELECT DISTINCT paciente_id FROM paciente_identificadores
                        WHERE tipo=? AND valor_normalizado=? AND activo=1 AND paciente_id<>?
                        """,
                        (tipo, valor, paciente_id),
                    ).fetchall()
                    if ajenos:
                        conn.rollback()
                        raise ValueError(
                            f"El {tipo} ya está asignado a otra ficha."
                        )

                cur.execute(
                    """
                    UPDATE pacientes SET
                        nombre=?,cedula=?,telefono=?,direccion=?,nacionalidad=?,ars=?,nss=?,
                        nss_clean=?,cedula_clean=?,telefono_clean=?,provisional=?,
                        requiere_revision=0,updated_at=datetime('now','localtime')
                    WHERE id=?
                    """,
                    (
                        nombre,
                        cedula_clean if is_valid_cedula_key(cedula_clean) else None,
                        telefono_db,
                        direccion,
                        nacionalidad,
                        ars,
                        nss_clean if is_valid_nss_key(nss_clean) else None,
                        nss_clean if is_valid_nss_key(nss_clean) else None,
                        cedula_clean if is_valid_cedula_key(cedula_clean) else None,
                        telefono_clean if len(telefono_clean) == 10 else None,
                        int(not identificadores),
                        paciente_id,
                    ),
                )
                for tipo in ("NSS", "CEDULA"):
                    cur.execute(
                        "DELETE FROM paciente_identificadores WHERE paciente_id=? AND tipo=?",
                        (paciente_id, tipo),
                    )
                for tipo, valor in identificadores:
                    cur.execute(
                        """
                        INSERT INTO paciente_identificadores(
                            paciente_id,tipo,valor_normalizado,activo,conflicto
                        ) VALUES (?,?,?,1,0)
                        ON CONFLICT(paciente_id,tipo,valor_normalizado)
                        DO UPDATE SET activo=1,conflicto=0
                        """,
                        (paciente_id, tipo, valor),
                    )

            after = dict(cur.execute("SELECT * FROM atenciones WHERE id=?", (int(atencion_id),)).fetchone())
            self._registrar_auditoria_conn(
                conn,
                int(atencion_id),
                "MODIFICACION",
                correction_reason,
                actor,
                dict(snapshot),
                after,
                "OPERADOR",
            )
            event_ref = enqueue_billing_event(
                conn,
                attention_id=int(atencion_id),
                actor=actor,
                actor_role=self.session_context.role,
                session_id=self.session_context.session_id,
            )

            conn.commit()
            self._emit_contract_event("attention_updated", event_ref)
            return rowcount

    def eliminar_ars_global(self, ars_actual: str):
        actual = normalizar_seguro(ars_actual, "9999")
        if not actual: return 0
        total = 0
        with closing(self._connect()) as conn:
            conn.row_factory = sqlite3.Row
            cur = conn.cursor()
            cur.execute("SELECT DISTINCT ars, nss FROM pacientes WHERE TRIM(IFNULL(ars,'')) <> ''")
            raw_to_update = [r["ars"] for r in cur.fetchall() if normalizar_seguro(r["ars"], r["nss"]) == actual]
            for raw in set(raw_to_update):
                cur.execute("UPDATE pacientes SET ars='SIN SEGURO',updated_at=datetime('now','localtime') WHERE ars=?", (raw,))
                total += cur.rowcount
            conn.commit()
        return total

    def restaurar_atencion_snapshot(self, atencion: dict):
        raise RuntimeError(
            "La reinserción heredada está deshabilitada: las anulaciones conservan el registro y su ID."
        )


    def resumen_turno_actual(self):
        turno_cfg = cargar_turno_config() or cargar_turno_config(permitir_vencido=True)
        base = {"total": 0, "sin_seguro": 0, "GENERAL": 0, "PEDIATRIA": 0, "GINECOLOGIA": 0, "URGENCIAS": 0, "CONSULTAS": 0}
        if not turno_cfg:
            return base
        contexto = self.buscar_contexto_turno_existente(turno_cfg)
        if not contexto:
            excel = resumen_excel_actual_simple(turno_cfg=turno_cfg)
            if int(excel.get("total", 0) or 0):
                excel["URGENCIAS"] = 0
                excel["CONSULTAS"] = 0
                excel["_fuente"] = "EXCEL_RECUPERADO"
                return excel
            return base
        filas = self.obtener_atenciones_para_rango_real(
            turno_id=int(contexto["turno_id"])
        )
        resumen = dict(base)
        for f in filas:
            tipo = (f.get("tipo_atencion") or "EMERGENCIA").strip().upper()
            # El total del turno representa EXCLUSIVAMENTE emergencias.
            # URGENCIA y CONSULTA se contabilizan siempre por separado.
            if tipo == "URGENCIA":
                resumen["URGENCIAS"] += 1
                continue
            if tipo == "CONSULTA":
                resumen["CONSULTAS"] += 1
                continue
            if tipo != "EMERGENCIA":
                continue

            resumen["total"] += 1
            if normalizar_seguro(f.get("ars", ""), f.get("nss", "")) == "SIN SEGURO":
                resumen["sin_seguro"] += 1
            hoja = (f.get("hoja", "") or "").upper()
            if hoja in resumen:
                resumen[hoja] += 1
        resumen["_fuente"] = "BD"

        # Si se recuperó un Excel del mismo turno y contiene más filas que la BD,
        # se conserva visible su conteo en vez de presentar un cero engañoso.
        excel = resumen_excel_actual_simple(turno_cfg=turno_cfg)
        # Si existen urgencias/consultas, el Excel heredado no puede sustituir
        # el total porque no contiene el tipo de atención. La BD es la fuente
        # capaz de separar correctamente esos conteos.
        hay_conteos_separados = bool(resumen.get("URGENCIAS", 0) or resumen.get("CONSULTAS", 0))
        if (not hay_conteos_separados) and int(excel.get("total", 0) or 0) > int(resumen.get("total", 0) or 0):
            excel["URGENCIAS"] = 0
            excel["CONSULTAS"] = 0
            excel["_fuente"] = "EXCEL_RECUPERADO"
            return excel
        if (not hay_conteos_separados) and int(excel.get("total", 0) or 0) == int(resumen.get("total", 0) or 0):
            resumen["_fuente"] = "BD_EXCEL"
        return resumen

    def obtener_atenciones_para_reporte(self, fecha_inicio=None, fecha_fin=None):
        data = self.obtener_atenciones_para_rango_real(fecha_inicio, fecha_fin)
        for item in data:
            dt_real = item.get("dt_real")
            item["dt_operativo"] = (
                dt_real - timedelta(days=1) if dt_real and dt_real.time() < time(8, 0) else dt_real
            )
        return sorted(data, key=lambda item: int(item.get("id", 0)), reverse=True)

    def obtener_atenciones_para_rango_real(
        self,
        fecha_inicio=None,
        fecha_fin=None,
        *,
        turno_id=None,
        dia_operativo_id=None,
        operational_turn_id=None,
        operational_source_id=None,
    ):
        with closing(self._connect()) as conn:
            conn.row_factory = sqlite3.Row
            where = ["a.estado='ACTIVA'"]
            params = []
            if fecha_inicio:
                where.append("datetime(a.created_at) >= datetime(?)")
                params.append(fecha_inicio.strftime("%Y-%m-%d %H:%M:%S"))
            if fecha_fin:
                where.append("datetime(a.created_at) < datetime(?)")
                params.append(fecha_fin.strftime("%Y-%m-%d %H:%M:%S"))
            if turno_id is not None:
                where.append("a.turno_id=?")
                params.append(int(turno_id))
            if operational_turn_id is not None:
                # ``turno_id`` is a legacy SQLite key.  The central turn
                # identity is durable across replicas and must therefore be
                # used for current-turn datasets.
                where.append("COALESCE(a.operational_turn_id,a.turno_id)=?")
                params.append(int(operational_turn_id))
            if operational_source_id:
                where.append("a.operational_source_id=?")
                params.append(str(operational_source_id))
            if dia_operativo_id is not None:
                where.append("a.dia_operativo_id=?")
                params.append(int(dia_operativo_id))
            rows = conn.execute(
                """
                SELECT a.id,a.paciente_id,a.dia_operativo_id,a.turno_id,
                       a.operational_turn_id,a.operational_source_id,
                       a.fecha,a.hora,a.created_at,
                       a.nombre,a.hoja,a.ars,a.nss,a.cedula,a.edad_num,a.unidad,a.tipo_atencion,
                       t.representante,t.tipo_turno,d.fecha_base
                FROM atenciones a
                JOIN turnos t ON t.id=a.turno_id
                JOIN dias_operativos d ON d.id=a.dia_operativo_id
                WHERE """ + " AND ".join(where) + " ORDER BY a.id DESC",
                params,
            ).fetchall()

            data = []
            for r in rows:
                item = dict(r)
                try:
                    dt_real = datetime.strptime(str(item.get("created_at") or "")[:19], "%Y-%m-%d %H:%M:%S")
                except ValueError:
                    dt_real = construir_datetime_real(item.get("fecha", ""), item.get("hora", ""))

                canon = normalizar_seguro(item.get("ars", ""), item.get("nss", ""))
                item["dt_real"] = dt_real
                item["ars_normalizado"] = canon
                item["ars_display"] = seguro_para_mostrar(canon)
                item["hoja_normalizada"] = (item.get("hoja", "") or "").strip().upper() or "SIN ESPECIALIDAD"
                item["tipo_atencion"] = (item.get("tipo_atencion") or "EMERGENCIA").strip().upper()
                data.append(item)

            return sorted(data, key=lambda item: int(item.get("id", 0)), reverse=True)

    def obtener_metadatos_reporte(self, registros):
        representantes = sorted(
            {
                limpiar_nombre_representante(row.get("representante") or "")
                for row in registros
                if limpiar_nombre_representante(row.get("representante") or "")
            }
        )
        turnos = sorted(
            {
                (
                    row.get("fecha_base") or "",
                    normalizar_turno_codigo(row.get("tipo_turno") or "8AM_8AM"),
                )
                for row in registros
            }
        )
        if len(representantes) == 1:
            representante = representantes[0]
        elif representantes:
            representante = "Varios: " + ", ".join(representantes)
        else:
            representante = ""
        if len(turnos) == 1:
            turno_resumen = obtener_datos_turno_visual(
                date.fromisoformat(turnos[0][0]), turnos[0][1]
            )["turno_resumen"]
        elif turnos:
            turno_resumen = f"{len(turnos)} turnos en el período"
        else:
            turno_resumen = None
        return turno_resumen, representante


# -------------------------------
# EXCEL
# -------------------------------
def guardar_excel_seguro(wb, ruta_excel=EXCEL_PATH, accion="guardar el Excel", interactivo=True):
    while True:
        try:
            wb.save(ruta_excel)
            return True
        except PermissionError:
            if not interactivo:
                try:
                    wb.close()
                except Exception:
                    pass
                raise
            try:
                retry = messagebox.askretrycancel(
                    "Excel abierto",
                    "El listado de Excel está abierto.\n\n"
                    "Cierre el archivo y presione 'Reintentar'.\n\n"
                    f"Acción pendiente: {accion}."
                )
            except Exception:
                retry = False
            if retry:
                continue
            try: wb.close()
            except Exception: pass
            return False
        except Exception as e:
            try: wb.close()
            except Exception: pass
            if not interactivo:
                raise
            try: messagebox.showwarning("Excel", f"No se pudo {accion}:\n{e}")
            except Exception: print(f"[AVISO] No se pudo {accion}: {e}")
            return False


def aplicar_formato_excel(ws):
    ws.column_dimensions["A"].width = 5
    ws.column_dimensions["B"].width = 70
    ws.column_dimensions["C"].width = 22
    ws.column_dimensions["D"].width = 18

    for row in range(1, 6):
        ws.row_dimensions[row].height = 18

    thin = XLSide(style="thin", color="D9D9D9")
    border = XLBorder(left=thin, right=thin, top=thin, bottom=thin)

    for row in range(1, 5):
        cell = ws[f"A{row}"]
        cell.font = XLFont(name="Calibri", size=11, bold=False)
        cell.alignment = XLAlignment(horizontal="left", vertical="center")

    for col in range(1, 5):
        cell = ws.cell(row=5, column=col)
        cell.font = XLFont(name="Calibri", size=11, bold=False)
        cell.alignment = XLAlignment(horizontal="left", vertical="center")
        cell.border = border

    for row in range(6, ws.max_row + 1):
        ws.row_dimensions[row].height = 18
        for col in range(1, 5):
            cell = ws.cell(row=row, column=col)
            cell.font = XLFont(name="Calibri", size=11, bold=False)
            cell.border = border
            if col == 1:
                cell.alignment = XLAlignment(horizontal="right", vertical="center")
            else:
                cell.alignment = XLAlignment(horizontal="left", vertical="center", wrap_text=False)

    try:
        orientacion_excel = str(app_setting("print_excel_orientation", "Horizontal")).lower()
        ws.page_setup.orientation = "landscape" if orientacion_excel.startswith("h") else "portrait"
        ws.page_setup.fitToWidth = 1
        ws.page_setup.fitToHeight = 0
        ws.sheet_properties.pageSetUpPr.fitToPage = True
        ws.print_area = f"A1:D{max(ws.max_row, 6)}"
        ws.freeze_panes = "A6"
    except Exception:
        pass


def es_error_excel_corrupto(exc) -> bool:
    msg = str(exc or "").lower()
    return (
        isinstance(exc, (zipfile.BadZipFile, zlib.error))
        or "decompress" in msg
        or "decompressing data" in msg
        or "invalid distance" in msg
        or "bad crc" in msg
        or "file is not a zip file" in msg
    )


def recrear_excel_basico_por_corrupcion():
    try:
        if os.path.exists(EXCEL_PATH):
            corrupt_name = os.path.join(
                os.path.dirname(EXCEL_PATH),
                f"LISTADO_CORRUPTO_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            )
            try:
                shutil.move(EXCEL_PATH, corrupt_name)
            except Exception:
                try:
                    os.remove(EXCEL_PATH)
                except Exception:
                    pass

        wb = Workbook()
        ws = wb.active
        ws.title = "Pacientes"

        ws.merge_cells('A1:D1')
        ws.merge_cells('A2:D2')
        ws.merge_cells('A3:D3')
        ws.merge_cells('A4:D4')

        ws['A1'] = "ASISTENCIA DE PACIENTES A EMERGENCIA"
        ws['A2'] = "ASEGURADOS Y NO ASEGURADOS"
        ws['A3'] = ""
        ws['A4'] = ""

        ws['A5'] = "NO."
        ws['B5'] = "NOMBRE"
        ws['C5'] = "ESPECIALIDAD"
        ws['D5'] = "ARS"

        aplicar_formato_excel(ws)
        guardar_excel_seguro(wb, EXCEL_PATH, "recrear el listado de Excel")
        return True
    except Exception:
        return False


def abrir_excel_workbook_seguro(ruta_excel=None, mostrar_error=True, **kwargs):
    if ruta_excel is None:
        ruta_excel = EXCEL_PATH

    try:
        return openpyxl.load_workbook(ruta_excel, **kwargs)
    except Exception as e:
        if es_error_excel_corrupto(e):
            if not mostrar_error:
                raise
            messagebox.showwarning(
                "Excel dañado",
                "El listado de Excel presentó un error de compresión o corrupción.\n\n"
                "Se creará un Excel nuevo y se intentará reconstruir con los datos del turno actual."
            )
            recrear_excel_basico_por_corrupcion()
            return openpyxl.load_workbook(ruta_excel, **kwargs)
        raise

def verificar_o_crear_excel():
    if not os.path.exists(EXCEL_PATH):
        wb = Workbook()
        ws = wb.active
        ws.title = "Pacientes"

        ws.merge_cells('A1:D1')
        ws.merge_cells('A2:D2')
        ws.merge_cells('A3:D3')
        ws.merge_cells('A4:D4')

        ws['A1'] = "ASISTENCIA DE PACIENTES A EMERGENCIA"
        ws['A2'] = "ASEGURADOS Y NO ASEGURADOS"
        ws['A3'] = ""
        ws['A4'] = ""

        ws['A5'] = "NO."
        ws['B5'] = "NOMBRE"
        ws['C5'] = "ESPECIALIDAD"
        ws['D5'] = "ARS"

        aplicar_formato_excel(ws)
        guardar_excel_seguro(wb, EXCEL_PATH, "crear el listado de Excel")


def actualizar_encabezado_excel(representante: str, turno_label: str, fecha_label: str):
    if not os.path.exists(EXCEL_PATH):
        verificar_o_crear_excel()

    wb = abrir_excel_workbook_seguro(EXCEL_PATH)
    ws = wb.active

    for rng in ('A1:D1', 'A2:D2', 'A3:D3', 'A4:D4'):
        try:
            ws.unmerge_cells(rng)
        except Exception:
            pass
        ws.merge_cells(rng)

    ws['A1'] = "ASISTENCIA DE PACIENTES A EMERGENCIA"
    ws['A2'] = "ASEGURADOS Y NO ASEGURADOS"
    ws['A3'] = f"{limpiar_nombre_representante(representante)} {fecha_label}".strip()
    ws['A4'] = turno_label

    ws['A5'] = "NO."
    ws['B5'] = "NOMBRE"
    ws['C5'] = "ESPECIALIDAD"
    ws['D5'] = "ARS"

    aplicar_formato_excel(ws)
    guardar_excel_seguro(wb, EXCEL_PATH, "actualizar el encabezado del Excel")


def limpiar_filas_excel():
    if not os.path.exists(EXCEL_PATH):
        verificar_o_crear_excel()

    wb = abrir_excel_workbook_seguro(EXCEL_PATH)
    ws = wb.active
    max_row = ws.max_row

    if max_row >= 6:
        ws.delete_rows(6, max_row - 5)

    ws['A5'] = "NO."
    ws['B5'] = "NOMBRE"
    ws['C5'] = "ESPECIALIDAD"
    ws['D5'] = "ARS"

    aplicar_formato_excel(ws)
    guardar_excel_seguro(wb, EXCEL_PATH, "limpiar las filas del Excel")


def _obtener_numeros_usados_excel(ws):
    usados = set()
    for fila in range(6, ws.max_row + 1):
        valor = ws.cell(row=fila, column=1).value
        try:
            if valor is not None and str(valor).strip() != "":
                usados.add(int(valor))
        except Exception:
            pass
    return usados


def _primer_numero_libre(usados):
    n = 1
    while n in usados:
        n += 1
    return n


def _identidad_turno_central(db, *, turn_id=None, operational_source_id=None):
    """Obtiene una identidad central completa; no inventa un rango por fecha."""
    snapshot_provider = getattr(db, "get_operational_station_snapshot", None)
    snapshot = {}
    if callable(snapshot_provider):
        snapshot = dict(snapshot_provider() or {})
    effective_turn_id = turn_id if turn_id is not None else snapshot.get("turn_id")
    effective_source_id = (
        operational_source_id
        if operational_source_id is not None
        else snapshot.get("operational_source_id")
    )
    try:
        effective_turn_id = int(effective_turn_id)
    except (TypeError, ValueError):
        effective_turn_id = 0
    effective_source_id = str(effective_source_id or "").strip()
    return effective_turn_id, effective_source_id


def _dataset_turno_central(db, *, turn_id=None, operational_source_id=None):
    builder = getattr(db, "build_turn_dataset", None)
    if not callable(builder):
        return None
    effective_turn_id, effective_source_id = _identidad_turno_central(
        db,
        turn_id=turn_id,
        operational_source_id=operational_source_id,
    )
    if effective_turn_id <= 0 or not effective_source_id:
        raise TurnoNoVigenteError("TURN_ID_NOT_AVAILABLE")
    rows = list(
        builder(
            turn_id=effective_turn_id,
            operational_source_id=effective_source_id,
        )
        or []
    )
    return rows, effective_turn_id, effective_source_id


def construir_hoja_listado_pacientes(
    ws,
    filas,
    *,
    encabezado_linea_3="",
    encabezado_linea_4="",
    revision=None,
):
    """Build the single official operational patient-listing format."""
    filas = list(filas or [])
    for rng in ('A1:D1', 'A2:D2', 'A3:D3', 'A4:D4'):
        try:
            ws.unmerge_cells(rng)
        except (KeyError, ValueError):
            pass
        ws.merge_cells(rng)
    ws['A1'] = "ASISTENCIA DE PACIENTES A EMERGENCIA"
    ws['A2'] = "ASEGURADOS Y NO ASEGURADOS"
    ws['A3'] = str(encabezado_linea_3 or "")
    ws['A4'] = str(encabezado_linea_4 or "")
    ws['A5'] = "NO."
    ws['B5'] = "NOMBRE"
    ws['C5'] = "ESPECIALIDAD"
    ws['D5'] = "ARS"
    ws['E5'] = "GLOBAL_ATTENTION_ID"
    ws['F1'] = str(revision if revision is not None else _admission_dataset_revision(filas))
    ws.column_dimensions['E'].hidden = True
    ws.column_dimensions['F'].hidden = True
    if ws.max_row >= 6:
        ws.delete_rows(6, ws.max_row - 5)
    for numero, fila in enumerate(filas, start=1):
        ws.append([
            numero,
            str(fila.get("nombre") or fila.get("patient_name") or "SIN NOMBRE").upper(),
            fila.get("hoja_normalizada", fila.get("specialty", fila.get("hoja", ""))),
            fila.get("ars_display", fila.get("canonical_ars", fila.get("ars", "SIN SEGURO"))),
            str(fila.get("global_attention_id") or ""),
        ])
    aplicar_formato_excel(ws)
    return len(filas)


def _construir_workbook_turno(db: DatabaseManager, turno_cfg: dict):
    if not turno_cfg:
        raise TurnoNoVigenteError("No hay turno vigente para reconstruir el listado.")

    datos_turno = obtener_datos_turno_visual(turno_cfg["fecha_base"], turno_cfg["turno_codigo"])
    inicio, fin = obtener_rango_turno_efectivo(turno_cfg)
    central_dataset = _dataset_turno_central(db)
    if central_dataset is not None:
        filas, _, _ = central_dataset
    else:
        contexto = db.buscar_contexto_turno_existente(turno_cfg)
        turn_id = int(contexto["turno_id"]) if contexto else None
        filas = db.obtener_atenciones_para_rango_real(
            fecha_inicio=None if contexto else inicio,
            fecha_fin=None if contexto else fin,
            turno_id=turn_id,
        )

    # El documento derivado se construye desde cero. No se abre ni se apropia
    # del libro que el usuario pueda tener abierto en Microsoft Excel.
    wb = Workbook()
    wb.active.title = "Pacientes"
    ws = wb.active
    total = construir_hoja_listado_pacientes(
        ws,
        filas,
        encabezado_linea_3=(
            f"{limpiar_nombre_representante(turno_cfg.get('representante', ''))} "
            f"{datos_turno['fecha_label']}"
        ).strip(),
        encabezado_linea_4=datos_turno["turno_label"],
    )
    return wb, total


def reconstruir_excel_turno(db: DatabaseManager, turno_cfg: dict):
    wb, total = _construir_workbook_turno(db, turno_cfg)
    if total == 0:
        revision = str(wb.active['F1'].value or "")
        wb.close()
        export_state = _read_excel_export_state()
        if (
            export_state.get("excel_dataset_revision") == revision
            and export_state.get("excel_status") == "SKIPPED_EMPTY"
            and int(export_state.get("patient_count") or 0) == 0
        ):
            return 0
        _write_excel_export_state(
            revision=revision,
            status="SKIPPED_EMPTY",
            patient_count=0,
        )
        APP_LOG.info("ADMISSION_EXCEL_SKIPPED_EMPTY patient_count=0")
        return 0
    revision = str(wb.active['F1'].value or "")
    export_state = _read_excel_export_state()
    if (
        export_state.get("excel_dataset_revision") == revision
        and export_state.get("excel_status") == "SYNCED"
        and os.path.isfile(EXCEL_PATH)
    ):
        wb.close()
        return total
    temp_excel = EXCEL_LATEST_PATH + ".tmp.xlsx"
    try:
        wb.save(temp_excel)
        wb.close()
        os.replace(temp_excel, EXCEL_LATEST_PATH)
    except Exception:
        try:
            wb.close()
        except Exception:
            pass
        try:
            if os.path.exists(temp_excel):
                os.remove(temp_excel)
        except OSError:
            pass
        raise
    try:
        _update_canonical_excel(EXCEL_LATEST_PATH, EXCEL_PATH)
    except Exception as exc:
        if not _excel_file_in_use(exc):
            raise
        _write_excel_export_state(
            revision=revision,
            status="FILE_LOCKED",
            patient_count=total,
        )
        enqueue_excel_export_job(
            f"dataset:{revision}",
            int((db.buscar_contexto_turno_existente(turno_cfg) or {}).get("turno_id") or 0),
            turno_cfg,
            source_file=EXCEL_LATEST_PATH,
        )
        APP_LOG.warning("EXCEL_UPDATE_PENDING_FILE_LOCKED revision=%s", revision)
        return total
    _write_excel_export_state(
        revision=revision,
        status="SYNCED",
        patient_count=total,
    )
    return total


def _excel_file_in_use(exc: BaseException) -> bool:
    return isinstance(exc, PermissionError) or int(getattr(exc, "winerror", 0) or 0) in (5, 32)


def _admission_dataset_revision(rows) -> str:
    """Hash estable del contenido/orden; no depende de metadatos binarios XLSX."""
    normalized = []
    for position, raw in enumerate(rows, start=1):
        row = dict(raw or {})
        normalized.append({
            "position": position,
            "global_attention_id": str(row.get("global_attention_id") or ""),
            "attention_id": int(row.get("attention_id") or row.get("id") or 0),
            "name": str(row.get("nombre") or row.get("patient_name") or "").strip().upper(),
            "specialty": str(
                row.get("hoja_normalizada") or row.get("specialty") or row.get("hoja") or ""
            ).strip().upper(),
            "ars": str(
                row.get("ars_display") or row.get("canonical_ars") or row.get("ars") or "SIN SEGURO"
            ).strip().upper(),
        })
    payload = json.dumps(normalized, ensure_ascii=False, sort_keys=True, separators=(",", ":"))
    return hashlib.sha256(payload.encode("utf-8")).hexdigest()


def _excel_dataset_revision(path: str) -> str:
    try:
        workbook = load_workbook(path, read_only=True, data_only=True)
        try:
            stored = str(workbook.active['F1'].value or "").strip()
            if stored:
                return stored
        finally:
            workbook.close()
    except (OSError, ValueError, TypeError):
        pass
    digest = hashlib.sha256()
    with open(path, "rb") as handle:
        for block in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(block)
    return digest.hexdigest()


def _read_excel_export_state() -> dict:
    try:
        with open(EXCEL_EXPORT_STATE_PATH, "r", encoding="utf-8") as handle:
            return dict(json.load(handle) or {})
    except (OSError, ValueError, TypeError):
        return {}


def _write_excel_export_state(*, revision: str, status: str, patient_count: int) -> None:
    payload = {
        "excel_dataset_revision": str(revision or ""),
        "excel_last_generated_at": datetime.now().isoformat(timespec="seconds"),
        "excel_status": str(status or "ERROR"),
        "patient_count": int(patient_count or 0),
    }
    temp_path = EXCEL_EXPORT_STATE_PATH + ".tmp"
    with open(temp_path, "w", encoding="utf-8") as handle:
        json.dump(payload, handle, ensure_ascii=False, sort_keys=True, indent=2)
    os.replace(temp_path, EXCEL_EXPORT_STATE_PATH)


def excel_canonical_in_use(path: str = EXCEL_PATH) -> bool:
    """Comprueba el lock sin abrir, cerrar ni automatizar Microsoft Excel."""
    if not os.path.isfile(path) or platform.system() != "Windows":
        return False
    create_file = ctypes.windll.kernel32.CreateFileW
    create_file.restype = ctypes.c_void_p
    handle = create_file(
        os.path.abspath(path),
        0x00010000,  # DELETE: el permiso requerido por os.replace
        0x00000001 | 0x00000002 | 0x00000004,
        None,
        3,           # OPEN_EXISTING
        0x80,        # FILE_ATTRIBUTE_NORMAL
        None,
    )
    invalid = ctypes.c_void_p(-1).value
    if handle in (None, invalid):
        return int(ctypes.windll.kernel32.GetLastError() or 0) in (5, 32)
    ctypes.windll.kernel32.CloseHandle(handle)
    return False


def _turno_cfg_to_json(turno_cfg: dict) -> str:
    fecha_base = turno_cfg.get("fecha_base")
    inicio_real_dt = turno_cfg.get("inicio_real_dt")
    payload = {
        "representante": limpiar_nombre_representante(turno_cfg.get("representante", "")),
        "turno_codigo": normalizar_turno_codigo(turno_cfg.get("turno_codigo", "8AM_8AM")),
        "fecha_base": fecha_base.isoformat() if isinstance(fecha_base, date) else str(fecha_base or ""),
        "inicio_real": str(turno_cfg.get("inicio_real") or ""),
        "inicio_real_dt": (
            inicio_real_dt.isoformat(timespec="seconds")
            if isinstance(inicio_real_dt, datetime)
            else str(inicio_real_dt or "")
        ),
    }
    return json.dumps(payload, ensure_ascii=False, sort_keys=True)


def _turno_cfg_from_json(payload: str) -> dict:
    raw = json.loads(payload or "{}")
    fecha_texto = str(raw.get("fecha_base") or "")
    inicio_texto = str(raw.get("inicio_real_dt") or raw.get("inicio_real") or "")
    try:
        fecha_base = date.fromisoformat(fecha_texto[:10])
    except ValueError:
        fecha_base = parse_fecha_ddmmyyyy(fecha_texto)
    try:
        inicio_real_dt = datetime.fromisoformat(inicio_texto)
    except ValueError:
        inicio_real_dt = parse_datetime_local(inicio_texto)
    if not fecha_base:
        raise TurnoNoVigenteError("La tarea de Excel no contiene una fecha de turno válida.")
    return {
        "representante": limpiar_nombre_representante(raw.get("representante", "")),
        "turno_codigo": normalizar_turno_codigo(raw.get("turno_codigo", "8AM_8AM")),
        "fecha_base": fecha_base,
        "inicio_real": str(raw.get("inicio_real") or ""),
        "inicio_real_dt": inicio_real_dt,
    }


def _ensure_excel_export_queue() -> None:
    os.makedirs(os.path.dirname(EXCEL_EXPORT_QUEUE_PATH), exist_ok=True)
    os.makedirs(EXCEL_VERSIONED_DIR, exist_ok=True)
    with sqlite3.connect(EXCEL_EXPORT_QUEUE_PATH, timeout=15) as conn:
        conn.execute(
            """
            CREATE TABLE IF NOT EXISTS excel_export_jobs(
                job_id TEXT PRIMARY KEY,
                transition_id TEXT NOT NULL UNIQUE,
                turn_id INTEGER,
                turno_json TEXT NOT NULL,
                source_file TEXT,
                canonical_target TEXT NOT NULL,
                status TEXT NOT NULL DEFAULT 'PENDING',
                attempts INTEGER NOT NULL DEFAULT 0,
                last_error_code TEXT,
                created_at TEXT NOT NULL,
                updated_at TEXT NOT NULL,
                completed_at TEXT,
                CHECK(status IN ('PENDING','RUNNING','COMPLETED','ERROR'))
            )
            """
        )
        # Una caída no deja el job tomado indefinidamente.
        conn.execute(
            "UPDATE excel_export_jobs SET status='PENDING',updated_at=? WHERE status='RUNNING'",
            (datetime.now().isoformat(timespec="seconds"),),
        )


def enqueue_excel_export_job(
    transition_id: str,
    turn_id: int,
    turno_cfg: dict,
    *,
    source_file: str = "",
) -> str:
    import uuid

    _ensure_excel_export_queue()
    transition_id = str(transition_id or uuid.uuid4())
    job_id = str(uuid.uuid4())
    ahora = datetime.now().isoformat(timespec="seconds")
    with sqlite3.connect(EXCEL_EXPORT_QUEUE_PATH, timeout=15) as conn:
        conn.execute(
            """
            INSERT INTO excel_export_jobs(
                job_id,transition_id,turn_id,turno_json,source_file,
                canonical_target,status,attempts,last_error_code,
                created_at,updated_at,completed_at
            ) VALUES(?,?,?,?,?,?,'PENDING',0,NULL,?,?,NULL)
            ON CONFLICT(transition_id) DO NOTHING
            """,
            (
                job_id, transition_id, int(turn_id or 0),
                _turno_cfg_to_json(turno_cfg), str(source_file or "") or None,
                EXCEL_PATH, ahora, ahora,
            ),
        )
        row = conn.execute(
            "SELECT job_id FROM excel_export_jobs WHERE transition_id=?",
            (transition_id,),
        ).fetchone()
    return str(row[0] if row else job_id)


def _versioned_excel_path(turn_id: int, transition_id: str) -> str:
    os.makedirs(EXCEL_VERSIONED_DIR, exist_ok=True)
    safe_transition = re.sub(r"[^A-Za-z0-9]", "", str(transition_id or ""))[:12] or "local"
    return os.path.join(
        EXCEL_VERSIONED_DIR,
        f"LISTADO_DE_PACIENTES_TURNO_{int(turn_id or 0)}_{safe_transition}.xlsx",
    )


def _generate_versioned_excel(
    db: DatabaseManager,
    turno_cfg: dict,
    *,
    turn_id: int,
    transition_id: str,
) -> tuple[str, int]:
    target = _versioned_excel_path(turn_id, transition_id)
    if os.path.isfile(target):
        return target, 1 if excel_tiene_registros(target) else 0
    wb, total = _construir_workbook_turno(db, turno_cfg)
    if total == 0:
        wb.close()
        APP_LOG.info(
            "SHIFT_EXCEL_SKIPPED_EMPTY turn_id=%s patient_count=0",
            int(turn_id or 0),
        )
        return "", 0
    temp_target = target + ".tmp.xlsx"
    try:
        wb.save(temp_target)
        wb.close()
        os.replace(temp_target, target)
    except Exception:
        try:
            wb.close()
        except Exception:
            pass
        try:
            if os.path.exists(temp_target):
                os.remove(temp_target)
        except OSError:
            pass
        raise
    return target, total


def _update_canonical_excel(source_file: str, canonical_target: str = EXCEL_PATH) -> None:
    temp_target = canonical_target + ".pending.xlsx"
    try:
        shutil.copy2(source_file, temp_target)
        os.replace(temp_target, canonical_target)
    except Exception:
        try:
            if os.path.exists(temp_target):
                os.remove(temp_target)
        except OSError:
            pass
        raise


def synchronize_latest_excel() -> str:
    """Publica latest sin tomar control de Excel; FILE_LOCKED no es un fallo clínico."""
    if not os.path.isfile(EXCEL_LATEST_PATH):
        return "NO_LATEST"
    revision = _excel_dataset_revision(EXCEL_LATEST_PATH)
    state = _read_excel_export_state()
    if (
        state.get("excel_dataset_revision") == revision
        and state.get("excel_status") == "SYNCED"
        and os.path.isfile(EXCEL_PATH)
    ):
        return "SYNCED"
    try:
        _update_canonical_excel(EXCEL_LATEST_PATH, EXCEL_PATH)
    except Exception as exc:
        if not _excel_file_in_use(exc):
            raise
        _write_excel_export_state(
            revision=revision,
            status="FILE_LOCKED",
            patient_count=int(state.get("patient_count") or 0),
        )
        return "FILE_LOCKED"
    _write_excel_export_state(
        revision=revision,
        status="SYNCED",
        patient_count=int(state.get("patient_count") or 0),
    )
    return "SYNCED"


def pending_excel_export_jobs() -> int:
    _ensure_excel_export_queue()
    with sqlite3.connect(EXCEL_EXPORT_QUEUE_PATH, timeout=15) as conn:
        return int(
            conn.execute(
                "SELECT COUNT(*) FROM excel_export_jobs WHERE status='PENDING'"
            ).fetchone()[0]
        )


def process_excel_export_jobs(db: DatabaseManager, *, limit: int = 3) -> dict:
    """Procesa efectos post-commit; nunca cambia ni revierte un turno."""
    _ensure_excel_export_queue()
    resultado = {
        "completed": 0,
        "pending": 0,
        "errors": 0,
        "processed": 0,
        "skipped_empty": 0,
    }
    with sqlite3.connect(EXCEL_EXPORT_QUEUE_PATH, timeout=15) as conn:
        conn.row_factory = sqlite3.Row
        jobs = conn.execute(
            """
            SELECT * FROM excel_export_jobs
            WHERE status='PENDING'
            ORDER BY created_at,job_id LIMIT ?
            """,
            (max(1, int(limit or 1)),),
        ).fetchall()

    for job in jobs:
        job_id = str(job["job_id"])
        ahora = datetime.now().isoformat(timespec="seconds")
        with sqlite3.connect(EXCEL_EXPORT_QUEUE_PATH, timeout=15) as conn:
            claimed = conn.execute(
                """
                UPDATE excel_export_jobs
                SET status='RUNNING',attempts=attempts+1,updated_at=?
                WHERE job_id=? AND status='PENDING'
                """,
                (ahora, job_id),
            ).rowcount
        if claimed != 1:
            continue
        resultado["processed"] += 1
        try:
            source_file = str(job["source_file"] or "")
            total = None
            if not source_file or not os.path.isfile(source_file):
                source_file, total = _generate_versioned_excel(
                    db,
                    _turno_cfg_from_json(str(job["turno_json"])),
                    turn_id=int(job["turn_id"] or 0),
                    transition_id=str(job["transition_id"]),
                )
                with sqlite3.connect(EXCEL_EXPORT_QUEUE_PATH, timeout=15) as conn:
                    conn.execute(
                        "UPDATE excel_export_jobs SET source_file=?,updated_at=? WHERE job_id=?",
                        (source_file, ahora, job_id),
                    )
            if total == 0 or (
                source_file
                and os.path.isfile(source_file)
                and not excel_tiene_registros(source_file)
            ):
                with sqlite3.connect(EXCEL_EXPORT_QUEUE_PATH, timeout=15) as conn:
                    conn.execute(
                        """UPDATE excel_export_jobs
                           SET status='COMPLETED',last_error_code='SKIPPED_EMPTY',
                               updated_at=?,completed_at=? WHERE job_id=?""",
                        (ahora, ahora, job_id),
                    )
                resultado["completed"] += 1
                resultado["skipped_empty"] += 1
                continue
            _update_canonical_excel(source_file, str(job["canonical_target"] or EXCEL_PATH))
            state = _read_excel_export_state()
            _write_excel_export_state(
                revision=_excel_dataset_revision(source_file),
                status="SYNCED",
                patient_count=int(state.get("patient_count") or 0),
            )
        except Exception as exc:
            file_in_use = _excel_file_in_use(exc)
            error_code = (
                "EXCEL_EXPORT_DEFERRED_FILE_IN_USE"
                if file_in_use else f"EXCEL_EXPORT_{type(exc).__name__.upper()}"
            )
            with sqlite3.connect(EXCEL_EXPORT_QUEUE_PATH, timeout=15) as conn:
                conn.execute(
                    """
                    UPDATE excel_export_jobs
                    SET status=?,last_error_code=?,updated_at=? WHERE job_id=?
                    """,
                    ("PENDING" if file_in_use else "ERROR", error_code, ahora, job_id),
                )
            if file_in_use:
                resultado["pending"] += 1
                APP_LOG.warning("EXCEL_EXPORT_DEFERRED_FILE_IN_USE transition=%s", job["transition_id"])
            else:
                resultado["errors"] += 1
                APP_LOG.exception("Falló el efecto post-commit de Excel transition=%s", job["transition_id"])
            continue
        with sqlite3.connect(EXCEL_EXPORT_QUEUE_PATH, timeout=15) as conn:
            conn.execute(
                """
                UPDATE excel_export_jobs
                SET status='COMPLETED',last_error_code=NULL,updated_at=?,completed_at=?
                WHERE job_id=?
                """,
                (ahora, ahora, job_id),
            )
        resultado["completed"] += 1
    return resultado


def actualizar_representante_turno_actual(
    db: DatabaseManager,
    representante: str,
) -> dict:
    """Cambia solo el responsable del turno, su encabezado y reportes futuros."""
    representante = limpiar_nombre_representante(representante)
    if not es_representante_valido(representante):
        raise ValueError(
            "Escriba un nombre válido. 'No disponible' no puede guardarse como usuario."
        )
    turno_cfg = cargar_turno_config(permitir_vencido=True)
    if not turno_cfg:
        raise TurnoNoVigenteError("No existe un turno configurado para actualizar.")
    contexto = db.buscar_contexto_turno_existente(turno_cfg)
    if not contexto:
        raise TurnoNoVigenteError("No se encontró el turno asociado a la configuración.")

    datos_turno = obtener_datos_turno_visual(
        turno_cfg["fecha_base"], turno_cfg["turno_codigo"]
    )
    # Se comprueba primero que el Excel pueda actualizarse. Si está abierto, no
    # se modifica la configuración ni la base y el operador puede reintentar.
    actualizar_encabezado_excel(
        representante,
        datos_turno["turno_label"],
        datos_turno["fecha_label"],
    )
    inicio_real = turno_cfg.get("inicio_real_dt")
    if not guardar_turno_config(
        representante,
        turno_cfg["turno_codigo"],
        turno_cfg["fecha_base"],
        inicio_real=inicio_real,
    ):
        raise OSError("No se pudo actualizar la configuración del turno.")
    if not db.actualizar_representante_turno(
        int(contexto["turno_id"]), representante
    ):
        raise RuntimeError("No se pudo actualizar el representante en el turno.")

    guardar_representante_catalogo(representante, db)
    actualizado = cargar_turno_config(permitir_vencido=True) or dict(turno_cfg)
    actualizado["representante"] = representante
    actualizado["turno_id"] = int(contexto["turno_id"])
    return actualizado


def agregar_excel_temporal(nombre, especialidad, ars_canonico):
    try:
        if not os.path.exists(EXCEL_PATH):
            verificar_o_crear_excel()

        wb = abrir_excel_workbook_seguro(EXCEL_PATH)
        ws = wb.active

        usados = _obtener_numeros_usados_excel(ws)
        siguiente_numero = _primer_numero_libre(usados)

        ws.append([
            siguiente_numero,
            (nombre or '').upper(),
            especialidad,
            seguro_para_mostrar(ars_canonico)
        ])
        aplicar_formato_excel(ws)
        if not guardar_excel_seguro(wb, EXCEL_PATH, "agregar el paciente al Excel"):
            return False
        return True
    except PermissionError:
        messagebox.showwarning("Archivo abierto", "Cierre el Excel para actualizar.")
        return False
    except Exception as e:
        messagebox.showerror("Error", f"Error al guardar Excel: {str(e)}")
        return False


# -------------------------------
# UTILIDADES GENERALES
# -------------------------------
def formatear_cedula(cedula):
    cedula = (cedula or "").strip()
    if not cedula or is_all_zeros(cedula):
        return "N/A"
    if len(cedula) != 11 or not cedula.isdigit():
        return "N/A"
    return f"{cedula[:3]}-{cedula[3:10]}-{cedula[10:]}"

def formatear_telefono(telefono):
    telefono = (telefono or "").strip()
    if len(telefono) != 10 or not telefono.isdigit():
        return "N/A"
    return f"{telefono[:3]}-{telefono[3:6]}-{telefono[6:]}"

def escribir_log_impresion(mensaje):
    try:
        os.makedirs(LOGS_DIR, exist_ok=True)
        log_path = os.path.join(LOGS_DIR, "impresion.log")
        with open(log_path, "a", encoding="utf-8") as f:
            f.write(f"[{datetime.now().strftime('%d/%m/%Y %I:%M:%S %p')}] {mensaje}\n")
    except Exception:
        pass

def imprimir_pdf(ruta_pdf, copias=1, mostrar_error=False):
    try:
        escribir_log_impresion("Solicitud de impresión recibida.")

        if not ruta_pdf or not os.path.exists(ruta_pdf):
            escribir_log_impresion(f"ERROR: PDF no encontrado: {ruta_pdf}")
            if mostrar_error:
                messagebox.showwarning("Impresión", "No se encontró el PDF para imprimir.")
            return False

        ruta_abs = os.path.abspath(ruta_pdf)
        copias = max(1, int(copias or 1))
        sis = platform.system()

        escribir_log_impresion(f"PDF: {ruta_abs}")
        escribir_log_impresion(f"Copias solicitadas: {copias}")
        escribir_log_impresion(f"Sistema: {sis}")

        if sis == "Windows":
            exe_dir = os.path.dirname(sys.executable) if getattr(sys, "frozen", False) else BASE_DIR
            meipass_dir = getattr(sys, "_MEIPASS", exe_dir)
            cwd_dir = os.getcwd()

            nombres_sumatra = [
                "SumatraPDF.exe",
                "SumatraPDF-3.6.1-64.exe",
                "SumatraPDF.exe.exe",
            ]

            script_dir = os.path.dirname(os.path.abspath(sys.argv[0])) if sys.argv and sys.argv[0] else BASE_DIR
            appdata_dir = os.path.dirname(EXCEL_PATH)

            carpetas_busqueda = []
            for carpeta in [
                exe_dir,
                meipass_dir,
                BASE_DIR,
                cwd_dir,
                script_dir,
                appdata_dir,
                os.path.dirname(os.path.abspath(ruta_abs)),
            ]:
                if carpeta and carpeta not in carpetas_busqueda:
                    carpetas_busqueda.append(carpeta)

            posibles_sumatra = []
            for carpeta in carpetas_busqueda:
                for nombre in nombres_sumatra:
                    posibles_sumatra.append(os.path.join(carpeta, nombre))

            escribir_log_impresion("Buscando SumatraPDF en:")
            for p in posibles_sumatra:
                escribir_log_impresion(f"  - {p}")

            global SUMATRA_PATH_CACHE
            sumatra_path = SUMATRA_PATH_CACHE if SUMATRA_PATH_CACHE and os.path.exists(SUMATRA_PATH_CACHE) else None
            if not sumatra_path:
                for p in posibles_sumatra:
                    if p and os.path.exists(p):
                        sumatra_path = p
                        SUMATRA_PATH_CACHE = p
                        break

            if not sumatra_path:
                escribir_log_impresion("ERROR: No se encontró SumatraPDF en ninguna ruta.")
                if mostrar_error:
                    messagebox.showwarning(
                        "SumatraPDF no encontrado",
                        "No se encontró SumatraPDF.exe.\n\n"
                        "Coloque SumatraPDF.exe en la misma carpeta de la app o inclúyalo al compilar.\n\n"
                        "Se creó/actualizó el archivo debug_impresion.txt para diagnóstico."
                    )
                return False

            escribir_log_impresion(f"SumatraPDF encontrado: {sumatra_path}")

            # Sumatra no devuelve un callback de la impresora. Si Windows tiene
            # impresora predeterminada y acepta el proceso, la hoja quedó enviada
            # a la cola; no se debe convertir ese envío en un falso fallo.
            try:
                required = ctypes.c_uint(0)
                available = ctypes.windll.winspool.GetDefaultPrinterW(None, ctypes.byref(required))
                if not available and required.value == 0:
                    raise RuntimeError("No hay una impresora predeterminada disponible.")
            except AttributeError:
                pass

            ok_general = False
            for i in range(copias):
                comando = [
                    sumatra_path,
                    "-print-to-default",
                    "-silent",
                    ruta_abs
                ]
                escribir_log_impresion(f"Ejecutando copia {i + 1}/{copias}: {' '.join(comando)}")

                try:
                    subprocess.Popen(
                        comando,
                        stdout=subprocess.DEVNULL,
                        stderr=subprocess.DEVNULL,
                        creationflags=getattr(subprocess, "CREATE_NO_WINDOW", 0),
                    )
                    escribir_log_impresion("Trabajo aceptado y enviado a la aplicación de impresión.")
                    ok_general = True
                except Exception as e:
                    escribir_log_impresion(f"ERROR ejecutando SumatraPDF: {str(e)}")

                if copias > 1 and i < copias - 1:
                    try:
                        _time.sleep(0.7)
                    except Exception:
                        pass

            if ok_general:
                escribir_log_impresion("Impresión enviada a la cola correctamente.")
            else:
                escribir_log_impresion("ERROR: No se pudo enviar la impresión a la cola.")

            if mostrar_error and not ok_general:
                messagebox.showwarning(
                    "Impresión",
                    "No se pudo enviar la hoja a la impresora.\n\n"
                    "Revise que exista una impresora predeterminada y que esté disponible.\n\n"
                    "También revise debug_impresion.txt junto a la app."
                )

            return ok_general

        ok_general = False
        for i in range(copias):
            if sis == "Darwin":
                ok = subprocess.run(["lp", ruta_abs], check=False).returncode == 0
            elif sis == "Linux":
                ok = subprocess.run(["lpr", ruta_abs], check=False).returncode == 0
            else:
                ok = False

            ok_general = ok_general or ok

            if copias > 1 and i < copias - 1:
                try:
                    _time.sleep(0.7)
                except Exception:
                    pass

        return ok_general

    except Exception as e:
        escribir_log_impresion(f"ERROR general en imprimir_pdf: {str(e)}")
        if mostrar_error:
            messagebox.showwarning(
                "Impresión",
                f"No se pudo imprimir automáticamente el PDF:\n{str(e)}"
            )
        return False

def resumen_excel_actual_simple(turno_cfg=None):
    resumen = {
        "total": 0,
        "sin_seguro": 0,
        "GENERAL": 0,
        "PEDIATRIA": 0,
        "GINECOLOGIA": 0,
    }
    try:
        if not os.path.exists(EXCEL_PATH):
            return resumen

        wb = abrir_excel_workbook_seguro(EXCEL_PATH, read_only=True, data_only=True)
        ws = wb.active

        if turno_cfg:
            encabezado_fecha = str(ws["A3"].value or "")
            encabezado_turno = str(ws["A4"].value or "")
            fecha_esperada = turno_cfg["fecha_base"].strftime("%d/%m/%Y")
            turno_esperado = normalizar_turno_codigo(
                turno_cfg.get("turno_codigo", "8AM_8AM")
            )
            if (
                fecha_esperada not in encabezado_fecha
                or normalizar_turno_codigo(encabezado_turno) != turno_esperado
            ):
                try:
                    wb.close()
                except Exception:
                    pass
                return resumen

        for fila in range(6, ws.max_row + 1):
            nombre = str(ws.cell(row=fila, column=2).value or "").strip()
            if not nombre:
                continue
            esp = str(ws.cell(row=fila, column=3).value or "").strip().upper()
            ars = str(ws.cell(row=fila, column=4).value or "").strip()

            resumen["total"] += 1
            if esp in resumen:
                resumen[esp] += 1

            if normalizar_seguro(ars, "") == "SIN SEGURO":
                resumen["sin_seguro"] += 1

        try:
            wb.close()
        except Exception:
            pass
    except Exception:
        pass
    return resumen

def excel_tiene_registros(ruta_excel=None):
    if ruta_excel is None:
        ruta_excel = EXCEL_PATH

    try:
        if not os.path.exists(ruta_excel):
            return False

        wb = abrir_excel_workbook_seguro(ruta_excel, read_only=True, data_only=True)
        ws = wb.active

        for fila in range(6, ws.max_row + 1):
            nombre = ws.cell(row=fila, column=2).value
            if nombre is not None and str(nombre).strip():
                try:
                    wb.close()
                except Exception:
                    pass
                return True

        try:
            wb.close()
        except Exception:
            pass
        return False
    except Exception:
        return False

def reintentar_si_excel_abierto(accion):
    while True:
        resp = messagebox.askretrycancel(
            "Excel abierto",
            "El listado de Excel está abierto.\n\n"
            "Cierre el archivo y presione “Reintentar”."
        )
        if not resp:
            return False
        try:
            return bool(accion())
        except PermissionError:
            continue
        except Exception as e:
            messagebox.showwarning("Aviso", f"No se pudo completar la acción:\n{str(e)}")
            return False

def imprimir_excel(ruta_excel=None, copias=1, *, permitir_reintento=True):
    if ruta_excel is None:
        ruta_excel = EXCEL_PATH

    try:
        if not os.path.exists(ruta_excel):
            return False

        if not excel_tiene_registros(ruta_excel):
            return False

        ruta_abs = os.path.abspath(ruta_excel)
        sis = platform.system()
        copias = max(1, int(copias or 1))

        for _ in range(copias):
            if sis == "Windows":
                os.startfile(ruta_abs, "print")
            elif sis == "Darwin":
                subprocess.run(["lp", ruta_abs], check=False)
            elif sis == "Linux":
                subprocess.run(["lpr", ruta_abs], check=False)

        return True

    except PermissionError:
        if permitir_reintento:
            return reintentar_si_excel_abierto(
                lambda: imprimir_excel(
                    ruta_excel, copias, permitir_reintento=permitir_reintento
                )
            )
        return False
    except Exception as e:
        messagebox.showwarning(
            "Aviso",
            f"No se pudo imprimir automáticamente el listado de Excel:\n{str(e)}"
        )
        return False

def abrir_pdf(ruta_pdf, mostrar_error=True):
    try:
        sis = platform.system()
        if sis == "Windows":
            os.startfile(ruta_pdf)
        elif sis == "Darwin":
            subprocess.run(["open", ruta_pdf], check=False)
        elif sis == "Linux":
            subprocess.run(["xdg-open", ruta_pdf], check=False)
        return True
    except Exception as e:
        APP_LOG.exception("PDF_OPEN_ERROR path=%s", os.path.basename(str(ruta_pdf or "")))
        if mostrar_error:
            messagebox.showerror("Error", f"No se pudo abrir el PDF: {str(e)}")
        return False


def sanitize_filename(name: str) -> str:
    keep = "-_.() "
    return "".join(c for c in (name or "") if c.isalnum() or c in keep).strip().replace("  ", " ")


def poner_hora_entre_parentesis(texto: str) -> str:
    if not texto:
        return ""
    return re.sub(r'(?<!\()(\b\d{1,2}:\d{2}\s?[AP]M\b)(?!\))', r'(\1)', texto)


def draw_text_auto(c, x, y, text, base_font="Helvetica", base_size=12, max_len=THRESHOLD_LEN, min_size=8):
    if text is None:
        return base_size
    txt = str(text).strip()
    size = float(base_size)

    if len(txt) > max_len:
        factor = max(min_size / base_size, max_len / len(txt))
        size = max(round(base_size * factor, 1), min_size)

    c.setFont(base_font, size)
    c.drawString(x, y, txt)
    c.setFont("Helvetica", 12)
    return size


def draw_direccion_auto(c, x, y, text, max_width=200, base_font="Helvetica", base_size=12, min_size=6, normal_limit=18):
    txt = str(text or "").strip()
    if not txt:
        return base_size

    size = float(base_size)
    
    if len(txt) > normal_limit:
        exceso = len(txt) - normal_limit
        size = max(base_size - (exceso * 0.3), min_size)

    c.setFont(base_font, size)

    while size > min_size and c.stringWidth(txt, base_font, size) > max_width:
        size -= 0.2
        c.setFont(base_font, size)

    c.drawString(x, y, txt)
    c.setFont("Helvetica", 12)
    return size


def preparar_datos_pdf(datos):
    ars_canon = normalizar_seguro(datos.get('Aseguradora (ARS)', ''), datos.get('NSS', ''))
    usar_guiones = bool(app_setting("pdf_nss_guiones", True))
    modo_ars = str(app_setting("pdf_ars_display_mode", "Abreviada"))
    ars_display = ars_canon if modo_ars.lower().startswith("completa") else seguro_para_mostrar(ars_canon)

    return {
        **datos,
        "Cédula": formatear_cedula(datos.get('Cédula', '')),
        "Teléfono": formatear_telefono(datos.get('Teléfono', '')),
        "NSS": formatear_nss_para_pdf(datos.get('NSS', ''), ars_canon) if usar_guiones else str(datos.get('NSS', '') or '').strip().upper(),
        "Edad": f"{datos.get('Edad_num', 0)}{datos.get('Unidad', 'Años')[0].upper()}",
        "ARS_CANONICO": ars_canon,
        "ARS_DISPLAY": ars_display,
    }


_PDF_TEMPLATE_BYTES_CACHE = {}
_PDF_TEMPLATE_CACHE_LOCK = threading.Lock()


def _read_pdf_template_bytes(ruta_hoja):
    """Return immutable template bytes without sharing PyPDF2 objects across workers."""
    stat = os.stat(ruta_hoja)
    fingerprint = (stat.st_mtime_ns, stat.st_size)
    with _PDF_TEMPLATE_CACHE_LOCK:
        cached = _PDF_TEMPLATE_BYTES_CACHE.get(ruta_hoja)
        if cached and cached[0] == fingerprint:
            return cached[1]
    with open(ruta_hoja, "rb") as template_stream:
        template_bytes = template_stream.read()
    with _PDF_TEMPLATE_CACHE_LOCK:
        _PDF_TEMPLATE_BYTES_CACHE[ruta_hoja] = (fingerprint, template_bytes)
    return template_bytes


def _append_overlay_to_cloned_page(writer, page, overlay_page, prefix):
    """Append a ReportLab overlay without rewriting the template image streams."""
    overlay_resources = overlay_page.get("/Resources", DictionaryObject()).get_object()
    overlay_fonts = overlay_resources.get("/Font", DictionaryObject()).get_object()
    page_resources = page.get("/Resources", DictionaryObject()).get_object()
    resources = DictionaryObject()
    resources.update(page_resources)
    fonts = DictionaryObject()
    existing_fonts = page_resources.get("/Font")
    if existing_fonts:
        fonts.update(existing_fonts.get_object())

    content = overlay_page.get_contents().get_data()
    for font_name, font_ref in overlay_fonts.items():
        replacement = NameObject(f"/{prefix}_{str(font_name).lstrip('/')}")
        fonts[replacement] = writer._add_object(font_ref.get_object())
        content = content.replace(
            str(font_name).encode("ascii"), str(replacement).encode("ascii")
        )
    resources[NameObject("/Font")] = fonts
    page[NameObject("/Resources")] = resources

    overlay_stream = DecodedStreamObject()
    overlay_stream.set_data(b"q\n" + content + b"\nQ\n")
    overlay_ref = writer._add_object(overlay_stream)
    existing_content = page.get("/Contents")
    contents = ArrayObject()
    if isinstance(existing_content, ArrayObject):
        contents.extend(existing_content)
    elif existing_content:
        contents.append(existing_content)
    contents.append(overlay_ref)
    page[NameObject("/Contents")] = contents


def crear_pdf_temporal(hoja, datos, mostrar_error=True):
    ruta_hoja = RUTA_HOJAS.get(hoja)
    writer = PdfWriter()
    if not ruta_hoja or not os.path.exists(ruta_hoja):
        APP_LOG.error("Plantilla no encontrada para %s: %s", hoja, ruta_hoja)
        if mostrar_error:
            messagebox.showerror("Error", f"Plantilla no encontrada: {ruta_hoja}")
        return None
    temp_overlay = None
    try:
        template_started = _time.perf_counter()
        reader = PdfReader(io.BytesIO(_read_pdf_template_bytes(ruta_hoja)))
        APP_LOG.info(
            "PDF_TEMPLATE_READ_MS sheet=%s elapsed_ms=%.1f",
            hoja,
            (_time.perf_counter() - template_started) * 1000.0,
        )
        fd_overlay, temp_overlay = tempfile.mkstemp(suffix=".pdf")
        os.close(fd_overlay)
        c = canvas.Canvas(temp_overlay, pagesize=letter)
        first_page = reader.pages[0]
        page_width = float(first_page.mediabox[2])
        page_height = float(first_page.mediabox[3])
        c.setPageSize((page_width, page_height))
        c.setFont("Helvetica", 12)
        c.setFillColorRGB(0, 0, 0)

        d = preparar_datos_pdf(datos)

        if hoja == "GENERAL":
            c.drawString(495, 680, f"{d['Fecha']}")
            c.drawString(275, 665, f"{d['Hora']}")
            draw_text_auto(c, 63, 579, f"{d['Nombre']}", base_size=12, max_len=32, min_size=8)
            c.drawString(53, 560, f"{d['Sexo']}")
            c.drawString(135, 560, f"{d['Edad']}")
            
            draw_text_auto(c, 83, 542, f"{d['ARS_DISPLAY']}", base_size=12, max_len=20, min_size=8)
            
            c.drawString(228, 542, f"{d['NSS']}")
            c.drawString(495, 560, f"{d['Cédula']}")
            c.drawString(340, 493, f"{d['Teléfono']}")
            draw_direccion_auto(c, 70, 510, f"{d['Dirección']}", max_width=450, base_size=12, min_size=7)
            draw_text_auto(c, 100, 495, f"{d['Nacionalidad']}", base_size=12, max_len=18, min_size=8)

        elif hoja == "GINECOLOGIA":
            c.drawString(182, 720, f"{d['Fecha']}")
            c.drawString(273, 720, f"{d['Hora']}")
            draw_text_auto(c, 128, 705, f"{d['Nombre']}", base_size=12, max_len=30, min_size=8)
            c.drawString(406, 707, f"{d['Edad']}")
            
            ars_visual = d['ARS_DISPLAY']
            if ars_visual == "S.CONTRIBUTIVO":
                ars_visual = "S.CONT"
            elif ars_visual == "S.PENSIONADOS":
                ars_visual = "S.PENSIONAD"
                
            draw_text_auto(c, 110, 692, f"{ars_visual}", base_size=9.5, max_len=20, min_size=6)
            
            c.drawString(330, 693, f"{d['Teléfono']}")
            
            c.drawString(177, 693, f"{d['NSS']}")
            
            c.drawString(120, 667, f"{d['Cédula']}")
            
            draw_direccion_auto(c, 133, 679, f"{d['Dirección']}", max_width=185, base_size=12, min_size=6, normal_limit=17)
            
            draw_text_auto(c, 330, 680, f"{d['Nacionalidad']}", base_size=12, max_len=16, min_size=8)

        elif hoja == "PEDIATRIA":
            draw_text_auto(c, 110, 639, f"{d['Nombre']}", base_size=12, max_len=30, min_size=8)
            c.drawString(305, 625, f"{d['Teléfono']}")
            c.drawString(418, 639, f"{d['Edad']}")
            
            c.drawString(165, 624, f"{d['NSS']}")
            
            ars_visual = d['ARS_DISPLAY']
            if ars_visual == "S.CONTRIBUTIVO":
                ars_visual = "S.CONT"
            elif ars_visual == "S.PENSIONADOS":
                ars_visual = "S.PENSIONAD"
                
            draw_text_auto(c, 95, 624, f"{ars_visual}", base_size=9.5, max_len=13, min_size=5.5)
            
            draw_direccion_auto(c, 121, 610, f"{d['Dirección']}", max_width=185, base_size=12, min_size=6, normal_limit=17)
            
            draw_text_auto(c, 320, 610, f"{d['Nacionalidad']}", base_size=12, max_len=14, min_size=8)
            c.drawString(187, 657, f"{d['Fecha']}")
            c.drawString(297, 657, f"{d['Hora']}")
            c.drawString(469, 637, f"{d['Sexo']}")

        overlay_started = _time.perf_counter()
        c.save()

        overlay_pages = [PdfReader(temp_overlay).pages[0]]
        temp_overlay2 = None
        
        if hoja == "GENERAL" and len(reader.pages) >= 2:
            try:
                fd_overlay2, temp_overlay2 = tempfile.mkstemp(suffix=".pdf")
                os.close(fd_overlay2)
                c2 = canvas.Canvas(temp_overlay2, pagesize=letter)
                c2.setPageSize((page_width, page_height))
                c2.setFont("Helvetica", 10)
                c2.setFillColorRGB(0, 0, 0)
                
                campos_examen = [
                    {"nombre": "Cabeza", "x_check": 499, "y_check": 609, "x_texto": 510, "y_texto": 610},
                    {"nombre": "Cuello", "x_check": 499, "y_check": 584, "x_texto": 510, "y_texto": 585},
                    {"nombre": "Corazón", "x_check": 499, "y_check": 557, "x_texto": 510, "y_texto": 558},
                    {"nombre": "Tórax", "x_check": 499, "y_check": 529, "x_texto": 510, "y_texto": 530},
                    {"nombre": "Abdomen", "x_check": 499, "y_check": 500, "x_texto": 510, "y_texto": 501},
                    {"nombre": "Genitales", "x_check": 499, "y_check": 469, "x_texto": 510, "y_texto": 470},
                    {"nombre": "Pulmones", "x_check": 499, "y_check": 441, "x_texto": 510, "y_texto": 442},
                    {"nombre": "Extremidades", "x_check": 499, "y_check": 415, "x_texto": 510, "y_texto": 416},
                    {"nombre": "Ex. Neurológico", "x_check": 499, "y_check": 389, "x_texto": 510, "y_texto": 390},
                    {"nombre": "Tacto Rectal", "x_check": 499, "y_check": 363, "x_texto": 510, "y_texto": 364},
                    {"nombre": "Tacto Vaginal", "x_check": 499, "y_check": 337, "x_texto": 510, "y_texto": 338},
                ]
                
                checkbox_size = 9
                
                for campo in campos_examen:
                    c2.rect(campo["x_check"], campo["y_check"] - 2, checkbox_size, checkbox_size, stroke=1, fill=0)
                    c2.drawString(campo["x_texto"], campo["y_texto"] - 2, "Sin patología aparente")
                
                c2.save()
                overlay_pages.append(PdfReader(temp_overlay2).pages[0])
            except Exception:
                pass
            finally:
                if temp_overlay2 and os.path.exists(temp_overlay2):
                    try:
                        os.remove(temp_overlay2)
                    except Exception:
                        pass

        APP_LOG.info(
            "PDF_OVERLAY_MS sheet=%s elapsed_ms=%.1f",
            hoja,
            (_time.perf_counter() - overlay_started) * 1000.0,
        )
        merge_started = _time.perf_counter()
        # Clone the immutable template once.  The overlay streams are appended
        # directly so the image-heavy template streams remain untouched.
        # Each worker receives a fresh reader, so no PyPDF2 objects are shared.
        writer.clone_document_from_reader(reader)
        for idx, overlay_page in enumerate(overlay_pages):
            _append_overlay_to_cloned_page(
                writer, writer.pages[idx], overlay_page, f"ADMISSION_OVERLAY_{idx}"
            )

        fd_final, temp_final = tempfile.mkstemp(
            suffix=f"_{sanitize_filename(d.get('Nombre', 'PACIENTE'))}_{hoja}.pdf"
        )
        os.close(fd_final)
        with open(temp_final, "wb") as out:
            writer.write(out)
        APP_LOG.info(
            "PDF_MERGE_WRITE_MS sheet=%s elapsed_ms=%.1f",
            hoja,
            (_time.perf_counter() - merge_started) * 1000.0,
        )

        try:
            os.remove(temp_overlay)
        except Exception:
            pass

        return temp_final

    except Exception as e:
        APP_LOG.exception("Error generando PDF temporal")
        if mostrar_error:
            messagebox.showerror("Error", f"Error generando PDF temporal: {str(e)}")
        if temp_overlay and os.path.exists(temp_overlay):
            try:
                os.remove(temp_overlay)
            except Exception:
                pass
        return None


def archivar_pdf_atencion(ruta_temporal, atencion_id, fecha=None):
    """Compatibilidad: se deshabilitó el archivo permanente de hojas individuales."""
    APP_LOG.info("Se omitió el archivo permanente de la hoja de atención #%s", atencion_id)
    return None


def regenerar_pdf_archivado(db: DatabaseManager, atencion_id: int, mostrar_error=False):
    """Compatibilidad: las atenciones ya no generan ni archivan PDF individual."""
    return None


def programar_limpieza_pdf_temporal(ruta, espera_segundos=900, reintentos=8):
    """Retira una hoja temporal cuando el visor o la cola de impresión la haya liberado."""
    ruta = os.path.abspath(str(ruta or ""))
    if not ruta:
        return

    def _limpiar(intentos_restantes):
        try:
            if os.path.isfile(ruta):
                os.remove(ruta)
            return
        except PermissionError:
            if intentos_restantes > 0:
                threading.Timer(60, _limpiar, args=(intentos_restantes - 1,)).start()
                return
        except Exception:
            pass
        APP_LOG.warning("No se pudo retirar todavía una hoja temporal de atención")

    temporizador = threading.Timer(max(1, int(espera_segundos)), _limpiar, args=(int(reintentos),))
    temporizador.daemon = True
    temporizador.start()


def eliminar_archivo_sensible(ruta):
    """Sobrescribe y retira un archivo regular; falla de forma explícita si no puede hacerlo."""
    ruta = os.path.abspath(ruta)
    if not os.path.isfile(ruta):
        return
    size = os.path.getsize(ruta)
    with open(ruta, "r+b", buffering=0) as stream:
        bloque = b"\0" * (1024 * 1024)
        restante = size
        while restante:
            chunk = bloque if restante >= len(bloque) else bloque[:restante]
            stream.write(chunk)
            restante -= len(chunk)
        stream.flush()
        os.fsync(stream.fileno())
    os.remove(ruta)


def crear_selector_fecha(parent, width=16):
    widget = TBDateEntry(
        parent,
        bootstyle="primary",
        dateformat="%d/%m/%Y",
        firstweekday=0,
        width=width,
        startdate=datetime.now()
    )
    return widget


def obtener_fecha_selector(widget):
    try:
        return widget.entry.get().strip()
    except Exception:
        try:
            return widget.get().strip()
        except Exception:
            return ""


def establecer_fecha_selector(widget, fecha_obj: date):
    texto = fecha_obj.strftime("%d/%m/%Y")
    try:
        widget.entry.delete(0, tk.END)
        widget.entry.insert(0, texto)
        return
    except Exception:
        pass

    try:
        widget.delete(0, tk.END)
        widget.insert(0, texto)
    except Exception:
        pass


def construir_resumen_desde_registros(registros, periodo_texto, turno_resumen=None, representante=""):
    registros = list(registros or [])
    conteo_seguro = {}
    conteo_esp = {}
    urgencias = 0
    consultas = 0
    total_emergencia = 0

    for r in registros:
        tipo = (r.get("tipo_atencion") or "EMERGENCIA").strip().upper()
        if tipo == "URGENCIA":
            urgencias += 1
            continue
        if tipo == "CONSULTA":
            consultas += 1
            continue
        if tipo != "EMERGENCIA":
            continue

        total_emergencia += 1
        seguro = r.get("ars_display", "SIN SEGURO")
        especialidad = r.get("hoja_normalizada", "SIN ESPECIALIDAD")
        conteo_seguro[seguro] = conteo_seguro.get(seguro, 0) + 1
        conteo_esp[especialidad] = conteo_esp.get(especialidad, 0) + 1

    cantidad_sin_seguro = conteo_seguro.get("SIN SEGURO", 0)

    por_seguro = sorted(conteo_seguro.items(), key=lambda x: (-x[1], x[0]))
    por_especialidad = sorted(conteo_esp.items(), key=lambda x: (-x[1], x[0]))

    return {
        "periodo_texto": poner_hora_entre_parentesis(periodo_texto),
        "total_general": total_emergencia,
        "cantidad_sin_seguro": cantidad_sin_seguro,
        "cantidad_urgencias": urgencias,
        "cantidad_consultas": consultas,
        "por_seguro": por_seguro,
        "por_especialidad": por_especialidad,
        "turno_resumen": turno_resumen,
        "representante": (representante or "").strip(),
        "registros": registros,
    }


class EmptyAdmissionReportError(ValueError):
    """El conjunto exacto que alimenta el reporte no contiene pacientes."""


def reportable_patient_count(resumen):
    data = dict(resumen or {})
    if "total_patients" in data:
        return max(0, int(data.get("total_patients") or 0))
    return max(
        0,
        int(data.get("total_general") or 0)
        + int(data.get("cantidad_urgencias") or 0)
        + int(data.get("cantidad_consultas") or 0),
    )

def construir_resumen_desde_excel_actual(turno_cfg: dict, periodo_texto: str):
    registros = []
    try:
        if not os.path.exists(EXCEL_PATH) or not excel_tiene_registros(EXCEL_PATH):
            return None

        wb = abrir_excel_workbook_seguro(EXCEL_PATH, read_only=True, data_only=True)
        ws = wb.active

        for fila in range(6, ws.max_row + 1):
            nombre = str(ws.cell(row=fila, column=2).value or "").strip()
            hoja = str(ws.cell(row=fila, column=3).value or "").strip().upper()
            ars = str(ws.cell(row=fila, column=4).value or "").strip()

            if not nombre:
                continue

            ars_canon = normalizar_seguro(ars, "999999999")
            registros.append({
                "nombre": nombre,
                "hoja": hoja,
                "hoja_normalizada": hoja or "SIN ESPECIALIDAD",
                "ars": ars_canon,
                "ars_display": seguro_para_mostrar(ars_canon),
                "nss": "",
                "cedula": "",
                "fecha": "",
                "hora": "",
            })

        try:
            wb.close()
        except Exception:
            pass

        if not registros:
            return None

        datos_turno = obtener_datos_turno_visual(turno_cfg["fecha_base"], turno_cfg["turno_codigo"])
        return construir_resumen_desde_registros(
            registros,
            periodo_texto,
            turno_resumen=datos_turno["turno_resumen"],
            representante=turno_cfg.get("representante", "")
        )
    except Exception:
        return None


def construir_resumen_turno(
    db: DatabaseManager,
    turno_cfg: dict,
    fin_override: datetime = None,
    *,
    turn_id=None,
    operational_source_id=None,
):
    if not turno_cfg:
        return None

    inicio, fin = obtener_rango_turno_efectivo(turno_cfg, fin_override=fin_override)
    central_dataset = _dataset_turno_central(
        db,
        turn_id=turn_id,
        operational_source_id=operational_source_id,
    )
    if central_dataset is not None:
        registros, effective_turn_id, effective_source_id = central_dataset
        APP_LOG.info(
            "SHIFT_REPORT_DATASET turn_id=%s source=%s count=%s",
            effective_turn_id,
            effective_source_id,
            len(registros),
        )
    else:
        contexto = db.buscar_contexto_turno_existente(turno_cfg)
        registros = db.obtener_atenciones_para_rango_real(
            inicio,
            fin,
            turno_id=int(contexto["turno_id"]) if contexto else None,
        )
    datos_turno = obtener_datos_turno_visual(turno_cfg["fecha_base"], turno_cfg["turno_codigo"])
    periodo_texto = f"{inicio.strftime('%d/%m/%Y %I:%M %p')} a {fin.strftime('%d/%m/%Y %I:%M %p')}"

    return construir_resumen_desde_registros(
        registros,
        periodo_texto,
        turno_resumen=datos_turno["turno_resumen"],
        representante=turno_cfg.get("representante", "")
    )


def crear_pdf_reporte(resumen, destino=None):
    if reportable_patient_count(resumen) == 0:
        raise EmptyAdmissionReportError(
            "No hay pacientes que coincidan con los criterios seleccionados."
        )
    os.makedirs(REPORTES_DIR, exist_ok=True)

    if destino:
        pdf_path = destino
    else:
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        pdf_path = output_report_path(f"Reporte_Pacientes_{ts}.pdf")

    orientacion_pdf = str(app_setting("print_pdf_orientation", "Horizontal")).lower()
    pagesize_reporte = (letter[1], letter[0]) if orientacion_pdf.startswith("h") else letter
    c = canvas.Canvas(pdf_path, pagesize=pagesize_reporte)
    width, height = pagesize_reporte
    margen_x = 40
    margen_y = 42
    usable_w = width - (margen_x * 2)
    titulo_institucional = "REPORTE ESTADÍSTICO DE EMERGENCIA"
    subtitulo = "HOSPITAL PROVINCIAL DR. ÁNGEL CONTRERAS"
    representante_mayus = (resumen.get("representante") or "").strip().upper()

    def draw_center_text(txt, y, font="Helvetica-Bold", size=13):
        c.setFont(font, size)
        c.drawCentredString(width / 2, y, txt)

    def draw_table_header(y, headers, widths):
        h = 22
        x = margen_x
        c.setFillColorRGB(0.89, 0.93, 0.98)
        c.rect(x, y - h, sum(widths), h, stroke=1, fill=1)
        c.setFillColorRGB(0, 0, 0)
        c.setFont("Helvetica-Bold", 10)
        cx = x
        for idx, header in enumerate(headers):
            c.rect(cx, y - h, widths[idx], h, stroke=1, fill=0)
            c.drawCentredString(cx + widths[idx] / 2, y - 15, str(header))
            cx += widths[idx]
        return y - h

    def draw_table_row(y, values, widths, bold=False):
        h = 20
        x = margen_x
        c.setFont("Helvetica-Bold" if bold else "Helvetica", 9.5)
        cx = x
        for idx, value in enumerate(values):
            c.rect(cx, y - h, widths[idx], h, stroke=1, fill=0)
            text = str(value)
            if idx == len(values) - 1:
                c.drawCentredString(cx + widths[idx] / 2, y - 14, text)
            else:
                size = 9.5
                while size > 7 and c.stringWidth(text, "Helvetica-Bold" if bold else "Helvetica", size) > widths[idx] - 8:
                    size -= 0.5
                c.setFont("Helvetica-Bold" if bold else "Helvetica", size)
                c.drawString(cx + 5, y - 14, text)
                c.setFont("Helvetica-Bold" if bold else "Helvetica", 9.5)
            cx += widths[idx]
        return y - h

    def draw_section_title(txt, y):
        c.setFillColorRGB(0.08, 0.18, 0.30)
        c.rect(margen_x, y - 21, usable_w, 21, stroke=0, fill=1)
        c.setFillColorRGB(1, 1, 1)
        c.setFont("Helvetica-Bold", 11)
        c.drawString(margen_x + 8, y - 15, txt)
        c.setFillColorRGB(0, 0, 0)
        return y - 28

    def draw_summary_box(y):
        box_h = 48
        gap = 8
        box_w = (usable_w - gap * 3) / 4
        total_pacientes = int(
            resumen.get("total_patients", reportable_patient_count(resumen)) or 0
        )
        sin_seguro = int(
            resumen.get("uninsured_patients", resumen.get("cantidad_sin_seguro", 0)) or 0
        )
        items = [
            ("TOTAL PACIENTES", total_pacientes),
            ("ASEGURADOS", resumen.get("insured_patients", total_pacientes - sin_seguro)),
            ("SIN SEGURO", sin_seguro),
            ("MEDICINA GENERAL", resumen.get("general_patients", 0)),
        ]
        x = margen_x
        for title, value in items:
            c.rect(x, y - box_h, box_w, box_h, stroke=1, fill=0)
            c.setFont("Helvetica-Bold", 9)
            c.drawCentredString(x + box_w / 2, y - 15, str(title))
            c.setFont("Helvetica-Bold", 13 if isinstance(value, int) else 9.5)
            c.drawCentredString(x + box_w / 2, y - 35, str(value))
            x += box_w + gap
        return y - box_h - 18

    def dibujar_encabezado(page_no=1):
        y = height - 38
        if os.path.exists(LOGO_PATH):
            try:
                logo_w = usable_w + 20
                logo_h = 82
                x_logo = margen_x - 10
                c.drawImage(LOGO_PATH, x_logo, height - logo_h - 6,
                            width=logo_w, height=logo_h, preserveAspectRatio=True, anchor='c', mask='auto')
                y = height - logo_h - 22
            except Exception:
                pass

        draw_center_text(titulo_institucional, y, size=14)
        y -= 16
        draw_center_text(subtitulo, y, font="Helvetica-Bold", size=11)
        y -= 18

        c.setFont("Helvetica", 9.5)
        c.drawString(margen_x, y, f"Período: {poner_hora_entre_parentesis(resumen.get('periodo_texto', ''))}")
        y -= 13
        if resumen.get("turno_resumen"):
            c.drawString(margen_x, y, f"Turno: {resumen.get('turno_resumen')}")
            y -= 13
        c.drawString(
            margen_x,
            y,
            "Filtros: "
            f"ARS {resumen.get('ars_mode', ARS_ALL)} · "
            f"Cobertura {resumen.get('coverage_filter', COVERAGE_ALL)} · "
            f"Especialidad {resumen.get('specialty_filter', SPECIALTY_ALL)}",
        )
        y -= 13
        if resumen.get("selected_ars_label") and resumen.get("ars_mode") != ARS_ALL:
            c.drawString(
                margen_x,
                y,
                f"ARS seleccionadas: {resumen.get('selected_ars_label')}",
            )
            y -= 13
        if representante_mayus:
            c.drawString(margen_x, y, f"Representante(s): {representante_mayus}")
            y -= 13
        c.drawRightString(width - margen_x, height - 30, f"Página {page_no}")
        c.line(margen_x, y - 4, width - margen_x, y - 4)
        return y - 20

    page_no = 1
    y = dibujar_encabezado(page_no)
    y = draw_summary_box(y)

    representatives_by_turn = resumen.get("representatives_by_turn") or ()
    if representatives_by_turn:
        y = draw_section_title("REPRESENTANTES POR TURNO", y)
        representative_widths = [usable_w * 0.48, usable_w * 0.52]
        y = draw_table_header(y, ["Turno", "Representante(s)"], representative_widths)
        for turn_label, representatives in representatives_by_turn:
            if y < margen_y + 60:
                c.showPage()
                page_no += 1
                y = dibujar_encabezado(page_no)
                y = draw_section_title("REPRESENTANTES POR TURNO", y)
                y = draw_table_header(
                    y, ["Turno", "Representante(s)"], representative_widths
                )
            y = draw_table_row(
                y,
                [turn_label, ", ".join(representatives)],
                representative_widths,
            )
        y -= 14

    por_seguro = resumen.get("por_seguro", []) or []
    sin_seguro_rows = [(s, n) for s, n in por_seguro if str(s).upper() == "SIN SEGURO"]
    asegurados_rows = [(s, n) for s, n in por_seguro if str(s).upper() != "SIN SEGURO"]

    y = draw_section_title("PACIENTES ASEGURADOS POR ARS", y)
    widths = [usable_w * 0.76, usable_w * 0.24]
    y = draw_table_header(y, ["ARS / Seguro", "Cantidad"], widths)
    if asegurados_rows:
        for seguro, cantidad in asegurados_rows:
            if y < margen_y + 80:
                c.showPage(); page_no += 1; y = dibujar_encabezado(page_no)
                y = draw_section_title("PACIENTES ASEGURADOS POR ARS", y)
                y = draw_table_header(y, ["ARS / Seguro", "Cantidad"], widths)
            y = draw_table_row(y, [seguro, cantidad], widths)
    else:
        y = draw_table_row(y, ["Sin registros de asegurados", 0], widths)

    y -= 14
    if y < margen_y + 100:
        c.showPage(); page_no += 1; y = dibujar_encabezado(page_no)

    y = draw_section_title("PACIENTES SIN SEGURO", y)
    y = draw_table_header(y, ["Categoría", "Cantidad"], widths)
    cantidad_ss = sin_seguro_rows[0][1] if sin_seguro_rows else resumen.get("cantidad_sin_seguro", 0)
    y = draw_table_row(y, ["SIN SEGURO", cantidad_ss], widths, bold=True)

    y -= 14
    if y < margen_y + 120:
        c.showPage(); page_no += 1; y = dibujar_encabezado(page_no)

    y = draw_section_title("PACIENTES POR ESPECIALIDAD", y)
    y = draw_table_header(y, ["Especialidad", "Cantidad"], widths)
    por_especialidad = resumen.get("por_especialidad", []) or []
    if por_especialidad:
        for esp, cantidad in por_especialidad:
            if y < margen_y + 60:
                c.showPage(); page_no += 1; y = dibujar_encabezado(page_no)
                y = draw_section_title("PACIENTES POR ESPECIALIDAD", y)
                y = draw_table_header(y, ["Especialidad", "Cantidad"], widths)
            y = draw_table_row(y, [esp, cantidad], widths)
    else:
        y = draw_table_row(y, ["Sin registros de especialidad", 0], widths)

    if y < margen_y + 60:
        c.showPage(); page_no += 1; y = dibujar_encabezado(page_no)
    y -= 20
    c.setFont("Helvetica-Bold", 12)
    c.drawString(
        margen_x,
        y,
        f"Total de pacientes del período: {reportable_patient_count(resumen)}",
    )
    if representante_mayus:
        c.drawRightString(width - margen_x, y, f"REPRESENTANTE(S): {representante_mayus}")

    c.save()
    return pdf_path


def crear_excel_reporte_estadistico(resumen, destino=None):
    if reportable_patient_count(resumen) == 0:
        raise EmptyAdmissionReportError(
            "No hay pacientes que coincidan con los criterios seleccionados."
        )
    os.makedirs(REPORTES_DIR, exist_ok=True)
    if destino:
        xlsx_path = destino
    else:
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        xlsx_path = output_report_path(f"Reporte_Estadistico_{ts}.xlsx")
    registros = list(resumen.get("records") or resumen.get("registros") or [])
    if len(registros) != reportable_patient_count(resumen):
        raise RuntimeError(
            "El total del resumen no coincide con el dataset del listado."
        )
    wb = Workbook()
    listado = wb.active
    listado.title = "LISTADO DE PACIENTES"
    construir_hoja_listado_pacientes(
        listado,
        registros,
        encabezado_linea_3=resumen.get("period_label") or resumen.get("periodo_texto") or "",
        encabezado_linea_4=resumen.get("turn_label") or resumen.get("turno_resumen") or "",
    )

    resumen_ws = wb.create_sheet("RESUMEN ESTADÍSTICO")
    total = int(resumen.get("total_patients", reportable_patient_count(resumen)) or 0)
    sin_seguro = int(
        resumen.get("uninsured_patients", resumen.get("cantidad_sin_seguro", 0)) or 0
    )
    asegurados = int(resumen.get("insured_patients", total - sin_seguro) or 0)
    porcentaje_asegurados = float(
        resumen.get("insured_percentage", (asegurados * 100 / total) if total else 0)
        or 0
    )
    porcentaje_sin_seguro = float(
        resumen.get("uninsured_percentage", (sin_seguro * 100 / total) if total else 0)
        or 0
    )
    generated_at = resumen.get("generated_at") or datetime.now()
    if isinstance(generated_at, datetime):
        generated_label = generated_at.strftime("%d/%m/%Y %I:%M %p")
    else:
        generated_label = str(generated_at)
    summary_rows = [
        ("REPORTE ESTADÍSTICO", ""),
        ("Total de pacientes", total),
        ("Total asegurados", asegurados),
        ("Total sin seguro", sin_seguro),
        ("Porcentaje asegurados", porcentaje_asegurados / 100),
        ("Porcentaje sin seguro", porcentaje_sin_seguro / 100),
        ("Medicina General", int(resumen.get("general_patients", 0) or 0)),
        ("Pediatría", int(resumen.get("pediatric_patients", 0) or 0)),
        ("Ginecología", int(resumen.get("gynecology_patients", 0) or 0)),
        ("Turno seleccionado", resumen.get("turn_label") or resumen.get("turno_resumen") or "Todos los turnos"),
        ("Período utilizado", resumen.get("period_label") or resumen.get("periodo_texto") or ""),
        ("Fecha desde", _report_excel_datetime(resumen.get("start_at"))),
        ("Fecha hasta", _report_excel_datetime(resumen.get("end_at"))),
        ("Modo ARS", resumen.get("ars_mode") or ARS_ALL),
        ("ARS seleccionadas", resumen.get("selected_ars_label") or "Ninguna"),
        ("Cobertura seleccionada", resumen.get("coverage_filter") or COVERAGE_ALL),
        ("Especialidad seleccionada", resumen.get("specialty_filter") or SPECIALTY_ALL),
        ("Representante(s)", resumen.get("representante") or "No disponible"),
        ("Fecha/hora de generación", generated_label),
    ]
    for label, value in summary_rows:
        resumen_ws.append([label, value])
    resumen_ws.append([])
    resumen_ws.append(["CONTEO POR ESPECIALIDAD", "Cantidad"])
    for especialidad, cantidad in (
        resumen.get("by_specialty") or resumen.get("por_especialidad") or []
    ):
        resumen_ws.append([especialidad, int(cantidad)])
    resumen_ws.append([])
    resumen_ws.append(["CONTEO POR ARS", "Cantidad"])
    for ars, cantidad in resumen.get("by_ars") or resumen.get("por_seguro") or []:
        resumen_ws.append([ars, int(cantidad)])
    representatives_by_turn = resumen.get("representatives_by_turn") or ()
    if representatives_by_turn:
        resumen_ws.append([])
        resumen_ws.append(["REPRESENTANTES POR TURNO", "Representante(s)"])
        for turn_label, representatives in representatives_by_turn:
            resumen_ws.append([turn_label, ", ".join(representatives)])
    _aplicar_formato_resumen_estadistico(resumen_ws)
    if resumen_ws["B2"].value != len(registros):
        wb.close()
        raise RuntimeError(
            "El resumen estadístico y el listado no contienen el mismo universo."
        )
    if not guardar_excel_seguro(
        wb, xlsx_path, "exportar el reporte estadístico a Excel"
    ):
        raise RuntimeError("No se pudo guardar el reporte estadístico en Excel.")
    return xlsx_path


def _report_excel_datetime(value):
    if isinstance(value, datetime):
        return value.strftime("%d/%m/%Y %I:%M %p")
    return str(value or "")


def _aplicar_formato_resumen_estadistico(ws):
    thin = XLSide(style="thin", color="D9E2EC")
    border = XLBorder(left=thin, right=thin, top=thin, bottom=thin)
    ws.column_dimensions["A"].width = 34
    ws.column_dimensions["B"].width = 42
    ws.freeze_panes = "A2"
    for row in ws.iter_rows():
        for cell in row[:2]:
            cell.font = XLFont(name="Calibri", size=11)
            cell.alignment = XLAlignment(vertical="center", wrap_text=True)
            if any(value is not None for value in (row[0].value, row[1].value)):
                cell.border = border
    section_labels = {
        "REPORTE ESTADÍSTICO",
        "CONTEO POR ESPECIALIDAD",
        "CONTEO POR ARS",
        "REPRESENTANTES POR TURNO",
    }
    for row_number in range(1, ws.max_row + 1):
        if ws.cell(row_number, 1).value not in section_labels:
            continue
        for cell in ws[row_number][:2]:
            cell.font = XLFont(name="Calibri", size=12, bold=True, color="FFFFFF")
            cell.fill = openpyxl.styles.PatternFill("solid", fgColor="1F4E78")
    for row in range(2, min(7, ws.max_row) + 1):
        ws.cell(row, 2).number_format = "0.00%" if row in (5, 6) else "0"


# -------------------------------
# APP
# -------------------------------
class App:
    def __init__(self, *, standalone=False, root=None, context=None):
        self._standalone = bool(standalone)
        if self._standalone and root is not None:
            raise ValueError("El modo standalone no acepta una raíz embebida.")
        if context is None:
            if not self._standalone:
                raise ValueError("La instancia embebida de Admisión requiere AdmissionContext.")
            context = create_standalone_context(
                session_context=load_session_context(),
                main_app_gateway=MainAppGateway.from_environment(),
                admission_database_factory=lambda session: DatabaseManager(
                    session_context=session,
                    event_bus=None,
                ),
                logger=APP_LOG,
            )
        if not isinstance(context, AdmissionContext):
            raise TypeError("context debe ser una instancia de AdmissionContext.")
        self.context = context
        self._shutdown_complete = False
        self._admission_visible = True
        self._worker_threads = set()
        self._worker_lock = threading.Lock()
        self._date_after_id = None
        self._summary_after_id = None
        self._session_after_id = None
        self._excel_export_after_id = None
        self._excel_export_running = False
        self.session_context = context.session_context
        self.main_app_gateway = context.main_app_gateway
        self.connection_factory = context.connection_factory
        self.user_id = context.user_id
        self.device_id = context.device_id
        self.device_name = context.device_name
        self.shared_configuration = context.configuration
        self.current_shift_context = context.current_shift
        self.shared_logger = context.logger
        self.event_bus = context.event_bus
        try:
            migrate_legacy_files((
                "pacientes.db",
                "LISTADO DE PACIENTES EN EMERGENCIA.xlsx",
                "turnos_config.json",
                "app_settings.json",
                "ars_catalogo.json",
                "nss_formatos_ars.json",
                "representantes.json",
                "security.json",
            ))
        except Exception:
            APP_LOG.exception("No se pudo completar la preparacion del directorio de datos")
        self.db = context.create_admission_database()
        try:
            self.db.backup_manager.ensure_daily()
        except Exception:
            APP_LOG.exception("No se pudo crear el respaldo diario verificado")
        self.security = AdminSecurity(
            SECURITY_CONFIG_PATH,
            os.path.join(LOGS_DIR, "security_audit.jsonl"),
        )
        self._admin_authorized_until = None
        self._admin_authorized_actor = ""
        self.app_settings = cargar_app_settings()
        self._host_theme_controlled = bool(
            not self._standalone and getattr(context, "embedded", False)
        )
        configured_host_theme = self.shared_configuration.get("host_theme_is_dark")
        self._host_theme_is_dark = (
            bool(configured_host_theme)
            if self._host_theme_controlled and configured_host_theme is not None
            else None
        )
        configured_theme = self.shared_configuration.get("host_visual_theme")
        self._host_visual_theme = (
            dict(configured_theme)
            if self._host_theme_controlled and isinstance(configured_theme, dict)
            else None
        )
        self._embedded_visual_notifier = None
        self._embedded_theme_indicators = []
        self._embedded_entry_height_floor = {}
        self._entry_geometry_sync_scheduled = False
        self._responsive_layout_profile = None
        self._host_layout_snapshot = None
        self._last_responsive_log_signature = None
        self._asegurar_preferencias_impresion_hoja()
        verificar_o_crear_excel()
        self._excel_pendiente_turno_manual = excel_requiere_turno_manual()

        self._temp_files = set()
        self._updating_period_dates = False
        self._undo_stack = []
        self._undo_limit = 20
        self._turn_change_in_progress = False
        self._turn_change_committing = False
        self._primary_transfer_in_progress = False

        # FASE 3: Cache de rendimiento
        self._cache_ars = []
        self._cache_ars_time = 0
        self._cache_resumen_turno = None
        self._cache_resumen_time = 0
        self._cache_especialidades = ["GENERAL", "PEDIATRIA", "GINECOLOGIA"]
        self._sumatra_path_cache = None

        self.root = root or tb.Window(
            themename="superhero",
            owns_application_loop=self._standalone,
        )
        # qt_compat invokes this GUI-only hook immediately before a Toplevel
        # becomes visible.  It prevents a dark construction palette from
        # flashing in a dialog opened while the host is in light mode.
        self.root._admission_theme_applier = self._apply_theme_to_new_window
        if self._standalone:
            self.root.title("Generador de Formularios de Emergencia - Hospital General")
            self.root.geometry(self.app_settings.get("window_size", "1280x740"))
            self.root.minsize(1220, 700)
            self.root.resizable(True, True)
        self.root.configure(bg="#07111f")
        self.root.protocol("WM_DELETE_WINDOW", self.on_close)

        style = tb.Style()
        style.configure("TLabel", font=("Arial", 11), foreground="#EAF2FF")
        style.configure("Muted.TLabel", font=("Arial", 10), foreground="#A9B8CC")
        style.configure("Title.TLabel", font=("Arial", 18, "bold"), foreground="#F5F9FF")
        style.configure("Subtitle.TLabel", font=("Arial", 12), foreground="#B7C6DA")
        style.configure("Section.TLabel", font=("Arial", 13, "bold"), foreground="#5CB6FF")
        style.configure("TButton", font=("Arial", 11, "bold"))
        style.configure("TEntry", font=("Arial", 11), fieldbackground="#111E2E", foreground="#F5F9FF")
        style.configure("TCombobox", font=("Arial", 11), fieldbackground="#111E2E", foreground="#F5F9FF")
        style.configure("Card.TFrame", background="#0E1B2B")
        style.configure("Root.TFrame", background="#07111f")
        try:
            style.configure("Treeview", font=("Arial", 10), rowheight=28)
            style.configure("Treeview.Heading", font=("Arial", 10, "bold"))
            style.configure(
                "Modern.Treeview",
                font=("Arial", 10),
                rowheight=29,
                background="#0B1624",
                foreground="#EAF2FF",
                fieldbackground="#0B1624",
                bordercolor="#203348",
                lightcolor="#203348",
                darkcolor="#203348"
            )
            style.configure(
                "Modern.Treeview.Heading",
                font=("Arial", 10, "bold"),
                background="#12243A",
                foreground="#FFFFFF",
                bordercolor="#203348"
            )
            style.map(
                "Modern.Treeview",
                background=[("selected", "#1D6EFF")],
                foreground=[("selected", "#FFFFFF")]
            )
        except Exception:
            pass

        self.style = style
        self._configurar_estilos_desde_preferencias()

        self.val_cedula = (self.root.register(lambda P, *_: self.validar_numerico(P, 'cedula')), '%P')
        self.val_telefono = (self.root.register(lambda P, *_: self.validar_numerico(P, 'telefono')), '%P')

        self.main = tb.Frame(self.root, padding=14, style="Root.TFrame")
        self.main.pack(fill="both", expand=True)

        header = tb.Frame(self.main, style="Root.TFrame")
        header.pack(fill="x", pady=(0, 14))
        header.columnconfigure(1, weight=2, minsize=520)
        self.header = header

        logo_box = tb.Frame(header, padding=10, style="Card.TFrame")
        logo_box.grid(row=0, column=0, rowspan=2, sticky="nw", padx=(0, 14))
        try:
            logo_original = tk.PhotoImage(file=resource_path("istipo_hospitales.png"))
            factor = max(1, int(max(logo_original.width(), logo_original.height()) / 48))
            self.header_logo_image = logo_original.subsample(factor, factor)
            tb.Label(logo_box, image=self.header_logo_image, background="#0E1B2B").pack()
        except Exception:
            tb.Label(
                logo_box,
                text="HG",
                font=("Arial", 18, "bold"),
                foreground="#58A6FF",
                background="#0E1B2B",
            ).pack()

        title_lbl = tb.Label(
            header,
            text="GENERADOR DE FORMULARIOS DE\nEMERGENCIA",
            style="Title.TLabel",
            background="#07111f",
            wraplength=560,
            justify="left"
        )
        title_lbl.grid(row=0, column=1, sticky="w")
        self.title_lbl = title_lbl
        self.subtitle_lbl = tb.Label(header, text="Sistema de Admisión en Emergencia", style="Subtitle.TLabel", background="#07111f")
        self.subtitle_lbl.grid(row=1, column=1, sticky="w")

        info_header = tb.Frame(header, style="Root.TFrame")
        info_header.grid(row=0, column=2, rowspan=2, sticky="e")
        self.info_header = info_header

        fecha_card = tb.Frame(info_header, padding=(14, 10), style="Card.TFrame")
        fecha_card.pack(side="left", padx=6)
        tb.Label(fecha_card, text="▣  Fecha actual", style="Muted.TLabel", background="#0E1B2B")\
            .pack(anchor="w")
        self.fecha_actual_var = tk.StringVar(value=datetime.now().strftime("%d/%m/%Y"))
        tb.Label(fecha_card, textvariable=self.fecha_actual_var, font=("Arial", 11, "bold"), foreground="#FFFFFF", background="#0E1B2B")\
            .pack(anchor="w")

        central_cfg = self._turno_config_desde_snapshot_operacional()
        central_available = bool(
            str((self.current_shift_context or {}).get("operational_session_id") or "").strip()
        )
        turno_cfg = central_cfg if central_available else cargar_turno_config()
        turno_txt = (
            self._descripcion_turno_snapshot_operacional()
            if central_available
            else descripcion_turno_config(turno_cfg)
        )
        self.turno_header_var = tk.StringVar(value=turno_txt)
        turno_card = tb.Frame(info_header, padding=(14, 10), style="Card.TFrame")
        turno_card.pack(side="left", padx=6)
        tb.Label(turno_card, text="◷  Turno actual", style="Muted.TLabel", background="#0E1B2B")\
            .pack(anchor="w")
        tb.Label(turno_card, textvariable=self.turno_header_var, font=("Arial", 11, "bold"), foreground="#FFFFFF", background="#0E1B2B")\
            .pack(anchor="w")

        user_card = tb.Frame(info_header, padding=(14, 10), style="Card.TFrame")
        user_card.pack(side="left", padx=6)
        tb.Label(
            user_card,
            text="Usuario",
            style="Muted.TLabel",
            background="#0E1B2B",
        ).pack(anchor="w")
        tb.Label(
            user_card,
            text=f"{self.session_context.display_name} · {self.session_context.role}",
            font=("Arial", 10, "bold"),
            foreground="#FFFFFF",
            background="#0E1B2B",
        ).pack(anchor="w")

        self.boton_cambiar_turno = tb.Button(
            info_header, text="Cambiar Turno", bootstyle=SECONDARY,
            command=self.request_change_admission_turn, width=17
        )
        self.boton_cambiar_turno.pack(side="left", padx=(8, 0), ipady=5)

        self.actions_menu_button = tb.Menubutton(
            header, text="Menú", bootstyle=SECONDARY, width=12
        )
        self.actions_menu_button.grid(row=0, column=3, rowspan=2, sticky="e", padx=(10, 0))
        self.actions_menu = tk.Menu(self.actions_menu_button, tearoff=0)
        self.change_turn_menu_action = self.actions_menu.add_command(
            label="Cambiar turno", command=self.request_change_admission_turn
        )
        self.transfer_primary_menu_action = None
        if normalize_role(getattr(self.session_context, "role", "")) == ROLE_ADMIN:
            self.transfer_primary_menu_action = self.actions_menu.add_command(
                label="Transferir acceso principal",
                command=self.request_transfer_admission_primary,
            )
        self.actions_menu.add_command(label="Historial", command=self.abrir_historial)
        if self._puede(CAP_VIEW_REPORTS):
            self.actions_menu.add_command(label="Reporte estadístico", command=self.abrir_ventana_reporte)
        if self._puede(CAP_OPEN_EXCEL):
            self.actions_menu.add_command(label="Listado de Excel", command=self._abrir_excel_actual)
        self.actions_menu.add_separator()
        if self._puede(CAP_EDIT_RECORDS):
            self.actions_menu.add_command(label="Editar paciente", command=self._abrir_edicion_paciente)
        self.actions_menu.add_command(label="Impresiones y documentos pendientes", command=self.abrir_trabajos_salida_pendientes)
        if self._puede(CAP_INTERNAL_CONFIG):
            self.actions_menu.add_command(label="Configuración", command=self._abrir_configuracion_interna)
        self.actions_menu.aboutToShow.connect(self._refresh_actions_menu_state)
        self.actions_menu_button.configure(menu=self.actions_menu)

        content_area = tb.Frame(self.main, style="Root.TFrame")
        content_area.pack(fill="both", expand=True)
        content_area.columnconfigure(0, weight=1)
        content_area.columnconfigure(1, weight=0)
        content_area.columnconfigure(2, weight=0)
        content_area.rowconfigure(0, weight=1)
        self.content_area = content_area

        # PySide6: el formulario principal se monta directamente en el grid.
        # El Canvas de Tk era útil para pantallas pequeñas, pero al traducirlo a
        # QScrollArea alteraba el ancho real de la tarjeta y provocaba el gran
        # vacío central / superposiciones observadas en V5. La versión original
        # de referencia usa esta distribución directa en escritorio.
        self.frame = tb.Frame(content_area, padding=18, style="Card.TFrame")
        self.frame.grid(row=0, column=0, sticky="nsew", padx=(0, 0))
        self.form_host = self.frame
        self.form_canvas = self.frame  # alias compatible para preferencias visuales
        self.form_scrollbar = None
        self._form_canvas_window = None
        for col in range(6):
            self.frame.columnconfigure(col, weight=1)

        # Separador físico real entre DATOS DEL PACIENTE y ACCIONES RÁPIDAS.
        # No depende de un borde QSS del panel: Qt siempre reserva esta columna.
        self.quick_separator = ttk.Separator(content_area, orient="vertical")
        self.quick_separator.grid(row=0, column=1, sticky="ns", padx=(10, 12))
        try:
            self.quick_separator.setMinimumWidth(2)
            self.quick_separator.setMaximumWidth(2)
        except Exception:
            pass

        self.quick_panel = tb.Frame(content_area, padding=14, style="Card.TFrame")
        self.quick_panel.grid(row=0, column=2, sticky="ns")
        self.quick_panel.configure(width=350)
        try:
            self.quick_panel.grid_propagate(False)
        except Exception:
            pass
        self._crear_panel_acciones_rapidas(self.quick_panel)

        title_row = tb.Frame(self.frame, style="Card.TFrame")
        title_row.grid(row=0, column=0, columnspan=6, sticky="ew", pady=(0, 12))
        title_row.columnconfigure(0, weight=1)
        self.section_patient_label = tb.Label(
            title_row, text="👤  DATOS DEL PACIENTE", style="Section.TLabel", background="#0E1B2B"
        )
        self.section_patient_label.grid(row=0, column=0, sticky="w")
        self.boton_historial = tb.Button(title_row, text="Historial", command=self.abrir_historial, width=20, bootstyle=INFO)
        self.boton_historial.grid(row=0, column=1, sticky="e")
        sep = ttk.Separator(self.frame, orient="horizontal")
        sep.grid(row=1, column=0, columnspan=6, sticky="ew", pady=(0, 14))

        def lbl(text, row, col, colspan=1):
            widget = tb.Label(
                self.frame,
                text=text,
                font=("Arial", 10, "bold"),
                foreground="#EAF2FF",
                background="#0E1B2B",
            )
            widget.grid(row=row, column=col, columnspan=colspan, sticky="w", padx=(4, 10), pady=(2, 3))
            return widget

        self.lbl_nombre = lbl("Nombre  ·  Nombre completo del paciente", 2, 0, 3)
        self.lbl_nombre.setWordWrap(True)
        self.entry_nombre = tb.Entry(self.frame)
        self.entry_nombre.grid(row=3, column=0, columnspan=3, sticky="ew", padx=(4, 24), pady=(0, 10), ipady=6)

        self.lbl_sexo = lbl("Sexo", 2, 3, 2)
        self.var_sexo = tk.StringVar(value="Femenino")
        
        sexo_frame = tb.Frame(self.frame)
        self.sexo_frame = sexo_frame
        sexo_frame.grid(row=3, column=3, columnspan=3, sticky="w", padx=(4, 4), pady=(0, 10))

        self.var_embarazada = tk.BooleanVar(value=False)

        self.lbl_sexo_m = tb.Radiobutton(sexo_frame, text="Masculino", variable=self.var_sexo, value="Masculino")
        self.lbl_sexo_m.pack(side="left", padx=(0, 10))

        self.lbl_sexo_f = tb.Radiobutton(sexo_frame, text="Femenino", variable=self.var_sexo, value="Femenino")
        self.lbl_sexo_f.pack(side="left", padx=(0, 10))

        self.check_embarazada = tb.Checkbutton(sexo_frame, text="Embarazada", variable=self.var_embarazada, bootstyle=INFO)

        self.var_sexo.trace_add("write", lambda *args: self.actualizar_embarazada())
        self.var_embarazada.trace_add("write", lambda *args: self.actualizar_embarazada())
        self.actualizar_embarazada()

        self.lbl_edad = lbl("Edad  ·  Ej: 25", 4, 0)
        self.entry_edad = tb.Entry(self.frame)
        self.entry_edad.grid(row=5, column=0, sticky="ew", padx=(4, 8), pady=(0, 10), ipady=6)
        self.unidad_edad = tk.StringVar(value="Años")
        self.combo_unidad = tb.Combobox(self.frame, textvariable=self.unidad_edad,
                                        values=["Días", "Meses", "Años"], state="readonly")
        self.combo_unidad.grid(row=5, column=1, columnspan=2, sticky="ew", padx=(0, 24), pady=(0, 10), ipady=6)

        self.var_urgencia = tk.BooleanVar(value=False)
        self.check_urgencia = tb.Checkbutton(
            self.frame,
            text="Atención de urgencia (conteo aparte)",
            variable=self.var_urgencia,
            bootstyle=INFO
        )
        self.check_urgencia.grid(row=5, column=3, columnspan=3, sticky="w", padx=(4, 4), pady=(0, 10))

        self.lbl_cedula = lbl("Cédula  ·  11 dígitos", 6, 0, 3)
        self.entry_cedula = tb.Entry(self.frame, validate="key", validatecommand=self.val_cedula)
        self.entry_cedula.grid(row=7, column=0, columnspan=3, sticky="ew", padx=(4, 24), pady=(0, 10), ipady=6)
        self.entry_cedula.bind("<KeyRelease>", lambda e: (self.limitar_caracteres(self.entry_cedula, 11), self._try_autocomplete_cedula()))
        self.entry_cedula.bind("<FocusOut>", self.auto_completar)
        self.entry_cedula.bind("<Return>", self.auto_completar)

        self.lbl_telefono = lbl("Teléfono  ·  10 dígitos", 6, 3, 3)
        self.entry_telefono = tb.Entry(self.frame, validate="key", validatecommand=self.val_telefono)
        self.entry_telefono.grid(row=7, column=3, columnspan=3, sticky="ew", padx=(4, 4), pady=(0, 10), ipady=6)
        self.entry_telefono.bind("<KeyRelease>", lambda e: self.limitar_caracteres(self.entry_telefono, 10))

        self.lbl_direccion = lbl("Dirección", 8, 0, 6)
        self.entry_direccion = tb.Entry(self.frame)
        self.entry_direccion.grid(row=9, column=0, columnspan=6, sticky="ew", padx=(4, 4), pady=(0, 10), ipady=6)

        self.lbl_nacionalidad = lbl("Nacionalidad", 10, 0, 3)
        self.entry_nacionalidad = tb.Entry(self.frame)
        self.entry_nacionalidad.grid(row=11, column=0, columnspan=3, sticky="ew", padx=(4, 24), pady=(0, 10), ipady=6)

        self.lbl_ars = lbl("Aseguradora (ARS)  ·  Escriba: SUB, HUMANO, MAPFRE…", 10, 3, 3)
        self.lbl_ars.setWordWrap(True)
        self.entry_ars = tb.Entry(self.frame)
        self.entry_ars.grid(row=11, column=3, columnspan=3, sticky="ew", padx=(4, 4), pady=(0, 0), ipady=6)
        self.entry_ars.bind("<KeyRelease>", self._on_ars_keyrelease)
        self.entry_ars.bind("<FocusOut>", lambda e: (self._actualizar_deteccion_seguro(), self.root.after(180, self._ocultar_sugerencias_ars)))
        self.entry_ars.bind("<Down>", self._focus_sugerencias_ars)

        self.ars_detectado_var = tk.StringVar(value="Detectado como: pendiente")
        self.ars_detectado_label = tb.Label(
            self.frame,
            textvariable=self.ars_detectado_var,
            font=("Arial", 9, "bold"),
            foreground="#8FA6BF",
            background="#0E1B2B"
        )
        self.ars_detectado_label.grid(row=12, column=3, columnspan=3, sticky="w", padx=(4, 4), pady=(2, 4))

        self.ars_suggestions = tk.Listbox(
            self.frame,
            height=5,
            bg="#0B1624",
            fg="#EAF2FF",
            selectbackground="#1D6EFF",
            selectforeground="#FFFFFF",
            highlightthickness=1,
            highlightbackground="#254260",
            relief="flat",
            font=("Arial", 10)
        )
        self.ars_suggestions.grid(row=13, column=3, columnspan=3, sticky="ew", padx=(4, 4), pady=(2, 8))
        self.ars_suggestions.grid_remove()
        self.ars_suggestions.bind("<ButtonRelease-1>", self._seleccionar_sugerencia_ars)
        self.ars_suggestions.bind("<Return>", self._seleccionar_sugerencia_ars)
        self.ars_suggestions.bind("<Escape>", lambda e: self._ocultar_sugerencias_ars())
        self._ars_catalogo = self._obtener_catalogo_ars()

        self.lbl_nss = lbl("NSS  ·  Número de seguro o SIN SEGURO", 13, 0, 1)
        self.lbl_nss.setWordWrap(True)

        self.nss_detectado_var = tk.StringVar(value="NSS: pendiente")
        self.nss_detectado_label = tb.Label(
            self.frame,
            textvariable=self.nss_detectado_var,
            font=("Arial", 9, "bold"),
            foreground="#8FA6BF",
            background="#0E1B2B"
        )
        self.nss_detectado_label.grid(row=13, column=1, columnspan=2, sticky="w", padx=(4, 24), pady=(0, 2))

        self.entry_nss = tb.Entry(self.frame)
        self.entry_nss.grid(row=14, column=0, columnspan=3, sticky="ew", padx=(4, 24), pady=(0, 6), ipady=6)
        self.entry_nss.bind("<KeyRelease>", lambda e: (self._actualizar_deteccion_seguro(), self._try_autocomplete_nss()))
        self.entry_nss.bind("<FocusOut>", self.auto_completar_por_nss)
        self.entry_nss.bind("<Return>", self.auto_completar_por_nss)

        self.form_actions_separator = ttk.Separator(self.frame, orient="horizontal")
        self.form_actions_separator.grid(row=15, column=0, columnspan=6, sticky="ew", pady=(6, 6))

        btns = tb.Frame(self.frame, style="Card.TFrame")
        self.form_buttons = btns
        btns.grid(row=16, column=0, columnspan=6, sticky="e", pady=(0, 10))
        self.boton_limpiar = tb.Button(btns, text="Limpiar", command=self.limpiar_campos, width=18, bootstyle=DANGER)
        self.boton_limpiar.pack(side="left", padx=(0, 12), ipady=4)
        self.boton_generar_pdf = tb.Button(btns, text="Generar PDF", command=self.generar_pdf, width=22, bootstyle=PRIMARY)
        self.boton_generar_pdf.pack(side="left", ipady=4)

        status = tb.Frame(self.root, padding=(18, 8), style="Root.TFrame")
        status.pack(fill="x", side="bottom", before=self.main)

        self.shortcuts_var = tk.StringVar(
            value="ⓘ  Ctrl+Z: deshacer | F5: cambio de turno | Ctrl+H: historial | Ctrl+L: limpiar | Enter: generar si está completo"
        )
        self.shortcuts_label = tb.Label(
            status,
            textvariable=self.shortcuts_var,
            style="Muted.TLabel",
            background="#07111f"
        )
        self.shortcuts_label.pack(side="left")

        self.connection_var = tk.StringVar(value="Conectado a: verificando…")
        self.connection_label = tb.Label(
            status,
            textvariable=self.connection_var,
            font=("Arial", 10, "bold"),
            foreground="#72E39B",
            background="#07111f"
        )
        self.connection_label.pack(side="right")

        self.status_var = tk.StringVar(value="✓  Listo para generar formulario")
        self.status_label = tb.Label(
            status,
            textvariable=self.status_var,
            font=("Arial", 10, "bold"),
            foreground="#5CB6FF",
            background="#07111f"
        )
        self.status_label.pack(side="right", padx=(0, 24))

        self.notif_frame = tb.Frame(self.root, padding=(10, 6), style="Root.TFrame")
        self.notif_frame.pack(fill="x", side="bottom")
        self.notif_frame.pack_forget()
        self.notif_label = tb.Label(self.notif_frame, text="", anchor="w", background="#07111f")
        self.notif_label.pack(side="left", fill="x", expand=True)
        self.btn_deshacer = tb.Button(self.notif_frame, text="Deshacer", bootstyle=WARNING)
        self.btn_deshacer.pack(side="right", padx=6)

        self._notif_after_id = None
        self._ultimo_atencion_id = None
        self._last_report_summary = None

        self.historial_win = None
        self.historial_sin_seguro_win = None
        self.reporte_win = None
        self.turno_win = None
        self.dialogo_unico_win = None
        self.edicion_paciente_win = None
        self.configuracion_interna_win = None
        self.salida_pendiente_win = None
        self.trabajos_salida_win = None
        self._output_payloads = {}
        self._output_inflight = set()
        self._output_lock = threading.Lock()

        self.menu_contextual = tk.Menu(self.root, tearoff=0)
        self.menu_contextual.add_command(label="Copiar", command=self._copiar)
        self.menu_contextual.add_command(label="Pegar", command=self._pegar)
        self.menu_contextual.add_command(label="Cortar", command=self._cortar)
        for w in [self.entry_nombre, self.entry_edad, self.entry_cedula, self.entry_telefono,
                  self.entry_direccion, self.entry_nacionalidad, self.entry_ars, self.entry_nss]:
            w.bind("<Button-3>", self.mostrar_menu_contextual)

        self.root.bind('<F5>', self.request_change_admission_turn)
        self._suspend_autocomplete = False
        self._last_autofill_identity = None
        self._last_autofill_at = 0.0
        self._autofill_cloud_pending = set()
        self._patient_fields_modified_by_user = set()
        self._patient_autofill_baseline = {}
        self._verified_cloud_patient = None
        self._verified_cloud_identity = None
        self._local_patient_revision = 0
        self._final_revalidation_in_progress = False
        self._final_revalidation_ready = False
        self._field_focus_started_at = 0.0
        self.entry_nombre.focus_set()

        self.all_entries = [
            self.entry_nombre,
            self.entry_edad,
            self.entry_cedula,
            self.entry_telefono,
            self.entry_direccion,
            self.entry_nacionalidad,
            self.entry_ars,
            self.entry_nss
        ]
        self._initial_styles = {}
        self._pending_restores = {}
        self._capture_initial_styles()
        self._configurar_accesibilidad_teclado()

        # FASE 9B: Vincular FocusIn/Out para Ctrl+Z por campo
        for _w in (self.entry_nombre, self.entry_edad, self.entry_cedula,
                   self.entry_telefono, self.entry_direccion,
                   self.entry_nacionalidad, self.entry_ars, self.entry_nss):
            try:
                _w.bind("<FocusIn>",  self._on_field_focus_in)
                _w.bind("<FocusOut>", self._on_field_focus_out)
            except Exception:
                pass

        self._aplicar_preferencias_en_vivo()
        self._set_turn_change_controls_enabled(True)
        self._responsive_after_id = None
        self.root.bind("<Configure>", self._programar_modo_responsivo, add="+")
        try:
            self.root.after(400, self._actualizar_resumen_turno_panel)
            self._summary_after_id = self.root.after(1800, self._programar_refresco_resumen_en_vivo)
            self._date_after_id = self.root.after(1000, self._actualizar_fecha_actual)
            self.root.after(2200, self._avisar_trabajos_salida_pendientes)
            self.root.after(2600, self._retry_excel_export_jobs)
        except Exception:
            pass

        if turno_cfg:
            try:
                self.root.after(1200, lambda cfg=turno_cfg: self._reconstruir_excel_inicio_diferido(cfg))
            except Exception:
                pass
        elif getattr(self, "_excel_pendiente_turno_manual", False):
            try:
                self.root.after(1200, self._avisar_turno_manual_pendiente)
            except Exception:
                pass

        self._main_session_failures = 0
        if self.session_context.launched_from_billing:
            self._session_after_id = self.root.after(1200, self._verificar_sesion_principal)
            self.root.after(1600, self._asegurar_turno_de_sesion)

    def _asegurar_preferencias_impresion_hoja(self):
        try:
            changed = False
            if self.app_settings.get("auto_print") is not True:
                self.app_settings["auto_print"] = True
                changed = True
            if self.app_settings.get("print_auto_hoja") is not True:
                self.app_settings["print_auto_hoja"] = True
                changed = True
            if str(self.app_settings.get("print_behavior_hoja", "")).strip() not in ["Solo imprimir", "Imprimir y abrir PDF"]:
                self.app_settings["print_behavior_hoja"] = "Imprimir y abrir PDF"
                changed = True
            try:
                copias = int(self.app_settings.get("print_copies_hoja", 1) or 1)
                if copias < 1:
                    self.app_settings["print_copies_hoja"] = 1
                    changed = True
            except Exception:
                self.app_settings["print_copies_hoja"] = 1
                changed = True
            if changed:
                guardar_app_settings(self.app_settings)
        except Exception:
            pass

    def _reconstruir_excel_inicio_diferido(self, turno_cfg):
        try:
            verificar_o_crear_excel()
            canonical_builder = getattr(
                self.db, "build_current_admission_list_dataset", None
            )
            if callable(canonical_builder):
                filas_bd = list(canonical_builder() or [])
            else:
                contexto = self.db.buscar_contexto_turno_existente(turno_cfg)
                filas_bd = (
                    self.db.obtener_atenciones_para_rango_real(
                        turno_id=int(contexto["turno_id"])
                    )
                    if contexto
                    else []
                )
            filas_excel = int(
                resumen_excel_actual_simple(turno_cfg).get("total", 0) or 0
            )
            if filas_bd:
                reconstruir_excel_turno(self.db, turno_cfg)
                self.set_status(
                    f"Listado verificado: {len(filas_bd)} paciente(s) sincronizados.",
                    "ok",
                )
            elif filas_excel:
                self.set_status(
                    f"Listado recuperado: {filas_excel} paciente(s) visibles; "
                    "no se sobrescribió el Excel.",
                    "warning",
                )
            else:
                self.set_status("Inicio rápido: Excel verificado", "ok")
            self._actualizar_turno_visual_en_vivo()
            self._refrescar_resumen_en_vivo()
        except PermissionError:
            self.set_status("Excel abierto. Cierre el listado para actualizarlo.", "warning")
        except Exception as e:
            self.set_status(f"Aviso al verificar Excel: {e}", "warning")

    def _avisar_turno_manual_pendiente(self):
        aviso = (
            "El Excel conserva pacientes, pero no hay un turno configurado. "
            "El listado no fue modificado; use Cambiar turno para confirmarlo manualmente."
        )
        self.set_status(aviso, "warning")
        self._mostrar_notificacion(aviso, autohide_ms=15000, tipo="warning")

    def _paleta_visual_actual(self):
        """
        FASE 13: Paleta sobria con colores institucionales.
        """
        if self._host_theme_controlled and self._host_theme_is_dark is not None:
            theme = "oscuro" if self._host_theme_is_dark else "claro"
        else:
            theme = str(self.app_settings.get("theme", "oscuro") or "oscuro").lower()
        if self._host_theme_controlled and self._host_visual_theme:
            return self._paleta_con_alto_contraste(dict(self._host_visual_theme))

        high = bool(self.app_settings.get("high_contrast", False))
        accent_pref = self.app_settings.get("accent_color", "Azul hospitalario")
        accent_hex = resolver_color_principal(accent_pref)

        accent_dark_soft = mezclar_color_hex(accent_hex, "#111E2E", 0.18)
        accent_dark_heading = mezclar_color_hex(accent_hex, "#172A3E", 0.25)
        accent_dark_selected = mezclar_color_hex(accent_hex, "#2D5F93", 0.55)

        accent_light_soft = mezclar_color_hex(accent_hex, "#FFFFFF", 0.10)
        accent_light_heading = mezclar_color_hex(accent_hex, "#DDE6F0", 0.18)
        accent_light_selected = mezclar_color_hex(accent_hex, "#4B83C2", 0.50)

        accent_high_dark_soft = mezclar_color_hex(accent_hex, "#10151C", 0.22)
        accent_high_light_soft = mezclar_color_hex(accent_hex, "#FFFFFF", 0.12)

        if high and theme == "claro":
            return self._normalizar_paleta_visual({
                "mode": "claro_alto",
                "root": "#FAFAFA",
                "card": "#FFFFFF",
                "card2": accent_high_light_soft,
                "entry": "#FFFFFF",
                "tree": "#FFFFFF",
                "heading": mezclar_color_hex(accent_hex, "#E1E6EE", 0.18),
                "text": "#111827",
                "muted": "#374151",
                "accent": accent_hex,
                "accent2": COLOR_INFO,
                "border": "#111827",
                "separator": "#6B7280",
                "selected_bg": mezclar_color_hex(accent_hex, "#E6B800", 0.30),
                "selected_fg": "#111827",
                "button_fg": "#FFFFFF",
                "danger": "#8B4C4C",
                "warning": "#7A6B4A",
                "success": "#4F7B55",
                "info": COLOR_INFO,
            })

        if high and theme != "claro":
            return self._normalizar_paleta_visual({
                "mode": "oscuro_alto",
                "root": "#080A0D",
                "card": "#10151C",
                "card2": accent_high_dark_soft,
                "entry": "#080A0D",
                "tree": "#080A0D",
                "heading": mezclar_color_hex(accent_hex, "#1B2532", 0.22),
                "text": "#F8FAFC",
                "muted": "#D7DEE8",
                "accent": accent_hex,
                "accent2": COLOR_INFO,
                "border": "#E5E7EB",
                "separator": "#B7C0CA",
                "selected_bg": mezclar_color_hex(accent_hex, "#D8B536", 0.30),
                "selected_fg": "#111827",
                "button_fg": "#FFFFFF",
                "danger": "#C85858",
                "warning": "#D7A24A",
                "success": "#5BAA70",
                "info": COLOR_INFO,
            })

        if theme == "claro":
            return self._normalizar_paleta_visual({
                "mode": "claro",
                "root": "#EEF2F6",
                "card": "#FFFFFF",
                "card2": accent_light_soft,
                "entry": "#FFFFFF",
                "tree": "#FFFFFF",
                "heading": accent_light_heading,
                "text": "#1F2A37",
                "muted": "#566678",
                "accent": accent_hex,
                "accent2": COLOR_INFO,
                "border": "#A9BACB",
                "separator": "#CCD6E2",
                "selected_bg": accent_light_selected,
                "selected_fg": "#FFFFFF",
                "button_fg": "#FFFFFF",
                "danger": "#9B6464",
                "warning": "#8B7B5A",
                "success": "#5F7B69",
                "info": COLOR_INFO,
            })

        return self._normalizar_paleta_visual({
            "mode": "oscuro",
            "root": "#0A1420",
            "card": "#111E2E",
            "card2": accent_dark_soft,
            "entry": "#0F1B2A",
            "tree": "#0B1724",
            "heading": accent_dark_heading,
            "text": "#E5EEF8",
            "muted": "#AAB8C8",
            "accent": accent_hex,
            "accent2": COLOR_INFO,
            "border": "#36516A",
            "separator": "#40566D",
            "selected_bg": accent_dark_selected,
            "selected_fg": "#FFFFFF",
            "button_fg": "#FFFFFF",
            "danger": COLOR_DANGER,
            "warning": COLOR_WARNING,
            "success": COLOR_SUCCESS,
            "info": COLOR_INFO,
        })

    def _paleta_con_alto_contraste(self, palette):
        """Derive only contrast-sensitive fields from the host token contract."""
        if not bool(self.app_settings.get("high_contrast", False)):
            return palette
        result = dict(palette)
        dark = str(result.get("mode") or "").startswith("oscuro")
        if dark:
            result.update(
                {
                    "text": "#FFFFFF",
                    "text_primary": "#FFFFFF",
                    "muted": "#E1EAF4",
                    "text_muted": "#E1EAF4",
                    "border": "#D8E4F0",
                    "border_strong": "#FFFFFF",
                    "separator": "#B8C8D8",
                    "selected_bg": "#2E7DBA",
                    "selection_bg": "#2E7DBA",
                }
            )
        else:
            result.update(
                {
                    "text": "#111827",
                    "text_primary": "#111827",
                    "muted": "#253342",
                    "text_muted": "#253342",
                    "border": "#253342",
                    "border_strong": "#111827",
                    "separator": "#5A6877",
                    "selected_bg": "#0B5CA6",
                    "selection_bg": "#0B5CA6",
                }
            )
        return result

    @staticmethod
    def _normalizar_paleta_visual(palette):
        """Complete the standalone palette with the host semantic contract."""
        result = dict(palette)
        dark = str(result.get("mode") or "").startswith("oscuro")
        result.setdefault("window_bg", result["root"])
        result.setdefault("content_bg", result["root"])
        result.setdefault("panel_bg", result["card"])
        result.setdefault("panel_elevated_bg", result["card2"])
        result.setdefault("input_bg", result["entry"])
        result.setdefault("input_hover_bg", result["entry"])
        result.setdefault("input_disabled_bg", result["card2"])
        result.setdefault("table_bg", result["tree"])
        result.setdefault("table_alt_bg", result["card2"])
        result.setdefault("table_header_bg", result["heading"])
        result.setdefault("popup_bg", result["card"])
        result.setdefault("menu_bg", result["card"])
        result.setdefault("text_primary", result["text"])
        result.setdefault("text_secondary", result["muted"])
        result.setdefault("text_muted", result["muted"])
        result.setdefault("text_disabled", "#8EA1B2" if dark else "#7C8B9A")
        result.setdefault("text_on_accent", result["button_fg"])
        result.setdefault("selection_bg", result["selected_bg"])
        result.setdefault("selection_text", result["selected_fg"])
        result.setdefault("border_strong", result["border"])
        result.setdefault("border_focus", result["accent"])
        result.setdefault("accent_hover", result["accent"])
        result.setdefault("checkbox_indicator_bg", result["entry"])
        result.setdefault("checkbox_indicator_border", result["border"])
        result.setdefault("checkbox_checked_bg", result["accent"])
        result.setdefault("checkbox_checkmark", result["button_fg"])
        result.setdefault("danger_bg", "#3A2027" if dark else "#FDECEC")
        result.setdefault("danger_text", "#FFB9C0" if dark else "#8D1914")
        result.setdefault("warning_bg", "#3A2E14" if dark else "#FFF3D6")
        result.setdefault("warning_text", "#FFE3A3" if dark else "#704600")
        result.setdefault("success_bg", "#163A2D" if dark else "#E6F5EC")
        result.setdefault("success_text", "#C0F5D9" if dark else "#175C36")
        result.setdefault("info_bg", "#17354A" if dark else "#EDF4FD")
        result.setdefault("info_text", "#D0ECFF" if dark else "#17445A")
        button_defaults = {
            "primary": (result["accent"], result["accent"], result["button_fg"]),
            "secondary": (result["card2"], result["card2"], result["text"]),
            "neutral": (result["card2"], result["card2"], result["text"]),
            "success": (result["success"], result["success"], result["button_fg"]),
            "warning": (result["warning"], result["warning"], result["button_fg"]),
            "danger": (result["danger"], result["danger"], result["button_fg"]),
        }
        for role, (base, hover, foreground) in button_defaults.items():
            result.setdefault(f"button_{role}_bg", base)
            result.setdefault(f"button_{role}_hover", hover)
            result.setdefault(f"button_{role}_text", foreground)
        return result

    def set_embedded_visual_notifier(self, callback):
        """Register the host-only hook used for non-data visual recovery."""
        self._embedded_visual_notifier = callback

    def _notify_embedded_visual_event(self, event_name):
        callback = self._embedded_visual_notifier
        if not callable(callback):
            return
        try:
            callback(str(event_name))
        except Exception:
            APP_LOG.debug("No se pudo notificar el evento visual %s", event_name)

    def _embedded_theme_label(self):
        return "Noche" if self._host_theme_is_dark else "Día"

    def _refresh_embedded_theme_indicators(self):
        if not self._host_theme_controlled:
            return
        label = f"Controlado por Facturación ({self._embedded_theme_label()})"
        active_indicators = []
        for indicator in self._embedded_theme_indicators:
            try:
                indicator.set(label)
                active_indicators.append(indicator)
            except (RuntimeError, tk.TclError):
                continue
        self._embedded_theme_indicators = active_indicators

    def _apply_preferences_to_open_windows(self):
        for attr in (
            "historial_win",
            "historial_sin_seguro_win",
            "reporte_win",
            "turno_win",
            "dialogo_unico_win",
            "configuracion_interna_win",
            "edicion_paciente_win",
            "salida_pendiente_win",
            "trabajos_salida_win",
        ):
            try:
                window = getattr(self, attr, None)
                if window is not None and window.winfo_exists():
                    self._aplicar_preferencias_a_widgets(window)
                    refresh = getattr(window, "_admission_theme_refresh", None)
                    if callable(refresh):
                        refresh()
            except (RuntimeError, tk.TclError):
                continue

    def _apply_theme_to_new_window(self, window):
        """Polish a newly-created dialog before its first visible frame."""
        if self._shutdown_complete:
            return
        self._aplicar_preferencias_a_widgets(window)
        refresh = getattr(window, "_admission_theme_refresh", None)
        if callable(refresh):
            refresh()

    def apply_host_theme(self, is_dark: bool, *, theme_tokens=None) -> bool:
        """Adopt the main application's theme in embedded mode only."""
        if not self._host_theme_controlled:
            return False
        is_dark = bool(is_dark)
        changed = self._host_theme_is_dark is not is_dark
        self._host_theme_is_dark = is_dark
        if isinstance(theme_tokens, dict):
            self._host_visual_theme = dict(theme_tokens)
            self.shared_configuration["host_visual_theme"] = dict(theme_tokens)
        else:
            # Los hosts anteriores solo comunican el modo. Nunca se debe
            # conservar una paleta previa de otro modo en ese contrato.
            self._host_visual_theme = None
            self.shared_configuration.pop("host_visual_theme", None)
        self.shared_configuration["host_theme_is_dark"] = is_dark
        self._aplicar_preferencias_en_vivo(forzar_todo=True)
        self._refresh_embedded_theme_indicators()
        return changed

    def ensure_embedded_entry_geometry(self, *, establish_baseline=False):
        """Keep form controls above their font-derived usable height."""
        entries = (
            "entry_nombre",
            "entry_edad",
            "entry_cedula",
            "entry_telefono",
            "entry_direccion",
            "entry_nacionalidad",
            "entry_ars",
            "entry_nss",
        )
        for entry_name in entries:
            entry = getattr(self, entry_name, None)
            if entry is None:
                continue
            try:
                content_height = entry.fontMetrics().height() + 14
                valid_height = max(
                    int(entry.minimumSizeHint().height()),
                    int(entry.sizeHint().height()),
                    content_height,
                )
                established_floor = self._embedded_entry_height_floor.get(entry_name)
                if established_floor is None:
                    established_floor = max(valid_height, int(entry.minimumHeight()))
                if establish_baseline:
                    established_floor = max(
                        int(established_floor),
                        int(entry.minimumHeight()),
                        int(entry.height()),
                    )
                self._embedded_entry_height_floor[entry_name] = established_floor
                final_height = max(valid_height, int(established_floor))
                entry.setMinimumHeight(final_height)
                entry._compat_ipady_minimum_height = max(
                    int(getattr(entry, "_compat_ipady_minimum_height", 0) or 0),
                    final_height,
                )
                entry.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Fixed)
            except RuntimeError:
                continue

    def _schedule_embedded_entry_geometry_sync(self):
        """Apply the font-derived floor after Qt finishes polishing a style."""
        if self._entry_geometry_sync_scheduled:
            return
        self._entry_geometry_sync_scheduled = True

        def apply_geometry_floor():
            self._entry_geometry_sync_scheduled = False
            if not self._shutdown_complete:
                self.ensure_embedded_entry_geometry()

        QTimer.singleShot(0, apply_geometry_floor)

    def _font_size_pref(self):
        try:
            return max(10, min(18, int(self.app_settings.get("font_size", 11))))
        except Exception:
            return 11

    def _configurar_estilos_desde_preferencias(self):
        pal = self._paleta_visual_actual()
        fs = self._font_size_pref()
        style = getattr(self, "style", None) or tb.Style()

        qt_compat.set_compat_theme_tokens(pal)

        try:
            self.root.configure(bg=pal["root"])
            self.root.option_add("*Font", f"Arial {fs}")
            self.root.option_add("*TCombobox*Listbox.background", pal["entry"])
            self.root.option_add("*TCombobox*Listbox.foreground", pal["text"])
            self.root.option_add("*TCombobox*Listbox.selectBackground", pal["selected_bg"])
            self.root.option_add("*TCombobox*Listbox.selectForeground", pal["selected_fg"])
            self.root.option_add("*TCombobox*Listbox.font", f"Arial {fs}")
            self.root.option_add("*Listbox.background", pal["entry"])
            self.root.option_add("*Listbox.foreground", pal["text"])
            self.root.option_add("*Listbox.selectBackground", pal["selected_bg"])
            self.root.option_add("*Listbox.selectForeground", pal["selected_fg"])
        except Exception:
            pass

        try:
            style.configure("Root.TFrame", background=pal["root"])
            style.configure("Card.TFrame", background=pal["card"])
            style.configure("TFrame", background=pal["card"])

            style.configure(
                "TLabelframe",
                background=pal["card"],
                foreground=pal["text"],
                bordercolor=pal["border"],
                lightcolor=pal["border"],
                darkcolor=pal["border"]
            )
            style.configure(
                "TLabelframe.Label",
                _qt_transparent=True,
                foreground=pal["accent"],
                font=("Arial", fs, "bold")
            )

            # Plain labels must never paint their own card.  Semantic panels
            # opt into a background explicitly; normal labels stay transparent.
            style.configure("TLabel", font=("Arial", fs), foreground=pal["text"], _qt_transparent=True)
            style.configure("Muted.TLabel", font=("Arial", max(fs - 1, 10)), foreground=pal["muted"], _qt_transparent=True)
            style.configure("Title.TLabel", font=("Arial", fs + 7, "bold"), foreground=pal["text"], _qt_transparent=True)
            style.configure("Subtitle.TLabel", font=("Arial", fs + 1), foreground=pal["muted"], _qt_transparent=True)
            style.configure("Section.TLabel", font=("Arial", fs + 2, "bold"), foreground=pal["accent"], _qt_transparent=True)

            button_common = {
                "font": ("Arial", fs, "bold"),
                "disabled_background": pal.get("input_disabled_bg", pal["card2"]),
                "disabled_foreground": pal.get("text_disabled", pal["muted"]),
                "focus_border": pal.get("border_focus", pal["accent"]),
                "bordercolor": pal["border"],
            }
            style.configure("TButton", foreground=pal["button_fg"], **button_common)
            for style_name, role_key, fallback in (
                ("primary.TButton", "primary", COLOR_PRIMARY),
                ("success.TButton", "success", COLOR_SUCCESS),
                ("warning.TButton", "warning", COLOR_WARNING),
                ("danger.TButton", "danger", COLOR_DANGER),
                ("info.TButton", "primary", COLOR_INFO),
                ("secondary.TButton", "secondary", pal["card2"]),
            ):
                style.configure(
                    style_name,
                    background=pal.get(f"button_{role_key}_bg", fallback),
                    foreground=pal.get(f"button_{role_key}_text", pal["button_fg"]),
                    hoverbackground=pal.get(
                        f"button_{role_key}_hover", pal.get("accent_hover", pal["accent"])
                    ),
                    pressedbackground=pal.get(
                        f"button_{role_key}_hover", pal.get("accent_hover", pal["accent"])
                    ),
                    **button_common,
                )

            style.configure(
                "TEntry",
                font=("Arial", fs),
                fieldbackground=pal["entry"],
                foreground=pal["text"],
                bordercolor=pal["border"],
                lightcolor=pal["border"],
                darkcolor=pal["border"],
                insertcolor=pal["text"],
                hoverbackground=pal.get("input_hover_bg", pal["entry"]),
                disabled_background=pal.get("input_disabled_bg", pal["card2"]),
                focus_border=pal.get("border_focus", pal["accent"]),
            )
            style.map(
                "TEntry",
                fieldbackground=[("focus", pal["entry"]), ("!disabled", pal["entry"])],
                foreground=[("!disabled", pal["text"])]
            )

            style.configure(
                "TCombobox",
                font=("Arial", fs),
                fieldbackground=pal["entry"],
                background=pal["entry"],
                foreground=pal["text"],
                bordercolor=pal["border"],
                lightcolor=pal["border"],
                darkcolor=pal["border"],
                arrowcolor=pal["accent"],
                hoverbackground=pal.get("input_hover_bg", pal["entry"]),
                disabled_background=pal.get("input_disabled_bg", pal["card2"]),
                focus_border=pal.get("border_focus", pal["accent"]),
            )
            style.map(
                "TCombobox",
                fieldbackground=[("readonly", pal["entry"]), ("!disabled", pal["entry"])],
                background=[("readonly", pal["entry"]), ("!disabled", pal["entry"])],
                foreground=[("readonly", pal["text"]), ("!disabled", pal["text"])],
                selectbackground=[("readonly", pal["selected_bg"]), ("!disabled", pal["selected_bg"])],
                selectforeground=[("readonly", pal["selected_fg"]), ("!disabled", pal["selected_fg"])],
                arrowcolor=[("readonly", pal["accent"]), ("!disabled", pal["accent"])]
            )

            choice_style = {
                "font": ("Arial", fs),
                "foreground": pal["text"],
                "background": "transparent",
                "indicator_bg": pal["entry"],
                "indicator_border": pal.get("checkbox_indicator_border", pal["border"]),
                "indicator_checked": pal.get("checkbox_checked_bg", pal["accent"]),
                "focus_border": pal.get("border_focus", pal["accent"]),
                "disabledforeground": pal.get("text_disabled", pal["muted"]),
                "disabled_background": pal.get("input_disabled_bg", pal["card2"]),
            }
            style.configure("TCheckbutton", **choice_style)
            style.configure("TRadiobutton", **choice_style)

            style.configure("TNotebook", background=pal["card"], bordercolor=pal["border"])
            style.configure(
                "TNotebook.Tab",
                font=("Arial", max(fs - 1, 10), "bold"),
                padding=(10, 5),
                background=pal["card2"],
                foreground=pal["text"]
            )
            style.map(
                "TNotebook.Tab",
                background=[("selected", pal["heading"]), ("active", pal["card2"])],
                foreground=[("selected", pal["text"]), ("active", pal["text"])]
            )

            style.configure(
                "Treeview",
                font=("Arial", fs),
                rowheight=max(int(self.app_settings.get("table_row_height", 29) or 29), fs + 19),
                background=pal["tree"],
                foreground=pal["text"],
                fieldbackground=pal["tree"],
                bordercolor=pal["border"],
                lightcolor=pal["border"],
                darkcolor=pal["border"]
            )
            style.configure(
                "Treeview.Heading",
                font=("Arial", fs, "bold"),
                background=pal["heading"],
                foreground=pal["text"],
                bordercolor=pal["border"],
                lightcolor=pal["border"],
                darkcolor=pal["border"]
            )
            style.map(
                "Treeview",
                background=[("selected", pal["selected_bg"])],
                foreground=[("selected", pal["selected_fg"])]
            )

            style.configure(
                "Modern.Treeview",
                font=("Arial", fs),
                rowheight=max(int(self.app_settings.get("table_row_height", 29) or 29), fs + 20),
                background=pal["tree"],
                foreground=pal["text"],
                fieldbackground=pal["tree"],
                bordercolor=pal["border"],
                lightcolor=pal["border"],
                darkcolor=pal["border"]
            )
            style.configure(
                "Modern.Treeview.Heading",
                font=("Arial", fs, "bold"),
                background=pal["heading"],
                foreground=pal["text"],
                bordercolor=pal["border"],
                lightcolor=pal["border"],
                darkcolor=pal["border"]
            )
            style.map(
                "Modern.Treeview",
                background=[("selected", pal["selected_bg"])],
                foreground=[("selected", pal["selected_fg"])]
            )

            style.configure("TSeparator", background=pal["separator"])
        except Exception:
            pass
        self._configurar_colores_menu(pal, fs)

    def _configurar_colores_menu(self, pal, font_size):
        """Apply one high-contrast palette to enabled, hover and disabled items."""
        menu_qss = (
            f"QMenu {{ background-color: {pal['card']}; color: {pal['text']}; "
            f"border: 1px solid {pal['border']}; padding: 5px; "
            f"font-family: 'Segoe UI'; font-size: {int(font_size)}pt; }}"
            f"QMenu::item {{ color: {pal['text']}; background-color: transparent; "
            "padding: 8px 30px 8px 12px; margin: 1px 3px; border-radius: 3px; }}"
            f"QMenu::item:selected {{ background-color: {pal['selected_bg']}; "
            f"color: {pal['selected_fg']}; }}"
            f"QMenu::item:disabled {{ color: {pal['muted']}; "
            "background-color: transparent; }}"
            f"QMenu::separator {{ height: 1px; background-color: {pal['separator']}; "
            "margin: 5px 8px; }}"
            f"QMenu::indicator {{ width: 15px; height: 15px; }}"
            f"QMenu::right-arrow {{ color: {pal['text']}; }}"
        )
        for menu_name in ("actions_menu", "menu_contextual"):
            menu = getattr(self, menu_name, None)
            if menu is not None:
                try:
                    menu.setStyleSheet(menu_qss)
                except Exception:
                    pass

    def _aplicar_preferencias_a_widgets(self, base=None):
        pal = self._paleta_visual_actual()
        fs = self._font_size_pref()
        base = base or self.root

        old_root_colors = {
            "#07111f", "#07111F", "#000000", "#050505", "#080A0D", "#F3F6FA",
            "#f3f6fa", "#EEF3F8", "#eef3f8", "#FDFDFD", "#fdfdfd", "#FAFAFA", "#fafafa"
        }
        old_card_colors = {
            "#0E1B2B", "#0e1b2b", "#111E2E", "#111e2e", "#10151C", "#10151c",
            "#0A0A0A", "#0a0a0a", "#FFFFFF", "#ffffff", "#F6F9FC", "#f6f9fc",
            "#121212", "#132337", "#17283B", "#17283b"
        }
        old_entry_colors = {
            "#111E2E", "#111e2e", "#0F1D2E", "#0f1d2e", "#0B1624", "#0b1624",
            "#FFFFFF", "#ffffff", "#000000", "#080A0D", "#080a0d", "#0F1B2A", "#0f1b2a"
        }

        def safe_conf(w, **kwargs):
            try:
                w.configure(**kwargs)
            except Exception:
                pass

        def get_bg(w):
            for opt in ("background", "bg"):
                try:
                    return str(w.cget(opt))
                except Exception:
                    continue
            return ""

        def decide_bg(w, text=""):
            current_bg = get_bg(w)
            upper = (text or "").upper()

            if isinstance(w, (tk.Tk, tk.Toplevel)):
                return pal["root"]

            if any(x in upper for x in ["GENERADOR DE FORMULARIOS", "SISTEMA DE ADMISIÓN", "SISTEMA DE ADMISION"]):
                return pal["root"]

            if current_bg in old_root_colors:
                return pal["root"]

            if isinstance(w, (tk.Entry, tk.Text, tk.Listbox)):
                return pal["entry"]

            if current_bg in old_entry_colors and not isinstance(w, (tk.Label, ttk.Label)):
                return pal["entry"]

            if current_bg in old_card_colors:
                return pal["card"]

            return pal["card"]

        def semantic_button_palette(widget):
            role = str(getattr(widget, "_bootstyle", "") or "").lower()
            role_key = {
                "primary": "primary",
                "success": "success",
                "warning": "warning",
                "danger": "danger",
                "info": "primary",
                "secondary": "secondary",
            }.get(role, "secondary")
            return (
                pal.get(f"button_{role_key}_bg", pal["card2"]),
                pal.get(f"button_{role_key}_text", pal["text"]),
            )

        def walk(w):
            try:
                text = str(w.cget("text"))
            except Exception:
                text = ""

            bg = decide_bg(w, text)

            if isinstance(w, (tk.Tk, tk.Toplevel)):
                safe_conf(w, background=pal["root"], bg=pal["root"])

            elif isinstance(w, tk.Canvas):
                safe_conf(w, background=pal["root"], bg=pal["root"])
                try:
                    w.viewport().setStyleSheet(
                        f"background:{pal['root']};border:none;"
                    )
                    w.verticalScrollBar().setStyleSheet(
                        f"QScrollBar:vertical{{background:{pal.get('scrollbar_track', pal['root'])};width:12px;margin:0;}}"
                        f"QScrollBar::handle:vertical{{background:{pal.get('scrollbar_handle', pal['border'])};min-height:24px;border-radius:5px;}}"
                        f"QScrollBar::handle:vertical:hover{{background:{pal.get('scrollbar_hover', pal['accent'])};}}"
                        "QScrollBar::add-line:vertical,QScrollBar::sub-line:vertical{height:0;}"
                    )
                except Exception:
                    pass

            elif isinstance(w, tk.Frame):
                safe_conf(
                    w,
                    background=bg,
                    bg=bg,
                    highlightbackground=pal["border"],
                    highlightcolor=pal["accent"],
                )

            elif isinstance(w, tk.LabelFrame):
                safe_conf(
                    w,
                    background=pal["card"],
                    bg=pal["card"],
                    foreground=pal["text"],
                    fg=pal["text"],
                    font=("Arial", fs, "bold"),
                    highlightbackground=pal["border"],
                    highlightcolor=pal["accent"],
                )

            elif isinstance(w, tk.Label):
                upper = text.upper()
                fsize, weight = fs, "normal"
                fg = pal["text"]

                if any(x in upper for x in ["GENERADOR", "CONFIGURACIÓN", "CONFIGURACION", "HISTORIAL", "REPORTE"]):
                    fsize, weight = fs + 4, "bold"
                    fg = pal["text"]
                elif any(x in upper for x in ["DATOS", "ACCIONES", "INFORMACIÓN", "INFORMACION", "RESUMEN"]):
                    fsize, weight = fs + 1, "bold"
                    fg = pal["accent"]

                # A QLabel is text by default, not a card.  The old walker
                # reapplied a dark background captured at construction to
                # every label, which caused the visible dark rectangles in
                # light mode.  Badges/status cards opt in elsewhere.
                safe_conf(
                    w,
                    _qt_transparent=True,
                    foreground=fg,
                    fg=fg,
                    font=("Arial", fsize, weight),
                    activebackground="transparent",
                    activeforeground=pal["accent"],
                )

            elif isinstance(w, ttk.Label):
                upper = text.upper()
                style_name = "TLabel"
                fg = pal["text"]
                if any(x in upper for x in ["GENERADOR", "CONFIGURACIÓN", "CONFIGURACION", "HISTORIAL", "REPORTE"]):
                    style_name = "Title.TLabel"
                    bg = pal["root"]
                elif any(x in upper for x in ["DATOS", "ACCIONES", "INFORMACIÓN", "INFORMACION", "RESUMEN"]):
                    style_name = "Section.TLabel"
                    fg = pal["accent"]

                safe_conf(w, style=style_name, background=bg, foreground=fg)

            elif isinstance(w, (tk.Entry, tk.Text, tk.Listbox)):
                safe_conf(
                    w,
                    background=pal["entry"],
                    bg=pal["entry"],
                    foreground=pal["text"],
                    fg=pal["text"],
                    insertbackground=pal["text"],
                    font=("Arial", fs),
                    highlightbackground=pal["border"],
                    highlightcolor=pal["accent"],
                    selectbackground=pal["selected_bg"],
                    selectforeground=pal["selected_fg"]
                )

            elif isinstance(w, tk.Button):
                button_bg, button_fg = semantic_button_palette(w)
                safe_conf(
                    w,
                    background=button_bg,
                    bg=button_bg,
                    foreground=button_fg,
                    fg=button_fg,
                    activebackground=pal.get("accent_hover", pal["accent"]),
                    activeforeground=pal["selected_fg"],
                    font=("Arial", fs, "bold"),
                    highlightbackground=pal["border"],
                    highlightcolor=pal["accent"]
                )

            elif isinstance(w, ttk.Combobox):
                safe_conf(
                    w,
                    background=pal["entry"],
                    fieldbackground=pal["entry"],
                    foreground=pal["text"],
                    bordercolor=pal["border"],
                    selectbackground=pal["selected_bg"],
                    selectforeground=pal["selected_fg"],
                    disabled_background=pal.get("input_disabled_bg", pal["card2"]),
                )

            elif isinstance(w, (ttk.Checkbutton, ttk.Radiobutton)):
                safe_conf(
                    w,
                    _qt_transparent=True,
                    foreground=pal["text"],
                    indicator_bg=pal["entry"],
                    indicator_border=pal.get("checkbox_indicator_border", pal["border"]),
                    indicator_checked=pal.get("checkbox_checked_bg", pal["accent"]),
                    focus_border=pal.get("border_focus", pal["accent"]),
                    disabledforeground=pal.get("text_disabled", pal["muted"]),
                    disabled_background=pal.get("input_disabled_bg", pal["card2"]),
                )

            try:
                if isinstance(w, ttk.Treeview):
                    w.configure(style="Modern.Treeview")
            except Exception:
                pass

            try:
                for child in w.winfo_children():
                    walk(child)
            except Exception:
                pass

        try:
            walk(base)
        except Exception:
            pass

    def _aplicar_preferencias_en_vivo(self, ventana_actual=None, forzar_todo=False):
        """
        FASE 13: Aplica preferencias sin recorrer todos los widgets cada vez.
        - Al iniciar: root.
        - Al guardar preferencias: usar forzar_todo=True.
        """
        self._configurar_estilos_desde_preferencias()

        if forzar_todo:
            self._aplicar_preferencias_a_widgets(self.root)
            for attr in (
                "historial_win",
                "historial_sin_seguro_win",
                "reporte_win",
                "turno_win",
                "dialogo_unico_win",
                "configuracion_interna_win",
                "edicion_paciente_win",
                "salida_pendiente_win",
                "trabajos_salida_win",
            ):
                try:
                    w = getattr(self, attr, None)
                    if w is not None and w.winfo_exists():
                        self._aplicar_preferencias_a_widgets(w)
                        refresh = getattr(w, "_admission_theme_refresh", None)
                        if callable(refresh):
                            refresh()
                except Exception:
                    pass
        elif ventana_actual is not None:
            self._aplicar_preferencias_a_widgets(ventana_actual)

        if self._standalone:
            try:
                self.root.geometry(self.app_settings.get("window_size", "1280x740"))
            except Exception:
                pass

        try:
            self.root.updateGeometry()
            root_layout = self.root.layout()
            if root_layout is not None:
                root_layout.invalidate()
        except Exception:
            pass
        self._aplicar_modo_responsivo()

    def _aplicar_paridad_visual_inicio(self, profile=None):
        """
        Paridad visual del inicio sobre PySide6.

        Esta rutina solo toca geometría, fuentes y QSS de widgets ya existentes.
        No lee/escribe pacientes ni modifica turnos, Excel, PDF, permisos, sesión
        o reglas de validación.
        """
        try:
            pal = self._paleta_visual_actual()
            dark = str(pal.get("mode") or "").startswith("oscuro")
            root_bg = pal["root"]
            card_bg = pal["card"]
            entry_bg = pal["entry"]
            text_fg = pal["text"]
            muted = pal["muted"]
            border = pal["border"]
            accent = pal["accent"]
            line = pal["separator"]
            panel_border = border
            elevated_bg = pal["card2"]
            selected_bg = pal["selected_bg"]
            selected_fg = pal["selected_fg"]

            try:
                ancho_root = max(1, int(self.root.winfo_width() or 0))
                alto_root = max(1, int(self.root.winfo_height() or 0))
            except Exception:
                ancho_root, alto_root = 1365, 768
            profile = profile or self._responsive_layout_profile
            if profile is None:
                profile = resolve_admission_layout_profile(
                    ancho_root,
                    alto_root,
                    profile_preference=PROFILE_AUTO,
                    density_preference=DENSITY_AUTO,
                    text_percent=max(
                        85,
                        min(125, int(self._font_size_pref() / 11 * 100)),
                    ),
                )
                self._responsive_layout_profile = profile
            vista_grande = profile.layout_mode in ("WIDE", "NORMAL")

            form_gap_y = profile.vertical_gap
            form_gap_x = profile.horizontal_gap
            entry_h = profile.input_min_height
            title_size = profile.title_point_size
            label_size = profile.label_point_size
            value_size = profile.value_point_size
            side_width = profile.side_panel_width

            self.content_area.configure(background=root_bg)
            self.frame.configure(style="Card.TFrame", background=card_bg)
            self.form_host.configure(style="Card.TFrame", background=card_bg)
            self.quick_panel.configure(style="Card.TFrame", background=card_bg)
            self.quick_panel.setMinimumWidth(profile.side_panel_min_width)
            self.quick_panel.setMaximumWidth(
                profile.side_panel_max_width or 16777215
            )
            self.quick_panel.setSizePolicy(
                QSizePolicy.Preferred, QSizePolicy.Expanding
            )
            if side_width:
                self.quick_panel.configure(width=side_width)
            self.quick_panel.setStyleSheet(f"background-color: {card_bg};")
            if getattr(self, "quick_separator", None) is not None:
                self.quick_separator.setStyleSheet(f"background-color: {panel_border}; border: none;")
                try:
                    self.quick_separator.setMinimumWidth(2)
                    self.quick_separator.setMaximumWidth(2)
                except Exception:
                    pass

            form_layout = getattr(self.frame, "_compat_layout", None)
            if form_layout is not None:
                try:
                    form_layout.setAlignment(Qt.AlignTop)
                    form_layout.setVerticalSpacing(form_gap_y)
                    form_layout.setHorizontalSpacing(form_gap_x)
                    form_layout.setContentsMargins(
                        profile.form_padding,
                        max(5, profile.form_padding - 2),
                        profile.form_padding,
                        max(5, profile.form_padding - 2),
                    )
                except Exception:
                    pass
            quick_layout = getattr(self.quick_panel, "_compat_layout", None)
            if quick_layout is not None:
                try:
                    quick_layout.setAlignment(Qt.AlignTop)
                    quick_layout.setSpacing(max(4, form_gap_y))
                    quick_layout.setContentsMargins(
                        max(9, profile.form_padding),
                        max(5, profile.vertical_gap),
                        max(7, profile.form_padding - 2),
                        max(5, profile.vertical_gap),
                    )
                except Exception:
                    pass

            for col, weight in enumerate((3, 2, 2, 3, 2, 2)):
                try:
                    self.frame.columnconfigure(col, weight=weight)
                except Exception:
                    pass

            try:
                label_cls = self.lbl_nombre.__class__
                for lab in self.frame.findChildren(label_cls):
                    try:
                        lab.configure(_qt_transparent=True)
                    except Exception:
                        pass
            except Exception:
                pass

            field_labels = (
                self.lbl_nombre, self.lbl_sexo, self.lbl_edad, self.lbl_cedula,
                self.lbl_telefono, self.lbl_direccion, self.lbl_nacionalidad,
                self.lbl_ars, self.lbl_nss,
            )
            for lab in field_labels:
                try:
                    lab.configure(
                        _qt_transparent=True, foreground=text_fg,
                        font=("Segoe UI", label_size, "bold")
                    )
                    lab.setWordWrap(False)
                    lab.setSizePolicy(QSizePolicy.MinimumExpanding, QSizePolicy.Preferred)
                    required_width = lab.fontMetrics().horizontalAdvance(
                        lab.text()
                    ) + 8
                    form_width = max(
                        240,
                        profile.available_width
                        - profile.side_panel_width
                        - (profile.outer_margin + profile.form_padding) * 2,
                    )
                    label_limit = form_width if not profile.two_columns else form_width // 2
                    lab.setMinimumWidth(min(required_width, label_limit))
                except Exception:
                    pass
            if not profile.two_columns:
                self.lbl_nss.setWordWrap(True)

            self.section_patient_label.configure(
                _qt_transparent=True, foreground=accent,
                font=("Segoe UI", 12 if vista_grande else 11, "bold")
            )
            self.ars_detectado_label.configure(
                _qt_transparent=True, font=("Segoe UI", 10 if vista_grande else 9, "bold"),
                foreground="#6DE4A7" if dark else accent
            )
            self.nss_detectado_label.configure(
                _qt_transparent=True, font=("Segoe UI", 10 if vista_grande else 9, "bold"),
                foreground="#F2B35A" if dark else accent
            )

            entry_qss = (
                f"QLineEdit{{background:{entry_bg};color:{text_fg};border:1px solid {border};"
                f"border-radius:2px;padding:5px 8px;font-family:'Segoe UI';font-size:{value_size}pt;}}"
                f"QLineEdit:focus{{border:1px solid {accent};background:{entry_bg};}}"
                f"QLineEdit:disabled{{color:{muted};background:{elevated_bg};}}"
            )
            for widget in (
                self.entry_nombre, self.entry_edad, self.entry_cedula,
                self.entry_telefono, self.entry_direccion, self.entry_nacionalidad,
                self.entry_ars, self.entry_nss,
            ):
                widget.setStyleSheet(entry_qss)
                try:
                    widget.setMinimumHeight(entry_h)
                    widget.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Fixed)
                except Exception:
                    pass
            self.ensure_embedded_entry_geometry()
            self._schedule_embedded_entry_geometry_sync()

            self.combo_unidad.setStyleSheet(
                f"QComboBox{{background:{entry_bg};color:{text_fg};border:1px solid {border};"
                f"border-radius:2px;padding:5px 32px 5px 8px;font-family:'Segoe UI';font-size:{value_size}pt;}}"
                f"QComboBox:focus{{border:1px solid {accent};}}"
                "QComboBox::drop-down{border:0;width:30px;subcontrol-origin:padding;subcontrol-position:top right;}"
                f"QComboBox QAbstractItemView{{background:{entry_bg};color:{text_fg};border:1px solid {border};"
                f"selection-background-color:{selected_bg};selection-color:{selected_fg};}}"
            )
            self.combo_unidad.setMinimumHeight(entry_h)
            self.combo_unidad.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Fixed)

            choice_style = {
                "_qt_transparent": True,
                "foreground": text_fg,
                "font": ("Segoe UI", value_size),
                "indicator_bg": entry_bg,
                "indicator_border": pal.get("checkbox_indicator_border", border),
                "indicator_checked": pal.get("checkbox_checked_bg", accent),
                "radio_checkmark": True,
                "focus_border": pal.get("border_focus", accent),
                "disabledforeground": pal.get("text_disabled", muted),
                "disabled_background": pal.get("input_disabled_bg", elevated_bg),
            }
            for widget in (self.lbl_sexo_m, self.lbl_sexo_f, self.check_urgencia, self.check_embarazada):
                # qt_compat owns the full-control/indicator boundary.  Do not
                # replace it with a second compound QSS here: doing so made
                # Windows compose a native outer frame around the label.
                widget.configure(**choice_style)
            self.sexo_frame.configure(background=card_bg)
            self.form_buttons.configure(background=card_bg)

            self.ars_suggestions.setStyleSheet(
                f"QListWidget{{background:{entry_bg};color:{text_fg};border:1px solid {border};"
                "font-family:'Segoe UI';font-size:10pt;padding:1px;}"
                f"QListWidget::item{{padding:3px 5px;}}"
                f"QListWidget::item:selected{{background:{selected_bg};color:{selected_fg};}}"
            )
            try:
                self.ars_suggestions.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Fixed)
                self.ars_suggestions._sync_compat_height()
            except Exception:
                pass

            for separator in (getattr(self, "form_actions_separator", None),):
                if separator is not None:
                    separator.setStyleSheet(f"background:{line};border:none;")
                    try:
                        separator.setMaximumHeight(1)
                    except Exception:
                        pass

            # Encabezado estético y compacto.
            self.title_lbl.setText("GENERADOR DE FORMULARIOS DE\nEMERGENCIA")
            self.title_lbl.configure(_qt_transparent=True, font=("Segoe UI", title_size, "bold"), foreground=text_fg)
            try:
                # El salto de línea es deliberado. Se desactiva wordWrap para
                # impedir que Qt cree una tercera línea y recorte EMERGENCIA.
                self.title_lbl.setWordWrap(False)
                self.title_lbl.setAlignment(Qt.AlignLeft | Qt.AlignVCenter)
            except Exception:
                pass
            title_min_width = max(
                280,
                min(460, int(profile.available_width * 0.34)),
            )
            self.title_lbl.setMinimumWidth(title_min_width)
            self.title_lbl.setMaximumWidth(max(title_min_width, 560))
            self.title_lbl.setMinimumHeight(58 if vista_grande else 52)
            self.subtitle_lbl.configure(
                _qt_transparent=True,
                font=("Segoe UI", max(9, value_size), "normal"),
                foreground=muted,
            )

            # Botones: QSS deliberadamente simple para evitar warnings del parser
            # Qt en Windows. Tipografía/tamaño se aplican con QFont/geometry.
            def _set_button_font(widget, size=10, bold=True):
                try:
                    f = QFont("Segoe UI", int(size))
                    f.setBold(bool(bold))
                    widget.setFont(f)
                except Exception:
                    pass

            header_button_bg = pal.get("button_secondary_bg", elevated_bg)
            header_button_text = pal.get("button_secondary_text", text_fg)
            header_btn_qss = (
                f"QPushButton {{ background-color: {header_button_bg}; color: {header_button_text}; "
                f"border: 1px solid {border}; border-radius: 3px; padding: 8px 14px; }} "
                f"QPushButton:hover {{ background-color: {pal.get('button_secondary_hover', pal['heading'])}; border-color: {accent}; }} "
                f"QPushButton:pressed {{ background-color: {selected_bg}; color: {selected_fg}; }}"
                f"QPushButton:disabled {{ background-color: {pal.get('input_disabled_bg', elevated_bg)}; color: {pal.get('text_disabled', muted)}; border-color: {border}; }}"
            )
            if getattr(self, "boton_cambiar_turno", None) is not None:
                _set_button_font(self.boton_cambiar_turno, 10, True)
                self.boton_cambiar_turno.setStyleSheet(header_btn_qss)
                self.boton_cambiar_turno.setText("Cambiar Turno")
                try:
                    self.boton_cambiar_turno.setMinimumWidth(
                        142 if profile.compact_labels else 170
                    )
                    self.boton_cambiar_turno.setMinimumHeight(profile.button_height)
                    apply_admission_button_icon(
                        self.boton_cambiar_turno,
                        "turno.svg",
                        pal,
                        "secondary",
                        size=16,
                    )
                except Exception:
                    pass
            if getattr(self, "actions_menu_button", None) is not None:
                _set_button_font(self.actions_menu_button, 10, True)
                self.actions_menu_button.setStyleSheet(header_btn_qss)
                try:
                    self.actions_menu_button.setMinimumWidth(
                        92 if profile.compact_labels else 112
                    )
                    self.actions_menu_button.setMinimumHeight(profile.button_height)
                    apply_admission_button_icon(
                        self.actions_menu_button,
                        "menu.svg",
                        pal,
                        "secondary",
                        size=16,
                    )
                except Exception:
                    pass

            history_bg = pal.get("button_primary_bg", pal["info"])
            history_text = pal.get("button_primary_text", pal["button_fg"])
            history_qss = (
                f"QPushButton {{ background-color: {history_bg}; color: {history_text}; "
                f"border: 1px solid {history_bg}; border-radius: 2px; padding: 9px 20px; }} "
                f"QPushButton:hover {{ background-color: {pal.get('button_primary_hover', pal['accent'])}; border-color: {accent}; color: {history_text}; }} "
                f"QPushButton:pressed {{ background-color: {selected_bg}; color: {selected_fg}; }}"
                f"QPushButton:disabled {{ background-color: {pal.get('input_disabled_bg', elevated_bg)}; color: {pal.get('text_disabled', muted)}; border-color: {border}; }}"
            )
            _set_button_font(self.boton_historial, 10, True)
            self.boton_historial.setStyleSheet(history_qss)
            self.boton_historial.setText("Historial")
            try:
                self.boton_historial.setMinimumHeight(profile.button_height)
                apply_admission_button_icon(
                    self.boton_historial,
                    "history.svg",
                    pal,
                    "primary",
                    size=16,
                )
            except Exception:
                pass

            clear_bg = pal.get("button_danger_bg", pal["danger"])
            clear_text = pal.get("button_danger_text", pal["button_fg"])
            pdf_bg = pal.get("button_primary_bg", accent)
            pdf_text = pal.get("button_primary_text", pal["button_fg"])
            clear_qss = (
                f"QPushButton {{ background-color: {clear_bg}; color: {clear_text}; "
                f"border: 1px solid {clear_bg}; border-radius: 2px; padding: 9px 20px; }} "
                f"QPushButton:hover {{ background-color: {pal.get('button_danger_hover', pal['danger'])}; border-color: {clear_bg}; color: {clear_text}; }} "
                f"QPushButton:pressed {{ background-color: {pal['danger_bg']}; color: {pal['danger_text']}; }}"
                f"QPushButton:disabled {{ background-color: {pal.get('input_disabled_bg', elevated_bg)}; color: {pal.get('text_disabled', muted)}; border-color: {border}; }}"
            )
            pdf_qss = (
                f"QPushButton {{ background-color: {pdf_bg}; color: {pdf_text}; "
                f"border: 1px solid {pdf_bg}; border-radius: 2px; padding: 9px 20px; }} "
                f"QPushButton:hover {{ background-color: {pal.get('button_primary_hover', accent)}; border-color: {accent}; color: {pdf_text}; }} "
                f"QPushButton:pressed {{ background-color: {selected_bg}; color: {selected_fg}; }} "
                f"QPushButton:disabled {{ color: {pal.get('text_disabled', muted)}; border-color: {border}; background-color: {pal.get('input_disabled_bg', elevated_bg)}; }}"
            )
            _set_button_font(self.boton_limpiar, 10, True)
            _set_button_font(self.boton_generar_pdf, 10, True)
            self.boton_limpiar.setStyleSheet(clear_qss)
            self.boton_generar_pdf.setStyleSheet(pdf_qss)
            self.boton_limpiar.setText("Limpiar")
            self.boton_generar_pdf.setText("Generar PDF")
            try:
                apply_admission_button_icon(
                    self.boton_limpiar,
                    "clear.svg",
                    pal,
                    "danger",
                    size=18,
                )
                apply_admission_button_icon(
                    self.boton_generar_pdf,
                    "pdf.svg",
                    pal,
                    "primary",
                    size=16,
                )
            except Exception:
                pass

            roles = {
                "report": "primary",
                "excel": "success",
                "uninsured": "warning",
                "edit": "primary",
                "config": "secondary",
            }
            texts = {
                "report": "Reporte estadístico   ›",
                "excel": "Abrir Listado en Excel   ›",
                "uninsured": "Ver Historial sin Seguro   ›",
                "edit": "Editar paciente   ›",
                "config": "Configuración interna   ›",
            }
            icon_map = {
                "report": "report.svg",
                "excel": "excel.svg",
                "uninsured": "uninsured.svg",
                "edit": "edit.svg",
                "config": "config.svg",
            }
            for role, btn in getattr(self, "quick_action_buttons", {}).items():
                button_role = roles.get(role, "secondary")
                color = pal.get(f"button_{button_role}_bg", elevated_bg)
                hover_color = pal.get(f"button_{button_role}_hover", pal["heading"])
                button_text = pal.get(f"button_{button_role}_text", text_fg)
                btn.setText(texts.get(role, btn.text()))
                _set_button_font(btn, value_size, True)
                try:
                    btn.setMinimumHeight(profile.button_height)
                except Exception:
                    pass
                btn.setStyleSheet(
                    f"QPushButton {{ background-color: {color}; color: {button_text}; "
                    f"border: 1px solid {color}; border-radius: 2px; padding: 10px 14px; }} "
                    f"QPushButton:hover {{ background-color: {hover_color}; border-color: {color}; color: {button_text}; }} "
                    f"QPushButton:pressed {{ background-color: {selected_bg}; color: {selected_fg}; }}"
                    f"QPushButton:disabled {{ background-color: {pal.get('input_disabled_bg', elevated_bg)}; color: {pal.get('text_disabled', muted)}; border-color: {border}; }}"
                )
                try:
                    apply_admission_button_icon(
                        btn,
                        icon_map.get(role, "report.svg"),
                        pal,
                        button_role,
                        size=16,
                    )
                except Exception:
                    pass

            # Secciones laterales resaltadas y claramente separadas.
            if getattr(self, "quick_actions_title", None) is not None:
                self.quick_actions_title.configure(
                    foreground=accent, font=("Segoe UI", 13 if vista_grande else 12, "bold"), _qt_transparent=True
                )
            if getattr(self, "quick_info_area", None) is not None:
                self.quick_info_area.setStyleSheet(
                    f"background-color: {card_bg}; border-top: 1px solid {panel_border};"
                )
            for attr in ("quick_info_title", "quick_summary_title"):
                lab = getattr(self, attr, None)
                if lab is not None:
                    lab.configure(foreground=muted, font=("Segoe UI", 11 if vista_grande else 10, "bold"), _qt_transparent=True)
                    try:
                        lab.setAlignment(Qt.AlignLeft | Qt.AlignVCenter)
                        lab.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Preferred)
                    except Exception:
                        pass
            for attr in ("quick_representante_label", "quick_turno_label", "quick_summary_label"):
                lab = getattr(self, attr, None)
                if lab is not None:
                    lab.configure(foreground=text_fg, font=("Segoe UI", 11 if vista_grande else 10, "normal"), _qt_transparent=True)
                    try:
                        lab.setAlignment(Qt.AlignLeft | Qt.AlignTop)
                        lab.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Preferred)
                        lab.setMinimumWidth(0)
                        if profile.side_panel_max_width:
                            lab.setMaximumWidth(profile.side_panel_max_width)
                    except Exception:
                        pass

            if getattr(self, "quick_total_label", None) is not None:
                try:
                    self.quick_total_label.configure(
                        foreground=accent,
                        font=("Segoe UI", 13 if vista_grande else 12, "bold"),
                        _qt_transparent=True,
                    )
                    self.quick_total_label.setAlignment(Qt.AlignLeft | Qt.AlignVCenter)
                    self.quick_total_label.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Preferred)
                    self.quick_total_label.setMinimumWidth(0)
                    if profile.side_panel_max_width:
                        self.quick_total_label.setMaximumWidth(
                            profile.side_panel_max_width
                        )
                except Exception:
                    pass

            try:
                self.quick_representante_label.setAlignment(Qt.AlignLeft | Qt.AlignTop)
                self.quick_turno_label.setAlignment(Qt.AlignLeft | Qt.AlignTop)
                self.quick_summary_label.setWordWrap(True)
                self.quick_summary_label.setMinimumWidth(0)
                if profile.side_panel_max_width:
                    self.quick_summary_label.setMaximumWidth(
                        profile.side_panel_max_width
                    )
                self.quick_summary_label.setAlignment(Qt.AlignLeft | Qt.AlignTop)
            except Exception:
                pass

            self.shortcuts_label.configure(_qt_transparent=True, font=("Segoe UI", 10 if vista_grande else 9, "normal"), foreground=muted)
            self.connection_label.configure(_qt_transparent=True, font=("Segoe UI", 10 if vista_grande else 9, "bold"), foreground=pal["success"])
            self.status_label.configure(_qt_transparent=True, font=("Segoe UI", 10 if vista_grande else 9, "bold"), foreground=accent)
        except Exception:
            pass

    def apply_embedded_responsive_layout(
        self, available_width, available_height, host_snapshot=None
    ):
        """Resolve V15 geometry from the embedded viewport and host preference."""
        if not self._host_theme_controlled:
            return self._aplicar_modo_responsivo()
        self._host_layout_snapshot = host_snapshot
        profile_preference = getattr(
            host_snapshot, "configured_profile", None
        ) or getattr(host_snapshot, "applied_profile", PROFILE_AUTO)
        density_preference = getattr(
            host_snapshot, "configured_density", None
        ) or getattr(host_snapshot, "density", DENSITY_AUTO)
        text_percent = int(getattr(host_snapshot, "text_percent", 100) or 100)
        logical_dpi = float(
            getattr(host_snapshot, "logical_dpi", 96.0) or 96.0
        )
        self._responsive_layout_profile = resolve_admission_layout_profile(
            available_width,
            available_height,
            logical_dpi=logical_dpi,
            profile_preference=profile_preference,
            density_preference=density_preference,
            text_percent=text_percent,
        )
        profile = self._responsive_layout_profile
        signature = (
            profile.available_width,
            profile.available_height,
            round(profile.dpi_scale, 2),
            profile.layout_mode,
            profile.density,
            profile.text_percent,
        )
        if signature != self._last_responsive_log_signature:
            APP_LOG.info(
                "ADMISSION_RESPONSIVE_PROFILE viewport=%sx%s dpi=%.2f "
                "profile=%s density=%s font_scale=%.2f input_height=%s",
                profile.available_width,
                profile.available_height,
                profile.dpi_scale,
                profile.layout_mode,
                profile.density,
                profile.font_scale,
                profile.input_min_height,
            )
            self._last_responsive_log_signature = signature
        return self._aplicar_modo_responsivo()

    def _programar_modo_responsivo(self, event=None):
        if self._shutdown_complete or not self._admission_visible:
            return
        if event is not None and event.widget is not self.root:
            return
        previo = getattr(self, "_responsive_after_id", None)
        if previo:
            try:
                self.root.after_cancel(previo)
            except Exception:
                pass
        self._responsive_after_id = self.root.after(80, self._aplicar_modo_responsivo)

    def _standalone_responsive_preferences(self):
        if bool(self.app_settings.get("small_screen_mode", False)):
            return "MUY_COMPACTO", "MUY_COMPACTA"
        if bool(self.app_settings.get("compact_mode", False)):
            return "COMPACTO", "COMPACTA"
        return PROFILE_AUTO, DENSITY_AUTO

    def _resolve_current_responsive_profile(self, width, height):
        current = self._responsive_layout_profile
        if (
            current is not None
            and current.available_width == width
            and current.available_height == height
        ):
            return current
        host_snapshot = self._host_layout_snapshot
        if self._host_theme_controlled and host_snapshot is not None:
            profile_preference = getattr(
                host_snapshot, "configured_profile", None
            ) or getattr(host_snapshot, "applied_profile", PROFILE_AUTO)
            density_preference = getattr(
                host_snapshot, "configured_density", None
            ) or getattr(host_snapshot, "density", DENSITY_AUTO)
            text_percent = int(
                getattr(host_snapshot, "text_percent", 100) or 100
            )
            logical_dpi = float(
                getattr(host_snapshot, "logical_dpi", 96.0) or 96.0
            )
        else:
            profile_preference, density_preference = (
                self._standalone_responsive_preferences()
            )
            text_percent = max(
                85,
                min(125, int(self._font_size_pref() / 11 * 100)),
            )
            logical_dpi = 96.0
        return resolve_admission_layout_profile(
            width,
            height,
            logical_dpi=logical_dpi,
            profile_preference=profile_preference,
            density_preference=density_preference,
            text_percent=text_percent,
        )

    @staticmethod
    def _set_grid_visibility(widget, visible):
        widget.grid() if visible else widget.grid_remove()

    def _aplicar_modo_responsivo(self):
        try:
            self._responsive_after_id = None
            if self._shutdown_complete or not self._admission_visible:
                return
            ancho = max(1, int(self.root.winfo_width()))
            alto = max(1, int(self.root.winfo_height()))
            forzar_pequeno = bool(self.app_settings.get("small_screen_mode", False))
            profile = self._resolve_current_responsive_profile(ancho, alto)
            self._responsive_layout_profile = profile
            mostrar_panel = (
                bool(self.app_settings.get("show_side_panel", True))
                and profile.show_side_panel
                and not forzar_pequeno
            )

            self._set_grid_visibility(self.quick_panel, mostrar_panel)
            self._set_grid_visibility(
                self.info_header,
                profile.show_header_info and not forzar_pequeno,
            )

            self.main.configure(padding=profile.outer_margin)
            self.frame.configure(padding=profile.form_padding)
            self.form_host.grid_configure(
                padx=(0, profile.horizontal_gap if mostrar_panel else 0)
            )
            self.title_lbl.configure(wraplength=0)
            self._configurar_columnas_formulario(
                una_columna=not profile.two_columns,
                profile=profile,
            )

            pal = self._paleta_visual_actual()
            self.form_canvas.configure(background=pal["card"])
            padding_boton = (
                max(6, profile.horizontal_gap),
                max(3, profile.vertical_gap - 1),
            )
            self.style.configure("TButton", padding=padding_boton)
            self._aplicar_paridad_visual_inicio(profile)
        except (AttributeError, tk.TclError):
            pass

    def _configurar_columnas_formulario(self, una_columna=False, profile=None):
        layout_signature = (
            bool(una_columna),
            getattr(profile, "horizontal_gap", None),
            getattr(profile, "vertical_gap", None),
        )
        if getattr(self, "_formulario_layout_signature", None) == layout_signature:
            return
        self._formulario_layout_signature = layout_signature
        self._formulario_una_columna = bool(una_columna)
        if una_columna:
            layout = (
                (self.lbl_nombre, 2, 0, 6, "w", (4, 4)),
                (self.entry_nombre, 3, 0, 6, "ew", (4, 4)),
                (self.lbl_sexo, 4, 0, 6, "w", (4, 4)),
                (self.sexo_frame, 5, 0, 6, "w", (4, 4)),
                (self.lbl_edad, 6, 0, 6, "w", (4, 4)),
                (self.entry_edad, 7, 0, 4, "ew", (4, 8)),
                (self.combo_unidad, 7, 4, 2, "ew", (0, 4)),
                (self.check_urgencia, 8, 0, 6, "w", (4, 4)),
                (self.lbl_cedula, 9, 0, 6, "w", (4, 4)),
                (self.entry_cedula, 10, 0, 6, "ew", (4, 4)),
                (self.lbl_telefono, 11, 0, 6, "w", (4, 4)),
                (self.entry_telefono, 12, 0, 6, "ew", (4, 4)),
                (self.lbl_direccion, 13, 0, 6, "w", (4, 4)),
                (self.entry_direccion, 14, 0, 6, "ew", (4, 4)),
                (self.lbl_nacionalidad, 15, 0, 6, "w", (4, 4)),
                (self.entry_nacionalidad, 16, 0, 6, "ew", (4, 4)),
                (self.lbl_ars, 17, 0, 6, "w", (4, 4)),
                (self.entry_ars, 18, 0, 6, "ew", (4, 4)),
                (self.ars_detectado_label, 19, 0, 6, "w", (4, 4)),
                (self.ars_suggestions, 20, 0, 6, "ew", (4, 4)),
                (self.lbl_nss, 21, 0, 4, "w", (4, 4)),
                (self.nss_detectado_label, 21, 4, 2, "w", (4, 4)),
                (self.entry_nss, 22, 0, 6, "ew", (4, 4)),
                (self.form_actions_separator, 23, 0, 6, "ew", (4, 4)),
                (self.form_buttons, 24, 0, 6, "e", (4, 4)),
            )
        else:
            layout = (
                (self.lbl_nombre, 2, 0, 3, "w", (4, 10)),
                (self.entry_nombre, 3, 0, 3, "ew", (4, 24)),
                (self.lbl_sexo, 2, 3, 2, "w", (4, 10)),
                (self.sexo_frame, 3, 3, 3, "w", (4, 4)),
                (self.lbl_edad, 4, 0, 1, "w", (4, 10)),
                (self.entry_edad, 5, 0, 1, "ew", (4, 8)),
                (self.combo_unidad, 5, 1, 2, "ew", (0, 24)),
                (self.check_urgencia, 5, 3, 3, "w", (4, 4)),
                (self.lbl_cedula, 6, 0, 3, "w", (4, 10)),
                (self.entry_cedula, 7, 0, 3, "ew", (4, 24)),
                (self.lbl_telefono, 6, 3, 3, "w", (4, 10)),
                (self.entry_telefono, 7, 3, 3, "ew", (4, 4)),
                (self.lbl_direccion, 8, 0, 6, "w", (4, 10)),
                (self.entry_direccion, 9, 0, 6, "ew", (4, 4)),
                (self.lbl_nacionalidad, 10, 0, 3, "w", (4, 10)),
                (self.entry_nacionalidad, 11, 0, 3, "ew", (4, 24)),
                (self.lbl_ars, 10, 3, 3, "w", (4, 10)),
                (self.entry_ars, 11, 3, 3, "ew", (4, 4)),
                (self.ars_detectado_label, 12, 3, 3, "w", (4, 4)),
                (self.ars_suggestions, 13, 3, 3, "ew", (4, 4)),
                (self.lbl_nss, 13, 0, 2, "w", (4, 10)),
                (self.nss_detectado_label, 13, 2, 1, "w", (4, 24)),
                (self.entry_nss, 14, 0, 3, "ew", (4, 24)),
                (self.form_actions_separator, 15, 0, 6, "ew", (4, 4)),
                (self.form_buttons, 16, 0, 6, "e", (4, 4)),
            )
        suggestions_visible = self.ars_suggestions.winfo_ismapped()
        for widget, row, column, span, sticky, padx in layout:
            widget.grid_configure(
                row=row,
                column=column,
                columnspan=span,
                sticky=sticky,
                padx=padx,
            )
        if profile is not None:
            gap_x = max(6, int(profile.horizontal_gap))
            gap_y = max(3, int(profile.vertical_gap))
            left_entries = (
                self.entry_nombre,
                self.entry_edad,
                self.combo_unidad,
                self.entry_cedula,
                self.entry_nacionalidad,
                self.entry_nss,
            )
            for entry in left_entries:
                entry.grid_configure(padx=(4, gap_x), pady=(0, gap_y))
            for entry in (
                self.entry_telefono,
                self.entry_direccion,
                self.entry_ars,
            ):
                entry.grid_configure(padx=(4, 4), pady=(0, gap_y))
            for label in (
                self.lbl_nombre,
                self.lbl_sexo,
                self.lbl_edad,
                self.lbl_cedula,
                self.lbl_telefono,
                self.lbl_direccion,
                self.lbl_nacionalidad,
                self.lbl_ars,
                self.lbl_nss,
            ):
                label.grid_configure(pady=(1, 2))
        if not suggestions_visible:
            self.ars_suggestions.grid_remove()


    def _post_to_ui(self, callback):
        if self._shutdown_complete or not callable(callback):
            return None
        try:
            return self.root.after(
                0,
                lambda: None if self._shutdown_complete else callback(),
            )
        except Exception:
            return None

    def _start_worker(self, target, *, name=None, kwargs=None):
        if self._shutdown_complete:
            return None

        def runner():
            try:
                target(**(kwargs or {}))
            finally:
                with self._worker_lock:
                    self._worker_threads.discard(threading.current_thread())

        hilo = threading.Thread(target=runner, name=name, daemon=True)
        with self._worker_lock:
            self._worker_threads.add(hilo)
        hilo.start()
        return hilo

    def _ejecutar_en_segundo_plano(self, mensaje, funcion, al_terminar=None, al_error=None):
        def worker():
            try:
                resultado = funcion()
                if al_terminar:
                    self._post_to_ui(lambda: al_terminar(resultado))
            except Exception as e:
                APP_LOG.exception("Falló una tarea de interfaz ejecutada en segundo plano")
                if al_error:
                    self._post_to_ui(lambda err=e: al_error(err))
                else:
                    self._post_to_ui(
                        lambda err=e: messagebox.showerror("Error", str(err), parent=self.root)
                    )

        try:
            self.set_status(mensaje, "process")
        except Exception:
            pass

        return self._start_worker(worker, name="admission-background")

    def _retry_excel_export_jobs(self, delay_ms=0):
        """Reintento ligero; nunca bloquea la GUI ni repite la transición."""
        if self._shutdown_complete:
            return
        if delay_ms and delay_ms > 0:
            try:
                if self._excel_export_after_id:
                    self.root.after_cancel(self._excel_export_after_id)
                self._excel_export_after_id = self.root.after(
                    int(delay_ms), self._retry_excel_export_jobs
                )
            except Exception:
                self._excel_export_after_id = None
            return
        self._excel_export_after_id = None
        if self._excel_export_running:
            return
        try:
            if pending_excel_export_jobs() <= 0:
                return
        except Exception:
            APP_LOG.exception("No se pudo consultar la cola post-commit de Excel")
            return

        self._excel_export_running = True

        def finished(result):
            self._excel_export_running = False
            pending = int((result or {}).get("pending") or 0)
            completed = int((result or {}).get("completed") or 0)
            errors = int((result or {}).get("errors") or 0)
            if pending:
                self.set_status(
                    "Turno aplicado; actualización de Excel pendiente porque el archivo está abierto.",
                    "warning",
                )
                self._retry_excel_export_jobs(30000)
            elif errors:
                self.set_status(
                    "Turno aplicado; el efecto de Excel quedó registrado con error técnico.",
                    "warning",
                )
            elif completed:
                self.set_status("Listado de Excel actualizado después del cambio de turno.", "ok")

        def failed(exc):
            self._excel_export_running = False
            APP_LOG.error(
                "Falló el procesador post-commit de Excel: %s",
                type(exc).__name__,
            )
            self.set_status(
                "Turno aplicado; la actualización de Excel sigue pendiente.",
                "warning",
            )
            self._retry_excel_export_jobs(30000)

        self._ejecutar_en_segundo_plano(
            "Actualizando Excel en segundo plano...",
            lambda: process_excel_export_jobs(self.db),
            al_terminar=finished,
            al_error=failed,
        )

    def _run_turn_post_commit_effects(
        self, turno_saliente, turno_cfg_nuevo, momento_cambio
    ):
        """Efectos derivados. Ninguno modifica el resultado del cambio central."""
        warnings = []
        outgoing_has_patients = True
        transition = getattr(self.db, "last_transition_result", None)
        outgoing_turn_id = getattr(transition, "old_turn_id", None)
        outgoing_source_id = self._snapshot_operacional_integrado().get(
            "operational_source_id"
        )
        if turno_saliente:
            try:
                outgoing_summary = construir_resumen_turno(
                    self.db,
                    turno_saliente,
                    fin_override=momento_cambio,
                    turn_id=outgoing_turn_id,
                    operational_source_id=outgoing_source_id,
                )
                outgoing_has_patients = (
                    reportable_patient_count(outgoing_summary) > 0
                )
            except Exception:
                outgoing_has_patients = True
                APP_LOG.exception(
                    "No se pudo verificar si el turno saliente contiene pacientes"
                )
        if turno_saliente and bool(self.app_settings.get("turnos_generate_report", True)):
            try:
                self._generar_y_abrir_reporte_turno(
                    turno_saliente,
                    fin_corte=momento_cambio,
                    turn_id=outgoing_turn_id,
                    operational_source_id=outgoing_source_id,
                )
            except Exception as exc:
                warnings.append("REPORTE_TURNO")
                APP_LOG.exception("Efecto post-commit: reporte de turno")
        if (
            turno_saliente
            and outgoing_has_patients
            and bool(self.app_settings.get("turnos_save_excel_copy", True))
        ):
            try:
                guardar_copia_excel_turno(turno_saliente, EXCEL_PATH)
            except Exception:
                warnings.append("COPIA_EXCEL_SALIENTE")
                APP_LOG.exception("Efecto post-commit: copia del Excel saliente")
        if (
            outgoing_has_patients
            and
            self.app_settings.get("auto_print", True)
            and bool(self.app_settings.get("print_auto_excel_turno", True))
        ):
            try:
                if excel_canonical_in_use(EXCEL_PATH):
                    warnings.append("IMPRESION_EXCEL_ARCHIVO_EN_USO")
                elif not imprimir_excel(
                    EXCEL_PATH,
                    copias=max(
                        1, int(self.app_settings.get("print_copies_excel", 2) or 2)
                    ),
                    permitir_reintento=False,
                ):
                    warnings.append("IMPRESION_EXCEL")
            except Exception:
                warnings.append("IMPRESION_EXCEL")
                APP_LOG.exception("Efecto post-commit: impresión del Excel")
        if bool(self.app_settings.get("turnos_open_archive_folder", False)):
            try:
                carpeta = (
                    carpeta_archivo_turno(turno_saliente)
                    if turno_saliente else ARCHIVO_DIARIO_DIR
                )
                if platform.system() == "Windows":
                    os.startfile(carpeta)
                elif platform.system() == "Darwin":
                    subprocess.run(["open", carpeta], check=False)
                else:
                    subprocess.run(["xdg-open", carpeta], check=False)
            except Exception:
                warnings.append("OPEN_ARCHIVE_FOLDER")
                APP_LOG.exception("Efecto post-commit: abrir carpeta de archivo")
        self._post_to_ui(lambda: self._retry_excel_export_jobs())
        return tuple(warnings)

    def _actor_actual(self):
        if self.session_context.audit_actor:
            return self.session_context.audit_actor
        turno = cargar_turno_config(permitir_vencido=True) or {}
        return limpiar_nombre_representante(turno.get("representante", "")) or os.environ.get("USERNAME", "OPERADOR")

    def _snapshot_operacional_integrado(self):
        provider = getattr(self.db, "get_operational_station_snapshot", None)
        if not callable(provider):
            return {}
        try:
            return dict(provider() or {})
        except Exception:
            APP_LOG.exception("No se pudo leer el snapshot operacional en memoria")
            return {}

    @staticmethod
    def _revision_snapshot_operacional(snapshot):
        return (
            int(snapshot.get("operational_revision") or 0),
            int(snapshot.get("generation") or 0),
            int(snapshot.get("lease_generation") or 0),
        )

    def apply_operational_snapshot(self, snapshot):
        """Apply the already-fetched central state to all operational widgets.

        The method only receives an in-memory snapshot from the hybrid runtime;
        it performs no PostgreSQL, SQLite, or JSON I/O on the GUI thread.
        """
        state = dict(snapshot or {})
        session_id = str(state.get("operational_session_id") or "").strip()
        if not session_id:
            return False

        current = dict(getattr(self, "current_shift_context", {}) or {})
        current_session_id = str(current.get("operational_session_id") or "").strip()
        if (
            current_session_id == session_id
            and self._revision_snapshot_operacional(state)
            < self._revision_snapshot_operacional(current)
        ):
            APP_LOG.info(
                "OP_SNAPSHOT_STALE_IGNORED ui=admission old_revision=%s new_revision=%s",
                current.get("operational_revision", 0),
                state.get("operational_revision", 0),
            )
            return False

        current.update(
            {
                "operational_session_id": session_id,
                "operational_source_id": state.get("operational_source_id") or "",
                "owner_user_id": state.get("active_user_id") or "",
                "owner_username": state.get("active_username") or "",
                "representative_display_name": (
                    state.get("active_user_display_name")
                    or state.get("active_username")
                    or ""
                ),
                "turn_id": state.get("turn_id"),
                "turn_code": state.get("turn_code") or "",
                "turn_started_at": state.get("turn_started_at") or "",
                "turn_ends_at": state.get("turn_ends_at") or "",
                "generation": state.get("generation") or 0,
                "operational_revision": state.get("operational_revision") or 0,
                "primary_device_id": state.get("primary_device_id") or "",
                "primary_login_session_id": (
                    state.get("primary_login_session_id") or ""
                ),
                "lease_generation": state.get("lease_generation") or 0,
            }
        )
        self.current_shift_context = current
        context = getattr(self, "context", None)
        if context is not None:
            context.current_shift = dict(current)
        APP_LOG.info(
            "OP_SNAPSHOT_UI_APPLIED ui=admission revision=%s turn_id=%s representative_id=%s",
            current["operational_revision"],
            current["turn_id"],
            current["owner_user_id"],
        )
        self._actualizar_turno_visual_en_vivo()
        return True

    def _adoptar_snapshot_remoto_si_no_primaria(self, snapshot=None):
        """Adopta estado central en UI; nunca ejecuta una transición local."""
        state = dict(snapshot or self._snapshot_operacional_integrado())
        station_role = str(state.get("role") or "").strip().upper()
        if not station_role or station_role == "PRIMARY":
            return False
        return self.apply_operational_snapshot(state)

    def _turno_pertenece_a_sesion(self, turno=None):
        turno = turno or cargar_turno_config(permitir_vencido=True) or {}
        central = dict(self.current_shift_context or {})
        central_user_id = str(central.get("owner_user_id") or "").strip()
        session_user_id = str(self.user_id or "").strip()
        if central_user_id and session_user_id:
            return central_user_id == session_user_id
        central_username = str(central.get("owner_username") or "").strip()
        session_username = str(self.session_context.username or "").strip()
        if central_username and session_username:
            return central_username.casefold() == session_username.casefold()
        # El nombre visible nunca es una identidad ni una autorización.
        return False

    def _asegurar_turno_de_sesion(self):
        if not self.session_context.launched_from_billing:
            return True

        snapshot = self._snapshot_operacional_integrado()
        if str(snapshot.get("operational_session_id") or "").strip():
            # The authenticated login never owns or creates an operational turn.
            # Adopt the central identity on every station, including PRIMARY.
            self.apply_operational_snapshot(snapshot)
            return bool(
                snapshot.get("writable", snapshot.get("write_allowed", False))
            )

        # A missing central snapshot must be handled by the normal connection
        # state.  It must not open the legacy turn dialog or create a turn whose
        # identity is inferred from the logged-in user.
        return normalize_role(getattr(self.session_context, "role", "")) == ROLE_ADMIN

    def _verificar_sesion_principal(self):
        self._session_after_id = None
        if self._shutdown_complete:
            return
        if not self.session_context.launched_from_billing:
            return
        try:
            estado = self.main_app_gateway.session_status()
            if estado.username.casefold() != self.session_context.username.casefold():
                raise MainAppGatewayError("La identidad de la sesión principal cambió.")
            self._main_session_failures = 0
        except MainAppGatewayError:
            self._main_session_failures += 1
            if self._main_session_failures >= 2:
                messagebox.showinfo(
                    "Sesión finalizada",
                    "Facturación se cerró o cambió de usuario. Admisión también se cerrará para proteger el turno.",
                    parent=self.root,
                )
                self.on_close()
                return
        try:
            self._session_after_id = self.root.after(5000, self._verificar_sesion_principal)
        except tk.TclError:
            pass

    def _puede(self, capability):
        return self.session_context.allows(capability)

    def _exigir_permiso(self, capability, accion):
        if self._puede(capability):
            return True
        messagebox.showwarning(
            "Acceso restringido",
            f"Tu rol no puede {accion}. Si necesitas hacerlo, solicita apoyo al administrador.",
            parent=self.root,
        )
        return False

    def _solicitar_autorizacion_admin(self, accion, parent=None, force=False):
        parent = parent or self.root
        actor = self._actor_actual()
        if (
            not force
            and self._admin_authorized_until
            and datetime.now() < self._admin_authorized_until
            and self._admin_authorized_actor
        ):
            return self._admin_authorized_actor
        try:
            configured = self.security.is_configured()
        except (ConfigError, SecurityError) as exc:
            messagebox.showerror("Seguridad", str(exc), parent=parent)
            return None
        if not configured:
            if not messagebox.askyesno(
                "Configurar administración",
                "Esta acción requiere un PIN administrativo y todavía no existe uno.\n\n"
                "¿Desea configurarlo ahora?",
                parent=parent,
            ):
                return None
            first = simpledialog.askstring(
                "Nuevo PIN administrativo",
                "Defina un PIN numérico de al menos 6 dígitos:",
                show="*",
                parent=parent,
            )
            if first is None:
                return None
            second = simpledialog.askstring(
                "Confirmar PIN",
                "Repita el PIN administrativo:",
                show="*",
                parent=parent,
            )
            if first != second:
                messagebox.showerror("Seguridad", "Los PIN no coinciden.", parent=parent)
                return None
            try:
                self.security.setup(first, actor=actor)
            except SecurityError as exc:
                messagebox.showerror("Seguridad", str(exc), parent=parent)
                return None

        pin = simpledialog.askstring(
            "Autorización administrativa",
            f"Ingrese el PIN para autorizar:\n{accion}",
            show="*",
            parent=parent,
        )
        if pin is None:
            return None
        try:
            if not self.security.verify(pin, actor=actor, action=accion):
                messagebox.showerror("Seguridad", "PIN administrativo incorrecto.", parent=parent)
                return None
        except SecurityError as exc:
            messagebox.showerror("Seguridad", str(exc), parent=parent)
            return None
        self._admin_authorized_until = datetime.now() + timedelta(minutes=5)
        self._admin_authorized_actor = actor
        return actor

    def set_status(self, mensaje, tipo="info"):
        colores = {
            "info": "#72E39B",
            "ok": "#72E39B",
            "warning": "#FFD166",
            "error": "#FF6B6B",
            "process": "#5CB6FF",
        }
        try:
            self.status_var.set(mensaje)
            self.status_label.configure(foreground=colores.get(tipo, "#72E39B"))
            self.root.update_idletasks()
        except Exception:
            pass

    def _configurar_accesibilidad_teclado(self):
        self._tab_order = [
            self.entry_nombre,
            self.lbl_sexo_m,
            self.lbl_sexo_f,
            self.entry_edad,
            self.combo_unidad,
            self.check_urgencia,
            self.entry_cedula,
            self.entry_telefono,
            self.entry_direccion,
            self.entry_nacionalidad,
            self.entry_ars,
            self.entry_nss,
            self.boton_limpiar,
            self.boton_generar_pdf,
            self.boton_historial,
            self.actions_menu_button,
        ]
        for w in self._tab_order:
            try:
                w.configure(takefocus=True)
            except Exception:
                pass
            try:
                w.bind("<Tab>", lambda e, widget=w: self._focus_siguiente(widget), add="+")
                w.bind("<Shift-Tab>", lambda e, widget=w: self._focus_anterior(widget), add="+")
                w.bind("<ISO_Left_Tab>", lambda e, widget=w: self._focus_anterior(widget), add="+")
            except Exception:
                pass

        self.root.bind("<Control-l>", lambda e: (self.limpiar_campos(), self.set_status("✓  Campos limpiados", "ok"), "break"))
        self.root.bind("<Control-L>", lambda e: (self.limpiar_campos(), self.set_status("✓  Campos limpiados", "ok"), "break"))
        self.root.bind("<Control-h>", lambda e: (self.abrir_historial(), "break"))
        self.root.bind("<Control-H>", lambda e: (self.abrir_historial(), "break"))
        self.root.bind("<Control-z>", self._undo_last_action)
        self.root.bind("<Control-Z>", self._undo_last_action)
        def registrar_desde_teclado(_event=None):
            self.generar_pdf()
            return "break"

        self.root.bind("<Control-Return>", registrar_desde_teclado)
        self.root.bind("<F9>", registrar_desde_teclado)
        self.root.bind("<Return>", self._enter_generar_si_completo, add="+")

    def _focus_siguiente(self, widget):
        try:
            idx = self._tab_order.index(widget)
            self._tab_order[(idx + 1) % len(self._tab_order)].focus_set()
            return "break"
        except Exception:
            return None

    def _focus_anterior(self, widget):
        try:
            idx = self._tab_order.index(widget)
            self._tab_order[(idx - 1) % len(self._tab_order)].focus_set()
            return "break"
        except Exception:
            return None

    def _campos_minimos_completos(self):
        nombre = (self.entry_nombre.get() or "").strip()
        telefono = (self.entry_telefono.get() or "").strip()
        edad = (self.entry_edad.get() or "").strip()
        unidad = self.unidad_edad.get()
        return bool(nombre and telefono.isdigit() and len(telefono) == 10 and edad.isdigit() and unidad in ("Días", "Meses", "Años"))

    def _enter_generar_si_completo(self, event=None):
        try:
            if isinstance(event.widget, tk.Listbox):
                return None
        except Exception:
            pass
        if self._campos_minimos_completos():
            self.generar_pdf()
            return "break"
        self.set_status("Complete nombre, edad y teléfono válido para generar con Enter.", "warning")
        return None

    def _obtener_ars_cache(self, forzar=False):
        try:
            ahora = datetime.now().timestamp()
            if (not forzar) and self._cache_ars and (ahora - float(self._cache_ars_time or 0) < 30):
                return list(self._cache_ars)
            self._cache_ars = self.db.listar_ars_distintas()
            self._cache_ars_time = ahora
            return list(self._cache_ars)
        except Exception:
            return []

    def _invalidar_cache_ars(self):
        self._cache_ars = []
        self._cache_ars_time = 0
        self._invalidar_cache_resumen_turno()

    def _invalidar_cache_resumen_turno(self):
        self._cache_resumen_turno = None
        self._cache_resumen_time = 0

    def _obtener_resumen_turno_cache(self, forzar=False):
        ahora = datetime.now().timestamp()
        if (not forzar) and self._cache_resumen_turno and (ahora - float(self._cache_resumen_time or 0) < 10):
            return dict(self._cache_resumen_turno)

        resumen = self.db.resumen_turno_actual()
        self._cache_resumen_turno = dict(resumen)
        self._cache_resumen_time = ahora
        self._guardar_resumen_turno_json(resumen)
        return resumen

    def _guardar_resumen_turno_json(self, resumen: dict):
        """
        FASE 7: resumen_turno.json para estadísticas rápidas.
        """
        try:
            payload = {
                "actualizado": datetime.now().strftime("%d/%m/%Y %I:%M:%S %p"),
                "total": int(resumen.get("total", 0) or 0),
                "sin_seguro": int(resumen.get("sin_seguro", 0) or 0),
                "general": int(resumen.get("GENERAL", 0) or 0),
                "pediatria": int(resumen.get("PEDIATRIA", 0) or 0),
                "ginecologia": int(resumen.get("GINECOLOGIA", 0) or 0),
                "urgencias": int(resumen.get("URGENCIAS", 0) or 0),
                "consultas": int(resumen.get("CONSULTAS", 0) or 0),
            }
            with open(RESUMEN_TURNO_PATH, "w", encoding="utf-8") as f:
                json.dump(payload, f, ensure_ascii=False, indent=2)
        except Exception:
            pass

    def _invalidar_caches_datos(self):
        self._invalidar_cache_ars()
        self._invalidar_cache_resumen_turno()


    def _turno_config_desde_snapshot_operacional(self):
        """Build a display-only V15 turn config from the central snapshot.

        ``turnos_config.json`` is a local mirror and can legitimately lag on a
        freshly attached secondary.  The central snapshot already adopted by
        the host must therefore win for representative and turn presentation.
        """
        context = dict(getattr(self, "current_shift_context", {}) or {})
        codigo = str(context.get("turn_code") or "").strip()
        if not codigo:
            return None
        started_at = context.get("turn_started_at")
        if isinstance(started_at, str):
            try:
                started_at = datetime.fromisoformat(
                    started_at.replace("Z", "+00:00")
                )
            except ValueError:
                started_at = None
        if isinstance(started_at, datetime) and started_at.tzinfo is not None:
            started_at = started_at.astimezone().replace(tzinfo=None)
        if not isinstance(started_at, datetime):
            return None
        return {
            "turno_codigo": codigo,
            "fecha_base": started_at.date(),
            "inicio_real_dt": started_at,
        }

    def _descripcion_turno_snapshot_operacional(self):
        """Returns the central turn label without consulting local turn JSON."""
        context = dict(getattr(self, "current_shift_context", {}) or {})
        if not str(context.get("operational_session_id") or "").strip():
            return "No configurado"
        codigo = str(context.get("turn_code") or "").strip()
        if not codigo:
            return "No configurado"
        return {
            "8AM_8AM": "8:00 AM → 8:00 AM",
            "8AM_8PM": "8:00 AM → 8:00 PM",
            "8PM_8AM": "8:00 PM → 8:00 AM",
        }.get(normalizar_turno_codigo(codigo), codigo)


    def _actualizar_turno_visual_en_vivo(self):
        try:
            central_cfg = self._turno_config_desde_snapshot_operacional()
            central = dict(self.current_shift_context or {})
            central_available = bool(
                str(central.get("operational_session_id") or "").strip()
            )
            if central_available:
                turno_txt = self._descripcion_turno_snapshot_operacional()
                representante = limpiar_nombre_representante(
                    central.get("representative_display_name")
                    or central.get("owner_username")
                ) or "No configurado"
            else:
                cfg = central_cfg or cargar_turno_config(permitir_vencido=True)
                turno_txt = descripcion_turno_config(cfg)
                representante = limpiar_nombre_representante(
                    (cfg.get("representante") if cfg else "")
                ) or "No configurado"

            if hasattr(self, "turno_header_var"):
                self.turno_header_var.set(turno_txt)
            if hasattr(self, "turno_panel_var"):
                self.turno_panel_var.set(f"Turno:\n{turno_txt}")
            if hasattr(self, "representante_panel_var"):
                self.representante_panel_var.set(f"Representante:\n{limpiar_nombre_representante(representante) or 'No configurado'}")

            self._refrescar_resumen_en_vivo()
            self._set_turn_change_controls_enabled(True)
            refresh_primary = getattr(self, "_refresh_primary_config_panel", None)
            if callable(refresh_primary):
                refresh_primary()
            try:
                self.root.update_idletasks()
            except Exception:
                pass
        except Exception:
            pass

    def _actualizar_fecha_actual(self):
        self._date_after_id = None
        if self._shutdown_complete or not self._admission_visible:
            return
        try:
            if hasattr(self, "fecha_actual_var"):
                self.fecha_actual_var.set(datetime.now().strftime("%d/%m/%Y"))
        except Exception:
            pass
        try:
            self._date_after_id = self.root.after(60000, self._actualizar_fecha_actual)
        except Exception:
            pass

    def _crear_panel_acciones_rapidas(self, parent):
        """
        Panel derecho refinado para la vista PySide6. Conserva exactamente los
        mismos callbacks y permisos; únicamente mejora jerarquía visual, tamaños
        y separación respecto al formulario principal.
        """
        for w in parent.winfo_children():
            try:
                w.destroy()
            except Exception:
                pass

        pal = self._paleta_visual_actual()
        self.quick_action_buttons = {}

        actions_area = tb.Frame(parent, padding=(0, 0), style="Card.TFrame")
        actions_area.pack(side="top", fill="x")
        self.quick_actions_title = tb.Label(
            actions_area, text="⚡  ACCIONES RÁPIDAS", font=("Segoe UI", 12, "bold"),
            foreground=pal["accent"], background=pal["card"]
        )
        self.quick_actions_title.pack(anchor="w", pady=(0, 10))

        def add_action(texto, comando, bootstyle=SECONDARY):
            btn = tb.Button(
                actions_area, text=f"{texto}   ›", bootstyle=bootstyle,
                command=comando, width=30
            )
            btn.pack(fill="x", pady=4, ipady=4)
            role_map = {
                "Reporte estadístico": "report",
                "Abrir Listado en Excel": "excel",
                "Ver Historial sin Seguro": "uninsured",
                "Editar paciente": "edit",
                "Configuración interna": "config",
            }
            self.quick_action_buttons[role_map.get(texto, texto)] = btn
            return btn

        if self._puede(CAP_VIEW_REPORTS):
            add_action("Reporte estadístico", self.abrir_ventana_reporte, INFO)
        if self._puede(CAP_OPEN_EXCEL):
            add_action("Abrir Listado en Excel", self._abrir_excel_actual, SUCCESS)
        add_action("Ver Historial sin Seguro", self.abrir_historial_sin_seguros, WARNING)
        if self._puede(CAP_EDIT_RECORDS):
            add_action("Editar paciente", self._abrir_edicion_paciente, PRIMARY)
        if self._puede(CAP_INTERNAL_CONFIG):
            add_action("Configuración interna", self._abrir_configuracion_interna, SECONDARY)

        # Panel informativo claramente separado del formulario.
        self.quick_info_area = tb.Frame(parent, padding=(0, 0), style="Card.TFrame")
        self.quick_info_area.pack(side="top", fill="x", pady=(18, 0))

        self.quick_info_title = tb.Label(
            self.quick_info_area, text="INFORMACIÓN", font=("Segoe UI", 11, "bold"),
            foreground=pal["muted"], background=pal["card"]
        )
        self.quick_info_title.pack(anchor="w", fill="x", pady=(0, 8))

        try:
            central = dict(getattr(self, "current_shift_context", {}) or {})
            central_available = bool(
                str(central.get("operational_session_id") or "").strip()
            )
            if central_available:
                representante = limpiar_nombre_representante(
                    central.get("representative_display_name")
                    or central.get("owner_username")
                ) or "No configurado"
                turno = self._descripcion_turno_snapshot_operacional()
            else:
                cfg = cargar_turno_config(permitir_vencido=True)
                representante = limpiar_nombre_representante(
                    (cfg.get("representante") if cfg else "")
                ) or "No configurado"
                turno = descripcion_turno_config(cfg)
        except Exception:
            representante = "No configurado"
            turno = "No configurado"

        self.representante_panel_var = tk.StringVar(value=f"Representante:\n{representante}")
        self.turno_panel_var = tk.StringVar(value=f"Turno:\n{turno}")
        self.quick_representante_label = tb.Label(
            self.quick_info_area, textvariable=self.representante_panel_var, style="Muted.TLabel",
            background=pal["card"], wraplength=330, justify="left"
        )
        self.quick_representante_label.pack(anchor="w", fill="x", pady=(0, 4))
        self.quick_turno_label = tb.Label(
            self.quick_info_area, textvariable=self.turno_panel_var, style="Muted.TLabel",
            background=pal["card"], wraplength=330, justify="left"
        )
        self.quick_turno_label.pack(anchor="w", fill="x", pady=(0, 12))

        self.quick_summary_title = tb.Label(
            self.quick_info_area, text="RESUMEN DEL TURNO", font=("Segoe UI", 11, "bold"),
            foreground=pal["muted"], background=pal["card"]
        )
        self.quick_summary_title.pack(anchor="w", fill="x", pady=(0, 7))
        self.turno_total_var = tk.StringVar(value="Total pacientes: —")
        self.quick_total_label = tb.Label(
            self.quick_info_area, textvariable=self.turno_total_var,
            background=pal["card"], justify="left", anchor="w"
        )
        self.quick_total_label.pack(anchor="w", fill="x", pady=(0, 4))
        self.turno_resumen_var = tk.StringVar(value="Cargando resumen…")
        self.quick_summary_label = tb.Label(
            self.quick_info_area, textvariable=self.turno_resumen_var, style="Muted.TLabel",
            background=pal["card"], justify="left", wraplength=300
        )
        self.quick_summary_label.pack(anchor="w", fill="x")

        self._actualizar_resumen_turno_panel()
        self._aplicar_paridad_visual_inicio()

    def apply_turn_summary(self, resumen, *, reason="background_refresh"):
        """Aplica al panel el resultado canónico ya calculado en background."""
        values = dict(resumen or {})
        self._cache_resumen_turno = values
        self._cache_resumen_time = datetime.now().timestamp()
        APP_LOG.info(
            "TURN_SUMMARY_PANEL_APPLY reason=%s total=%s general=%s pediatria=%s "
            "ginecologia=%s urgencias=%s consultas=%s status=%s error_code=%s "
            "turn_id=%s generation=%s operational_revision=%s",
            str(reason or "background_refresh"),
            int(values.get("total", 0) or 0),
            int(values.get("GENERAL", 0) or 0),
            int(values.get("PEDIATRIA", 0) or 0),
            int(values.get("GINECOLOGIA", 0) or 0),
            int(values.get("URGENCIAS", 0) or 0),
            int(values.get("CONSULTAS", 0) or 0),
            str(values.get("_status") or "UNKNOWN"),
            str(values.get("_error_code") or "-"),
            values.get("_turn_id"),
            values.get("_generation"),
            values.get("_operational_revision"),
        )
        self._actualizar_resumen_turno_panel(resumen=values)

    def _actualizar_resumen_turno_panel(self, forzar=False, resumen=None):
        try:
            r = dict(resumen) if resumen is not None else self._obtener_resumen_turno_cache(forzar=forzar)
            fuente = "Base de datos"
            if r.get("_fuente") == "BD_EXCEL":
                fuente = "BD y Excel sincronizados"
            elif os.path.exists(EXCEL_PATH):
                fuente = "BD · Excel pendiente de actualización"
            if r.get("_fuente") == "EXCEL_RECUPERADO":
                fuente = "Excel recuperado · revisión de sincronización"
            if r.get("_status") == "STALE":
                fuente = "Último resumen válido · conexión temporalmente no verificada"
            elif r.get("_status") == "INVALID_REFRESH":
                return

            if not bool(self.app_settings.get("show_turno_summary", True)):
                total_texto = ""
                texto = "Resumen oculto por preferencias."
            else:
                # Total = solo EMERGENCIAS. Urgencias y consultas son conteos aparte.
                total_emergencias = sum(
                    int(r.get(categoria, 0) or 0)
                    for categoria in ("GENERAL", "PEDIATRIA", "GINECOLOGIA")
                )
                total_texto = f"Total pacientes: {total_emergencias}"
                texto = (
                    f"Sin seguro: {r.get('sin_seguro', 0)}\n"
                    f"General: {r.get('GENERAL', 0)}\n"
                    f"Pediatría: {r.get('PEDIATRIA', 0)}\n"
                    f"Ginecología: {r.get('GINECOLOGIA', 0)}\n"
                    f"Urgencias: {r.get('URGENCIAS', 0)}\n"
                    f"Consultas: {r.get('CONSULTAS', 0)}"
                )

            if hasattr(self, "turno_total_var"):
                self.turno_total_var.set(total_texto)
            if hasattr(self, "turno_resumen_var"):
                self.turno_resumen_var.set(texto)

            if hasattr(self, "connection_var"):
                self.connection_var.set(f"Conectado a: {fuente}")

            try:
                self.root.update_idletasks()
            except Exception:
                pass

        except Exception:
            if hasattr(self, "turno_total_var"):
                self.turno_total_var.set("")
            if hasattr(self, "turno_resumen_var"):
                self.turno_resumen_var.set("No disponible")
            if hasattr(self, "connection_var"):
                self.connection_var.set("Conectado a: no disponible")

    def _programar_refresco_resumen_en_vivo(self):
        """Solicita una lectura inicial; los cambios posteriores son event-driven."""
        self._summary_after_id = None
        if self._shutdown_complete or not self._admission_visible:
            return
        try:
            if getattr(self, "root", None) and self.root.winfo_exists():
                if self.root.state() != "withdrawn":
                    self._refrescar_resumen_en_vivo()
        except Exception:
            pass

    def _refrescar_resumen_en_vivo(self, delay_ms=0):
        try:
            if delay_ms and delay_ms > 0:
                self.root.after(
                    delay_ms,
                    lambda: self.request_turn_summary_refresh("v15_ui_event"),
                )
                return
            self.request_turn_summary_refresh("v15_ui_event")
        except Exception:
            pass

    def request_turn_summary_refresh(self, reason="attention_event"):
        """Única entrada event-driven para recalcular los conteos del turno."""
        self._invalidar_cache_resumen_turno()
        coordinator = getattr(self, "_hybrid_refresh_controller", None)
        if coordinator is not None:
            coordinator.request_summary(str(reason or "attention_event"))
            return
        self._actualizar_resumen_turno_panel(forzar=True)

    def _abrir_excel_actual(self):
        if not self._exigir_permiso(CAP_OPEN_EXCEL, "abrir el listado administrativo de Excel"):
            return
        try:
            latest_status = synchronize_latest_excel()
            if latest_status == "FILE_LOCKED":
                self.set_status(
                    "Excel abierto; la versión actualizada queda pendiente.",
                    "warning",
                )
            turno_cfg = cargar_turno_config(permitir_vencido=True)
            if turno_cfg:
                preview_wb, patient_count = _construir_workbook_turno(
                    self.db, turno_cfg
                )
                preview_wb.close()
                if patient_count == 0:
                    messagebox.showinfo(
                        "Listado en Excel",
                        "No hay pacientes para incluir en este listado.",
                    )
                    self.set_status("Listado omitido: no hay pacientes.", "warning")
                    return
            verificar_o_crear_excel()
            ruta = os.path.abspath(EXCEL_PATH)
            if platform.system() == "Windows":
                os.startfile(ruta)
            elif platform.system() == "Darwin":
                subprocess.run(["open", ruta], check=False)
            else:
                subprocess.run(["xdg-open", ruta], check=False)
            self.set_status("Listado de Excel abierto", "ok")
        except Exception as e:
            self.set_status("Error: Excel abierto o no disponible", "error")
            messagebox.showerror("Error", f"No se pudo abrir el listado de Excel:\n{str(e)}")

    def abrir_ventana_reporte(self):
        if not self._exigir_permiso(CAP_VIEW_REPORTS, "consultar reportes administrativos"):
            return
        win = self._crear_toplevel_estable("Reporte estadístico", "1240x850", "reporte_win")
        if win is None:
            return
        self._bind_esc_cerrar(win)
        cont = tb.Frame(win, padding=14, style="Root.TFrame")
        cont.pack(fill="both", expand=True)
        self._crear_header_ventana(
            cont,
            "Reporte estadístico",
            "Genera reportes por turno, período, ARS, especialidad y cobertura.",
            "📊"
        )
        try:
            win.minsize(1120, 760)
            win.resizable(True, True)
        except Exception:
            pass
        barra = tb.Frame(cont, padding=(8, 8), style="Card.TFrame")
        barra.pack(side="bottom", fill="x", pady=(8, 0))
        panel = tb.Frame(cont, padding=12, style="Card.TFrame")
        panel.pack(fill="x", pady=(0, 8))
        panel.columnconfigure(0, weight=3)
        panel.columnconfigure(1, weight=0)
        panel.columnconfigure(2, weight=2)

        filtros = tb.Frame(panel, style="Card.TFrame")
        filtros.grid(row=0, column=0, sticky="nsew", padx=(0, 12))
        filtros.columnconfigure(1, weight=1)
        filtros.columnconfigure(3, weight=1)
        tb.Label(filtros, text="⚲  Filtros", font=("Arial", 11, "bold")).grid(
            row=0, column=0, columnspan=4, sticky="w", padx=4, pady=(0, 8)
        )
        periodo_var = tk.StringVar(value="Diario")
        turno_var = tk.StringVar(value="Turno actual")
        especialidad_var = tk.StringVar(value=SPECIALTY_ALL)
        cobertura_var = tk.StringVar(value=COVERAGE_ALL)
        fecha_inicio = crear_selector_fecha(filtros, width=14)
        fecha_fin = crear_selector_fecha(filtros, width=14)
        fecha_base = fecha_base_operativa_actual()
        establecer_fecha_selector(fecha_inicio, fecha_base)
        establecer_fecha_selector(fecha_fin, fecha_base)

        tb.Label(filtros, text="Tipo de período").grid(row=1, column=0, sticky="w", padx=4, pady=4)
        combo_periodo = tb.Combobox(
            filtros,
            textvariable=periodo_var,
            state="readonly",
            values=["Diario", "Semanal", "Mensual", "Anual", "Rango"],
            width=22,
        )
        combo_periodo.grid(row=1, column=1, columnspan=3, sticky="ew", padx=4, pady=4)
        tb.Label(filtros, text="Turno").grid(row=2, column=0, sticky="w", padx=4, pady=4)
        combo_turno = tb.Combobox(
            filtros,
            textvariable=turno_var,
            state="readonly",
            values=["Turno actual", "Turno anterior", "Todos los turnos"],
            width=22,
        )
        combo_turno.grid(row=2, column=1, columnspan=3, sticky="ew", padx=4, pady=4)
        tb.Label(filtros, text="Desde").grid(row=3, column=0, sticky="w", padx=4, pady=4)
        fecha_inicio.grid(row=3, column=1, sticky="ew", padx=4, pady=4)
        tb.Label(filtros, text="Hasta").grid(row=3, column=2, sticky="w", padx=4, pady=4)
        fecha_fin.grid(row=3, column=3, sticky="ew", padx=4, pady=4)
        tb.Label(filtros, text="Especialidad").grid(row=4, column=0, sticky="w", padx=4, pady=4)
        tb.Combobox(
            filtros,
            textvariable=especialidad_var,
            state="readonly",
            values=[SPECIALTY_ALL, "GENERAL", "PEDIATRIA", "GINECOLOGIA", "OTRAS"],
            width=22,
        ).grid(row=4, column=1, columnspan=3, sticky="ew", padx=4, pady=4)
        tb.Label(filtros, text="Cobertura").grid(row=5, column=0, sticky="w", padx=4, pady=4)
        tb.Combobox(
            filtros,
            textvariable=cobertura_var,
            state="readonly",
            values=[COVERAGE_ALL, COVERAGE_INSURED, COVERAGE_UNINSURED],
            width=22,
        ).grid(row=5, column=1, columnspan=3, sticky="ew", padx=4, pady=4)

        ttk.Separator(panel, orient="vertical").grid(
            row=0, column=1, sticky="ns", padx=(0, 12)
        )
        ars_panel = tb.Frame(panel, style="Card.TFrame")
        ars_panel.grid(row=0, column=2, sticky="nsew")
        ars_panel.columnconfigure(0, weight=1)
        tb.Label(ars_panel, text="ARS", font=("Arial", 11, "bold")).grid(
            row=0, column=0, sticky="w", padx=4, pady=(0, 8)
        )
        ars_mode_var = tk.StringVar(value=ARS_ALL)
        ars_search_var = tk.StringVar(value="")
        tb.Label(ars_panel, text="Modo ARS").grid(row=1, column=0, sticky="w", padx=4)
        ars_mode_combo = tb.Combobox(
            ars_panel,
            textvariable=ars_mode_var,
            state="readonly",
            values=[ARS_ALL, ARS_INCLUDE, ARS_EXCLUDE],
            width=20,
        )
        ars_mode_combo.grid(row=2, column=0, sticky="ew", padx=4, pady=(2, 6))
        ars_search = tb.Entry(ars_panel, textvariable=ars_search_var)
        ars_search.grid(row=3, column=0, sticky="ew", padx=4, pady=(0, 6))
        try:
            ars_search.insert(0, "")
            ars_search.setPlaceholderText("Buscar ARS…")
        except Exception:
            pass
        ars_canvas = tk.Canvas(ars_panel, height=110)
        try:
            ars_canvas.setMinimumHeight(120)
            ars_canvas.setMaximumHeight(120)
            ars_canvas.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Fixed)
        except Exception:
            pass
        ars_canvas.grid(row=4, column=0, sticky="nsew", padx=4)
        ars_checks_frame = tb.Frame(ars_canvas, style="Card.TFrame")
        ars_canvas.create_window((0, 0), window=ars_checks_frame, anchor="nw")
        ars_catalog = sorted(set(self._obtener_catalogo_ars()) | {"SIN SEGURO"})
        ars_vars = {}
        ars_checks = {}
        for row_index, ars_name in enumerate(ars_catalog):
            variable = tk.BooleanVar(value=False)
            checkbox = tb.Checkbutton(
                ars_checks_frame,
                text=ars_name,
                variable=variable,
                bootstyle=INFO,
            )
            checkbox.grid(row=row_index, column=0, sticky="w", padx=3, pady=1)
            ars_vars[ars_name] = variable
            ars_checks[ars_name] = checkbox

        tb.Label(
            panel,
            text="ⓘ Los turnos usan identidad operacional persistida y una ventana de 8:00 AM a 8:00 AM.",
            style="Muted.TLabel",
        ).grid(row=1, column=0, columnspan=3, sticky="w", padx=4, pady=(10, 0))

        cards = tb.Frame(cont, style="Root.TFrame")
        cards.pack(fill="x", pady=(0, 8))
        card_specs = (
            ("total", "👥", "Total pacientes"),
            ("insured", "🛡", "Asegurados"),
            ("uninsured", "👤", "Sin seguro"),
            ("general", "⚕", "General"),
            ("pediatric", "🧒", "Pediatría"),
            ("gynecology", "♀", "Ginecología"),
        )
        card_vars = {}
        for column, (key, icon, title) in enumerate(card_specs):
            cards.columnconfigure(column, weight=1)
            card = tb.Frame(cards, padding=10, style="Card.TFrame")
            card.grid(row=0, column=column, sticky="nsew", padx=3)
            value_var = tk.StringVar(value="0")
            detail_var = tk.StringVar(value="0% del total")
            tb.Label(card, text=f"{icon}  {title}", anchor="center").pack(fill="x")
            tb.Label(card, textvariable=value_var, font=("Arial", 18, "bold"), anchor="center").pack(fill="x")
            tb.Label(card, textvariable=detail_var, style="Muted.TLabel", anchor="center").pack(fill="x")
            card_vars[key] = (value_var, detail_var)

        preview_card = tb.Frame(cont, padding=6, style="Card.TFrame")
        preview_card.pack(fill="both", expand=True)
        tb.Label(preview_card, text="◉  Vista previa del reporte", font=("Arial", 11, "bold")).pack(
            anchor="w", padx=4, pady=(0, 5)
        )
        vista = ttk.Treeview(
            preview_card,
            columns=("seccion", "concepto", "cantidad", "ars", "turno", "observacion"),
            show="headings",
            height=7,
            style="Modern.Treeview",
        )
        vista.pack(fill="both", expand=True)
        vista.heading("seccion", text="Sección")
        vista.heading("concepto", text="Concepto")
        vista.heading("cantidad", text="Cantidad")
        vista.heading("ars", text="ARS")
        vista.heading("turno", text="Turno")
        vista.heading("observacion", text="Observación")
        for column, width in (
            ("seccion", 145), ("concepto", 245), ("cantidad", 90),
            ("ars", 120), ("turno", 160), ("observacion", 260),
        ):
            vista.column(column, width=width, anchor="center" if column in {"cantidad", "ars"} else "w")

        estado_var = tk.StringVar(value="Seleccione filtros y presione Generar reporte.")
        snapshot_store = ReportSnapshotStore()
        ultimo = {
            "dataset": None,
            "snapshot_store": snapshot_store,
            "pdf": "",
            "excel": "",
        }
        pdf_button = None
        excel_button = None

        def _selected_ars():
            return tuple(name for name, variable in ars_vars.items() if variable.get())

        def _period_from_controls():
            first = parse_fecha_ddmmyyyy(obtener_fecha_selector(fecha_inicio)) or fecha_base_operativa_actual()
            last = parse_fecha_ddmmyyyy(obtener_fecha_selector(fecha_fin)) or first
            return build_operational_period(periodo_var.get(), first, last)

        def _refresh_ars_search(*_):
            visible = set(search_ars_catalog(ars_catalog, ars_search_var.get()))
            for name, checkbox in ars_checks.items():
                if name in visible:
                    checkbox.grid()
                else:
                    checkbox.grid_remove()

        def _load_dataset_into_ui(dataset):
            vista.delete(*vista.get_children())
            for preview_row in dataset.preview_rows:
                vista.insert("", "end", values=preview_row)
            summary = dataset.summary
            total = int(summary["total_patients"])
            values = {
                "total": (total, 100.0 if total else 0.0),
                "insured": (summary["insured_patients"], summary["insured_percentage"]),
                "uninsured": (summary["uninsured_patients"], summary["uninsured_percentage"]),
                "general": (summary["general_patients"], (summary["general_patients"] * 100 / total) if total else 0),
                "pediatric": (summary["pediatric_patients"], (summary["pediatric_patients"] * 100 / total) if total else 0),
                "gynecology": (summary["gynecology_patients"], (summary["gynecology_patients"] * 100 / total) if total else 0),
            }
            for key, (count, percentage) in values.items():
                card_vars[key][0].set(f"{int(count):,}")
                card_vars[key][1].set(f"{float(percentage):.2f}% del total")

        def _set_export_enabled(enabled):
            state = "normal" if enabled else "disabled"
            for button in (pdf_button, excel_button):
                if button is not None:
                    button.configure(state=state)

        def _mark_snapshot_stale(*_):
            if not snapshot_store.mark_stale():
                return
            _set_export_enabled(False)
            estado_var.set("Filtros modificados · genere nuevamente antes de exportar.")

        def _friendly_turn_label(scope, selected_turn, turn_period):
            if selected_turn is None:
                return scope
            return (
                f"{scope} · {turn_period.start_at:%d/%m/%Y %I:%M %p} → "
                f"{turn_period.end_at:%d/%m/%Y %I:%M %p}"
            )

        def _turns_for_snapshot(source, start_at, end_at):
            selected = source.get("selected_turn")
            if selected is not None:
                return (selected,)
            result = []
            for turn in source.get("turns") or ():
                turn_period = build_turn_operational_period(
                    turn.get("started_at"), fallback_date=start_at.date()
                )
                if turn_period.start_at < end_at and start_at < turn_period.end_at:
                    result.append(turn)
            return tuple(result)

        def _log_report_output_error(code, operation, error):
            APP_LOG.error(
                "STATISTICAL_REPORT_OUTPUT_FAILED category=REPORT_EXPORT_ERROR "
                "code=%s operation=%s "
                "exception_type=%s safe_error_message=%s",
                code,
                operation,
                type(error).__name__,
                "No fue posible generar el archivo solicitado.",
            )

        def generar():
            try:
                period = _period_from_controls()
                scope = turno_var.get()
                ars_mode = ars_mode_var.get()
                selected_ars = _selected_ars()
                specialty = especialidad_var.get()
                coverage = cobertura_var.get()
                snapshot = dict(self.db.get_operational_station_snapshot() or {})
                estado_var.set("Cargando datos del reporte…")
                self.set_status("Generando reporte en segundo plano…", "process")
                win.update_idletasks()
                def _trabajo():
                    source_id = str(snapshot.get("operational_source_id") or "").strip()
                    current_turn_id = int(snapshot.get("turn_id") or 0)
                    if not source_id or current_turn_id <= 0:
                        raise RuntimeError(
                            "No existe una identidad operacional central activa."
                        )
                    source = self.db.load_statistical_report_source(
                        operational_source_id=source_id,
                        turn_scope=scope,
                        current_turn_id=current_turn_id,
                        start_at=period.start_at,
                        end_at=period.end_at,
                    )
                    turn_id = source.get("turn_id")
                    selected_turn = source.get("selected_turn")
                    if selected_turn is not None:
                        turn_period = build_turn_operational_period(
                            selected_turn.get("started_at"),
                            fallback_date=period.start_at.date(),
                        )
                        start_at, end_at = turn_period.start_at, turn_period.end_at
                    else:
                        turn_period = period
                        start_at, end_at = period.start_at, period.end_at
                    turn_label = _friendly_turn_label(
                        scope, selected_turn, turn_period
                    )
                    filters = AdmissionReportFilters(
                        start_at=start_at,
                        end_at=end_at,
                        period_label=(
                            turn_label if turn_id is not None else period.label
                        ),
                        turn_label=turn_label,
                        operational_source_id=source_id,
                        turn_id=turn_id,
                        specialty=specialty,
                        coverage=coverage,
                        ars_mode=ars_mode,
                        selected_ars=selected_ars,
                    )
                    dataset = build_admission_report_dataset(
                        source.get("records") or (),
                        filters,
                        turns=_turns_for_snapshot(source, start_at, end_at),
                    )
                    APP_LOG.info(
                        "ADMISSION_STATISTICAL_REPORT_DATASET turn_id=%s rows=%s diagnostics=%s",
                        turn_id,
                        len(dataset.records),
                        dict(dataset.diagnostics),
                    )
                    return dataset

                def _ok(dataset):
                    snapshot_store.replace(dataset)
                    ultimo.update({"dataset": dataset, "pdf": "", "excel": ""})
                    _load_dataset_into_ui(dataset)
                    _set_export_enabled(bool(dataset.records))
                    total = int(dataset.summary["total_patients"])
                    estado_var.set(f"Reporte listo · {total:,} paciente(s) del dataset canónico.")
                    self.set_status("Reporte cargado", "ok")
                def _error(e):
                    code = getattr(e, "code", None) or (
                        "REPORT_DATA_ERROR"
                        if isinstance(e, (TypeError, ValueError))
                        else "REPORT_QUERY_ERROR"
                    )
                    estado_var.set(f"No se pudo actualizar el reporte · {code}.")
                    messagebox.showerror(
                        "Reporte",
                        f"No se pudo generar el reporte.\nCódigo: {code}\n{str(e)}",
                        parent=win,
                    )
                self._ejecutar_en_segundo_plano("Generando reporte…", _trabajo, _ok, _error)
            except Exception as e:
                estado_var.set("Error al generar reporte.")
                messagebox.showerror("Reporte", f"No se pudo generar el reporte:\n{str(e)}", parent=win)

        def guardar_pdf():
            try:
                dataset = snapshot_store.require_exportable()
            except SnapshotStaleError as exc:
                messagebox.showinfo("Reporte", str(exc), parent=win)
                return
            if not dataset.records:
                messagebox.showinfo("Reporte", "El reporte no contiene pacientes.", parent=win)
                return
            def _trabajo():
                return crear_pdf_reporte(dataset.summary)
            def _ok(ruta):
                ultimo["pdf"] = ruta
                abrir_pdf(ruta)
                estado_var.set(f"PDF creado: {os.path.basename(ruta)}")
                self.set_status("PDF del reporte generado", "ok")
            def _error(error):
                _log_report_output_error("REPORT_PDF_ERROR", "pdf", error)
                messagebox.showerror("Reporte", f"No se pudo crear el PDF:\n{error}", parent=win)
            self._ejecutar_en_segundo_plano("Creando PDF…", _trabajo, _ok, _error)

        def guardar_excel_reporte():
            try:
                dataset = snapshot_store.require_exportable()
            except SnapshotStaleError as exc:
                messagebox.showinfo("Reporte", str(exc), parent=win)
                return
            if not dataset.records:
                messagebox.showinfo("Reporte", "El reporte no contiene pacientes.", parent=win)
                return
            destination = filedialog.asksaveasfilename(
                parent=win,
                title="Exportar reporte estadístico",
                defaultextension=".xlsx",
                filetypes=[("Excel", "*.xlsx")],
            )
            if not destination:
                return
            def _trabajo():
                return crear_excel_reporte_estadistico(dataset.summary, destino=destination)
            def _ok(ruta):
                ultimo["excel"] = ruta
                estado_var.set(f"Excel creado: {os.path.basename(ruta)}")
                self.set_status("Excel del reporte exportado", "ok")
            def _error(error):
                _log_report_output_error("REPORT_EXCEL_ERROR", "excel", error)
                messagebox.showerror("Reporte", f"No se pudo exportar el Excel:\n{error}", parent=win)
            self._ejecutar_en_segundo_plano("Exportando Excel…", _trabajo, _ok, _error)

        def limpiar_filtros():
            periodo_var.set("Diario")
            turno_var.set("Turno actual")
            especialidad_var.set(SPECIALTY_ALL)
            cobertura_var.set(COVERAGE_ALL)
            ars_mode_var.set(ARS_ALL)
            ars_search_var.set("")
            establecer_fecha_selector(fecha_inicio, fecha_base_operativa_actual())
            establecer_fecha_selector(fecha_fin, fecha_base_operativa_actual())
            for variable in ars_vars.values():
                variable.set(False)
            snapshot_store.clear()
            ultimo.update({"dataset": None, "pdf": "", "excel": ""})
            _set_export_enabled(False)
            vista.delete(*vista.get_children())
            for value_var, detail_var in card_vars.values():
                value_var.set("0")
                detail_var.set("0% del total")
            estado_var.set("Filtros restablecidos.")

        def _actualizar_fechas_por_periodo(*_):
            modo = periodo_var.get()
            if modo == "Diario":
                base = parse_fecha_ddmmyyyy(obtener_fecha_selector(fecha_inicio)) or datetime.now().date()
                establecer_fecha_selector(fecha_inicio, base)
                establecer_fecha_selector(fecha_fin, base)
            elif modo == "Semanal":
                base = parse_fecha_ddmmyyyy(obtener_fecha_selector(fecha_inicio)) or datetime.now().date()
                inicio_date = base - timedelta(days=base.weekday())
                fin_date = inicio_date + timedelta(days=6)
                establecer_fecha_selector(fecha_inicio, inicio_date)
                establecer_fecha_selector(fecha_fin, fin_date)
            elif modo == "Mensual":
                base = parse_fecha_ddmmyyyy(obtener_fecha_selector(fecha_inicio)) or datetime.now().date()
                inicio_date = base.replace(day=1)
                if inicio_date.month == 12:
                    siguiente = date(inicio_date.year + 1, 1, 1)
                else:
                    siguiente = date(inicio_date.year, inicio_date.month + 1, 1)
                establecer_fecha_selector(fecha_inicio, inicio_date)
                establecer_fecha_selector(fecha_fin, date(siguiente.year, siguiente.month, 1) - timedelta(days=1))
            elif modo == "Anual":
                base = parse_fecha_ddmmyyyy(obtener_fecha_selector(fecha_inicio)) or datetime.now().date()
                inicio_date = date(base.year, 1, 1)
                fin_date = date(base.year, 12, 31)
                establecer_fecha_selector(fecha_inicio, inicio_date)
                establecer_fecha_selector(fecha_fin, fin_date)
            else:
                periodo_var.set("Rango")
            _mark_snapshot_stale()

        def _marcar_rango(*_):
            periodo_var.set("Rango")
            _mark_snapshot_stale()

        combo_periodo.bind("<<ComboboxSelected>>", _actualizar_fechas_por_periodo)
        fecha_inicio.bind("<<DateEntrySelected>>", _marcar_rango)
        fecha_inicio.bind("<KeyRelease>", _marcar_rango)
        fecha_fin.bind("<<DateEntrySelected>>", _marcar_rango)
        fecha_fin.bind("<KeyRelease>", _marcar_rango)
        ars_search_var.changed.connect(lambda _value: _refresh_ars_search())
        for filter_variable in (
            periodo_var,
            turno_var,
            especialidad_var,
            cobertura_var,
            ars_mode_var,
            *ars_vars.values(),
        ):
            filter_variable.changed.connect(lambda _value: _mark_snapshot_stale())
        button_row = tb.Frame(barra, style="Card.TFrame")
        button_row.pack(fill="x")
        tb.Button(button_row, text="📊  Generar reporte", bootstyle=PRIMARY, command=generar, width=20).pack(side="left", padx=5, ipady=5)
        pdf_button = tb.Button(button_row, text="📄  Crear / abrir PDF", bootstyle=SUCCESS, command=guardar_pdf, width=20)
        pdf_button.pack(side="left", padx=5, ipady=5)
        excel_button = tb.Button(button_row, text="📗  Exportar Excel", bootstyle=INFO, command=guardar_excel_reporte, width=18)
        excel_button.pack(side="left", padx=5, ipady=5)
        _set_export_enabled(False)
        tb.Button(button_row, text="Limpiar filtros", bootstyle=SECONDARY, command=limpiar_filtros, width=16).pack(side="left", padx=5, ipady=5)
        tb.Button(button_row, text="Cerrar", bootstyle=SECONDARY, command=win.destroy, width=12).pack(side="right", padx=5, ipady=5)
        tb.Label(barra, textvariable=estado_var, style="Muted.TLabel").pack(
            fill="x", padx=5, pady=(6, 0)
        )
        win.report_controls = {
            "period": combo_periodo,
            "turn": combo_turno,
            "specialty_var": especialidad_var,
            "coverage_var": cobertura_var,
            "ars_mode": ars_mode_combo,
            "ars_search": ars_search,
            "ars_vars": ars_vars,
            "cards": card_vars,
            "preview": vista,
            "filters_frame": filtros,
            "filter_panel": panel,
            "ars_panel": ars_panel,
            "ars_canvas": ars_canvas,
            "cards_frame": cards,
            "preview_card": preview_card,
            "generate": generar,
            "clear": limpiar_filtros,
            "pdf_button": pdf_button,
            "excel_button": excel_button,
            "state": ultimo,
        }


    def _generar_reporte_del_dia_rapido(self):
        try:
            snapshot = dict(self.db.get_operational_station_snapshot() or {})
            source_id = str(snapshot.get("operational_source_id") or "").strip()
            turn_id = int(snapshot.get("turn_id") or 0)
            if not source_id or turn_id <= 0:
                messagebox.showwarning(
                    "Reporte",
                    "No hay una identidad de turno central disponible.",
                    parent=self.root,
                )
                return
            period = build_turn_operational_period(
                snapshot.get("turn_started_at"),
                fallback_date=fecha_base_operativa_actual(),
            )

            def _trabajo():
                source = self.db.load_statistical_report_source(
                    operational_source_id=source_id,
                    turn_scope="Turno actual",
                    current_turn_id=turn_id,
                    start_at=period.start_at,
                    end_at=period.end_at,
                )
                selected_turn = source.get("selected_turn") or {}
                effective_period = build_turn_operational_period(
                    selected_turn.get("started_at"),
                    fallback_date=period.start_at.date(),
                )
                turn_label = (
                    "Turno actual · "
                    f"{effective_period.start_at:%d/%m/%Y %I:%M %p} → "
                    f"{effective_period.end_at:%d/%m/%Y %I:%M %p}"
                )
                filters = AdmissionReportFilters(
                    start_at=effective_period.start_at,
                    end_at=effective_period.end_at,
                    period_label=effective_period.label,
                    turn_label=turn_label,
                    operational_source_id=source_id,
                    turn_id=turn_id,
                )
                dataset = build_admission_report_dataset(
                    source.get("records") or (),
                    filters,
                    turns=(selected_turn,),
                )
                if not dataset.records:
                    return "", dataset
                return crear_pdf_reporte(dataset.summary), dataset

            def _ok(result):
                ruta, dataset = result
                if not ruta:
                    messagebox.showinfo(
                        "Sin registros",
                        "No hay registros para el turno operacional actual.",
                        parent=self.root,
                    )
                    self.set_status("Reporte omitido: turno sin pacientes", "warning")
                    return
                abrir_pdf(ruta)
                self.set_status(
                    f"Reporte generado · {len(dataset.records)} paciente(s)", "ok"
                )

            def _error(error):
                messagebox.showerror(
                    "Error",
                    f"No se pudo generar el reporte del turno:\n{error}",
                    parent=self.root,
                )

            self._ejecutar_en_segundo_plano(
                "Generando reporte del turno…", _trabajo, _ok, _error
            )
        except Exception as e:
            messagebox.showerror("Error", f"No se pudo generar el reporte del día:\n{str(e)}")

    def _obtener_catalogo_ars(self):
        catalogo = set()
        try:
            for k, v in SEGUROS_DISPLAY.items():
                if k and k != "SIN SEGURO":
                    catalogo.add(seguro_para_mostrar(k))
                    catalogo.add(k)
                if v and v != "SIN SEGURO":
                    catalogo.add(v)
        except Exception:
            pass
        try:
            for a in self.db.listar_ars_distintas():
                if a and a != "(Todas)":
                    catalogo.add(seguro_para_mostrar(a))
                    catalogo.add(a)
        except Exception:
            pass
        canonicos = [
            "SENASA SUBSIDIADO", "SENASA CONTRIBUTIVO", "SENASA PENSIONADOS",
            "MAPFRE/PALIC", "HUMANO", "PRIMERA", "SEMMA", "RENACER",
            "RESERVAS", "ASEMAP", "UNIVERSAL", "MONUMENTAL", "ABEL GONZALEZ/SIMAG",
            "METASALUD"
        ]
        catalogo.update(canonicos)
        return sorted({c.strip() for c in catalogo if c and c.strip()})

    def _configurar_label_deteccion(self, label, estado="neutral"):
        try:
            colores = {
                "ok": "#72E39B",
                "warning": "#F6B860",
                "error": "#FF7A7A",
                "neutral": "#8FA6BF",
            }
            label.configure(foreground=colores.get(estado, colores["neutral"]))
        except Exception:
            pass

    def _on_ars_keyrelease(self, event=None):
        self._actualizar_deteccion_seguro()
        return self._actualizar_sugerencias_ars(event)

    def _actualizar_deteccion_seguro(self, event=None):
        try:
            ars_txt = (self.entry_ars.get() or "").strip()
            nss_txt = (self.entry_nss.get() or "").strip().upper()
        except Exception:
            return

        try:
            if ars_txt:
                canon_alias = _mejor_seguro_por_similitud(ars_txt)
                if not canon_alias:
                    canon_alias = normalizar_seguro(ars_txt, nss_txt)
            else:
                canon_alias = "SIN SEGURO" if nss_txt in ["SIN SEGURO", "NO", "N/S", "NS", "N\\S"] else ""

            if not ars_txt and not nss_txt:
                self.ars_detectado_var.set("Detectado como: pendiente")
                self._configurar_label_deteccion(self.ars_detectado_label, "neutral")
            else:
                display = seguro_para_mostrar(canon_alias or "SIN SEGURO")
                self.ars_detectado_var.set(f"Detectado como: {display}")

                if display == "SIN SEGURO":
                    self._configurar_label_deteccion(self.ars_detectado_label, "warning")
                else:
                    self._configurar_label_deteccion(self.ars_detectado_label, "ok")

            if not nss_txt:
                if ars_txt and canon_alias and canon_alias != "SIN SEGURO":
                    self.nss_detectado_var.set("NSS: pendiente; sin NSS se guardará como SIN SEGURO")
                    self._configurar_label_deteccion(self.nss_detectado_label, "warning")
                else:
                    self.nss_detectado_var.set("NSS: pendiente")
                    self._configurar_label_deteccion(self.nss_detectado_label, "neutral")
            elif nss_txt in ["SIN SEGURO", "NO", "N/S", "NS", "N\\S"]:
                self.nss_detectado_var.set("NSS: marcado como SIN SEGURO")
                self._configurar_label_deteccion(self.nss_detectado_label, "warning")
            elif not nss_txt.isdigit():
                self.nss_detectado_var.set("NSS: inválido; debe ser numérico o SIN SEGURO")
                self._configurar_label_deteccion(self.nss_detectado_label, "error")
            elif is_all_zeros(nss_txt):
                self.nss_detectado_var.set("NSS: inválido; no puede ser todo ceros")
                self._configurar_label_deteccion(self.nss_detectado_label, "error")
            else:
                self.nss_detectado_var.set("NSS: válido")
                self._configurar_label_deteccion(self.nss_detectado_label, "ok")

        except Exception:
            pass

    def _actualizar_sugerencias_ars(self, event=None):
        texto = (self.entry_ars.get() or "").strip()
        if not texto:
            self._ocultar_sugerencias_ars()
            return
        q = _limpiar_texto_seguro(texto)
        resultados = []
        for opcion in self._ars_catalogo:
            limpio = _limpiar_texto_seguro(opcion)
            if not limpio:
                continue
            score = SequenceMatcher(None, q, limpio).ratio()
            if q in limpio or limpio.startswith(q) or score >= 0.45:
                resultados.append((score + (0.4 if q in limpio else 0), opcion))
        resultados = [op for _, op in sorted(resultados, reverse=True)[:8]]
        self.ars_suggestions.delete(0, tk.END)
        if not resultados:
            self._ocultar_sugerencias_ars()
            return
        for item in resultados:
            self.ars_suggestions.insert(tk.END, item)
        self.ars_suggestions.grid()

    def _seleccionar_sugerencia_ars(self, event=None):
        try:
            idx = self.ars_suggestions.curselection()
            if not idx:
                return
            val = self.ars_suggestions.get(idx[0])
            self.entry_ars.delete(0, tk.END)
            self.entry_ars.insert(0, val)
            self._actualizar_deteccion_seguro()
            self._ocultar_sugerencias_ars()
            self.entry_nss.focus_set()
        except Exception:
            pass

    def _focus_sugerencias_ars(self, event=None):
        try:
            if self.ars_suggestions.winfo_ismapped() and self.ars_suggestions.size() > 0:
                self.ars_suggestions.focus_set()
                self.ars_suggestions.selection_clear(0, tk.END)
                self.ars_suggestions.selection_set(0)
                self.ars_suggestions.activate(0)
                return "break"
        except Exception:
            pass

    def _ocultar_sugerencias_ars(self):
        try:
            self.ars_suggestions.grid_remove()
        except Exception:
            pass

    def _capture_initial_styles(self):
        for e in self.all_entries:
            try:
                self._initial_styles[e] = {
                    "bootstyle": e.cget("bootstyle") if "bootstyle" in e.keys() else "",
                    "style": e.cget("style") if "style" in e.keys() else "",
                }
            except Exception:
                self._initial_styles[e] = {"bootstyle": "", "style": ""}

    def _restore_widget_style(self, widget):
        try:
            if widget in self._pending_restores and self._pending_restores[widget]:
                try:
                    self.root.after_cancel(self._pending_restores[widget])
                except Exception:
                    pass
                self._pending_restores[widget] = None
            init = self._initial_styles.get(widget, {"bootstyle": "", "style": ""})
            try:
                widget.configure(bootstyle=init.get("bootstyle", ""))
            except Exception:
                pass
            try:
                widget.configure(style=init.get("style", ""))
            except Exception:
                pass
            try:
                if widget in getattr(self, "all_entries", []):
                    self._aplicar_paridad_visual_inicio()
            except Exception:
                pass
        except Exception:
            pass

    def _restore_all_styles(self):
        for w, after_id in list(self._pending_restores.items()):
            if after_id:
                try:
                    self.root.after_cancel(after_id)
                except Exception:
                    pass
                self._pending_restores[w] = None
        for e in self.all_entries:
            self._restore_widget_style(e)
        try:
            self._aplicar_paridad_visual_inicio()
        except Exception:
            pass
        self.root.update_idletasks()

    def _ventana_activa(self, ventana):
        try:
            return ventana is not None and ventana.winfo_exists()
        except Exception:
            return False

    def _enfocar_ventana(self, ventana):
        try:
            ventana.deiconify()
            ventana.lift()
            ventana.focus_force()
        except Exception:
            pass

    def _crear_toplevel_estable(self, titulo, geometry, attr_name):
        ventana_existente = getattr(self, attr_name, None)
        if self._ventana_activa(ventana_existente):
            self._enfocar_ventana(ventana_existente)
            return None

        win = Toplevel(self.root)
        win.title(titulo)
        win.geometry(geometry)
        win.configure(bg=self._paleta_visual_actual()["root"])
        try:
            base_geo = geometry.split("+")[0]
            mw, mh = [int(x) for x in base_geo.lower().split("x")[:2]]
            win.minsize(min(max(mw, 760), 1180), min(max(mh, 460), 800))
        except Exception:
            try:
                win.minsize(900, 480)
            except Exception:
                pass
        win.transient(self.root)
        win.bind("<Escape>", lambda e: _cerrar())

        def _cerrar():
            try:
                setattr(self, attr_name, None)
                win.destroy()
            except Exception:
                setattr(self, attr_name, None)

        win.protocol("WM_DELETE_WINDOW", _cerrar)
        setattr(self, attr_name, win)
        try:
            win.lift()
            win.focus_set()
            win.after(10, lambda w=win: (w.lift(), w.focus_set()))
        except Exception:
            pass
        return win

    @staticmethod
    def _calcular_tamano_configuracion(ancho_disponible, alto_disponible):
        """Return a centered 80% window that remains usable at 1366x768."""
        ancho = max(1, int(ancho_disponible or 1))
        alto = max(1, int(alto_disponible or 1))
        return (
            min(ancho - 24, max(920, round(ancho * 0.80))),
            min(alto - 40, max(600, round(alto * 0.80))),
        )

    def _aplicar_tamano_configuracion(self, win):
        screen_width = win.winfo_screenwidth()
        screen_height = win.winfo_screenheight()
        vroot_width = getattr(win, "winfo_vrootwidth", None)
        vroot_height = getattr(win, "winfo_vrootheight", None)
        disponible_w = int(
            (vroot_width() if callable(vroot_width) else 0) or screen_width
        )
        disponible_h = int(
            (vroot_height() if callable(vroot_height) else 0) or screen_height
        )
        ancho, alto = self._calcular_tamano_configuracion(
            disponible_w, disponible_h
        )
        vroot_x = getattr(win, "winfo_vrootx", None)
        vroot_y = getattr(win, "winfo_vrooty", None)
        origen_x = int(vroot_x() if callable(vroot_x) else 0)
        origen_y = int(vroot_y() if callable(vroot_y) else 0)
        x = origen_x + max(0, (disponible_w - ancho) // 2)
        y = origen_y + max(0, (disponible_h - alto) // 2)
        win.geometry(f"{ancho}x{alto}+{x}+{y}")
        win.minsize(min(920, ancho), min(600, alto))

    def _crear_header_ventana(self, parent, titulo, subtitulo="", icono="▣"):
        pal = self._paleta_visual_actual()
        header = tb.Frame(parent, padding=(12, 10), style="Card.TFrame")
        header.pack(fill="x", pady=(0, 12))
        header.columnconfigure(1, weight=1)
        header.rowconfigure(0, minsize=32)

        icon_label = tb.Label(
            header,
            text=icono,
            font=("Segoe UI Symbol", 24, "bold"),
            foreground=pal["accent"],
            background=pal["card"]
        )
        icon_label.grid(
            row=0, column=0, rowspan=2, sticky="nsw", padx=(2, 14), pady=2
        )

        title_label = tb.Label(
            header,
            text=titulo,
            font=("Segoe UI", 16, "bold"),
            foreground=pal["text"],
            background=pal["card"],
            wraplength=760,
            justify="left",
            anchor="w"
        )
        title_label.grid(row=0, column=1, sticky="ew")

        subtitle_label = None
        if subtitulo:
            subtitle_label = tb.Label(
                header,
                text=subtitulo,
                style="Muted.TLabel",
                background=pal["card"],
                wraplength=760,
                justify="left",
                anchor="w"
            )
            subtitle_label.grid(row=1, column=1, sticky="ew", pady=(3, 0))

        def ajustar_header(event):
            wrap = max(320, int(event.width) - 100)
            title_label.configure(wraplength=wrap)
            if subtitle_label is not None:
                subtitle_label.configure(wraplength=wrap)

        header.bind("<Configure>", ajustar_header, add="+")
        return header

    def _crear_card(self, parent, padding=12):
        return tb.Frame(parent, padding=padding, style="Card.TFrame")

    def _mostrar_dialogo_modal_unico(self, titulo, mensaje):
        if self._ventana_activa(self.dialogo_unico_win):
            self._enfocar_ventana(self.dialogo_unico_win)
            return

        win = Toplevel(self.root)
        self.dialogo_unico_win = win
        win.title(titulo)
        win.resizable(False, False)
        win.transient(self.root)
        win.grab_set()

        win.bind("<Escape>", lambda e: self._cerrar_dialogo_modal_unico())

        frame = tb.Frame(win, padding=16)
        frame.pack(fill="both", expand=True)

        tb.Label(frame, text=mensaje, wraplength=360, justify="left").pack(anchor="w", pady=(0, 14))
        tb.Button(frame, text="Cerrar", bootstyle=PRIMARY, command=lambda: self._cerrar_dialogo_modal_unico()).pack()

        def on_close():
            self._cerrar_dialogo_modal_unico()

        win.protocol("WM_DELETE_WINDOW", on_close)
        win.update_idletasks()

        w = win.winfo_width()
        h = win.winfo_height()
        x = self.root.winfo_rootx() + (self.root.winfo_width() // 2) - (w // 2)
        y = self.root.winfo_rooty() + (self.root.winfo_height() // 2) - (h // 2)
        try:
            win.geometry(f"+{max(x, 0)}+{max(y, 0)}")
        except Exception:
            pass

        try:
            win.focus_force()
        except Exception:
            pass

    def _cerrar_dialogo_modal_unico(self):
        try:
            if self.dialogo_unico_win and self.dialogo_unico_win.winfo_exists():
                self.dialogo_unico_win.grab_release()
                self.dialogo_unico_win.destroy()
        except Exception:
            pass
        finally:
            self.dialogo_unico_win = None

    def _register_temp(self, path: str):
        if path and os.path.exists(path):
            self._temp_files.add(path)

    def _cleanup_temp_files(self):
        for p in list(self._temp_files):
            try:
                if os.path.exists(p):
                    os.remove(p)
            except Exception:
                pass
            finally:
                self._temp_files.discard(p)

    def shutdown(self):
        if self._shutdown_complete:
            return
        self._shutdown_complete = True
        self.deactivate()
        if self._session_after_id:
            try:
                self.root.after_cancel(self._session_after_id)
            except Exception:
                pass
        self._session_after_id = None
        if self._excel_export_after_id:
            try:
                self.root.after_cancel(self._excel_export_after_id)
            except Exception:
                pass
        self._excel_export_after_id = None
        try:
            self.root.close_owned_toplevels()
        except Exception:
            pass
        self._cleanup_temp_files()
        try:
            self.root.destroy()
        except Exception:
            pass

    def on_close(self):
        self.shutdown()

    def activate(self):
        """Resume work that is useful only while the embedded page is visible."""
        if self._shutdown_complete:
            return
        was_visible = self._admission_visible
        self._admission_visible = True
        self._programar_modo_responsivo()
        if not was_visible:
            self._date_after_id = self.root.after(0, self._actualizar_fecha_actual)
            self._summary_after_id = self.root.after(0, self._programar_refresco_resumen_en_vivo)
        self._retry_excel_export_jobs()

    def deactivate(self):
        """Pause presentation timers without ending session or application."""
        self._admission_visible = False
        for attr_name in ("_responsive_after_id", "_date_after_id", "_summary_after_id"):
            timer_id = getattr(self, attr_name, None)
            if timer_id:
                try:
                    self.root.after_cancel(timer_id)
                except Exception:
                    pass
            setattr(self, attr_name, None)

    def _highlight_error(self, widget, mensaje: str):
        if widget not in self._initial_styles:
            try:
                self._initial_styles[widget] = {
                    "bootstyle": widget.cget("bootstyle") if "bootstyle" in widget.keys() else "",
                    "style": widget.cget("style") if "style" in widget.keys() else "",
                }
            except Exception:
                self._initial_styles[widget] = {"bootstyle": "", "style": ""}

        try:
            widget.configure(bootstyle=DANGER)
        except Exception:
            try:
                widget.configure(style="danger.TEntry")
            except Exception:
                pass

        widget.focus_set()
        self.set_status(f"Corrija el campo marcado: {mensaje}", "error")
        self._mostrar_notificacion(mensaje, autohide_ms=9000, tipo="error")

        try:
            after_id = self.root.after(2000, lambda w=widget: self._restore_widget_style(w))
            self._pending_restores[widget] = after_id
        except Exception:
            pass

    def _ocultar_notificacion(self):
        if self._notif_after_id:
            try:
                self.root.after_cancel(self._notif_after_id)
            except Exception:
                pass
            self._notif_after_id = None
        self.notif_frame.pack_forget()
        self.notif_label.config(text="")
        self.btn_deshacer.config(command=lambda: None)

    def _push_undo_action(self, descripcion: str, callback):
        if not callable(callback):
            return
        try:
            self._undo_stack.append({"descripcion": descripcion or "última acción", "callback": callback})
            if len(self._undo_stack) > getattr(self, "_undo_limit", 20):
                self._undo_stack = self._undo_stack[-self._undo_limit:]
        except Exception:
            self._undo_stack = [{"descripcion": descripcion or "última acción", "callback": callback}]

    def _on_field_focus_in(self, event):
        """
        FASE 9B: Captura el valor del campo al recibir foco.
        """
        try:
            self._field_focus_started_at = _time.perf_counter()
            w = event.widget
            if isinstance(w, (tk.Entry, tb.Entry)):
                w._undo_prev_value = w.get()
                w._undo_has_prev = True
        except Exception:
            pass

    def _on_field_focus_out(self, event):
        """
        FASE 9B: Si el valor del campo cambió, registra undo de campo.
        """
        try:
            w = event.widget
            if not isinstance(w, (tk.Entry, tb.Entry)):
                return
            if not getattr(w, "_undo_has_prev", False):
                return
            prev = getattr(w, "_undo_prev_value", None)
            curr = w.get()
            if prev is not None and prev != curr:
                def _undo_campo(ww=w, valor_anterior=prev):
                    try:
                        ww.delete(0, tk.END)
                        ww.insert(0, valor_anterior)
                        ww.focus_set()
                    except Exception:
                        pass
                    return "break"

                nombre_campo = "campo"
                try:
                    mapping = {
                        "entry_nombre": "Nombre",
                        "entry_edad": "Edad",
                        "entry_cedula": "Cédula",
                        "entry_telefono": "Teléfono",
                        "entry_direccion": "Dirección",
                        "entry_nacionalidad": "Nacionalidad",
                        "entry_ars": "ARS",
                        "entry_nss": "NSS",
                    }
                    nombre_campo = mapping.get(str(w), "campo")
                except Exception:
                    pass

                self._push_undo_action(f"cambio en {nombre_campo}", _undo_campo)
            w._undo_has_prev = False
        except Exception:
            pass

    def _undo_last_action(self, event=None):
        """
        FASE 9: Deshacer global con Ctrl+Z.
        - Restaura último campo modificado / texto eliminado / formulario limpiado.
        - Devuelve el foco al widget original cuando aplica.
        """
        try:
            if not getattr(self, "_undo_stack", None):
                self.set_status("No hay acciones para deshacer", "warning")
                try:
                    self.root.bell()
                except Exception:
                    pass
                return "break"

            accion = self._undo_stack.pop()
            callback = accion.get("callback")
            descripcion = accion.get("descripcion", "última acción")

            try:
                foco_antes = self.root.focus_get()
            except Exception:
                foco_antes = None

            if callable(callback):
                try:
                    resultado = callback()
                except Exception as e:
                    self.set_status(f"Error al deshacer: {e}", "error")
                    return "break"

                try:
                    foco_actual = self.root.focus_get()
                except Exception:
                    foco_actual = None

                if foco_actual in (None, "") and foco_antes is not None:
                    try:
                        if foco_antes.winfo_exists():
                            foco_antes.focus_set()
                    except Exception:
                        pass

                try:
                    self._invalidar_caches_datos()
                except Exception:
                    pass
                self.set_status(f"Deshecho: {descripcion}", "ok")
                self._mostrar_notificacion(
                    f"Deshecho: {descripcion}",
                    on_undo=None,
                    autohide_ms=3000
                )
            return "break"
        except Exception as e:
            self.set_status(f"No se pudo deshacer: {str(e)}", "error")
            messagebox.showerror("Deshacer", f"No se pudo deshacer la última acción:\n{str(e)}")
            return "break"


    def _mostrar_notificacion(self, texto: str, on_undo=None, autohide_ms=7000, tipo="info"):
        colores = {
            "info": self._paleta_visual_actual()["text"],
            "ok": "#72E39B",
            "warning": "#FFD166",
            "error": "#FF8A80",
        }
        self.notif_label.config(text=texto, foreground=colores.get(tipo, colores["info"]))
        if on_undo:
            self.btn_deshacer.config(command=lambda: (on_undo(), self._ocultar_notificacion()))
            self.btn_deshacer.pack(side="right", padx=6)
        else:
            self.btn_deshacer.pack_forget()
        self.notif_frame.pack(fill="x", side="bottom")
        if self._notif_after_id:
            try:
                self.root.after_cancel(self._notif_after_id)
            except Exception:
                pass
        self._notif_after_id = self.root.after(autohide_ms, self._ocultar_notificacion)

    def validar_numerico(self, value, tipo):
        if value == "":
            return True
        if tipo == 'cedula':
            return value.isdigit() and len(value) <= 11
        elif tipo == 'telefono':
            return value.isdigit() and len(value) <= 10
        return True

    def limitar_caracteres(self, entry, max_len):
        current = entry.get()
        if len(current) > max_len:
            entry.delete(max_len, tk.END)

    def actualizar_embarazada(self):
        try:
            es_femenino = (self.var_sexo.get() == "Femenino")
            if es_femenino:
                if not self.check_embarazada.winfo_ismapped():
                    self.check_embarazada.pack(side="left", padx=(10, 0))
            else:
                if self.check_embarazada.winfo_ismapped():
                    self.check_embarazada.pack_forget()
                self.var_embarazada.set(False)
        except Exception:
            pass

    def determinar_hoja(self):
        if self.var_sexo.get() == "Femenino" and self.var_embarazada.get():
            return "GINECOLOGIA"

        try:
            edad = int(self.entry_edad.get())
            unidad = self.unidad_edad.get()[0].upper()
        except Exception:
            return None

        if unidad == "D":
            meses = edad / 30
        elif unidad == "M":
            meses = edad
        else:
            meses = edad * 12

        if meses <= 180:
            return "PEDIATRIA"
        return "GENERAL"

    def _registro_afecta_excel_turno(self, atencion):
        if not atencion:
            return False
        if (atencion.get("tipo_atencion") or "EMERGENCIA").strip().upper() in ("URGENCIA", "CONSULTA"):
            return False
        return self._registro_esta_en_turno_actual(atencion)

    def _cambio_requiere_reconstruir_excel(self, antes: dict, despues: dict) -> bool:
        """
        FASE 8: Reconstruir Excel SOLO si cambia un campo crítico.
        Campos críticos: nombre, hoja, ARS, tipo_atencion, fecha.
        Campos NO críticos: telefono, direccion, sexo, nacionalidad, cedula, edad, nss, hora.
        URGENCIA nunca reconstruye.
        """
        if not antes and not despues:
            return False

        antes = antes or {}
        despues = despues or {}

        tipo_despues = str(
            despues.get("TipoAtencion", despues.get("tipo_atencion",
            antes.get("tipo_atencion", "EMERGENCIA"))) or "EMERGENCIA"
        ).strip().upper()

        if tipo_despues in ("URGENCIA", "CONSULTA"):
            return False

        campos_clave = [
            ("nombre", "Nombre"),
            ("hoja",   "Hoja"),
            ("ars",    "Aseguradora (ARS)"),
            ("tipo_atencion", "TipoAtencion"),
            ("fecha",  "Fecha"),
        ]

        for k_antes, k_despues in campos_clave:
            a = str(antes.get(k_antes, "") or "").strip().upper()
            d = str(despues.get(k_despues, despues.get(k_antes, "")) or "").strip().upper()
            if k_antes == "ars":
                a = normalizar_seguro(a, antes.get("nss", ""))
                d = normalizar_seguro(d, despues.get("NSS", despues.get("nss", "")))
            if a != d:
                return True

        if self._registro_afecta_excel_turno(antes):
            return True

        try:
            tmp = dict(antes)
            tmp.update({
                "fecha": despues.get("Fecha", despues.get("fecha", antes.get("fecha", ""))),
                "hora":  despues.get("Hora",  despues.get("hora",  antes.get("hora",  ""))),
                "tipo_atencion": tipo_despues,
            })
            if self._registro_afecta_excel_turno(tmp):
                return True
        except Exception:
            pass

        return False

    def _reconstruir_excel_si_necesario(self, razon="", antes=None, despues=None, forzar=False):
        try:
            turno_cfg = cargar_turno_config()
            if not turno_cfg:
                return False

            if not forzar:
                if antes is not None or despues is not None:
                    if not self._cambio_requiere_reconstruir_excel(antes or {}, despues or {}):
                        return False

            reconstruir_excel_turno(self.db, turno_cfg)
            self._refrescar_resumen_en_vivo()
            return True
        except PermissionError:
            self.set_status("Excel abierto. Cierre el listado para actualizarlo.", "warning")
            raise
        except Exception as e:
            self.set_status(f"Aviso al actualizar Excel: {e}", "warning")
            return False

    def _registro_esta_en_turno_actual(self, atencion):
        if not atencion:
            return False
        snapshot_provider = getattr(self.db, "get_operational_station_snapshot", None)
        if callable(snapshot_provider):
            current_turn_id, current_source_id = _identidad_turno_central(self.db)
            if current_turn_id <= 0 or not current_source_id:
                return False
            current = dict(atencion)
            if (
                current.get("operational_turn_id") is None
                or not current.get("operational_source_id")
            ) and current.get("id"):
                current = dict(
                    self.db.obtener_atencion_por_id(int(current["id"])) or current
                )
            try:
                return (
                    int(current.get("operational_turn_id")) == current_turn_id
                    and str(current.get("operational_source_id") or "")
                    == current_source_id
                )
            except (TypeError, ValueError):
                return False
        turno_cfg = cargar_turno_config()
        if not turno_cfg:
            return False
        contexto = self.db.buscar_contexto_turno_existente(turno_cfg)
        if not contexto:
            return False
        return int(atencion.get("turno_id") or 0) == int(contexto["turno_id"])

    def _generar_y_abrir_reporte_turno(
        self,
        turno_cfg,
        fin_corte=None,
        *,
        turn_id=None,
        operational_source_id=None,
    ):
        if not turno_cfg:
            return ""

        resumen = construir_resumen_turno(
            self.db,
            turno_cfg,
            fin_override=fin_corte or datetime.now(),
            turn_id=turn_id,
            operational_source_id=operational_source_id,
        )
        if not resumen:
            self.set_status("No se pudo generar el reporte: no se encontraron datos del turno.", "warning")
            return ""
        if reportable_patient_count(resumen) == 0:
            contexto = self.db.buscar_contexto_turno_existente(turno_cfg) or {}
            APP_LOG.info(
                "SHIFT_REPORT_SKIPPED_EMPTY turn_id=%s patient_count=0",
                int(contexto.get("turno_id") or 0),
            )
            self.set_status(
                "Turno cerrado correctamente. No se generÃ³ reporte porque el turno no contiene pacientes.",
                "ok",
            )
            return ""

        ruta = crear_pdf_reporte(resumen)

        try:
            guardar_copia_reporte_turno(ruta, turno_cfg)
        except Exception as e:
            messagebox.showwarning(
                "Archivo diario",
                f"El reporte se generó, pero no se pudo guardar la copia clasificada:\n{str(e)}"
            )

        abrir_pdf(ruta)
        if self.app_settings.get("auto_print", True) and bool(self.app_settings.get("print_auto_reporte_turno", True)):
            copias_reporte = max(1, int(self.app_settings.get("print_copies_reporte", 2) or 2))
            imprimir_pdf(ruta, copias=copias_reporte, mostrar_error=True)
            self.set_status(f"Reporte impreso {copias_reporte} vez/veces", "ok")
        else:
            self.set_status("Reporte generado (impresión automática desactivada)", "ok")
        return ruta

    @staticmethod
    def _patient_mapping(patient):
        try:
            return dict(patient or {})
        except (TypeError, ValueError):
            return {}

    @staticmethod
    def _replace_entry_value(widget, value):
        blocker = QSignalBlocker(widget)
        try:
            widget.delete(0, tk.END)
            widget.insert(0, str(value or ""))
        finally:
            del blocker

    def _capture_patient_dirty_fields(self):
        widgets = {
            "nombre": self.entry_nombre,
            "telefono": self.entry_telefono,
            "direccion": self.entry_direccion,
            "nacionalidad": self.entry_nacionalidad,
            "ars": self.entry_ars,
            "cedula": self.entry_cedula,
            "nss": self.entry_nss,
        }
        for field, widget in widgets.items():
            if field not in self._patient_autofill_baseline:
                continue
            current = str(widget.get() or "")
            if current != str(self._patient_autofill_baseline.get(field) or ""):
                self._patient_fields_modified_by_user.add(field)

    def _apply_patient_ui(
        self, patient, *, include_cedula=False, include_nss=False,
        respect_dirty=False,
    ):
        """Aplica un snapshot sin cascadas; un resultado cloud no pisa edición humana."""
        data = self._patient_mapping(patient)
        if not data:
            return False
        if respect_dirty:
            self._capture_patient_dirty_fields()
        fields = {
            "nombre": (self.entry_nombre, data.get("nombre")),
            "telefono": (self.entry_telefono, data.get("telefono")),
            "direccion": (self.entry_direccion, data.get("direccion")),
            "nacionalidad": (self.entry_nacionalidad, data.get("nacionalidad")),
            "ars": (self.entry_ars, seguro_para_mostrar(data.get("ars") or "")),
        }
        if include_cedula:
            fields["cedula"] = (self.entry_cedula, data.get("cedula"))
        if include_nss:
            fields["nss"] = (self.entry_nss, str(data.get("nss") or "").upper())
        self._suspend_autocomplete = True
        try:
            for field, (widget, value) in fields.items():
                if respect_dirty and field in self._patient_fields_modified_by_user:
                    continue
                self._replace_entry_value(widget, value)
        finally:
            self._suspend_autocomplete = False
        self._patient_autofill_baseline = {
            field: str(widget.get() or "") for field, (widget, _value) in fields.items()
        }
        self._actualizar_deteccion_seguro()
        return True

    def _schedule_cloud_patient_lookup(self, kind, value):
        runtime = getattr(self.db, "_runtime", None)
        token = (str(kind), str(value))
        if runtime is None or bool(getattr(runtime, "offline", False)):
            return
        if token in self._autofill_cloud_pending:
            return
        self._autofill_cloud_pending.add(token)

        def worker():
            started = _time.perf_counter()
            try:
                result = runtime.verify_patient_with_cloud(
                    cedula=value if kind == "CEDULA" else "",
                    nss=value if kind == "NSS" else "",
                    timeout_ms=1500,
                )
                error = None
            except Exception as exc:
                result, error = None, exc
            elapsed_ms = (_time.perf_counter() - started) * 1000.0

            def finish():
                self._autofill_cloud_pending.discard(token)
                APP_LOG.info(
                    "PATIENT_CLOUD_VERIFY kind=%s CLOUD_VERIFY_MS=%.3f success=%s",
                    kind, elapsed_ms, bool(result) and error is None,
                )
                if error is not None:
                    APP_LOG.warning(
                        "PATIENT_AUTOFILL_BACKGROUND_FAILED kind=%s error=%s",
                        kind, type(error).__name__,
                    )
                    return
                current = re.sub(
                    r"\D", "", self.entry_cedula.get() if kind == "CEDULA" else self.entry_nss.get()
                )
                if not result or current != value:
                    return
                cloud_revision = int(dict(result).get("server_revision") or 0)
                self._verified_cloud_patient = dict(result)
                self._verified_cloud_identity = token
                if cloud_revision >= int(self._local_patient_revision or 0):
                    self._apply_patient_ui(
                        result,
                        include_cedula=kind == "NSS",
                        include_nss=kind == "CEDULA",
                        respect_dirty=True,
                    )
                    self.set_status("✓ Datos del paciente verificados en nube", "ok")

            try:
                self.root.after(0, finish)
            except tk.TclError:
                pass

        threading.Thread(
            target=worker,
            daemon=True,
            name=f"patient-cloud-{kind.lower()}",
        ).start()

    def _autofill_patient(self, kind, value, *, input_method):
        if self._suspend_autocomplete:
            return None
        identity = (str(kind), str(value))
        if identity == self._last_autofill_identity and (
            _time.perf_counter() - self._last_autofill_at
        ) < 1.5:
            return None
        started = _time.perf_counter()
        normalized_ms = (_time.perf_counter() - started) * 1000.0
        lookup_started = _time.perf_counter()
        try:
            patient = (
                self.db.buscar_paciente(value)
                if kind == "CEDULA"
                else self.db.buscar_por_nss(value)
            )
        except sqlite3.OperationalError as exc:
            patient = None
            APP_LOG.warning(
                "PATIENT_LOCAL_LOOKUP_BUSY kind=%s error=%s", kind, type(exc).__name__
            )
        local_ms = (_time.perf_counter() - lookup_started) * 1000.0
        self._patient_fields_modified_by_user.clear()
        self._verified_cloud_patient = None
        self._verified_cloud_identity = None
        self._local_patient_revision = int(
            self._patient_mapping(patient).get("server_revision") or 0
        )
        apply_started = _time.perf_counter()
        applied = self._apply_patient_ui(
            patient,
            include_cedula=kind == "NSS",
            include_nss=kind == "CEDULA",
        )
        apply_ms = (_time.perf_counter() - apply_started) * 1000.0
        self._last_autofill_identity = identity
        self._last_autofill_at = _time.perf_counter()
        total_ms = (self._last_autofill_at - started) * 1000.0
        focus_ms = (
            (self._last_autofill_at - self._field_focus_started_at) * 1000.0
            if self._field_focus_started_at else 0.0
        )
        APP_LOG.info(
            "PATIENT_LOOKUP input_method=%s input_type=%s source=LOCAL "
            "DOCUMENT_NORMALIZE_MS=%.3f LOCAL_QUERY_MS=%.3f FORM_APPLY_MS=%.3f "
            "FOCUS_CHANGE_MS=%.3f TOTAL_LOCAL_AUTOFILL_MS=%.3f found=%s",
            input_method, kind, normalized_ms, local_ms, apply_ms, focus_ms,
            total_ms, applied,
        )
        self._schedule_cloud_patient_lookup(kind, value)
        return patient

    def auto_completar(self, event=None, *, input_method="CLICK"):
        del event
        cedula = re.sub(r"\D", "", self.entry_cedula.get() or "")
        if len(cedula) != 11 or is_all_zeros(cedula):
            return None
        return self._autofill_patient("CEDULA", cedula, input_method=input_method)

    def auto_completar_por_nss(self, event=None, *, input_method="CLICK"):
        del event
        nss = re.sub(r"\D", "", self.entry_nss.get() or "")
        if not is_valid_nss_key(nss):
            return None
        return self._autofill_patient("NSS", nss, input_method=input_method)

    def _try_autocomplete_cedula(self):
        return self.auto_completar(input_method="ENTER")

    def _try_autocomplete_nss(self):
        return self.auto_completar_por_nss(input_method="ENTER")

    def _begin_final_patient_revalidation(self):
        cedula = re.sub(r"\D", "", self.entry_cedula.get() or "")
        nss = re.sub(r"\D", "", self.entry_nss.get() or "")
        kind, value = ("CEDULA", cedula) if len(cedula) == 11 else ("NSS", nss)
        token = (kind, value)
        if not value:
            return False
        if self._verified_cloud_identity == token and self._verified_cloud_patient:
            self._apply_patient_ui(
                self._verified_cloud_patient,
                include_cedula=kind == "NSS",
                include_nss=kind == "CEDULA",
                respect_dirty=True,
            )
            return False
        runtime = getattr(self.db, "_runtime", None)
        if runtime is None or bool(getattr(runtime, "offline", False)):
            APP_LOG.info("FINAL_PATIENT_REVALIDATION source=LOCAL_OFFLINE")
            return False
        if self._final_revalidation_in_progress:
            return True
        self._final_revalidation_in_progress = True
        self.set_status("Verificando datos finales del paciente…", "process")

        def worker():
            started = _time.perf_counter()
            try:
                result = runtime.verify_patient_with_cloud(
                    cedula=value if kind == "CEDULA" else "",
                    nss=value if kind == "NSS" else "",
                    timeout_ms=750,
                )
                error = None
            except Exception as exc:
                result, error = None, exc
            elapsed_ms = (_time.perf_counter() - started) * 1000.0

            def finish():
                self._final_revalidation_in_progress = False
                if result:
                    self._verified_cloud_patient = dict(result)
                    self._verified_cloud_identity = token
                    self._apply_patient_ui(
                        result,
                        include_cedula=kind == "NSS",
                        include_nss=kind == "CEDULA",
                        respect_dirty=True,
                    )
                APP_LOG.info(
                    "FINAL_PATIENT_REVALIDATION_MS=%.3f source=%s error=%s",
                    elapsed_ms, "CLOUD" if result else "LOCAL_FALLBACK",
                    type(error).__name__ if error else "",
                )
                self._final_revalidation_ready = True
                self.generar_pdf()

            try:
                self.root.after(0, finish)
            except tk.TclError:
                self._final_revalidation_in_progress = False

        threading.Thread(
            target=worker, daemon=True, name="patient-final-revalidation"
        ).start()
        return True

    def _guardar_formulario_actual(self):
        """
        FASE 9/10: Guarda estado completo del formulario + widget con foco.
        """
        try:
            try:
                widget_con_foco = self.root.focus_get()
            except Exception:
                widget_con_foco = None

            self._ultimo_formulario = {
                "nombre":      self.entry_nombre.get(),
                "edad":        self.entry_edad.get(),
                "edad_unit":   self.unidad_edad.get(),
                "cedula":      self.entry_cedula.get(),
                "telefono":    self.entry_telefono.get(),
                "direccion":   self.entry_direccion.get(),
                "nacionalidad":self.entry_nacionalidad.get(),
                "ars":         self.entry_ars.get(),
                "nss":         self.entry_nss.get(),
                "sexo":        self.var_sexo.get(),
                "embarazada":  self.var_embarazada.get(),
                "urgencia":    self.var_urgencia.get(),
                "hoja":        self.determinar_hoja() or "GENERAL",
                "_widget_foco": widget_con_foco,
            }
        except Exception as e:
            self.set_status(f"Aviso: no se pudo guardar formulario para Ctrl+Z: {e}", "warning")

    def _restaurar_formulario(self):
        """
        FASE 9/10: Restaura el formulario + devuelve el foco al widget original.
        """
        try:
            form = getattr(self, "_ultimo_formulario", None)
            if not form:
                self.set_status("No hay formulario anterior para restaurar", "warning")
                return

            self.entry_nombre.delete(0, tk.END);       self.entry_nombre.insert(0, form.get("nombre", ""))
            self.entry_edad.delete(0, tk.END);         self.entry_edad.insert(0, form.get("edad", ""))
            self.unidad_edad.set(form.get("edad_unit", "Años"))
            self.entry_cedula.delete(0, tk.END);       self.entry_cedula.insert(0, form.get("cedula", ""))
            self.entry_telefono.delete(0, tk.END);     self.entry_telefono.insert(0, form.get("telefono", ""))
            self.entry_direccion.delete(0, tk.END);    self.entry_direccion.insert(0, form.get("direccion", ""))
            self.entry_nacionalidad.delete(0, tk.END); self.entry_nacionalidad.insert(0, form.get("nacionalidad", ""))
            self.entry_ars.delete(0, tk.END);          self.entry_ars.insert(0, form.get("ars", ""))
            self.entry_nss.delete(0, tk.END);          self.entry_nss.insert(0, form.get("nss", ""))
            self.var_sexo.set(form.get("sexo") or "Femenino")
            self.var_embarazada.set(form.get("embarazada", False))
            self.var_urgencia.set(form.get("urgencia", False))

            self._restore_all_styles()
            try:
                self._actualizar_deteccion_seguro()
            except Exception:
                pass
            try:
                self.actualizar_embarazada()
            except Exception:
                pass

            try:
                widget_foco = form.get("_widget_foco")
                if widget_foco is not None and widget_foco.winfo_exists():
                    widget_foco.focus_set()
                else:
                    self.entry_nombre.focus_set()
            except Exception:
                self.entry_nombre.focus_set()

            self.set_status("✓  Formulario restaurado", "ok")
        except Exception as e:
            self.set_status(f"No se pudo restaurar formulario: {e}", "error")

    def limpiar_campos(self):
        """
        FASE 9/10: Antes de limpiar, guarda el estado y registra undo.
        """
        estado_previo = (
            (self.entry_nombre.get() or "").strip()
            + (self.entry_edad.get() or "").strip()
            + (self.entry_cedula.get() or "").strip()
            + (self.entry_telefono.get() or "").strip()
            + (self.entry_nss.get() or "").strip()
        )

        if estado_previo:
            self._guardar_formulario_actual()

            def _undo():
                self._restaurar_formulario()
            self._push_undo_action("restaurar formulario limpiado", _undo)

        self.entry_nombre.delete(0, tk.END)
        self.entry_edad.delete(0, tk.END)
        self.entry_cedula.delete(0, tk.END)
        self.entry_telefono.delete(0, tk.END)
        self.entry_direccion.delete(0, tk.END)
        self.entry_nacionalidad.delete(0, tk.END)
        self.entry_ars.delete(0, tk.END)
        self.entry_nss.delete(0, tk.END)
        self._patient_fields_modified_by_user.clear()
        self._patient_autofill_baseline = {}
        self._verified_cloud_patient = None
        self._verified_cloud_identity = None
        self._local_patient_revision = 0
        self._final_revalidation_ready = False
        self.var_sexo.set("Femenino")
        self.var_embarazada.set(False)
        try:
            self.var_urgencia.set(False)
        except Exception:
            pass
        self.unidad_edad.set("Años")
        try:
            self.ars_detectado_var.set("Detectado como: pendiente")
            self.nss_detectado_var.set("NSS: pendiente")
            self._configurar_label_deteccion(self.ars_detectado_label, "neutral")
            self._configurar_label_deteccion(self.nss_detectado_label, "neutral")
        except Exception:
            pass
        self.entry_nombre.focus_set()

    def _validar_campos_o_alertar(self):
        nombre = (self.entry_nombre.get() or "").strip()
        telefono = (self.entry_telefono.get() or "").strip()
        cedula = (self.entry_cedula.get() or "").strip()
        nss = (self.entry_nss.get() or "").strip().upper()
        direccion = (self.entry_direccion.get() or "").strip()
        nacionalidad = (self.entry_nacionalidad.get() or "").strip()
        ars = (self.entry_ars.get() or "").strip()
        sexo = self.var_sexo.get()

        if not nombre:
            self._highlight_error(self.entry_nombre, "El nombre es obligatorio.")
            return None

        if sexo not in ("Masculino", "Femenino"):
            self._highlight_error(
                self.lbl_sexo_m,
                "Seleccione explícitamente el sexo del paciente.",
            )
            return None

        if telefono:
            if not (telefono.isdigit() and len(telefono) == 10):
                self._highlight_error(self.entry_telefono, "El teléfono debe tener exactamente 10 dígitos numéricos.")
                return None
        elif not bool(self.app_settings.get("validation_allow_missing_phone", False)):
            self._highlight_error(self.entry_telefono, "El teléfono es obligatorio según las preferencias actuales.")
            return None

        edad_txt = (self.entry_edad.get() or "").strip()
        if edad_txt == "" or not edad_txt.isdigit():
            self._highlight_error(self.entry_edad, "La edad es obligatoria y debe ser un número entero.")
            return None
        edad = int(edad_txt)
        if edad < 0 or edad > 130:
            self._highlight_error(self.entry_edad, "La edad debe estar entre 0 y 130.")
            return None

        unidad = self.unidad_edad.get()
        if unidad not in ("Días", "Meses", "Años"):
            self._highlight_error(self.combo_unidad, "Selecciona una unidad de edad válida (Días/Meses/Años).")
            return None

        if cedula:
            if not (cedula.isdigit() and len(cedula) == 11):
                self._highlight_error(self.entry_cedula, "La cédula debe tener 11 dígitos.")
                return None
        elif not bool(self.app_settings.get("validation_allow_missing_cedula", True)):
            self._highlight_error(self.entry_cedula, "La cédula es obligatoria según las preferencias actuales.")
            return None

        if nss:
            if not nss.isdigit() and nss not in ["N/S", "N\\S", "NS", "NO", "SIN SEGURO"]:
                self._highlight_error(
                    self.entry_nss,
                    "El NSS debe ser numérico. Si el paciente no tiene seguro, deje el campo vacío o escriba SIN SEGURO."
                )
                return None
            if is_all_zeros(nss):
                self._highlight_error(self.entry_nss, "El NSS no puede ser todo ceros. Si no tiene seguro, deje el campo vacío o escriba SIN SEGURO.")
                return None

        if bool(self.app_settings.get("validation_block_short_ars", True)) and ars_es_corta_invalida(ars):
            self._highlight_error(
                self.entry_ars,
                "La ARS escrita es demasiado corta o no reconocida. Escriba al menos 4 caracteres o use una referencia válida como SUB, HUMANO, MAPFRE o SIN SEGURO."
            )
            return None

        invertido, msg_invertido = _detectar_campos_invertidos(
            self.entry_nss.get(), self.entry_ars.get()
        )
        if invertido:
            self._highlight_error(self.entry_nss, msg_invertido)
            return None

        hoja = self.determinar_hoja()
        if hoja is None:
            self._highlight_error(self.entry_edad, "Edad inválida para determinar la hoja.")
            return None

        avisos = []
        if bool(self.app_settings.get("validation_warn_nss_incomplete", True)):
            nss_digits = re.sub(r"\D", "", nss)
            if nss_digits and len(nss_digits) < 8:
                avisos.append("El NSS parece incompleto o demasiado corto.")

        ars_canon_tmp = normalizar_seguro(ars, nss)
        if bool(self.app_settings.get("validation_warn_ars_sin_seguro", True)) and ars_canon_tmp == "SIN SEGURO":
            avisos.append("La ARS fue detectada como SIN SEGURO.")

        if bool(self.app_settings.get("rn_warn", True)) and nombre_tiene_prefijo_rn(nombre):
            avisos.append("El nombre tiene prefijo RN-. Según preferencias, puede guardarse sin RN- en la base de datos.")

        if avisos:
            if not messagebox.askyesno("Revisión de datos", "Revise estos datos antes de continuar:\n\n- " + "\n- ".join(avisos) + "\n\n¿Desea continuar?"):
                return None

        ahora = datetime.now()
        hora_12 = ahora.strftime("%I:%M")
        am_pm = "AM" if ahora.hour < 12 else "PM"

        datos = {
            "Fecha": ahora.strftime("%d/%m/%Y"),
            "Hora": f"{hora_12} {am_pm}",
            "Nombre": nombre,
            "Sexo": sexo,
            "Edad_num": edad,
            "Unidad": unidad,
            "Cédula": cedula,
            "Teléfono": telefono,
            "Dirección": direccion,
            "Nacionalidad": nacionalidad,
            "Aseguradora (ARS)": ars,
            "NSS": nss,
            "TipoAtencion": "URGENCIA" if getattr(self, "var_urgencia", tk.BooleanVar(value=False)).get() else "EMERGENCIA",
        }
        return datos, hoja

    def _buscar_duplicado_turno_actual(self, datos):
        nss = (datos.get("NSS") or "").strip().upper()
        ced = (datos.get("Cédula") or "").strip()
        nombre = (datos.get("Nombre") or "").strip()
        telefono = (datos.get("Teléfono") or "").strip()
        if (
            not is_valid_nss_key(nss)
            and not is_valid_cedula_key(ced)
            and not (nombre and len(re.sub(r"\D", "", telefono)) == 10)
        ):
            return None

        turno_cfg = cargar_turno_config()
        if not turno_cfg:
            return None

        inicio_turno, fin_turno = obtener_rango_turno_efectivo(turno_cfg)
        contexto = self.db.obtener_contexto_turno(turno_cfg)
        return self.db.buscar_atencion_en_turno(
            nss,
            ced,
            inicio_turno,
            fin_turno,
            turno_id=contexto["turno_id"],
            dia_operativo_id=contexto["dia_operativo_id"],
            nombre=nombre,
            telefono=telefono,
        )

    def _dialogo_atencion_existente(self, atencion):
        resultado = {"accion": "cancelar"}
        win = Toplevel(self.root)
        win.title("Atención ya registrada")
        win.geometry("720x360")
        win.resizable(False, False)
        win.transient(self.root)
        win.grab_set()

        cont = tb.Frame(win, padding=20, style="Root.TFrame")
        cont.pack(fill="both", expand=True)
        tb.Label(
            cont,
            text="Este paciente ya tiene una hoja en el turno actual",
            font=("Arial", 15, "bold"),
            foreground="#FFFFFF",
        ).pack(anchor="w", pady=(0, 12))
        detalle = (
            f"Atención #{atencion.get('id')}\n"
            f"Paciente: {(atencion.get('nombre') or '').upper()}\n"
            f"Registrada: {atencion.get('fecha', '')} {atencion.get('hora', '')}\n"
            f"Especialidad: {atencion.get('hoja', '')}"
        )
        tb.Label(cont, text=detalle, justify="left", wraplength=670).pack(anchor="w", pady=(0, 18))
        tb.Label(
            cont,
            text="No se creará otra atención ni se reemplazará el historial.",
            bootstyle=WARNING,
        ).pack(anchor="w", pady=(0, 18))

        botones = tb.Frame(cont, style="Root.TFrame")
        botones.pack(fill="x", side="bottom")

        def elegir(accion):
            resultado["accion"] = accion
            win.destroy()

        tb.Button(botones, text="Reimprimir", bootstyle=SUCCESS, command=lambda: elegir("reimprimir")).pack(side="left", padx=(0, 8))
        tb.Button(botones, text="Abrir hoja", bootstyle=INFO, command=lambda: elegir("abrir")).pack(side="left", padx=(0, 8))
        tb.Button(botones, text="Editar", bootstyle=SECONDARY, command=lambda: elegir("editar")).pack(side="left", padx=(0, 8))
        tb.Button(
            botones,
            text="Registrar reingreso",
            bootstyle=WARNING,
            command=lambda: elegir("reingreso"),
        ).pack(side="left", padx=(0, 8))
        tb.Button(botones, text="Cancelar", command=lambda: elegir("cancelar")).pack(side="right")
        win.protocol("WM_DELETE_WINDOW", lambda: elegir("cancelar"))
        win.wait_window()
        return resultado["accion"]

    def _atender_duplicado_turno(self, atencion):
        accion = self._dialogo_atencion_existente(atencion)
        if accion == "editar":
            self._abrir_editor_atencion(int(atencion["id"]))
            return None
        if accion == "reingreso":
            actor = self._solicitar_autorizacion_admin(
                "AUTORIZAR_REINGRESO", parent=self.root, force=True
            )
            if not actor:
                return None
            motivo = simpledialog.askstring(
                "Motivo del reingreso",
                "Explique por qué corresponde crear una segunda hoja en este día operativo:",
                parent=self.root,
            )
            motivo = (motivo or "").strip()
            if len(motivo) < 8:
                messagebox.showwarning(
                    "Reingreso",
                    "El motivo debe contener al menos 8 caracteres.",
                    parent=self.root,
                )
                return None
            return {
                "EsReingreso": True,
                "AtencionOrigenId": int(atencion["id"]),
                "MotivoReingreso": motivo,
                "AutorizadoPor": actor,
            }
        if accion not in {"abrir", "reimprimir"}:
            return None

        if accion == "abrir":
            self._abrir_hoja_temporal_atencion(int(atencion["id"]))
            return None

        copias = max(1, int(self.app_settings.get("print_copies_hoja", 1) or 1))
        self.set_status(f"Reimprimiendo atención #{atencion['id']}...", "process")
        if self._imprimir_hoja_temporal_atencion(int(atencion["id"]), copias=copias, mostrar_error=True):
            self.set_status(f"Atención #{atencion['id']} enviada a impresión", "ok")
        else:
            self.set_status("La atención existe, pero falló la reimpresión", "error")
        return None

    def _iniciar_salida_atencion(
        self,
        atencion_id,
        hoja,
        datos_pdf,
        turno_cfg,
        abrir_pdf_final=False,
        flow_started_at=None,
    ):
        attention_id = int(atencion_id)
        with self._output_lock:
            if attention_id in self._output_inflight:
                APP_LOG.warning("PDF_OUTPUT_DUPLICATE_SUPPRESSED attention_id=%s", attention_id)
                return False
            self._output_inflight.add(attention_id)

        payload = {
            "atencion_id": attention_id,
            "hoja": hoja,
            "datos_pdf": MappingProxyType(dict(datos_pdf)),
            "turno_cfg": dict(turno_cfg) if turno_cfg else None,
            "abrir_pdf_final": bool(abrir_pdf_final),
            "flow_started_at": float(flow_started_at or _time.perf_counter()),
        }
        self._output_payloads[attention_id] = payload
        APP_LOG.info(
            "PDF_FLOW_START attention_id=%s elapsed_ms=0.0 thread=%s",
            attention_id,
            threading.current_thread().name,
        )
        try:
            self.boton_generar_pdf.configure(state=tk.DISABLED)
            self.set_status(f"Atención #{attention_id} guardada; generando hoja...", "process")
        except Exception:
            pass
        self._start_worker(
            self._procesar_salida_atencion,
            kwargs=payload,
            name=f"salida-atencion-{attention_id}",
        )
        return True

    def _release_output_inflight(self, atencion_id):
        with self._output_lock:
            self._output_inflight.discard(int(atencion_id))

    def _should_print_immediate_attention(self, trabajo):
        estado = str((trabajo or {}).get("impresion_estado") or "").upper()
        if estado in {"PENDIENTE", "FALLIDO", "PROCESANDO"}:
            return True
        return bool(
            self.app_settings.get("auto_print", False)
            and self.app_settings.get("print_auto_hoja", False)
        )

    def _render_immediate_attention_pdf(
        self,
        atencion_id,
        hoja,
        datos_pdf,
        *,
        should_open,
        should_print,
        flow_started_at,
    ):
        hoja_actual = str(hoja or "GENERAL").upper()
        if hoja_actual not in RUTA_HOJAS:
            hoja_actual = "GENERAL"
        self.db.actualizar_trabajo_salida(atencion_id, "pdf", "PROCESANDO")
        render_started = _time.perf_counter()
        APP_LOG.info(
            "PDF_RENDER_START attention_id=%s elapsed_ms=%.1f thread=%s",
            atencion_id,
            (render_started - flow_started_at) * 1000.0,
            threading.current_thread().name,
        )
        ruta_pdf = crear_pdf_temporal(hoja_actual, datos_pdf, mostrar_error=False)
        render_ms = (_time.perf_counter() - render_started) * 1000.0
        if not ruta_pdf or not os.path.isfile(ruta_pdf) or os.path.getsize(ruta_pdf) <= 0:
            raise RuntimeError("No fue posible generar temporalmente la hoja.")
        self.db.actualizar_trabajo_salida(atencion_id, "pdf", "COMPLETADO")
        APP_LOG.info(
            "PDF_RENDER_DONE attention_id=%s elapsed_ms=%.1f thread=%s",
            atencion_id,
            render_ms,
            threading.current_thread().name,
        )
        return AttentionOutputResult(
            attention_id=int(atencion_id),
            pdf_path=ruta_pdf,
            render_ms=render_ms,
            flow_started_at=flow_started_at,
            should_open=bool(should_open),
            should_print=bool(should_print),
        )

    def _open_immediate_attention_pdf(self, output):
        if not output.should_open:
            return True
        dispatch_started = _time.perf_counter()
        APP_LOG.info(
            "PDF_OPEN_DISPATCH_START attention_id=%s elapsed_ms=%.1f thread=%s",
            output.attention_id,
            (dispatch_started - output.flow_started_at) * 1000.0,
            threading.current_thread().name,
        )
        opened = abrir_pdf(output.pdf_path, mostrar_error=False)
        elapsed_ms = (_time.perf_counter() - dispatch_started) * 1000.0
        APP_LOG.info(
            "PDF_OPEN_DISPATCH_DONE attention_id=%s elapsed_ms=%.1f thread=%s",
            output.attention_id,
            elapsed_ms,
            threading.current_thread().name,
        )
        return opened

    def _apply_immediate_pdf_open_result(self, output, opened):
        self._notify_embedded_visual_event("pdf_complete")
        if opened:
            self.set_status(
                f"Atención #{output.attention_id} guardada · PDF generado",
                "ok",
            )
        else:
            self.set_status(
                f"Atención #{output.attention_id} guardada; no fue posible abrir la hoja.",
                "warning",
            )

    def _procesar_salida_atencion(
        self,
        atencion_id,
        hoja,
        datos_pdf,
        turno_cfg,
        abrir_pdf_final=False,
        flow_started_at=None,
    ):
        """Render the immediate sheet first; secondary work never delays its opening."""
        flow_started_at = float(flow_started_at or _time.perf_counter())
        errores = {}
        trabajo = self.db.obtener_trabajo_salida(atencion_id) or {}
        output = None
        should_print = self._should_print_immediate_attention(trabajo)
        try:
            output = self._render_immediate_attention_pdf(
                atencion_id,
                hoja,
                datos_pdf,
                should_open=abrir_pdf_final,
                should_print=should_print,
                flow_started_at=flow_started_at,
            )
            opened = self._open_immediate_attention_pdf(output)
            self._post_to_ui(
                lambda result=output, was_opened=opened:
                self._apply_immediate_pdf_open_result(result, was_opened)
            )
        except Exception as exc:
            errores["PDF"] = str(exc)
            APP_LOG.exception("PDF_RENDER_ERROR attention_id=%s", atencion_id)
            self.db.actualizar_trabajo_salida(atencion_id, "pdf", "FALLIDO", error=str(exc))

        if output and output.should_print:
            try:
                self.db.actualizar_trabajo_salida(atencion_id, "impresion", "PROCESANDO")
                print_started = _time.perf_counter()
                APP_LOG.info(
                    "PDF_PRINT_START attention_id=%s elapsed_ms=%.1f thread=%s",
                    atencion_id,
                    (print_started - flow_started_at) * 1000.0,
                    threading.current_thread().name,
                )
                if not imprimir_pdf(
                    output.pdf_path,
                    copias=max(1, int(self.app_settings.get("print_copies_hoja", 1) or 1)),
                ):
                    raise RuntimeError("No fue posible enviar la hoja a la impresora.")
                self.db.actualizar_trabajo_salida(
                    atencion_id, "impresion", "ENVIADO_A_IMPRESORA", incrementar_intento=True
                )
                APP_LOG.info(
                    "PDF_PRINT_DONE attention_id=%s elapsed_ms=%.1f thread=%s",
                    atencion_id,
                    (_time.perf_counter() - print_started) * 1000.0,
                    threading.current_thread().name,
                )
            except Exception as exc:
                errores["Impresión"] = str(exc)
                APP_LOG.exception("PRINT_ERROR attention_id=%s", atencion_id)
                self.db.actualizar_trabajo_salida(
                    atencion_id, "impresion", "FALLIDO", error=str(exc), incrementar_intento=True
                )

        if output:
            programar_limpieza_pdf_temporal(output.pdf_path, espera_segundos=90)

        if trabajo.get("excel_estado") != "COMPLETADO":
            try:
                excel_started = _time.perf_counter()
                APP_LOG.info(
                    "PDF_EXCEL_START attention_id=%s elapsed_ms=%.1f thread=%s",
                    atencion_id,
                    (excel_started - flow_started_at) * 1000.0,
                    threading.current_thread().name,
                )
                self.db.actualizar_trabajo_salida(atencion_id, "excel", "PROCESANDO")
                if not turno_cfg:
                    turno_cfg = self.db.obtener_turno_config_atencion(atencion_id)
                if not turno_cfg:
                    raise TurnoNoVigenteError("No se pudo reconstruir el contexto del turno.")
                reconstruir_excel_turno(self.db, turno_cfg)
                self.db.actualizar_trabajo_salida(atencion_id, "excel", "COMPLETADO")
                APP_LOG.info(
                    "PDF_EXCEL_DONE attention_id=%s elapsed_ms=%.1f thread=%s",
                    atencion_id,
                    (_time.perf_counter() - excel_started) * 1000.0,
                    threading.current_thread().name,
                )
            except Exception as exc:
                errores["Excel"] = str(exc)
                APP_LOG.exception("EXCEL_ERROR attention_id=%s", atencion_id)
                self.db.actualizar_trabajo_salida(
                    atencion_id, "excel", "FALLIDO", error=str(exc)
                )

        trabajo = self.db.obtener_trabajo_salida(atencion_id) or {}
        if trabajo.get("excel_estado") == "COMPLETADO" and not errores:
            self.db.limpiar_error_trabajo_salida(atencion_id)
        trabajo = self.db.obtener_trabajo_salida(atencion_id) or {}
        resultado = {
            "atencion_id": int(atencion_id),
            "errores": errores,
            "trabajo": trabajo,
            "output_created": bool(output),
        }
        APP_LOG.info(
            "PDF_OUTPUT_JOB_DONE attention_id=%s elapsed_ms=%.1f thread=%s",
            atencion_id,
            (_time.perf_counter() - flow_started_at) * 1000.0,
            threading.current_thread().name,
        )
        try:
            self._post_to_ui(lambda r=resultado: self._finalizar_salida_atencion(r))
        except Exception:
            self._release_output_inflight(atencion_id)
            APP_LOG.warning("La interfaz se cerró antes de finalizar la salida #%s", atencion_id)

    def _finalizar_salida_atencion(self, resultado):
        atencion_id = int(resultado["atencion_id"])
        trabajo = resultado.get("trabajo") or {}
        errores = resultado.get("errores") or {}
        try:
            self.boton_generar_pdf.configure(state=tk.NORMAL)
        except Exception:
            pass
        if "PDF" in errores:
            self.set_status(
                f"Atención #{atencion_id} fue guardada, pero no fue posible generar la hoja.",
                "warning",
            )
        elif errores:
            etapas = ", ".join(errores)
            self.set_status(
                f"Atención #{atencion_id} guardada; pendiente: {etapas}", "warning"
            )
            self._dialogo_salida_pendiente(atencion_id, None, errores)
        elif trabajo.get("excel_estado") == "COMPLETADO":
            self.db.notify_detail_sheet_generated(atencion_id)
            if trabajo.get("impresion_estado") == "ENVIADO_A_IMPRESORA":
                self.set_status("Hoja enviada a la impresora correctamente.", "ok")
            else:
                self.set_status(
                    f"Atención #{atencion_id} guardada; listado actualizado", "ok"
                )
            self._output_payloads.pop(atencion_id, None)
        else:
            etapas = ", ".join(errores) or "una etapa de salida"
            self.set_status(
                f"Atención #{atencion_id} guardada; pendiente: {etapas}", "warning"
            )
            self._dialogo_salida_pendiente(atencion_id, None, errores)

        self._release_output_inflight(atencion_id)

        self._invalidar_caches_datos()
        self._refrescar_resumen_en_vivo()
        self._retry_excel_export_jobs()

    def _dialogo_salida_pendiente(self, atencion_id, ruta_pdf, errores):
        win = self._crear_toplevel_estable(
            f"Impresion o documento pendiente #{atencion_id}", "640x360", "salida_pendiente_win"
        )
        if win is None:
            return
        cont = tb.Frame(win, padding=20, style="Root.TFrame")
        cont.pack(fill="both", expand=True)
        tb.Label(
            cont,
            text=f"La atención #{atencion_id} está guardada",
            font=("Arial", 15, "bold"),
        ).pack(anchor="w", pady=(0, 10))
        detalle = "\n".join(f"{etapa}: {mensaje}" for etapa, mensaje in errores.items())
        tb.Label(
            cont,
            text=(detalle or "Queda una etapa pendiente.")
            + "\n\nReintentar no crea otra atención ni otro número.",
            justify="left",
            wraplength=590,
        ).pack(anchor="w", fill="x")
        botones = tb.Frame(cont, style="Root.TFrame")
        botones.pack(side="bottom", fill="x")

        def reintentar():
            try:
                win.destroy()
            except Exception:
                pass
            self._reintentar_trabajo_salida(atencion_id)

        tb.Button(
            botones, text="Reintentar pendientes", bootstyle=SUCCESS, command=reintentar
        ).pack(side="left", padx=(0, 8))
        tb.Button(
            botones,
            text="Abrir hoja",
            bootstyle=INFO,
            command=lambda: self._abrir_hoja_temporal_atencion(atencion_id),
        ).pack(side="left", padx=(0, 8))
        tb.Button(botones, text="Continuar", command=win.destroy).pack(side="right")

    def _reintentar_trabajo_salida(self, atencion_id):
        payload = getattr(self, "_output_payloads", {}).get(int(atencion_id))
        if payload:
            self._iniciar_salida_atencion(**payload)
            return
        atencion = self.db.obtener_atencion_por_id(int(atencion_id))
        if not atencion or str(atencion.get("estado") or "").upper() != "ACTIVA":
            messagebox.showwarning("Documento pendiente", "La atención ya no está activa.")
            return
        self._iniciar_salida_atencion(
            int(atencion_id),
            atencion.get("hoja") or "GENERAL",
            self._snapshot_a_datos(atencion),
            self.db.obtener_turno_config_atencion(int(atencion_id)),
            bool(self.app_settings.get("pdf_open_after_generate", True)),
        )

    def _avisar_trabajos_salida_pendientes(self):
        try:
            pendientes = self.db.listar_trabajos_salida_pendientes(limite=100)
        except Exception:
            APP_LOG.exception("No se pudo consultar la cola de salidas")
            return
        if pendientes:
            self.set_status(
                f"Hay {len(pendientes)} atención(es) con impresión o documento pendiente", "warning"
            )

    def abrir_trabajos_salida_pendientes(self):
        win = self._crear_toplevel_estable(
            "Impresiones y documentos pendientes", "980x560", "trabajos_salida_win"
        )
        if win is None:
            return
        cont = tb.Frame(win, padding=14, style="Root.TFrame")
        cont.pack(fill="both", expand=True)
        tb.Label(
            cont,
            text="Atenciones guardadas con tareas pendientes",
            font=("Arial", 15, "bold"),
        ).pack(anchor="w", pady=(0, 8))
        tb.Label(
            cont,
            text="Reintentar continúa desde la etapa fallida y nunca crea otra atención.",
            style="Muted.TLabel",
        ).pack(anchor="w", pady=(0, 12))

        columnas = ("id", "paciente", "excel", "pdf", "impresion", "error")
        tree = ttk.Treeview(cont, columns=columnas, show="headings", style="Modern.Treeview")
        titulos = {
            "id": "Atención",
            "paciente": "Paciente",
            "excel": "Excel",
            "pdf": "PDF",
            "impresion": "Impresión",
            "error": "Último error",
        }
        anchos = {"id": 80, "paciente": 220, "excel": 95, "pdf": 95, "impresion": 105, "error": 300}
        for columna in columnas:
            tree.heading(columna, text=titulos[columna])
            tree.column(columna, width=anchos[columna], anchor="w")
        tree.pack(fill="both", expand=True)

        def cargar():
            for item in tree.get_children():
                tree.delete(item)
            pendientes = self.db.listar_trabajos_salida_pendientes(limite=200)
            for trabajo in pendientes:
                tree.insert(
                    "",
                    "end",
                    iid=str(trabajo["atencion_id"]),
                    values=(
                        trabajo["atencion_id"],
                        trabajo.get("nombre") or "",
                        trabajo.get("excel_estado") or "",
                        trabajo.get("pdf_estado") or "",
                        trabajo.get("impresion_estado") or "",
                        trabajo.get("ultimo_error") or "",
                    ),
                )
            if not pendientes:
                self.set_status("No hay impresiones ni documentos pendientes", "ok")

        def reintentar_seleccion():
            seleccion = tree.selection()
            if not seleccion:
                messagebox.showwarning(
                    "Impresiones y documentos pendientes", "Seleccione una atención para reintentar.", parent=win
                )
                return
            atencion_id = int(seleccion[0])
            win.destroy()
            self._reintentar_trabajo_salida(atencion_id)

        botones = tb.Frame(cont, style="Root.TFrame")
        botones.pack(fill="x", pady=(10, 0))
        tb.Button(
            botones,
            text="Reintentar selección",
            bootstyle=SUCCESS,
            command=reintentar_seleccion,
        ).pack(side="left")
        tb.Button(botones, text="Actualizar", command=cargar).pack(side="left", padx=8)
        tb.Button(botones, text="Cerrar", command=win.destroy).pack(side="right")
        tree.bind("<Double-1>", lambda _event: reintentar_seleccion())
        cargar()

    def generar_pdf(self):
        salida_iniciada = False
        flow_started_at = _time.perf_counter()
        self._notify_embedded_visual_event("before_pdf")
        APP_LOG.info(
            "PDF_FLOW_START attention_id=pending elapsed_ms=0.0 thread=%s",
            threading.current_thread().name,
        )
        try:
            self.boton_generar_pdf.config(state=tk.DISABLED)
            self.set_status("Validando atención...", "process")

            validacion = self._validar_campos_o_alertar()
            if not validacion:
                return
            if self._final_revalidation_ready:
                self._final_revalidation_ready = False
                validacion = self._validar_campos_o_alertar()
                if not validacion:
                    return
            elif self._begin_final_patient_revalidation():
                return
            else:
                validacion = self._validar_campos_o_alertar()
                if not validacion:
                    return
            datos, hoja = validacion

            ars_canon = normalizar_seguro(datos.get('Aseguradora (ARS)', ''), datos.get('NSS', ''))
            es_sin_seguro = (ars_canon == "SIN SEGURO")

            datos['Aseguradora (ARS)'] = ars_canon

            turno_cfg = cargar_turno_config()
            if not turno_cfg:
                self.set_status("Debe abrir el turno operativo actual", "warning")
                messagebox.showwarning(
                    "Turno requerido",
                    "No existe un turno vigente. Abra el turno actual antes de registrar al paciente.",
                    parent=self.root,
                )
                self._dialogo_turno()
                return
            if not self._turno_pertenece_a_sesion(turno_cfg):
                self.set_status("El turno pertenece a otro usuario", "warning")
                self._asegurar_turno_de_sesion()
                return

            duplicado = self._buscar_duplicado_turno_actual(datos)
            if duplicado is not None:
                try:
                    self.root.lift()
                    self.root.focus_force()
                except Exception:
                    pass
                reingreso = self._atender_duplicado_turno(duplicado)
                if not reingreso:
                    return
                datos.update(reingreso)

            if datos.get("EsReingreso"):
                msg = (
                    "Se registrará un reingreso autorizado y se conservará la referencia "
                    f"a la atención #{datos.get('AtencionOrigenId')}.\n\n¿Desea continuar?"
                )
            elif es_sin_seguro:
                msg = (
                    "Este paciente será registrado como SIN SEGURO.\n\n"
                    "Se guardará en el historial sin seguro, se agregará al Excel "
                    "y quedará disponible para facturación."
                )
            else:
                msg = "¿Guardar esta atención y actualizar el listado?"

            if bool(self.app_settings.get("validation_confirm_before_generate", True)):
                if not messagebox.askyesno("Confirmación", msg):
                    return

            datos_db = dict(datos)
            if nombre_tiene_prefijo_rn(datos_db.get("Nombre", "")):
                if bool(self.app_settings.get("rn_strip_db", True)):
                    datos_db["Nombre"] = limpiar_nombre_rn_para_db(datos_db.get("Nombre", ""))

            self.set_status("Guardando atención en la base de datos...", "process")
            save_started = _time.perf_counter()
            APP_LOG.info("ATTENTION_SAVE_START operation=attention-local-save")
            atencion_id = self.db.guardar_atencion(
                datos_db, hoja, turno_cfg=turno_cfg
            )
            APP_LOG.info(
                "ATTENTION_SAVE_LOCAL_DONE operation=attention-local-save elapsed_ms=%.1f",
                (_time.perf_counter() - save_started) * 1000.0,
            )
            APP_LOG.info(
                "PDF_LOCAL_SAVE_DONE attention_id=%s elapsed_ms=%.1f thread=%s",
                atencion_id,
                (_time.perf_counter() - flow_started_at) * 1000.0,
                threading.current_thread().name,
            )
            self._ultimo_atencion_id = atencion_id
            revision_nss_id = self.db.obtener_revision_nss_atencion(atencion_id)
            self.set_status(
                f"Atención #{atencion_id} guardada; actualizando listado...", "process"
            )
            self._iniciar_salida_atencion(
                atencion_id,
                hoja,
                datos_db,
                turno_cfg,
                abrir_pdf_final=bool(self.app_settings.get("pdf_open_after_generate", True)),
                flow_started_at=flow_started_at,
            )
            salida_iniciada = True

            self.limpiar_campos()
            self._restore_all_styles()
            self._refrescar_resumen_en_vivo()
            if revision_nss_id:
                aviso = (
                    f"Atención #{atencion_id} guardada; la hoja continúa normalmente. "
                    "El NSS fue enviado a revisión administrativa."
                )
                self.set_status(aviso, "warning")
                self._mostrar_notificacion(aviso, autohide_ms=12000, tipo="warning")

        except sqlite3.IntegrityError:
            APP_LOG.exception("Se bloqueó una atención duplicada por la restricción del turno")
            duplicado = self._buscar_duplicado_turno_actual(locals().get("datos", {}))
            if duplicado:
                self._atender_duplicado_turno(duplicado)
            else:
                messagebox.showwarning("Atención duplicada", "No se creó otra hoja para este paciente en el turno actual.")
        except (TurnoNoVigenteError, ValueError) as e:
            APP_LOG.warning("Registro rechazado: %s", e)
            self.set_status(str(e), "warning")
            messagebox.showwarning("No se guardó la atención", str(e), parent=self.root)
        except Exception as e:
            APP_LOG.exception("Error al generar la hoja de emergencia")
            self.set_status(f"Error: {str(e)}", "error")
            messagebox.showerror("Error", str(e))
        finally:
            if not salida_iniciada and not self._final_revalidation_in_progress:
                self.boton_generar_pdf.config(state=tk.NORMAL)

    # ─── HISTORIALES ───────────────────────────────────────────────────────
    def abrir_historial(self):
        win = self._crear_toplevel_estable("Historial de Atenciones", "1160x720", "historial_win")
        if win is None:
            return
        pal = self._paleta_visual_actual()

        # FASE 3: Constante para el caché del menú
        MENU_CACHE_SECONDS = 60

        self._bind_esc_cerrar(win)

        cont = tb.Frame(win, padding=14, style="Root.TFrame")
        cont.pack(fill="both", expand=True)

        self._crear_header_ventana(
            cont,
            "Historial de Atenciones",
            "Consulta, edita y anula atenciones cuando sea necesario.",
            "📁"
        )

        frm_bus = tb.Frame(cont, padding=12, style="Card.TFrame")
        frm_bus.pack(fill="x", pady=(0, 10))

        tb.Label(frm_bus, text="Buscar", font=("Arial", 10, "bold"), foreground=pal["text"], background=pal["card"]).pack(side="left", padx=(0, 8))
        self.var_bus = tk.StringVar()
        ent_bus = tb.Entry(frm_bus, textvariable=self.var_bus, width=44)
        ent_bus.pack(side="left", ipady=4)
        win.after(80, lambda: (ent_bus.focus_set(), ent_bus.icursor("end")))

        # El historial abre acotado al turno vigente registrado en la base.
        filtro_rapido_var = tk.StringVar(value="Este turno")
        ars_filtro_var = tk.StringVar(value="(Todas)")
        esp_filtro_var = tk.StringVar(value="(Todas)")
        fecha_filtro = crear_selector_fecha(frm_bus, width=12)

        tb.Label(frm_bus, text="Filtro", font=("Arial", 10, "bold"), foreground=pal["text"], background=pal["card"]).pack(side="left", padx=(10, 4))

        filtro_label_var = tk.StringVar(value="Todos  ▾")
        filtro_btn = tk.Menubutton(
            frm_bus,
            textvariable=filtro_label_var,
            relief="solid",
            bd=1,
            width=24,
            anchor="w",
            bg=pal["entry"],
            fg=pal["text"],
            activebackground=pal["selected_bg"],
            activeforeground=pal["selected_fg"],
            font=("Arial", 10)
        )
        filtro_btn.pack(side="left", padx=4, ipady=3)

        filtro_menu = tk.Menu(
            filtro_btn,
            tearoff=0,
            bg=pal["card"],
            fg=pal["text"],
            activebackground=pal["selected_bg"],
            activeforeground=pal["selected_fg"],
            font=("Arial", 10)
        )
        ars_menu = tk.Menu(
            filtro_menu,
            tearoff=0,
            bg=pal["card"],
            fg=pal["text"],
            activebackground=pal["selected_bg"],
            activeforeground=pal["selected_fg"],
            font=("Arial", 10)
        )
        esp_menu = tk.Menu(
            filtro_menu,
            tearoff=0,
            bg=pal["card"],
            fg=pal["text"],
            activebackground=pal["selected_bg"],
            activeforeground=pal["selected_fg"],
            font=("Arial", 10)
        )
        filtro_btn.configure(menu=filtro_menu)

        def _refresh_history_theme():
            current = self._paleta_visual_actual()
            controls = (filtro_btn,)
            for control in controls:
                try:
                    control.configure(
                        bg=current["entry"],
                        fg=current["text"],
                        activebackground=current["selected_bg"],
                        activeforeground=current["selected_fg"],
                        highlightbackground=current["border"],
                        highlightcolor=current["accent"],
                    )
                except Exception:
                    pass
            for menu in (filtro_menu, ars_menu, esp_menu):
                try:
                    menu.configure(
                        bg=current["card"],
                        fg=current["text"],
                        activebackground=current["selected_bg"],
                        activeforeground=current["selected_fg"],
                    )
                except Exception:
                    pass

        win._admission_theme_refresh = _refresh_history_theme
        _refresh_history_theme()

        try:
            fecha_filtro.pack_forget()
        except Exception:
            pass

        turno_info_var = tk.StringVar(value="Turno seleccionado: cargando…")
        resultados_var = tk.StringVar(value="")
        turno_info = tb.Frame(cont, padding=(10, 6), style="Card.TFrame")
        turno_info.pack(fill="x", pady=(0, 8))
        tb.Label(turno_info, textvariable=turno_info_var, style="Muted.TLabel").pack(side="left")
        tb.Label(turno_info, textvariable=resultados_var, style="Muted.TLabel").pack(side="right")

        def _resolver_turno_historial():
            modo = filtro_rapido_var.get()
            if modo not in {"Este turno", "Turno anterior"}:
                turno_info_var.set("Filtro sin restricción de turno.")
                return None
            turnos = self.db.obtener_turnos_historial()
            turno = turnos.get("actual" if modo == "Este turno" else "anterior")
            if not turno:
                turno_info_var.set(f"{modo}: no hay un turno registrado para mostrar.")
                return None
            inicio = parse_datetime_local(turno.get("fecha_inicio"))
            fin = parse_datetime_local(turno.get("fecha_fin"))
            inicio_txt = format_datetime_local(inicio) if inicio else str(turno.get("fecha_inicio") or "")
            fin_txt = format_datetime_local(fin) if fin else str(turno.get("fecha_fin") or "")
            turno_info_var.set(f"{modo}: {inicio_txt} → {fin_txt}")
            return int(turno["id"])

        def _actualizar_texto_boton_filtro():
            modo = filtro_rapido_var.get()
            if modo == "Por ARS":
                seleccionado = ars_filtro_var.get() or "(Todas)"
                filtro_label_var.set(f"Por ARS: {seleccionado}  ▾")
            elif modo == "Por especialidad":
                seleccionado = esp_filtro_var.get() or "(Todas)"
                filtro_label_var.set(f"Por especialidad: {seleccionado}  ▾")
            elif modo == "Por fecha":
                filtro_label_var.set("Por fecha  ▾")
            else:
                filtro_label_var.set(f"{modo}  ▾")

        def _actualizar_visibilidad_filtros(*_):
            try:
                fecha_filtro.pack_forget()
            except Exception:
                pass
            if filtro_rapido_var.get() == "Por fecha":
                fecha_filtro.pack(side="left", padx=4)
            try:
                _actualizar_texto_boton_filtro()
                _resolver_turno_historial()
            except Exception:
                pass

        menu_state = {"last_build": 0, "busy": False}

        def _programar_busqueda_filtro():
            try:
                win.after(180, buscar)
            except Exception:
                try:
                    buscar()
                except Exception:
                    pass

        def _seleccionar_filtro_simple(modo):
            filtro_rapido_var.set(modo)
            if modo != "Por ARS":
                ars_filtro_var.set("(Todas)")
            if modo != "Por especialidad":
                esp_filtro_var.set("(Todas)")
            _actualizar_visibilidad_filtros()
            _programar_busqueda_filtro()

        def _seleccionar_ars_filtro(valor):
            filtro_rapido_var.set("Por ARS")
            ars_filtro_var.set(valor)
            esp_filtro_var.set("(Todas)")
            _actualizar_visibilidad_filtros()
            _programar_busqueda_filtro()

        def _seleccionar_especialidad_filtro(valor):
            filtro_rapido_var.set("Por especialidad")
            esp_filtro_var.set(valor)
            ars_filtro_var.set("(Todas)")
            _actualizar_visibilidad_filtros()
            _programar_busqueda_filtro()

        def _reconstruir_menu_filtros(_=None, forzar=False):
            """
            FASE 3: Reconstruye el menú usando caché de 60 segundos.
            """
            if menu_state.get("busy"):
                return
            ahora = _time.time()
            if (not forzar) and (ahora - float(menu_state.get("last_build") or 0) < MENU_CACHE_SECONDS):
                return

            menu_state["busy"] = True
            try:
                filtro_menu.delete(0, "end")
                ars_menu.delete(0, "end")
                esp_menu.delete(0, "end")

                filtro_menu.add_command(label="Este turno", command=lambda: _seleccionar_filtro_simple("Este turno"))
                filtro_menu.add_command(label="Turno anterior", command=lambda: _seleccionar_filtro_simple("Turno anterior"))
                filtro_menu.add_separator()
                filtro_menu.add_command(label="Todos", command=lambda: _seleccionar_filtro_simple("Todos"))
                filtro_menu.add_command(label="Hoy", command=lambda: _seleccionar_filtro_simple("Hoy"))
                filtro_menu.add_command(label="Sin seguro", command=lambda: _seleccionar_filtro_simple("Sin seguro"))
                filtro_menu.add_separator()

                ars_menu.add_command(label="(Todas)", command=lambda: _seleccionar_ars_filtro("(Todas)"))
                ars_menu.add_separator()

                ars_items = self._obtener_ars_cache(forzar=False)
                if not ars_items:
                    ars_items = sorted([a for a in DEFAULT_ARS_CATALOGO.keys() if a != "SIN SEGURO"])

                ars_items = [a for a in ars_items if a and str(a).strip()]
                total_ars = len(ars_items)

                for ars_val in ars_items[:80]:
                    ars_menu.add_command(label=ars_val, command=lambda v=ars_val: _seleccionar_ars_filtro(v))

                if total_ars > 80:
                    ars_menu.add_separator()
                    ars_menu.add_command(
                        label=f"Hay {total_ars} ARS. Use Buscar para filtrar más.",
                        state="disabled"
                    )

                filtro_menu.add_cascade(label="Por ARS", menu=ars_menu)

                for esp_val in ["(Todas)", "GENERAL", "PEDIATRIA", "GINECOLOGIA"]:
                    esp_menu.add_command(label=esp_val, command=lambda v=esp_val: _seleccionar_especialidad_filtro(v))
                filtro_menu.add_cascade(label="Por especialidad", menu=esp_menu)

                filtro_menu.add_command(label="Por fecha", command=lambda: _seleccionar_filtro_simple("Por fecha"))
                menu_state["last_build"] = ahora
            finally:
                menu_state["busy"] = False

        try:
            filtro_menu.configure(postcommand=lambda: _reconstruir_menu_filtros(forzar=False))
        except Exception:
            pass

        try:
            win.after(50, lambda: _reconstruir_menu_filtros(forzar=False))
        except Exception:
            pass
        _actualizar_texto_boton_filtro()

        cols = ("id", "fecha", "hora", "nombre", "hoja", "ars", "nss", "cedula", "tipo")
        tree = ttk.Treeview(cont, columns=cols, show="headings", height=10, style="Modern.Treeview")
        tree.pack(fill="both", expand=True, pady=(0, 10))

        cols_def = [
            ("id", "ID", 60, "center"),
            ("fecha", "Fecha", 90, "center"),
            ("hora", "Hora", 90, "center"),
            ("nombre", "Nombre", 250, "w"),
            ("hoja", "Especialidad", 120, "center"),
            ("ars", "Seguro", 160, "center"),
            ("nss", "NSS", 125, "center"),
            ("cedula", "Cédula", 125, "center"),
            ("tipo", "Tipo", 100, "center"),
        ]
        for c, title, w, anchor in cols_def:
            tree.heading(c, text=title)
            tree.column(c, width=w, anchor=anchor)

        # FASE 1: 100 / 150
        page_state = {
            "offset": 0,
            "first_limit": min(100, max(50, int(self.app_settings.get("hist_initial_limit", 100) or 100))),
            "next_limit":  min(150, max(80, int(self.app_settings.get("hist_next_limit",  150) or 150))),
            "loading": False,
            "done": False,
            "load_id": 0,
            "rows": [],
            "fingerprint": (),
        }

        def _insertar_mensaje_tabla(mensaje):
            for i in tree.get_children():
                tree.delete(i)
            tree.insert("", "end", values=("", "", "", mensaje, "", "", "", "", ""))

        def _deduplicar_filas_historial(filas):
            deduplicadas = []
            indices = {}
            for raw in filas or ():
                fila = dict(raw or {})
                global_id = str(fila.get("global_attention_id") or "").replace("-", "").lower()
                identidad = global_id or f"local:{fila.get('id') or ''}"
                if identidad in indices:
                    deduplicadas[indices[identidad]] = fila
                else:
                    indices[identidad] = len(deduplicadas)
                    deduplicadas.append(fila)
            return deduplicadas

        def cargar_pagina(
            reset=False,
            *,
            completion=None,
            cache_only=False,
            show_loading=True,
        ):
            if page_state["loading"]:
                APP_LOG.info(
                    "HISTORY_REFRESH_SKIPPED reason=page_busy source=%s",
                    "cache" if cache_only else "central",
                )
                if callable(completion):
                    completion(changed=False)
                return
            page_state["loading"] = True
            page_state["load_id"] += 1
            current_load = page_state["load_id"]
            selected_attention_id = ""
            scroll_position = 0.0
            try:
                selected_items = tree.selection()
                if selected_items:
                    selected_attention_id = str(
                        tree.item(selected_items[0], "values")[0] or ""
                    )
                scroll_position = float(tree.yview()[0] or 0.0)
            except Exception:
                pass

            if reset:
                page_state["offset"] = 0
                page_state["done"] = False
                if show_loading and not page_state["fingerprint"]:
                    _insertar_mensaje_tabla("Cargando datos…")

            limit = page_state["first_limit"] if reset else page_state["next_limit"]
            fecha_txt = obtener_fecha_selector(fecha_filtro) if filtro_rapido_var.get() == "Por fecha" else None
            modo_actual = filtro_rapido_var.get()
            turno_id_seleccionado = _resolver_turno_historial()

            def _finalizar_carga(filas):
                if current_load != page_state["load_id"]:
                    return
                filas = _deduplicar_filas_historial(filas)
                combined_rows = list(filas) if reset else list(page_state["rows"]) + list(filas)
                combined_rows = _deduplicar_filas_historial(combined_rows)
                new_fingerprint = history_rows_fingerprint(combined_rows)
                changed = new_fingerprint != page_state["fingerprint"]
                if reset and changed:
                    for i in tree.get_children():
                        tree.delete(i)
                rows_to_insert = filas if (not reset or changed) else ()
                for f in rows_to_insert:
                    seguro_canon = normalizar_seguro(f.get("ars", ""), f.get("nss", ""))
                    tree.insert(
                        "",
                        "end",
                        values=(
                            f["id"],
                            f["fecha"],
                            f["hora"],
                            f["nombre"],
                            f["hoja"],
                            seguro_para_mostrar(seguro_canon),
                            f["nss"],
                            f.get("cedula", ""),
                            (f.get("tipo_atencion") or "EMERGENCIA")
                        )
                    )
                if selected_attention_id and changed:
                    for item in tree.get_children():
                        values = tree.item(item, "values")
                        if values and str(values[0] or "") == selected_attention_id:
                            tree.selection_set(item)
                            tree.focus(item)
                            break
                if changed:
                    try:
                        tree.yview_moveto(scroll_position)
                    except Exception:
                        pass
                page_state["rows"] = combined_rows
                page_state["fingerprint"] = new_fingerprint
                page_state["offset"] = len(combined_rows)
                if len(filas) < limit:
                    page_state["done"] = True
                if reset and not filas and changed:
                    tree.insert("", "end", values=("", "", "", "No se encontraron registros.", "", "", "", "", ""))
                cantidad = len(combined_rows)
                resultados_var.set(
                    "Sin resultados" if not filas and reset else f"Mostrando {cantidad} resultado(s)"
                )
                page_state["loading"] = False
                APP_LOG.info(
                    "HISTORY_REFRESH_CHANGED source=%s changed=%s rows=%s",
                    "cache" if cache_only else "central",
                    str(bool(changed)).lower(),
                    cantidad,
                )
                if callable(completion):
                    completion(changed=changed, rows=cantidad)

            def _error_carga(err):
                if current_load != page_state["load_id"]:
                    return
                page_state["loading"] = False
                APP_LOG.error(
                    "HISTORY_LOAD_ERROR type=%s message=%s",
                    type(err).__name__,
                    err,
                    exc_info=(type(err), err, getattr(err, "__traceback__", None)),
                )
                if reset and not page_state["fingerprint"]:
                    _insertar_mensaje_tabla(
                        "No se pudo cargar el historial. Presione Buscar para reintentar."
                    )
                if page_state["fingerprint"]:
                    resultados_var.set("Mostrando caché local · actualización pendiente")
                    self.set_status(
                        "Historial local visible; actualización central pendiente.",
                        "warning",
                    )
                else:
                    resultados_var.set("Error de carga · puede reintentar")
                    self.set_status("Error cargando historial.", "error")
                if callable(completion):
                    completion(changed=False, error=err)

            def _hacer_carga():
                if page_state["done"] and not reset:
                    return []
                requires_local_turn = not bool(
                    getattr(self.db, "uses_central_history", False)
                )
                if (
                    modo_actual in {"Este turno", "Turno anterior"}
                    and turno_id_seleccionado is None
                    and requires_local_turn
                ):
                    return []
                method_name = "listar_atenciones_filtradas"
                method_values = {
                    "filtro_texto": self.var_bus.get().strip(),
                    "modo": modo_actual,
                    "ars": ars_filtro_var.get(),
                    "especialidad": esp_filtro_var.get(),
                    "fecha_txt": fecha_txt,
                    "limite": limit,
                    "offset": page_state["offset"],
                    "turno_id": turno_id_seleccionado,
                }
                # FASE 2: Rama dedicada para Sin seguro
                if modo_actual == "Sin seguro":
                    method_name = "listar_atenciones_sin_seguro"
                    method_values = {
                        "filtro_texto": self.var_bus.get().strip(),
                        "limite": limit,
                        "offset": page_state["offset"],
                    }
                if cache_only:
                    cache_reader = getattr(self.db, "list_history_cache_local", None)
                    if callable(cache_reader):
                        return cache_reader(method_name, **method_values)
                return getattr(self.db, method_name)(**method_values)

            self._ejecutar_en_segundo_plano(
                "Cargando historial…",
                _hacer_carga,
                al_terminar=_finalizar_carga,
                al_error=_error_carga,
            )

        history_refresh_gate = None

        def _start_history_refresh(_reason, done):
            cargar_pagina(
                reset=True,
                completion=done,
                cache_only=True,
                show_loading=False,
            )

        history_refresh_gate = CoalescedRefreshGate(
            schedule=lambda delay, callback: win.after(delay, callback),
            cancel=lambda token: win.after_cancel(token),
            start=_start_history_refresh,
            debounce_ms=200,
            logger=APP_LOG,
            log_prefix="HISTORY_REFRESH",
        )

        def request_history_refresh(reason="manual", *, immediate=False):
            return history_refresh_gate.request(reason, immediate=immediate)

        self.request_history_refresh = request_history_refresh

        def buscar():
            request_history_refresh("manual_search", immediate=True)

        def _al_mover_scroll(event=None):
            try:
                win.after(80, lambda: (not page_state["loading"] and not page_state["done"] and tree.yview()[1] >= 0.98 and cargar_pagina(reset=False)))
            except Exception:
                pass

        try:
            fecha_filtro.bind("<<DateEntrySelected>>", lambda e: buscar())
            fecha_filtro.bind("<Return>", lambda e: buscar())
        except Exception:
            pass
        tree.bind("<MouseWheel>", _al_mover_scroll)
        tree.bind("<Button-5>", _al_mover_scroll)

        tb.Button(frm_bus, text="🔎  Buscar", bootstyle=PRIMARY, command=buscar).pack(side="left", padx=6, ipady=3)
        tb.Button(
            frm_bus,
            text="Mostrar todo",
            bootstyle=SECONDARY,
            command=lambda: [
                self.var_bus.set(""),
                filtro_rapido_var.set("Todos"),
                ars_filtro_var.set("(Todas)"),
                esp_filtro_var.set("(Todas)"),
                _actualizar_visibilidad_filtros(),
                _reconstruir_menu_filtros(),
                buscar()
            ]
        ).pack(side="left", padx=6, ipady=3)
        if self._puede(CAP_VIEW_REPORTS):
            tb.Button(frm_bus, text="📊  Reporte estadístico", bootstyle=INFO, command=self.abrir_ventana_reporte).pack(side="right", padx=(8, 0), ipady=3)

        ent_bus.bind("<Return>", lambda e: buscar())
        _actualizar_visibilidad_filtros()

        frm_btn = tb.Frame(cont, padding=(12, 10), style="Card.TFrame")
        frm_btn.pack(side="bottom", fill="x", pady=(8, 0))

        tb.Button(frm_btn, text="📄  Abrir hoja", bootstyle=SUCCESS, command=lambda: self.ver_pdf_seleccionado(tree)).pack(side="left", padx=4, ipady=4)
        if self._puede(CAP_EDIT_RECORDS):
            tb.Button(frm_btn, text="🖉  Editar atención", bootstyle=SECONDARY, command=lambda: self._abrir_editor_atencion_desde_tree(tree, buscar)).pack(side="left", padx=4, ipady=4)
        if self._puede(CAP_VOID_RECORDS):
            tb.Button(frm_btn, text="Anular seleccionado", bootstyle=DANGER, command=lambda: self.eliminar_atencion_seleccionada(tree, reordenar_ids=True, refrescar_callback=buscar)).pack(side="left", padx=4, ipady=4)
        tb.Button(frm_btn, text="🛡  Historial sin seguros", bootstyle=WARNING, command=self.abrir_historial_sin_seguros).pack(side="left", padx=4, ipady=4)

        def _cache_loaded(**_result):
            return None

        try:
            win.after(
                10,
                lambda: cargar_pagina(
                    reset=True,
                    completion=_cache_loaded,
                    cache_only=True,
                ),
            )
        except Exception:
            cargar_pagina(
                reset=True,
                completion=_cache_loaded,
                cache_only=True,
            )

        # The shared sync coordinator emits this only after applying remote
        # attention events.  Keep the current filter and selection while the
        # open history receives updates from another station.
        event_bus = getattr(getattr(self, "context", None), "event_bus", None)
        refresh_signal = getattr(event_bus, "history_refresh_requested", None)
        refresh_callback = None
        if refresh_signal is not None and hasattr(refresh_signal, "connect"):
            def refresh_callback():
                try:
                    if win.winfo_exists():
                        request_history_refresh("attention_event")
                except Exception:
                    pass

            refresh_signal.connect(refresh_callback)
            try:
                def disconnect_history_refresh(_event):
                    history_refresh_gate.close()
                    try:
                        refresh_signal.disconnect(refresh_callback)
                    except (RuntimeError, TypeError):
                        pass
                    if getattr(self, "request_history_refresh", None) is request_history_refresh:
                        try:
                            delattr(self, "request_history_refresh")
                        except AttributeError:
                            pass

                win.bind(
                    "<Destroy>",
                    disconnect_history_refresh,
                    add="+",
                )
            except Exception:
                pass
        coordinator = getattr(self, "_hybrid_coordinator", None)
        if coordinator is not None:
            try:
                coordinator._schedule()
            except Exception:
                pass

        menu_historial = tk.Menu(win, tearoff=0)

        def _seleccionar_fila_click_derecho(event):
            item = tree.identify_row(event.y)
            if item:
                tree.selection_set(item)
                tree.focus(item)
            return item

        def _identidad_seleccionada_para_config():
            sel = tree.selection()
            if not sel:
                self._mostrar_dialogo_modal_unico("Historial", "Seleccione un registro para editar.")
                return ""
            vals = tree.item(sel[0], "values")
            nss = (vals[6] if len(vals) > 6 else "") or ""
            ced = (vals[7] if len(vals) > 7 else "") or ""
            if is_valid_nss_key(str(nss)):
                return str(nss).strip()
            if is_valid_cedula_key(str(ced)):
                return str(ced).strip()
            self._mostrar_dialogo_modal_unico(
                "Editar paciente",
                "Este registro no tiene NSS ni cédula válida para buscarlo en configuración."
            )
            return ""

        def _editar_paciente_desde_historial():
            identidad = _identidad_seleccionada_para_config()
            if identidad:
                self._abrir_edicion_paciente(prefill_identidad=identidad)

        if self._puede(CAP_EDIT_RECORDS):
            menu_historial.add_command(label="🖉 Editar atención", command=lambda: self._abrir_editor_atencion_desde_tree(tree, buscar))
            menu_historial.add_command(label="⚙ Editar datos del paciente", command=_editar_paciente_desde_historial)
            menu_historial.add_separator()
        menu_historial.add_command(label="📄 Abrir hoja", command=lambda: self.ver_pdf_seleccionado(tree))
        if self._puede(CAP_VOID_RECORDS):
            menu_historial.add_command(label="Anular atención", command=lambda: self.eliminar_atencion_seleccionada(tree, reordenar_ids=True, refrescar_callback=buscar))

        def _mostrar_menu_historial(event):
            if _seleccionar_fila_click_derecho(event):
                menu_historial.tk_popup(event.x_root, event.y_root)

        tree.bind("<Button-3>", _mostrar_menu_historial)
        tree.bind("<Double-1>", lambda e: self.ver_pdf_seleccionado(tree))

    def cargar_tabla(self, tree, filtro):
        for i in tree.get_children():
            tree.delete(i)
        filas = self.db.listar_atenciones(filtro_texto=filtro, limite=200, offset=0)
        for f in filas:
            seguro_canon = normalizar_seguro(f.get("ars", ""), f.get("nss", ""))
            tree.insert(
                "",
                "end",
                values=(
                    f["id"],
                    f["fecha"],
                    f["hora"],
                    f["nombre"],
                    f["hoja"],
                    seguro_para_mostrar(seguro_canon),
                    f["nss"],
                    f.get("cedula", "")
                )
            )

    def cargar_tabla_filtrada(self, tree, filtro=None, modo="Todos", ars=None, especialidad=None, fecha_txt=None):
        for i in tree.get_children(): tree.delete(i)
        filas = self.db.listar_atenciones_filtradas(filtro_texto=filtro, modo=modo or "Todos", ars=ars, especialidad=especialidad, fecha_txt=fecha_txt, limite=200, offset=0)
        for f in filas:
            seguro_canon = normalizar_seguro(f.get("ars", ""), f.get("nss", ""))
            tree.insert("", "end", values=(f["id"], f["fecha"], f["hora"], f["nombre"], f["hoja"], seguro_para_mostrar(seguro_canon), f["nss"], f.get("cedula", "")))

    def _snapshot_a_datos(self, atencion):
        return {
            "Fecha": atencion["fecha"],
            "Hora": atencion["hora"],
            "Nombre": atencion["nombre"],
            "Sexo": atencion.get("sexo", "") or "Femenino",
            "Edad_num": int(atencion.get("edad_num") or 0),
            "Unidad": atencion.get("unidad", "Años"),
            "Cédula": atencion.get("cedula", ""),
            "Teléfono": atencion.get("telefono", "") or "",
            "Dirección": atencion.get("direccion", ""),
            "Nacionalidad": atencion.get("nacionalidad", ""),
            "Aseguradora (ARS)": atencion.get("ars", ""),
            "NSS": atencion.get("nss", ""),
            "TipoAtencion": atencion.get("tipo_atencion", "EMERGENCIA")
        }

    def _crear_hoja_temporal_atencion(self, atencion_id, mostrar_error=True):
        """Reconstruye la hoja desde los datos actuales sin usar PDFs archivados."""
        atencion = self.db.obtener_atencion_por_id(int(atencion_id))
        if not atencion:
            if mostrar_error:
                messagebox.showerror("Historial", "No se encontró la atención seleccionada.")
            return None
        hoja = str(atencion.get("hoja") or "GENERAL").upper()
        if hoja not in RUTA_HOJAS:
            hoja = "GENERAL"
        return crear_pdf_temporal(hoja, self._snapshot_a_datos(atencion), mostrar_error=mostrar_error)

    def _abrir_hoja_temporal_atencion(self, atencion_id):
        ruta_temporal = self._crear_hoja_temporal_atencion(atencion_id, mostrar_error=True)
        if not ruta_temporal:
            return False
        try:
            abrir_pdf(ruta_temporal)
            return True
        finally:
            programar_limpieza_pdf_temporal(ruta_temporal)

    def _imprimir_hoja_temporal_atencion(self, atencion_id, copias=1, mostrar_error=False):
        ruta_temporal = self._crear_hoja_temporal_atencion(atencion_id, mostrar_error=mostrar_error)
        if not ruta_temporal:
            return False
        try:
            return imprimir_pdf(ruta_temporal, copias=copias, mostrar_error=mostrar_error)
        finally:
            programar_limpieza_pdf_temporal(ruta_temporal, espera_segundos=90)

    def ver_pdf_seleccionado(self, tree):
        sel = tree.selection()
        if not sel:
            self._mostrar_dialogo_modal_unico("Historial", "Seleccione un registro.")
            return

        vals = tree.item(sel[0], "values")
        atencion_id = int(vals[0])
        atencion = self.db.obtener_atencion_por_id(atencion_id)
        if not atencion:
            messagebox.showerror("Error", "No se encontró la atención seleccionada.")
            return

        self._abrir_hoja_temporal_atencion(atencion_id)

    def eliminar_atencion_seleccionada(self, tree, reordenar_ids=False, refrescar_callback=None):
        if not self._exigir_permiso(CAP_VOID_RECORDS, "anular una atención"):
            return
        sel = tree.selection()
        if not sel:
            self._mostrar_dialogo_modal_unico("Historial", "Seleccione un registro para anular.")
            return

        actor_autorizado = self._actor_actual()

        vals = tree.item(sel[0], "values")
        atencion_id = int(vals[0])

        atencion = self.db.obtener_atencion_por_id(atencion_id)
        if not atencion:
            messagebox.showwarning("Aviso", "No se encontró el registro.")
            return

        afecta_excel = self._registro_esta_en_turno_actual(atencion)

        mensaje_eliminar = (
            "¿Anular esta atención?\n\n"
            f"Paciente: {(atencion.get('nombre') or '').upper()}\n"
            f"Fecha: {atencion.get('fecha', '')} {atencion.get('hora', '')}\n"
            f"Especialidad: {atencion.get('hoja', '')}\n\n"
            "También se actualizará el Excel si pertenece al turno actual."
        )
        if not messagebox.askyesno("Confirmación", mensaje_eliminar):
            return

        motivo = simpledialog.askstring(
            "Motivo de anulación",
            "Indique brevemente por qué se anula esta atención:",
            parent=self.root,
        )
        if motivo is None:
            return
        motivo = motivo.strip()
        if len(motivo) < 5:
            messagebox.showwarning(
                "Motivo requerido",
                "Escriba un motivo de al menos 5 caracteres para conservar la auditoría.",
                parent=tree.winfo_toplevel(),
            )
            return
        turno_cfg = cargar_turno_config() or {}
        usuario = actor_autorizado or turno_cfg.get("representante", "")

        try:
            anulada = self.db.borrar_atencion(
                atencion_id, motivo=motivo, usuario=usuario
            )
        except ValueError as exc:
            messagebox.showwarning("No se puede anular", str(exc), parent=self.root)
            return

        if anulada:

            if afecta_excel and (atencion.get("tipo_atencion") or "EMERGENCIA").strip().upper() not in ("URGENCIA", "CONSULTA"):
                self._reconstruir_excel_si_necesario("anular atención", antes=atencion, despues={})

            if refrescar_callback:
                try:
                    refrescar_callback()
                except Exception:
                    self.cargar_tabla_filtrada(tree)
            else:
                self.cargar_tabla_filtrada(tree)

            self._actualizar_resumen_turno_panel()
            self._mostrar_notificacion(f"Atención #{atencion_id} anulada.")
        else:
            messagebox.showwarning("Aviso", "No se pudo anular. Intente nuevamente.")

    def abrir_historial_sin_seguros(self):
        win = self._crear_toplevel_estable("Historial sin seguros", "1120x700", "historial_sin_seguro_win")
        if win is None:
            return
        pal = self._paleta_visual_actual()

        self._bind_esc_cerrar(win)

        cont = tb.Frame(win, padding=14, style="Root.TFrame")
        cont.pack(fill="both", expand=True)

        self._crear_header_ventana(
            cont,
            "Historial sin seguros",
            "Pacientes registrados sin cobertura o sin NSS válido. Busca por nombre, consulta, abre PDF o anula.",
            "🛡"
        )

        frm_bus = tb.Frame(cont, padding=12, style="Card.TFrame")
        frm_bus.pack(fill="x", pady=(0, 10))

        tb.Label(frm_bus, text="Buscar", font=("Arial", 10, "bold"), foreground=pal["text"], background=pal["card"]).pack(side="left", padx=(0, 8))
        var_bus = tk.StringVar()
        ent_bus = tb.Entry(frm_bus, textvariable=var_bus, width=44)
        ent_bus.pack(side="left", ipady=4)
        win.after(80, lambda: (ent_bus.focus_set(), ent_bus.icursor("end")))

        btn_buscar_nombre = tb.Button(frm_bus, text="🔎  Buscar nombre", bootstyle=PRIMARY, command=lambda: None)
        btn_buscar_nombre.pack(side="left", padx=8, ipady=3)

        btn_mostrar_todo = tb.Button(frm_bus, text="Mostrar todo", bootstyle=SECONDARY, command=lambda: None)
        btn_mostrar_todo.pack(side="left", padx=4, ipady=3)

        cols = ("id", "fecha", "hora", "nombre", "hoja", "ars", "nss", "cedula", "tipo")
        tree = ttk.Treeview(cont, columns=cols, show="headings", height=10, style="Modern.Treeview")
        tree.pack(fill="both", expand=True, pady=(0, 10))

        cols_def = [
            ("id", "ID", 60, "center"),
            ("fecha", "Fecha", 90, "center"),
            ("hora", "Hora", 90, "center"),
            ("nombre", "Nombre", 260, "w"),
            ("hoja", "Especialidad", 125, "center"),
            ("ars", "Seguro", 130, "center"),
            ("nss", "NSS", 125, "center"),
            ("cedula", "Cédula", 125, "center"),
            ("tipo", "Tipo", 100, "center"),
        ]
        for c, title, w, anchor in cols_def:
            tree.heading(c, text=title)
            tree.column(c, width=w, anchor=anchor)

        # FASE 1: 100 / 150
        page_state = {
            "offset": 0,
            "first_limit": min(100, max(50, int(self.app_settings.get("hist_initial_limit", 100) or 100))),
            "next_limit":  min(150, max(80, int(self.app_settings.get("hist_next_limit",  150) or 150))),
            "loading": False,
            "done": False,
            "load_id": 0,
        }

        def _insertar_mensaje_tabla(mensaje):
            for i in tree.get_children():
                tree.delete(i)
            tree.insert("", "end", values=("", "", "", mensaje, "", "", "", "", ""))

        def cargar(reset=True):
            if page_state["loading"]:
                return
            page_state["loading"] = True
            page_state["load_id"] += 1
            current_load = page_state["load_id"]

            if reset:
                page_state["offset"] = 0
                page_state["done"] = False
                _insertar_mensaje_tabla("Cargando datos…")

            limit = page_state["first_limit"] if reset else page_state["next_limit"]

            def _finalizar_carga(filas):
                if current_load != page_state["load_id"]:
                    return
                if reset:
                    for i in tree.get_children():
                        tree.delete(i)
                for f in filas:
                    tree.insert(
                        "",
                        "end",
                        values=(
                            f["id"],
                            f["fecha"],
                            f["hora"],
                            f["nombre"],
                            f["hoja"],
                            "SIN SEGURO",
                            f["nss"],
                            f.get("cedula", ""),
                            (f.get("tipo_atencion") or "EMERGENCIA")
                        )
                    )
                page_state["offset"] += len(filas)
                if len(filas) < limit:
                    page_state["done"] = True
                if reset and not filas:
                    tree.insert("", "end", values=("", "", "", "No se encontraron registros sin seguro.", "", "", "", "", ""))
                page_state["loading"] = False

            def _error_carga(err):
                if current_load != page_state["load_id"]:
                    return
                page_state["loading"] = False
                self.set_status("Error cargando historial sin seguro.", "error")

            def _hacer_carga():
                if page_state["done"] and not reset:
                    return []
                return self.db.listar_atenciones_sin_seguro(
                    var_bus.get().strip() or None,
                    limite=limit,
                    offset=page_state["offset"],
                )

            self._ejecutar_en_segundo_plano(
                "Cargando historial sin seguro…",
                _hacer_carga,
                al_terminar=_finalizar_carga,
                al_error=_error_carga,
            )

        def _al_mover_scroll_ss(event=None):
            try:
                win.after(80, lambda: (not page_state["loading"] and not page_state["done"] and tree.yview()[1] >= 0.98 and cargar(reset=False)))
            except Exception:
                pass

        tree.bind("<MouseWheel>", _al_mover_scroll_ss)
        tree.bind("<Button-5>", _al_mover_scroll_ss)

        def buscar():
            cargar(reset=True)

        def mostrar_todo():
            var_bus.set("")
            cargar(reset=True)

        try:
            btn_buscar_nombre.configure(command=buscar)
            btn_mostrar_todo.configure(command=mostrar_todo)
            ent_bus.bind("<Return>", lambda e: buscar())
        except Exception:
            pass

        frm_btn = tb.Frame(cont, padding=(12, 10), style="Card.TFrame")
        frm_btn.pack(side="bottom", fill="x", pady=(8, 0))
        tb.Button(frm_btn, text="📄  Abrir hoja", bootstyle=SUCCESS, command=lambda: self.ver_pdf_seleccionado(tree)).pack(side="left", padx=4, ipady=4)
        if self._puede(CAP_EDIT_RECORDS):
            tb.Button(frm_btn, text="🖉  Editar atención", bootstyle=SECONDARY, command=lambda: self._abrir_editor_atencion_desde_tree(tree, cargar)).pack(side="left", padx=4, ipady=4)
            tb.Button(frm_btn, text="⚙ Editar paciente", bootstyle=INFO, command=lambda: self._abrir_edicion_paciente(prefill_identidad=(tree.item(tree.selection()[0], "values")[6] or tree.item(tree.selection()[0], "values")[7]) if tree.selection() else "")).pack(side="left", padx=4, ipady=4)
        if self._puede(CAP_VOID_RECORDS):
            tb.Button(frm_btn, text="Anular seleccionado", bootstyle=DANGER, command=lambda: self.eliminar_atencion_seleccionada(tree, reordenar_ids=False, refrescar_callback=cargar)).pack(side="left", padx=4, ipady=4)

        menu_sin_seguro = tk.Menu(win, tearoff=0)

        def _seleccionar_fila_ss(event):
            item = tree.identify_row(event.y)
            if item:
                tree.selection_set(item)
                tree.focus(item)
            return item

        def _editar_paciente_ss():
            sel = tree.selection()
            if not sel:
                return
            vals = tree.item(sel[0], "values")
            identidad = ""
            if len(vals) > 6 and is_valid_nss_key(str(vals[6])):
                identidad = str(vals[6]).strip()
            elif len(vals) > 7 and is_valid_cedula_key(str(vals[7])):
                identidad = str(vals[7]).strip()
            if identidad:
                self._abrir_edicion_paciente(prefill_identidad=identidad)
            else:
                self._mostrar_dialogo_modal_unico("Editar paciente", "Este registro no tiene NSS ni cédula válida para buscarlo en configuración.")

        if self._puede(CAP_EDIT_RECORDS):
            menu_sin_seguro.add_command(label="🖉 Editar atención", command=lambda: self._abrir_editor_atencion_desde_tree(tree, cargar))
            menu_sin_seguro.add_command(label="⚙ Editar datos del paciente", command=_editar_paciente_ss)
            menu_sin_seguro.add_separator()
        menu_sin_seguro.add_command(label="📄 Abrir hoja", command=lambda: self.ver_pdf_seleccionado(tree))
        if self._puede(CAP_VOID_RECORDS):
            menu_sin_seguro.add_command(label="Anular atención", command=lambda: self.eliminar_atencion_seleccionada(tree, reordenar_ids=False, refrescar_callback=cargar))

        def _mostrar_menu_ss(event):
            if _seleccionar_fila_ss(event):
                menu_sin_seguro.tk_popup(event.x_root, event.y_root)

        tree.bind("<Button-3>", _mostrar_menu_ss)
        tree.bind("<Double-1>", lambda e: self.ver_pdf_seleccionado(tree))

        try:
            win.after(80, lambda: cargar(reset=True))
        except Exception:
            cargar(reset=True)

    @staticmethod
    def _bind_esc_cerrar(win: Toplevel):
        win.bind("<Escape>", lambda e: win.destroy())
    # ────────────────────────────────────────────────────────────────────────

    def _abrir_editor_atencion_desde_tree(self, tree, on_saved=None):
        if not self._exigir_permiso(CAP_EDIT_RECORDS, "modificar una atención registrada"):
            return
        sel = tree.selection()
        if not sel:
            self._mostrar_dialogo_modal_unico("Historial", "Seleccione una atención para editar.")
            return
        vals = tree.item(sel[0], "values")
        self._abrir_editor_atencion(int(vals[0]), on_saved=on_saved)

    def _abrir_editor_atencion(self, atencion_id: int, on_saved=None):
        at = self.db.obtener_atencion_por_id(atencion_id)
        if not at:
            messagebox.showerror("Error", "No se encontró la atención seleccionada.")
            return

        win = Toplevel(self.root)
        win.title(f"Editar atención #{atencion_id}")
        win.geometry("820x760")
        win.minsize(800, 720)
        win.transient(self.root)
        self._bind_esc_cerrar(win)
        try:
            win.lift()
            win.focus_set()
            win.after(20, win.focus_set)
        except Exception:
            pass

        cont = tb.Frame(win, padding=14, style="Root.TFrame")
        cont.pack(fill="both", expand=True)

        self._crear_header_ventana(
            cont,
            "Editar atención específica",
            "Corrige este registro del historial, incluyendo edad, unidad, seguro y tipo de atención.",
            "🖉"
        )

        footer = tb.Frame(cont, padding=(8, 10), style="Card.TFrame")
        footer.pack(side="bottom", fill="x", pady=(8, 0))

        form = tb.Frame(cont, padding=12, style="Card.TFrame")
        form.pack(fill="both", expand=True)
        form.columnconfigure(1, weight=1)
        form.columnconfigure(3, weight=1)

        campos = {}
        campos_def = [
            ("Nombre", at.get("nombre", ""), "entry"),
            ("Sexo", at.get("sexo", "Femenino") or "Femenino", "sexo"),
            ("Fecha", at.get("fecha", ""), "entry"),
            ("Hora", at.get("hora", ""), "entry"),
            ("Hoja", at.get("hoja", ""), "hoja"),
            ("Edad", at.get("edad_num", ""), "entry"),
            ("Unidad", at.get("unidad", "Años"), "unidad"),
            ("TipoAtencion", at.get("tipo_atencion", "EMERGENCIA"), "tipo"),
            ("Aseguradora (ARS)", seguro_para_mostrar(normalizar_seguro(at.get("ars", ""), at.get("nss", ""))), "entry"),
            ("NSS", at.get("nss", ""), "entry"),
            ("Cédula", at.get("cedula", ""), "entry"),
            ("Teléfono", at.get("telefono", ""), "entry"),
            ("Dirección", at.get("direccion", ""), "entry"),
            ("Nacionalidad", at.get("nacionalidad", ""), "entry"),
        ]

        for idx, (label, value, kind) in enumerate(campos_def):
            row = idx // 2
            col = (idx % 2) * 2
            visible_label = "Tipo de atención" if label == "TipoAtencion" else label

            tb.Label(form, text=visible_label, background="#0E1B2B", font=("Arial", 10, "bold")).grid(
                row=row * 2, column=col, sticky="w", padx=6, pady=(4, 2)
            )

            if kind == "hoja":
                var = tk.StringVar(value=(value or "GENERAL").upper())
                ent = tb.Combobox(form, textvariable=var, state="readonly", values=["GENERAL", "PEDIATRIA", "GINECOLOGIA"])
                ent.after_idle(lambda ent=ent, v=(value or "GENERAL"): ent.current(["GENERAL", "PEDIATRIA", "GINECOLOGIA"].index(v.upper())))
            elif kind == "unidad":
                val_unidad = value if value in ["Días", "Meses", "Años"] else "Años"
                var = tk.StringVar(value=val_unidad)
                ent = tb.Combobox(form, textvariable=var, state="readonly", values=["Días", "Meses", "Años"])
                ent.after_idle(lambda ent=ent, v=val_unidad: ent.current(["Días", "Meses", "Años"].index(v)))
            elif kind == "tipo":
                tipos_atencion = ["EMERGENCIA", "URGENCIA", "CONSULTA"]
                valor_tipo = (value or "EMERGENCIA").strip().upper()
                if valor_tipo not in tipos_atencion:
                    valor_tipo = "EMERGENCIA"
                var = tk.StringVar(value=valor_tipo)
                ent = tb.Combobox(form, textvariable=var, state="readonly", values=tipos_atencion)
                ent.after_idle(lambda ent=ent, v=valor_tipo, tipos=tipos_atencion: ent.current(tipos.index(v)))
            elif kind == "sexo":
                val_sexo = value if value in ["Masculino", "Femenino"] else "Femenino"
                var = tk.StringVar(value=val_sexo)
                ent = tb.Combobox(
                    form,
                    textvariable=var,
                    state="readonly",
                    values=["Femenino", "Masculino"],
                )
                ent.after_idle(
                    lambda ent=ent, v=val_sexo: ent.current(
                        ["Femenino", "Masculino"].index(v)
                    )
                )
            else:
                ent = tb.Entry(form)
                ent.insert(0, value or "")

            ent.grid(row=row * 2 + 1, column=col, sticky="ew", padx=6, pady=(0, 6), ipady=3)
            campos[label] = ent

        try:
            campos["Cédula"].bind("<KeyRelease>", lambda e: self.limitar_caracteres(campos["Cédula"], 11))
            campos["Teléfono"].bind("<KeyRelease>", lambda e: self.limitar_caracteres(campos["Teléfono"], 10))
        except Exception:
            pass

        def _get(label):
            return campos[label].get().strip()

        def guardar():
            if not _get("Nombre"):
                messagebox.showerror("Validación", "El nombre no puede quedar vacío.")
                campos["Nombre"].focus_set()
                return

            edad_txt = _get("Edad")
            if not edad_txt.isdigit():
                messagebox.showerror("Validación", "La edad debe ser numérica.")
                campos["Edad"].focus_set()
                return

            edad_num = int(edad_txt)
            if edad_num < 0 or edad_num > 130:
                messagebox.showerror("Validación", "La edad debe estar entre 0 y 130.")
                campos["Edad"].focus_set()
                return

            hoja = _get("Hoja").upper().strip()
            # FASE 11: La hoja no puede quedar vacía
            if not hoja or hoja not in ["GENERAL", "PEDIATRIA", "GINECOLOGIA"]:
                messagebox.showerror("Validación", "Debe seleccionar una especialidad válida (GENERAL, PEDIATRIA o GINECOLOGIA).")
                campos["Hoja"].focus_set()
                return

            ced = _get("Cédula").replace("-", "")
            tel = _get("Teléfono").replace("-", "")

            if ced and not is_valid_cedula_key(ced):
                messagebox.showerror("Validación", "La cédula debe tener 11 dígitos o dejarse vacía.")
                campos["Cédula"].focus_set()
                return

            if tel and not (tel.isdigit() and len(tel) == 10):
                messagebox.showerror("Validación", "El teléfono debe tener 10 dígitos o dejarse vacío.")
                campos["Teléfono"].focus_set()
                return

            nss = _get("NSS").upper()
            if nss and not is_valid_nss_key(nss) and nss not in ["SIN SEGURO", "NO", "N/S", "NS", "N\\S"]:
                messagebox.showerror("Validación", "El NSS debe ser numérico o SIN SEGURO.")
                campos["NSS"].focus_set()
                return

            nuevos = {
                "Nombre": _get("Nombre"),
                "Sexo": _get("Sexo") or "Femenino",
                "Fecha": _get("Fecha"),
                "Hora": _get("Hora"),
                "Hoja": hoja,
                "Edad_num": edad_num,
                "Unidad": _get("Unidad") or "Años",
                "TipoAtencion": _get("TipoAtencion") or "EMERGENCIA",
                "Aseguradora (ARS)": _get("Aseguradora (ARS)"),
                "NSS": nss,
                "Cédula": ced,
                "Teléfono": tel,
                "Dirección": _get("Dirección"),
                "Nacionalidad": _get("Nacionalidad"),
            }

            motivo_rectificacion = simpledialog.askstring(
                "Motivo de rectificación",
                "Indique el motivo de esta corrección (mínimo 5 caracteres):",
                parent=win,
            )
            if motivo_rectificacion is None:
                return
            motivo_rectificacion = motivo_rectificacion.strip()
            if len(motivo_rectificacion) < 5:
                messagebox.showerror(
                    "Motivo requerido",
                    "Debe indicar un motivo de al menos 5 caracteres.",
                    parent=win,
                )
                return

            try:
                snapshot_antes = dict(at)
                self.db.actualizar_atencion_especifica(
                    atencion_id,
                    nuevos,
                    usuario=self.session_context.audit_actor,
                    motivo=motivo_rectificacion,
                )
                revision_nss_id = self.db.obtener_revision_nss_atencion(atencion_id)
                self._invalidar_cache_ars()

                def _undo_edit_atencion():
                    datos_anteriores = {
                        "Nombre": snapshot_antes.get("nombre", ""),
                        "Sexo": snapshot_antes.get("sexo", "Femenino") or "Femenino",
                        "Fecha": snapshot_antes.get("fecha", ""),
                        "Hora": snapshot_antes.get("hora", ""),
                        "Hoja": snapshot_antes.get("hoja", ""),
                        "Edad_num": int(snapshot_antes.get("edad_num") or 0),
                        "Unidad": snapshot_antes.get("unidad", "Años"),
                        "TipoAtencion": snapshot_antes.get("tipo_atencion", "EMERGENCIA"),
                        "Aseguradora (ARS)": snapshot_antes.get("ars", ""),
                        "NSS": snapshot_antes.get("nss", ""),
                        "Cédula": snapshot_antes.get("cedula", ""),
                        "Teléfono": snapshot_antes.get("telefono", ""),
                        "Dirección": snapshot_antes.get("direccion", ""),
                        "Nacionalidad": snapshot_antes.get("nacionalidad", ""),
                    }
                    self.db.actualizar_atencion_especifica(
                        atencion_id,
                        datos_anteriores,
                        usuario=self.session_context.audit_actor,
                        motivo="Reversión de edición desde historial",
                    )
                    turno_cfg_undo = cargar_turno_config()
                    if turno_cfg_undo:
                        reconstruir_excel_turno(self.db, turno_cfg_undo)
                    self._actualizar_resumen_turno_panel()
                    if on_saved:
                        try:
                            on_saved()
                        except Exception:
                            pass
                    messagebox.showinfo("Deshacer", f"Edición de la atención #{atencion_id} revertida.")

                self._push_undo_action(f"edición de atención #{atencion_id}", _undo_edit_atencion)

                self._reconstruir_excel_si_necesario("editar atención", antes=snapshot_antes, despues=nuevos)
                self._invalidar_caches_datos()
                self._refrescar_resumen_en_vivo()
                if on_saved:
                    on_saved()

                self._mostrar_notificacion(f"Atención #{atencion_id} actualizada. Ctrl+Z para deshacer.", on_undo=_undo_edit_atencion, autohide_ms=5000)
                if revision_nss_id:
                    self._mostrar_notificacion(
                        f"Atención #{atencion_id} actualizada. El conflicto NSS "
                        "fue enviado a revisión administrativa sin detener el flujo.",
                        autohide_ms=12000,
                        tipo="warning",
                    )
                messagebox.showinfo(
                    "Guardado",
                    "Atención actualizada correctamente. No se generó ningún PDF individual.",
                )
                try:
                    win.destroy()
                except Exception:
                    pass
            except PermissionError:
                messagebox.showwarning("Excel abierto", "Cierre el Excel para reconstruir el listado del turno.")
            except Exception as e:
                messagebox.showerror("Error", f"No se pudo guardar la atención:\n{str(e)}")

        tb.Button(footer, text="💾  Guardar cambios", bootstyle=SUCCESS, command=guardar, width=22).pack(side="left", padx=5, ipady=5)
        tb.Button(footer, text="Cerrar", bootstyle=SECONDARY, command=win.destroy, width=12).pack(side="right", padx=5, ipady=5)

    def _abrir_edicion_paciente(self, prefill_identidad=None):
        """
        Ventana independiente y ligera para editar pacientes.
        """
        if not self._exigir_permiso(CAP_EDIT_RECORDS, "modificar los datos de un paciente"):
            return
        win = self._crear_toplevel_estable("Editar paciente", "1160x760", "edicion_paciente_win")
        if win is None:
            return

        try:
            win.minsize(1100, 720)
        except Exception:
            pass

        cont = tb.Frame(win, padding=12, style="Root.TFrame")
        cont.pack(fill="both", expand=True)

        self._crear_header_ventana(
            cont,
            "Editar paciente",
            "Busca por NSS, cédula, nombre o teléfono y corrige datos sin generar una hoja nueva.",
            "🖉"
        )

        footer_edicion = tb.Frame(cont, padding=(8, 8), style="Card.TFrame")
        footer_edicion.pack(side="bottom", fill="x", pady=(8, 0))

        patient_card = tb.Frame(cont, padding=10, style="Card.TFrame")
        patient_card.pack(fill="both", expand=True)
        patient_card.columnconfigure(1, weight=1)
        patient_card.columnconfigure(3, weight=1)

        tb.Label(
            patient_card,
            text="Buscar por NSS, cédula, nombre o teléfono:",
            background="#0E1B2B",
            font=("Arial", 10, "bold")
        ).grid(row=0, column=0, sticky="w", padx=5, pady=(0, 6))

        buscar_var = tk.StringVar()
        ent_buscar = tb.Entry(patient_card, textvariable=buscar_var, width=34)
        ent_buscar.grid(row=0, column=1, sticky="ew", padx=5, pady=(0, 6), ipady=3)

        btn_buscar = tb.Button(patient_card, text="🔎  Buscar", bootstyle=PRIMARY)
        btn_buscar.grid(row=0, column=2, sticky="w", padx=5, pady=(0, 6), ipady=3)

        estado_var = tk.StringVar(value="Ingrese un dato del paciente y presione Buscar.")
        tb.Label(
            patient_card,
            textvariable=estado_var,
            style="Muted.TLabel",
            background="#0E1B2B"
        ).grid(row=1, column=0, columnspan=4, sticky="w", padx=5, pady=(0, 6))

        resultados_tree = ttk.Treeview(
            patient_card,
            columns=("id", "fecha", "nombre", "nss", "cedula", "telefono"),
            show="headings",
            height=4,
            style="Modern.Treeview"
        )
        resultados_tree.grid(row=2, column=0, columnspan=4, sticky="ew", padx=5, pady=(0, 8))
        for col, title, width in [
            ("id", "ID", 55),
            ("fecha", "Fecha", 90),
            ("nombre", "Nombre", 310),
            ("nss", "NSS", 140),
            ("cedula", "Cédula", 130),
            ("telefono", "Teléfono", 120),
        ]:
            resultados_tree.heading(col, text=title)
            resultados_tree.column(col, width=width, anchor="center" if col != "nombre" else "w")

        ttk.Separator(patient_card, orient="horizontal").grid(row=3, column=0, columnspan=4, sticky="ew", pady=(2, 8))

        campos = {}
        ayudas = {
            "Nombre": "Nombre completo del paciente",
            "Cédula": "11 dígitos",
            "Teléfono": "10 dígitos",
            "NSS": "Número de seguro o SIN SEGURO",
            "Dirección": "Dirección del paciente",
            "Nacionalidad": "Nacionalidad",
            "Aseguradora (ARS)": "Ej.: SUB, HUMANO, MAPFRE, AVANZADA…",
        }

        filas = [
            ("Nombre", "Cédula"),
            ("Teléfono", "NSS"),
            ("Aseguradora (ARS)", "Nacionalidad"),
            ("Dirección", None),
        ]

        row_base = 4
        for i, (campo_izq, campo_der) in enumerate(filas):
            r = row_base + i

            tb.Label(
                patient_card,
                text=f"{campo_izq}:",
                font=("Arial", 10, "bold"),
                background="#0E1B2B"
            ).grid(row=r, column=0, sticky="w", padx=(5, 4), pady=5)

            ent_izq = tb.Entry(patient_card)
            ent_izq.grid(row=r, column=1, sticky="ew", padx=(0, 12), pady=5, ipady=4)
            try:
                ent_izq.insert(0, ayudas.get(campo_izq, ""))
                ent_izq.delete(0, tk.END)
            except Exception:
                pass
            campos[campo_izq] = ent_izq

            if campo_der:
                tb.Label(
                    patient_card,
                    text=f"{campo_der}:",
                    font=("Arial", 10, "bold"),
                    background="#0E1B2B"
                ).grid(row=r, column=2, sticky="w", padx=(5, 4), pady=5)

                ent_der = tb.Entry(patient_card)
                ent_der.grid(row=r, column=3, sticky="ew", padx=(0, 5), pady=5, ipady=4)
                campos[campo_der] = ent_der
            else:
                ent_izq.grid(row=r, column=1, columnspan=3, sticky="ew", padx=(0, 5), pady=5, ipady=4)

        ayuda_var = tk.StringVar(
            value="Consejo: si un paciente estaba SIN SEGURO y luego aparece vigente, coloque NSS y ARS aquí; quedará actualizado como asegurado."
        )
        tb.Label(
            patient_card,
            textvariable=ayuda_var,
            style="Muted.TLabel",
            background="#0E1B2B",
            wraplength=980,
            justify="left"
        ).grid(row=row_base + len(filas), column=0, columnspan=4, sticky="w", padx=5, pady=(8, 0))

        original_identidad = {"valor": "", "paciente_id": None}

        def _llenar_formulario_paciente(data, ident):
            original_identidad["valor"] = ident
            original_identidad["paciente_id"] = data.get("paciente_id")
            valores = {
                "Nombre": data.get("nombre", ""),
                "Cédula": data.get("cedula", ""),
                "Teléfono": data.get("telefono", ""),
                "NSS": data.get("nss", ""),
                "Dirección": data.get("direccion", ""),
                "Nacionalidad": data.get("nacionalidad", ""),
                "Aseguradora (ARS)": seguro_para_mostrar(normalizar_seguro(data.get("ars", ""), data.get("nss", ""))),
            }
            for k, ent in campos.items():
                ent.delete(0, tk.END)
                ent.insert(0, valores.get(k, ""))

            estado_var.set(f"Registro cargado. Última atención ID: {data.get('id', 'N/A')}")
            try:
                btn_eliminar.configure(state=tk.NORMAL if original_identidad["paciente_id"] else tk.DISABLED)
            except (NameError, tk.TclError):
                pass
            self.set_status("Paciente cargado para edición", "ok")

        def cargar_paciente():
            ident = buscar_var.get().strip()
            if not ident:
                messagebox.showwarning("Buscar", "Ingrese NSS, cédula, nombre o teléfono para buscar.")
                return

            for i in resultados_tree.get_children():
                resultados_tree.delete(i)
            resultados_tree.insert("", "end", values=("", "", "Cargando datos…", "", "", ""))
            estado_var.set("Cargando datos del paciente…")
            win.update_idletasks()

            def _buscar():
                try:
                    for i in resultados_tree.get_children():
                        resultados_tree.delete(i)

                    resultados = self.db.buscar_pacientes_avanzado(ident)
                    if not resultados:
                        data = self.db.buscar_paciente_para_edicion(ident)
                        if not data:
                            estado_var.set("No se encontraron coincidencias.")
                            self.set_status("No se encontró paciente para editar", "warning")
                            return
                        resultados = [data]

                    for r in resultados:
                        resultados_tree.insert(
                            "",
                            "end",
                            values=(
                                r.get("id", ""),
                                r.get("fecha", ""),
                                r.get("nombre", ""),
                                r.get("nss", ""),
                                r.get("cedula", ""),
                                r.get("telefono", "")
                            )
                        )

                    primero = resultados[0]
                    ident_carga = (
                        f"A:{primero.get('id')}" if primero.get("id")
                        else str(primero.get("nss") or primero.get("cedula") or ident)
                    )
                    _llenar_formulario_paciente(primero, ident_carga)
                    estado_var.set(f"{len(resultados)} coincidencia(s). Seleccione una fila para cargar otra.")
                except Exception as e:
                    estado_var.set("Error al buscar paciente.")
                    messagebox.showerror("Error", f"No se pudo buscar el paciente:\n{str(e)}")

            try:
                win.after(40, _buscar)
            except Exception:
                _buscar()

        def seleccionar_resultado_paciente(_=None):
            sel = resultados_tree.selection()
            if not sel:
                return
            vals = resultados_tree.item(sel[0], "values")
            ident = f"A:{vals[0]}" if vals[0] else (vals[3] or vals[4])
            data = self.db.buscar_paciente_para_edicion(str(ident))
            if data:
                _llenar_formulario_paciente(data, str(ident))

        resultados_tree.bind("<<TreeviewSelect>>", seleccionar_resultado_paciente)

        def guardar_edicion():
            if not original_identidad["valor"]:
                messagebox.showwarning("Guardar", "Primero busque y cargue un paciente.")
                return

            nuevos = {k: ent.get().strip() for k, ent in campos.items()}

            snapshot_paciente = self.db.buscar_paciente_para_edicion(original_identidad["valor"])
            if not snapshot_paciente:
                messagebox.showwarning("Guardar", "No se pudo cargar los datos actuales del paciente.")
                return

            def _get_nombre(d):
                return (d.get("nombre", d.get("Nombre", "")) or "").strip()

            def _get_cedula(d):
                return (d.get("cedula", d.get("Cédula", "")) or "").strip().replace("-", "")

            def _get_telefono(d):
                return (d.get("telefono", d.get("Teléfono", "")) or "").strip().replace("-", "")

            def _get_nss(d):
                return re.sub(r"\D", "", (d.get("nss", d.get("NSS", "")) or "").strip().upper())

            def _get_direccion(d):
                return (d.get("direccion", d.get("Dirección", "")) or "").strip()

            def _get_nacionalidad(d):
                return (d.get("nacionalidad", d.get("Nacionalidad", "")) or "").strip()

            def _get_ars(d, nss_ref=""):
                return normalizar_seguro(
                    d.get("ars", d.get("ARS", d.get("Aseguradora (ARS)", "")) or ""),
                    nss_ref or ""
                )

            nuevo_nombre = _get_nombre(nuevos)
            viejo_nombre = _get_nombre(snapshot_paciente)

            nueva_cedula = _get_cedula(nuevos)
            vieja_cedula = _get_cedula(snapshot_paciente)

            nuevo_telefono = _get_telefono(nuevos)
            viejo_telefono = _get_telefono(snapshot_paciente)

            nuevo_nss = _get_nss(nuevos)
            viejo_nss = _get_nss(snapshot_paciente)

            nueva_direccion = _get_direccion(nuevos)
            vieja_direccion = _get_direccion(snapshot_paciente)

            nueva_nacionalidad = _get_nacionalidad(nuevos)
            vieja_nacionalidad = _get_nacionalidad(snapshot_paciente)

            nueva_ars = _get_ars(nuevos, nuevo_nss)
            vieja_ars = _get_ars(snapshot_paciente, viejo_nss)

            hay_cambios = (
                nuevo_nombre != viejo_nombre or
                nueva_cedula != vieja_cedula or
                nuevo_telefono != viejo_telefono or
                nuevo_nss != viejo_nss or
                nueva_ars != vieja_ars or
                nueva_direccion != vieja_direccion or
                nueva_nacionalidad != vieja_nacionalidad
            )

            if not hay_cambios:
                messagebox.showinfo("Guardar", "No se detectaron cambios. No se guardó nada.")
                return

            if not nuevos.get("Nombre"):
                messagebox.showerror("Validación", "El nombre no puede quedar vacío.")
                campos["Nombre"].focus_set()
                return

            ced = nuevos.get("Cédula", "").replace("-", "")
            tel = nuevos.get("Teléfono", "").replace("-", "")
            nss = nuevos.get("NSS", "").upper()
            ars = nuevos.get("Aseguradora (ARS)", "")

            if ced and not is_valid_cedula_key(ced):
                messagebox.showerror("Validación", "La cédula debe tener 11 dígitos o dejarse vacía.")
                campos["Cédula"].focus_set()
                return

            if tel and not (tel.isdigit() and len(tel) == 10):
                messagebox.showerror("Validación", "El teléfono debe tener 10 dígitos o dejarse vacío.")
                campos["Teléfono"].focus_set()
                return

            if nss and not is_valid_nss_key(nss) and nss not in ["SIN SEGURO", "NO", "N/S", "NS", "N\\S"]:
                messagebox.showerror(
                    "Validación",
                    "El NSS debe ser numérico. Si no tiene seguro, escriba SIN SEGURO o déjelo vacío."
                )
                campos["NSS"].focus_set()
                return

            if ars_es_corta_invalida(ars):
                messagebox.showerror(
                    "Validación",
                    "La ARS es demasiado corta o no reconocida. No se guardan ARS accidentales menores de 4 letras."
                )
                campos["Aseguradora (ARS)"].focus_set()
                return

            try:
                alcance = messagebox.askyesnocancel(
                    "Alcance de la modificación",
                    "¿Desea actualizar también la ficha actual del paciente?\n\n"
                    "Sí: esta atención y la ficha actual\n"
                    "No: solamente esta atención\n"
                    "Cancelar: no guardar",
                    parent=win,
                )
                if alcance is None:
                    return
                actualizar_ficha = alcance is True
                snapshot_paciente = self.db.buscar_paciente_para_edicion(original_identidad["valor"])
                at_count, pac_count = self.db.actualizar_datos_paciente_por_identidad(
                    original_identidad["valor"],
                    nuevos,
                    actualizar_ficha=actualizar_ficha,
                )

                def _undo_edit_paciente():
                    if not snapshot_paciente:
                        messagebox.showwarning("Deshacer", "No se encontró una copia anterior para revertir.")
                        return
                    anteriores = {
                        "Nombre": snapshot_paciente.get("nombre", ""),
                        "Cédula": snapshot_paciente.get("cedula", ""),
                        "Teléfono": snapshot_paciente.get("telefono", ""),
                        "NSS": snapshot_paciente.get("nss", ""),
                        "Dirección": snapshot_paciente.get("direccion", ""),
                        "Nacionalidad": snapshot_paciente.get("nacionalidad", ""),
                        "Aseguradora (ARS)": snapshot_paciente.get("ars", ""),
                    }
                    ident_undo = (
                        snapshot_paciente.get("nss")
                        or snapshot_paciente.get("cedula")
                        or f"A:{snapshot_paciente.get('id')}"
                    )
                    self.db.actualizar_datos_paciente_por_identidad(
                        str(ident_undo),
                        anteriores,
                        actualizar_ficha=actualizar_ficha,
                    )
                    self._reconstruir_excel_si_necesario("deshacer edición paciente", antes=nuevos, despues=anteriores, forzar=False)
                    self._invalidar_caches_datos()
                    self._refrescar_resumen_en_vivo()
                    messagebox.showinfo("Deshacer", "Edición del paciente revertida.")

                self._push_undo_action("edición de paciente", _undo_edit_paciente)

                self._reconstruir_excel_si_necesario("editar paciente", antes=snapshot_paciente or {}, despues=nuevos, forzar=False)
                self._invalidar_caches_datos()
                self._ars_catalogo = self._obtener_catalogo_ars()
                self._refrescar_resumen_en_vivo()

                estado_var.set(f"Guardado: {at_count} atención(es) y {pac_count} ficha(s) actualizada(s).")
                self.set_status("Datos del paciente actualizados", "ok")
                self._mostrar_notificacion("Datos del paciente actualizados. Ctrl+Z para deshacer.", on_undo=_undo_edit_paciente, autohide_ms=5000)
                messagebox.showinfo("Guardado", "Datos actualizados correctamente. No se generó una hoja nueva.")
            except PermissionError:
                messagebox.showwarning(
                    "Excel abierto",
                    "El listado de Excel está abierto.\n\nCierre el archivo y vuelva a intentar guardar."
                )
            except Exception as e:
                messagebox.showerror("Error", f"No se pudo guardar la edición:\n{str(e)}")

        btn_buscar.configure(command=cargar_paciente)
        ent_buscar.bind("<Return>", lambda e: cargar_paciente())

        try:
            campos["Cédula"].bind("<KeyRelease>", lambda e: self.limitar_caracteres(campos["Cédula"], 11))
            campos["Teléfono"].bind("<KeyRelease>", lambda e: self.limitar_caracteres(campos["Teléfono"], 10))
        except Exception:
            pass

        def limpiar_campos_edicion():
            campos["Nombre"].delete(0, tk.END)
            campos["Cédula"].delete(0, tk.END)
            campos["Teléfono"].delete(0, tk.END)
            campos["NSS"].delete(0, tk.END)
            campos["Dirección"].delete(0, tk.END)
            campos["Nacionalidad"].delete(0, tk.END)
            campos["Aseguradora (ARS)"].delete(0, tk.END)
            buscar_var.set("")
            for i in resultados_tree.get_children():
                resultados_tree.delete(i)
            original_identidad.update({"valor": "", "paciente_id": None})
            try:
                btn_eliminar.configure(state=tk.DISABLED)
            except (NameError, tk.TclError):
                pass
            estado_var.set("Campos limpiados. Busque un paciente para cargar datos.")
            self.set_status("Campos limpiados en edición paciente", "ok")

        def eliminar_paciente_total():
            paciente_id = original_identidad.get("paciente_id")
            if not paciente_id:
                messagebox.showwarning("Eliminar paciente", "Primero busque y cargue el paciente que desea eliminar.", parent=win)
                return

            try:
                resumen = self.db.previsualizar_eliminacion_paciente(paciente_id)
            except Exception as e:
                APP_LOG.exception("No se pudo preparar la eliminación total del paciente")
                messagebox.showerror("Eliminar paciente", f"No se pudo consultar el paciente:\n{str(e)}", parent=win)
                return

            if not resumen:
                messagebox.showwarning("Eliminar paciente", "El paciente ya no existe en la base de datos.", parent=win)
                return
            if not resumen.get("seguro"):
                messagebox.showwarning(
                    "Identidad insuficiente",
                    resumen.get("motivo", "No se puede identificar al paciente con seguridad."),
                    parent=win,
                )
                return

            paciente = resumen["paciente"]
            nombre = (paciente.get("nombre") or "SIN NOMBRE").upper()
            total_atenciones = len(resumen["atenciones"])
            total_fichas = int(resumen.get("fichas", 0) or 0)
            total_auditorias = int(resumen.get("auditorias", 0) or 0)
            advertencia = (
                "Esta eliminación es permanente y está destinada a retirar pacientes de prueba.\n\n"
                f"Paciente: {nombre}\n"
                f"NSS: {paciente.get('nss') or 'N/A'}\n"
                f"Cédula: {paciente.get('cedula') or 'N/A'}\n"
                f"Atenciones que se eliminarán: {total_atenciones}\n"
                f"Fichas que se eliminarán: {total_fichas}\n"
                f"Registros clínicos de auditoría que se retirarán: {total_auditorias}\n"
                "Se conservará un evento administrativo anonimizado y un respaldo verificable.\n\n"
                "Los ID restantes no serán renumerados. ¿Desea continuar?"
            )
            if not messagebox.askyesno("Eliminar paciente por completo", advertencia, icon="warning", parent=win):
                return

            if not messagebox.askyesno(
                "Confirmar dato de prueba",
                "Confirme que esta ficha pertenece exclusivamente a un paciente de prueba.\n\n"
                "Una atención clínica real debe anularse, no purgarse.",
                icon="warning",
                parent=win,
            ):
                return

            motivo = simpledialog.askstring(
                "Motivo obligatorio",
                "Indique por qué se confirma que esta ficha es de prueba:",
                parent=win,
            )
            if motivo is None or len(motivo.strip()) < 8:
                messagebox.showwarning(
                    "Motivo requerido",
                    "La purga requiere un motivo de al menos 8 caracteres.",
                    parent=win,
                )
                return

            actor = self._solicitar_autorizacion_admin(
                "PURGAR_PACIENTE_DE_PRUEBA", parent=win, force=True
            )
            if not actor:
                return

            confirmacion = simpledialog.askstring(
                "Confirmación final",
                f"Para confirmar, escriba exactamente ELIMINAR {paciente_id}:",
                parent=win,
            )
            if (confirmacion or "").strip().upper() != f"ELIMINAR {paciente_id}":
                messagebox.showinfo("Cancelado", "No se eliminó ningún dato.", parent=win)
                return

            afecta_turno = any(
                self._registro_esta_en_turno_actual(atencion)
                for atencion in resumen["atenciones"]
            )
            try:
                resultado = self.db.eliminar_paciente_completo(
                    paciente_id,
                    motivo.strip(),
                    actor,
                    confirmado_prueba=True,
                )
                if not resultado or not resultado.get("seguro"):
                    messagebox.showwarning("Eliminar paciente", "No se pudo completar la eliminación.", parent=win)
                    return
                self.security.audit(
                    "PATIENT_PURGED",
                    actor=actor,
                    success=not bool(resultado.get("documentos_pendientes")),
                    detail=str(resultado.get("purga_event_hash") or ""),
                )

                aviso_excel = ""
                if afecta_turno:
                    turno_cfg = cargar_turno_config()
                    if turno_cfg:
                        try:
                            reconstruir_excel_turno(self.db, turno_cfg)
                        except Exception:
                            APP_LOG.exception("El paciente se eliminó, pero no se pudo reconstruir el Excel")
                            aviso_excel = (
                                "\n\nEl paciente sí fue eliminado, pero el Excel no pudo actualizarse. "
                                "Cierre el archivo y use la reconstrucción del turno."
                            )

                self._invalidar_caches_datos()
                self._refrescar_resumen_en_vivo()
                limpiar_campos_edicion()
                pendientes = resultado.get("documentos_pendientes") or []
                mensaje = (
                    f"Se eliminó la ficha de prueba seleccionada.\n\n"
                    f"Atenciones: {resultado.get('atenciones_eliminadas', 0)}\n"
                    f"Fichas: {resultado.get('fichas_eliminadas', 0)}\n"
                    f"Auditorías redactadas: {resultado.get('auditorias_redactadas', 0)}\n"
                    f"Documentos eliminados: {resultado.get('documentos_eliminados', 0)}\n\n"
                    f"Los demás ID permanecen sin cambios.{aviso_excel}"
                )
                if pendientes:
                    messagebox.showwarning(
                        "Purga incompleta",
                        mensaje
                        + "\n\nUno o más documentos quedaron aislados en cuarentena. "
                        "Revise el registro técnico antes de considerar completada la purga.",
                        parent=win,
                    )
                else:
                    messagebox.showinfo("Paciente eliminado", mensaje, parent=win)
            except PermissionError:
                messagebox.showwarning("Eliminar paciente", "No se pudo acceder a la base de datos.", parent=win)
            except Exception as e:
                messagebox.showerror("Eliminar paciente", f"No se pudo eliminar el paciente:\n{str(e)}", parent=win)

        tb.Button(footer_edicion, text="🧹  Limpiar campos", bootstyle=SECONDARY, command=limpiar_campos_edicion, width=18).pack(side="left", padx=5, ipady=5)
        tb.Button(footer_edicion, text="💾  Guardar cambios", bootstyle=SUCCESS, command=guardar_edicion, width=22).pack(side="left", padx=5, ipady=5)
        btn_eliminar = tb.Button(
            footer_edicion,
            text="Eliminar paciente de prueba",
            bootstyle=DANGER,
            command=eliminar_paciente_total,
            width=25,
            state=tk.DISABLED,
        )
        btn_eliminar.pack(side="left", padx=5, ipady=5)
        tb.Button(footer_edicion, text="Cerrar", bootstyle=SECONDARY, command=win.destroy, width=12).pack(side="right", padx=5, ipady=5)

        if prefill_identidad:
            try:
                buscar_var.set(str(prefill_identidad).strip())
                estado_var.set("Cargando datos…")
                win.after(160, cargar_paciente)
            except Exception:
                pass

        try:
            ent_buscar.focus_set()
        except Exception:
            pass

    def _abrir_configuracion_interna(self, prefill_identidad=None):
        if not self._exigir_permiso(CAP_INTERNAL_CONFIG, "abrir la configuración interna"):
            return
        if prefill_identidad:
            self._abrir_edicion_paciente(prefill_identidad=prefill_identidad)
            return

        config_open_started = _time.perf_counter()
        APP_LOG.info(
            "CONFIG_OPEN_START thread=%s",
            threading.current_thread().name,
        )

        existing = getattr(self, "configuracion_interna_win", None)
        if self._ventana_activa(existing) and not bool(
            getattr(existing, "_config_content_ready", False)
        ):
            APP_LOG.error("CONFIG_DIALOG_STALE_EMPTY_RECREATED")
            try:
                existing.destroy()
            finally:
                self.configuracion_interna_win = None

        win = self._crear_toplevel_estable("Configuración interna", "1280x800", "configuracion_interna_win")
        if win is None:
            return
        self._aplicar_tamano_configuracion(win)

        try:
            self.configuracion_interna_win = win
        except Exception:
            pass

        cont = tb.Frame(win, padding=14, style="Root.TFrame")
        cont.pack(fill="both", expand=True)

        self._crear_header_ventana(
            cont,
            "Configuración interna",
            "Administra ARS, usuarios de turno, respaldos, revisión NSS y preferencias.",
            "⚙"
        )

        notebook = ttk.Notebook(cont)
        notebook.pack(fill="both", expand=True)


        loading_tab = tb.Frame(notebook, padding=18, style="Card.TFrame")
        notebook.add(loading_tab, text="Cargando…")
        tb.Label(
            loading_tab,
            text="Configuración lista. Cargando secciones…",
            style="Muted.TLabel",
            background="#0E1B2B",
        ).pack(anchor="center", pady=24)
        win._config_content_ready = False

        def _build_config_content():
            if not win.winfo_exists():
                return
            build_started = _time.perf_counter()
            try:
                forget = getattr(notebook, "forget", None)
                if callable(forget):
                    forget(loading_tab)
                else:
                    loading_tab.destroy()

                # ---------------- TAB ARS ----------------
                tab_ars = tb.Frame(notebook, padding=12, style="Card.TFrame")
                notebook.add(tab_ars, text="Administrar ARS")
                tab_ars.columnconfigure(0, weight=1)
                tab_ars.columnconfigure(1, weight=1)

                info_ars = tk.StringVar(value="Cargando datos de ARS…")
                tb.Label(tab_ars, textvariable=info_ars, style="Muted.TLabel", background="#0E1B2B").grid(row=0, column=0, columnspan=2, sticky="w", pady=(0, 8))

                tree_ars = ttk.Treeview(tab_ars, columns=("ars", "cantidad"), show="headings", height=14, style="Modern.Treeview")
                tree_ars.grid(row=1, column=0, columnspan=2, sticky="nsew", pady=(0, 8))
                tree_ars.heading("ars", text="ARS / Seguro")
                tree_ars.heading("cantidad", text="Cantidad")
                tree_ars.column("ars", width=420, anchor="w")
                tree_ars.column("cantidad", width=120, anchor="center")
                tab_ars.rowconfigure(1, weight=1)

                tb.Label(tab_ars, text="ARS actual:", background="#0E1B2B").grid(row=2, column=0, sticky="w", padx=5, pady=(8, 2))
                ars_actual_var = tk.StringVar()
                ent_actual = tb.Entry(tab_ars, textvariable=ars_actual_var)
                ent_actual.grid(row=3, column=0, sticky="ew", padx=5, pady=(0, 6), ipady=4)

                tb.Label(tab_ars, text="Reemplazar por:", background="#0E1B2B").grid(row=2, column=1, sticky="w", padx=5, pady=(8, 2))
                ars_nueva_var = tk.StringVar()
                ent_nueva = tb.Entry(tab_ars, textvariable=ars_nueva_var)
                ent_nueva.grid(row=3, column=1, sticky="ew", padx=5, pady=(0, 6), ipady=4)

                def cargar_ars():
                    for i in tree_ars.get_children():
                        tree_ars.delete(i)
                    tree_ars.insert("", "end", values=("Cargando datos…", ""))
                    info_ars.set("Cargando datos de ARS…")

                    def _cargar():
                        rows = list(self.db.listar_ars_conteo())
                        catalogo = self._obtener_catalogo_ars()
                        return rows, catalogo

                    def _finalizar(resultado):
                        if not win.winfo_exists():
                            return
                        rows, catalogo = resultado
                        tree_ars.delete(*tree_ars.get_children())
                        for ars, cantidad in rows:
                            tree_ars.insert("", "end", values=(seguro_para_mostrar(ars), cantidad))
                        self._ars_catalogo = catalogo
                        self._invalidar_cache_ars()
                        info_ars.set("Seleccione una ARS para corregirla o use las acciones de limpieza.")

                    def _error(err):
                        info_ars.set("No se pudieron cargar las ARS.")
                        messagebox.showerror("ARS", f"No se pudo cargar la lista de ARS:\n{str(err)}")

                    self._ejecutar_en_segundo_plano(
                        "Cargando ARS…",
                        _cargar,
                        al_terminar=_finalizar,
                        al_error=_error,
                    )

                def seleccionar_ars(_evt=None):
                    sel = tree_ars.selection()
                    if not sel:
                        return
                    vals = tree_ars.item(sel[0], "values")
                    ars_actual_var.set(vals[0])

                def reemplazar_ars():
                    actual = ars_actual_var.get().strip()
                    nueva = ars_nueva_var.get().strip()
                    if not actual or not nueva:
                        messagebox.showwarning("ARS", "Indique la ARS actual y la nueva ARS.")
                        return
                    if ars_es_corta_invalida(nueva):
                        messagebox.showerror("ARS", "La nueva ARS es demasiado corta o no reconocida. Escriba al menos 4 letras o una referencia válida.")
                        ent_nueva.focus_set()
                        return
                    if not messagebox.askyesno(
                        "Confirmación",
                        f"¿Actualizar {actual} a {nueva} en las fichas actuales?\n\n"
                        "Las atenciones históricas conservarán el valor registrado en su momento.",
                    ):
                        return
                    total = self.db.reemplazar_ars_global(actual, nueva)
                    self.security.audit(
                        "ARS_CURRENT_RECORDS_REPLACED",
                        actor=self._admin_authorized_actor or self._actor_actual(),
                        success=True,
                        detail=f"{actual}->{nueva}; fichas={total}",
                    )
                    cargar_ars()
                    self.set_status(f"ARS actualizadas: {total}", "ok")
                    messagebox.showinfo("ARS", f"Registros actualizados: {total}")

                def normalizar_todo():
                    total = self.db.normalizar_todas_ars()
                    self.security.audit(
                        "ARS_CURRENT_RECORDS_NORMALIZED",
                        actor=self._admin_authorized_actor or self._actor_actual(),
                        success=True,
                        detail=f"fichas={total}",
                    )
                    cargar_ars()
                    self.set_status(f"ARS normalizadas: {total}", "ok")
                    messagebox.showinfo("ARS", f"ARS normalizadas: {total}")

                def limpiar_cortas():
                    total = self.db.limpiar_ars_cortas_invalidas()
                    self.security.audit(
                        "ARS_INVALID_CURRENT_RECORDS_CLEANED",
                        actor=self._admin_authorized_actor or self._actor_actual(),
                        success=True,
                        detail=f"fichas={total}",
                    )
                    cargar_ars()
                    self.set_status(f"ARS inválidas limpiadas: {total}", "ok")
                    messagebox.showinfo("ARS", f"ARS inválidas convertidas a SIN SEGURO: {total}")

                tree_ars.bind("<<TreeviewSelect>>", seleccionar_ars)

                menu_ars = tk.Menu(win, tearoff=0)
                menu_ars.add_command(label="✏ Cargar para reemplazar", command=seleccionar_ars)
                def eliminar_ars_sel():
                    sel = tree_ars.selection()
                    if not sel:
                        messagebox.showwarning("ARS", "Seleccione una ARS para eliminar."); return
                    actual = tree_ars.item(sel[0], "values")[0]
                    if not messagebox.askyesno(
                        "Eliminar ARS",
                        f"¿Convertir las fichas actuales de '{actual}' a SIN SEGURO?\n\n"
                        "Las atenciones históricas no se modificarán.",
                    ):
                        return
                    total = self.db.eliminar_ars_global(actual)
                    self.security.audit(
                        "ARS_CURRENT_RECORDS_REMOVED",
                        actor=self._admin_authorized_actor or self._actor_actual(),
                        success=True,
                        detail=f"ars={actual}; fichas={total}",
                    )
                    cargar_ars(); self._actualizar_resumen_turno_panel()
                    messagebox.showinfo("ARS", f"Registros convertidos a SIN SEGURO: {total}")
                menu_ars.add_command(label="🗑 Eliminar ARS (pasar a SIN SEGURO)", command=eliminar_ars_sel)
                def mostrar_menu_ars(event):
                    item = tree_ars.identify_row(event.y)
                    if item:
                        tree_ars.selection_set(item); tree_ars.focus(item); menu_ars.tk_popup(event.x_root, event.y_root)
                tree_ars.bind("<Button-3>", mostrar_menu_ars)

                acciones_ars = tb.Frame(tab_ars, padding=(0, 10), style="Card.TFrame")
                acciones_ars.grid(row=4, column=0, columnspan=2, sticky="ew")
                tb.Button(acciones_ars, text="↻  Actualizar lista", bootstyle=SECONDARY, command=cargar_ars, width=18).pack(side="left", padx=5, ipady=5)
                tb.Button(acciones_ars, text="✓  Normalizar ARS", bootstyle=INFO, command=normalizar_todo, width=18).pack(side="left", padx=5, ipady=5)
                tb.Button(acciones_ars, text="⌫  Limpiar cortas", bootstyle=WARNING, command=limpiar_cortas, width=18).pack(side="left", padx=5, ipady=5)
                tb.Button(acciones_ars, text="✓  Reemplazar", bootstyle=SUCCESS, command=reemplazar_ars, width=16).pack(side="left", padx=5, ipady=5)
                tb.Button(acciones_ars, text="🗑️  Eliminar ARS", bootstyle=DANGER, command=eliminar_ars_sel, width=16).pack(side="left", padx=5, ipady=5)
                tb.Button(acciones_ars, text="Cerrar", bootstyle=SECONDARY, command=win.destroy, width=12).pack(side="right", padx=5, ipady=5)


                # ---------------- TAB CATÁLOGO ARS ----------------
                tab_catalogo = tb.Frame(notebook, padding=12, style="Card.TFrame")
                notebook.add(tab_catalogo, text="Catálogo ARS")
                tab_catalogo.columnconfigure(0, weight=1); tab_catalogo.columnconfigure(1, weight=1); tab_catalogo.rowconfigure(1, weight=1)
                tb.Label(tab_catalogo, text="ARS oficial", background="#0E1B2B").grid(row=0, column=0, sticky="w", padx=5, pady=4)
                tb.Label(tab_catalogo, text="Alias aceptados (separados por coma)", background="#0E1B2B").grid(row=0, column=1, sticky="w", padx=5, pady=4)
                tree_cat = ttk.Treeview(tab_catalogo, columns=("oficial", "alias"), show="headings", height=12, style="Modern.Treeview")
                tree_cat.grid(row=1, column=0, columnspan=2, sticky="nsew", pady=(0, 8))
                tree_cat.heading("oficial", text="ARS oficial"); tree_cat.heading("alias", text="Alias aceptados")
                tree_cat.column("oficial", width=300, anchor="w"); tree_cat.column("alias", width=520, anchor="w")
                oficial_var = tk.StringVar(); alias_var = tk.StringVar()
                ent_oficial = tb.Entry(tab_catalogo, textvariable=oficial_var); ent_alias = tb.Entry(tab_catalogo, textvariable=alias_var)
                ent_oficial.grid(row=2, column=0, sticky="ew", padx=5, pady=4, ipady=4); ent_alias.grid(row=2, column=1, sticky="ew", padx=5, pady=4, ipady=4)
                catalogo_data = {k: list(v) for k, v in cargar_catalogo_ars().items()}
                def cargar_catalogo_tree():
                    tree_cat.delete(*tree_cat.get_children())
                    tree_cat.insert("", "end", values=("Cargando catálogo…", ""))

                    def _cargar():
                        return [(k, catalogo_data[k]) for k in sorted(catalogo_data)]

                    def _finalizar(resultado):
                        if not win.winfo_exists():
                            return
                        tree_cat.delete(*tree_cat.get_children())
                        for k, aliases in resultado:
                            tree_cat.insert("", "end", values=(k, ", ".join(aliases)))

                    self._ejecutar_en_segundo_plano(
                        "Cargando catálogo ARS…",
                        _cargar,
                        al_terminar=_finalizar,
                    )
                def seleccionar_catalogo(_=None):
                    sel = tree_cat.selection()
                    if not sel: return
                    vals = tree_cat.item(sel[0], "values"); oficial_var.set(vals[0]); alias_var.set(vals[1])
                def guardar_alias():
                    k = _limpiar_texto_seguro(oficial_var.get())
                    if not k: messagebox.showwarning("Catálogo", "Escriba una ARS oficial."); return
                    aliases = [a.strip() for a in alias_var.get().split(",") if a.strip()]
                    catalogo_data[k] = aliases; guardar_catalogo_ars(catalogo_data); cargar_catalogo_tree(); self._ars_catalogo = self._obtener_catalogo_ars()
                    messagebox.showinfo("Catálogo", "Alias guardados correctamente.")
                def eliminar_alias():
                    k = _limpiar_texto_seguro(oficial_var.get())
                    if k in catalogo_data and messagebox.askyesno("Catálogo", f"¿Eliminar {k} del catálogo editable?"):
                        catalogo_data.pop(k, None); guardar_catalogo_ars(catalogo_data); oficial_var.set(""); alias_var.set(""); cargar_catalogo_tree()
                tree_cat.bind("<<TreeviewSelect>>", seleccionar_catalogo)
                barra_cat = tb.Frame(tab_catalogo, style="Card.TFrame"); barra_cat.grid(row=3, column=0, columnspan=2, sticky="ew", pady=8)
                tb.Button(barra_cat, text="💾  Guardar alias", bootstyle=SUCCESS, command=guardar_alias).pack(side="left", padx=5, ipady=5)
                tb.Button(barra_cat, text="🗑️  Eliminar alias", bootstyle=DANGER, command=eliminar_alias).pack(side="left", padx=5, ipady=5)
                tb.Button(barra_cat, text="↻  Recargar", bootstyle=SECONDARY, command=cargar_catalogo_tree).pack(side="left", padx=5, ipady=5)


                # ---------------- TAB FORMATO NSS PDF ----------------
                tab_nss = tb.Frame(notebook, padding=12, style="Card.TFrame")
                notebook.add(tab_nss, text="Formato NSS PDF")
                tab_nss.columnconfigure(0, weight=1)
                tab_nss.columnconfigure(1, weight=1)
                tab_nss.rowconfigure(1, weight=1)

                info_nss_var = tk.StringVar(
                    value="Elija una ARS y escriba un ejemplo de cómo debe verse el NSS en el PDF. Ej.: 00896-00258-03."
                )
                tb.Label(tab_nss, textvariable=info_nss_var, style="Muted.TLabel", background="#0E1B2B").grid(
                    row=0, column=0, columnspan=2, sticky="w", pady=(0, 8)
                )

                tree_nss_fmt = ttk.Treeview(
                    tab_nss,
                    columns=("ars", "ejemplo", "patron"),
                    show="headings",
                    height=12,
                    style="Modern.Treeview"
                )
                tree_nss_fmt.grid(row=1, column=0, columnspan=2, sticky="nsew", pady=(0, 8))
                tree_nss_fmt.heading("ars", text="ARS")
                tree_nss_fmt.heading("ejemplo", text="Ejemplo en PDF")
                tree_nss_fmt.heading("patron", text="Patrón automático")
                tree_nss_fmt.column("ars", width=260, anchor="w")
                tree_nss_fmt.column("ejemplo", width=300, anchor="w")
                tree_nss_fmt.column("patron", width=150, anchor="center")

                nss_fmt_data = cargar_formatos_nss_ars()
                ars_disponibles = sorted(set(self._obtener_ars_cache(forzar=True)) | set(nss_fmt_data.keys()) | {"RENACER"})

                nss_ars_var = tk.StringVar()
                nss_ejemplo_var = tk.StringVar(value="00896-00258-03")
                nss_patron_detectado_var = tk.StringVar(value="")
                nss_prueba_sin_guiones_var = tk.StringVar(value="008960025803")

                tb.Label(tab_nss, text="ARS:", background="#0E1B2B").grid(row=2, column=0, sticky="w", padx=5, pady=(8, 2))
                combo_nss_ars = tb.Combobox(tab_nss, textvariable=nss_ars_var, values=ars_disponibles, width=34)
                combo_nss_ars.grid(row=3, column=0, sticky="ew", padx=5, pady=(0, 6), ipady=4)

                tb.Label(tab_nss, text="Ejemplo de NSS con guiones para esa ARS:", background="#0E1B2B").grid(row=2, column=1, sticky="w", padx=5, pady=(8, 2))
                ent_nss_ejemplo = tb.Entry(tab_nss, textvariable=nss_ejemplo_var)
                ent_nss_ejemplo.grid(row=3, column=1, sticky="ew", padx=5, pady=(0, 6), ipady=4)

                tb.Label(tab_nss, text="NSS de prueba sin guiones:", background="#0E1B2B").grid(row=4, column=0, sticky="w", padx=5, pady=(8, 2))
                ent_nss_prueba = tb.Entry(tab_nss, textvariable=nss_prueba_sin_guiones_var)
                ent_nss_prueba.grid(row=5, column=0, sticky="ew", padx=5, pady=(0, 6), ipady=4)

                tb.Label(tab_nss, textvariable=nss_patron_detectado_var, style="Muted.TLabel", background="#0E1B2B").grid(
                    row=5, column=1, sticky="w", padx=5, pady=(0, 6)
                )

                def _ejemplo_desde_patron(patron):
                    base = re.sub(r"\D", "", nss_prueba_sin_guiones_var.get().strip()) or "008960025803"
                    return aplicar_patron_nss(base, patron)

                def refrescar_formatos_nss():
                    tree_nss_fmt.delete(*tree_nss_fmt.get_children())
                    ars_todas = sorted(set(ars_disponibles) | set(nss_fmt_data.keys()))
                    combo_nss_ars.configure(values=ars_todas)
                    for ars_key in ars_todas:
                        patron = nss_fmt_data.get(ars_key, "")
                        ejemplo = _ejemplo_desde_patron(patron) if patron else "Sin formato"
                        tree_nss_fmt.insert("", "end", values=(ars_key, ejemplo, patron or ""))
                    probar_ejemplo_nss()

                def seleccionar_formato_nss(_=None):
                    sel = tree_nss_fmt.selection()
                    if not sel:
                        return
                    vals = tree_nss_fmt.item(sel[0], "values")
                    nss_ars_var.set(vals[0])
                    if vals[1] and vals[1] != "Sin formato":
                        nss_ejemplo_var.set(vals[1])
                    probar_ejemplo_nss()

                def probar_ejemplo_nss(*_):
                    ejemplo = nss_ejemplo_var.get().strip()
                    patron = patron_desde_ejemplo_nss(ejemplo)
                    prueba = re.sub(r"\D", "", nss_prueba_sin_guiones_var.get().strip())
                    if not patron:
                        nss_patron_detectado_var.set("Resultado: escriba el ejemplo con guiones. Ej.: 00896-00258-03")
                        return
                    resultado = aplicar_patron_nss(prueba, patron) if prueba else ejemplo
                    nss_patron_detectado_var.set(f"Patrón detectado: {patron}   |   Vista PDF: {resultado}")

                def guardar_formato_nss():
                    ars_key = _limpiar_texto_seguro(nss_ars_var.get())
                    ejemplo = nss_ejemplo_var.get().strip()
                    patron = patron_desde_ejemplo_nss(ejemplo)

                    if not ars_key:
                        messagebox.showwarning("Formato NSS", "Seleccione o escriba la ARS.")
                        combo_nss_ars.focus_set()
                        return

                    if not patron:
                        messagebox.showwarning(
                            "Formato NSS",
                            "Escriba un ejemplo con guiones para que el sistema detecte el formato.\n\n"
                            "Ejemplo: 00896-00258-03"
                        )
                        ent_nss_ejemplo.focus_set()
                        return

                    nss_fmt_data[ars_key] = patron
                    guardar_formatos_nss_ars(nss_fmt_data)
                    if ars_key not in ars_disponibles:
                        ars_disponibles.append(ars_key)
                    refrescar_formatos_nss()
                    self.set_status("Formato NSS guardado", "ok")
                    messagebox.showinfo(
                        "Formato NSS",
                        f"Formato guardado para {ars_key}.\n\n"
                        f"Ejemplo: {ejemplo}\n"
                        f"Patrón detectado: {patron}\n\n"
                        "Solo afectará el NSS mostrado en PDF."
                    )

                def eliminar_formato_nss():
                    ars_key = _limpiar_texto_seguro(nss_ars_var.get())
                    if not ars_key or ars_key not in nss_fmt_data:
                        messagebox.showwarning("Formato NSS", "Seleccione o escriba una ARS configurada.")
                        return
                    if not messagebox.askyesno("Formato NSS", f"¿Eliminar el formato de NSS para {ars_key}?"):
                        return
                    nss_fmt_data.pop(ars_key, None)
                    guardar_formatos_nss_ars(nss_fmt_data)
                    nss_ejemplo_var.set("")
                    refrescar_formatos_nss()

                tree_nss_fmt.bind("<<TreeviewSelect>>", seleccionar_formato_nss)
                ent_nss_ejemplo.bind("<KeyRelease>", probar_ejemplo_nss)
                ent_nss_prueba.bind("<KeyRelease>", lambda e: refrescar_formatos_nss())

                barra_nss = tb.Frame(tab_nss, style="Card.TFrame")
                barra_nss.grid(row=6, column=0, columnspan=2, sticky="ew", pady=8)
                tb.Button(barra_nss, text="💾  Guardar formato", bootstyle=SUCCESS, command=guardar_formato_nss, width=18).pack(side="left", padx=5, ipady=5)
                tb.Button(barra_nss, text="🗑️  Eliminar formato", bootstyle=DANGER, command=eliminar_formato_nss, width=18).pack(side="left", padx=5, ipady=5)
                tb.Button(barra_nss, text="↻  Recargar", bootstyle=SECONDARY, command=refrescar_formatos_nss, width=14).pack(side="left", padx=5, ipady=5)


                # ---------------- TAB REVISION NSS ----------------
                tab_revision_nss = tb.Frame(notebook, padding=12, style="Card.TFrame")
                notebook.add(tab_revision_nss, text="Revisión NSS")
                tab_revision_nss.columnconfigure(0, weight=1)
                tab_revision_nss.rowconfigure(1, weight=1)
                revision_nss_status = tk.StringVar(
                    value="Casos sin cédula cuyo NSS aparece en fichas con datos diferentes."
                )
                tb.Label(
                    tab_revision_nss,
                    textvariable=revision_nss_status,
                    style="Muted.TLabel",
                    wraplength=1050,
                ).grid(row=0, column=0, sticky="w", pady=(0, 8))

                columnas_revision = (
                    "caso","nss","ficha_nueva","paciente_nuevo","ficha_existente",
                    "paciente_existente","atencion","fecha",
                )
                revision_nss_tree = ttk.Treeview(
                    tab_revision_nss,
                    columns=columnas_revision,
                    show="headings",
                    style="Modern.Treeview",
                    height=14,
                )
                for columna, titulo, ancho in (
                    ("caso","Caso",65),("nss","NSS",130),("ficha_nueva","Ficha nueva",90),
                    ("paciente_nuevo","Paciente nuevo",210),("ficha_existente","Ficha existente",105),
                    ("paciente_existente","Paciente existente",210),("atencion","Atención",80),
                    ("fecha","Fecha",100),
                ):
                    revision_nss_tree.heading(columna, text=titulo)
                    revision_nss_tree.column(columna, width=ancho, anchor="w")
                revision_nss_tree.grid(row=1, column=0, sticky="nsew")
                revisiones_nss = {}


                def cargar_revisiones_nss():
                    revisiones_nss.clear()
                    revision_nss_tree.delete(*revision_nss_tree.get_children())
                    revision_nss_status.set("Cargando revisión NSS en segundo plano…")
                    started = _time.perf_counter()
                    APP_LOG.info("CONFIG_SECTION_LOAD_START section=revision_nss")

                    def _cargar():
                        return list(self.db.listar_revisiones_nss(True, 1000))

                    def _finalizar(rows):
                        if not win.winfo_exists():
                            return
                        revision_nss_tree.delete(*revision_nss_tree.get_children())
                        for row in rows:
                            revision_id = int(row["id"])
                            revisiones_nss[revision_id] = row
                            revision_nss_tree.insert(
                                "", "end", iid=str(revision_id),
                                values=(
                                    revision_id, row.get("nss_normalizado") or "",
                                    row.get("paciente_nuevo_id") or "", row.get("nombre_nuevo") or "",
                                    row.get("paciente_referencia_id") or "", row.get("nombre_referencia") or "",
                                    row.get("atencion_id") or "", row.get("fecha") or "",
                                ),
                            )
                        revision_nss_status.set(
                            f"{len(revisiones_nss)} caso(s) pendiente(s). Esta revisión nunca detiene la admisión."
                        )
                        APP_LOG.info(
                            "CONFIG_SECTION_LOAD_DONE section=revision_nss elapsed_ms=%.1f",
                            (_time.perf_counter() - started) * 1000.0,
                        )

                    def _error(exc):
                        if not win.winfo_exists():
                            return
                        revision_nss_tree.delete(*revision_nss_tree.get_children())
                        revision_nss_status.set(
                            "No fue posible actualizar esta sección. Use Recargar para reintentar."
                        )
                        APP_LOG.error(
                            "CONFIG_SECTION_LOAD_ERROR section=revision_nss elapsed_ms=%.1f type=%s",
                            (_time.perf_counter() - started) * 1000.0,
                            type(exc).__name__,
                        )

                    self._ejecutar_en_segundo_plano(
                        "Cargando revisión NSS…", _cargar,
                        al_terminar=_finalizar, al_error=_error,
                    )

                def revision_nss_seleccionada():
                    seleccion = revision_nss_tree.selection()
                    if not seleccion:
                        messagebox.showwarning(
                            "Revisión NSS","Seleccione un caso.",parent=win
                        )
                        return None
                    return revisiones_nss.get(int(seleccion[0]))

                def resolver_nss(tipo, descripcion):
                    row = revision_nss_seleccionada()
                    if not row:
                        return
                    motivo = simpledialog.askstring(
                        "Motivo administrativo",
                        f"Acción: {descripcion}\n\nExplique el criterio utilizado:",
                        parent=win,
                    )
                    motivo = (motivo or "").strip()
                    if len(motivo) < 8:
                        messagebox.showwarning(
                            "Revisión NSS","El motivo debe tener al menos 8 caracteres.",parent=win
                        )
                        return
                    if not messagebox.askyesno(
                        "Confirmar revisión NSS",
                        f"Caso #{row['id']} · NSS {row['nss_normalizado']}\n\n{descripcion}\n\n¿Continuar?",
                        parent=win,
                    ):
                        return
                    try:
                        actor = self._admin_authorized_actor or self._actor_actual()
                        self.db.resolver_revision_nss(int(row["id"]),tipo,actor,motivo)
                        self.security.audit(
                            "NSS_REVIEW_RESOLVED",actor=actor,success=True,
                            detail=f"revision={row['id']}; resolucion={tipo}",
                        )
                        cargar_revisiones_nss()
                        self._invalidar_caches_datos()
                        self.set_status(f"Revisión NSS #{row['id']} resuelta", "ok")
                    except Exception as exc:
                        APP_LOG.exception("No se pudo resolver la revisión NSS #%s", row["id"])
                        messagebox.showerror("Revisión NSS",str(exc),parent=win)

                def abrir_atencion_revision_nss():
                    row = revision_nss_seleccionada()
                    if not row:
                        return
                    if not row.get("atencion_id"):
                        messagebox.showinfo(
                            "Revisión NSS","Este caso no tiene una atención asociada.",parent=win
                        )
                        return
                    self._abrir_editor_atencion(int(row["atencion_id"]))

                acciones_revision_nss = tb.Frame(tab_revision_nss, style="Card.TFrame")
                acciones_revision_nss.grid(row=2, column=0, sticky="ew", pady=(10, 0))
                tb.Button(
                    acciones_revision_nss,text="Conservar ambas fichas",bootstyle=SECONDARY,
                    command=lambda: resolver_nss("MANTENER_AMBOS","Conservar las dos fichas con el mismo NSS"),
                ).pack(side="left", padx=4)
                tb.Button(
                    acciones_revision_nss,text="Quitar NSS de ficha nueva",bootstyle=WARNING,
                    command=lambda: resolver_nss("DESVINCULAR_NSS","Retirar el NSS de la ficha y atención nuevas"),
                ).pack(side="left", padx=4)
                tb.Button(
                    acciones_revision_nss,text="Fusionar y eliminar duplicada",bootstyle=DANGER,
                    command=lambda: resolver_nss("FUSIONAR_CON_EXISTENTE","Fusionar con la ficha existente y eliminar la ficha duplicada"),
                ).pack(side="left", padx=4)
                tb.Button(
                    acciones_revision_nss,text="Abrir atención",
                    command=abrir_atencion_revision_nss,
                ).pack(side="left", padx=4)
                tb.Button(
                    acciones_revision_nss,text="Actualizar",command=cargar_revisiones_nss
                ).pack(side="right", padx=4)

                # ---------------- TAB PREFERENCIAS ----------------
                # ---------------- TAB RESPALDOS ----------------
                tab_backups = tb.Frame(notebook, padding=12, style="Card.TFrame")
                notebook.add(tab_backups, text="Respaldos")
                tab_backups.columnconfigure(0, weight=1)
                tab_backups.rowconfigure(1, weight=1)
                backup_status = tk.StringVar(
                    value="Respaldos verificados; se eliminan automáticamente después de 4 días"
                )
                tb.Label(tab_backups, textvariable=backup_status, style="Muted.TLabel").grid(
                    row=0, column=0, sticky="w", pady=(0, 8)
                )
                backup_tree = ttk.Treeview(
                    tab_backups,
                    columns=("fecha", "motivo", "estado"),
                    show="headings",
                    height=15,
                    style="Modern.Treeview",
                )
                backup_tree.heading("fecha", text="Fecha")
                backup_tree.heading("motivo", text="Motivo")
                backup_tree.heading("estado", text="Verificación")
                backup_tree.column("fecha", width=175, anchor="w")
                backup_tree.column("motivo", width=360, anchor="w")
                backup_tree.column("estado", width=130, anchor="center")
                backup_tree.grid(row=1, column=0, sticky="nsew")


                def refrescar_respaldos():
                    for item in backup_tree.get_children():
                        backup_tree.delete(item)
                    backup_status.set("Cargando respaldos en segundo plano…")
                    started = _time.perf_counter()
                    APP_LOG.info("CONFIG_SECTION_LOAD_START section=backups")

                    def _cargar():
                        result = []
                        for folder in self.db.backup_manager.list_backups():
                            try:
                                manifest = self.db.backup_manager.verify(folder)
                                created = str(manifest.get("created_at", "")).replace("T", " ")
                                reason = str(manifest.get("reason", ""))
                                status = "Válido"
                            except Exception as exc:
                                created = folder.name[:15]
                                reason = str(exc)
                                status = "Inválido"
                            result.append((str(folder), created, reason, status))
                        return result

                    def _finalizar(rows):
                        if not win.winfo_exists():
                            return
                        backup_tree.delete(*backup_tree.get_children())
                        for folder, created, reason, status in rows:
                            backup_tree.insert("", "end", iid=folder, values=(created, reason, status))
                        backup_status.set(f"{len(rows)} respaldo(s) disponible(s)")
                        APP_LOG.info(
                            "CONFIG_SECTION_LOAD_DONE section=backups elapsed_ms=%.1f",
                            (_time.perf_counter() - started) * 1000.0,
                        )

                    def _error(exc):
                        if not win.winfo_exists():
                            return
                        backup_status.set(
                            "No fue posible actualizar esta sección. Las demás opciones continúan disponibles."
                        )
                        APP_LOG.error(
                            "CONFIG_SECTION_LOAD_ERROR section=backups elapsed_ms=%.1f type=%s",
                            (_time.perf_counter() - started) * 1000.0,
                            type(exc).__name__,
                        )

                    self._ejecutar_en_segundo_plano(
                        "Cargando respaldos…", _cargar,
                        al_terminar=_finalizar, al_error=_error,
                    )

                def respaldo_seleccionado():
                    selection = backup_tree.selection()
                    return selection[0] if selection else ""

                def crear_respaldo_manual():
                    try:
                        folder = self.db.backup_manager.create(
                            "respaldo_manual", label=f"actor={self._actor_actual()}"
                        )
                        backup_status.set(f"Respaldo creado y verificado: {os.path.basename(folder)}")
                        refrescar_respaldos()
                    except Exception as exc:
                        APP_LOG.exception("Falló el respaldo manual")
                        messagebox.showerror("Respaldos", str(exc), parent=win)

                def verificar_respaldo_ui():
                    folder = respaldo_seleccionado()
                    if not folder:
                        backup_status.set("Seleccione un respaldo para verificar")
                        return
                    try:
                        self.db.backup_manager.verify(folder)
                        backup_status.set("El respaldo seleccionado es íntegro y restaurable")
                    except Exception as exc:
                        backup_status.set(f"Respaldo inválido: {exc}")

                def restaurar_respaldo_ui():
                    folder = respaldo_seleccionado()
                    if not folder:
                        backup_status.set("Seleccione un respaldo para restaurar")
                        return
                    actor = self._solicitar_autorizacion_admin(
                        "RESTAURAR_BASE_DE_DATOS", parent=win, force=True
                    )
                    if not actor:
                        return
                    if not messagebox.askyesno(
                        "Restaurar base de datos",
                        "La base actual se respaldará antes de restaurar. La aplicación se reiniciará.\n\n"
                        "ADVERTENCIA: un respaldo anterior puede reintroducir pacientes de prueba "
                        "que fueron purgados después de su creación.\n\n"
                        "¿Desea continuar?",
                        icon="warning",
                        parent=win,
                    ):
                        return
                    confirmacion = simpledialog.askstring(
                        "Confirmación de restauración",
                        "Escriba RESTAURAR para confirmar que comprende el riesgo:",
                        parent=win,
                    )
                    if (confirmacion or "").strip().upper() != "RESTAURAR":
                        messagebox.showinfo("Restauración cancelada", "No se modificó la base.", parent=win)
                        return
                    try:
                        self.db.backup_manager.restore_database(folder)
                        self.security.audit(
                            "DATABASE_RESTORED", actor=actor, success=True, detail=os.path.basename(folder)
                        )
                        messagebox.showinfo(
                            "Restauración completada",
                            "La base fue restaurada y verificada. La aplicación se reiniciará.",
                            parent=win,
                        )
                        command = [sys.executable] + ([] if getattr(sys, "frozen", False) else [os.path.abspath(__file__)])
                        subprocess.Popen(command, cwd=os.path.dirname(sys.executable) if getattr(sys, "frozen", False) else BASE_DIR)
                        self.root.destroy()
                    except Exception as exc:
                        APP_LOG.exception("Falló la restauración de la base")
                        self.security.audit("DATABASE_RESTORED", actor=actor, success=False, detail=str(exc))
                        messagebox.showerror("Restauración", str(exc), parent=win)

                backup_actions = tb.Frame(tab_backups, style="Card.TFrame")
                backup_actions.grid(row=2, column=0, sticky="ew", pady=(10, 0))
                tb.Button(backup_actions, text="Crear respaldo", command=crear_respaldo_manual, bootstyle=SUCCESS).pack(side="left", padx=4)
                tb.Button(backup_actions, text="Verificar", command=verificar_respaldo_ui, bootstyle=INFO).pack(side="left", padx=4)
                tb.Button(backup_actions, text="Restaurar", command=restaurar_respaldo_ui, bootstyle=DANGER).pack(side="left", padx=4)

                # ---------------- TAB USUARIOS ----------------
                tab_usuarios = tb.Frame(notebook, padding=12, style="Card.TFrame")
                notebook.add(tab_usuarios, text="Representante del turno")
                tab_usuarios.columnconfigure(0, weight=3)
                tab_usuarios.columnconfigure(1, weight=2)
                tab_usuarios.rowconfigure(1, weight=1)

                usuarios_estado = tk.StringVar(
                    value="Cargando usuarios activos del sistema…"
                )
                tb.Label(
                    tab_usuarios,
                    textvariable=usuarios_estado,
                    style="Muted.TLabel",
                    background="#0E1B2B",
                    wraplength=1040,
                    justify="left",
                ).grid(row=0, column=0, columnspan=2, sticky="ew", pady=(0, 10))

                catalogo_card = tb.Frame(
                    tab_usuarios, padding=14, style="Root.TFrame"
                )
                catalogo_card.grid(
                    row=1, column=0, sticky="nsew", padx=(0, 8)
                )
                catalogo_card.columnconfigure(0, weight=1)
                catalogo_card.rowconfigure(2, weight=1)
                tb.Label(
                    catalogo_card,
                    text="REPRESENTANTES DE FACTURACIÓN",
                    font=("Arial", 13, "bold"),
                    foreground="#FFFFFF",
                    background="#07111f",
                ).grid(row=0, column=0, sticky="w")
                tb.Label(
                    catalogo_card,
                    text=(
                        "Selecciona un usuario habilitado en la aplicación principal. "
                        "Esta pantalla no crea, edita ni elimina usuarios."
                    ),
                    style="Muted.TLabel",
                    background="#07111f",
                    wraplength=590,
                    justify="left",
                ).grid(row=1, column=0, sticky="ew", pady=(3, 10))

                usuarios_tree = ttk.Treeview(
                    catalogo_card,
                    columns=("nombre", "usuario", "rol"),
                    show="headings",
                    height=12,
                    style="Modern.Treeview",
                )
                usuarios_tree.heading("nombre", text="Nombre del representante")
                usuarios_tree.heading("usuario", text="Usuario")
                usuarios_tree.heading("rol", text="Rol")
                usuarios_tree.column("nombre", width=310, anchor="w")
                usuarios_tree.column("usuario", width=150, anchor="w")
                usuarios_tree.column("rol", width=165, anchor="w")
                usuarios_tree.grid(row=2, column=0, sticky="nsew")

                seleccion_var = tk.StringVar(
                    value="Selecciona el representante correcto para habilitar la corrección."
                )
                tb.Label(
                    catalogo_card,
                    textvariable=seleccion_var,
                    style="Muted.TLabel",
                    background="#07111f",
                    wraplength=620,
                    justify="left",
                ).grid(row=3, column=0, sticky="ew", pady=(10, 0))
                rep_retry_button = tb.Button(
                    catalogo_card,
                    text="Reintentar",
                    bootstyle=INFO,
                    state="disabled",
                )
                rep_retry_button.grid(row=4, column=0, sticky="w", pady=(8, 0))

                turno_host = tb.Frame(tab_usuarios, style="Root.TFrame")
                turno_host.grid(row=1, column=1, sticky="nsew", padx=(8, 0))
                turno_host.columnconfigure(0, weight=1)
                turno_host.rowconfigure(0, weight=1)
                turno_canvas = tk.Canvas(
                    turno_host,
                    borderwidth=0,
                    highlightthickness=0,
                    background="#07111f",
                )
                turno_scroll = ttk.Scrollbar(
                    turno_host, orient="vertical", command=turno_canvas.yview
                )
                turno_canvas.configure(yscrollcommand=turno_scroll.set)
                turno_canvas.grid(row=0, column=0, sticky="nsew")
                turno_scroll.grid(row=0, column=1, sticky="ns")
                turno_card = tb.Frame(turno_canvas, padding=16, style="Root.TFrame")
                turno_window = turno_canvas.create_window(
                    (0, 0), window=turno_card, anchor="nw"
                )
                turno_card.bind(
                    "<Configure>",
                    lambda _event: turno_canvas.configure(
                        scrollregion=turno_canvas.bbox("all")
                    ),
                    add="+",
                )
                turno_canvas.bind(
                    "<Configure>",
                    lambda event: turno_canvas.itemconfigure(
                        turno_window, width=max(1, event.width)
                    ),
                    add="+",
                )
                turno_card.columnconfigure(0, weight=1)
                tb.Label(
                    turno_card,
                    text="TURNO ACTUAL",
                    font=("Arial", 13, "bold"),
                    foreground="#FFFFFF",
                    background="#07111f",
                ).grid(row=0, column=0, sticky="w")
                tb.Label(
                    turno_card,
                    text="● ACTIVO",
                    font=("Arial", 9, "bold"),
                    foreground="#72E39B",
                    background="#07111f",
                ).grid(row=0, column=1, sticky="e")

                tb.Label(
                    turno_card,
                    text="Representante registrado",
                    style="Muted.TLabel",
                    background="#07111f",
                ).grid(row=1, column=0, columnspan=2, sticky="w", pady=(18, 3))
                usuario_turno_nombre_var = tk.StringVar(value="No configurado")
                tb.Label(
                    turno_card,
                    textvariable=usuario_turno_nombre_var,
                    font=("Arial", 16, "bold"),
                    foreground="#5CB6FF",
                    background="#07111f",
                    wraplength=390,
                    justify="left",
                ).grid(row=2, column=0, columnspan=2, sticky="w")

                usuario_turno_detalle_var = tk.StringVar(value="")
                tb.Label(
                    turno_card,
                    textvariable=usuario_turno_detalle_var,
                    background="#07111f",
                    foreground="#EAF2FF",
                    wraplength=390,
                    justify="left",
                ).grid(row=3, column=0, columnspan=2, sticky="w", pady=(8, 18))

                primary_card = tb.Frame(turno_card, padding=12, style="Card.TFrame")
                primary_card.grid(
                    row=4, column=0, columnspan=2, sticky="ew", pady=(0, 14)
                )
                primary_card.columnconfigure(0, weight=1)
                tb.Label(
                    primary_card,
                    text="CONTROL DE SESIÓN PRINCIPAL",
                    font=("Arial", 10, "bold"),
                    foreground="#5CB6FF",
                    background="#0E1B2B",
                ).grid(row=0, column=0, sticky="w")
                primary_status_var = tk.StringVar(value="Cargando estado operacional…")
                tb.Label(
                    primary_card,
                    textvariable=primary_status_var,
                    style="Muted.TLabel",
                    background="#0E1B2B",
                    wraplength=360,
                    justify="left",
                ).grid(row=1, column=0, sticky="ew", pady=(6, 10))
                self._primary_transfer_config_button = tb.Button(
                    primary_card,
                    text="HACER ESTA COMPUTADORA PRINCIPAL",
                    bootstyle=WARNING,
                    command=lambda: self.request_transfer_admission_primary(parent=win),
                    state="disabled",
                )
                self._primary_transfer_config_button.grid(
                    row=2, column=0, sticky="ew", ipady=6
                )
                primary_hint_var = tk.StringVar(value="")
                tb.Label(
                    primary_card,
                    textvariable=primary_hint_var,
                    style="Muted.TLabel",
                    background="#0E1B2B",
                    wraplength=360,
                    justify="left",
                ).grid(row=3, column=0, sticky="ew", pady=(5, 0))

                def refrescar_control_primary():
                    runtime = getattr(self.db, "_runtime", None)
                    snapshot = dict(runtime.state() or {}) if runtime is not None else {}
                    representante_actual = str(
                        snapshot.get("active_user_display_name")
                        or snapshot.get("active_username")
                        or "No configurado"
                    )
                    usuario_turno_nombre_var.set(representante_actual)
                    if snapshot.get("turn_id") is not None:
                        usuario_turno_detalle_var.set(
                            f"Turno central: {snapshot.get('turn_id')}"
                            f" · {snapshot.get('turn_code') or 'Sin código'}\n"
                            f"Inicio: {snapshot.get('turn_started_at') or 'No configurado'}\n"
                            f"Fin: {snapshot.get('turn_ends_at') or 'No configurado'}"
                        )
                    else:
                        usuario_turno_detalle_var.set(
                            "No existe un turno central configurado."
                        )
                    role = str(snapshot.get("role") or "NONE").upper()
                    role_visible = "PRINCIPAL" if role == "PRIMARY" else "SECUNDARIA"
                    local_device = str(
                        getattr(runtime, "device_id", "") or "No disponible"
                    )
                    primary_device = str(
                        snapshot.get("primary_device_id") or "Sin asignar"
                    )
                    primary_status_var.set(
                        f"Estado de esta estación: {role_visible}\n"
                        f"Estación principal actual: {primary_device}\n"
                        f"Esta estación: {local_device}"
                    )
                    enabled = bool(
                        normalize_role(self.session_context.role) == ROLE_ADMIN
                        and runtime is not None
                        and not bool(getattr(runtime, "offline", False))
                        and role != "PRIMARY"
                        and not self._primary_transfer_in_progress
                    )
                    self._primary_transfer_config_button.configure(
                        state="normal" if enabled else "disabled"
                    )
                    hint = (
                        "Esta computadora ya posee la sesión principal."
                        if role == "PRIMARY"
                        else "Solo cambia qué estación posee PRIMARY; no modifica el turno ni el representante."
                    )
                    primary_hint_var.set(hint)
                    try:
                        self._primary_transfer_config_button.setToolTip(hint)
                    except Exception:
                        pass

                self._refresh_primary_config_panel = refrescar_control_primary

                proteccion_card = tb.Frame(turno_card, padding=12, style="Card.TFrame")
                proteccion_card.grid(
                    row=5, column=0, columnspan=2, sticky="ew", pady=(0, 14)
                )
                tb.Label(
                    proteccion_card,
                    text="Cambio controlado y auditable",
                    font=("Arial", 10, "bold"),
                    foreground="#72E39B",
                    background="#0E1B2B",
                ).pack(anchor="w")
                tb.Label(
                    proteccion_card,
                    text=(
                        "Corrige únicamente el representante operacional. No reinicia "
                        "el turno, no mueve PRIMARY y no modifica pacientes ni atenciones. "
                        "El Administrador selecciona el usuario destino y confirma con "
                        "su propia contraseña."
                    ),
                    style="Muted.TLabel",
                    background="#0E1B2B",
                    wraplength=360,
                    justify="left",
                ).pack(anchor="w", pady=(4, 0))

                def usuario_actual():
                    runtime = getattr(self.db, "_runtime", None)
                    snapshot = dict(runtime.state() or {}) if runtime is not None else {}
                    return {
                        "user_id": str(snapshot.get("active_user_id") or ""),
                        "username": str(snapshot.get("active_username") or ""),
                        "display_name": str(
                            snapshot.get("active_user_display_name")
                            or snapshot.get("active_username")
                            or ""
                        ),
                        "turn_id": snapshot.get("turn_id"),
                        "turn_started_at": snapshot.get("turn_started_at"),
                        "turn_ends_at": snapshot.get("turn_ends_at"),
                    }

                representantes_por_item = {}

                corregir_turno_btn = tb.Button(
                    turno_card,
                    text="Confirmar y corregir representante",
                    bootstyle=PRIMARY,
                    width=30,
                    state="disabled",
                )
                corregir_turno_btn.grid(
                    row=6, column=0, columnspan=2, sticky="ew", ipady=7
                )

                def representante_seleccionado():
                    seleccion = usuarios_tree.selection()
                    if not seleccion:
                        return None
                    return representantes_por_item.get(seleccion[0])


                def actualizar_seleccion(_event=None):
                    representante = representante_seleccionado()
                    actual = usuario_actual()
                    es_admin = normalize_role(self.session_context.role) == ROLE_ADMIN
                    if representante is None:
                        seleccion_var.set(
                            "Selecciona el representante correcto para habilitar la corrección."
                        )
                        corregir_turno_btn.configure(state="disabled")
                        return
                    seleccion_var.set(
                        f"Seleccionado: {representante.full_name} · usuario {representante.username}. "
                        "Puede asignarse aunque no tenga una sesión abierta."
                    )
                    representative_user_id = str(
                        getattr(representante, "user_id", "") or ""
                    )
                    current_user_id = str(actual.get("user_id") or "")
                    mismo = bool(
                        representative_user_id
                        and current_user_id
                        and representative_user_id == current_user_id
                    ) or (
                        not representative_user_id
                        and str(representante.username or "").casefold()
                        == str(actual.get("username") or "").casefold()
                    )
                    if not es_admin:
                        seleccion_var.set(
                            "Solo un Administrador puede corregir el representante del turno."
                        )
                    elif mismo:
                        seleccion_var.set(
                            f"{representante.full_name} ya es el representante operativo actual."
                        )
                    corregir_turno_btn.configure(
                        state="normal" if es_admin and not mismo else "disabled"
                    )


                # Cada petición tiene un final único: SUCCESS, ERROR o TIMEOUT.
                # Un proveedor antiguo que permanezca bloqueado no puede volver a
                # modificar la pestaña después de que su petición expiró.
                rep_load_state = {
                    "request_id": 0,
                    "timeout_id": None,
                    "finished": False,
                    "in_flight": False,
                }

                def _cancelar_timeout_representantes():
                    timeout_id = rep_load_state.get("timeout_id")
                    rep_load_state["timeout_id"] = None
                    if timeout_id is None:
                        return
                    try:
                        win.after_cancel(timeout_id)
                    except Exception:
                        pass

                def _mostrar_representantes(representantes):
                    usuarios_tree.delete(*usuarios_tree.get_children())
                    representantes_por_item.clear()
                    for indice, representante in enumerate(representantes):
                        item_id = f"usuario-{indice}"
                        representantes_por_item[item_id] = representante
                        rol_visible = representante.role.replace("_", " ").title()
                        usuarios_tree.insert(
                            "", "end", iid=item_id,
                            values=(
                                representante.full_name,
                                representante.username,
                                rol_visible,
                            ),
                        )
                    actualizar_seleccion()

                def cargar_usuarios():
                    if rep_load_state["in_flight"]:
                        APP_LOG.info(
                            "CONFIG_USERS_LOAD_IGNORED_ACTIVE thread=%s",
                            threading.current_thread().name,
                        )
                        return
                    rep_load_state["request_id"] += 1
                    request_id = rep_load_state["request_id"]
                    rep_load_state["finished"] = False
                    rep_load_state["in_flight"] = True
                    _cancelar_timeout_representantes()
                    rep_retry_button.configure(state="disabled")
                    actual = usuario_actual()
                    usuario_turno_nombre_var.set(
                        actual.get("display_name") or "No configurado"
                    )
                    if actual.get("turn_id") is not None:
                        usuario_turno_detalle_var.set(
                            f"Turno central: {actual.get('turn_id')}\n"
                            f"Inicio: {actual.get('turn_started_at') or 'No configurado'}\n"
                            f"Fin: {actual.get('turn_ends_at') or 'No configurado'}"
                        )
                    else:
                        usuario_turno_detalle_var.set(
                            "No existe un turno central configurado."
                        )
                    started = _time.perf_counter()
                    APP_LOG.info(
                        "CONFIG_USERS_LOAD_START elapsed_ms=0.0 thread=%s",
                        threading.current_thread().name,
                    )
                    cache_provider = getattr(
                        self.main_app_gateway, "cached_representatives", None
                    )
                    cached = list(cache_provider() or ()) if callable(cache_provider) else []
                    if cached:
                        APP_LOG.info(
                            "CONFIG_USERS_CACHE_USED count=%s thread=%s",
                            len(cached), threading.current_thread().name,
                        )
                        _mostrar_representantes(cached)
                        usuarios_estado.set(
                            f"{len(cached)} usuario(s) activos en caché. Actualizando…"
                        )
                    else:
                        usuarios_tree.delete(*usuarios_tree.get_children())
                        representantes_por_item.clear()
                        actualizar_seleccion()
                        usuarios_estado.set("Cargando usuarios activos del sistema…")

                    def _cargar_representantes():
                        APP_LOG.info(
                            "CONFIG_USERS_QUERY_START elapsed_ms=%.1f thread=%s",
                            (_time.perf_counter() - started) * 1000.0,
                            threading.current_thread().name,
                        )
                        if os.environ.get("HOSPITAL_OFFLINE", "").strip().casefold() in {
                            "1", "true", "yes", "si", "sí", "on",
                        }:
                            if cached:
                                return cached
                            raise RuntimeError("Sin conexión para actualizar los usuarios del sistema.")
                        representatives = list(self.main_app_gateway.list_representatives())
                        APP_LOG.info(
                            "CONFIG_USERS_QUERY_DONE elapsed_ms=%.1f count=%s thread=%s",
                            (_time.perf_counter() - started) * 1000.0,
                            len(representatives),
                            threading.current_thread().name,
                        )
                        return representatives

                    def _finalizar(representantes):
                        if request_id != rep_load_state["request_id"] or not win.winfo_exists():
                            return
                        if rep_load_state["finished"]:
                            rep_load_state["in_flight"] = False
                            rep_retry_button.configure(state="normal")
                            return
                        rep_load_state["finished"] = True
                        rep_load_state["in_flight"] = False
                        _cancelar_timeout_representantes()
                        _mostrar_representantes(representantes)
                        usuarios_estado.set(
                            f"{len(representantes)} usuario(s) activo(s) del sistema. "
                            "El Administrador puede seleccionar cualquiera sin iniciar sesión como ese usuario."
                        )
                        actualizar_seleccion()
                        APP_LOG.info(
                            "CONFIG_USERS_LOAD_DONE elapsed_ms=%.1f count=%s thread=%s",
                            (_time.perf_counter() - started) * 1000.0,
                            len(representantes),
                            threading.current_thread().name,
                        )

                    def _error(exc):
                        if request_id != rep_load_state["request_id"] or not win.winfo_exists():
                            return
                        if rep_load_state["finished"]:
                            rep_load_state["in_flight"] = False
                            rep_retry_button.configure(state="normal")
                            return
                        rep_load_state["finished"] = True
                        rep_load_state["in_flight"] = False
                        _cancelar_timeout_representantes()
                        APP_LOG.error(
                            "CONFIG_USERS_LOAD_ERROR elapsed_ms=%.1f type=%s thread=%s",
                            (_time.perf_counter() - started) * 1000.0,
                            type(exc).__name__, threading.current_thread().name,
                            exc_info=(type(exc), exc, exc.__traceback__),
                        )
                        usuarios_estado.set(
                            "No fue posible cargar los usuarios del sistema."
                        )
                        rep_retry_button.configure(state="normal")

                    def _timeout_representantes():
                        if (
                            request_id != rep_load_state["request_id"]
                            or rep_load_state["finished"]
                            or not win.winfo_exists()
                        ):
                            return
                        rep_load_state["finished"] = True
                        rep_load_state["timeout_id"] = None
                        APP_LOG.warning(
                            "CONFIG_USERS_LOAD_TIMEOUT elapsed_ms=%.1f thread=%s",
                            (_time.perf_counter() - started) * 1000.0,
                            threading.current_thread().name,
                        )
                        usuarios_estado.set(
                            "No fue posible cargar los usuarios del sistema. "
                            "La consulta anterior aún se está cerrando."
                        )
                        # La consulta no se mata desde Tk. Mantener single-flight
                        # evita que un clic repetido acumule workers PostgreSQL.
                        rep_retry_button.configure(state="disabled")

                    rep_load_state["timeout_id"] = win.after(
                        5000, _timeout_representantes
                    )

                    self._ejecutar_en_segundo_plano(
                        "Cargando usuarios del sistema…",
                        _cargar_representantes,
                        _finalizar,
                        _error,
                    )

                rep_retry_button.configure(command=cargar_usuarios)

                def abrir_dialogo_confirmacion(representante):
                    dialogo = Toplevel(win)
                    dialogo.title("Confirmar cambio de representante")
                    dialogo.geometry("720x540")
                    dialogo.minsize(680, 520)
                    dialogo.transient(win)
                    dialogo.grab_set()
                    dialogo.configure(bg="#07111f")
                    self._bind_esc_cerrar(dialogo)

                    contenido = tb.Frame(dialogo, padding=18, style="Root.TFrame")
                    contenido.pack(fill="both", expand=True)
                    tb.Label(
                        contenido,
                        text="Confirmar cambio de representante",
                        font=("Arial", 16, "bold"),
                        foreground="#FFFFFF",
                        background="#07111f",
                    ).pack(anchor="w")
                    tb.Label(
                        contenido,
                        text=(
                            "Selecciona un Administrador habilitado e introduce su "
                            "contraseña para autorizar este cambio. El Administrador "
                            "no necesita tener una sesión abierta."
                        ),
                        style="Muted.TLabel",
                        background="#07111f",
                        wraplength=510,
                        justify="left",
                    ).pack(anchor="w", pady=(4, 16))

                    destino_card = tb.Frame(contenido, padding=12, style="Card.TFrame")
                    destino_card.pack(fill="x", pady=(0, 12))
                    tb.Label(
                        destino_card,
                        text="Representante que quedará asignado",
                        background="#0E1B2B",
                        font=("Arial", 10, "bold"),
                    ).pack(anchor="w", pady=(0, 5))
                    tb.Label(
                        destino_card,
                        text=f"{representante.full_name}  ·  {representante.username}",
                        background="#0E1B2B",
                        foreground="#5CB6FF",
                        font=("Arial", 12, "bold"),
                    ).pack(anchor="w")

                    credenciales = tb.Frame(contenido, padding=12, style="Card.TFrame")
                    credenciales.pack(fill="x")
                    tb.Label(
                        credenciales,
                        text="Administrador que autoriza",
                        background="#0E1B2B",
                        font=("Arial", 10, "bold"),
                    ).pack(anchor="w")
                    admin_cache_provider = getattr(
                        self.main_app_gateway, "cached_active_administrators", None
                    )
                    administradores = list(
                        admin_cache_provider() or ()
                    ) if callable(admin_cache_provider) else []
                    admin_labels = [
                        f"{admin.full_name} · Administrador"
                        for admin in administradores
                    ]
                    admin_var = tk.StringVar(value="")
                    admin_combo = tb.Combobox(
                        credenciales,
                        textvariable=admin_var,
                        values=admin_labels,
                        state="readonly",
                    )
                    admin_combo.pack(fill="x", ipady=5, pady=(4, 5))
                    current_username = str(
                        getattr(self.session_context, "username", "") or ""
                    ).strip().casefold()
                    default_index = next(
                        (
                            index for index, admin in enumerate(administradores)
                            if str(admin.username or "").strip().casefold()
                            == current_username
                        ),
                        0 if administradores else -1,
                    )
                    if default_index >= 0:
                        admin_combo.current(default_index)
                    admin_status_var = tk.StringVar(
                        value=(
                            "El Administrador seleccionado no necesita una sesión abierta."
                            if administradores
                            else "No hay cuentas de Administrador habilitadas para autorizar esta operación."
                        )
                    )
                    tb.Label(
                        credenciales,
                        textvariable=admin_status_var,
                        background="#0E1B2B",
                        style="Muted.TLabel",
                        wraplength=610,
                        justify="left",
                    ).pack(anchor="w", pady=(0, 10))
                    tb.Label(
                        credenciales,
                        text="Contraseña del administrador",
                        background="#0E1B2B",
                        font=("Arial", 10, "bold"),
                    ).pack(anchor="w")
                    clave_var = tk.StringVar(value="")
                    clave_entry = tb.Entry(credenciales, textvariable=clave_var, show="●")
                    clave_entry.pack(fill="x", ipady=5, pady=(4, 0))


                    def guardar_dialogo():
                        selected_index = admin_combo.current()
                        selected_admin = (
                            administradores[selected_index]
                            if 0 <= selected_index < len(administradores)
                            else None
                        )
                        password_holder = {"value": clave_var.get()}
                        clave_var.set("")
                        if selected_admin is None or not password_holder["value"]:
                            password_holder["value"] = ""
                            messagebox.showwarning(
                                "Confirmar identidad",
                                (
                                    "Selecciona un Administrador habilitado e introduce su contraseña."
                                    if administradores
                                    else "No hay cuentas de Administrador habilitadas para autorizar esta operación."
                                ),
                                parent=dialogo,
                            )
                            clave_entry.focus_set()
                            return

                        try:
                            aplicar_btn.configure(state="disabled")
                        except Exception:
                            pass

                        def _aplicar_cambio():
                            try:
                                autorizado_por, confirmado = self.main_app_gateway.authorize_admin_action(
                                    selected_admin_user_id=selected_admin.user_id,
                                    selected_admin_username=selected_admin.username,
                                    password=password_holder.pop("value", ""),
                                    action="CORRECT_ADMISSION_REPRESENTATIVE",
                                    target_user_id=representante.user_id,
                                    target_username=representante.username,
                                )
                            finally:
                                password_holder["value"] = ""
                            runtime = getattr(self.db, "_runtime", None)
                            if runtime is None:
                                raise RuntimeError(
                                    "La integración central de Admisión no está disponible."
                                )
                            changed = self.db.admin_correct_current_turn_representative(
                                confirmado,
                                authorizing_admin=autorizado_por,
                            )
                            return autorizado_por, confirmado, changed

                        def _finalizar(resultado):
                            clave_var.set("")
                            if not win.winfo_exists() or not dialogo.winfo_exists():
                                return
                            autorizado_por, confirmado, changed = resultado
                            self.security.audit(
                                "TURN_REPRESENTATIVE_ADMIN_CORRECTED",
                                actor=autorizado_por.username,
                                success=True,
                                detail=(
                                    f"requesting_user_id={getattr(self.context, 'user_id', '')}; "
                                    f"authorizing_admin_user_id={autorizado_por.user_id}; "
                                    f"target_representative_user_id={confirmado.user_id}; "
                                    f"operational_revision={getattr(changed, 'operational_revision', '')}"
                                ),
                            )

                            # El runtime ya posee el snapshot devuelto por el
                            # commit central. Aplicarlo evita que la UI mezcle
                            # este representante con un turno/local JSON viejo.
                            self.apply_operational_snapshot(
                                self._snapshot_operacional_integrado()
                            )
                            self._invalidar_caches_datos()
                            cargar_usuarios()
                            refrescar_control_primary()
                            try:
                                dialogo.grab_release()
                            except Exception:
                                pass
                            dialogo.destroy()
                            messagebox.showinfo(
                                "Representante actualizado",
                                "El representante operativo fue corregido centralmente. "
                                "El usuario seleccionado no necesita haber iniciado sesión. "
                                "No se cambió el turno ni la computadora PRIMARY.",
                                parent=win,
                            )

                        def _error(exc):
                            clave_var.set("")
                            try:
                                aplicar_btn.configure(state="normal")
                            except Exception:
                                pass
                            APP_LOG.error(
                                "REPRESENTATIVE_ADMIN_CORRECTION_ERROR type=%s message=%s",
                                type(exc).__name__, str(exc),
                                exc_info=(type(exc), exc, exc.__traceback__),
                            )
                            messagebox.showwarning(
                                "No se autorizó el cambio",
                                str(exc) or "No fue posible aplicar la corrección.",
                                parent=dialogo,
                            )
                            clave_entry.focus_set()

                        self._ejecutar_en_segundo_plano(
                            "Aplicando representante…",
                            _aplicar_cambio,
                            al_terminar=_finalizar,
                            al_error=_error,
                        )

                    botones = tb.Frame(contenido, style="Root.TFrame")
                    botones.pack(side="bottom", fill="x", pady=(16, 0))
                    aplicar_btn = tb.Button(
                        botones,
                        text="Validar y aplicar cambio",
                        bootstyle=SUCCESS,
                        command=guardar_dialogo,
                        width=20,
                    )
                    aplicar_btn.pack(side="left")
                    if not administradores:
                        aplicar_btn.configure(state="disabled")
                    tb.Button(
                        botones,
                        text="Cancelar",
                        bootstyle=SECONDARY,
                        command=dialogo.destroy,
                        width=12,
                    ).pack(side="right")
                    clave_entry.bind("<Return>", lambda _event: guardar_dialogo())
                    dialogo.after(80, clave_entry.focus_set)

                def corregir_turno_actual():
                    representante = representante_seleccionado()
                    if representante is None:
                        messagebox.showwarning(
                            "Corregir representante",
                            "Selecciona primero un representante de la lista.",
                            parent=win,
                        )
                        return
                    abrir_dialogo_confirmacion(representante)

                corregir_turno_btn.configure(command=corregir_turno_actual)
                usuarios_tree.bind("<<TreeviewSelect>>", actualizar_seleccion)
                usuarios_tree.bind("<Double-1>", lambda _event: corregir_turno_actual())

                # ---------------- TAB PREFERENCIAS ----------------
                tab_pref = tb.Frame(notebook, padding=12, style="Card.TFrame")
                notebook.add(tab_pref, text="Preferencias")
                pref = dict(self.app_settings)

                pref_nb = ttk.Notebook(tab_pref)
                pref_nb.pack(fill="both", expand=True)

                def add_labeled_combo(parent, row, label, var, values, width=24):
                    tb.Label(parent, text=label).grid(row=row, column=0, sticky="w", padx=6, pady=6)
                    cb = tb.Combobox(parent, textvariable=var, state="readonly", values=values, width=width)
                    cb.grid(row=row, column=1, sticky="w", padx=6, pady=6)
                    return cb

                def add_labeled_entry(parent, row, label, var, width=10):
                    tb.Label(parent, text=label).grid(row=row, column=0, sticky="w", padx=6, pady=6)
                    ent = tb.Entry(parent, textvariable=var, width=width)
                    ent.grid(row=row, column=1, sticky="w", padx=6, pady=6)
                    return ent

                font_var = tk.StringVar(value=str(pref.get("font_size", 11)))
                theme_var = tk.StringVar(value=pref.get("theme", "oscuro"))
                contrast_var = tk.BooleanVar(value=bool(pref.get("high_contrast", False)))
                autosize_var = tk.StringVar(value=pref.get("window_size", "1280x740"))
                auto_print_var = tk.BooleanVar(value=bool(pref.get("auto_print", True)))

                print_auto_hoja_var = tk.BooleanVar(value=bool(pref.get("print_auto_hoja", True)))
                print_auto_reporte_var = tk.BooleanVar(value=bool(pref.get("print_auto_reporte_turno", True)))
                print_auto_excel_var = tk.BooleanVar(value=bool(pref.get("print_auto_excel_turno", True)))
                copies_hoja_var = tk.StringVar(value=str(pref.get("print_copies_hoja", 1)))
                copies_reporte_var = tk.StringVar(value=str(pref.get("print_copies_reporte", 2)))
                copies_excel_var = tk.StringVar(value=str(pref.get("print_copies_excel", 2)))
                pdf_orientation_var = tk.StringVar(value=pref.get("print_pdf_orientation", "Horizontal"))
                excel_orientation_var = tk.StringVar(value=pref.get("print_excel_orientation", "Horizontal"))
                print_behavior_var = tk.StringVar(value=pref.get("print_behavior_hoja", "Imprimir y abrir PDF"))

                confirm_var = tk.BooleanVar(value=bool(pref.get("validation_confirm_before_generate", True)))
                warn_nss_var = tk.BooleanVar(value=bool(pref.get("validation_warn_nss_incomplete", True)))
                warn_ars_var = tk.BooleanVar(value=bool(pref.get("validation_warn_ars_sin_seguro", True)))
                block_short_ars_var = tk.BooleanVar(value=bool(pref.get("validation_block_short_ars", True)))
                allow_cedula_var = tk.BooleanVar(value=bool(pref.get("validation_allow_missing_cedula", True)))
                allow_phone_var = tk.BooleanVar(value=bool(pref.get("validation_allow_missing_phone", False)))
                warn_dup_var = tk.BooleanVar(value=bool(pref.get("validation_warn_duplicate_turno", True)))

                rn_strip_var = tk.BooleanVar(value=bool(pref.get("rn_strip_db", True)))
                rn_show_pdf_var = tk.BooleanVar(value=bool(pref.get("rn_show_pdf", True)))
                rn_warn_var = tk.BooleanVar(value=bool(pref.get("rn_warn", True)))
                rn_format_var = tk.StringVar(value=pref.get("rn_format_display", "RN- NOMBRE DE LA MADRE"))

                button_size_var = tk.StringVar(value=pref.get("button_size", "Normal"))
                table_row_height_var = tk.StringVar(value=str(pref.get("table_row_height", 29)))
                compact_mode_var = tk.BooleanVar(value=bool(pref.get("compact_mode", False)))
                small_screen_var = tk.BooleanVar(value=bool(pref.get("small_screen_mode", False)))
                show_side_panel_var = tk.BooleanVar(value=bool(pref.get("show_side_panel", True)))
                show_summary_var = tk.BooleanVar(value=bool(pref.get("show_turno_summary", True)))
                accent_color_var = tk.StringVar(value=nombre_color_principal(pref.get("accent_color", "Azul hospitalario")))

                hist_initial_var = tk.StringVar(value=str(pref.get("hist_initial_limit", 100)))
                hist_next_var = tk.StringVar(value=str(pref.get("hist_next_limit", 150)))
                hist_filter_var = tk.StringVar(value=pref.get("hist_default_filter", "Todos"))
                hist_order_var = tk.StringVar(value=pref.get("hist_order", "Más reciente primero"))

                turno_default_var = tk.StringVar(value=pref.get("turno_default", "8AM_8AM"))
                ask_rep_var = tk.BooleanVar(value=bool(pref.get("turnos_ask_representante_start", False)))
                gen_report_var = tk.BooleanVar(value=bool(pref.get("turnos_generate_report", True)))
                save_excel_copy_var = tk.BooleanVar(value=bool(pref.get("turnos_save_excel_copy", True)))
                print_empty_report_var = tk.BooleanVar(value=bool(pref.get("turnos_print_empty_report", False)))
                open_archive_var = tk.BooleanVar(value=bool(pref.get("turnos_open_archive_folder", False)))

                pdf_nss_guiones_var = tk.BooleanVar(value=bool(pref.get("pdf_nss_guiones", True)))
                pdf_ars_mode_var = tk.StringVar(value=pref.get("pdf_ars_display_mode", "Abreviada"))
                pdf_nombre_font_var = tk.StringVar(value=str(pref.get("pdf_nombre_font_size", 12)))
                pdf_dir_font_var = tk.StringVar(value=str(pref.get("pdf_direccion_font_size", 12)))
                pdf_open_var = tk.BooleanVar(value=bool(pref.get("pdf_open_after_generate", True)))
                pdf_keep_temp_var = tk.BooleanVar(value=bool(pref.get("pdf_keep_temp", False)))

                p_print = tb.Frame(pref_nb, padding=14, style="Card.TFrame")
                pref_nb.add(p_print, text="Impresión")
                p_print.columnconfigure(1, weight=1)

                tb.Checkbutton(p_print, text="Activar impresión automática (reportes y Excel)", variable=auto_print_var).grid(row=0, column=0, columnspan=2, sticky="w", padx=6, pady=6)
                tb.Checkbutton(p_print, text="Permitir impresión operativa de hoja", variable=print_auto_hoja_var).grid(row=1, column=0, columnspan=2, sticky="w", padx=6, pady=6)
                add_labeled_entry(p_print, 2, "Copias (hoja):", copies_hoja_var)
                add_labeled_combo(p_print, 3, "Al generar hoja:", print_behavior_var, ["Solo imprimir", "Imprimir y abrir PDF"], 26)
        
                tb.Checkbutton(p_print, text="Imprimir reporte al cambiar turno", variable=print_auto_reporte_var).grid(row=4, column=0, columnspan=2, sticky="w", padx=6, pady=6)
                add_labeled_entry(p_print, 5, "Copias (reporte):", copies_reporte_var)
                add_labeled_combo(p_print, 6, "Orientación reporte PDF:", pdf_orientation_var, ["Horizontal", "Vertical"], 18)
        
                tb.Checkbutton(p_print, text="Imprimir Excel al cambiar turno", variable=print_auto_excel_var).grid(row=7, column=0, columnspan=2, sticky="w", padx=6, pady=6)
                add_labeled_entry(p_print, 8, "Copias (Excel):", copies_excel_var)
                add_labeled_combo(p_print, 9, "Orientación Excel:", excel_orientation_var, ["Horizontal", "Vertical"], 18)

                p_val = tb.Frame(pref_nb, padding=14, style="Card.TFrame")
                pref_nb.add(p_val, text="Validación")
                checks_val = [
                    ("Confirmar antes de generar PDF", confirm_var),
                    ("Alertar si NSS parece incompleto", warn_nss_var),
                    ("Alertar si ARS se detecta como SIN SEGURO", warn_ars_var),
                    ("Bloquear ARS de una letra o solo números", block_short_ars_var),
                    ("Permitir paciente sin cédula", allow_cedula_var),
                    ("Permitir paciente sin teléfono", allow_phone_var),
                    ("Alertar posible duplicado en turno actual", warn_dup_var),
                ]
                for i, (txt, var) in enumerate(checks_val):
                    tb.Checkbutton(p_val, text=txt, variable=var).grid(row=i, column=0, sticky="w", padx=6, pady=5)

                p_rn = tb.Frame(pref_nb, padding=14, style="Card.TFrame")
                pref_nb.add(p_rn, text="Recién nacido")
                tb.Checkbutton(p_rn, text="No guardar RN- en la base de datos", variable=rn_strip_var).grid(row=0, column=0, sticky="w", padx=6, pady=6)
                tb.Checkbutton(p_rn, text="Mostrar RN- solo en el PDF", variable=rn_show_pdf_var).grid(row=1, column=0, sticky="w", padx=6, pady=6)
                tb.Checkbutton(p_rn, text="Mostrar aviso cuando el nombre tenga RN-", variable=rn_warn_var).grid(row=2, column=0, sticky="w", padx=6, pady=6)
                add_labeled_combo(p_rn, 3, "Formato visual:", rn_format_var, ["RN- NOMBRE DE LA MADRE", "RN DE NOMBRE DE LA MADRE", "RECIÉN NACIDO DE NOMBRE"], 32)

                p_visual = tb.Frame(pref_nb, padding=14, style="Card.TFrame")
                pref_nb.add(p_visual, text="Visual")
                add_labeled_combo(p_visual, 0, "Tamaño de letra:", font_var, ["10", "11", "12", "13", "14", "15", "16", "18"], 18)
                if self._host_theme_controlled:
                    host_theme_var = tk.StringVar(
                        value=f"Controlado por Facturación ({self._embedded_theme_label()})"
                    )
                    self._embedded_theme_indicators.append(host_theme_var)
                    tb.Label(p_visual, text="Tema:").grid(row=1, column=0, sticky="w", padx=6, pady=6)
                    tb.Label(p_visual, textvariable=host_theme_var, style="Muted.TLabel").grid(
                        row=1, column=1, sticky="w", padx=6, pady=6
                    )
                    tb.Label(p_visual, text="Tamaño de ventana:").grid(row=2, column=0, sticky="w", padx=6, pady=6)
                    tb.Label(
                        p_visual,
                        text="Controlado por el contenedor de Facturación",
                        style="Muted.TLabel",
                    ).grid(row=2, column=1, sticky="w", padx=6, pady=6)
                else:
                    add_labeled_combo(p_visual, 1, "Tema:", theme_var, ["oscuro", "claro"], 18)
                    add_labeled_combo(p_visual, 2, "Tamaño de ventana:", autosize_var, ["1220x700", "1280x740", "1366x768", "1440x820", "1600x900"], 18)
                add_labeled_combo(p_visual, 3, "Tamaño de botones:", button_size_var, ["Compacto", "Normal", "Grande"], 18)
                add_labeled_entry(p_visual, 4, "Alto de filas:", table_row_height_var, 10)
                add_labeled_combo(p_visual, 5, "Color principal:", accent_color_var, list(ACCENT_COLOR_PRESETS.keys()), 22)
                tb.Checkbutton(p_visual, text="Modo compacto", variable=compact_mode_var).grid(row=6, column=0, sticky="w", padx=6, pady=6)
                tb.Checkbutton(p_visual, text="Modo pantalla pequeña", variable=small_screen_var).grid(row=6, column=1, sticky="w", padx=6, pady=6)
                tb.Checkbutton(p_visual, text="Modo alto contraste", variable=contrast_var).grid(row=7, column=0, sticky="w", padx=6, pady=6)
                tb.Checkbutton(p_visual, text="Mostrar panel derecho", variable=show_side_panel_var).grid(row=8, column=0, sticky="w", padx=6, pady=6)
                tb.Checkbutton(p_visual, text="Mostrar resumen del turno", variable=show_summary_var).grid(row=8, column=1, sticky="w", padx=6, pady=6)

                p_hist = tb.Frame(pref_nb, padding=14, style="Card.TFrame")
                pref_nb.add(p_hist, text="Historial")
                add_labeled_entry(p_hist, 0, "Registros al abrir:", hist_initial_var, 10)
                add_labeled_entry(p_hist, 1, "Cargar más:", hist_next_var, 10)
                add_labeled_combo(p_hist, 2, "Filtro inicial:", hist_filter_var, ["Todos", "Hoy", "Turno actual", "Sin seguro", "Por ARS", "Por especialidad", "Por fecha"], 20)
                add_labeled_combo(p_hist, 3, "Orden:", hist_order_var, ["Más reciente primero", "Más antiguo primero"], 22)

                p_turno = tb.Frame(pref_nb, padding=14, style="Card.TFrame")
                pref_nb.add(p_turno, text="Turnos")
                add_labeled_combo(p_turno, 0, "Turno predeterminado:", turno_default_var, ["8AM_8AM", "8AM_8PM", "8PM_8AM"], 18)
                tb.Checkbutton(p_turno, text="Preguntar representante al iniciar", variable=ask_rep_var).grid(row=1, column=0, sticky="w", padx=6, pady=6)
                tb.Checkbutton(p_turno, text="Generar reporte al cambiar turno", variable=gen_report_var).grid(row=2, column=0, sticky="w", padx=6, pady=6)
                tb.Checkbutton(p_turno, text="Guardar copia del Excel al cambiar turno", variable=save_excel_copy_var).grid(row=3, column=0, sticky="w", padx=6, pady=6)
                tb.Checkbutton(p_turno, text="Imprimir reporte aunque no haya pacientes", variable=print_empty_report_var).grid(row=4, column=0, sticky="w", padx=6, pady=6)
                tb.Checkbutton(p_turno, text="Abrir carpeta del archivo diario después del cambio", variable=open_archive_var).grid(row=5, column=0, sticky="w", padx=6, pady=6)

                p_pdf = tb.Frame(pref_nb, padding=14, style="Card.TFrame")
                pref_nb.add(p_pdf, text="PDF")
                tb.Checkbutton(p_pdf, text="Mostrar NSS con guiones según ARS", variable=pdf_nss_guiones_var).grid(row=0, column=0, sticky="w", padx=6, pady=6)
                add_labeled_combo(p_pdf, 1, "Mostrar ARS:", pdf_ars_mode_var, ["Abreviada", "Completa"], 18)
                add_labeled_entry(p_pdf, 2, "Letra nombre:", pdf_nombre_font_var, 10)
                add_labeled_entry(p_pdf, 3, "Letra dirección:", pdf_dir_font_var, 10)
                tb.Checkbutton(p_pdf, text="Abrir PDF después de generarlo", variable=pdf_open_var).grid(row=4, column=0, sticky="w", padx=6, pady=6)

                pref_estado_var = tk.StringVar(value="Organiza aquí impresión, validación, RN, visual, historial, turnos y PDF.")

                def _int_pref(var, default, min_value=1):
                    try:
                        value = int(str(var.get()).strip())
                        return max(min_value, value)
                    except Exception:
                        return default

                def guardar_pref():
                    updates = {
                        "font_size": _int_pref(font_var, 11, 10),
                        "high_contrast": bool(contrast_var.get()),
                        "auto_print": bool(auto_print_var.get()),

                        "print_auto_hoja": bool(print_auto_hoja_var.get()),
                        "print_auto_reporte_turno": bool(print_auto_reporte_var.get()),
                        "print_auto_excel_turno": bool(print_auto_excel_var.get()),
                        "print_copies_hoja": _int_pref(copies_hoja_var, 1, 1),
                        "print_copies_reporte": _int_pref(copies_reporte_var, 2, 1),
                        "print_copies_excel": _int_pref(copies_excel_var, 2, 1),
                        "print_pdf_orientation": pdf_orientation_var.get(),
                        "print_excel_orientation": excel_orientation_var.get(),
                        "print_behavior_hoja": print_behavior_var.get(),

                        "validation_confirm_before_generate": bool(confirm_var.get()),
                        "validation_warn_nss_incomplete": bool(warn_nss_var.get()),
                        "validation_warn_ars_sin_seguro": bool(warn_ars_var.get()),
                        "validation_block_short_ars": bool(block_short_ars_var.get()),
                        "validation_allow_missing_cedula": bool(allow_cedula_var.get()),
                        "validation_allow_missing_phone": bool(allow_phone_var.get()),
                        "validation_warn_duplicate_turno": bool(warn_dup_var.get()),

                        "rn_strip_db": bool(rn_strip_var.get()),
                        "rn_show_pdf": bool(rn_show_pdf_var.get()),
                        "rn_warn": bool(rn_warn_var.get()),
                        "rn_format_display": rn_format_var.get(),

                        "button_size": button_size_var.get(),
                        "table_row_height": _int_pref(table_row_height_var, 29, 20),
                        "compact_mode": bool(compact_mode_var.get()),
                        "small_screen_mode": bool(small_screen_var.get()),
                        "show_side_panel": bool(show_side_panel_var.get()),
                        "show_turno_summary": bool(show_summary_var.get()),
                        "accent_color": nombre_color_principal(accent_color_var.get()),

                        "hist_initial_limit": _int_pref(hist_initial_var, 100, 50),
                        "hist_next_limit": _int_pref(hist_next_var, 150, 50),
                        "hist_default_filter": hist_filter_var.get(),
                        "hist_order": hist_order_var.get(),

                        "turno_default": normalizar_turno_codigo(turno_default_var.get()),
                        "turnos_ask_representante_start": bool(ask_rep_var.get()),
                        "turnos_generate_report": bool(gen_report_var.get()),
                        "turnos_save_excel_copy": bool(save_excel_copy_var.get()),
                        "turnos_print_empty_report": bool(print_empty_report_var.get()),
                        "turnos_open_archive_folder": bool(open_archive_var.get()),

                        "pdf_nss_guiones": bool(pdf_nss_guiones_var.get()),
                        "pdf_ars_display_mode": pdf_ars_mode_var.get(),
                        "pdf_nombre_font_size": _int_pref(pdf_nombre_font_var, 12, 6),
                        "pdf_direccion_font_size": _int_pref(pdf_dir_font_var, 12, 6),
                        "pdf_open_after_generate": bool(pdf_open_var.get()),
                        "pdf_keep_temp": bool(pdf_keep_temp_var.get()),
                    }
                    if not self._host_theme_controlled:
                        updates["theme"] = theme_var.get()
                        updates["window_size"] = autosize_var.get()
                    self.app_settings.update(updates)

                    if not guardar_app_settings(self.app_settings):
                        messagebox.showwarning("Preferencias", "No se pudieron guardar las preferencias en disco.")
                        return

                    self._aplicar_preferencias_en_vivo(win)
                    self._actualizar_resumen_turno_panel()
                    pref_estado_var.set("Preferencias guardadas y aplicadas.")
                    self.set_status("Preferencias aplicadas", "ok")
                    messagebox.showinfo(
                        "Preferencias",
                        "Preferencias guardadas correctamente. Algunos cambios visuales estructurales pueden requerir cerrar y abrir la app."
                    )

                bottom_pref = tb.Frame(tab_pref, padding=(8, 10), style="Card.TFrame")
                bottom_pref.pack(fill="x", side="bottom", pady=(10, 0))
                tb.Label(bottom_pref, textvariable=pref_estado_var, style="Muted.TLabel").pack(side="left", padx=5)
                tb.Button(bottom_pref, text="💾  Guardar preferencias", bootstyle=SUCCESS, command=guardar_pref, width=22).pack(side="right", padx=5, ipady=5)

                _config_section_loaders = {
                    "Administrar ARS": cargar_ars,
                    "Catálogo ARS": cargar_catalogo_tree,
                    "Formato NSS PDF": refrescar_formatos_nss,
                    "Revisión NSS": cargar_revisiones_nss,
                    "Respaldos": refrescar_respaldos,
                    "Representante del turno": lambda: (
                        refrescar_control_primary(), cargar_usuarios()
                    ),
                }
                _config_sections_loaded = set()

                def _cargar_seccion_configuracion(_event=None, *, force=False):
                    try:
                        selected = notebook.select()
                        section = str(notebook.tab(selected, "text") or "")
                    except Exception:
                        section = "Administrar ARS"
                    loader = _config_section_loaders.get(section)
                    if loader is None:
                        return
                    if section in _config_sections_loaded and not force:
                        return
                    _config_sections_loaded.add(section)
                    started = _time.perf_counter()
                    APP_LOG.info("CONFIG_SECTION_LOAD_START section=%s", section)
                    try:
                        loader()
                    except Exception as exc:
                        _config_sections_loaded.discard(section)
                        APP_LOG.error(
                            "CONFIG_SECTION_LOAD_ERROR section=%s elapsed_ms=%.1f type=%s",
                            section,
                            (_time.perf_counter() - started) * 1000.0,
                            type(exc).__name__,
                            exc_info=(type(exc), exc, exc.__traceback__),
                        )

                notebook.bind("<<NotebookTabChanged>>", _cargar_seccion_configuracion)
                win.after(1, _cargar_seccion_configuracion)

                win._config_content_ready = True
                self._aplicar_preferencias_a_widgets(win)
                APP_LOG.info(
                    "CONFIG_UI_BUILD_MS elapsed_ms=%.1f",
                    (_time.perf_counter() - build_started) * 1000.0,
                )
            except Exception as exc:
                win._config_content_ready = False
                APP_LOG.error(
                    "CONFIG_SECTION_LOAD_ERROR section=ui_build elapsed_ms=%.1f type=%s message=%s",
                    (_time.perf_counter() - build_started) * 1000.0,
                    type(exc).__name__,
                    str(exc),
                    exc_info=(type(exc), exc, exc.__traceback__),
                )
                try:
                    error_tab = tb.Frame(notebook, padding=18, style="Card.TFrame")
                    notebook.add(error_tab, text="Error")
                    tb.Label(
                        error_tab,
                        text=(
                            "No fue posible construir una sección de Configuración. "
                            "Cierre y vuelva a abrir esta ventana."
                        ),
                        style="Muted.TLabel",
                        background="#0E1B2B",
                        wraplength=720,
                        justify="left",
                    ).pack(anchor="w", pady=16)
                except Exception:
                    pass

        def _config_first_frame():
            if not win.winfo_exists():
                return
            self._aplicar_preferencias_a_widgets(win)
            elapsed_ms = (_time.perf_counter() - config_open_started) * 1000.0
            APP_LOG.info("CONFIG_FIRST_FRAME elapsed_ms=%.1f", elapsed_ms)
            APP_LOG.info("CONFIG_SHOW_MS elapsed_ms=%.1f", elapsed_ms)
            win.after(1, _build_config_content)

        win.after_idle(_config_first_frame)
        return



    def mostrar_menu_contextual(self, event):
        try:
            self._widget_actual = event.widget
            self.menu_contextual.post(event.x_root, event.y_root)
        except Exception:
            pass

    def _copiar(self):
        try:
            self.root.clipboard_clear()
            self.root.clipboard_append(self._widget_actual.selection_get())
        except Exception:
            pass

    def _pegar(self):
        try:
            self._widget_actual.insert(tk.INSERT, self.root.clipboard_get())
        except Exception:
            pass

    def _cortar(self):
        try:
            sel = self._widget_actual.selection_get()
            self.root.clipboard_clear()
            self.root.clipboard_append(sel)
            self._widget_actual.delete("sel.first", "sel.last")
        except Exception:
            pass

    def _set_turn_change_controls_enabled(self, enabled: bool):
        allowed = False
        if enabled:
            try:
                allowed = bool(self.can_change_admission_turn()[0])
            except Exception:
                allowed = False
        try:
            self.boton_cambiar_turno.configure(
                state="normal" if enabled and allowed else "disabled"
            )
        except Exception:
            pass
        if enabled:
            self._refresh_actions_menu_state()

    def _refresh_actions_menu_state(self):
        """Refresh controls from the adopted in-memory operational snapshot."""
        try:
            self._configurar_colores_menu(
                self._paleta_visual_actual(), self._font_size_pref()
            )
        except Exception:
            pass
        allowed, _reason_code, _message = self.can_change_admission_turn()
        try:
            self.change_turn_menu_action.setEnabled(bool(allowed))
        except Exception:
            pass
        transfer_enabled = False
        try:
            runtime = getattr(self.db, "_runtime", None)
            snapshot = dict(runtime.state() or {}) if runtime is not None else {}
            transfer_enabled = bool(
                normalize_role(self.session_context.role) == ROLE_ADMIN
                and runtime is not None
                and not bool(getattr(runtime, "offline", False))
                and not self._primary_transfer_in_progress
                and str(snapshot.get("role") or "").upper() != "PRIMARY"
            )
            if self.transfer_primary_menu_action is not None:
                self.transfer_primary_menu_action.setEnabled(transfer_enabled)
            config_button = getattr(
                self, "_primary_transfer_config_button", None
            )
            if config_button is not None:
                config_button.configure(
                    state="normal" if transfer_enabled else "disabled"
                )
        except Exception:
            pass

    def can_change_admission_turn(self, *, allow_open_dialog=False):
        if (
            (self._turn_change_in_progress and not allow_open_dialog)
            or self._turn_change_committing
        ):
            return False, "TRANSITION_IN_PROGRESS", "Ya existe un cambio de turno en curso."
        runtime = getattr(self.db, "_runtime", None)
        if runtime is None:
            return (
                False,
                "CENTRAL_OFFLINE",
                "No está disponible el contexto operacional central.",
            )
        if bool(getattr(runtime, "offline", False)):
            return (
                False,
                "CENTRAL_OFFLINE",
                "Se requiere conexión central para cambiar el turno.",
            )
        snapshot = dict(runtime.state() or {})
        if not bool(snapshot.get("can_change_turn")):
            is_primary = str(snapshot.get("role") or "").upper() == "PRIMARY"
            return (
                False,
                "ROLE_NOT_ALLOWED" if is_primary else "NOT_PRIMARY",
                "Solo un usuario operativo autorizado en la estación PRIMARY "
                "puede cambiar el turno de Admisión."
                if is_primary
                else "Solo la estación PRIMARY activa puede cambiar el turno.",
            )
        if str(snapshot.get("role") or "").upper() != "PRIMARY":
            same_device = str(snapshot.get("primary_device_id") or "") == str(
                getattr(runtime, "device_id", "") or ""
            )
            return (
                False,
                "PRIMARY_LEASE_STALE" if same_device else "NOT_PRIMARY",
                "La sesión principal necesita actualizarse."
                if same_device
                else "Solo la estación PRIMARY activa puede cambiar el turno.",
            )
        try:
            runtime.require_primary_turn_change()
            return True, "ALLOWED", ""
        except Exception as exc:
            return False, "PRIMARY_LEASE_STALE", str(exc)

    def request_change_admission_turn(self, *_):
        """Canonical entry point used by the button, menu and F5."""
        if self._turn_change_in_progress:
            return "break"
        allowed, _reason_code, reason = self.can_change_admission_turn()
        if not allowed:
            messagebox.showwarning("Cambiar turno", reason, parent=self.root)
            return "break"
        if not messagebox.askyesno(
            "Confirmación",
            "¿Está seguro de que desea cambiar de turno y reiniciar los datos del Excel?",
            parent=self.root,
        ):
            return "break"
        self._turn_change_in_progress = True
        self._set_turn_change_controls_enabled(False)
        try:
            self._dialogo_turno()
        except Exception as exc:
            self._turn_change_in_progress = False
            self._set_turn_change_controls_enabled(True)
            APP_LOG.exception("No se pudo abrir el diálogo de cambio de turno")
            messagebox.showerror(
                "Cambiar turno",
                f"No se pudo abrir la configuración del turno:\n{exc}",
                parent=self.root,
            )
        return "break"

    def request_transfer_admission_primary(self, *, parent=None):
        """Transfer only the central PRIMARY lease from the current Admin login."""
        parent = parent or self.root
        if normalize_role(self.session_context.role) != ROLE_ADMIN:
            messagebox.showwarning(
                "Transferir acceso principal",
                "Solo un Administrador puede realizar esta operación.",
                parent=parent,
            )
            return
        runtime = getattr(self.db, "_runtime", None)
        if runtime is None or bool(getattr(runtime, "offline", False)):
            messagebox.showwarning(
                "Transferir acceso principal",
                "Se requiere conexión central para transferir PRIMARY.",
                parent=parent,
            )
            return
        if self._primary_transfer_in_progress:
            return
        before = dict(runtime.state() or {})
        if str(before.get("role") or "").upper() == "PRIMARY":
            messagebox.showinfo(
                "Transferir acceso principal",
                "Esta estación ya posee el acceso principal.",
                parent=parent,
            )
            return
        summary = (
            "Computadora actual:\n"
            f"{getattr(runtime, 'device_id', '') or 'No disponible'}\n\n"
            f"Rol actual:\n{str(before.get('role') or 'NONE').upper()}\n\n"
            "Computadora principal actual:\n"
            f"{before.get('primary_device_id') or 'Sin asignar'}\n\n"
            "Usuario operativo:\n"
            f"{before.get('active_user_display_name') or before.get('active_username') or 'No disponible'}\n\n"
            f"Turno:\n{before.get('turn_id') or 'No disponible'}"
        )
        messagebox.showinfo(
            "TRANSFERIR SESIÓN PRINCIPAL", summary, parent=parent
        )
        actor = self._solicitar_autorizacion_admin(
            "TRANSFERIR_ACCESO_PRINCIPAL", parent=parent, force=True
        )
        if not actor:
            return
        reason = simpledialog.askstring(
            "Motivo de la transferencia",
            "Indique el motivo de la transferencia de PRIMARY:",
            parent=parent,
        )
        if reason is None:
            return
        reason = reason.strip()
        if not reason:
            messagebox.showwarning(
                "Motivo requerido", "Debe indicar un motivo breve.", parent=parent
            )
            return
        if not messagebox.askyesno(
            "Transferir acceso principal",
            "Esta operación transferirá el acceso PRINCIPAL de Admisión "
            "a esta computadora.\n\n"
            "La estación principal anterior perderá su sesión actual.\n\n"
            "El turno y el usuario operativo de Admisión NO serán "
            "modificados.\n\n¿Desea continuar?",
            parent=parent,
        ):
            return
        self._primary_transfer_in_progress = True
        self._refresh_actions_menu_state()

        def transfer():
            return runtime.force_transfer_admission_primary(reason=reason)

        def finish(changed):
            self._primary_transfer_in_progress = False
            self._refresh_actions_menu_state()
            self._set_turn_change_controls_enabled(True)
            self.set_status("Conectado · Principal · Sincronizado", "ok")
            refresh_panel = getattr(self, "_refresh_primary_config_panel", None)
            if callable(refresh_panel):
                refresh_panel()
            if (
                changed.turn_id != before.get("turn_id")
                or changed.generation != before.get("generation")
            ):
                APP_LOG.critical(
                    "PRIMARY_FORCE_TRANSFER_INVARIANT_FAILED turn_before=%s "
                    "turn_after=%s generation_before=%s generation_after=%s",
                    before.get("turn_id"), changed.turn_id,
                    before.get("generation"), changed.generation,
                )
                messagebox.showerror(
                    "Transferir acceso principal",
                    "La transferencia no conservó el estado operativo esperado.",
                    parent=parent,
                )
                return
            messagebox.showinfo(
                "Transferencia completada",
                "Esta estación es ahora PRIMARY. El turno y el usuario "
                "operativo permanecen sin cambios.",
                parent=parent,
            )

        def failed(exc):
            self._primary_transfer_in_progress = False
            self._refresh_actions_menu_state()
            self.set_status("No se pudo transferir PRIMARY", "error")
            refresh_panel = getattr(self, "_refresh_primary_config_panel", None)
            if callable(refresh_panel):
                refresh_panel()
            APP_LOG.error(
                "Falló la transferencia administrativa de PRIMARY",
                exc_info=(type(exc), exc, exc.__traceback__),
            )
            messagebox.showerror(
                "Transferir acceso principal", str(exc), parent=parent
            )

        self._ejecutar_en_segundo_plano(
            "Transfiriendo acceso PRIMARY...",
            transfer,
            al_terminar=finish,
            al_error=failed,
        )

    def request_force_primary_transfer(self):
        """Compatibility alias for builds that still reference the old name."""
        return self.request_transfer_admission_primary()

    def reiniciar_datos_excel(self, *_):
        """Compatibility alias for integrations using the legacy method name."""
        return self.request_change_admission_turn()

    def _dialogo_turno(self):
        # This slot can still be reached by legacy callbacks.  A station that
        # already received a central SECONDARY snapshot must not fall back to
        # local shift configuration or present the transition dialog.
        snapshot = self._snapshot_operacional_integrado()
        if str(snapshot.get("role") or "").strip().upper() == "SECONDARY":
            self.apply_operational_snapshot(snapshot)
            self._turn_change_in_progress = False
            return

        win = self._crear_toplevel_estable("Configurar turno", "680x460", "turno_win")
        if win is None:
            self._turn_change_in_progress = False
            self._set_turn_change_controls_enabled(True)
            return

        def release_turn_action(event=None):
            if event is not None and getattr(event, "widget", None) is not win:
                return
            self._turn_change_in_progress = False
            self._turn_change_committing = False
            self._set_turn_change_controls_enabled(True)

        win.bind("<Destroy>", release_turn_action, add="+")

        self._bind_esc_cerrar(win)

        cont = tb.Frame(win, padding=14, style="Root.TFrame")
        cont.pack(fill="both", expand=True)

        runtime_turno = getattr(self.db, "_runtime", None)
        snapshot_turno = dict(runtime_turno.state() or {}) if runtime_turno is not None else {}
        relevo_formal = bool(
            runtime_turno is not None
            and runtime_turno.is_primary_shift_handover()
        )
        usuario_autenticado = limpiar_nombre_representante(
            getattr(self.session_context, "display_name", "")
            or getattr(self.session_context, "username", "")
        )
        subtitulo_turno = (
            "Confirma el horario. El usuario autenticado asumirá el nuevo turno."
            if relevo_formal
            else "Selecciona el horario. El representante operacional actual se conserva."
        )
        self._crear_header_ventana(
            cont,
            "Configurar turno",
            subtitulo_turno,
            "⚙"
        )

        form_card = tb.Frame(cont, padding=12, style="Card.TFrame")
        form_card.pack(fill="both", expand=True)

        usuario_sesion = limpiar_nombre_representante(
            snapshot_turno.get("active_user_display_name")
            or snapshot_turno.get("active_username")
        )
        rep_box = tb.Frame(form_card, padding=12, style="Card.TFrame")
        rep_box.grid(row=0, column=0, columnspan=2, sticky="ew", pady=(0, 12))
        tb.Label(
            rep_box,
            text="RESPONSABLE DEL TURNO",
            font=("Arial", 9, "bold"),
            foreground="#67B7FF",
            background="#0E1B2B",
        ).pack(anchor="w")
        tb.Label(
            rep_box,
            text=(usuario_autenticado if relevo_formal else usuario_sesion)
            or "Representante no configurado",
            font=("Arial", 13, "bold"),
            foreground="#FFFFFF",
            background="#0E1B2B",
        ).pack(anchor="w", pady=(3, 2))
        tb.Label(
            rep_box,
            text=(
                "Relevo formal: al confirmar, este usuario será el responsable del "
                "nuevo turno y se cerrará la sesión secundaria del responsable saliente."
                if relevo_formal
                else "El cambio de horario no modifica al representante operacional actual."
            ),
            wraplength=560,
            justify="left",
            foreground="#BDD6F4",
            background="#0E1B2B",
        ).pack(anchor="w")

        tb.Label(form_card, text="Horario del turno:", background="#0E1B2B", foreground="#EAF2FF").grid(row=1, column=0, sticky="w", pady=6)

        central_turn_code = normalizar_turno_codigo(
            snapshot_turno.get("turn_code") or self.app_settings.get("turno_default", "8AM_8AM")
        )
        turno_var = tk.StringVar(value=central_turn_code)
        combo_turno = tb.Combobox(
            form_card,
            textvariable=turno_var,
            state="readonly",
            values=[
                "8:00 AM → 8:00 AM",
                "8:00 AM → 8:00 PM",
                "8:00 PM → 8:00 AM",
            ],
            width=30
        )
        combo_turno.grid(row=1, column=1, sticky="w", pady=6)
        _td = central_turn_code
        combo_turno.set("8:00 AM → 8:00 PM" if _td == "8AM_8PM" else ("8:00 PM → 8:00 AM" if _td == "8PM_8AM" else "8:00 AM → 8:00 AM"))

        aviso_var = tk.StringVar(value="")
        aviso_lbl = tb.Label(form_card, textvariable=aviso_var, bootstyle=INFO, wraplength=470, justify="left")
        aviso_lbl.grid(row=2, column=0, columnspan=2, sticky="w", pady=(8, 10))

        vista_turno_var = tk.StringVar(value="")
        vista_fecha_var = tk.StringVar(value="")
        vista_inicio_real_var = tk.StringVar(value="")

        tb.Label(form_card, text="Vista previa del turno:", bootstyle=SECONDARY, background="#0E1B2B").grid(row=3, column=0, sticky="w", pady=(5, 2))
        tb.Label(form_card, textvariable=vista_turno_var, font=("Arial", 11, "bold"), background="#0E1B2B", foreground="#FFFFFF").grid(row=3, column=1, sticky="w", pady=(5, 2))

        tb.Label(form_card, text="Vista previa de fecha:", bootstyle=SECONDARY, background="#0E1B2B").grid(row=4, column=0, sticky="w", pady=(2, 2))
        tb.Label(form_card, textvariable=vista_fecha_var, font=("Arial", 11), background="#0E1B2B", foreground="#EAF2FF").grid(row=4, column=1, sticky="w", pady=(2, 2))

        tb.Label(form_card, text="Conteo real desde:", bootstyle=SECONDARY, background="#0E1B2B").grid(row=5, column=0, sticky="w", pady=(2, 8))
        tb.Label(form_card, textvariable=vista_inicio_real_var, font=("Arial", 11), background="#0E1B2B", foreground="#EAF2FF").grid(row=5, column=1, sticky="w", pady=(2, 8))

        fecha_base = datetime.now().date()

        def es_domingo(f: date) -> bool:
            return f.weekday() == 6

        def es_sabado(f: date) -> bool:
            return f.weekday() == 5

        def normalizar_turno_desde_combo(texto_combo: str) -> str:
            mapa = {
                "8:00 AM → 8:00 AM": "8AM_8AM",
                "8:00 AM → 8:00 PM": "8AM_8PM",
                "8:00 PM → 8:00 AM": "8PM_8AM",
            }
            return mapa.get(texto_combo, "8AM_8AM")

        def actualizar_vista_previa():
            turno_codigo = normalizar_turno_desde_combo(combo_turno.get())
            datos_turno = obtener_datos_turno_visual(fecha_base, turno_codigo)
            vista_turno_var.set(datos_turno["turno_label"])
            vista_fecha_var.set(datos_turno["fecha_label"])
            vista_inicio_real_var.set(datetime.now().strftime("%d/%m/%Y %I:%M %p"))

        def refrescar_turnos():
            combo_turno.configure(values=[
                "8:00 AM → 8:00 AM",
                "8:00 AM → 8:00 PM",
                "8:00 PM → 8:00 AM",
            ])
            aviso_var.set(
                (
                    "El nuevo conteo quedará separado y el usuario autenticado asumirá "
                    "el nuevo turno. Los tres turnos canónicos están disponibles todos los días."
                    if relevo_formal
                    else "El nuevo conteo quedará separado y conservará al representante "
                    "operacional actual. Los tres turnos canónicos están disponibles todos los días."
                )
            )
            actualizar_vista_previa()

        combo_turno.bind("<<ComboboxSelected>>", lambda e: actualizar_vista_previa())
        refrescar_turnos()

        def _aplicar_cambio():
            runtime_actual = getattr(self.db, "_runtime", None)
            snapshot_actual = dict(runtime_actual.state() or {}) if runtime_actual is not None else {}
            representante = limpiar_nombre_representante(
                snapshot_actual.get("active_user_display_name")
                or snapshot_actual.get("active_username")
            )
            if not es_representante_valido(representante):
                messagebox.showwarning(
                    "Representante",
                    "Configure primero un representante operacional. El usuario "
                    "autenticado no se asigna automáticamente al turno.",
                    parent=win,
                )
                return

            turno_codigo = normalizar_turno_desde_combo(combo_turno.get())

            momento_cambio = datetime.now()
            turno_saliente = cargar_turno_config(permitir_vencido=True)
            candidato = {
                "representante": representante,
                "turno_codigo": turno_codigo,
                "fecha_base": fecha_base,
                "inicio_real": format_datetime_local(momento_cambio),
                "inicio_real_dt": momento_cambio,
            }
            administrative_override = not turno_config_es_vigente(
                candidato, momento=momento_cambio
            )
            relevo_formal_actual = bool(
                runtime_actual is not None
                and runtime_actual.is_primary_shift_handover()
                and not administrative_override
            )
            if relevo_formal_actual:
                nuevo_representante = limpiar_nombre_representante(
                    getattr(self.session_context, "display_name", "")
                    or getattr(self.session_context, "username", "")
                )
                if not es_representante_valido(nuevo_representante):
                    messagebox.showwarning(
                        "Cambio de turno",
                        "No fue posible identificar al usuario operativo autenticado.",
                        parent=win,
                    )
                    return
                representante = nuevo_representante
                candidato["representante"] = representante
            elif not administrative_override:
                messagebox.showerror(
                    "No se puede realizar el relevo",
                    "No se puede realizar el relevo.\n"
                    "El usuario seleccionado ya es el representante del turno actual.\n"
                    "Seleccione al usuario que recibirá el próximo turno.",
                    parent=win,
                )
                return
            override_reason = ""
            if administrative_override:
                if normalize_role(self.session_context.role) != ROLE_ADMIN:
                    messagebox.showwarning(
                        "Cambio de turno",
                        "Solo un Administrador puede aplicar una corrección de turno fuera de horario.",
                        parent=win,
                    )
                    return
                if not messagebox.askyesno(
                    "Corrección administrativa de turno",
                    "Este turno no coincide con el horario que correspondería "
                    "automáticamente en este momento.\n\n"
                    "El representante y PRIMARY no cambiarán.\n\n"
                    "¿Desea aplicarlo como corrección administrativa?",
                    parent=win,
                ):
                    return
                override_reason = simpledialog.askstring(
                    "Razón de corrección",
                    "Razón para la auditoría:",
                    initialvalue="Corrección administrativa de turno",
                    parent=win,
                )
                if override_reason is None:
                    return
                override_reason = (
                    override_reason.strip() or "Corrección administrativa de turno"
                )
                candidato["administrative_override"] = True
                candidato["override_reason"] = override_reason

            try:
                self.db.backup_manager.create(
                    "cierre_turno",
                    label=(
                        f"representante={representante}; "
                        f"momento={momento_cambio.isoformat(timespec='seconds')}"
                    ),
                )
            except Exception as exc:
                APP_LOG.exception("No se pudo crear el respaldo previo al cambio de turno")
                messagebox.showerror(
                    "Respaldo requerido",
                    f"No se cambiará el turno porque falló el respaldo previo:\n{exc}",
                    parent=win,
                )
                return

            transition = self.db.perform_explicit_turn_handoff(
                shift_metadata=candidato
            )
            committed = bool(getattr(transition, "committed", False))
            if not committed:
                raise RuntimeError("La transición central no fue confirmada.")
            changed_session = getattr(transition, "operational_session", None)
            central_turn_id = int(getattr(changed_session, "turn_id", 0) or 0)
            if central_turn_id <= 0:
                raise RuntimeError("La transición central no devolvió un turn_id válido.")
            transition_id = str(getattr(transition, "transition_id", "") or "")
            if not transition_id:
                import uuid
                transition_id = str(uuid.uuid4())

            # Desde aquí la operación central está confirmada. Ninguna
            # actualización de Excel/PDF/impresión puede convertirla en FAIL.
            post_commit_warnings = []
            turno_cfg_nuevo = None
            nuevo_turno_local_id = None
            try:
                saved = guardar_turno_config(
                    representante,
                    turno_codigo,
                    fecha_base,
                    inicio_real=momento_cambio,
                    administrative_override=administrative_override,
                    override_reason=override_reason,
                )
                if not saved:
                    raise RuntimeError("No se pudo guardar el espejo de configuración.")
                if turno_saliente and not administrative_override:
                    self.db.cerrar_turno_existente(
                        turno_saliente,
                        momento_cambio,
                        actor=self.session_context.username,
                        actor_role=self.session_context.role,
                        session_id=self.session_context.session_id,
                    )
                guardar_representante_catalogo(representante, self.db)
                turno_cfg_nuevo = cargar_turno_config(
                    permitir_vencido=administrative_override
                )
                if not turno_cfg_nuevo:
                    raise TurnoNoVigenteError("El espejo del turno no quedó disponible.")
                nuevo_turno_local_id = self.db.obtener_o_crear_turno(
                    turno_cfg_nuevo,
                    administrative_override=administrative_override,
                )
            except Exception:
                post_commit_warnings.append("LOCAL_MIRROR")
                APP_LOG.exception(
                    "Turno central confirmado; espejo local pendiente transition=%s turn_id=%s",
                    transition_id,
                    central_turn_id,
                )

            try:
                if turno_cfg_nuevo:
                    enqueue_excel_export_job(
                        transition_id, central_turn_id, turno_cfg_nuevo
                    )
            except Exception:
                post_commit_warnings.append("EXCEL_QUEUE")
                APP_LOG.exception(
                    "Turno confirmado; no se pudo encolar el efecto de Excel transition=%s",
                    transition_id,
                )

            try:
                self.apply_operational_snapshot(
                    self._snapshot_operacional_integrado()
                )
            except Exception:
                post_commit_warnings.append("UI_REFRESH")
                APP_LOG.exception("Turno confirmado; refresco visual pendiente")
            try:
                if nuevo_turno_local_id:
                    self.db.notify_shift_changed(nuevo_turno_local_id)
            except Exception:
                post_commit_warnings.append("SHIFT_EVENT")
                APP_LOG.exception("Turno confirmado; notificación local pendiente")

            try:
                self.turno_win = None
                win.destroy()
            except Exception:
                self.turno_win = None
            if administrative_override:
                APP_LOG.info(
                    "TURN_ADMIN_OVERRIDE_POST_COMMIT_SKIPPED turn_id=%s reason=%s",
                    central_turn_id,
                    override_reason,
                )
            else:
                try:
                    if turno_cfg_nuevo:
                        self.root.after(
                            0,
                            lambda saliente=turno_saliente, nuevo=turno_cfg_nuevo, momento=momento_cambio: (
                                self._run_turn_post_commit_effects(saliente, nuevo, momento)
                            ),
                        )
                except Exception:
                    post_commit_warnings.append("POST_COMMIT_SCHEDULE")
                    APP_LOG.exception("Turno confirmado; efectos post-commit pendientes")

            try:
                if post_commit_warnings:
                    self.set_status(
                        "Cambio de turno aplicado; algunos efectos auxiliares quedaron pendientes.",
                        "warning",
                    )
                else:
                    self.set_status(
                        (
                            "Corrección administrativa de turno aplicada. "
                            "No se generó reporte de cierre."
                            if administrative_override
                            else (
                                "Relevo de turno aplicado. Actualizando Excel en segundo plano."
                                if relevo_formal_actual
                                else "Cambio de turno aplicado. Actualizando Excel en segundo plano."
                            )
                        ),
                        "ok",
                    )
            except Exception:
                APP_LOG.exception("Turno confirmado; no se pudo actualizar el estado visual")
            try:
                messagebox.showinfo(
                    "Cambio aplicado",
                    (
                        "Corrección administrativa de turno aplicada correctamente.\n\n"
                        "El representante y PRIMARY se conservaron. No se generó reporte de cierre."
                        if administrative_override
                        else (
                            "Relevo formal aplicado correctamente.\n\n"
                            "El usuario autenticado quedó como responsable del nuevo turno y "
                            "la sesión secundaria del responsable saliente fue cerrada."
                            if relevo_formal_actual
                            else "Cambio de turno aplicado correctamente.\n\n"
                            "El conteo del nuevo turno comenzará desde la hora real del cambio. "
                            "El listado de Excel se actualizará en segundo plano."
                        )
                    ),
                    parent=self.root,
                )
            except Exception:
                APP_LOG.exception("Turno confirmado; no se pudo mostrar el aviso final")
            return True

        aplicando = {"activo": False}

        def aplicar_una_vez():
            if aplicando["activo"]:
                return
            allowed, _reason_code, reason = self.can_change_admission_turn(
                allow_open_dialog=True
            )
            if not allowed:
                messagebox.showwarning("Cambiar turno", reason, parent=win)
                return
            aplicando["activo"] = True
            self._turn_change_committing = True
            aplicar_btn.configure(state="disabled", text="Aplicando cambio...")
            aviso_var.set("Aplicando cambio...")
            try:
                _aplicar_cambio()
            except Exception as exc:
                APP_LOG.exception("Falló la transición del turno principal")
                if win.winfo_exists():
                    messagebox.showerror(
                        "No se pudo aplicar",
                        "El cambio no se confirmó. Revise el error e intente nuevamente.\n\n"
                        + str(exc),
                        parent=win,
                    )
            finally:
                self._turn_change_committing = False
                if win.winfo_exists():
                    aplicando["activo"] = False
                    aplicar_btn.configure(state="normal", text="Aplicar")

        aplicar_btn = tb.Button(
            form_card, text="Aplicar", bootstyle=SUCCESS, command=aplicar_una_vez
        )
        aplicar_btn.grid(row=6, column=0, pady=12)
        tb.Button(form_card, text="Cancelar", bootstyle=SECONDARY, command=win.destroy).grid(row=6, column=1, sticky="w", pady=12)

    def run(self):
        if not self._standalone:
            raise RuntimeError(
                "App.run() solo está permitido en el wrapper standalone de Admisión."
            )
        self.root.mainloop()


# -------------------------------
# MAIN
# -------------------------------
if __name__ == "__main__":
    _instance_guard = None
    try:
        if SELF_TEST_MODE:
            manager = DatabaseManager(session_context=load_session_context())
            with closing(manager._connect()) as conn:
                integrity = conn.execute("PRAGMA integrity_check").fetchone()
                version = conn.execute("SELECT version FROM schema_version WHERE id=1").fetchone()
            if not integrity or integrity[0] != "ok" or not version or int(version[0]) != LATEST_SCHEMA_VERSION:
                raise RuntimeError("La autoprueba de SQLite no fue satisfactoria.")
            missing_templates = [path for path in RUTA_HOJAS.values() if not os.path.isfile(path)]
            if missing_templates:
                raise FileNotFoundError("Faltan plantillas: " + ", ".join(missing_templates))
            packaged_logo = resource_path("istipo_hospitales.png")
            if not os.path.isfile(packaged_logo):
                raise FileNotFoundError("El logo principal no quedó incluido en el ejecutable.")
            raise SystemExit(0)
        create_standalone_application()
        _instance_guard = SingleInstanceGuard()
        if not _instance_guard.acquire():
            messagebox.showinfo(
                "Admisión ya está abierta",
                "Ya existe una sesión de Admisión abierta en este equipo. "
                "Utiliza esa ventana para continuar.",
            )
            raise SystemExit(0)
        app = App(standalone=True)
        app.run()
    except Exception as exc:
        APP_LOG.exception("Error fatal durante el inicio de la aplicación")
        try:
            messagebox.showerror("Error al iniciar", f"No se pudo iniciar la aplicación:\n{str(exc)}")
        except Exception:
            pass
        raise
    finally:
        if _instance_guard is not None:
            _instance_guard.release()
        if SELF_TEST_DATA_DIR:
            shutil.rmtree(SELF_TEST_DATA_DIR, ignore_errors=True)
