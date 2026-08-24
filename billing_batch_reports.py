from __future__ import annotations

import os
import re
from collections import defaultdict
from datetime import datetime

from openpyxl import Workbook
from openpyxl.drawing.image import Image as ExcelImage
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
from reportlab.lib.pagesizes import landscape, letter
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import inch
from reportlab.platypus import (
    Image, LongTable, PageBreak, Paragraph, SimpleDocTemplate, Spacer, Table,
    TableStyle,
)


BLUE = "1F67B1"
DARK_BLUE = "123A63"
LIGHT_BLUE = "EAF2FB"
GREEN = "198754"
LIGHT_GREEN = "E8F5EE"
GRAY = "66788A"
LIGHT_GRAY = "EEF2F5"
CURRENCY_FORMAT = '"RD$" #,##0.00'


def _period_label(row: dict) -> str:
    return f"{int(row['period_month']):02d}-{int(row['period_year']):04d}"


def _display_datetime(value) -> str:
    text = str(value or "").strip()
    if not text:
        return ""
    for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%dT%H:%M:%S", "%Y-%m-%d"):
        try:
            parsed = datetime.strptime(text[:19], fmt)
            return parsed.strftime("%d-%m-%Y %H:%M" if "%H" in fmt else "%d-%m-%Y")
        except ValueError:
            continue
    return text


def _report_summary(rows: list[dict]) -> dict:
    by_ars = defaultdict(lambda: {"batches": 0, "receipts": 0, "total": 0.0})
    by_period = defaultdict(lambda: {"batches": 0, "receipts": 0, "total": 0.0})
    for row in rows:
        ars = str(row.get("ars") or "")
        period = _period_label(row)
        count = int(row.get("receipt_count") or 0)
        total = float(row.get("sent_total") or 0)
        for bucket, key in ((by_ars, ars), (by_period, period)):
            bucket[key]["batches"] += 1
            bucket[key]["receipts"] += count
            bucket[key]["total"] += total
    return {
        "batch_count": len(rows),
        "receipt_count": sum(int(row.get("receipt_count") or 0) for row in rows),
        "total": sum(float(row.get("sent_total") or 0) for row in rows),
        "by_ars": dict(sorted(by_ars.items(), key=lambda item: item[0].casefold())),
        "by_period": dict(sorted(by_period.items(), reverse=True)),
    }


def _safe_path(base_path: str, extension: str) -> str:
    base = os.path.abspath(str(base_path or "").strip())
    if base.lower().endswith((".pdf", ".xlsx")):
        base = os.path.splitext(base)[0]
    os.makedirs(os.path.dirname(base), exist_ok=True)
    return base + extension


def create_sent_batches_xlsx(
    base_path: str, rows: list[dict], generated_by: str, logo_path: str | None = None,
) -> str:
    if not rows:
        raise ValueError("No hay listados ENVIADOS para generar el reporte.")
    output_path = _safe_path(base_path, ".xlsx")
    summary = _report_summary(rows)
    workbook = Workbook()
    report = workbook.active
    report.title = "Resumen"
    detail = workbook.create_sheet("Listados enviados")
    thin = Side(style="thin", color="CBD5E1")
    border = Border(bottom=thin)

    for sheet in (report, detail):
        sheet.sheet_view.showGridLines = False
        sheet.freeze_panes = "A8"
        sheet.page_setup.orientation = "landscape"
        sheet.page_setup.fitToWidth = 1
        sheet.page_setup.fitToHeight = 0
        sheet.sheet_properties.pageSetUpPr.fitToPage = True

    report.merge_cells("A1:H2")
    report["A1"] = "REPORTE DE LISTADOS ARS ENVIADOS"
    report["A1"].font = Font(name="Arial", size=18, bold=True, color="FFFFFF")
    report["A1"].fill = PatternFill("solid", fgColor=BLUE)
    report["A1"].alignment = Alignment(horizontal="center", vertical="center")
    report.row_dimensions[1].height = 26
    if logo_path and os.path.isfile(logo_path):
        try:
            logo = ExcelImage(logo_path)
            logo.width, logo.height = 112, 46
            report.add_image(logo, "A1")
        except Exception:
            pass
    report["A4"] = "Fecha de generación"
    report["B4"] = datetime.now()
    report["B4"].number_format = "dd-mm-yyyy hh:mm"
    report["D4"] = "Usuario generador"
    report["E4"] = str(generated_by or "Sistema")
    for cell in (report["A4"], report["D4"]):
        cell.font = Font(bold=True, color=DARK_BLUE)

    cards = [
        ("A6", "TOTAL DE LISTADOS", summary["batch_count"]),
        ("C6", "TOTAL DE RECIBOS", summary["receipt_count"]),
        ("E6", "TOTAL GENERAL ENVIADO", summary["total"]),
    ]
    for anchor, label, value in cards:
        col = report[anchor].column
        report.merge_cells(start_row=6, start_column=col, end_row=6, end_column=col + 1)
        report.merge_cells(start_row=7, start_column=col, end_row=7, end_column=col + 1)
        report.cell(6, col, label)
        report.cell(7, col, value)
        for row_number in (6, 7):
            for column in range(col, col + 2):
                cell = report.cell(row_number, column)
                cell.fill = PatternFill("solid", fgColor=LIGHT_BLUE if row_number == 6 else "FFFFFF")
                cell.alignment = Alignment(horizontal="center", vertical="center")
        report.cell(6, col).font = Font(bold=True, color=DARK_BLUE)
        report.cell(7, col).font = Font(size=15, bold=True, color=GREEN)
        if label.startswith("TOTAL GENERAL"):
            report.cell(7, col).number_format = CURRENCY_FORMAT

    def write_summary_table(start_row: int, title: str, values: dict):
        report.merge_cells(start_row=start_row, start_column=1, end_row=start_row, end_column=4)
        report.cell(start_row, 1, title)
        report.cell(start_row, 1).font = Font(bold=True, color="FFFFFF")
        report.cell(start_row, 1).fill = PatternFill("solid", fgColor=DARK_BLUE)
        headers = ["Grupo", "Listados", "Recibos", "Total enviado"]
        for column, header in enumerate(headers, 1):
            cell = report.cell(start_row + 1, column, header)
            cell.font = Font(bold=True, color=DARK_BLUE)
            cell.fill = PatternFill("solid", fgColor=LIGHT_GRAY)
        cursor = start_row + 2
        for label, values_row in values.items():
            report.cell(cursor, 1, label)
            report.cell(cursor, 2, values_row["batches"])
            report.cell(cursor, 3, values_row["receipts"])
            report.cell(cursor, 4, values_row["total"])
            report.cell(cursor, 4).number_format = CURRENCY_FORMAT
            for column in range(1, 5):
                report.cell(cursor, column).border = border
            cursor += 1
        return cursor + 1

    next_row = write_summary_table(10, "SUBTOTALES POR ARS", summary["by_ars"])
    write_summary_table(next_row, "SUBTOTALES POR PERÍODO", summary["by_period"])
    for column, width in {"A": 28, "B": 16, "C": 16, "D": 20, "E": 18, "F": 18, "G": 16, "H": 16}.items():
        report.column_dimensions[column].width = width

    detail.merge_cells("A1:J2")
    detail["A1"] = "DETALLE DE LISTADOS ARS ENVIADOS"
    detail["A1"].font = Font(name="Arial", size=17, bold=True, color="FFFFFF")
    detail["A1"].fill = PatternFill("solid", fgColor=BLUE)
    detail["A1"].alignment = Alignment(horizontal="center", vertical="center")
    detail.merge_cells("A4:B4")
    detail["A4"] = f"Fecha de generación: {datetime.now():%d-%m-%Y %H:%M}"
    detail.merge_cells("D4:E4")
    detail["D4"] = f"Usuario generador: {generated_by or 'Sistema'}"
    for cell in (detail["A4"], detail["D4"]):
        cell.font = Font(bold=True, color=DARK_BLUE)
    headers = [
        "Período", "ARS", "Listado", "Versión", "Nº factura", "NCF",
        "Fecha de envío", "Confirmado por", "Recibos", "Total enviado",
    ]
    header_row = 7
    for column, header in enumerate(headers, 1):
        cell = detail.cell(header_row, column, header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor=DARK_BLUE)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for offset, row in enumerate(rows, header_row + 1):
        values = [
            _period_label(row), row.get("ars") or "", int(row["id"]),
            int(row.get("version") or 1), row.get("invoice_number") or "",
            row.get("ncf") or "", _display_datetime(row.get("sent_at")),
            row.get("sent_by") or "", int(row.get("receipt_count") or 0),
            float(row.get("sent_total") or 0),
        ]
        for column, value in enumerate(values, 1):
            cell = detail.cell(offset, column, value)
            cell.border = border
            cell.alignment = Alignment(
                horizontal="right" if column in (9, 10) else "left",
                vertical="center",
            )
        detail.cell(offset, 10).number_format = CURRENCY_FORMAT
    total_row = header_row + 1 + len(rows)
    detail.cell(total_row, 8, "TOTAL")
    detail.cell(total_row, 9, f"=SUM(I{header_row + 1}:I{total_row - 1})")
    detail.cell(total_row, 10, f"=SUM(J{header_row + 1}:J{total_row - 1})")
    for column in range(8, 11):
        cell = detail.cell(total_row, column)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor=GREEN)
    detail.cell(total_row, 10).number_format = CURRENCY_FORMAT
    detail.auto_filter.ref = f"A{header_row}:J{total_row - 1}"
    widths = [12, 25, 11, 10, 16, 18, 20, 20, 12, 20]
    for column, width in enumerate(widths, 1):
        detail.column_dimensions[get_column_letter(column)].width = width
    detail.print_area = f"A1:J{total_row}"
    report.print_area = f"A1:H{report.max_row}"
    workbook.save(output_path)
    return output_path


def create_sent_batches_pdf(
    base_path: str, rows: list[dict], generated_by: str, logo_path: str | None = None,
) -> str:
    if not rows:
        raise ValueError("No hay listados ENVIADOS para generar el reporte.")
    output_path = _safe_path(base_path, ".pdf")
    summary = _report_summary(rows)
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        "BatchTitle", parent=styles["Title"], fontName="Helvetica-Bold",
        fontSize=17, leading=20, textColor=colors.HexColor("#123A63"),
        alignment=TA_CENTER,
    )
    small = ParagraphStyle(
        "BatchSmall", parent=styles["BodyText"], fontName="Helvetica",
        fontSize=8, leading=10, textColor=colors.HexColor("#334155"),
    )
    right = ParagraphStyle("BatchRight", parent=small, alignment=TA_RIGHT)
    doc = SimpleDocTemplate(
        output_path, pagesize=landscape(letter), rightMargin=0.38 * inch,
        leftMargin=0.38 * inch, topMargin=0.35 * inch, bottomMargin=0.42 * inch,
        title="Reporte de listados ARS enviados",
    )
    story = []
    header = []
    if logo_path and os.path.isfile(logo_path):
        try:
            header.append(Image(logo_path, width=1.35 * inch, height=0.55 * inch))
        except Exception:
            header.append("")
    else:
        header.append("")
    header.append(Paragraph("REPORTE DE LISTADOS ARS ENVIADOS", title_style))
    header.append(Paragraph(
        f"Generado: {datetime.now():%d-%m-%Y %H:%M}<br/>Usuario: {generated_by or 'Sistema'}",
        right,
    ))
    header_table = Table([header], colWidths=[1.5 * inch, 6.5 * inch, 2.1 * inch])
    header_table.setStyle(TableStyle([("VALIGN", (0, 0), (-1, -1), "MIDDLE")]))
    story.extend([header_table, Spacer(1, 0.15 * inch)])
    cards = Table(
        [["TOTAL DE LISTADOS", "TOTAL DE RECIBOS", "TOTAL GENERAL ENVIADO"],
         [summary["batch_count"], summary["receipt_count"], f"RD$ {summary['total']:,.2f}"]],
        colWidths=[3.25 * inch] * 3,
    )
    cards.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#EAF2FB")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.HexColor("#123A63")),
        ("TEXTCOLOR", (0, 1), (-1, 1), colors.HexColor("#198754")),
        ("FONTNAME", (0, 0), (-1, -1), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, 0), 8), ("FONTSIZE", (0, 1), (-1, 1), 14),
        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
        ("BOX", (0, 0), (-1, -1), 0.6, colors.HexColor("#B6C8DB")),
        ("INNERGRID", (0, 0), (-1, -1), 0.3, colors.HexColor("#D5E1EC")),
        ("TOPPADDING", (0, 0), (-1, -1), 6), ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
    ]))
    story.extend([cards, Spacer(1, 0.16 * inch)])
    detail_data = [[
        "Período", "ARS", "Listado", "Versión", "Nº factura", "NCF",
        "Fecha de envío", "Confirmado por", "Recibos", "Total enviado",
    ]]
    for row in rows:
        detail_data.append([
            _period_label(row), Paragraph(str(row.get("ars") or ""), small),
            str(row["id"]), str(row.get("version") or 1),
            Paragraph(str(row.get("invoice_number") or ""), small),
            Paragraph(str(row.get("ncf") or ""), small),
            _display_datetime(row.get("sent_at")),
            Paragraph(str(row.get("sent_by") or ""), small),
            str(row.get("receipt_count") or 0),
            f"RD$ {float(row.get('sent_total') or 0):,.2f}",
        ])
    detail_data.append([
        "", "", "", "", "", "", "", "TOTAL",
        str(summary["receipt_count"]), f"RD$ {summary['total']:,.2f}",
    ])
    detail = LongTable(
        detail_data, repeatRows=1,
        colWidths=[0.65 * inch, 1.45 * inch, 0.55 * inch, 0.5 * inch,
                   1.0 * inch, 1.15 * inch, 1.05 * inch, 1.1 * inch,
                   0.55 * inch, 1.05 * inch],
    )
    detail.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#123A63")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 7.2),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("ALIGN", (2, 1), (3, -1), "CENTER"),
        ("ALIGN", (8, 1), (9, -1), "RIGHT"),
        ("GRID", (0, 0), (-1, -2), 0.3, colors.HexColor("#CBD5E1")),
        ("LEFTPADDING", (0, 0), (-1, -1), 4),
        ("RIGHTPADDING", (0, 0), (-1, -1), 4),
        ("BACKGROUND", (7, -1), (-1, -1), colors.HexColor("#E8F5EE")),
        ("FONTNAME", (7, -1), (-1, -1), "Helvetica-Bold"),
        ("TOPPADDING", (0, 0), (-1, -1), 4),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
    ]))
    story.extend([detail, PageBreak()])

    def subtotal_table(title: str, values: dict):
        data = [[title, "Listados", "Recibos", "Total enviado"]]
        for label, item in values.items():
            data.append([
                label, item["batches"], item["receipts"], f"RD$ {item['total']:,.2f}"
            ])
        table = Table(data, repeatRows=1, colWidths=[3.8 * inch, 1.2 * inch, 1.2 * inch, 2 * inch])
        table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#123A63")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("ALIGN", (1, 1), (-1, -1), "RIGHT"),
            ("GRID", (0, 0), (-1, -1), 0.3, colors.HexColor("#CBD5E1")),
            ("FONTSIZE", (0, 0), (-1, -1), 8),
            ("TOPPADDING", (0, 0), (-1, -1), 5),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
        ]))
        return table

    story.append(Paragraph("RESUMEN DE ENVÍOS", title_style))
    story.append(Spacer(1, 0.12 * inch))
    story.append(subtotal_table("SUBTOTALES POR ARS", summary["by_ars"]))
    story.append(Spacer(1, 0.2 * inch))
    story.append(subtotal_table("SUBTOTALES POR PERÍODO", summary["by_period"]))

    def footer(canvas, document):
        canvas.saveState()
        canvas.setFont("Helvetica", 7)
        canvas.setFillColor(colors.HexColor("#66788A"))
        canvas.drawString(0.4 * inch, 0.2 * inch, "Hospital Provincial Dr. Ángel Contreras Mejía")
        canvas.drawRightString(10.6 * inch, 0.2 * inch, f"Página {document.page}")
        canvas.restoreState()

    doc.build(story, onFirstPage=footer, onLaterPages=footer)
    return output_path


def create_sent_batches_reports(
    base_path: str, rows: list[dict], generated_by: str, logo_path: str | None = None,
) -> tuple[str, str]:
    pdf_path = create_sent_batches_pdf(base_path, rows, generated_by, logo_path)
    try:
        xlsx_path = create_sent_batches_xlsx(base_path, rows, generated_by, logo_path)
    except Exception:
        try:
            os.remove(pdf_path)
        except OSError:
            pass
        raise
    return pdf_path, xlsx_path
