from __future__ import annotations

from datetime import date, datetime, timezone
from pathlib import Path
from types import SimpleNamespace
from unittest.mock import Mock

import pytest
from openpyxl import load_workbook
from PyPDF2 import PdfReader

from admission_statistical_reports import (
    ARS_ALL,
    ARS_EXCLUDE,
    ARS_INCLUDE,
    COVERAGE_ALL,
    COVERAGE_INSURED,
    COVERAGE_UNINSURED,
    SPECIALTY_ALL,
    AdmissionReportFilters,
    ReportSnapshotStore,
    SnapshotStaleError,
    build_admission_report_dataset,
    build_operational_period,
    build_turn_operational_period,
    coerce_hospital_datetime,
    search_ars_catalog,
)
from admission_v15_adapter import ReportReadError, _HybridDatabaseProxy


SOURCE_ID = "44444444-4444-4444-8444-444444444444"


def _row(
    index: int,
    *,
    occurred_at: str = "2026-08-27T22:30:00-04:00",
    ars: str = "HUMANO",
    specialty: str = "GENERAL",
    source_id: str = SOURCE_ID,
    turn_id: int = 316,
    status: str = "ACTIVA",
    deleted: bool = False,
):
    return {
        "attention_id": index,
        "global_attention_id": f"00000000-0000-4000-8000-{index:012d}",
        "patient_name": f"PACIENTE {index:03d}",
        "service_date": occurred_at[:10],
        "service_time": occurred_at[11:19],
        "created_at_effective_utc": occurred_at,
        "canonical_ars": ars,
        "coverage_status": ("UNINSURED_DECLARED" if ars == "SIN SEGURO" else "INSURED"),
        "specialty": specialty,
        "service_type": "EMERGENCIA",
        "operational_source_id": source_id,
        "turn_id": turn_id,
        "source_status": status,
        "is_deleted": deleted,
    }


def _filters(**overrides):
    values = {
        "start_at": datetime(2026, 8, 27, 8, 0),
        "end_at": datetime(2026, 8, 28, 8, 0),
        "period_label": "Día operativo 27/08/2026",
        "turn_label": "Turno actual",
        "operational_source_id": SOURCE_ID,
        "turn_id": 316,
        "specialty": SPECIALTY_ALL,
        "coverage": COVERAGE_ALL,
        "ars_mode": ARS_ALL,
        "selected_ars": (),
    }
    values.update(overrides)
    return AdmissionReportFilters(**values)


def _ars_fixture():
    distribution = {
        "APS": 25,
        "HUMANO": 20,
        "FUTURO": 15,
        "SEMMA": 10,
        "OTRA": 30,
    }
    rows = []
    sequence = 1
    for ars, count in distribution.items():
        for _ in range(count):
            rows.append(_row(sequence, ars=ars))
            sequence += 1
    return rows


def test_operational_period_is_always_eight_am_to_eight_am():
    daily = build_operational_period("DIARIO", date(2026, 8, 27))
    custom = build_operational_period("RANGO", date(2026, 8, 27), date(2026, 8, 29))

    assert daily.start_at == datetime(2026, 8, 27, 8, 0)
    assert daily.end_at == datetime(2026, 8, 28, 8, 0)
    assert custom.start_at == datetime(2026, 8, 27, 8, 0)
    assert custom.end_at == datetime(2026, 8, 30, 8, 0)


@pytest.mark.parametrize(
    ("mode", "selected", "expected_start", "expected_end"),
    (
        (
            "SEMANAL",
            date(2026, 8, 27),
            datetime(2026, 8, 24, 8),
            datetime(2026, 8, 31, 8),
        ),
        (
            "MENSUAL",
            date(2026, 8, 27),
            datetime(2026, 8, 1, 8),
            datetime(2026, 9, 1, 8),
        ),
        (
            "MENSUAL",
            date(2026, 12, 27),
            datetime(2026, 12, 1, 8),
            datetime(2027, 1, 1, 8),
        ),
        ("ANUAL", date(2026, 8, 27), datetime(2026, 1, 1, 8), datetime(2027, 1, 1, 8)),
    ),
)
def test_operational_period_modes_cover_complete_operational_days(
    mode, selected, expected_start, expected_end
):
    period = build_operational_period(mode, selected)

    assert period.start_at == expected_start
    assert period.end_at == expected_end


def test_reversed_custom_range_is_normalized_without_losing_operational_limits():
    period = build_operational_period("RANGO", date(2026, 8, 29), date(2026, 8, 27))

    assert period.start_at == datetime(2026, 8, 27, 8)
    assert period.end_at == datetime(2026, 8, 30, 8)


def test_persisted_turn_period_handles_pre_eight_start_and_safe_fallback():
    early = build_turn_operational_period(
        "2026-08-28 04:15:00", fallback_date=date(2026, 1, 1)
    )
    fallback = build_turn_operational_period(
        "not-a-date", fallback_date=date(2026, 8, 26)
    )

    assert early.start_at == datetime(2026, 8, 27, 8)
    assert early.end_at == datetime(2026, 8, 28, 8)
    assert fallback.start_at == datetime(2026, 8, 26, 8)


def test_hospital_datetime_accepts_datetime_offsets_and_legacy_text():
    aware = datetime(2026, 8, 28, 12, tzinfo=timezone.utc)

    assert coerce_hospital_datetime(aware) == datetime(2026, 8, 28, 8)
    assert coerce_hospital_datetime("28/08/2026 07:59") == datetime(2026, 8, 28, 7, 59)
    assert coerce_hospital_datetime("") is None


@pytest.mark.parametrize(
    "overrides",
    (
        {"end_at": datetime(2026, 8, 27, 8)},
        {"ars_mode": "DESCONOCIDO"},
        {"coverage": "DESCONOCIDA"},
        {"ars_mode": ARS_INCLUDE, "selected_ars": ()},
    ),
)
def test_invalid_report_filters_fail_before_querying(overrides):
    with pytest.raises(ValueError):
        _filters(**overrides)


def test_turn_boundary_uses_half_open_operational_window():
    rows = [
        _row(1, occurred_at="2026-08-27T22:30:00-04:00"),
        _row(2, occurred_at="2026-08-28T00:30:00-04:00"),
        _row(3, occurred_at="2026-08-28T04:30:00-04:00"),
        _row(4, occurred_at="2026-08-28T07:59:00-04:00"),
        _row(5, occurred_at="2026-08-28T08:01:00-04:00", turn_id=317),
    ]

    dataset = build_admission_report_dataset(rows, _filters())

    assert [row["attention_id"] for row in dataset.records] == [1, 2, 3, 4]
    assert dataset.summary["total_patients"] == 4


def test_ars_include_and_exclude_use_the_same_dataset_for_every_total():
    rows = _ars_fixture()

    included = build_admission_report_dataset(
        rows,
        _filters(ars_mode=ARS_INCLUDE, selected_ars=("APS", "HUMANO")),
    )
    excluded = build_admission_report_dataset(
        rows,
        _filters(ars_mode=ARS_EXCLUDE, selected_ars=("APS", "HUMANO")),
    )

    assert included.summary["total_patients"] == 45 == len(included.records)
    assert excluded.summary["total_patients"] == 55 == len(excluded.records)
    assert sum(dict(included.summary["by_ars"]).values()) == 45
    assert sum(dict(excluded.summary["by_ars"]).values()) == 55


def test_cancelled_deleted_wrong_turn_and_wrong_source_are_excluded():
    rows = [
        _row(1),
        _row(2, status="ANULADA"),
        _row(3, deleted=True),
        _row(4, turn_id=315),
        _row(5, source_id="55555555-5555-4555-8555-555555555555"),
    ]

    dataset = build_admission_report_dataset(rows, _filters())

    assert [row["attention_id"] for row in dataset.records] == [1]


def test_specialty_and_coverage_filters_normalize_real_world_values():
    rows = [
        _row(1, specialty="MEDICINA GENERAL", ars="HUMANO"),
        _row(2, specialty="PEDIÁTRICA", ars="SIN SEGURO"),
        _row(3, specialty="GINECOLOGIA", ars="FUTURO"),
    ]

    insured = build_admission_report_dataset(rows, _filters(coverage=COVERAGE_INSURED))
    uninsured = build_admission_report_dataset(
        rows, _filters(coverage=COVERAGE_UNINSURED)
    )
    pediatric = build_admission_report_dataset(rows, _filters(specialty="PEDIATRIA"))

    assert insured.summary["total_patients"] == 2
    assert uninsured.summary["total_patients"] == 1
    assert [row["attention_id"] for row in pediatric.records] == [2]


def test_summary_cards_preview_and_records_share_one_result():
    rows = [
        _row(1, specialty="GENERAL", ars="HUMANO"),
        _row(2, specialty="PEDIATRIA", ars="SIN SEGURO"),
        _row(3, specialty="GINECOLOGIA", ars="FUTURO"),
        _row(4, specialty="ORTOPEDIA", ars="APS"),
    ]

    dataset = build_admission_report_dataset(rows, _filters())
    summary = dataset.summary

    assert summary["total_patients"] == len(dataset.records) == 4
    assert summary["insured_patients"] == 3
    assert summary["uninsured_patients"] == 1
    assert summary["insured_percentage"] == 75.0
    assert summary["uninsured_percentage"] == 25.0
    assert summary["general_patients"] == 1
    assert summary["pediatric_patients"] == 1
    assert summary["gynecology_patients"] == 1
    assert any(
        row[1] == "Total pacientes" and row[2] == 4 for row in dataset.preview_rows
    )


def test_snapshot_preserves_historical_representatives_in_chronological_order():
    turns = (
        {
            "turn_id": 316,
            "started_at": datetime(2026, 8, 27, 8),
            "ends_at": datetime(2026, 8, 28, 8),
            "representatives": (
                {
                    "user_id": "7",
                    "username": "auxiliar.uno",
                    "display_name": "Ana Pérez",
                    "event_at": datetime(2026, 8, 27, 8),
                },
                {
                    "user_id": "7",
                    "username": "auxiliar.uno",
                    "display_name": "Ana Pérez",
                    "event_at": datetime(2026, 8, 27, 9),
                },
                {
                    "user_id": "8",
                    "username": "auxiliar.dos",
                    "display_name": "Brenda Soto",
                    "event_at": datetime(2026, 8, 27, 14),
                },
            ),
        },
    )

    snapshot = build_admission_report_dataset(_ars_fixture(), _filters(), turns=turns)

    assert snapshot.representatives == ("Ana Pérez", "Brenda Soto")
    assert snapshot.summary["representantes"] == ("Ana Pérez", "Brenda Soto")
    assert snapshot.summary["representante"] == "Ana Pérez, Brenda Soto"
    assert snapshot.summary["representatives_by_turn"] == (
        ("Turno actual", ("Ana Pérez", "Brenda Soto")),
    )


def test_snapshot_store_rejects_exports_after_any_filter_change():
    empty_store = ReportSnapshotStore()
    assert empty_store.snapshot is None
    assert empty_store.invalidate_if_changed(_filters()) is False
    with pytest.raises(SnapshotStaleError, match="Primero genere"):
        empty_store.require_exportable()

    original = _filters()
    snapshot = build_admission_report_dataset(_ars_fixture(), original)
    store = ReportSnapshotStore()
    store.replace(snapshot)

    assert store.require_exportable(original) is snapshot

    changed = _filters(coverage=COVERAGE_INSURED)
    store.invalidate_if_changed(changed)

    with pytest.raises(SnapshotStaleError):
        store.require_exportable(changed)


def test_snapshot_exports_survive_database_loss_without_new_queries(tmp_path):
    from ADMISION_PYSIDE6_V15 import facturacion_tabs_pyside6 as v15

    snapshot = build_admission_report_dataset(_ars_fixture(), _filters())
    pdf_path = tmp_path / "snapshot.pdf"
    excel_path = tmp_path / "snapshot.xlsx"

    assert v15.crear_pdf_reporte(snapshot.summary, destino=str(pdf_path)) == str(pdf_path)
    assert v15.crear_excel_reporte_estadistico(
        snapshot.summary, destino=str(excel_path)
    ) == str(excel_path)
    assert pdf_path.exists()
    assert excel_path.exists()


def test_malformed_historical_row_does_not_fail_the_whole_report():
    rows = [_row(1), {"attention_id": 2, "patient_name": None}]

    dataset = build_admission_report_dataset(
        rows,
        _filters(operational_source_id="", turn_id=None),
    )

    assert dataset.summary["total_patients"] == 1
    assert dataset.diagnostics["INVALID_TIMESTAMP"] == 1


def test_all_turn_range_other_specialties_and_malformed_identity_are_safe():
    malformed_identity = _row(1, specialty="ORTOPEDIA")
    malformed_identity["attention_id"] = "invalid"
    malformed_identity["turn_id"] = "invalid"
    malformed_identity["patient_name"] = ""
    malformed_identity["service_type"] = "UNKNOWN"
    rows = [
        malformed_identity,
        _row(2, specialty="GENERAL"),
        _row(3, specialty="ORTOPEDIA", occurred_at="2026-08-28T08:00:00-04:00"),
        _row(4, specialty="ORTOPEDIA", deleted="true"),
    ]

    dataset = build_admission_report_dataset(
        rows,
        _filters(
            operational_source_id="",
            turn_id=None,
            specialty="OTRAS",
        ),
    )

    assert len(dataset.records) == 1
    assert dataset.records[0]["attention_id"] == 0
    assert dataset.records[0]["turn_id"] is None
    assert dataset.records[0]["patient_name"] == "SIN NOMBRE"
    assert dataset.records[0]["service_type"] == "EMERGENCIA"
    assert dataset.diagnostics == {
        "INCLUDED": 1,
        "WRONG_SPECIALTY": 1,
        "OUTSIDE_PERIOD": 1,
        "CANCELLED": 1,
    }


def test_ars_catalog_search_is_accent_and_case_insensitive():
    assert search_ars_catalog(["HUMANO", "FUTURO", "SENASA CONTRIBUTIVO"], "hum") == (
        "HUMANO",
    )
    assert search_ars_catalog(["SEGURO MÉDICO", "APS"], "medico") == ("SEGURO MÉDICO",)
    assert search_ars_catalog(["HUMANO", "", "APS", "HUMANO"], "") == (
        "APS",
        "HUMANO",
    )


class _Cursor:
    def __init__(self, rows):
        self.rows = rows

    def fetchall(self):
        return self.rows


class _CapturingConnection:
    def __init__(self, rows):
        self.rows = rows
        self.query = ""
        self.params = ()

    def __enter__(self):
        return self

    def __exit__(self, *_args):
        return False

    def execute(self, query, params=()):
        self.query = str(query)
        self.params = tuple(params)
        return _Cursor(self.rows)


class _ReportSourceConnection:
    def __init__(self):
        self.queries = []

    def __enter__(self):
        return self

    def __exit__(self, *_args):
        return False

    def execute(self, query, params=()):
        query_text = str(query)
        self.queries.append((query_text, tuple(params)))
        if "admission_operational_turn_intervals" in query_text:
            return _Cursor(
                [
                    {
                        "turn_id": 316,
                        "operational_source_id": SOURCE_ID,
                        "started_at": datetime(2026, 8, 27, 8),
                        "ends_at": datetime(2026, 8, 28, 8),
                        "status": "CURRENT",
                        "representatives": [
                            {
                                "user_id": "7",
                                "username": "auxiliar",
                                "display_name": "Ana Pérez",
                                "event_at": "2026-08-27T08:00:00",
                            }
                        ],
                    }
                ]
            )
        return _Cursor([_row(1)])


class _LocalDatabase:
    def obtener_atenciones_para_rango_real(self, *_args, **_kwargs):
        return []


def test_central_report_reader_uses_projection_and_operational_identity():
    connection = _CapturingConnection([_row(1)])
    runtime = SimpleNamespace(
        offline=False,
        logger=None,
        host=SimpleNamespace(connection_factory=lambda: connection),
        operational_session=SimpleNamespace(
            turn_id=316,
            operational_source_id=SOURCE_ID,
        ),
    )
    proxy = _HybridDatabaseProxy(_LocalDatabase(), runtime)

    rows = proxy.list_statistical_report_records(
        operational_source_id=SOURCE_ID,
        turn_id=316,
        start_at=datetime(2026, 8, 27, 8, 0),
        end_at=datetime(2026, 8, 28, 8, 0),
    )

    assert len(rows) == 1
    assert "FROM admission_attention_projection p" in connection.query
    assert "COALESCE(p.is_deleted,FALSE)=FALSE" in connection.query
    assert "p.operational_source_id::TEXT=%s" in connection.query
    assert "p.turn_id=%s" in connection.query
    assert connection.params[:2] == (SOURCE_ID, 316)
    assert "source_instance_id" not in connection.query.split("WHERE", 1)[1]


def test_report_source_uses_one_pooled_connection_for_turns_and_records():
    connection = _ReportSourceConnection()
    factory_calls = 0

    def connection_factory():
        nonlocal factory_calls
        factory_calls += 1
        return connection

    runtime = SimpleNamespace(
        offline=False,
        logger=None,
        host=SimpleNamespace(connection_factory=connection_factory),
        operational_session=SimpleNamespace(
            turn_id=316,
            operational_source_id=SOURCE_ID,
            turn_started_at=datetime(2026, 8, 27, 8),
            turn_ends_at=datetime(2026, 8, 28, 8),
        ),
        _is_temporary_connection_error=lambda exc: isinstance(exc, TimeoutError),
        _report_retry_delay_seconds=0,
    )
    proxy = _HybridDatabaseProxy(_LocalDatabase(), runtime)

    source = proxy.load_statistical_report_source(
        operational_source_id=SOURCE_ID,
        turn_scope="Turno actual",
        current_turn_id=316,
        start_at=datetime(2026, 8, 27, 8),
        end_at=datetime(2026, 8, 28, 8),
    )

    assert factory_calls == 1
    assert len(connection.queries) == 2
    assert source["turn_id"] == 316
    assert source["records"][0]["attention_id"] == 1
    assert source["turns"][0]["representatives"][0]["display_name"] == "Ana Pérez"


def test_transient_report_timeout_retries_once_then_succeeds():
    successful_connection = _ReportSourceConnection()
    calls = 0
    logger = Mock()

    def connection_factory():
        nonlocal calls
        calls += 1
        if calls == 1:
            raise TimeoutError("connection timed out")
        return successful_connection

    runtime = SimpleNamespace(
        offline=False,
        logger=logger,
        host=SimpleNamespace(connection_factory=connection_factory),
        operational_session=SimpleNamespace(
            turn_id=316,
            operational_source_id=SOURCE_ID,
            turn_started_at=datetime(2026, 8, 27, 8),
            turn_ends_at=datetime(2026, 8, 28, 8),
        ),
        _is_temporary_connection_error=lambda exc: isinstance(exc, TimeoutError),
        _report_retry_delay_seconds=0,
    )
    proxy = _HybridDatabaseProxy(_LocalDatabase(), runtime)

    source = proxy.load_statistical_report_source(
        operational_source_id=SOURCE_ID,
        turn_scope="Turno actual",
        current_turn_id=316,
        start_at=datetime(2026, 8, 27, 8),
        end_at=datetime(2026, 8, 28, 8),
    )

    assert calls == 2
    assert len(source["records"]) == 1
    logger.error.assert_called_once()


def test_non_transient_report_query_error_is_classified_without_retry():
    calls = 0

    def connection_factory():
        nonlocal calls
        calls += 1
        raise RuntimeError("column does_not_exist")

    runtime = SimpleNamespace(
        offline=False,
        logger=None,
        host=SimpleNamespace(connection_factory=connection_factory),
        operational_session=SimpleNamespace(turn_id=316),
        _is_temporary_connection_error=lambda _exc: False,
        _report_retry_delay_seconds=0,
    )
    proxy = _HybridDatabaseProxy(_LocalDatabase(), runtime)

    with pytest.raises(ReportReadError) as captured:
        proxy.load_statistical_report_source(
            operational_source_id=SOURCE_ID,
            turn_scope="Turno actual",
            current_turn_id=316,
            start_at=datetime(2026, 8, 27, 8),
            end_at=datetime(2026, 8, 28, 8),
        )

    assert calls == 1
    assert captured.value.code == "REPORT_QUERY_ERROR"


@pytest.mark.parametrize(
    ("error", "temporary", "expected"),
    (
        (TimeoutError("timed out"), True, "REPORT_CONNECTION_TIMEOUT"),
        (ConnectionError("connection refused"), True, "REPORT_DATABASE_UNAVAILABLE"),
        (ValueError("invalid uuid"), False, "REPORT_DATA_ERROR"),
        (ReportReadError("REPORT_DATA_ERROR", "safe"), False, "REPORT_DATA_ERROR"),
    ),
)
def test_report_read_errors_have_safe_operational_categories(
    error, temporary, expected
):
    runtime = SimpleNamespace(
        _is_temporary_connection_error=lambda _exc: temporary,
    )
    proxy = _HybridDatabaseProxy(_LocalDatabase(), runtime)

    assert proxy._classify_report_read_error(error).code == expected


def test_all_turn_source_uses_period_prefilter_and_has_no_selected_turn():
    connection = _ReportSourceConnection()
    runtime = SimpleNamespace(
        offline=False,
        logger=None,
        host=SimpleNamespace(connection_factory=lambda: connection),
        operational_session=SimpleNamespace(turn_id=316),
        _is_temporary_connection_error=lambda _exc: False,
        _report_retry_delay_seconds=0,
    )
    proxy = _HybridDatabaseProxy(_LocalDatabase(), runtime)

    source = proxy.load_statistical_report_source(
        operational_source_id=SOURCE_ID,
        turn_scope="Todos los turnos",
        current_turn_id=316,
        start_at=datetime(2026, 8, 1, 8),
        end_at=datetime(2026, 9, 1, 8),
    )

    assert source["turn_id"] is None
    assert source["selected_turn"] is None
    record_query, record_params = connection.queries[-1]
    assert "p.service_date BETWEEN %s AND %s" in record_query
    assert record_params == (SOURCE_ID, "2026-07-31", "2026-09-01")


def test_current_turn_fallback_and_missing_previous_are_explicit():
    runtime = SimpleNamespace(
        operational_session=SimpleNamespace(
            turn_started_at=datetime(2026, 8, 27, 8),
            turn_ends_at=datetime(2026, 8, 28, 8),
        )
    )
    proxy = _HybridDatabaseProxy(_LocalDatabase(), runtime)

    fallback, turns = proxy._ensure_statistical_report_turn(
        [],
        source_id=SOURCE_ID,
        turn_scope="Turno actual",
        current_turn_id=316,
    )

    assert fallback["turn_id"] == 316
    assert turns == [fallback]
    with pytest.raises(ReportReadError) as captured:
        proxy._ensure_statistical_report_turn(
            [],
            source_id=SOURCE_ID,
            turn_scope="Turno anterior",
            current_turn_id=316,
        )
    assert captured.value.code == "REPORT_DATA_ERROR"


def test_offline_snapshot_loader_uses_local_source_and_persisted_turn():
    requested = {}

    class Local:
        def obtener_atenciones_para_rango_real(self, start, end, **kwargs):
            requested.update({"start": start, "end": end, **kwargs})
            return [_row(1)]

    runtime = SimpleNamespace(
        offline=True,
        logger=Mock(),
        operational_session=SimpleNamespace(
            turn_id=316,
            operational_source_id=SOURCE_ID,
            turn_started_at=datetime(2026, 8, 27, 8),
            turn_ends_at=datetime(2026, 8, 28, 8),
        ),
    )
    proxy = _HybridDatabaseProxy(Local(), runtime)

    source = proxy.load_statistical_report_source(
        operational_source_id=SOURCE_ID,
        turn_scope="Turno actual",
        current_turn_id=316,
        start_at=datetime(2026, 8, 27, 8),
        end_at=datetime(2026, 8, 28, 8),
    )

    assert len(source["records"]) == 1
    assert source["turn_id"] == 316
    assert requested["operational_turn_id"] == 316
    runtime.logger.info.assert_called()


def test_snapshot_loader_rejects_missing_operational_source_before_io():
    proxy = _HybridDatabaseProxy(
        _LocalDatabase(), SimpleNamespace(offline=False, logger=None)
    )

    with pytest.raises(ValueError):
        proxy.load_statistical_report_source(
            operational_source_id="",
            turn_scope="Turno actual",
            current_turn_id=316,
            start_at=datetime(2026, 8, 27, 8),
            end_at=datetime(2026, 8, 28, 8),
        )


def test_offline_report_reader_uses_local_replica_with_same_turn_identity():
    requested = {}

    class Local:
        def obtener_atenciones_para_rango_real(self, start, end, **kwargs):
            requested.update({"start": start, "end": end, **kwargs})
            return [_row(1)]

    runtime = SimpleNamespace(
        offline=True,
        logger=None,
        operational_session=SimpleNamespace(
            turn_id=316,
            operational_source_id=SOURCE_ID,
        ),
    )
    proxy = _HybridDatabaseProxy(Local(), runtime)

    rows = proxy.list_statistical_report_records(
        operational_source_id=SOURCE_ID,
        turn_id=316,
        start_at=datetime(2026, 8, 27, 8, 0),
        end_at=datetime(2026, 8, 28, 8, 0),
    )

    assert len(rows) == 1
    assert requested["operational_turn_id"] == 316
    assert requested["operational_source_id"] == SOURCE_ID


def test_report_excel_has_official_listing_and_one_summary_sheet(tmp_path):
    from ADMISION_PYSIDE6_V15 import facturacion_tabs_pyside6 as v15

    dataset = build_admission_report_dataset(_ars_fixture(), _filters())
    destination = tmp_path / "reporte.xlsx"

    assert v15.crear_excel_reporte_estadistico(
        dataset.summary, destino=str(destination)
    ) == str(destination)

    workbook = load_workbook(destination, data_only=True)
    try:
        assert workbook.sheetnames == [
            "LISTADO DE PACIENTES",
            "RESUMEN ESTADÍSTICO",
        ]
        listing = workbook["LISTADO DE PACIENTES"]
        summary = workbook["RESUMEN ESTADÍSTICO"]
        assert listing["A1"].value == "ASISTENCIA DE PACIENTES A EMERGENCIA"
        assert [listing.cell(5, column).value for column in range(1, 5)] == [
            "NO.",
            "NOMBRE",
            "ESPECIALIDAD",
            "ARS",
        ]
        assert listing.max_row - 5 == 100
        summary_values = {
            summary.cell(row, 1).value: summary.cell(row, 2).value
            for row in range(1, summary.max_row + 1)
        }
        assert summary_values["Total de pacientes"] == 100
        assert summary_values["Modo ARS"] == "TODAS"
        assert summary_values["Turno seleccionado"] == "Turno actual"
    finally:
        workbook.close()


def test_report_excel_include_filter_has_exactly_45_rows_and_summary_total(tmp_path):
    from ADMISION_PYSIDE6_V15 import facturacion_tabs_pyside6 as v15

    dataset = build_admission_report_dataset(
        _ars_fixture(),
        _filters(ars_mode=ARS_INCLUDE, selected_ars=("APS", "HUMANO")),
    )
    destination = tmp_path / "incluidas.xlsx"
    v15.crear_excel_reporte_estadistico(dataset.summary, destino=str(destination))

    workbook = load_workbook(destination, data_only=True)
    try:
        listing = workbook["LISTADO DE PACIENTES"]
        summary = workbook["RESUMEN ESTADÍSTICO"]
        values = {
            summary.cell(row, 1).value: summary.cell(row, 2).value
            for row in range(1, summary.max_row + 1)
        }
        assert listing.max_row - 5 == 45
        assert values["Total de pacientes"] == 45
        assert values["ARS seleccionadas"] == "APS, HUMANO"
    finally:
        workbook.close()


def test_report_outputs_count_each_attention_once_across_service_types(tmp_path):
    from ADMISION_PYSIDE6_V15 import facturacion_tabs_pyside6 as v15

    rows = [_row(1), _row(2), _row(3)]
    rows[0]["service_type"] = "EMERGENCIA"
    rows[1]["service_type"] = "URGENCIA"
    rows[2]["service_type"] = "CONSULTA"
    dataset = build_admission_report_dataset(rows, _filters())
    destination = tmp_path / "tipos.xlsx"

    assert v15.reportable_patient_count(dataset.summary) == 3
    v15.crear_excel_reporte_estadistico(dataset.summary, destino=str(destination))

    workbook = load_workbook(destination, data_only=True)
    try:
        listing = workbook["LISTADO DE PACIENTES"]
        summary = workbook["RESUMEN ESTADÍSTICO"]
        assert listing.max_row - 5 == 3
        assert summary["B2"].value == 3
    finally:
        workbook.close()


def test_report_pdf_uses_the_same_filtered_summary_and_declares_filters(tmp_path):
    from ADMISION_PYSIDE6_V15 import facturacion_tabs_pyside6 as v15

    dataset = build_admission_report_dataset(
        _ars_fixture(),
        _filters(ars_mode=ARS_EXCLUDE, selected_ars=("APS", "HUMANO")),
    )
    destination = tmp_path / "excluidas.pdf"

    v15.crear_pdf_reporte(dataset.summary, destino=str(destination))

    text = "\n".join(page.extract_text() or "" for page in PdfReader(destination).pages)
    assert "TOTAL PACIENTES" in text
    assert "55" in text
    assert "ARS EXCLUIR" in text
    assert "APS, HUMANO" in text
    assert "Total de pacientes del período: 55" in text


def test_outputs_show_friendly_turn_and_persisted_representatives_without_id(tmp_path):
    from ADMISION_PYSIDE6_V15 import facturacion_tabs_pyside6 as v15

    friendly_label = (
        "Turno actual · 27/08/2026 08:00 AM → 28/08/2026 08:00 AM"
    )
    turns = (
        {
            "turn_id": 316,
            "started_at": datetime(2026, 8, 27, 8),
            "representatives": (
                {
                    "user_id": "7",
                    "username": "auxiliar",
                    "display_name": "Ana Pérez",
                    "event_at": datetime(2026, 8, 27, 8),
                },
            ),
        },
    )
    dataset = build_admission_report_dataset(
        [_row(1)], _filters(turn_label=friendly_label), turns=turns
    )
    pdf_path = tmp_path / "representantes.pdf"
    excel_path = tmp_path / "representantes.xlsx"

    v15.crear_pdf_reporte(dataset.summary, destino=str(pdf_path))
    v15.crear_excel_reporte_estadistico(
        dataset.summary, destino=str(excel_path)
    )

    pdf_text = "\n".join(
        page.extract_text() or "" for page in PdfReader(pdf_path).pages
    )
    assert "ANA PÉREZ" in pdf_text
    assert "ID 316" not in pdf_text
    workbook = load_workbook(excel_path, data_only=True)
    try:
        listing = workbook["LISTADO DE PACIENTES"]
        summary = workbook["RESUMEN ESTADÍSTICO"]
        values = {
            summary.cell(row, 1).value: summary.cell(row, 2).value
            for row in range(1, summary.max_row + 1)
        }
        assert "ID 316" not in str(listing["A4"].value)
        assert values["Representante(s)"] == "Ana Pérez"
    finally:
        workbook.close()


def test_operational_and_report_excel_share_the_listing_builder_source():
    source = (
        Path(__file__).parents[1]
        / "ADMISION_PYSIDE6_V15"
        / "facturacion_tabs_pyside6.py"
    ).read_text(encoding="utf-8")

    operational = source[
        source.index("def _construir_workbook_turno") : source.index(
            "def reconstruir_excel_turno"
        )
    ]
    report = source[
        source.index("def crear_excel_reporte_estadistico") : source.index(
            "# -------------------------------\n# APP",
            source.index("def crear_excel_reporte_estadistico"),
        )
    ]

    assert "construir_hoja_listado_pacientes(" in operational
    assert "construir_hoja_listado_pacientes(" in report
