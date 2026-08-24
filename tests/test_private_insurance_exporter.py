import tempfile
import unittest
from pathlib import Path

from openpyxl import load_workbook

from private_insurance_exporter import (
    create_private_ars_workbook,
    safe_export_filename,
    validate_export_payload,
)


class PrivateInsuranceExporterTests(unittest.TestCase):
    def setUp(self):
        self.batch = {
            "id": 7,
            "version": 2,
            "period_year": 2026,
            "period_month": 6,
            "ars": "FUTURO",
            "ars_display_name": "FUTURO",
            "invoice_date": "2026-07-07",
            "invoice_number": "505",
            "ncf": "B0100000505",
            "ncf_expiration_date": "2026-12-31",
            "provider_code": "14427",
            "provider_name": "HOSPITAL PROVINCIAL DR. ÁNGEL CONTRERAS MEJÍA",
            "provider_rnc": "430-130-516",
            "ars_rnc": "101-55754-2",
            "ars_address": "C. JUAN SÁNCHEZ RAMÍREZ 19, SANTO DOMINGO",
            "director_name": "DRA. CATALINA FABIÁN",
            "director_title": "DIRECTORA",
            "service_description": "EMERGENCIA",
            "specialty_default": "EMERGENCIOLOGÍA",
            "discount": 0,
            "itbis": 0,
        }
        self.receipts = [
            {
                "recibo_id": 1,
                "numero": 100,
                "patient_snapshot": "PACIENTE UNO",
                "document_type_snapshot": "NSS",
                "document_number_snapshot": "123456789",
                "service_date_snapshot": "2026-06-03",
                "authorization_snapshot": "700579",
                "specialty_snapshot": "EMERGENCIOLOGÍA",
                "total_snapshot": 373.75,
            },
            {
                "recibo_id": 2,
                "numero": 101,
                "patient_snapshot": "PACIENTE DOS",
                "document_type_snapshot": "CÉDULA",
                "document_number_snapshot": "00100000011",
                "service_date_snapshot": "2026-06-04",
                "authorization_snapshot": "700580",
                "specialty_snapshot": "PEDIATRÍA",
                "total_snapshot": 1250.62,
            },
        ]

    def test_validation_reports_missing_fiscal_and_patient_fields(self):
        batch = dict(self.batch)
        batch["ncf"] = ""
        rows = [dict(self.receipts[0], document_number_snapshot="")]
        problems = validate_export_payload(batch, rows)
        self.assertTrue(any("NCF" in problem for problem in problems))
        self.assertTrue(any("NSS o cédula" in problem for problem in problems))

    def test_workbook_contains_linked_relation_and_global_invoice(self):
        with tempfile.TemporaryDirectory() as temp:
            output = Path(temp) / "futuro.xlsx"
            create_private_ars_workbook(output, self.batch, self.receipts)
            workbook = load_workbook(output, data_only=False)
            try:
                self.assertEqual(
                    workbook.sheetnames,
                    ["Relación Emergencias", "Factura Global"],
                )
                relation = workbook["Relación Emergencias"]
                invoice = workbook["Factura Global"]
                self.assertEqual(relation["A15"].value, 1)
                self.assertEqual(relation["B16"].value, "CÉDULA\n00100000011")
                self.assertEqual(relation["E15"].value, "700579")
                self.assertIsNone(relation["A17"].value)
                self.assertEqual(relation["E18"].value, "TOTAL")
                self.assertEqual(relation["F18"].value, "=SUM(F15:F16)")
                self.assertNotIn(
                    "EXPEDIENTE COMPLETO",
                    [cell.value for row in relation for cell in row],
                )
                self.assertEqual(invoice["A17"].value, 2)
                self.assertEqual(
                    invoice["C17"].value,
                    "='Relación Emergencias'!F18",
                )
                self.assertEqual(invoice["C22"].value, "=C19-C20+C21")
                self.assertEqual(
                    invoice["D10"].value,
                    "FECHA VENCIMIENTO NCF: 31-12-2026",
                )
                self.assertEqual(
                    relation.sheet_properties.pageSetUpPr.fitToPage,
                    True,
                )
                self.assertEqual(invoice.page_setup.orientation, "portrait")
            finally:
                workbook.close()

    def test_filename_is_safe_and_period_specific(self):
        self.assertEqual(
            safe_export_filename("ARS Futuro / Privado", 2026, 6),
            "LISTADO_EMERGENCIAS_ARS_FUTURO_PRIVADO_2026_06.xlsx",
        )


if __name__ == "__main__":
    unittest.main()
