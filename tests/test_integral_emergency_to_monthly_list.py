import os
import sqlite3
import tempfile
import unittest
import uuid
from datetime import datetime
from pathlib import Path
from unittest.mock import patch

import psycopg2
from openpyxl import Workbook, load_workbook

import CALCULOS_QT as app
from admission_bridge import AdmissionReadOnlyRepository
from admission_contract import READINESS_INCOMPLETE, READINESS_READY
from private_insurance_exporter import create_private_ars_workbook


def build_integral_admission_database(path: Path) -> None:
    connection = sqlite3.connect(path)
    connection.executescript(
        """
        CREATE TABLE schema_version(
            id INTEGER PRIMARY KEY,
            version INTEGER NOT NULL
        );
        INSERT INTO schema_version VALUES(1,14);
        CREATE TABLE app_metadata(
            clave TEXT PRIMARY KEY,
            valor TEXT NOT NULL,
            updated_at TEXT
        );
        CREATE TABLE turnos(
            id INTEGER PRIMARY KEY,
            fecha_inicio TEXT,
            fecha_inicio_real TEXT,
            estado TEXT
        );
        INSERT INTO turnos VALUES(
            77,'2026-07-20 08:00:00','2026-07-20 08:00:00','ABIERTO'
        );
        INSERT INTO app_metadata(clave,valor)
        VALUES('integration.source_instance_id','source-integral-e2e');
        CREATE TABLE atenciones(
            id INTEGER PRIMARY KEY,
            paciente_id INTEGER NOT NULL,
            nombre TEXT NOT NULL,
            fecha TEXT,
            hora TEXT,
            nss TEXT,
            nss_clean TEXT,
            cedula TEXT,
            cedula_clean TEXT,
            ars TEXT,
            tipo_atencion TEXT,
            estado TEXT,
            identidad_estado TEXT,
            requiere_revision INTEGER,
            created_at TEXT,
            updated_at TEXT,
            turno_id INTEGER
        );
        CREATE TABLE paciente_identificadores(
            id INTEGER PRIMARY KEY,
            paciente_id INTEGER NOT NULL,
            tipo TEXT,
            valor_normalizado TEXT,
            activo INTEGER,
            conflicto INTEGER
        );
        """
    )
    connection.executemany(
        """INSERT INTO atenciones VALUES(
               ?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?
           )""",
        [
            (
                1001, 501, "PACIENTE INTEGRAL", "2026-07-20", "09:00",
                "123-456-789", "123456789", "001-0000001-1",
                "00100000011", "HUMANO", "EMERGENCIA", "ACTIVA",
                "VALIDADA", 0, "2026-07-20 09:00:00", None, 77,
            ),
            (
                1002, 502, "PACIENTE INCOMPLETO", "2026-07-20", "09:10",
                "", "", "001-0000002-2", "00100000022", "HUMANO",
                "EMERGENCIA", "ACTIVA", "VALIDADA", 0,
                "2026-07-20 09:10:00", None, 77,
            ),
            (
                1003, 503, "PACIENTE URGENCIA", "2026-07-20", "09:20",
                "987654321", "987654321", "001-0000003-3",
                "00100000033", "HUMANO", "URGENCIA", "ACTIVA",
                "VALIDADA", 0, "2026-07-20 09:20:00", None, 77,
            ),
        ],
    )
    connection.commit()
    connection.close()


class IntegralEmergencyToMonthlyListTests(unittest.TestCase):
    def setUp(self):
        self.temp_dir = tempfile.TemporaryDirectory()
        self.sqlite_path = Path(self.temp_dir.name) / "pacientes.db"
        build_integral_admission_database(self.sqlite_path)

        self.admin_url = os.environ.get(
            "HOSPITAL_E2E_ADMIN_URL",
            "postgresql://preview_admin@127.0.0.1:55432/postgres",
        )
        self.database_name = "hospital_e2e_" + uuid.uuid4().hex[:12]
        try:
            admin = psycopg2.connect(self.admin_url)
        except psycopg2.Error as exc:
            self.temp_dir.cleanup()
            self.skipTest(f"PostgreSQL local para prueba integral no disponible: {exc}")
        admin.autocommit = True
        with admin.cursor() as cursor:
            cursor.execute(f'CREATE DATABASE "{self.database_name}"')
        admin.close()

        self.database_url = (
            "postgresql://preview_admin@127.0.0.1:55432/"
            + self.database_name
        )
        self.original_url = app.DB_URL
        self.original_pool = app.db_pool
        app.DB_URL = self.database_url
        app.db_pool = None
        app.db_init()

    def tearDown(self):
        if app.db_pool is not None:
            app.db_pool.closeall()
        app.db_pool = self.original_pool
        app.DB_URL = self.original_url
        try:
            admin = psycopg2.connect(self.admin_url)
            admin.autocommit = True
            with admin.cursor() as cursor:
                cursor.execute(
                    "SELECT pg_terminate_backend(pid) FROM pg_stat_activity "
                    "WHERE datname=%s",
                    (self.database_name,),
                )
                cursor.execute(f'DROP DATABASE IF EXISTS "{self.database_name}"')
            admin.close()
        finally:
            self.temp_dir.cleanup()

    def test_emergency_to_audit_to_monthly_ars_list(self):
        repository = AdmissionReadOnlyRepository(self.sqlite_path)

        reconciliation = app.load_admission_reconciliation(
            "2026-07-20",
            "2026-07-20",
            repository=repository,
        )

        self.assertEqual(reconciliation["summary"]["admissions"], 2)
        self.assertEqual(reconciliation["summary"]["ready"], 1)
        self.assertEqual(reconciliation["summary"]["incomplete"], 1)
        self.assertNotIn(
            1003,
            [record["attention_id"] for record in reconciliation["records"]],
        )
        ready = repository.get_billable_attention(1001)
        self.assertIsNotNone(ready)
        incomplete = repository.get_eligible_attention(1002)
        self.assertEqual(incomplete.billing_readiness, READINESS_INCOMPLETE)
        self.assertIsNone(repository.get_billable_attention(1002))

        billing_session_id = "integral-emergency-session"
        ready = app.claim_projected_billable_attention(
            ready.attention_id,
            ready.source_instance_id,
            username="admin",
            session_id=billing_session_id,
        )
        self.assertIsNotNone(ready)
        number = app.get_next_recibo_number()
        receipt_id = app.save_receipt_with_items(
            None,
            number,
            ready.name,
            ready.service_date,
            "DX PRUEBA INTEGRAL",
            ready.canonical_ars,
            100,
            300,
            "recibo_integral.pdf",
            "admin",
            0,
            app.now_str(),
            [
                (
                    "Medicamentos",
                    [
                        (
                            "MEDICAMENTO DE PRUEBA",
                            200,
                            1,
                            200,
                            "Medicamentos",
                        )
                    ],
                )
            ],
            "ASEGURADO",
            "AUT-E2E-001",
            ready,
            billing_session_id,
        )

        with app.db_connect() as connection:
            connection.execute(
                """INSERT INTO pdf_storage(filename,file_data)
                   VALUES(%s,%s)
                   ON CONFLICT(filename) DO UPDATE
                   SET file_data=EXCLUDED.file_data""",
                ("recibo_integral.pdf", b"%PDF-1.4 integral"),
            )
            connection.execute(
                "UPDATE recibos SET pdf_synced=1 WHERE id=%s",
                (receipt_id,),
            )

        actor = {"username": "admin", "role": app.ROLE_ADMIN}
        with patch.object(
            app,
            "AdmissionReadOnlyRepository",
            return_value=repository,
        ):
            result = app.change_receipt_billing_status(
                receipt_id,
                app.BILLING_INVOICED,
                actor,
                reference="PRUEBA-INTEGRAL",
                checklist={
                    key: True for key in app.AUDIT_CHECKLIST_ITEMS
                },
            )
        self.assertEqual(result["estado_facturacion"], app.BILLING_INVOICED)

        with app.db_connect() as connection:
            connection.execute(
                "UPDATE recibos SET specialty_snapshot=NULL WHERE id=%s",
                (receipt_id,),
            )

        billed_at = datetime.fromisoformat(
            str(result["estado_facturacion_at"])[:19]
        )
        batch_id = app.create_monthly_billing_batch(
            billed_at.year,
            billed_at.month,
            "HUMANO",
            actor,
            notes="Prueba automatizada integral",
        )
        candidates = app.obtener_candidatos_listado_ars(
            batch_id,
            date_from="2026-07-01",
            date_to="2026-07-31",
        )
        self.assertEqual(
            {int(row["recibo_id"]) for row in candidates},
            {receipt_id},
        )
        app.add_receipts_to_monthly_batch(
            batch_id,
            [receipt_id],
            actor,
            date_from="2026-07-01",
            date_to="2026-07-31",
        )
        rows = app.list_monthly_batch_receipts(batch_id)

        self.assertEqual(len(rows), 1)
        self.assertEqual(int(rows[0]["recibo_id"]), receipt_id)
        self.assertEqual(rows[0]["ars_snapshot"], "HUMANO")
        self.assertEqual(float(rows[0]["total_snapshot"]), 300.0)
        self.assertEqual(rows[0]["document_type_snapshot"], "NSS")
        self.assertEqual(rows[0]["document_number_snapshot"], "123456789")
        self.assertEqual(rows[0]["authorization_snapshot"], "AUT-E2E-001")
        self.assertEqual(rows[0]["service_date_snapshot"], "2026-07-20")
        self.assertIsNone(rows[0]["specialty_snapshot"])

        app.update_monthly_batch_configuration(
            batch_id,
            {
                "ars_display_name": "HUMANO",
                "invoice_number": "FAC-E2E-001",
                "ncf": "B0100000999",
                "invoice_date": "2026-07-31",
                "ncf_expiration_date": "2026-12-31",
                "provider_code": "14427",
                "provider_name": "HOSPITAL PROVINCIAL DR. ÁNGEL CONTRERAS MEJÍA",
                "provider_rnc": "430-130-516",
                "ars_rnc": "101-00000-1",
                "ars_address": "DIRECCIÓN DE PRUEBA",
                "director_name": "DRA. DIRECTORA DE PRUEBA",
                "director_title": "DIRECTORA",
                "service_description": "EMERGENCIA",
                "specialty_default": "EMERGENCIOLOGÍA",
                "discount": 0,
                "itbis": 0,
            },
            actor,
            save_as_profile=True,
        )
        app.update_monthly_batch_receipt_export_data(
            batch_id,
            receipt_id,
            document_type="CÉDULA",
            document_number="00100000011",
            authorization="AUT-CORREGIDA-001",
            specialty="EMERGENCIOLOGÍA",
            user=actor,
        )
        self.assertEqual(app.monthly_batch_export_problems(batch_id), [])

        export_path = Path(self.temp_dir.name) / "expediente_integral.xlsx"
        batch = app.get_monthly_billing_batch(batch_id)
        rows = app.list_monthly_batch_receipts(batch_id)
        create_private_ars_workbook(
            export_path,
            batch,
            rows,
            logo_path=app.LOGO_PATH,
        )
        app.record_monthly_batch_export(batch_id, str(export_path), actor)
        self.assertEqual(
            app.get_monthly_billing_batch(batch_id)["status"],
            app.BATCH_PENDING,
        )
        with app.db_connect() as connection:
            connection.execute(
                "UPDATE billing_batches SET invoice_number='FAC-CAMBIADA' WHERE id=%s",
                (batch_id,),
            )
        with self.assertRaisesRegex(ValueError, "cambió después"):
            app.confirm_monthly_batch_sent(batch_id, actor)
        with app.db_connect() as connection:
            connection.execute(
                "UPDATE billing_batches SET invoice_number='FAC-E2E-001' WHERE id=%s",
                (batch_id,),
            )
        app.record_monthly_batch_export(batch_id, str(export_path), actor)
        sent = app.confirm_monthly_batch_sent(batch_id, actor)
        self.assertEqual(sent["sent_by"], "admin")
        self.assertEqual(
            app.get_monthly_billing_batch(batch_id)["status"],
            app.BATCH_SENT,
        )
        frozen = app.get_monthly_billing_batch(batch_id)
        self.assertEqual(int(frozen["sent_receipt_count"]), 1)
        self.assertEqual(float(frozen["sent_total"]), 300.0)
        self.assertEqual(frozen["sent_invoice_number"], "FAC-E2E-001")
        self.assertEqual(frozen["sent_ncf"], "B0100000999")
        with app.db_connect() as connection:
            connection.execute("UPDATE recibos SET total=9999 WHERE id=%s", (receipt_id,))
        report_rows = app.list_sent_billing_batches_report(batch_ids=[batch_id])
        self.assertEqual(len(report_rows), 1)
        self.assertEqual(float(report_rows[0]["sent_total"]), 300.0)
        self.assertTrue(export_path.is_file())
        workbook = load_workbook(export_path, data_only=False)
        try:
            self.assertEqual(
                workbook.sheetnames,
                ["Relación Emergencias", "Factura Global"],
            )
            self.assertEqual(
                workbook["Relación Emergencias"]["B15"].value,
                "CÉDULA\n00100000011",
            )
            self.assertEqual(
                workbook["Relación Emergencias"]["E15"].value,
                "AUT-CORREGIDA-001",
            )
            self.assertTrue(
                str(workbook["Factura Global"]["C19"].value).startswith(
                    "='Relación Emergencias'!"
                )
            )
            self.assertGreaterEqual(
                len(workbook["Relación Emergencias"]._images), 1
            )
            self.assertGreaterEqual(
                len(workbook["Factura Global"]._images), 1
            )
        finally:
            workbook.close()

        with app.db_connect() as connection:
            projection = connection.execute(
                """SELECT attention_id,readiness,coverage_status,
                          source_instance_id
                   FROM admission_attention_projection
                   ORDER BY attention_id"""
            ).fetchall()
            linked = connection.execute(
                """SELECT admission_source_instance_id,
                          admission_snapshot_hash,
                          admission_readiness
                   FROM recibos WHERE id=%s""",
                (receipt_id,),
            ).fetchone()

        self.assertEqual(
            [str(row["readiness"]) for row in projection],
            [READINESS_READY, READINESS_INCOMPLETE],
        )
        self.assertEqual(
            linked["admission_source_instance_id"],
            "source-integral-e2e",
        )
        self.assertEqual(linked["admission_readiness"], READINESS_READY)
        self.assertEqual(len(linked["admission_snapshot_hash"]), 64)

        replacement_batch_id = app.create_monthly_billing_batch(
            billed_at.year,
            billed_at.month,
            "HUMANO",
            actor,
            notes="No debe duplicar expedientes enviados",
        )
        self.assertEqual(app.list_monthly_batch_receipts(replacement_batch_id), [])
        available_ids = {
            int(row["recibo_id"])
            for row in app.list_available_receipts_for_batch(
                replacement_batch_id, date_from=None, date_to=None
            )
        }
        self.assertNotIn(receipt_id, available_ids)

    def test_schema_initialization_and_indexes_are_idempotent(self):
        app.db_init()
        app.db_init()
        with app.db_connect() as connection:
            indexes = {
                row["indexname"]
                for row in connection.execute(
                    """SELECT indexname FROM pg_indexes
                       WHERE schemaname=current_schema()
                         AND indexname=ANY(%s)""",
                    ([
                        "idx_recibos_ars_service_billing",
                        "idx_recibos_admission_attention_lookup",
                        "idx_billing_batches_status_ars_period",
                        "idx_billing_batch_receipts_receipt_active",
                    ],),
                ).fetchall()
            }
        self.assertEqual(
            indexes,
            {
                "idx_recibos_ars_service_billing",
                "idx_recibos_admission_attention_lookup",
                "idx_billing_batches_status_ars_period",
                "idx_billing_batch_receipts_receipt_active",
            },
        )

    def test_bulk_add_filters_missing_specialty_and_sent_exclusion(self):
        actor = {"username": "admin", "role": app.ROLE_ADMIN}
        batch_id = app.create_monthly_billing_batch(2026, 7, "FUTURO", actor)
        with app.db_connect() as connection:
            receipt_ids = []
            for number, service_date, ars in (
                (880001, "2026-07-20", "FUTURO"),
                (880002, "2026-07-21", "FUTURO"),
                (880003, "2026-07-21", "OTRA ARS"),
                (880004, "2026-06-15", "FUTURO"),
                (880005, "2026-03-01", "FUTURO"),
            ):
                row = connection.execute(
                    """INSERT INTO recibos(
                           numero,nombre,fecha,ars,total,created_at,is_deleted,
                           estado_facturacion,estado_facturacion_at,
                           numero_autorizacion,service_type,specialty_snapshot,
                           admission_nss_snapshot,admission_ars_snapshot
                       ) VALUES(%s,%s,%s,%s,100,%s,0,%s,%s,%s,'EMERGENCIA',NULL,%s,%s)
                       RETURNING id""",
                    (
                        number, f"PACIENTE {number}", service_date, ars,
                        f"{service_date} 12:00:00", app.BILLING_INVOICED,
                        f"{service_date} 12:30:00", f"AUT-{number}",
                        str(number), ars,
                    ),
                ).fetchone()
                receipt_ids.append(int(row["id"]))

        filtered = app.list_available_receipts_for_batch(
            batch_id, date_from="2026-07-21", date_to="2026-07-21"
        )
        self.assertEqual(
            {int(row["recibo_id"]) for row in filtered}, {receipt_ids[1]}
        )
        all_candidates = app.list_available_receipts_for_batch(
            batch_id, date_from=None, date_to=None
        )
        future_ids = {int(row["recibo_id"]) for row in all_candidates}
        self.assertEqual(
            future_ids,
            {receipt_ids[0], receipt_ids[1], receipt_ids[3], receipt_ids[4]},
        )
        self.assertTrue(
            all(row["specialty_snapshot"] is None for row in all_candidates)
        )

        ninety_days = app.list_available_receipts_for_batch(
            batch_id, date_from="2026-04-23", date_to="2026-07-21"
        )
        self.assertEqual(
            {int(row["recibo_id"]) for row in ninety_days},
            {receipt_ids[0], receipt_ids[1], receipt_ids[3]},
        )

        added = app.add_receipts_to_monthly_batch(
            batch_id, [receipt_ids[0], receipt_ids[1]], actor
        )
        self.assertEqual(added, 2)
        self.assertEqual(len(app.list_monthly_batch_receipts(batch_id)), 2)
        remaining_ids = {
            int(row["recibo_id"])
            for row in app.list_available_receipts_for_batch(batch_id)
        }
        self.assertEqual(remaining_ids, {receipt_ids[3], receipt_ids[4]})
        with self.assertRaises(ValueError):
            app.add_receipts_to_monthly_batch(batch_id, [receipt_ids[0]], actor)

        export_path = Path(self.temp_dir.name) / "expediente-prueba.xlsx"
        Workbook().save(export_path)
        app.record_monthly_batch_export(batch_id, str(export_path), actor)
        second_batch = app.create_monthly_billing_batch(2026, 7, "FUTURO", actor)
        self.assertEqual(app.list_monthly_batch_receipts(second_batch), [])
        second_available = {
            int(row["recibo_id"])
            for row in app.list_available_receipts_for_batch(second_batch)
        }
        self.assertEqual(second_available, {receipt_ids[3], receipt_ids[4]})

    def test_delete_only_pending_batch_and_keep_original_receipts(self):
        actor = {"username": "admin", "role": app.ROLE_ADMIN}
        batch_id = app.create_monthly_billing_batch(2026, 7, "FUTURO", actor)
        with app.db_connect() as connection:
            receipt_count_before = int(
                connection.execute("SELECT COUNT(*) FROM recibos").fetchone()[0]
            )
        result = app.delete_monthly_billing_batch(batch_id, actor)
        self.assertEqual(int(result["batch"]["id"]), batch_id)
        self.assertIsNone(app.get_monthly_billing_batch(batch_id))
        with app.db_connect() as connection:
            receipt_count_after = int(
                connection.execute("SELECT COUNT(*) FROM recibos").fetchone()[0]
            )
        self.assertEqual(receipt_count_after, receipt_count_before)

        sent_id = app.create_monthly_billing_batch(2026, 7, "FUTURO", actor)
        with app.db_connect() as connection:
            connection.execute(
                """UPDATE billing_batches SET status=%s,sent_at=%s,sent_by=%s,
                          sent_receipt_count=0,sent_total=0,sent_invoice_number='',
                          sent_ncf='',sent_ars=ars,sent_period_year=period_year,
                          sent_period_month=period_month
                   WHERE id=%s""",
                (app.BATCH_SENT, app.now_str(), "admin", sent_id),
            )
        with self.assertRaisesRegex(ValueError, "enviado"):
            app.delete_monthly_billing_batch(sent_id, actor)

    def test_senasa_subsidiado_cannot_create_new_batch(self):
        actor = {"username": "admin", "role": app.ROLE_ADMIN}
        with self.assertRaisesRegex(ValueError, "SENASA SUBSIDIADO"):
            app.create_monthly_billing_batch(
                2026, 7, "  SeNaSa   SuBsIdIaDo  ", actor
            )

    def test_transitional_receipt_candidates_include_incomplete_history(self):
        actor = {"username": "admin", "role": app.ROLE_ADMIN}
        batch_id = app.create_monthly_billing_batch(2026, 7, "HUMANO", actor)
        with app.db_connect() as connection:
            receipt = connection.execute(
                """INSERT INTO recibos(
                       numero,nombre,fecha,ars,sala,total,pdf_filename,username,
                       created_at,is_deleted,estado_facturacion,service_type
                   ) VALUES(%s,%s,%s,%s,0,0,'','auxiliar',%s,0,%s,%s)
                   RETURNING id""",
                (
                    991001, "RECIBO HISTÓRICO INCOMPLETO", "2026-07-10",
                    "  humano  ", "2026-07-10 10:00:00",
                    app.BILLING_UNCLASSIFIED, "EMERGENCIA",
                ),
            ).fetchone()
            receipt_id = int(receipt["id"])

        candidates = app.obtener_candidatos_listado_ars(
            batch_id,
            date_from="2026-07-01",
            date_to="2026-07-31",
        )
        selected = next(row for row in candidates if int(row["recibo_id"]) == receipt_id)
        self.assertEqual(selected["candidate_kind"], "RECEIPT")
        self.assertEqual(float(selected["total_snapshot"]), 0.0)
        self.assertFalse(selected["document_number_snapshot"])
        self.assertFalse(selected["authorization_snapshot"])

        added = app.add_receipts_to_monthly_batch(
            batch_id,
            [receipt_id],
            actor,
            date_from="2026-07-01",
            date_to="2026-07-31",
        )
        self.assertEqual(added, 1)
        included = app.list_monthly_batch_receipts(batch_id)
        self.assertEqual(len(included), 1)
        self.assertEqual(int(included[0]["recibo_id"]), receipt_id)
        self.assertEqual(float(included[0]["total_snapshot"]), 0.0)

        remaining = app.obtener_candidatos_listado_ars(
            batch_id,
            date_from="2026-07-01",
            date_to="2026-07-31",
        )
        self.assertNotIn(
            receipt_id,
            [row.get("recibo_id") for row in remaining],
        )
        next_batch = app.create_monthly_billing_batch(2026, 7, "HUMANO", actor)
        next_candidates = app.obtener_candidatos_listado_ars(
            next_batch,
            date_from="2026-07-01",
            date_to="2026-07-31",
        )
        self.assertNotIn(
            receipt_id,
            [row.get("recibo_id") for row in next_candidates],
        )

    def test_historical_candidates_use_exact_normalized_ars_and_receipt_date(self):
        actor = {"username": "admin", "role": app.ROLE_ADMIN}
        with app.db_connect() as connection:
            rows = []
            for number, name, date_value, ars, status in (
                (992001, "FUTURO UNO", "2026-07-02", "FUTURO", app.BILLING_UNCLASSIFIED),
                (992002, "FUTURO DOS", "2026-07-20", "  futuro  ", app.BILLING_PENDING),
                (992003, "OTRA ARS", "2026-07-10", "UNIVERSAL", app.BILLING_INVOICED),
                (992004, "FUERA DE RANGO", "2026-06-30", "FUTURO", app.BILLING_NOT_INVOICED),
                (992005, "SIN FECHA VÁLIDA", "SIN-FECHA", "FUTURO", app.BILLING_UNCLASSIFIED),
            ):
                row = connection.execute(
                    """INSERT INTO recibos(
                           numero,nombre,fecha,ars,sala,total,pdf_filename,username,
                           created_at,is_deleted,estado_facturacion,service_type
                       ) VALUES(%s,%s,%s,%s,0,125.50,'','auxiliar',%s,0,%s,'EMERGENCIA')
                       RETURNING id""",
                    (number, name, date_value, ars, "2026-07-21 09:00:00", status),
                ).fetchone()
                rows.append(int(row["id"]))

        batch_id = app.create_monthly_billing_batch(2026, 7, "FUTURO", actor)
        ranged = app.obtener_candidatos_listado_ars(
            batch_id, date_from="2026-07-01", date_to="2026-07-31"
        )
        self.assertEqual(
            {int(row["recibo_id"]) for row in ranged},
            {rows[0], rows[1]},
        )
        unlimited = app.obtener_candidatos_listado_ars(batch_id)
        self.assertEqual(
            {int(row["recibo_id"]) for row in unlimited},
            {rows[0], rows[1], rows[3], rows[4]},
        )
        self.assertTrue(
            all(row["candidate_kind"] == "RECEIPT" for row in unlimited)
        )

    def test_saved_batches_migration_is_idempotent(self):
        migration = (
            Path(app.APP_DIR)
            / "migrations"
            / "20260721_saved_ars_batches_receipt_source.sql"
        ).read_text(encoding="utf-8")
        with app.db_connect() as connection:
            connection.execute(migration)
            connection.execute(migration)
            columns = {
                row[0]
                for row in connection.execute(
                    """SELECT column_name FROM information_schema.columns
                       WHERE table_name='billing_batches'"""
                ).fetchall()
            }
            indexes = {
                row[0]
                for row in connection.execute(
                    """SELECT indexname FROM pg_indexes
                       WHERE tablename='billing_batch_receipts'"""
                ).fetchall()
            }
        self.assertTrue({"sent_at", "sent_by"}.issubset(columns))
        self.assertIn("uq_billing_batch_receipts_receipt_global", indexes)

    def test_sent_batch_report_migration_is_idempotent(self):
        migration = (
            Path(app.APP_DIR) / "migrations" / "20260721_sent_batch_reports.sql"
        ).read_text(encoding="utf-8")
        with app.db_connect() as connection:
            connection.execute(migration)
            connection.execute(migration)
            columns = {
                row[0]
                for row in connection.execute(
                    """SELECT column_name FROM information_schema.columns
                       WHERE table_name='billing_batches'"""
                ).fetchall()
            }
            indexes = {
                row[0]
                for row in connection.execute(
                    """SELECT indexname FROM pg_indexes
                       WHERE tablename='billing_batches'"""
                ).fetchall()
            }
        self.assertTrue({
            "last_export_signature", "sent_receipt_count", "sent_total",
            "sent_invoice_number", "sent_ncf", "sent_ars",
            "sent_period_year", "sent_period_month",
        }.issubset(columns))
        self.assertIn("idx_billing_batches_sent_at", indexes)

    def test_two_workstations_keep_independent_sessions(self):
        app.register_active_session(
            "auxiliar", "station-admission",
            device_id="device-admission", device_name="PC Admisión",
        )
        app.register_active_session(
            "auxiliar", "station-billing",
            device_id="device-billing", device_name="PC Facturación",
        )

        with app.db_connect() as connection:
            count = connection.execute(
                "SELECT COUNT(*) FROM active_sessions "
                "WHERE username=%s AND is_active=1",
                ("auxiliar",),
            ).fetchone()[0]
        self.assertEqual(count, 2)

        app.end_active_session("auxiliar", "station-admission")
        self.assertIn("auxiliar", app.get_active_sessions_map())

        app.end_active_session("auxiliar", "station-billing")
        self.assertNotIn("auxiliar", app.get_active_sessions_map())

    def test_third_workstation_is_blocked_and_same_device_is_recovered(self):
        app.register_active_session(
            "auxiliar", "station-one",
            device_id="device-one", device_name="PC 1",
        )
        app.register_active_session(
            "auxiliar", "station-two",
            device_id="device-two", device_name="PC 2",
        )
        with self.assertRaises(app.ActiveUserSessionError):
            app.register_active_session(
                "auxiliar", "station-three",
                device_id="device-three", device_name="PC 3",
            )

        app.end_active_session("auxiliar", "station-one")
        recovered = app.register_active_session(
            "auxiliar", "station-two-recovered",
            device_id="device-two", device_name="PC 2",
            return_details=True,
        )
        self.assertTrue(recovered["recovered_same_device"])
        with app.db_connect() as connection:
            active_rows = connection.execute(
                """SELECT session_id,device_id FROM active_sessions
                   WHERE username=%s AND is_active=1""",
                ("auxiliar",),
            ).fetchall()
        self.assertEqual(len(active_rows), 1)
        self.assertEqual(active_rows[0]["session_id"], "station-two-recovered")


if __name__ == "__main__":
    unittest.main()
