import tempfile
import unittest
from pathlib import Path

from openpyxl import load_workbook
from PyPDF2 import PdfReader

from billing_batch_reports import create_sent_batches_reports


class SentBatchReportsTests(unittest.TestCase):
    def test_pdf_and_excel_include_only_frozen_operational_data(self):
        rows = [
            {
                "id": 10, "version": 1, "period_year": 2026, "period_month": 7,
                "ars": "FUTURO", "invoice_number": "FAC-10", "ncf": "B0100000010",
                "sent_at": "2026-07-20 14:10:00", "sent_by": "aux.uno",
                "receipt_count": 4, "sent_total": 1800.50,
            },
            {
                "id": 11, "version": 2, "period_year": 2026, "period_month": 7,
                "ars": "FUTURO", "invoice_number": "FAC-11", "ncf": "B0100000011",
                "sent_at": "2026-07-21 09:30:00", "sent_by": "aux.dos",
                "receipt_count": 3, "sent_total": 950.25,
            },
            {
                "id": 12, "version": 1, "period_year": 2026, "period_month": 6,
                "ars": "HUMANO", "invoice_number": "FAC-12", "ncf": "B0100000012",
                "sent_at": "2026-07-01 08:15:00", "sent_by": "admin",
                "receipt_count": 2, "sent_total": 725.00,
            },
        ]
        with tempfile.TemporaryDirectory() as temp_dir:
            base = str(Path(temp_dir) / "Reporte_Listados_Enviados_Prueba")
            pdf_path, xlsx_path = create_sent_batches_reports(
                base, rows, "admin", None,
            )
            self.assertTrue(Path(pdf_path).is_file())
            self.assertTrue(Path(xlsx_path).is_file())

            workbook = load_workbook(xlsx_path, data_only=False)
            try:
                self.assertEqual(workbook.sheetnames, ["Resumen", "Listados enviados"])
                detail = workbook["Listados enviados"]
                self.assertEqual(detail["A8"].value, "07-2026")
                self.assertEqual(detail["B8"].value, "FUTURO")
                self.assertEqual(detail["J8"].value, 1800.50)
                self.assertEqual(detail["I11"].value, "=SUM(I8:I10)")
                self.assertEqual(detail["J11"].value, "=SUM(J8:J10)")
                summary = workbook["Resumen"]
                self.assertEqual(summary["A7"].value, 3)
                self.assertEqual(summary["C7"].value, 9)
                self.assertAlmostEqual(summary["E7"].value, 3475.75)
            finally:
                workbook.close()

            text = "\n".join(
                page.extract_text() or "" for page in PdfReader(pdf_path).pages
            )
            self.assertIn("REPORTE DE LISTADOS ARS ENVIADOS", text)
            self.assertIn("FUTURO", text)
            self.assertIn("HUMANO", text)
            self.assertIn("RD$ 3,475.75", text)


if __name__ == "__main__":
    unittest.main()
