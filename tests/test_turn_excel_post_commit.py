from __future__ import annotations

import ctypes
import concurrent.futures
import importlib
import os
import sqlite3
from datetime import datetime, timezone
from pathlib import Path
from types import SimpleNamespace

import pytest
from openpyxl import load_workbook
from PyPDF2 import PdfReader

from admission_hybrid import build_admission_order_key
from admission_v15_adapter import DEFAULT_V15_ROOT, _load_v15_modules


class _TurnDatabase:
    def __init__(self, rows=None):
        self.rows = (
            rows
            if rows is not None
            else [
                {
                    "id": 1,
                    "nombre": "Paciente de prueba",
                    "tipo_atencion": "EMERGENCIA",
                    "hoja": "GENERAL",
                    "hoja_normalizada": "GENERAL",
                    "ars_display": "HUMANO",
                }
            ]
        )

    def buscar_contexto_turno_existente(self, _turno_cfg):
        return {"turno_id": 282}

    def obtener_atenciones_para_rango_real(self, **_kwargs):
        return list(self.rows)


class _CanonicalTurnDatabase(_TurnDatabase):
    def __init__(self, rows):
        super().__init__(rows)
        self.requested_identity = None
        self.dataset_calls = 0

    def get_operational_station_snapshot(self):
        return {
            "turn_id": 316,
            "operational_source_id": "source-current",
        }

    def build_turn_dataset(self, *, turn_id, operational_source_id):
        self.dataset_calls += 1
        self.requested_identity = (turn_id, operational_source_id)
        return sorted(self.rows, key=build_admission_order_key)

    def obtener_atenciones_para_rango_real(self, **_kwargs):
        raise AssertionError("Excel no debe construir otra vista desde SQLite")


def _v15_module():
    _load_v15_modules(Path(DEFAULT_V15_ROOT))
    return importlib.import_module("ADMISION_PYSIDE6_V15.facturacion_tabs_pyside6")


def _turn_config():
    now = datetime.now(timezone.utc).astimezone().replace(microsecond=0, tzinfo=None)
    return {
        "representante": "USUARIO PRUEBA",
        "turno_codigo": "8AM_8AM",
        "fecha_base": now.date(),
        "inicio_real": now.strftime("%d/%m/%Y %I:%M:%S %p"),
        "inicio_real_dt": now,
    }


def test_open_canonical_excel_is_deferred_without_repeating_transition(
    tmp_path, monkeypatch
):
    v15 = _v15_module()
    canonical = tmp_path / "LISTADO DE PACIENTES EN EMERGENCIA.xlsx"
    queue = tmp_path / "excel_export_jobs.sqlite3"
    versions = tmp_path / "EXCEL_POR_TURNO"
    monkeypatch.setattr(v15, "EXCEL_PATH", str(canonical))
    monkeypatch.setattr(v15, "EXCEL_EXPORT_QUEUE_PATH", str(queue))
    monkeypatch.setattr(v15, "EXCEL_VERSIONED_DIR", str(versions))

    transition_id = "77777777-7777-4777-8777-777777777777"
    first_job = v15.enqueue_excel_export_job(transition_id, 282, _turn_config())
    duplicate_job = v15.enqueue_excel_export_job(transition_id, 282, _turn_config())
    assert duplicate_job == first_job

    real_update = v15._update_canonical_excel

    def locked_canonical(_source_file, _canonical_target):
        raise PermissionError(13, "archivo en uso")

    monkeypatch.setattr(v15, "_update_canonical_excel", locked_canonical)
    deferred = v15.process_excel_export_jobs(_TurnDatabase())
    assert deferred == {
        "completed": 0,
        "pending": 1,
        "errors": 0,
        "processed": 1,
        "skipped_empty": 0,
    }
    assert not canonical.exists()
    versioned = list(versions.glob("LISTADO_DE_PACIENTES_TURNO_282_*.xlsx"))
    assert len(versioned) == 1

    with sqlite3.connect(queue) as conn:
        status, attempts, error_code = conn.execute(
            "SELECT status,attempts,last_error_code FROM excel_export_jobs"
        ).fetchone()
    assert status == "PENDING"
    assert attempts == 1
    assert error_code == "EXCEL_EXPORT_DEFERRED_FILE_IN_USE"

    monkeypatch.setattr(v15, "_update_canonical_excel", real_update)
    retried = v15.process_excel_export_jobs(_TurnDatabase())
    assert retried == {
        "completed": 1,
        "pending": 0,
        "errors": 0,
        "processed": 1,
        "skipped_empty": 0,
    }
    assert canonical.is_file()
    assert canonical.read_bytes() == versioned[0].read_bytes()
    with sqlite3.connect(queue) as conn:
        status, attempts, total = conn.execute(
            "SELECT status,attempts,(SELECT COUNT(*) FROM excel_export_jobs) "
            "FROM excel_export_jobs"
        ).fetchone()
    assert status == "COMPLETED"
    assert attempts == 2
    assert total == 1


def test_empty_turn_does_not_create_or_replace_excel(tmp_path, monkeypatch):
    v15 = _v15_module()
    canonical = tmp_path / "LISTADO DE PACIENTES EN EMERGENCIA.xlsx"
    canonical.write_bytes(b"archivo historico")
    queue = tmp_path / "excel_export_jobs.sqlite3"
    versions = tmp_path / "EXCEL_POR_TURNO"
    monkeypatch.setattr(v15, "EXCEL_PATH", str(canonical))
    monkeypatch.setattr(v15, "EXCEL_EXPORT_QUEUE_PATH", str(queue))
    monkeypatch.setattr(v15, "EXCEL_VERSIONED_DIR", str(versions))

    v15.enqueue_excel_export_job(
        "88888888-8888-4888-8888-888888888888", 283, _turn_config()
    )
    result = v15.process_excel_export_jobs(_TurnDatabase(rows=[]))

    assert result == {
        "completed": 1,
        "pending": 0,
        "errors": 0,
        "processed": 1,
        "skipped_empty": 1,
    }
    assert canonical.read_bytes() == b"archivo historico"
    assert not list(versions.glob("*.xlsx"))
    with sqlite3.connect(queue) as conn:
        status, code = conn.execute(
            "SELECT status,last_error_code FROM excel_export_jobs"
        ).fetchone()
    assert status == "COMPLETED"
    assert code == "SKIPPED_EMPTY"


def test_admission_report_helpers_reject_empty_dataset(tmp_path):
    v15 = _v15_module()
    empty = v15.construir_resumen_desde_registros([], "Turno de prueba")
    assert v15.reportable_patient_count(empty) == 0
    with pytest.raises(v15.EmptyAdmissionReportError):
        v15.crear_pdf_reporte(empty, destino=str(tmp_path / "vacio.pdf"))
    with pytest.raises(v15.EmptyAdmissionReportError):
        v15.crear_excel_reporte_estadistico(empty, destino=str(tmp_path / "vacio.xlsx"))
    assert not (tmp_path / "vacio.pdf").exists()
    assert not (tmp_path / "vacio.xlsx").exists()


def test_nonempty_admission_reports_are_still_generated(tmp_path):
    v15 = _v15_module()
    summary = v15.construir_resumen_desde_registros(
        [
            {
                "id": 1,
                "nombre": "Paciente de prueba",
                "tipo_atencion": "EMERGENCIA",
                "ars_display": "HUMANO",
                "hoja_normalizada": "GENERAL",
                "fecha": "2026-08-12",
                "hora": "12:00",
                "nss": "000000001",
                "cedula": "",
            }
        ],
        "Turno de prueba",
        representante="USUARIO PRUEBA",
    )
    pdf = tmp_path / "con_datos.pdf"
    excel = tmp_path / "con_datos.xlsx"
    assert v15.reportable_patient_count(summary) == 1
    assert v15.crear_pdf_reporte(summary, destino=str(pdf)) == str(pdf)
    assert v15.crear_excel_reporte_estadistico(summary, destino=str(excel)) == str(
        excel
    )
    assert pdf.stat().st_size > 0
    assert excel.stat().st_size > 0


def test_excel_uses_canonical_dataset_order_and_hidden_global_identity(
    tmp_path, monkeypatch
):
    v15 = _v15_module()
    canonical = tmp_path / "LISTADO DE PACIENTES EN EMERGENCIA.xlsx"
    latest = tmp_path / "LISTADO.latest.xlsx"
    state = tmp_path / "excel_export_state.json"
    queue = tmp_path / "excel_export_jobs.sqlite3"
    monkeypatch.setattr(v15, "EXCEL_PATH", str(canonical))
    monkeypatch.setattr(v15, "EXCEL_LATEST_PATH", str(latest))
    monkeypatch.setattr(v15, "EXCEL_EXPORT_STATE_PATH", str(state))
    monkeypatch.setattr(v15, "EXCEL_EXPORT_QUEUE_PATH", str(queue))
    rows = [
        {
            "id": 2,
            "global_attention_id": "bbbbbbbb-bbbb-4bbb-8bbb-bbbbbbbbbbbb",
            "nombre": "PACIENTE B",
            "hoja_normalizada": "PEDIATRIA",
            "ars_display": "HUMANO",
            "created_at_effective_utc": "2026-08-13T12:00:00.000+00:00",
            "origin_device_id": "PC-2",
            "device_local_sequence": 1,
        },
        {
            "id": 1,
            "global_attention_id": "aaaaaaaa-aaaa-4aaa-8aaa-aaaaaaaaaaaa",
            "nombre": "PACIENTE A",
            "hoja_normalizada": "GENERAL",
            "ars_display": "FUTURO",
            "created_at_effective_utc": "2026-08-13T12:00:00.000+00:00",
            "origin_device_id": "PC-1",
            "device_local_sequence": 1,
        },
    ]
    database = _CanonicalTurnDatabase(rows)

    assert v15.reconstruir_excel_turno(database, _turn_config()) == 2
    assert database.requested_identity == (316, "source-current")
    workbook = load_workbook(canonical, data_only=True)
    try:
        sheet = workbook.active
        assert [sheet.cell(row=row, column=1).value for row in (6, 7)] == [1, 2]
        assert [sheet.cell(row=row, column=2).value for row in (6, 7)] == [
            "PACIENTE A",
            "PACIENTE B",
        ]
        assert [sheet.cell(row=row, column=5).value for row in (6, 7)] == [
            "aaaaaaaa-aaaa-4aaa-8aaa-aaaaaaaaaaaa",
            "bbbbbbbb-bbbb-4bbb-8bbb-bbbbbbbbbbbb",
        ]
        assert sheet.column_dimensions["E"].hidden is True
        assert sheet.column_dimensions["F"].hidden is True
        assert len(str(sheet["F1"].value)) == 64
    finally:
        workbook.close()


def test_live_excel_lock_keeps_latest_and_defers_canonical_publish(
    tmp_path, monkeypatch
):
    v15 = _v15_module()
    canonical = tmp_path / "LISTADO DE PACIENTES EN EMERGENCIA.xlsx"
    canonical.write_bytes(b"version abierta")
    latest = tmp_path / "LISTADO.latest.xlsx"
    state = tmp_path / "excel_export_state.json"
    queue = tmp_path / "excel_export_jobs.sqlite3"
    versions = tmp_path / "EXCEL_POR_TURNO"
    monkeypatch.setattr(v15, "EXCEL_PATH", str(canonical))
    monkeypatch.setattr(v15, "EXCEL_LATEST_PATH", str(latest))
    monkeypatch.setattr(v15, "EXCEL_EXPORT_STATE_PATH", str(state))
    monkeypatch.setattr(v15, "EXCEL_EXPORT_QUEUE_PATH", str(queue))
    monkeypatch.setattr(v15, "EXCEL_VERSIONED_DIR", str(versions))
    database = _CanonicalTurnDatabase(
        [
            {
                "id": 1,
                "global_attention_id": "aaaaaaaa-aaaa-4aaa-8aaa-aaaaaaaaaaaa",
                "nombre": "PACIENTE A",
                "hoja_normalizada": "GENERAL",
                "ars_display": "FUTURO",
                "created_at_effective_utc": "2026-08-13T12:00:00+00:00",
                "origin_device_id": "PC-1",
                "device_local_sequence": 1,
            }
        ]
    )

    monkeypatch.setattr(
        v15,
        "_update_canonical_excel",
        lambda *_args: (_ for _ in ()).throw(PermissionError(13, "archivo en uso")),
    )
    assert v15.reconstruir_excel_turno(database, _turn_config()) == 1
    assert canonical.read_bytes() == b"version abierta"
    assert latest.is_file()
    export_state = v15._read_excel_export_state()
    assert export_state["excel_status"] == "FILE_LOCKED"
    assert export_state["patient_count"] == 1
    with sqlite3.connect(queue) as connection:
        source_file, status = connection.execute(
            "SELECT source_file,status FROM excel_export_jobs"
        ).fetchone()
    assert Path(source_file) == latest
    assert status == "PENDING"


def test_turn_dialog_commits_before_excel_and_does_not_use_excel_com():
    source_path = Path(DEFAULT_V15_ROOT) / "facturacion_tabs_pyside6.py"
    source = source_path.read_text(encoding="utf-8")
    dialog_start = source.index("        def refrescar_turnos():")
    start = source.index("        def _aplicar_cambio():", dialog_start)
    end = source.index("        aplicando =", start)
    apply_block = source[start:end]

    assert "reconstruir_excel_turno(self.db, turno_cfg_nuevo)" not in apply_block
    assert apply_block.index("capture_outgoing_turn_context(") < apply_block.index(
        "self.db.perform_explicit_turn_handoff("
    )
    assert "if outgoing_context is not None:" in apply_block
    assert apply_block.index("self.db.obtener_o_crear_turno(") < apply_block.index(
        "enqueue_excel_export_job("
    )
    assert apply_block.index("win.destroy()") < apply_block.index(
        "self._run_turn_post_commit_effects"
    )
    assert "win32com" not in source
    assert "GetActiveObject" not in source
    assert "Dispatch(" not in source
    assert "DispatchEx(" not in source
    assert "os.startfile(ruta)" in source


def test_outgoing_turn_context_is_captured_before_the_handoff():
    v15 = _v15_module()
    closed_at = datetime(2026, 9, 3, 8, 0, 0)
    started_at = datetime(2026, 9, 2, 8, 0, 0)
    context = v15.capture_outgoing_turn_context(
        {
            "representante": "REPRESENTANTE ANTERIOR",
            "turno_codigo": "8AM_8AM",
            "fecha_base": started_at.date(),
            "inicio_real_dt": started_at,
        },
        {
            "operational_source_id": "source-old",
            "turn_id": 410,
            "generation": 17,
            "operational_revision": 29,
            "active_user_id": "user-old",
            "active_user_display_name": "REPRESENTANTE ANTERIOR",
        },
        closed_at,
    )

    assert context.operational_source_id == "source-old"
    assert context.turn_id == 410
    assert context.generation == 17
    assert context.operational_revision == 29
    assert context.representative_id == "user-old"
    assert context.started_at == started_at
    assert context.closed_at == closed_at


def test_outgoing_context_rejects_missing_identity_and_transition_collision():
    v15 = _v15_module()
    started_at = datetime(2026, 9, 2, 8, 0, 0)
    with pytest.raises(v15.TurnoNoVigenteError):
        v15.capture_outgoing_turn_context(
            {
                "representante": "REPRESENTANTE",
                "turno_codigo": "8AM_8AM",
                "fecha_base": started_at.date(),
                "inicio_real_dt": started_at,
            },
            {"turn_id": None, "operational_source_id": ""},
            datetime(2026, 9, 3, 8, 0, 0),
        )
    with pytest.raises(v15.TurnoNoVigenteError, match="IDENTITY_INVALID"):
        v15.capture_outgoing_turn_context(
            {"inicio_real_dt": started_at},
            {"turn_id": "invalid", "operational_source_id": "source-old"},
            datetime(2026, 9, 3, 8, 0, 0),
        )
    context = v15.OutgoingTurnContext(
        operational_source_id="source-old",
        turn_id=410,
        generation=17,
        operational_revision=29,
        representative_id="user-old",
        representative_display_name="REPRESENTANTE ANTERIOR",
        started_at=started_at,
        closed_at=datetime(2026, 9, 3, 8, 0, 0),
        turn_code="8AM_8AM",
        base_date=started_at.date(),
    )
    transition = SimpleNamespace(
        old_turn_id=999,
        transition_id="99999999-9999-4999-8999-999999999999",
        new_turn_id=411,
        new_user_id="user-new",
        operational_session=None,
    )
    with pytest.raises(RuntimeError, match="turno saliente"):
        v15.bind_outgoing_turn_transition(context, transition)
    with pytest.raises(RuntimeError, match="identidad completa"):
        v15.bind_outgoing_turn_transition(
            context,
            SimpleNamespace(
                old_turn_id=410,
                transition_id="",
                new_turn_id=0,
                operational_session=None,
            ),
        )


def test_outgoing_context_parses_persisted_dates_and_binds_committed_transition():
    v15 = _v15_module()
    context = v15.capture_outgoing_turn_context(
        {
            "representante": "REPRESENTANTE ANTERIOR",
            "turno_codigo": "8AM_8AM",
            "fecha_base": datetime(2026, 9, 2, 0, 0, 0),
            "inicio_real": "02/09/2026 08:00 AM",
        },
        {
            "operational_source_id": "source-old",
            "turn_id": "410",
            "generation": "17",
            "operational_revision": "29",
            "representative_user_id": "user-old",
            "representative_display_name": "REPRESENTANTE ANTERIOR",
        },
        datetime(2026, 9, 3, 8, 0, 0),
    )
    transition = SimpleNamespace(
        old_turn_id=410,
        transition_id="99999999-9999-4999-8999-999999999999",
        new_turn_id=411,
        new_user_id="user-new",
        operational_session=None,
    )

    bound = v15.bind_outgoing_turn_transition(context, transition)

    assert bound.transition_id == transition.transition_id
    assert bound.new_turn_id == 411
    assert bound.new_representative_id == "user-new"
    assert bound.base_date == datetime(2026, 9, 2).date()
    assert bound.as_turn_config()["turno_codigo"] == "8AM_8AM"
    fallback_date_context = v15.capture_outgoing_turn_context(
        {"inicio_real_dt": datetime(2026, 9, 2, 8, 0, 0)},
        {"operational_source_id": "source-old", "turn_id": 410},
        datetime(2026, 9, 3, 8, 0, 0),
    )
    assert fallback_date_context.base_date == datetime(2026, 9, 2).date()


def test_outgoing_context_rejects_invalid_start_and_end():
    v15 = _v15_module()
    snapshot = {"operational_source_id": "source-old", "turn_id": 410}
    with pytest.raises(v15.TurnoNoVigenteError, match="START"):
        v15.capture_outgoing_turn_context(
            {"inicio_real": "invalid"},
            snapshot,
            datetime(2026, 9, 3, 8, 0, 0),
        )
    with pytest.raises(v15.TurnoNoVigenteError, match="END"):
        v15.capture_outgoing_turn_context(
            {"inicio_real_dt": datetime(2026, 9, 3, 8, 0, 0)},
            snapshot,
            datetime(2026, 9, 3, 8, 0, 0),
        )


def test_post_commit_scheduler_runs_reports_only_when_context_exists():
    v15 = _v15_module()
    scheduled = []
    root = SimpleNamespace(
        after=lambda delay, callback: scheduled.append((delay, callback))
    )
    warnings = []

    assert (
        v15.schedule_turn_closure_post_commit(
            root, lambda *_args: None, None, None, warnings
        )
        is False
    )
    assert warnings == ["POST_COMMIT_REPORT_CONTEXT"]
    context = SimpleNamespace(transition_id="transition")
    calls = []
    assert (
        v15.schedule_turn_closure_post_commit(
            root,
            lambda current, new: calls.append((current, new)),
            context,
            {"turn_id": 411},
            warnings,
        )
        is True
    )
    assert scheduled[-1][0] == 0
    scheduled[-1][1]()
    assert calls == [(context, {"turn_id": 411})]


def test_turn_closure_snapshot_queries_old_identity_once_and_is_immutable():
    v15 = _v15_module()
    started_at = datetime(2026, 9, 2, 8, 0, 0)
    closed_at = datetime(2026, 9, 3, 8, 0, 0)
    rows = [
        {
            "id": 9,
            "attention_id": 9,
            "global_attention_id": "aaaaaaaa-aaaa-4aaa-8aaa-aaaaaaaaaaaa",
            "nombre": "PACIENTE TURNO SALIENTE",
            "tipo_atencion": "EMERGENCIA",
            "hoja_normalizada": "GENERAL",
            "ars_display": "HUMANO",
            "created_at_effective_utc": "2026-09-02T14:30:00",
            "operational_source_id": "source-old",
            "turn_id": 410,
            "source_status": "ACTIVA",
            "is_deleted": False,
        }
    ]
    database = _CanonicalTurnDatabase(rows)
    context = v15.OutgoingTurnContext(
        operational_source_id="source-old",
        turn_id=410,
        generation=17,
        operational_revision=29,
        representative_id="user-old",
        representative_display_name="REPRESENTANTE ANTERIOR",
        started_at=started_at,
        closed_at=closed_at,
        turn_code="8AM_8AM",
        base_date=started_at.date(),
        transition_id="99999999-9999-4999-8999-999999999999",
        new_turn_id=411,
        new_representative_id="user-new",
    )

    snapshot = v15.build_turn_closure_report_snapshot(database, context)

    assert database.requested_identity == (410, "source-old")
    assert database.dataset_calls == 1
    assert snapshot.context == context
    assert snapshot.patient_count == 1
    assert snapshot.dataset.summary["turn_id"] == 410
    assert snapshot.dataset.summary["operational_source_id"] == "source-old"
    assert snapshot.global_attention_ids == ("aaaaaaaa-aaaa-4aaa-8aaa-aaaaaaaaaaaa",)
    with pytest.raises(TypeError):
        snapshot.dataset.records[0]["nombre"] = "MUTADO"
    anonymous_context = v15.replace(
        context,
        representative_id="",
        representative_display_name="",
        transition_id="aaaaaaaa-aaaa-4aaa-8aaa-aaaaaaaaaaaa",
    )
    anonymous = v15.build_turn_closure_report_snapshot(
        _CanonicalTurnDatabase(rows), anonymous_context
    )
    assert anonymous.dataset.representatives == ()


def test_turn_closure_snapshot_requires_transition_and_central_reader(monkeypatch):
    v15 = _v15_module()
    context = v15.OutgoingTurnContext(
        operational_source_id="source-old",
        turn_id=410,
        generation=17,
        operational_revision=29,
        representative_id="user-old",
        representative_display_name="REPRESENTANTE ANTERIOR",
        started_at=datetime(2026, 9, 2, 8, 0, 0),
        closed_at=datetime(2026, 9, 3, 8, 0, 0),
        turn_code="8AM_8AM",
        base_date=datetime(2026, 9, 2).date(),
    )
    with pytest.raises(ValueError, match="transition_id"):
        v15.build_turn_closure_report_snapshot(object(), context)
    bound = v15.replace(
        context,
        transition_id="99999999-9999-4999-8999-999999999999",
        new_turn_id=411,
    )
    monkeypatch.setattr(v15, "_dataset_turno_central", lambda *_args, **_kwargs: None)
    with pytest.raises(RuntimeError, match="central canónico"):
        v15.build_turn_closure_report_snapshot(object(), bound)
    monkeypatch.setattr(
        v15,
        "_dataset_turno_central",
        lambda *_args, **_kwargs: ([], 999, "another-source"),
    )
    with pytest.raises(RuntimeError, match="otra identidad"):
        v15.build_turn_closure_report_snapshot(object(), bound)


def test_turn_closure_pdf_and_excel_share_one_snapshot_and_are_idempotent(
    tmp_path, monkeypatch
):
    v15 = _v15_module()
    started_at = datetime(2026, 9, 2, 8, 0, 0)
    closed_at = datetime(2026, 9, 3, 8, 0, 0)
    context = v15.OutgoingTurnContext(
        operational_source_id="source-old",
        turn_id=410,
        generation=17,
        operational_revision=29,
        representative_id="user-old",
        representative_display_name="REPRESENTANTE ANTERIOR",
        started_at=started_at,
        closed_at=closed_at,
        turn_code="8AM_8AM",
        base_date=started_at.date(),
        transition_id="99999999-9999-4999-8999-999999999999",
        new_turn_id=411,
        new_representative_id="user-new",
    )
    database = _CanonicalTurnDatabase(
        [
            {
                "id": 9,
                "global_attention_id": "aaaaaaaa-aaaa-4aaa-8aaa-aaaaaaaaaaaa",
                "nombre": "PACIENTE TURNO SALIENTE",
                "tipo_atencion": "EMERGENCIA",
                "hoja_normalizada": "GENERAL",
                "ars_display": "HUMANO",
                "created_at_effective_utc": "2026-09-02T14:30:00",
                "operational_source_id": "source-old",
                "turn_id": 410,
                "source_status": "ACTIVA",
                "is_deleted": False,
            }
        ]
    )
    snapshot = v15.build_turn_closure_report_snapshot(database, context)
    summaries = []

    def fake_pdf(summary, destino=None):
        summaries.append(summary)
        Path(destino).write_bytes(b"%PDF-1.4\n% test\n")
        return str(destino)

    monkeypatch.setattr(v15, "crear_pdf_reporte", fake_pdf)
    first = v15.generate_turn_closure_report_files(
        snapshot,
        output_directory=tmp_path,
        generate_pdf=True,
        generate_excel=True,
    )
    second = v15.generate_turn_closure_report_files(
        snapshot,
        output_directory=tmp_path,
        generate_pdf=True,
        generate_excel=True,
    )

    assert summaries == [snapshot.dataset.summary]
    assert first.pdf_generated is True
    assert first.excel_generated is True
    assert second.pdf_generated is False
    assert second.excel_generated is False
    assert second.pdf_already_existed is True
    assert second.excel_already_existed is True
    assert first.patient_count == second.patient_count == 1
    assert first.dataset_revision == second.dataset_revision
    assert first.pdf_path == second.pdf_path
    assert first.excel_path == second.excel_path
    workbook = load_workbook(first.excel_path, data_only=True)
    try:
        listing = workbook["LISTADO DE PACIENTES"]
        assert listing["A6"].value == 1
        assert listing["B6"].value == "PACIENTE TURNO SALIENTE"
        assert listing["E6"].value == "aaaaaaaa-aaaa-4aaa-8aaa-aaaaaaaaaaaa"
        assert listing["F1"].value == first.dataset_revision
    finally:
        workbook.close()


def test_concurrent_report_retry_creates_each_old_turn_file_once(tmp_path, monkeypatch):
    v15 = _v15_module()
    started_at = datetime(2026, 9, 2, 8, 0, 0)
    context = v15.OutgoingTurnContext(
        operational_source_id="source-old",
        turn_id=410,
        generation=17,
        operational_revision=29,
        representative_id="user-old",
        representative_display_name="REPRESENTANTE ANTERIOR",
        started_at=started_at,
        closed_at=datetime(2026, 9, 3, 8, 0, 0),
        turn_code="8AM_8AM",
        base_date=started_at.date(),
        transition_id="99999999-9999-4999-8999-999999999999",
        new_turn_id=411,
        new_representative_id="user-new",
    )
    snapshot = v15.build_turn_closure_report_snapshot(
        _CanonicalTurnDatabase(
            [
                {
                    "id": 9,
                    "global_attention_id": "aaaaaaaa-aaaa-4aaa-8aaa-aaaaaaaaaaaa",
                    "nombre": "PACIENTE TURNO SALIENTE",
                    "created_at_effective_utc": "2026-09-02T14:30:00",
                    "operational_source_id": "source-old",
                    "turn_id": 410,
                }
            ]
        ),
        context,
    )
    generated_pdf_paths = []

    def fake_pdf(_summary, destino=None):
        generated_pdf_paths.append(str(destino))
        Path(destino).write_bytes(b"%PDF-1.4\n% test\n")
        return str(destino)

    monkeypatch.setattr(v15, "crear_pdf_reporte", fake_pdf)
    with concurrent.futures.ThreadPoolExecutor(max_workers=2) as executor:
        results = tuple(
            executor.map(
                lambda _index: v15.generate_turn_closure_report_files(
                    snapshot,
                    output_directory=tmp_path,
                    generate_pdf=True,
                    generate_excel=False,
                ),
                range(2),
            )
        )

    assert len(generated_pdf_paths) == 1
    assert sum(result.pdf_generated for result in results) == 1
    assert sum(result.pdf_already_existed for result in results) == 1


def test_real_turn_closure_pdf_and_excel_are_readable(tmp_path):
    v15 = _v15_module()
    started_at = datetime(2026, 9, 2, 8, 0, 0)
    context = v15.OutgoingTurnContext(
        operational_source_id="source-old",
        turn_id=410,
        generation=17,
        operational_revision=29,
        representative_id="user-old",
        representative_display_name="REPRESENTANTE ANTERIOR",
        started_at=started_at,
        closed_at=datetime(2026, 9, 3, 8, 0, 0),
        turn_code="8AM_8AM",
        base_date=started_at.date(),
        transition_id="99999999-9999-4999-8999-999999999999",
        new_turn_id=411,
        new_representative_id="user-new",
    )
    rows = [
        {
            "id": index,
            "global_attention_id": f"00000000-0000-4000-8000-{index:012d}",
            "nombre": f"PACIENTE {index}",
            "tipo_atencion": "EMERGENCIA",
            "hoja_normalizada": specialty,
            "ars_display": ars,
            "created_at_effective_utc": f"2026-09-02T{hour:02d}:30:00",
            "operational_source_id": "source-old",
            "turn_id": 410,
            "source_status": "ACTIVA",
            "is_deleted": False,
        }
        for index, specialty, ars, hour in (
            (1, "GENERAL", "HUMANO", 9),
            (2, "PEDIATRIA", "SIN SEGURO", 10),
        )
    ]
    snapshot = v15.build_turn_closure_report_snapshot(
        _CanonicalTurnDatabase(rows), context
    )

    files = v15.generate_turn_closure_report_files(
        snapshot,
        output_directory=tmp_path,
    )

    assert files.patient_count == 2
    assert len(PdfReader(files.pdf_path).pages) >= 1
    workbook = load_workbook(files.excel_path, data_only=True)
    try:
        sheet = workbook["LISTADO DE PACIENTES"]
        assert sheet.max_row == 7
        assert [sheet[f"B{row}"].value for row in (6, 7)] == [
            "PACIENTE 1",
            "PACIENTE 2",
        ]
    finally:
        workbook.close()


def test_empty_turn_closure_has_no_files_and_no_false_zero_artifact(tmp_path):
    v15 = _v15_module()
    started_at = datetime(2026, 9, 2, 8, 0, 0)
    context = v15.OutgoingTurnContext(
        operational_source_id="source-old",
        turn_id=410,
        generation=17,
        operational_revision=29,
        representative_id="user-old",
        representative_display_name="REPRESENTANTE ANTERIOR",
        started_at=started_at,
        closed_at=datetime(2026, 9, 3, 8, 0, 0),
        turn_code="8AM_8AM",
        base_date=started_at.date(),
        transition_id="99999999-9999-4999-8999-999999999999",
        new_turn_id=411,
        new_representative_id="user-new",
    )
    snapshot = v15.build_turn_closure_report_snapshot(
        _CanonicalTurnDatabase([]), context
    )

    files = v15.generate_turn_closure_report_files(snapshot, output_directory=tmp_path)

    assert files.patient_count == 0
    assert files.pdf_path == ""
    assert files.excel_path == ""
    assert list(tmp_path.iterdir()) == []
    with pytest.raises(v15.EmptyAdmissionReportError):
        v15.crear_excel_listado_turno_cerrado(snapshot, tmp_path / "empty.xlsx")


def test_turn_closure_excel_cleans_temporary_file_on_publish_error(
    tmp_path, monkeypatch
):
    v15 = _v15_module()
    started_at = datetime(2026, 9, 2, 8, 0, 0)
    context = v15.OutgoingTurnContext(
        operational_source_id="source-old",
        turn_id=410,
        generation=17,
        operational_revision=29,
        representative_id="user-old",
        representative_display_name="REPRESENTANTE ANTERIOR",
        started_at=started_at,
        closed_at=datetime(2026, 9, 3, 8, 0, 0),
        turn_code="8AM_8AM",
        base_date=started_at.date(),
        transition_id="99999999-9999-4999-8999-999999999999",
        new_turn_id=411,
    )
    snapshot = v15.build_turn_closure_report_snapshot(
        _CanonicalTurnDatabase(
            [
                {
                    "id": 1,
                    "global_attention_id": "aaaaaaaa-aaaa-4aaa-8aaa-aaaaaaaaaaaa",
                    "nombre": "PACIENTE",
                    "created_at_effective_utc": "2026-09-02T10:00:00",
                    "operational_source_id": "source-old",
                    "turn_id": 410,
                }
            ]
        ),
        context,
    )
    monkeypatch.setattr(
        v15.os,
        "replace",
        lambda *_args: (_ for _ in ()).throw(OSError("publish failed")),
    )

    with pytest.raises(OSError, match="publish failed"):
        v15.crear_excel_listado_turno_cerrado(snapshot, tmp_path / "failed.xlsx")

    assert not list(tmp_path.glob(".turn-closure-*.xlsx"))
    real_remove = v15.os.remove

    def fail_turn_closure_cleanup(path):
        if Path(path).name.startswith(".turn-closure-"):
            raise OSError("cleanup failed")
        return real_remove(path)

    monkeypatch.setattr(v15.os, "remove", fail_turn_closure_cleanup)
    with pytest.raises(OSError, match="publish failed"):
        v15.crear_excel_listado_turno_cerrado(snapshot, tmp_path / "failed-again.xlsx")
    temporary_files = list(tmp_path.glob(".turn-closure-*.xlsx"))
    assert temporary_files
    for temporary_file in temporary_files:
        temporary_file.unlink()
    monkeypatch.setattr(v15.os, "remove", real_remove)
    real_exists = v15.os.path.exists
    monkeypatch.setattr(
        v15.os.path,
        "exists",
        lambda path: (
            False if Path(path).name.startswith(".turn-closure-") else real_exists(path)
        ),
    )
    with pytest.raises(OSError, match="publish failed"):
        v15.crear_excel_listado_turno_cerrado(
            snapshot, tmp_path / "failed-hidden-temp.xlsx"
        )
    for temporary_file in tmp_path.glob(".turn-closure-*.xlsx"):
        temporary_file.unlink()


def test_nonempty_turn_closure_requires_transition_identifier(tmp_path):
    v15 = _v15_module()
    snapshot = SimpleNamespace(
        patient_count=1,
        dataset_revision="revision",
        context=SimpleNamespace(transition_id=""),
    )

    with pytest.raises(ValueError, match="transition_id"):
        v15.generate_turn_closure_report_files(snapshot, output_directory=tmp_path)


@pytest.mark.parametrize("system", ["Windows", "Darwin", "Linux"])
def test_generated_excel_open_uses_the_platform_launcher(tmp_path, monkeypatch, system):
    v15 = _v15_module()
    workbook = tmp_path / "turn.xlsx"
    workbook.write_bytes(b"xlsx")
    launched = []
    monkeypatch.setattr(v15.platform, "system", lambda: system)
    monkeypatch.setattr(
        v15.os, "startfile", lambda path: launched.append(("windows", path))
    )
    monkeypatch.setattr(
        v15.subprocess,
        "run",
        lambda command, check=False: launched.append((command[0], command[1])),
    )

    assert v15.abrir_excel_generado(workbook, mostrar_error=False) is True
    assert launched


def test_generated_excel_open_failure_is_nonfatal(tmp_path):
    v15 = _v15_module()

    assert (
        v15.abrir_excel_generado(tmp_path / "missing.xlsx", mostrar_error=False)
        is False
    )


def test_generated_excel_open_failure_can_show_friendly_warning(tmp_path, monkeypatch):
    v15 = _v15_module()
    warnings = []
    monkeypatch.setattr(
        v15.messagebox,
        "showwarning",
        lambda title, message: warnings.append((title, message)),
    )

    assert v15.abrir_excel_generado(tmp_path / "missing.xlsx") is False
    assert warnings and warnings[0][0] == "Listado Excel"


class _ImmediatePostCommitApp:
    def __init__(self, v15, *, fail=False):
        self.db = object()
        self.app_settings = {
            "turnos_generate_report": True,
            "turnos_save_excel_copy": True,
            "auto_print": True,
            "print_auto_reporte_turno": True,
            "print_auto_excel_turno": True,
            "print_copies_reporte": 2,
            "print_copies_excel": 2,
            "turnos_open_archive_folder": False,
        }
        self.statuses = []
        self.retry_excel_calls = 0
        self.scheduled = []
        self._fail = fail
        self.root = SimpleNamespace(
            after=lambda delay, callback: self.scheduled.append((delay, callback))
        )
        self._run_turn_post_commit_effects = lambda context, new=None, retry_attempt=0: (
            v15.App._run_turn_post_commit_effects(self, context, new, retry_attempt)
        )

    def set_status(self, message, level):
        self.statuses.append((message, level))

    def _retry_excel_export_jobs(self):
        self.retry_excel_calls += 1

    def _ejecutar_en_segundo_plano(
        self, _message, operation, al_terminar=None, al_error=None
    ):
        if self._fail:
            al_error(RuntimeError("central report query failed"))
            return "failed-worker"
        al_terminar(operation())
        return "completed-worker"


def test_post_commit_open_failures_do_not_repeat_handoff(monkeypatch):
    v15 = _v15_module()
    context = v15.OutgoingTurnContext(
        operational_source_id="source-old",
        turn_id=410,
        generation=17,
        operational_revision=29,
        representative_id="user-old",
        representative_display_name="REPRESENTANTE ANTERIOR",
        started_at=datetime(2026, 9, 2, 8, 0, 0),
        closed_at=datetime(2026, 9, 3, 8, 0, 0),
        turn_code="8AM_8AM",
        base_date=datetime(2026, 9, 2).date(),
        transition_id="99999999-9999-4999-8999-999999999999",
        new_turn_id=411,
        new_representative_id="user-new",
    )
    snapshot = SimpleNamespace(
        patient_count=2,
        dataset_revision="revision-1",
    )
    files = v15.TurnClosureReportFiles(
        pdf_path="C:/reports/old.pdf",
        excel_path="C:/reports/old.xlsx",
        patient_count=2,
        dataset_revision="revision-1",
        pdf_generated=True,
        excel_generated=True,
    )
    monkeypatch.setattr(
        v15, "build_turn_closure_report_snapshot", lambda *_args: snapshot
    )
    monkeypatch.setattr(
        v15, "generate_turn_closure_report_files", lambda *_args, **_kwargs: files
    )
    monkeypatch.setattr(v15, "abrir_pdf", lambda *_args, **_kwargs: False)
    monkeypatch.setattr(v15, "abrir_excel_generado", lambda *_args, **_kwargs: False)
    monkeypatch.setattr(v15, "imprimir_pdf", lambda *_args, **_kwargs: False)
    monkeypatch.setattr(v15, "imprimir_excel", lambda *_args, **_kwargs: False)
    current = _ImmediatePostCommitApp(v15)

    worker = v15.App._run_turn_post_commit_effects(current, context)

    assert worker == "completed-worker"
    assert current.retry_excel_calls == 1
    assert current.scheduled == []
    assert current.statuses[-1][1] == "warning"


def test_post_commit_generation_failure_retries_only_the_reports(monkeypatch):
    v15 = _v15_module()
    context = v15.OutgoingTurnContext(
        operational_source_id="source-old",
        turn_id=410,
        generation=17,
        operational_revision=29,
        representative_id="user-old",
        representative_display_name="REPRESENTANTE ANTERIOR",
        started_at=datetime(2026, 9, 2, 8, 0, 0),
        closed_at=datetime(2026, 9, 3, 8, 0, 0),
        turn_code="8AM_8AM",
        base_date=datetime(2026, 9, 2).date(),
        transition_id="99999999-9999-4999-8999-999999999999",
        new_turn_id=411,
        new_representative_id="user-new",
    )
    current = _ImmediatePostCommitApp(v15, fail=True)

    worker = v15.App._run_turn_post_commit_effects(current, context)

    assert worker == "failed-worker"
    assert current.retry_excel_calls == 1
    assert len(current.scheduled) == 1
    assert current.scheduled[0][0] == 30000
    assert current.statuses[-1][1] == "warning"


def test_post_commit_empty_snapshot_finishes_without_open_or_print(monkeypatch):
    v15 = _v15_module()
    context = v15.OutgoingTurnContext(
        operational_source_id="source-old",
        turn_id=410,
        generation=17,
        operational_revision=29,
        representative_id="user-old",
        representative_display_name="REPRESENTANTE ANTERIOR",
        started_at=datetime(2026, 9, 2, 8, 0, 0),
        closed_at=datetime(2026, 9, 3, 8, 0, 0),
        turn_code="8AM_8AM",
        base_date=datetime(2026, 9, 2).date(),
        transition_id="99999999-9999-4999-8999-999999999999",
        new_turn_id=411,
        new_representative_id="user-new",
    )
    snapshot = SimpleNamespace(patient_count=0, dataset_revision="empty")
    files = v15.TurnClosureReportFiles(
        pdf_path="",
        excel_path="",
        patient_count=0,
        dataset_revision="empty",
    )
    monkeypatch.setattr(
        v15, "build_turn_closure_report_snapshot", lambda *_args: snapshot
    )
    monkeypatch.setattr(
        v15, "generate_turn_closure_report_files", lambda *_args, **_kwargs: files
    )
    current = _ImmediatePostCommitApp(v15)
    current.app_settings["turnos_open_archive_folder"] = True

    v15.App._run_turn_post_commit_effects(current, context)

    assert current.statuses[-1] == (
        "Relevo aplicado; el turno saliente no contiene pacientes.",
        "ok",
    )


def test_post_commit_excel_only_with_auto_print_disabled(monkeypatch):
    v15 = _v15_module()
    context = v15.OutgoingTurnContext(
        operational_source_id="source-old",
        turn_id=410,
        generation=17,
        operational_revision=29,
        representative_id="user-old",
        representative_display_name="REPRESENTANTE ANTERIOR",
        started_at=datetime(2026, 9, 2, 8, 0, 0),
        closed_at=datetime(2026, 9, 3, 8, 0, 0),
        turn_code="8AM_8AM",
        base_date=datetime(2026, 9, 2).date(),
        transition_id="99999999-9999-4999-8999-999999999999",
        new_turn_id=411,
        new_representative_id="user-new",
    )
    snapshot = SimpleNamespace(patient_count=1, dataset_revision="revision")
    files = v15.TurnClosureReportFiles(
        pdf_path="",
        excel_path="C:/reports/old.xlsx",
        patient_count=1,
        dataset_revision="revision",
    )
    monkeypatch.setattr(
        v15, "build_turn_closure_report_snapshot", lambda *_args: snapshot
    )
    monkeypatch.setattr(
        v15, "generate_turn_closure_report_files", lambda *_args, **_kwargs: files
    )
    monkeypatch.setattr(v15, "abrir_excel_generado", lambda *_args, **_kwargs: True)
    current = _ImmediatePostCommitApp(v15)
    current.app_settings["auto_print"] = False

    v15.App._run_turn_post_commit_effects(current, context)

    assert current.statuses[-1][1] == "ok"
    pdf_files = v15.TurnClosureReportFiles(
        pdf_path="C:/reports/old.pdf",
        excel_path="",
        patient_count=1,
        dataset_revision="revision",
    )
    monkeypatch.setattr(
        v15, "generate_turn_closure_report_files", lambda *_args, **_kwargs: pdf_files
    )
    monkeypatch.setattr(v15, "abrir_pdf", lambda *_args, **_kwargs: True)
    v15.App._run_turn_post_commit_effects(current, context)
    assert current.statuses[-1][1] == "ok"


def test_second_report_failure_does_not_schedule_an_infinite_retry():
    v15 = _v15_module()
    context = v15.OutgoingTurnContext(
        operational_source_id="source-old",
        turn_id=410,
        generation=17,
        operational_revision=29,
        representative_id="user-old",
        representative_display_name="REPRESENTANTE ANTERIOR",
        started_at=datetime(2026, 9, 2, 8, 0, 0),
        closed_at=datetime(2026, 9, 3, 8, 0, 0),
        turn_code="8AM_8AM",
        base_date=datetime(2026, 9, 2).date(),
        transition_id="99999999-9999-4999-8999-999999999999",
        new_turn_id=411,
        new_representative_id="user-new",
    )
    current = _ImmediatePostCommitApp(v15, fail=True)

    v15.App._run_turn_post_commit_effects(current, context, retry_attempt=1)

    assert current.scheduled == []


@pytest.mark.parametrize("system", ["Windows", "Darwin", "Linux"])
def test_post_commit_can_open_archive_folder_for_each_platform(monkeypatch, system):
    v15 = _v15_module()
    context = v15.OutgoingTurnContext(
        operational_source_id="source-old",
        turn_id=410,
        generation=17,
        operational_revision=29,
        representative_id="user-old",
        representative_display_name="REPRESENTANTE ANTERIOR",
        started_at=datetime(2026, 9, 2, 8, 0, 0),
        closed_at=datetime(2026, 9, 3, 8, 0, 0),
        turn_code="8AM_8AM",
        base_date=datetime(2026, 9, 2).date(),
        transition_id="99999999-9999-4999-8999-999999999999",
        new_turn_id=411,
        new_representative_id="user-new",
    )
    snapshot = SimpleNamespace(patient_count=1, dataset_revision="revision")
    files = v15.TurnClosureReportFiles(
        pdf_path="C:/reports/old.pdf",
        excel_path="",
        patient_count=1,
        dataset_revision="revision",
    )
    launched = []
    monkeypatch.setattr(
        v15, "build_turn_closure_report_snapshot", lambda *_args: snapshot
    )
    monkeypatch.setattr(
        v15, "generate_turn_closure_report_files", lambda *_args, **_kwargs: files
    )
    monkeypatch.setattr(v15, "abrir_pdf", lambda *_args, **_kwargs: True)
    monkeypatch.setattr(v15, "imprimir_pdf", lambda *_args, **_kwargs: True)
    monkeypatch.setattr(v15.platform, "system", lambda: system)
    monkeypatch.setattr(
        v15.os, "startfile", lambda path: launched.append(("windows", path))
    )
    monkeypatch.setattr(
        v15.subprocess,
        "run",
        lambda command, check=False: launched.append((command[0], command[1])),
    )
    current = _ImmediatePostCommitApp(v15)
    current.app_settings["turnos_open_archive_folder"] = True

    v15.App._run_turn_post_commit_effects(current, context)

    assert launched


def test_report_retry_schedule_failure_is_nonfatal():
    v15 = _v15_module()
    context = v15.OutgoingTurnContext(
        operational_source_id="source-old",
        turn_id=410,
        generation=17,
        operational_revision=29,
        representative_id="user-old",
        representative_display_name="REPRESENTANTE ANTERIOR",
        started_at=datetime(2026, 9, 2, 8, 0, 0),
        closed_at=datetime(2026, 9, 3, 8, 0, 0),
        turn_code="8AM_8AM",
        base_date=datetime(2026, 9, 2).date(),
        transition_id="99999999-9999-4999-8999-999999999999",
        new_turn_id=411,
        new_representative_id="user-new",
    )
    current = _ImmediatePostCommitApp(v15, fail=True)
    current.root.after = lambda *_args: (_ for _ in ()).throw(
        RuntimeError("scheduler closed")
    )

    assert v15.App._run_turn_post_commit_effects(current, context) == "failed-worker"


def test_archive_folder_open_failure_does_not_change_handoff(monkeypatch):
    v15 = _v15_module()
    context = v15.OutgoingTurnContext(
        operational_source_id="source-old",
        turn_id=410,
        generation=17,
        operational_revision=29,
        representative_id="user-old",
        representative_display_name="REPRESENTANTE ANTERIOR",
        started_at=datetime(2026, 9, 2, 8, 0, 0),
        closed_at=datetime(2026, 9, 3, 8, 0, 0),
        turn_code="8AM_8AM",
        base_date=datetime(2026, 9, 2).date(),
        transition_id="99999999-9999-4999-8999-999999999999",
        new_turn_id=411,
        new_representative_id="user-new",
    )
    snapshot = SimpleNamespace(patient_count=1, dataset_revision="revision")
    files = v15.TurnClosureReportFiles(
        pdf_path="C:/reports/old.pdf",
        excel_path="",
        patient_count=1,
        dataset_revision="revision",
    )
    monkeypatch.setattr(
        v15, "build_turn_closure_report_snapshot", lambda *_args: snapshot
    )
    monkeypatch.setattr(
        v15, "generate_turn_closure_report_files", lambda *_args, **_kwargs: files
    )
    monkeypatch.setattr(v15, "abrir_pdf", lambda *_args, **_kwargs: True)
    monkeypatch.setattr(v15, "imprimir_pdf", lambda *_args, **_kwargs: True)
    monkeypatch.setattr(v15.platform, "system", lambda: "Windows")
    monkeypatch.setattr(
        v15.os,
        "startfile",
        lambda *_args: (_ for _ in ()).throw(OSError("shell unavailable")),
    )
    current = _ImmediatePostCommitApp(v15)
    current.app_settings["turnos_open_archive_folder"] = True

    assert v15.App._run_turn_post_commit_effects(current, context) == "completed-worker"


def test_empty_excel_dataset_is_recorded_only_once(tmp_path, monkeypatch):
    v15 = _v15_module()
    state_path = tmp_path / "excel_export_state.json"
    monkeypatch.setattr(v15, "EXCEL_EXPORT_STATE_PATH", str(state_path))
    database = _TurnDatabase(rows=[])
    messages = []
    monkeypatch.setattr(
        v15.APP_LOG,
        "info",
        lambda message, *args: messages.append(message % args if args else message),
    )

    assert v15.reconstruir_excel_turno(database, _turn_config()) == 0
    assert v15.reconstruir_excel_turno(database, _turn_config()) == 0

    assert sum("ADMISSION_EXCEL_SKIPPED_EMPTY" in value for value in messages) == 1
    assert v15._read_excel_export_state()["excel_status"] == "SKIPPED_EMPTY"


@pytest.mark.skipif(
    os.name != "nt", reason="Valida el bloqueo real de archivos de Windows"
)
def test_real_windows_file_lock_does_not_get_closed_by_exporter(tmp_path):
    v15 = _v15_module()
    source = tmp_path / "versionado.xlsx"
    canonical = tmp_path / "LISTADO DE PACIENTES EN EMERGENCIA.xlsx"
    source.write_bytes(b"nuevo")
    canonical.write_bytes(b"anterior")

    create_file = ctypes.windll.kernel32.CreateFileW
    create_file.restype = ctypes.c_void_p
    handle = create_file(
        str(canonical),
        0x80000000,  # GENERIC_READ
        0,  # sin compartir: equivalente a libro bloqueado
        None,
        3,  # OPEN_EXISTING
        0x80,  # FILE_ATTRIBUTE_NORMAL
        None,
    )
    assert handle not in (None, ctypes.c_void_p(-1).value)
    try:
        assert v15.excel_canonical_in_use(str(canonical)) is True
        with pytest.raises(OSError) as caught:
            v15._update_canonical_excel(str(source), str(canonical))
        assert v15._excel_file_in_use(caught.value)
        # La aplicación no cierra ni altera el handle perteneciente al usuario.
        assert ctypes.windll.kernel32.GetFileType(handle) != 0
    finally:
        assert ctypes.windll.kernel32.CloseHandle(handle)

    v15._update_canonical_excel(str(source), str(canonical))
    assert canonical.read_bytes() == b"nuevo"
