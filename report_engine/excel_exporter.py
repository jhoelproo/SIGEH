from __future__ import annotations

import os
import tempfile
from datetime import datetime


BLUE = "123F83"
LIGHT_BLUE = "EAF2FF"
GREEN = "0B7A5A"
PURPLE = "6B35C8"
ORANGE = "D16619"
TEXT = "24364B"
GRID = "D7E2F0"


def _safe_excel_value(value):
    """Evita inyección de fórmulas en textos procedentes de catálogos o usuarios."""
    if isinstance(value, str) and value[:1] in ("=", "+", "-", "@"):
        return "'" + value
    return value


def _excel_datetime(value):
    if isinstance(value, datetime):
        return value
    if isinstance(value, str):
        try:
            return datetime.fromisoformat(value.replace("Z", "+00:00")).replace(tzinfo=None)
        except ValueError:
            return _safe_excel_value(value)
    return value


def export_panel_xlsx(data: dict, output_path: str, generated_by: str, logo_path: str = "") -> str:
    """Exporta el panel a Excel con gráficos nativos y dos hojas auditables."""
    try:
        from openpyxl import Workbook, load_workbook
        from openpyxl.chart import BarChart, DoughnutChart, LineChart, Reference
        from openpyxl.chart.label import DataLabelList
        from openpyxl.drawing.image import Image
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
        from openpyxl.worksheet.table import Table, TableStyleInfo
    except ImportError as exc:
        raise RuntimeError(
            "La exportación Excel requiere openpyxl. Instálalo con: pip install openpyxl"
        ) from exc

    os.makedirs(os.path.dirname(os.path.abspath(output_path)), exist_ok=True)
    wb = Workbook()
    summary_ws = wb.active
    summary_ws.title = "Resumen"
    data_ws = wb.create_sheet("Datos")
    summary_ws.sheet_view.showGridLines = False
    data_ws.sheet_view.showGridLines = False

    thin = Side(style="thin", color=GRID)
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    currency_format = '"RD$" #,##0.00'
    percent_format = "0.0%"

    summary_ws.merge_cells("A1:L2")
    title_cell = summary_ws["A1"]
    title_cell.value = "HOSPITAL PROVINCIAL DR. ÁNGEL CONTRERAS MEJÍA"
    title_cell.fill = PatternFill("solid", fgColor=BLUE)
    title_cell.font = Font(name="Aptos Display", size=18, bold=True, color="FFFFFF")
    title_cell.alignment = Alignment(vertical="center", horizontal="center")
    summary_ws.row_dimensions[1].height = 28
    summary_ws.row_dimensions[2].height = 18

    if logo_path and os.path.isfile(logo_path):
        try:
            logo = Image(logo_path)
            logo.width = 92
            logo.height = 46
            summary_ws.add_image(logo, "A1")
        except Exception:
            pass

    summary_ws.merge_cells("A4:L4")
    summary_ws["A4"] = "PANEL DE REPORTES Y GRÁFICOS"
    summary_ws["A4"].font = Font(size=15, bold=True, color=BLUE)
    summary_ws["A4"].alignment = Alignment(horizontal="center")
    summary_ws.merge_cells("A5:L5")
    summary_ws["A5"] = (
        f"Período: {data['start_date']} al {data['end_date']}  |  "
        f"Generado por: {generated_by}  |  {datetime.now():%d/%m/%Y %H:%M}"
    )
    summary_ws["A5"].font = Font(size=10, color="52657A")
    summary_ws["A5"].alignment = Alignment(horizontal="center")

    def selection_text(selection, empty):
        if not isinstance(selection, dict) or not selection.get("values"):
            return empty
        mode = "Excluir" if selection.get("mode") in ("exclude", "excluir") else "Incluir"
        return f"{mode}: {', '.join(selection['values'])}"
    filters = data.get("filters", {})
    summary_ws.merge_cells("A6:L6")
    summary_ws["A6"] = (
        f"Análisis: {filters.get('analysis_label', 'Facturación confirmada')}  |  "
        f"Criterio: {filters.get('date_basis_label', 'Fecha de validación')}  |  "
        f"ARS: {selection_text(filters.get('ars'), 'Todas')}  |  "
        f"Facturadores: {selection_text(filters.get('user'), 'Todos')}  |  "
        f"Cobertura: {filters.get('coverage', 'Todas')}  |  "
        f"Comparación: {'Sí' if filters.get('compare_previous') else 'No'}"
    )
    summary_ws["A6"].font = Font(size=9, color="52657A")
    summary_ws["A6"].alignment = Alignment(horizontal="center")

    summary = data["summary"]
    view = data.get("view", {})
    show_ars_comparison = bool(view.get("show_ars_comparison"))
    cards = [
        ("A", "C", filters.get("receipt_label", "TOTAL DE RECIBOS"), summary["receipts"], BLUE, "#,##0"),
        ("D", "F", filters.get("total_label", "TOTAL EMITIDO"), summary["total"], GREEN, currency_format),
        ("G", "I", "PROMEDIO POR RECIBO", summary["average"], ORANGE, currency_format),
        ("J", "L", "SALA DE EMERGENCIA", summary.get("room", 0), PURPLE, currency_format),
    ]
    for first_col, last_col, label, value, color, number_format in cards:
        summary_ws.merge_cells(f"{first_col}7:{last_col}7")
        summary_ws.merge_cells(f"{first_col}8:{last_col}9")
        label_cell = summary_ws[f"{first_col}7"]
        value_cell = summary_ws[f"{first_col}8"]
        label_cell.value = label
        value_cell.value = value
        label_cell.font = Font(size=9, bold=True, color="62748A")
        value_cell.font = Font(size=15, bold=True, color=color)
        label_cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        value_cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        value_cell.number_format = number_format
        for row in summary_ws.iter_rows(
            min_row=7,
            max_row=9,
            min_col=label_cell.column,
            max_col=summary_ws[f"{last_col}7"].column,
        ):
            for cell in row:
                cell.fill = PatternFill("solid", fgColor="FFFFFF")
                cell.border = border
    for row in range(7, 10):
        summary_ws.row_dimensions[row].height = 25

    summary_headers = ["ARS", "Total emitido", "Recibos", "% del dinero", "% de recibos", "Promedio / recibo"]
    start_row = 45
    start_col = 1
    if show_ars_comparison:
        for col, header in enumerate(summary_headers, start_col):
            cell = summary_ws.cell(start_row, col, header)
            cell.fill = PatternFill("solid", fgColor=BLUE)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.alignment = Alignment(horizontal="center")
    summary_rows = (data.get("summary_table") or []) if show_ars_comparison else []
    for row_index, row in enumerate(summary_rows, start_row + 1):
        values = [
            row["label"], row["total"], row["receipts"], row["money_percentage"],
            row["receipt_percentage"], row.get("average", 0),
        ]
        for col, value in enumerate(values, start_col):
            cell = summary_ws.cell(row_index, col, value)
            cell.border = border
            cell.alignment = Alignment(vertical="center", horizontal="right" if col > start_col else "left")
        summary_ws.cell(row_index, start_col + 1).number_format = currency_format
        summary_ws.cell(row_index, start_col + 3).number_format = percent_format
        summary_ws.cell(row_index, start_col + 4).number_format = percent_format
        summary_ws.cell(row_index, start_col + 5).number_format = currency_format
    if summary_rows:
        table_ref = f"A{start_row}:F{start_row + len(summary_rows)}"
        table = Table(displayName="PanelSummaryTable", ref=table_ref)
        table.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium2", showFirstColumn=False, showLastColumn=False,
            showRowStripes=True, showColumnStripes=False,
        )
        summary_ws.add_table(table)

    helper_row = max(start_row + len(summary_rows) + 5, 65)
    trend_col = 14
    summary_ws.cell(helper_row, trend_col, "Período")
    summary_ws.cell(helper_row, trend_col + 1, "Total")
    summary_ws.cell(helper_row, trend_col + 2, "Recibos")
    summary_ws.cell(helper_row, trend_col + 3, "Promedio")
    for index, row in enumerate(data.get("trend", []), helper_row + 1):
        summary_ws.cell(index, trend_col, row["label"])
        summary_ws.cell(index, trend_col + 1, row["total"]).number_format = currency_format
        summary_ws.cell(index, trend_col + 2, row["receipts"])
        summary_ws.cell(index, trend_col + 3, row.get("average", 0)).number_format = currency_format

    category_col = 18
    summary_ws.cell(helper_row, category_col, "Categoría")
    summary_ws.cell(helper_row, category_col + 1, "Total")
    for index, row in enumerate(data.get("category_distribution", data.get("categories", [])), helper_row + 1):
        summary_ws.cell(index, category_col, row["label"])
        summary_ws.cell(index, category_col + 1, row["total"]).number_format = currency_format

    category_bar_col = 27
    summary_ws.cell(helper_row, category_bar_col, "Categoría")
    summary_ws.cell(helper_row, category_bar_col + 1, "Total")
    for index, row in enumerate(data.get("categories", []), helper_row + 1):
        summary_ws.cell(index, category_bar_col, row["label"])
        summary_ws.cell(index, category_bar_col + 1, row["total"]).number_format = currency_format

    comparison_col = 21
    summary_ws.cell(helper_row, comparison_col, "ARS")
    summary_ws.cell(helper_row, comparison_col + 1, "Recibos")
    summary_ws.cell(helper_row, comparison_col + 2, "Total")
    summary_ws.cell(helper_row, comparison_col + 3, "% dinero")
    summary_ws.cell(helper_row, comparison_col + 4, "% recibos")
    summary_ws.cell(helper_row, comparison_col + 5, "Promedio")
    for index, row in enumerate(data.get("comparison", []), helper_row + 1):
        summary_ws.cell(index, comparison_col, row["label"])
        summary_ws.cell(index, comparison_col + 1, row["receipts"])
        summary_ws.cell(index, comparison_col + 2, row["total"]).number_format = currency_format
        summary_ws.cell(index, comparison_col + 3, row["money_percentage"]).number_format = percent_format
        summary_ws.cell(index, comparison_col + 4, row["receipt_percentage"]).number_format = percent_format
        summary_ws.cell(index, comparison_col + 5, row.get("average", 0)).number_format = currency_format

    evolution_metric = view.get("evolution_metric", "total")
    ars_metric = view.get("ars_metric", "total")

    if data.get("category_distribution", data.get("categories")):
        doughnut = DoughnutChart()
        doughnut.title = "Distribución porcentual por categoría"
        distribution_rows = data.get("category_distribution", data.get("categories", []))
        values = Reference(summary_ws, min_col=category_col + 1, min_row=helper_row, max_row=helper_row + len(distribution_rows))
        labels = Reference(summary_ws, min_col=category_col, min_row=helper_row + 1, max_row=helper_row + len(distribution_rows))
        doughnut.add_data(values, titles_from_data=True)
        doughnut.set_categories(labels)
        doughnut.holeSize = 55
        doughnut.dataLabels = DataLabelList()
        doughnut.dataLabels.showPercent = True
        doughnut.visible_cells_only = False
        doughnut.height = 7.5
        doughnut.width = 13
        summary_ws.add_chart(doughnut, "A12")

    if data.get("categories"):
        category_bar = BarChart()
        category_bar.title = f"Distribución por categorías - {filters.get('total_label', 'TOTAL EMITIDO').title()}"
        category_bar.y_axis.title = "RD$"
        values = Reference(summary_ws, min_col=category_bar_col + 1, min_row=helper_row, max_row=helper_row + len(data["categories"]))
        labels = Reference(summary_ws, min_col=category_bar_col, min_row=helper_row + 1, max_row=helper_row + len(data["categories"]))
        category_bar.add_data(values, titles_from_data=True)
        category_bar.set_categories(labels)
        category_bar.legend = None
        category_bar.visible_cells_only = False
        category_bar.height = 7.5
        category_bar.width = 13
        summary_ws.add_chart(category_bar, "G12")

    comparison_column = comparison_col + 4 if ars_metric == "receipts" else comparison_col + 3
    if show_ars_comparison and data.get("comparison"):
        comparison_bar = BarChart()
        comparison_bar.title = f"Distribución porcentual por ARS - {view.get('ars_metric_label', 'Total emitido')}"
        comparison_bar.y_axis.title = "Participación"
        comparison_bar.y_axis.numFmt = "0%"
        values = Reference(summary_ws, min_col=comparison_column, min_row=helper_row, max_row=helper_row + len(data["comparison"]))
        labels = Reference(summary_ws, min_col=comparison_col, min_row=helper_row + 1, max_row=helper_row + len(data["comparison"]))
        comparison_bar.add_data(values, titles_from_data=True)
        comparison_bar.set_categories(labels)
        comparison_bar.legend = None
        comparison_bar.visible_cells_only = False
        comparison_bar.height = 7.5
        comparison_bar.width = 13
        summary_ws.add_chart(comparison_bar, "A28")

    trend_column = {
        "total": trend_col + 1,
        "receipts": trend_col + 2,
        "average": trend_col + 3,
    }.get(evolution_metric, trend_col + 1)
    if data.get("trend"):
        line = LineChart()
        line.title = f"Evolución diaria - {view.get('evolution_label', 'Total emitido')}"
        line.y_axis.title = "RD$" if evolution_metric in ("total", "average") else "Recibos"
        line.x_axis.title = "Día"
        values = Reference(summary_ws, min_col=trend_column, min_row=helper_row, max_row=helper_row + len(data["trend"]))
        labels = Reference(summary_ws, min_col=trend_col, min_row=helper_row + 1, max_row=helper_row + len(data["trend"]))
        line.add_data(values, titles_from_data=True)
        line.set_categories(labels)
        line.legend = None
        line.visible_cells_only = False
        line.height = 7.5
        line.width = 13
        if show_ars_comparison:
            summary_ws.add_chart(line, "G28")
        else:
            line.width = 26
            summary_ws.add_chart(line, "A28")

    for column in range(14, 29):
        summary_ws.column_dimensions[summary_ws.cell(1, column).column_letter].hidden = True

    detail_headers = [
        "Recibo", "Fecha servicio", "Fecha generada", "Usuario", "ARS", "Cobertura",
        "Categoría", "Artículo", "Cantidad", "Precio unitario", "Total",
        "Estado de facturación", "Fecha de validación", "Validado por", "Referencia",
        "Asignado a auditoría", "Riesgo", "Código no facturado",
        "Estado del documento", "N.º autorización",
    ]
    data_ws.append(detail_headers)
    for detail in data.get("details", []):
        data_ws.append([
            detail["receipt"], _safe_excel_value(detail.get("service_date", "")),
            _excel_datetime(detail["created_at"]), _safe_excel_value(detail["username"]),
            _safe_excel_value(detail["ars"]), _safe_excel_value(detail.get("coverage", "")),
            _safe_excel_value(detail["category"]), _safe_excel_value(detail["item"]), detail["quantity"],
            detail["unit_price"], detail["total"], _safe_excel_value(detail.get("billing_status", "")),
            _excel_datetime(detail.get("billing_status_at", "")),
            _safe_excel_value(detail.get("validated_by", "")),
            _safe_excel_value(detail.get("billing_reference", "")),
            _safe_excel_value(detail.get("audit_assignee", "")),
            int(detail.get("audit_risk", 0) or 0),
            _safe_excel_value(detail.get("not_invoiced_reason_code", "")),
            _safe_excel_value(detail.get("document_state", "")),
            _safe_excel_value(detail.get("authorization_number", "")),
        ])
    for cell in data_ws[1]:
        cell.fill = PatternFill("solid", fgColor=BLUE)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.alignment = Alignment(horizontal="center")
    for row in data_ws.iter_rows(min_row=2, max_col=20):
        for cell in row:
            cell.border = Border(bottom=Side(style="hair", color=GRID))
        row[9].number_format = currency_format
        row[10].number_format = currency_format
        row[2].number_format = "yyyy-mm-dd hh:mm:ss"
        row[12].number_format = "yyyy-mm-dd hh:mm:ss"
    if data_ws.max_row > 1:
        data_table = Table(displayName="PanelDetailTable", ref=f"A1:T{data_ws.max_row}")
        data_table.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium2", showFirstColumn=False, showLastColumn=False,
            showRowStripes=True, showColumnStripes=False,
        )
        data_ws.add_table(data_table)
    data_ws.freeze_panes = "A2"
    data_ws.auto_filter.ref = f"A1:R{max(1, data_ws.max_row)}"

    widths = {"A": 12, "B": 15, "C": 22, "D": 18, "E": 24, "F": 18,
              "G": 20, "H": 40, "I": 12, "J": 18, "K": 18, "L": 22,
              "M": 22, "N": 20, "O": 20, "P": 22, "Q": 12, "R": 24}
    for column, width in widths.items():
        data_ws.column_dimensions[column].width = width
    for column in "ABCDEFGHIJKLM":
        summary_ws.column_dimensions[column].width = 15
    summary_ws.freeze_panes = "A7"
    summary_ws.sheet_properties.pageSetUpPr.fitToPage = True
    summary_ws.page_setup.fitToWidth = 1
    summary_ws.page_setup.fitToHeight = 0
    summary_ws.print_options.horizontalCentered = True
    summary_ws.sheet_view.zoomScale = 85
    data_ws.sheet_view.zoomScale = 90

    shift = data.get("shift_billing") or {}
    if shift.get("by_ars"):
        shift_ws = wb.create_sheet("Turno actual")
        shift_ws.sheet_view.showGridLines = False
        shift_ws.merge_cells("A1:F2")
        shift_ws["A1"] = "SEGUIMIENTO DE FACTURACIÓN DEL TURNO ACTUAL"
        shift_ws["A1"].fill = PatternFill("solid", fgColor=BLUE)
        shift_ws["A1"].font = Font(size=16, bold=True, color="FFFFFF")
        shift_ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
        shift_ws.append([])
        shift_ws.append([
            "Pacientes admitidos", shift.get("admitted", 0),
            "Facturados", shift.get("invoiced", 0), "Pendientes", shift.get("pending", 0),
        ])
        shift_ws.append([
            "Requieren corrección", shift.get("correction", 0),
            "Facturados de listados anteriores", shift.get("previous_day_invoiced", 0),
            "Monto de listados anteriores", shift.get("previous_day_total", 0),
        ])
        shift_ws["F5"].number_format = currency_format
        shift_ws.append([])
        shift_ws.append(["ARS", "Admitidos", "Facturados", "Pendientes", "Requieren corrección"])
        for row in shift.get("by_ars", []):
            shift_ws.append([
                _safe_excel_value(row.get("ars", "")), row.get("admitted", 0),
                row.get("invoiced", 0), row.get("pending", 0), row.get("correction", 0),
            ])
        for cell in shift_ws[7]:
            cell.fill = PatternFill("solid", fgColor=BLUE)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.alignment = Alignment(horizontal="center")
        for column, width in {
            "A": 30, "B": 14, "C": 34, "D": 14, "E": 32, "F": 16,
        }.items():
            shift_ws.column_dimensions[column].width = width
        shift_ws.freeze_panes = "A8"

    output_path = os.path.abspath(output_path)
    output_dir = os.path.dirname(output_path)
    with tempfile.NamedTemporaryFile(
        prefix="panel_", suffix=".xlsx", dir=output_dir, delete=False
    ) as temporary:
        temporary_path = temporary.name
    try:
        wb.save(temporary_path)
        verification = load_workbook(temporary_path, read_only=False, data_only=False)
        try:
            expected_sheets = ["Resumen", "Datos"] + (
                ["Turno actual"] if shift.get("by_ars") else []
            )
            if verification.sheetnames != expected_sheets:
                raise RuntimeError("La estructura del libro exportado no es válida.")
            expected_charts = sum((
                bool(data.get("category_distribution", data.get("categories"))),
                bool(data.get("categories")),
                bool(show_ars_comparison and data.get("comparison")),
                bool(data.get("trend")),
            ))
            if len(verification["Resumen"]._charts) != expected_charts:
                raise RuntimeError(
                    f"No se pudieron validar los {expected_charts} gráficos del panel."
                )
        finally:
            verification.close()
        os.replace(temporary_path, output_path)
    finally:
        if os.path.exists(temporary_path):
            os.remove(temporary_path)
    return os.path.abspath(output_path)
